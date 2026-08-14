"""
bff/routers/optimization.py

Optimization endpoints:
  GET   /scenarios/{id}/optimization            → get optimization result
  POST  /scenarios/{id}/run-optimization        → async: run MILP/ALNS optimizer
"""

from __future__ import annotations

import traceback
import json
import csv
import logging
import math
import shutil
from dataclasses import is_dataclass, replace
from collections import Counter, defaultdict
from collections.abc import Mapping
import threading
import multiprocessing
import os
import time
from concurrent.futures import Executor, Future, ProcessPoolExecutor, ThreadPoolExecutor
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Dict, List, Literal, Optional

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field

from bff.dependencies import require_built
from bff.errors import AppErrorCode, make_error
from bff.mappers.scenario_to_problemdata import (
    ScenarioBuildReport,
    build_problem_data_from_scenario,
)
from bff.mappers.solver_results import (
    serialize_milp_result,
    serialize_simulation_result,
)
from bff.routers.graph import (
    _build_blocks_payload,
    _build_dispatch_plan_payload,
    _build_duties_payload,
    _build_graph_payload,
    _build_trips_payload,
)
from bff.services.experiment_reports import log_optimization_experiment
from bff.services.optimization_run.artifact_completeness import (
    SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE,
    SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_SCHEMA_VERSION,
    STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE,
    persist_frontend_run_artifact_audit,
    validate_solver_objective_accounting_reconciliation,
    validate_stage1_used_powertrain_composition_search,
)
from bff.services.optimization_run.canonical_graph import (
    canonical_datetime_from_min as _canonical_datetime_from_min,
    canonical_deadhead_distance_km as _canonical_deadhead_distance_km,
    canonical_estimated_deadhead_energy_kwh as _canonical_estimated_deadhead_energy_kwh,
    canonical_horizon_start_min as _canonical_horizon_start_min,
    canonical_output_base_date as _canonical_output_base_date,
    canonical_slot_datetime as _canonical_slot_datetime,
    canonical_vehicle_initial_soc_kwh as _canonical_vehicle_initial_soc_kwh,
)
from bff.services.optimization_run.cost_breakdown import (
    CANONICAL_LEDGER_COMPONENT_SOURCES,
    canonical_cost_breakdown_json as _canonical_cost_breakdown_json,
    canonical_cost_ledger_json as _canonical_cost_ledger_json,
    cost_breakdown as _cost_breakdown,
)
from bff.services.optimization_run.execute import (
    normalize_solver_mode as _normalize_solver_mode,
    parse_optimization_mode as _parse_optimization_mode,
    phase_from_solver_mode as _phase_from_solver_mode,
)
from bff.services.optimization_run.input_provenance import (
    MANIFEST_FILE as RUN_INPUT_MANIFEST_FILE,
    collect_git_state,
    persist_run_input_provenance,
)
from bff.services.optimization_run.rolling_chain import (
    DAY_AHEAD_EXPLORATORY_PROFILE,
    DEFAULT_FRONTEND_RUN_PROFILE,
    RollingChainExecutionError,
    execute_frontend_rolling_chain,
    finalize_frontend_rolling_evidence,
    frontend_rolling_is_required,
    normalize_frontend_run_profile,
    persist_frontend_day_ahead_rolling_contract,
    refresh_frontend_rolling_manifest,
)
from bff.services.optimization_run.thesis_ablation import (
    CSV_COLUMNS as THESIS_ABLATION_CSV_COLUMNS,
    ablation_candidate_csv_rows,
    build_day_ahead_ablation_candidates,
)
from bff.services.optimization_run.rich_outputs import (
    persist_json_outputs as _persist_json_outputs,
    run_stamp as _run_stamp,
    service_date_for_output as _service_date_for_output,
    write_csv_rows as _write_csv_rows,
)
from bff.services.optimization_run.weather import (
    configured_service_date as _configured_service_date,
    load_weather_proxy_for_bff as _load_weather_proxy_for_bff,
    preflight_weather_proxy_request as _preflight_weather_proxy_request,
    prepare_weather_policy_for_scenario as _prepare_weather_policy_for_scenario,
    resolve_weather_proxy_path as _resolve_weather_proxy_path,
    weather_policy_payload_from_problem_metadata as _weather_policy_payload_from_problem_metadata,
    weather_policy_requested as _weather_policy_requested,
    weather_proxy_http_error as _weather_proxy_http_error,
)
from bff.services.run_preparation import (
    get_or_build_run_preparation,
    load_prepared_input,
    materialize_scenario_from_prepared_input,
)
from bff.services.optimization_run.vehicle_timeline import vehicle_ids_with_timeline_activity
from bff.store import job_store, output_paths, scenario_store as store
from src.dispatch.models import hhmm_to_min
from src.optimization import (
    OptimizationConfig,
    OptimizationEngine,
    ProblemBuilder,
    ResultSerializer,
)
from src.optimization.common.energy_flow_accounting import (
    compute_pv_curtail_kwh,
    compute_pv_utilization_rate,
    normalize_pv_energy_breakdown,
)
from src.optimization.common.fleet_contract import (
    canonical_powertrain,
    resolve_scenario_fleet_contract,
)
from src.optimization.common.time_axis import normalize_timestep_min
from src.optimization.common.soc_helpers import (
    deadhead_distance_km,
    is_electric_vehicle,
    return_deadhead_min_to_home,
    vehicle_energy_rate_kwh_per_km,
)

from src.optimization.rolling.reoptimizer import (
    RollingReoptimizer,
    assignment_plan_from_serialized_result,
)
from src.optimization.rolling.acceptance import rolling_chain_acceptance_audit
from src.optimization.common.bess_terminal_policy import (
    resolve_bess_terminal_soc_target_kwh,
)
from src.preprocess.weather.operation_policy import apply_weather_policy_to_problem
from src.run_output_layout import allocate_run_dir
from src.pipeline.solve import solve_problem_data

logger = logging.getLogger(__name__)
router = APIRouter(tags=["optimization"])
_OPTIMIZATION_EXECUTOR: Optional[Executor] = None

# Freeze the source identity when this BFF worker imports the optimization
# router. Reading Git only when a request arrives is insufficient: a stale
# long-lived Python process can otherwise report the repository's newer HEAD
# while it is still executing modules loaded from an older commit.
_BFF_RUNTIME_STARTED_AT_UTC = datetime.now(timezone.utc).isoformat()
_BFF_RUNTIME_PROCESS_ID = os.getpid()
_BFF_RUNTIME_GIT_STATE = dict(collect_git_state())

# Interactive BFF/Tk launches are used for comparable research runs.  Keep
# this policy at the BFF boundary so a stale client payload cannot silently
# re-enable the Stage 1 early-stop rule or vary Gurobi parallelism.  The formal
# CLI runner remains independently configurable for non-interactive studies.
INTERACTIVE_RUNTIME_POLICY_VERSION = "interactive_runtime_controls_v1"
INTERACTIVE_STAGE1_BEST_OBJ_STOP_ENABLED = False
# Keep a fixed value for pair reproducibility while using enough of the
# research workstation to finish the integrated root relaxation and proof.
INTERACTIVE_GUROBI_THREADS = 4
INTERACTIVE_TERMINAL_SOC_POLICY_VERSION = "interactive_terminal_soc_controls_v1"
INTERACTIVE_BEV_TERMINAL_SOC_POLICY = "return_to_initial"
INTERACTIVE_OPERATION_TIME_WINDOW_CONTROLS_VERSION = (
    "interactive_operation_time_window_controls_v1"
)
FULL_DAY_OPERATION_START_TIME = "00:00"
FULL_DAY_OPERATION_END_TIME = "23:59"
# ``0`` is the canonical numeric full-network sentinel consumed by
# ``MILPModelBuilder._successor_limit``. Do not send a presentation label into
# the typed solver setting.
FORMAL_RESEARCH_MAX_SUCCESSORS_PER_TRIP = 0
FORMAL_RESEARCH_SUCCESSOR_POLICY = "full_network"
_FUEL_SUMMARY_FIELDS = (
    "vehicle_id",
    "vehicle_type",
    "fuel_liters",
    "trip_fuel_liters",
    "deadhead_fuel_liters",
    "refuel_liters",
    "unit",
)


def _research_git_state_is_ready(git_state: Dict[str, Any]) -> bool:
    """Return whether a Git record satisfies the formal-run start contract."""

    return bool(
        bool(git_state.get("git_state_available", False))
        and str(git_state.get("git_sha") or "").strip()
        and git_state.get("git_dirty") is False
    )


def _runtime_git_attestation(
    current_git_state: Dict[str, Any],
    *,
    runtime_git_state: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Compare the code state loaded at BFF startup with the current checkout."""

    startup = dict(runtime_git_state or _BFF_RUNTIME_GIT_STATE)
    startup_ready = _research_git_state_is_ready(startup)
    current_ready = _research_git_state_is_ready(current_git_state)
    state_matches = bool(
        startup_ready
        and current_ready
        and startup.get("git_sha") == current_git_state.get("git_sha")
        and startup.get("repository_root")
        == current_git_state.get("repository_root")
    )
    return {
        "schema_version": "bff_runtime_git_attestation_v1",
        "runtime_started_at_utc": _BFF_RUNTIME_STARTED_AT_UTC,
        "runtime_process_id": _BFF_RUNTIME_PROCESS_ID,
        "runtime_git_state_available": bool(
            startup.get("git_state_available", False)
        ),
        "runtime_git_sha": startup.get("git_sha"),
        "runtime_git_dirty": startup.get("git_dirty"),
        "runtime_git_state_error": startup.get("git_state_error"),
        "runtime_repository_root": startup.get("repository_root"),
        "runtime_git_state_ready": startup_ready,
        "runtime_git_state_matches_current": state_matches,
    }


def _require_matching_research_runtime_git_state(
    *,
    research_run: bool,
    current_git_state: Dict[str, Any],
) -> Dict[str, Any]:
    """Fail closed when a formal request reaches a stale BFF process."""

    attestation = _runtime_git_attestation(current_git_state)
    if not research_run:
        return attestation
    if attestation["runtime_git_state_matches_current"]:
        return attestation
    raise ValueError(
        "formal research run requires the BFF process to have started from "
        "the same clean Git commit as the current checkout; restart the BFF "
        f"after committing changes. runtime_sha={attestation.get('runtime_git_sha')}, "
        f"current_sha={current_git_state.get('git_sha')}, "
        f"runtime_dirty={attestation.get('runtime_git_dirty')}, "
        f"current_dirty={current_git_state.get('git_dirty')}"
    )


def _research_git_preflight_payload(
    git_state: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Build the user-facing preflight from the canonical provenance collector."""

    state = dict(git_state or collect_git_state())
    runtime_attestation = _runtime_git_attestation(state)
    return {
        "formal_research_ready": bool(
            _research_git_state_is_ready(state)
            and runtime_attestation["runtime_git_state_matches_current"]
        ),
        "git_state_available": bool(state.get("git_state_available", False)),
        "git_sha": state.get("git_sha"),
        "git_dirty": state.get("git_dirty"),
        "git_state_error": state.get("git_state_error"),
        "uncommitted_changes": list(state.get("status_porcelain") or ()),
        "repository_root": state.get("repository_root"),
        **runtime_attestation,
    }


def _require_clean_research_git_state(
    *, research_run: bool, git_state: Dict[str, Any]
) -> None:
    """Fail closed before a formal research solve on unversioned source."""

    if not research_run:
        return
    if _research_git_state_is_ready(git_state):
        return
    raise ValueError(
        "formal research run requires a clean Git worktree with an available "
        f"commit SHA; git_state_available={bool(git_state.get('git_state_available', False))}, "
        f"git_dirty={git_state.get('git_dirty')}, "
        f"git_state_error={git_state.get('git_state_error')!r}"
    )


def _require_research_git_preflight_before_job_creation(
    *, research_run: bool
) -> Optional[Dict[str, Any]]:
    """Reject a formal request synchronously, while retaining the worker guard."""

    if not research_run:
        return None
    git_state = collect_git_state()
    try:
        _require_clean_research_git_state(
            research_run=True,
            git_state=git_state,
        )
        _require_matching_research_runtime_git_state(
            research_run=True,
            current_git_state=git_state,
        )
    except ValueError as exc:
        preflight = _research_git_preflight_payload(git_state)
        raise HTTPException(
            status_code=409,
            detail=make_error(
                AppErrorCode.RESEARCH_GIT_STATE_INVALID,
                str(exc),
                **preflight,
            ),
        ) from exc
    return git_state


def _available_inventory_for_selected_depot(
    scenario: Dict[str, Any],
    *,
    depot_id: Optional[str],
) -> Dict[str, int]:
    """Return scenario-derived counts using the canonical fleet resolver."""

    contract = resolve_scenario_fleet_contract(
        scenario,
        selected_depot_ids=(str(depot_id or "").strip(),),
        research_run=True,
    )
    return dict(contract.inventory_by_powertrain)


def _apply_interactive_research_contract(
    scenario: Dict[str, Any],
    *,
    research_run: bool,
    depot_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Apply the fail-closed scenario-fleet and full-network formal-run contract."""

    if not research_run:
        return {
            "enabled": False,
            "scope": "non_research_interactive_run",
        }
    simulation_config = scenario.get("simulation_config")
    if not isinstance(simulation_config, dict):
        simulation_config = {}
        scenario["simulation_config"] = simulation_config
    scenario_overlay = scenario.get("scenario_overlay")
    if not isinstance(scenario_overlay, dict):
        scenario_overlay = {}
        scenario["scenario_overlay"] = scenario_overlay
    solver_config = scenario_overlay.get("solver_config")
    if not isinstance(solver_config, dict):
        solver_config = {}
        scenario_overlay["solver_config"] = solver_config

    fleet_contract = resolve_scenario_fleet_contract(
        scenario,
        selected_depot_ids=(str(depot_id or "").strip(),),
        research_run=True,
    )
    expected_inventory = dict(fleet_contract.inventory_by_powertrain)
    simulation_config["research_vehicle_inventory"] = expected_inventory
    simulation_config["research_vehicle_ids"] = list(
        fleet_contract.active_vehicle_ids
    )
    simulation_config["research_vehicle_id_hash"] = (
        fleet_contract.active_vehicle_id_hash
    )
    simulation_config["research_vehicle_parameter_hash"] = (
        fleet_contract.vehicle_parameter_hash
    )
    simulation_config["research_vehicle_initial_state_hash"] = (
        fleet_contract.initial_state_hash
    )
    simulation_config["research_fleet_contract_hash"] = (
        fleet_contract.fleet_contract_hash
    )
    simulation_config["scenario_fleet_contract"] = fleet_contract.to_dict(
        include_source_records=True
    )
    simulation_config["milp_max_successors_per_trip"] = (
        FORMAL_RESEARCH_MAX_SUCCESSORS_PER_TRIP
    )
    solver_config["milp_max_successors_per_trip"] = (
        FORMAL_RESEARCH_MAX_SUCCESSORS_PER_TRIP
    )
    return {
        "enabled": True,
        "scope": "interactive_formal_research_run",
        "expected_available_inventory": expected_inventory,
        "inventory_source": "selected_scenario_depot_available_vehicles",
        "inventory_depot_id": str(depot_id or "").strip() or None,
        "active_vehicle_ids": list(fleet_contract.active_vehicle_ids),
        "active_vehicle_id_hash": fleet_contract.active_vehicle_id_hash,
        "vehicle_parameter_hash": fleet_contract.vehicle_parameter_hash,
        "initial_state_hash": fleet_contract.initial_state_hash,
        "fleet_contract_hash": fleet_contract.fleet_contract_hash,
        "excluded_vehicle_records": [
            dict(item) for item in fleet_contract.excluded_vehicle_records
        ],
        "milp_successor_policy": FORMAL_RESEARCH_SUCCESSOR_POLICY,
        "milp_max_successors_per_trip": (
            FORMAL_RESEARCH_MAX_SUCCESSORS_PER_TRIP
        ),
        "successor_pruning_allowed": False,
        "fallback_allowed": False,
        "postsolve_repair_allowed": False,
        "clean_git_worktree_required": True,
    }


def _validate_git_state_after_solve(
    *,
    research_run: bool,
    before: Dict[str, Any],
    after: Dict[str, Any],
) -> bool:
    """Reject a research solve only when its source changes during the solve."""

    unchanged = all(
        before.get(key) == after.get(key)
        for key in (
            "git_state_available",
            "git_sha",
            "git_dirty",
            "worktree_patch_sha256",
        )
    )
    if research_run:
        _require_clean_research_git_state(
            research_run=True,
            git_state=after,
        )
        if not unchanged:
            raise ValueError(
                "research run source state changed during solve; "
                f"before_sha={before.get('git_sha')}, "
                f"after_sha={after.get('git_sha')}, "
                f"before_dirty={before.get('git_dirty')}, "
                f"after_dirty={after.get('git_dirty')}"
            )
        _require_matching_research_runtime_git_state(
            research_run=True,
            current_git_state=after,
        )
    return unchanged


def _interactive_runtime_controls_payload(
    *,
    requested_stage1_best_obj_stop_enabled: Any,
    requested_gurobi_threads: Any,
) -> Dict[str, Any]:
    """Describe the server-enforced solver controls for an interactive run."""

    requested = {
        "stage1_best_obj_stop_enabled": bool(
            requested_stage1_best_obj_stop_enabled
        ),
        "gurobi_threads": requested_gurobi_threads,
    }
    effective = {
        "stage1_best_obj_stop_enabled": INTERACTIVE_STAGE1_BEST_OBJ_STOP_ENABLED,
        "gurobi_threads": INTERACTIVE_GUROBI_THREADS,
    }
    return {
        "policy_version": INTERACTIVE_RUNTIME_POLICY_VERSION,
        "scope": "interactive_bff_run_optimization",
        "enforced": True,
        "requested": requested,
        "effective": effective,
        "override_applied": requested != effective,
        "reason": (
            "Interactive runs disable Stage 1 BestObjStop and use four fixed "
            "Gurobi threads so their solver controls are recorded consistently "
            "while Phase 4 can advance its global bound."
        ),
    }


def _apply_interactive_bev_terminal_soc_policy(
    scenario: Dict[str, Any],
) -> Dict[str, Any]:
    """Enforce energy-neutral BEV terminal SOC for interactive day-ahead runs.

    A fixed percentage terminal target lets vehicles with heterogeneous initial
    SOC contribute or retain inventory across the representative day.  That
    makes a daily high-PV/low-PV cost comparison non-neutral.  The interactive
    path therefore applies the per-vehicle ``return_to_initial`` policy after
    all scenario/weather overlays and before ``ProblemBuilder`` is called.
    The formal CLI remains explicit and independently configurable.
    """

    simulation_config = scenario.get("simulation_config")
    if not isinstance(simulation_config, dict):
        simulation_config = {}
        scenario["simulation_config"] = simulation_config

    scenario_overlay = scenario.get("scenario_overlay")
    charging_constraints: Dict[str, Any] = {}
    if isinstance(scenario_overlay, dict):
        candidate = scenario_overlay.get("charging_constraints")
        if isinstance(candidate, dict):
            charging_constraints = candidate

    terminal_fields = (
        "bev_terminal_soc_policy",
        "terminal_soc_policy",
        "final_soc_target_percent",
        "final_soc_target_tolerance_percent",
    )
    requested = {
        "simulation_config": {
            field: simulation_config.get(field) for field in terminal_fields
        },
        "charging_constraints": {
            field: charging_constraints.get(field) for field in terminal_fields
        },
    }

    # Set the policy in the source with builder precedence and clear legacy
    # fixed-target inputs.  ``None`` deliberately masks an inherited overlay
    # because ProblemBuilder's _first_present treats it as the explicit value.
    simulation_config["bev_terminal_soc_policy"] = INTERACTIVE_BEV_TERMINAL_SOC_POLICY
    simulation_config["terminal_soc_policy"] = INTERACTIVE_BEV_TERMINAL_SOC_POLICY
    simulation_config["final_soc_target_percent"] = None
    simulation_config["final_soc_target_tolerance_percent"] = None
    if charging_constraints:
        charging_constraints["bev_terminal_soc_policy"] = INTERACTIVE_BEV_TERMINAL_SOC_POLICY
        charging_constraints["terminal_soc_policy"] = INTERACTIVE_BEV_TERMINAL_SOC_POLICY
        charging_constraints["final_soc_target_percent"] = None
        charging_constraints["final_soc_target_tolerance_percent"] = None

    effective = {
        "bev_terminal_soc_policy": INTERACTIVE_BEV_TERMINAL_SOC_POLICY,
        "terminal_soc_policy": INTERACTIVE_BEV_TERMINAL_SOC_POLICY,
        "final_soc_target_percent": None,
        "final_soc_target_tolerance_percent": None,
    }
    requested_policy = requested["simulation_config"].get(
        "bev_terminal_soc_policy"
    ) or requested["charging_constraints"].get("bev_terminal_soc_policy")
    requested_target = requested["simulation_config"].get(
        "final_soc_target_percent"
    )
    if requested_target is None:
        requested_target = requested["charging_constraints"].get(
            "final_soc_target_percent"
        )
    requested_tolerance = requested["simulation_config"].get(
        "final_soc_target_tolerance_percent"
    )
    if requested_tolerance is None:
        requested_tolerance = requested["charging_constraints"].get(
            "final_soc_target_tolerance_percent"
        )
    return {
        "policy_version": INTERACTIVE_TERMINAL_SOC_POLICY_VERSION,
        "scope": "interactive_bff_run_optimization",
        "enforced": True,
        "requested": requested,
        "effective": effective,
        "override_applied": bool(
            requested_policy != INTERACTIVE_BEV_TERMINAL_SOC_POLICY
            or requested_target is not None
            or requested_tolerance is not None
        ),
        "reason": (
            "Interactive day-ahead runs enforce per-vehicle return_to_initial "
            "BEV terminal SOC so representative-day energy and cost comparisons "
            "do not consume or create BEV inventory."
        ),
    }


def _coerce_bool(value: Any, *, default: bool) -> bool:
    """Parse persisted/UI booleans without treating an invalid string as true."""

    if value is None:
        return bool(default)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "on"}:
            return True
        if normalized in {"0", "false", "no", "off"}:
            return False
        return bool(default)
    return bool(value)


def _apply_interactive_operation_time_window_controls(
    scenario: Dict[str, Any],
) -> Dict[str, Any]:
    """Materialize the front-end time-window intent before canonical build.

    The start/end pair is retained as the user's requested pair even when the
    checkbox is disabled.  In that case this function writes separate effective
    fields and a 24-hour horizon, while ``ProblemBuilder`` independently enforces
    the same full-day semantics.  Keeping both representations makes a later
    artifact review distinguish user input from solver input.
    """

    simulation_config = scenario.get("simulation_config")
    if not isinstance(simulation_config, dict):
        simulation_config = {}
        scenario["simulation_config"] = simulation_config

    requested_enabled = simulation_config.get("operation_time_window_enabled")
    if requested_enabled is None:
        requested_enabled = simulation_config.get("operationTimeWindowEnabled")
    # Prepared inputs created before this feature retain their established
    # scoped-time behavior.  Tk Prepare writes the flag explicitly.
    enabled = _coerce_bool(requested_enabled, default=True)
    requested_start_time = str(
        simulation_config.get("start_time") or "05:00"
    )
    requested_end_time = str(
        simulation_config.get("end_time") or "23:00"
    )
    planning_days = max(int(simulation_config.get("planning_days") or 1), 1)
    effective_start_time = (
        requested_start_time if enabled else FULL_DAY_OPERATION_START_TIME
    )
    effective_end_time = (
        requested_end_time if enabled else FULL_DAY_OPERATION_END_TIME
    )

    simulation_config["operation_time_window_enabled"] = enabled
    simulation_config["operation_time_window_effective_start_time"] = (
        effective_start_time
    )
    simulation_config["operation_time_window_effective_end_time"] = (
        effective_end_time
    )
    if not enabled:
        simulation_config["planning_horizon_hours"] = 24.0 * float(planning_days)

    return {
        "schema_version": INTERACTIVE_OPERATION_TIME_WINDOW_CONTROLS_VERSION,
        "scope": "interactive_bff_run_optimization",
        "requested": {
            "operation_time_window_enabled": requested_enabled,
            "start_time": requested_start_time,
            "end_time": requested_end_time,
        },
        "effective": {
            "operation_time_window_enabled": enabled,
            "start_time": effective_start_time,
            "end_time": effective_end_time,
            "planning_horizon_hours": (
                24.0 * float(planning_days)
                if not enabled
                else simulation_config.get("planning_horizon_hours")
            ),
        },
        "full_day_horizon_forced": not enabled,
        "reason": (
            "The start/end pair is inactive; canonical optimization uses the "
            "complete 00:00-23:59 calendar day."
            if not enabled
            else "The explicitly enabled start/end pair scopes the optimization horizon."
        ),
    }
_OPTIMIZATION_FUTURE: Optional[Future[Any]] = None
_OPTIMIZATION_FUTURE_LOCK = threading.RLock()


def _require_nonempty_prepared_scope(prep, *, action: str) -> None:
    if int(prep.scope_summary.get("trip_count") or 0) > 0:
        return
    raise HTTPException(
        status_code=409,
        detail=make_error(
            AppErrorCode.SCENARIO_INCOMPLETE,
            f"{action} failed: no trips matched the current depot / route / day-type selection.",
            scopeSummary=prep.scope_summary,
        ),
    )


def _request_timestep_min(*values: Any) -> Optional[int]:
    raw = next((value for value in values if value is not None), None)
    if raw is None:
        return None
    try:
        return normalize_timestep_min(raw, default=30)
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                str(exc),
                field="timestep_min",
                allowed=[5, 15, 30, 60],
            ),
        ) from exc


def _apply_timestep_min_to_scenario(scenario: Dict[str, Any], timestep_min: Optional[int]) -> None:
    if timestep_min is None:
        return
    sim_cfg = scenario.get("simulation_config")
    if not isinstance(sim_cfg, dict):
        sim_cfg = {}
        scenario["simulation_config"] = sim_cfg
    sim_cfg["time_step_min"] = timestep_min
    sim_cfg["timestep_min"] = timestep_min


def _apply_interactive_bev_utilization_policy(
    problem: Any,
    *,
    require_all_available_bevs: bool,
) -> Dict[str, Any]:
    """Reuse the formal minimum-BEV-use constraint for an explicit sensitivity."""

    metadata = getattr(problem, "metadata", None)
    if not isinstance(metadata, dict):
        raise ValueError(
            "Canonical problem metadata must be mutable before applying the "
            "BEV utilization sensitivity."
        )
    fleet_contract = dict(metadata.get("scenario_fleet_contract") or {})
    contract_parameters = list(
        fleet_contract.get("active_vehicle_parameters") or ()
    )
    contract_bev_ids = {
        str(item.get("vehicle_id") or "").strip()
        for item in contract_parameters
        if isinstance(item, dict)
        and str(item.get("powertrain") or "").strip().upper() == "BEV"
    }
    powertrain_by_vehicle_type = {
        str(getattr(item, "vehicle_type_id", "") or "").strip(): str(
            getattr(item, "powertrain_type", "") or ""
        )
        .strip()
        .upper()
        for item in tuple(getattr(problem, "vehicle_types", ()) or ())
    }

    def _problem_vehicle_powertrain(vehicle: Any) -> str:
        vehicle_id = str(getattr(vehicle, "vehicle_id", "") or "").strip()
        if vehicle_id in contract_bev_ids:
            return "BEV"
        vehicle_type = str(
            getattr(vehicle, "vehicle_type", "") or ""
        ).strip()
        catalog_powertrain = powertrain_by_vehicle_type.get(vehicle_type)
        if catalog_powertrain:
            return catalog_powertrain
        return canonical_powertrain(
            {
                "type": vehicle_type,
                "battery_capacity_kwh": getattr(
                    vehicle,
                    "battery_capacity_kwh",
                    None,
                ),
                "fuel_consumption_l_per_km": getattr(
                    vehicle,
                    "fuel_consumption_l_per_km",
                    None,
                ),
            },
            research_run=False,
        )

    available_bev_ids = sorted(
        str(getattr(vehicle, "vehicle_id", "") or "")
        for vehicle in tuple(getattr(problem, "vehicles", ()) or ())
        if bool(getattr(vehicle, "available", True))
        and _problem_vehicle_powertrain(vehicle) == "BEV"
    )
    if require_all_available_bevs and not available_bev_ids:
        raise ValueError(
            "All-available-BEV sensitivity was requested, but the canonical "
            "problem contains no available BEV."
        )
    minimum_used_bev_count = (
        len(available_bev_ids) if require_all_available_bevs else 0
    )
    metadata["minimum_used_bev_count"] = minimum_used_bev_count
    metadata["minimum_used_bev_count_policy_case"] = bool(
        require_all_available_bevs
    )
    metadata["minimum_used_bev_vehicle_ids"] = (
        available_bev_ids if require_all_available_bevs else []
    )
    return {
        "policy": "require_all_available_bevs_at_least_one_trip",
        "enabled": bool(require_all_available_bevs),
        "available_bev_count": len(available_bev_ids),
        "available_bev_ids": available_bev_ids,
        "minimum_used_bev_count": minimum_used_bev_count,
        "mathematical_effect": (
            "sum(used_vehicle[v] for available BEV v) >= "
            f"{minimum_used_bev_count}"
            if require_all_available_bevs
            else "no additional minimum-BEV-use constraint"
        ),
        "claim_scope": (
            "policy sensitivity; not the unconstrained total-cost optimum"
            if require_all_available_bevs
            else "baseline optimization policy"
        ),
    }


class RunOptimizationBody(BaseModel):
    mode: str = "thesis_mode"
    research_run: bool = False
    time_step_min: Optional[int] = None
    timestep_min: Optional[int] = None
    time_limit_seconds: int = 300
    stage1_time_limit_seconds: Optional[int] = None
    stage2_time_limit_seconds: Optional[int] = None
    # These interactive defaults are also enforced by _run_optimization so an
    # older client cannot restore the early stop or a machine-dependent thread
    # count through a request body.
    stage1_best_obj_stop_enabled: bool = INTERACTIVE_STAGE1_BEST_OBJ_STOP_ENABLED
    stage1_stage2_candidate_limit: int = Field(default=1, ge=1, le=100)
    stage1_composition_search_radius: int = Field(default=0, ge=0, le=100)
    stage1_bev_frontier_enabled: bool = False
    stage1_bev_frontier_min_count: int = Field(default=15, ge=0, le=200)
    stage1_bev_frontier_max_count: int = Field(default=35, ge=0, le=200)
    stage1_bev_frontier_target_time_limit_seconds: int = Field(
        default=120,
        ge=1,
        le=3600,
    )
    integrated_actual_cost_objective: bool = False
    integrated_ev_utilization_mode: Literal[
        "disabled",
        "minimum_ice_fuel_lexicographic",
    ] = "disabled"
    integrated_actual_cost_upper_bound_jpy: Optional[float] = Field(
        default=None,
        ge=0.0,
    )
    integrated_actual_cost_upper_bound_delta_ratio: Optional[float] = Field(
        default=None,
        ge=0.0,
        le=1.0,
    )
    co2_emissions_cap_kg: Optional[float] = Field(default=None, ge=0.0)
    gurobi_threads: Optional[int] = Field(
        default=INTERACTIVE_GUROBI_THREADS,
        ge=1,
    )
    run_profile: str = DEFAULT_FRONTEND_RUN_PROFILE
    run_hourly_rolling: bool = True
    rolling_execution_minutes: int = Field(default=60, ge=1)
    mip_gap: float = 0.01
    random_seed: int = 42
    prepared_input_id: Optional[str] = None
    service_id: Optional[str] = None
    depot_id: Optional[str] = None
    rebuild_dispatch: bool = True
    force_reprepare: bool = False
    use_existing_duties: bool = False
    alns_iterations: int = 500
    no_improvement_limit: int = 100
    destroy_fraction: float = 0.25
    weatherProxyForecastPath: Optional[str] = None
    enableWeatherOperationPolicy: Optional[bool] = None
    require_all_available_bevs: bool = False


class DelayEventBody(BaseModel):
    trip_id: str
    delay_min: float


class ReoptimizeBody(BaseModel):
    mode: str = "hybrid"
    research_run: bool = False
    current_time: str
    time_step_min: Optional[int] = None
    timestep_min: Optional[int] = None
    time_limit_seconds: int = 180
    mip_gap: float = 0.02
    random_seed: int = 42
    alns_iterations: int = 300
    no_improvement_limit: int = 100
    destroy_fraction: float = 0.25
    prepared_input_id: Optional[str] = None
    service_id: Optional[str] = None
    depot_id: Optional[str] = None
    actual_soc: Dict[str, float] = Field(default_factory=dict)
    actual_bess_soc_kwh: Dict[str, float] = Field(default_factory=dict)
    observed_on_peak_kw_by_depot: Dict[str, float] = Field(default_factory=dict)
    observed_off_peak_kw_by_depot: Dict[str, float] = Field(default_factory=dict)
    actual_location_node_id: Dict[str, str] = Field(default_factory=dict)
    delays: list[DelayEventBody] = Field(default_factory=list)
    updated_pv_profile: list[Dict[str, Any]] = Field(default_factory=list)
    reoptimization_strategy: str = "generic"
    execution_minutes: int = Field(default=60, ge=1)
    bess_terminal_policy: str = "scenario"


def _optimization_capabilities() -> Dict[str, Any]:
    return {
        "implemented": True,
        "async_job": True,
        "job_persistence": dict(job_store.JOB_PERSISTENCE_INFO),
        "supported_modes": [
            "thesis_mode",
            "debug_mode",
            "mode_milp_only",
            "mode_alns_only",
            "mode_ga_only",
            "mode_abc_only",
            "mode_hybrid",
            "phase1_charging_only",
            "phase2_assignment_only",
            "phase3_two_stage",
            "phase4_integrated",
            "diagnostic",
        ],
        "mode_aliases": {
            "milp": "mode_milp_only",
            "exact": "mode_milp_only",
            "thesis": "thesis_mode",
            "debug": "debug_mode",
            "alns": "mode_alns_only",
            "heuristic": "mode_alns_only",
            "ga": "mode_ga_only",
            "genetic": "mode_ga_only",
            "abc": "mode_abc_only",
            "colony": "mode_abc_only",
            "hybrid": "mode_hybrid",
            "phase1": "phase1_charging_only",
            "phase2": "phase2_assignment_only",
            "phase3": "phase3_two_stage",
            "phase4": "phase4_integrated",
            "diagnostic_mode": "diagnostic",
        },
        "deprecated_modes": {
            "mode_alns_milp": "mode_hybrid (auto-routed)",
            "mode_a_journey_charge": "BLOCKED - no longer supported",
            "mode_b_optimistic": "BLOCKED - no longer supported",
        },
        "default_mode": "thesis_mode",
        "interactive_runtime_controls": {
            "policy_version": INTERACTIVE_RUNTIME_POLICY_VERSION,
            "stage1_best_obj_stop_enabled": INTERACTIVE_STAGE1_BEST_OBJ_STOP_ENABLED,
            "gurobi_threads": INTERACTIVE_GUROBI_THREADS,
            "enforced_server_side": True,
        },
        "interactive_terminal_soc_controls": {
            "policy_version": INTERACTIVE_TERMINAL_SOC_POLICY_VERSION,
            "bev_terminal_soc_policy": INTERACTIVE_BEV_TERMINAL_SOC_POLICY,
            "enforced_server_side": True,
        },
        "interactive_rolling_controls": {
            "default_run_profile": DEFAULT_FRONTEND_RUN_PROFILE,
            "run_hourly_rolling": True,
            "rolling_execution_minutes": 60,
            "day_ahead_only_profile": DAY_AHEAD_EXPLORATORY_PROFILE,
            "enforced_server_side": True,
        },
        "authoritative_engine": "canonical (src/optimization/)",
        "supports_reoptimization": True,
        "max_concurrent_jobs": 1,
        "execution_model": f"{_executor_mode()}_pool",
        "notes": [
            "All supported modes use the canonical optimization engine (src/optimization/).",
            "thesis_mode runs the canonical two-stage MILP without unserved decision variables or postsolve repair.",
            "debug_mode keeps diagnostic unserved variables and is not eligible for research KPI claims.",
            "Phase 1 requires a concrete fixed assignment with duties before charging/PV/BESS optimization.",
            "Phase 2 returns assignment-only MILP results; charging/SOC feasibility is explicitly not evaluated.",
            "diagnostic is diagnostic-only and is never eligible for research KPI claims.",
            "mode_alns_milp is auto-routed to mode_hybrid (ALNS+MILP hybrid).",
            "Optimization runs against canonical CanonicalOptimizationProblem built from scenario.",
            "Dispatch artifacts can be rebuilt before solve when requested.",
            "Results are persisted to the scenario snapshot; job state is not.",
            "Optimization/re-optimization runs in a dedicated executor so API polling stays responsive.",
            "Only one optimization/re-optimization job is allowed at a time in this BFF process.",
            "Interactive /run-optimization enforces Stage 1 BestObjStop=OFF and Gurobi Threads=4; the formal CLI runner remains explicit.",
            "The default interactive profile executes the complete 60-minute rolling chain in the same job; day-ahead-only diagnostics require run_profile=day_ahead_exploratory.",
        ],
    }


def _executor_mode() -> str:
    mode = (os.getenv("BFF_OPT_EXECUTOR") or "").strip().lower()
    if mode in {"process", "thread"}:
        return mode
    # Windows + spawn で worker が即死するケースがあるため既定は thread。
    return "thread" if os.name == "nt" else "process"


def _get_optimization_executor() -> Executor:
    global _OPTIMIZATION_EXECUTOR
    with _OPTIMIZATION_FUTURE_LOCK:
        if _OPTIMIZATION_EXECUTOR is None:
            if _executor_mode() == "thread":
                _OPTIMIZATION_EXECUTOR = ThreadPoolExecutor(max_workers=1)
            else:
                _OPTIMIZATION_EXECUTOR = ProcessPoolExecutor(
                    max_workers=1,
                    mp_context=multiprocessing.get_context("spawn"),
                )
    return _OPTIMIZATION_EXECUTOR


def shutdown_optimization_executor() -> None:
    global _OPTIMIZATION_EXECUTOR, _OPTIMIZATION_FUTURE
    with _OPTIMIZATION_FUTURE_LOCK:
        executor = _OPTIMIZATION_EXECUTOR
        future = _OPTIMIZATION_FUTURE
        _OPTIMIZATION_EXECUTOR = None
        _OPTIMIZATION_FUTURE = None
    if future is not None and not future.done():
        future.cancel()
    if executor is not None:
        executor.shutdown(wait=True, cancel_futures=True)


def _register_optimization_future(
    future: Future[Any],
    *,
    job_id: str,
    scenario_id: str,
    service_id: str,
    depot_id: Optional[str],
    mode: str,
    stage: str,
) -> None:
    def _handle_completion(done: Future[Any]) -> None:
        try:
            exc = done.exception()
        except Exception as callback_exc:  # pragma: no cover - defensive
            exc = callback_exc
        if exc is None:
            return
        try:
            job_store.update_job(
                job_id,
                status="failed",
                progress=100,
                message="Optimization worker crashed.",
                error=str(exc),
                metadata=_job_metadata(
                    scenario_id=scenario_id,
                    service_id=service_id,
                    depot_id=depot_id,
                    stage=stage,
                    mode=mode,
                    extra={"worker_failure": True},
                ),
            )
        except KeyError:
            import logging
            logging.getLogger("bff.optimization").error(
                "Failed to update job %s status to failed (job may have been cleaned up). "
                "Worker error: %s",
                job_id, exc,
            )
            return

    future.add_done_callback(_handle_completion)


def _submit_optimization_job(
    *,
    fn,
    args: tuple[Any, ...],
    job_id: str,
    scenario_id: str,
    service_id: str,
    depot_id: Optional[str],
    mode: str,
    stage: str,
) -> bool:
    global _OPTIMIZATION_FUTURE
    with _OPTIMIZATION_FUTURE_LOCK:
        if _OPTIMIZATION_FUTURE is not None and not _OPTIMIZATION_FUTURE.done():
            return False
        future = _get_optimization_executor().submit(fn, *args)
        _OPTIMIZATION_FUTURE = future
        _register_optimization_future(
            future,
            job_id=job_id,
            scenario_id=scenario_id,
            service_id=service_id,
            depot_id=depot_id,
            mode=mode,
            stage=stage,
        )
        return True



def _not_found(scenario_id: str) -> HTTPException:
    return HTTPException(status_code=404, detail=f"Scenario '{scenario_id}' not found")


def _require_scenario(scenario_id: str) -> None:
    try:
        store.get_scenario(scenario_id)
        store.ensure_runtime_master_data(scenario_id)
    except KeyError:
        raise _not_found(scenario_id)
    except RuntimeError as e:
        if "artifacts are incomplete" in str(e):
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "INCOMPLETE_ARTIFACT",
                    "message": str(e)
                }
            )
        raise


def _resolve_dispatch_scope(
    scenario_id: str,
    *,
    service_id: Optional[str] = None,
    depot_id: Optional[str] = None,
    persist: bool = False,
) -> Dict[str, Any]:
    current = store.get_dispatch_scope(scenario_id)
    scope: Dict[str, Any] = {}
    if service_id is not None:
        scope["serviceId"] = service_id
    if depot_id is not None:
        scope["depotId"] = depot_id
    if not scope:
        return current
    if persist:
        return store.set_dispatch_scope(scenario_id, scope)
    doc = store.get_scenario_document_shallow(scenario_id)
    doc["dispatch_scope"] = {**current, **scope}
    return store._normalize_dispatch_scope(doc)

def _git_sha() -> str:
    """Return the current SHA when provenance collection succeeds.

    Legacy call sites consume a string, while rich artifacts additionally emit
    the full explicit state through :func:`collect_git_state`.
    """

    return str(collect_git_state().get("git_sha") or "")


def _prepared_inputs_root() -> Path:
    return output_paths.outputs_root() / "prepared_inputs"


def _persist_prepared_scope_artifacts(
    scenario_id: str,
    scenario_snapshot: Dict[str, Any],
    *,
    clear_stale_dispatch: bool = False,
) -> None:
    prepared_trips = list(scenario_snapshot.get("trips") or [])
    prepared_timetable_rows = list(
        scenario_snapshot.get("timetable_rows")
        or prepared_trips
    )
    prepared_stops = list(scenario_snapshot.get("stops") or [])
    prepared_stop_timetables = list(scenario_snapshot.get("stop_timetables") or [])
    if prepared_trips:
        store.set_field(scenario_id, "trips", prepared_trips)
    if prepared_timetable_rows:
        store.set_field(scenario_id, "timetable_rows", prepared_timetable_rows)
    if prepared_stops:
        store.set_field(scenario_id, "stops", prepared_stops)
    if prepared_stop_timetables:
        store.set_field(scenario_id, "stop_timetables", prepared_stop_timetables)
    if clear_stale_dispatch:
        store.set_field(scenario_id, "graph", {})
        store.set_field(scenario_id, "blocks", [])
        store.set_field(scenario_id, "duties", [])
        store.set_field(scenario_id, "dispatch_plan", {})


def _rebuild_dispatch_artifacts(
    scenario_id: str,
    service_id: str,
    depot_id: str,
) -> None:
    """Rebuild trips, graph, blocks, duties and dispatch_plan in one pass.

    Builds DispatchContext once and reuses it for all downstream steps,
    avoiding the O(n^2) graph analysis being repeated 4 times.
    """
    from src.dispatch.graph_builder import ConnectionGraphBuilder
    from src.dispatch.dispatcher import DispatchGenerator
    from src.dispatch.pipeline import TimetableDispatchPipeline
    from bff.routers.graph import (
        _build_dispatch_context,
        build_graph_response,
        trip_to_dict,
        vehicle_duty_to_dict,
    )

    context = _build_dispatch_context(scenario_id, service_id, depot_id)

    # Trips
    trips = [trip_to_dict(t) for t in context.trips]

    # Graph (O(n^2) — computed once)
    builder = ConnectionGraphBuilder()
    combined_graph: Dict[str, Any] = {
        "trips": trips,
        "arcs": [],
        "total_arcs": 0,
        "feasible_arcs": 0,
        "infeasible_arcs": 0,
        "reason_counts": {},
    }
    for vt in list(context.vehicle_profiles.keys()):
        analyzed_arcs = builder.analyze(context, vt)
        partial = build_graph_response(context.trips, analyzed_arcs)
        combined_graph["arcs"].extend(partial["arcs"])
        combined_graph["feasible_arcs"] += partial["feasible_arcs"]
        combined_graph["infeasible_arcs"] += partial["infeasible_arcs"]
        combined_graph["total_arcs"] += partial["total_arcs"]
        for rc, cnt in partial["reason_counts"].items():
            combined_graph["reason_counts"][rc] = (
                combined_graph["reason_counts"].get(rc, 0) + cnt
            )

    # Blocks (reuse context)
    generator = DispatchGenerator()
    vehicle_types = list(context.vehicle_profiles.keys())
    blocks: List[Dict[str, Any]] = []
    for vt in vehicle_types:
        for block in generator.generate_greedy_blocks(context, vt):
            blocks.append({
                "block_id": block.block_id,
                "vehicle_type": block.vehicle_type,
                "trip_ids": list(block.trip_ids),
            })

    # Duties (reuse context)
    pipeline = TimetableDispatchPipeline()
    duties: List[Dict[str, Any]] = []
    for vt in vehicle_types:
        result = pipeline.run(context, vt)
        for duty in result.duties:
            duties.append(vehicle_duty_to_dict(duty))

    # Dispatch plan (reuse blocks + duties)
    plan_blocks = [
        {
            "block_id": b["block_id"],
            "vehicle_type": b["vehicle_type"],
            "trip_ids": b["trip_ids"],
        }
        for b in blocks
    ]
    plan_duties = [
        {
            "duty_id": d["duty_id"],
            "vehicle_type": d.get("vehicle_type", "BEV"),
            "legs": d.get("legs", []),
        }
        for d in duties
    ]
    dispatch_plan = {
        "plans": [
            {
                "plan_id": f"plan_{vt}",
                "vehicle_type": vt,
                "blocks": [b for b in plan_blocks if b["vehicle_type"] == vt],
                "duties": [d for d in plan_duties if d["vehicle_type"] == vt],
                "charging_plan": [],
            }
            for vt in vehicle_types
        ],
        "total_plans": len(vehicle_types),
        "total_blocks": len(blocks),
        "total_duties": len(duties),
    }

    store.set_field(scenario_id, "trips", trips)
    store.set_field(scenario_id, "graph", combined_graph)
    store.set_field(scenario_id, "blocks", blocks)
    store.set_field(scenario_id, "duties", duties)
    store.set_field(scenario_id, "dispatch_plan", dispatch_plan)


def _job_metadata(
    *,
    scenario_id: str,
    service_id: str,
    depot_id: Optional[str],
    stage: str,
    mode: str,
    extra: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    return {
        "scenario_id": scenario_id,
        "service_id": service_id,
        "depot_id": depot_id,
        "stage": stage,
        "mode": mode,
        **(extra or {}),
    }


def _scenario_feed_context(scenario_id: str) -> Dict[str, Any]:
    return dict(store.get_feed_context(scenario_id) or {})


def _scoped_output_dir(
    *,
    root: str,
    feed_context: Dict[str, Any],
    scenario_id: str,
    stage: str,
    service_id: Optional[str] = None,
    depot_id: Optional[str] = None,
) -> str:
    return str(allocate_run_dir(root))


def _dated_scenario_run_dir(
    *,
    scenario: Dict[str, Any],
    scenario_id: str,
    mode: str,
    service_id: str,
    depot_id: Optional[str],
) -> Path:
    root = output_paths.outputs_root()
    return allocate_run_dir(root)


def _normalize_depot_slot_mapping(raw: Any) -> Dict[str, Dict[int, float]]:
    if not isinstance(raw, dict):
        return {}
    out: Dict[str, Dict[int, float]] = {}
    for depot_id, slot_map in raw.items():
        if isinstance(slot_map, dict):
            out[str(depot_id)] = {
                int(slot_idx): float(value or 0.0)
                for slot_idx, value in slot_map.items()
            }
        elif isinstance(slot_map, list):
            out[str(depot_id)] = {
                int(slot_idx): float(value or 0.0)
                for slot_idx, value in enumerate(slot_map)
            }
    return out


def _mapping_has_positive_flow(mapping: Dict[str, Dict[int, float]]) -> bool:
    return any(
        max(float(value or 0.0), 0.0) > 0.0
        for slot_map in mapping.values()
        for value in slot_map.values()
    )


def _depot_mapping_has_positive_flow(
    mapping: Dict[str, Dict[int, float]],
    depot_id: str,
) -> bool:
    return any(max(float(value or 0.0), 0.0) > 0.0 for value in (mapping.get(str(depot_id)) or {}).values())


def _canonical_vehicle_home_depot_map(problem) -> Dict[str, str]:
    return {
        str(getattr(vehicle, "vehicle_id", "") or ""): str(getattr(vehicle, "home_depot_id", "") or "")
        for vehicle in list(getattr(problem, "vehicles", ()) or ())
    }


def _canonical_charging_source_and_depot(
    problem,
    charging_slot,
) -> tuple[str, str]:
    vehicle_home_depot = _canonical_vehicle_home_depot_map(problem)
    fallback_depot = str(
        getattr(charging_slot, "charging_depot_id", "") or vehicle_home_depot.get(str(getattr(charging_slot, "vehicle_id", "") or ""), "")
    ).strip()
    raw = str(getattr(charging_slot, "charger_id", "") or "")
    if ":" in raw:
        source, depot_id = raw.split(":", 1)
        normalized_source = source.strip().lower()
        if normalized_source in {"grid", "pv", "bess"}:
            return normalized_source, depot_id.strip() or fallback_depot or "depot_default"
    return "grid", fallback_depot or "depot_default"


def _canonical_charging_slot_signature(charging_slot) -> tuple[str, int, str, str]:
    return (
        str(getattr(charging_slot, "vehicle_id", "") or ""),
        int(getattr(charging_slot, "slot_index", 0) or 0),
        str(getattr(charging_slot, "charger_id", "") or ""),
        str(getattr(charging_slot, "charging_depot_id", "") or ""),
    )


def _merge_depot_slot_flow_maps(
    base: Dict[str, Dict[int, float]],
    additions: Dict[str, Dict[int, float]],
) -> Dict[str, Dict[int, float]]:
    merged = {depot_id: dict(slot_map) for depot_id, slot_map in base.items()}
    for depot_id, slot_map in additions.items():
        if not slot_map:
            continue
        target = merged.setdefault(depot_id, {})
        for slot_idx, value in slot_map.items():
            target[slot_idx] = float(target.get(slot_idx, 0.0) or 0.0) + float(value or 0.0)
    return merged


def _bess_soc_bounds_for_asset(asset: Any) -> tuple[float, float, float] | None:
    capacity = max(float(getattr(asset, "bess_energy_kwh", 0.0) or 0.0), 0.0)
    configured_max = max(float(getattr(asset, "bess_soc_max_kwh", 0.0) or 0.0), 0.0)
    if configured_max > 0.0:
        max_soc = min(configured_max, capacity) if capacity > 0.0 else configured_max
    else:
        max_soc = capacity
    if max_soc <= 1.0e-9:
        return None
    min_soc = min(max(float(getattr(asset, "bess_soc_min_kwh", 0.0) or 0.0), 0.0), max_soc)
    initial_soc = min(max(float(getattr(asset, "bess_initial_soc_kwh", 0.0) or 0.0), min_soc), max_soc)
    return min_soc, max_soc, initial_soc


def _resolved_bess_terminal_target_kwh(asset: Any) -> float | None:
    bounds = _bess_soc_bounds_for_asset(asset)
    if bounds is None:
        return None
    min_soc, max_soc, initial_soc = bounds
    terminal_floor = min(
        max(
            float(getattr(asset, "bess_terminal_soc_min_kwh", 0.0) or 0.0),
            min_soc,
        ),
        max_soc,
    )
    return resolve_bess_terminal_soc_target_kwh(
        policy=getattr(asset, "bess_terminal_soc_policy", ""),
        initial_soc_kwh=initial_soc,
        configured_target_kwh=float(
            getattr(asset, "bess_terminal_soc_target_kwh", 0.0) or 0.0
        ),
        terminal_soc_floor_kwh=terminal_floor,
        maximum_soc_kwh=max_soc,
    )


def _complete_bess_soc_flow_context(problem: Any, flow_ctx: Dict[str, Any]) -> Dict[str, Any]:
    assets = dict(getattr(problem, "depot_energy_assets", {}) or {})
    if not assets:
        return flow_ctx

    completed = dict(flow_ctx)
    bess_soc = _normalize_depot_slot_mapping(completed.get("bess_soc_kwh_by_depot_slot", {}))
    bess_soc_start = _normalize_depot_slot_mapping(completed.get("bess_soc_start_kwh_by_depot_slot", {}))
    bess_soc_end = _normalize_depot_slot_mapping(completed.get("bess_soc_end_kwh_by_depot_slot", {}))
    slot_count = len(list(getattr(problem, "price_slots", ()) or ()))

    flow_keys = (
        "grid_to_bus_kwh_by_depot_slot",
        "pv_to_bus_kwh_by_depot_slot",
        "bess_to_bus_kwh_by_depot_slot",
        "pv_to_bess_kwh_by_depot_slot",
        "grid_to_bess_kwh_by_depot_slot",
        "pv_curtail_kwh_by_depot_slot",
        "contract_over_limit_kwh_by_depot_slot",
    )
    for asset in assets.values():
        slot_count = max(slot_count, len(tuple(getattr(asset, "pv_generation_kwh_by_slot", ()) or ())))
    for key in flow_keys:
        for slot_map in dict(completed.get(key) or {}).values():
            slot_count = max(slot_count, max((int(slot_idx) + 1 for slot_idx in dict(slot_map or {}).keys()), default=slot_count))
    for mapping in (bess_soc, bess_soc_start, bess_soc_end):
        for slot_map in mapping.values():
            slot_count = max(slot_count, max((int(slot_idx) + 1 for slot_idx in dict(slot_map or {}).keys()), default=slot_count))

    for depot_id, asset in assets.items():
        depot_key = str(depot_id or "")
        if not depot_key or not bool(getattr(asset, "bess_enabled", False)):
            continue
        bounds = _bess_soc_bounds_for_asset(asset)
        if bounds is None:
            continue
        min_soc, max_soc, initial_soc = bounds
        charge_eff = min(max(float(getattr(asset, "bess_charge_efficiency", 0.95) or 0.95), 1.0e-9), 1.0)
        discharge_eff = min(max(float(getattr(asset, "bess_discharge_efficiency", 0.95) or 0.95), 1.0e-9), 1.0)
        slot_indices = set(range(slot_count))
        for key in flow_keys:
            slot_indices.update(int(slot_idx) for slot_idx in dict((completed.get(key) or {}).get(depot_key, {}) or {}).keys())
        slot_indices.update(int(slot_idx) for slot_idx in dict(bess_soc.get(depot_key, {}) or {}).keys())
        slot_indices.update(int(slot_idx) for slot_idx in dict(bess_soc_start.get(depot_key, {}) or {}).keys())
        slot_indices.update(int(slot_idx) for slot_idx in dict(bess_soc_end.get(depot_key, {}) or {}).keys())
        if not slot_indices:
            continue

        soc = initial_soc
        start_map = bess_soc_start.setdefault(depot_key, {})
        end_map = bess_soc_end.setdefault(depot_key, {})
        soc_map = bess_soc.setdefault(depot_key, {})
        for slot_idx in sorted(slot_indices):
            explicit_start = start_map.get(slot_idx)
            start = min(max(float(explicit_start if explicit_start is not None else soc), min_soc), max_soc)
            explicit_end = end_map.get(slot_idx)
            if explicit_end is None:
                explicit_end = soc_map.get(slot_idx)
            if explicit_end is None:
                pv_to_bess = max(float(((completed.get("pv_to_bess_kwh_by_depot_slot") or {}).get(depot_key, {}) or {}).get(slot_idx, 0.0) or 0.0), 0.0)
                grid_to_bess = max(float(((completed.get("grid_to_bess_kwh_by_depot_slot") or {}).get(depot_key, {}) or {}).get(slot_idx, 0.0) or 0.0), 0.0)
                bess_to_bus = max(float(((completed.get("bess_to_bus_kwh_by_depot_slot") or {}).get(depot_key, {}) or {}).get(slot_idx, 0.0) or 0.0), 0.0)
                end = start + ((pv_to_bess + grid_to_bess) * charge_eff) - (bess_to_bus / discharge_eff)
            else:
                end = float(explicit_end or 0.0)
            end = min(max(end, min_soc), max_soc)
            start_map[slot_idx] = start
            end_map[slot_idx] = end
            soc_map[slot_idx] = end
            soc = end

    completed["bess_soc_kwh_by_depot_slot"] = bess_soc
    completed["bess_soc_start_kwh_by_depot_slot"] = bess_soc_start
    completed["bess_soc_end_kwh_by_depot_slot"] = bess_soc_end or bess_soc
    completed["bess_soc_kwh_semantics"] = "slot_end"
    return completed


def _canonical_energy_flow_context_from_snapshot(problem, plan, preserved_context) -> Dict[str, Any]:
    timestep_min = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
    timestep_h = timestep_min / 60.0

    raw_grid_to_bus = _normalize_depot_slot_mapping(
        preserved_context.get("grid_to_bus_kwh_by_depot_slot", {})
    )
    raw_pv_to_bus = _normalize_depot_slot_mapping(
        preserved_context.get("pv_to_bus_kwh_by_depot_slot", {})
    )
    raw_bess_to_bus = _normalize_depot_slot_mapping(
        preserved_context.get("bess_to_bus_kwh_by_depot_slot", {})
    )
    raw_pv_to_bess = _normalize_depot_slot_mapping(
        preserved_context.get("pv_to_bess_kwh_by_depot_slot", {})
    )
    raw_grid_to_bess = _normalize_depot_slot_mapping(
        preserved_context.get("grid_to_bess_kwh_by_depot_slot", {})
    )
    raw_pv_curtail = _normalize_depot_slot_mapping(
        preserved_context.get("pv_curtail_kwh_by_depot_slot", {})
    )
    raw_bess_soc = _normalize_depot_slot_mapping(
        preserved_context.get("bess_soc_kwh_by_depot_slot", {})
    )
    raw_bess_soc_start = _normalize_depot_slot_mapping(
        preserved_context.get("bess_soc_start_kwh_by_depot_slot", {})
    )
    raw_bess_soc_end = _normalize_depot_slot_mapping(
        preserved_context.get("bess_soc_end_kwh_by_depot_slot", {})
    )
    raw_contract_over_limit = _normalize_depot_slot_mapping(
        preserved_context.get("contract_over_limit_kwh_by_depot_slot", {})
    )

    preserved_slot_signatures = {
        tuple(item)
        for item in list(preserved_context.get("charging_slot_signatures") or [])
        if isinstance(item, (list, tuple)) and len(item) >= 4
    }

    derived_grid_to_bus: Dict[str, Dict[int, float]] = {}
    derived_pv_to_bus: Dict[str, Dict[int, float]] = {}
    derived_bess_to_bus: Dict[str, Dict[int, float]] = {}
    derived_depots: set[str] = set()
    for charging_slot in list(getattr(plan, "charging_slots", ()) or ()):
        if _canonical_charging_slot_signature(charging_slot) in preserved_slot_signatures:
            continue
        charge_kw = max(float(getattr(charging_slot, "charge_kw", 0.0) or 0.0), 0.0)
        discharge_kw = max(float(getattr(charging_slot, "discharge_kw", 0.0) or 0.0), 0.0)
        net_charge_kwh = max(charge_kw - discharge_kw, 0.0) * timestep_h
        if net_charge_kwh <= 0.0:
            continue
        source, depot_id = _canonical_charging_source_and_depot(problem, charging_slot)
        if source == "pv":
            target = derived_pv_to_bus
        elif source == "bess":
            target = derived_bess_to_bus
        else:
            target = derived_grid_to_bus
        slot_map = target.setdefault(str(depot_id), {})
        slot_idx = int(getattr(charging_slot, "slot_index", 0) or 0)
        slot_map[slot_idx] = slot_map.get(slot_idx, 0.0) + net_charge_kwh
        derived_depots.add(str(depot_id))

    effective_grid_to_bus = _merge_depot_slot_flow_maps(raw_grid_to_bus, derived_grid_to_bus)
    effective_pv_to_bus = _merge_depot_slot_flow_maps(raw_pv_to_bus, derived_pv_to_bus)
    effective_bess_to_bus = _merge_depot_slot_flow_maps(raw_bess_to_bus, derived_bess_to_bus)

    depot_limit_kw = {
        str(getattr(depot, "depot_id", "") or ""): float(getattr(depot, "import_limit_kw", 0.0) or 0.0)
        for depot in list(getattr(problem, "depots", ()) or ())
        if str(getattr(depot, "depot_id", "") or "")
    }
    depot_ids = set(depot_limit_kw.keys())
    for mapping in (
        effective_grid_to_bus,
        effective_pv_to_bus,
        effective_bess_to_bus,
        raw_pv_to_bess,
        raw_grid_to_bess,
        raw_pv_curtail,
        raw_bess_soc,
        raw_bess_soc_start,
        raw_bess_soc_end,
        raw_contract_over_limit,
    ):
        depot_ids.update(mapping.keys())
    depot_ids.update(
        str(getattr(vehicle, "home_depot_id", "") or "")
        for vehicle in list(getattr(problem, "vehicles", ()) or ())
        if str(getattr(vehicle, "home_depot_id", "") or "")
    )
    depot_ids.update(str(key) for key in dict(getattr(problem, "depot_energy_assets", {}) or {}).keys())

    pv_generation_kwh_by_depot_slot: Dict[str, Dict[int, float]] = {}
    for depot_id, asset in dict(getattr(problem, "depot_energy_assets", {}) or {}).items():
        generation = {}
        for slot_idx, value in enumerate(list(getattr(asset, "pv_generation_kwh_by_slot", ()) or ())):
            generation[int(slot_idx)] = max(float(value or 0.0), 0.0)
        pv_generation_kwh_by_depot_slot[str(depot_id)] = generation

    price_by_slot = {
        int(getattr(slot, "slot_index", 0) or 0): float(getattr(slot, "grid_buy_yen_per_kwh", 0.0) or 0.0)
        for slot in list(getattr(problem, "price_slots", ()) or ())
    }
    demand_flag_by_slot = {
        int(getattr(slot, "slot_index", 0) or 0): bool(float(getattr(slot, "demand_charge_weight", 0.0) or 0.0) > 0.0)
        for slot in list(getattr(problem, "price_slots", ()) or ())
    }

    derived_from_charging_slots = any(
        _mapping_has_positive_flow(mapping)
        for mapping in (derived_grid_to_bus, derived_pv_to_bus, derived_bess_to_bus)
    )
    provenance_note = str(preserved_context.get("source_provenance_note") or "").strip()
    if not provenance_note:
        provenance_note = (
            "Preserved per-source depot/slot energy-flow maps from the pre-postsolve plan; added postsolve charging slots were derived from the current charging slots."
            if derived_from_charging_slots
            else "Preserved per-source depot/slot energy-flow maps are present in the assignment plan."
        )

    flow_ctx = {
        "timestep_min": timestep_min,
        "timestep_h": timestep_h,
        "depot_ids": sorted(item for item in depot_ids if item),
        "grid_to_bus_kwh_by_depot_slot": effective_grid_to_bus,
        "pv_to_bus_kwh_by_depot_slot": effective_pv_to_bus,
        "bess_to_bus_kwh_by_depot_slot": effective_bess_to_bus,
        "pv_to_bess_kwh_by_depot_slot": raw_pv_to_bess,
        "grid_to_bess_kwh_by_depot_slot": raw_grid_to_bess,
        "pv_curtail_kwh_by_depot_slot": raw_pv_curtail,
        "bess_soc_kwh_by_depot_slot": raw_bess_soc,
        "bess_soc_start_kwh_by_depot_slot": raw_bess_soc_start,
        "bess_soc_end_kwh_by_depot_slot": raw_bess_soc_end or raw_bess_soc,
        "contract_over_limit_kwh_by_depot_slot": raw_contract_over_limit,
        "pv_generation_kwh_by_depot_slot": pv_generation_kwh_by_depot_slot,
        "bess_terminal_soc_target_kwh_by_depot": dict(
            preserved_context.get("bess_terminal_soc_target_kwh_by_depot") or {}
        ),
        "bess_terminal_soc_deviation_kwh_by_depot": dict(
            preserved_context.get("bess_terminal_soc_deviation_kwh_by_depot") or {}
        ),
        "depot_limit_kw": depot_limit_kw,
        "price_by_slot": price_by_slot,
        "demand_flag_by_slot": demand_flag_by_slot,
        "source_provenance_exact": bool(preserved_context.get("source_provenance_exact")) and not derived_from_charging_slots,
        "source_provenance_note": provenance_note,
        "derived_from_charging_slots": derived_from_charging_slots,
        "derived_depots": sorted(derived_depots),
    }
    return _complete_bess_soc_flow_context(problem, flow_ctx)


def _canonical_energy_flow_context(problem, plan) -> Dict[str, Any]:
    metadata = dict(getattr(plan, "metadata", {}) or {})
    preserved_context = dict(metadata.get("canonical_source_flow_context") or {})
    if preserved_context:
        return _canonical_energy_flow_context_from_snapshot(problem, plan, preserved_context)

    timestep_min = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
    timestep_h = timestep_min / 60.0

    raw_grid_to_bus = _normalize_depot_slot_mapping(getattr(plan, "grid_to_bus_kwh_by_depot_slot", {}))
    raw_pv_to_bus = _normalize_depot_slot_mapping(getattr(plan, "pv_to_bus_kwh_by_depot_slot", {}))
    raw_bess_to_bus = _normalize_depot_slot_mapping(getattr(plan, "bess_to_bus_kwh_by_depot_slot", {}))
    raw_pv_to_bess = _normalize_depot_slot_mapping(getattr(plan, "pv_to_bess_kwh_by_depot_slot", {}))
    raw_grid_to_bess = _normalize_depot_slot_mapping(getattr(plan, "grid_to_bess_kwh_by_depot_slot", {}))
    raw_pv_curtail = _normalize_depot_slot_mapping(getattr(plan, "pv_curtail_kwh_by_depot_slot", {}))
    raw_bess_soc = _normalize_depot_slot_mapping(getattr(plan, "bess_soc_kwh_by_depot_slot", {}))
    raw_bess_soc_start = _normalize_depot_slot_mapping(metadata.get("bess_soc_start_kwh_by_depot_slot", {}))
    raw_bess_soc_end = _normalize_depot_slot_mapping(metadata.get("bess_soc_end_kwh_by_depot_slot", {}))
    raw_contract_over_limit = _normalize_depot_slot_mapping(getattr(plan, "contract_over_limit_kwh_by_depot_slot", {}))

    derived_grid_to_bus: Dict[str, Dict[int, float]] = {}
    derived_pv_to_bus: Dict[str, Dict[int, float]] = {}
    derived_bess_to_bus: Dict[str, Dict[int, float]] = {}
    derived_depots: set[str] = set()
    for charging_slot in list(getattr(plan, "charging_slots", ()) or ()):
        charge_kw = max(float(getattr(charging_slot, "charge_kw", 0.0) or 0.0), 0.0)
        discharge_kw = max(float(getattr(charging_slot, "discharge_kw", 0.0) or 0.0), 0.0)
        net_charge_kwh = max(charge_kw - discharge_kw, 0.0) * timestep_h
        if net_charge_kwh <= 0.0:
            continue
        source, depot_id = _canonical_charging_source_and_depot(problem, charging_slot)
        if source == "pv":
            target = derived_pv_to_bus
        elif source == "bess":
            target = derived_bess_to_bus
        else:
            target = derived_grid_to_bus
        slot_map = target.setdefault(str(depot_id), {})
        slot_idx = int(getattr(charging_slot, "slot_index", 0) or 0)
        slot_map[slot_idx] = slot_map.get(slot_idx, 0.0) + net_charge_kwh
        derived_depots.add(str(depot_id))

    effective_grid_to_bus = dict(raw_grid_to_bus)
    effective_pv_to_bus = dict(raw_pv_to_bus)
    effective_bess_to_bus = dict(raw_bess_to_bus)
    for depot_id, slot_map in derived_grid_to_bus.items():
        if not _depot_mapping_has_positive_flow(raw_grid_to_bus, depot_id):
            effective_grid_to_bus[depot_id] = dict(slot_map)
    for depot_id, slot_map in derived_pv_to_bus.items():
        if not _depot_mapping_has_positive_flow(raw_pv_to_bus, depot_id):
            effective_pv_to_bus[depot_id] = dict(slot_map)
    for depot_id, slot_map in derived_bess_to_bus.items():
        if not _depot_mapping_has_positive_flow(raw_bess_to_bus, depot_id):
            effective_bess_to_bus[depot_id] = dict(slot_map)

    depot_limit_kw = {
        str(getattr(depot, "depot_id", "") or ""): float(getattr(depot, "import_limit_kw", 0.0) or 0.0)
        for depot in list(getattr(problem, "depots", ()) or ())
        if str(getattr(depot, "depot_id", "") or "")
    }
    depot_ids = set(depot_limit_kw.keys())
    for mapping in (
        effective_grid_to_bus,
        effective_pv_to_bus,
        effective_bess_to_bus,
        raw_pv_to_bess,
        raw_grid_to_bess,
        raw_pv_curtail,
        raw_bess_soc,
        raw_bess_soc_start,
        raw_bess_soc_end,
        raw_contract_over_limit,
    ):
        depot_ids.update(mapping.keys())
    depot_ids.update(
        str(getattr(vehicle, "home_depot_id", "") or "")
        for vehicle in list(getattr(problem, "vehicles", ()) or ())
        if str(getattr(vehicle, "home_depot_id", "") or "")
    )
    depot_ids.update(str(key) for key in dict(getattr(problem, "depot_energy_assets", {}) or {}).keys())

    pv_generation_kwh_by_depot_slot: Dict[str, Dict[int, float]] = {}
    for depot_id, asset in dict(getattr(problem, "depot_energy_assets", {}) or {}).items():
        generation = {}
        for slot_idx, value in enumerate(list(getattr(asset, "pv_generation_kwh_by_slot", ()) or ())):
            generation[int(slot_idx)] = max(float(value or 0.0), 0.0)
        pv_generation_kwh_by_depot_slot[str(depot_id)] = generation

    price_by_slot = {
        int(getattr(slot, "slot_index", 0) or 0): float(getattr(slot, "grid_buy_yen_per_kwh", 0.0) or 0.0)
        for slot in list(getattr(problem, "price_slots", ()) or ())
    }
    demand_flag_by_slot = {
        int(getattr(slot, "slot_index", 0) or 0): bool(float(getattr(slot, "demand_charge_weight", 0.0) or 0.0) > 0.0)
        for slot in list(getattr(problem, "price_slots", ()) or ())
    }

    explicit_source_split = any(
        _mapping_has_positive_flow(mapping)
        for mapping in (
            raw_grid_to_bus,
            raw_pv_to_bus,
            raw_bess_to_bus,
            raw_pv_to_bess,
            raw_grid_to_bess,
            raw_pv_curtail,
        )
    )
    derived_from_charging_slots = any(
        _mapping_has_positive_flow(mapping)
        for mapping in (derived_grid_to_bus, derived_pv_to_bus, derived_bess_to_bus)
    ) and not explicit_source_split

    if explicit_source_split:
        provenance_note = "Explicit per-source depot/slot energy-flow maps are present in the assignment plan."
    elif derived_from_charging_slots:
        provenance_note = (
            "Per-source depot/slot energy-flow maps are not present; grid-origin charging was derived from charging slots. "
            "PV/BESS source split remains zero unless the plan encodes it explicitly."
        )
    else:
        provenance_note = "No charging energy flow was recorded for this plan."

    flow_ctx = {
        "timestep_min": timestep_min,
        "timestep_h": timestep_h,
        "depot_ids": sorted(item for item in depot_ids if item),
        "grid_to_bus_kwh_by_depot_slot": effective_grid_to_bus,
        "pv_to_bus_kwh_by_depot_slot": effective_pv_to_bus,
        "bess_to_bus_kwh_by_depot_slot": effective_bess_to_bus,
        "pv_to_bess_kwh_by_depot_slot": raw_pv_to_bess,
        "grid_to_bess_kwh_by_depot_slot": raw_grid_to_bess,
        "pv_curtail_kwh_by_depot_slot": raw_pv_curtail,
        "bess_soc_kwh_by_depot_slot": raw_bess_soc,
        "bess_soc_start_kwh_by_depot_slot": raw_bess_soc_start,
        "bess_soc_end_kwh_by_depot_slot": raw_bess_soc_end or raw_bess_soc,
        "contract_over_limit_kwh_by_depot_slot": raw_contract_over_limit,
        "pv_generation_kwh_by_depot_slot": pv_generation_kwh_by_depot_slot,
        "bess_terminal_soc_target_kwh_by_depot": dict(
            metadata.get("bess_terminal_soc_target_kwh_by_depot") or {}
        ),
        "bess_terminal_soc_deviation_kwh_by_depot": dict(
            metadata.get("bess_terminal_soc_deviation_kwh_by_depot") or {}
        ),
        "depot_limit_kw": depot_limit_kw,
        "price_by_slot": price_by_slot,
        "demand_flag_by_slot": demand_flag_by_slot,
        "source_provenance_exact": not derived_from_charging_slots,
        "source_provenance_note": provenance_note,
        "derived_from_charging_slots": derived_from_charging_slots,
        "derived_depots": sorted(derived_depots),
    }
    return _complete_bess_soc_flow_context(problem, flow_ctx)


def _canonical_charging_output_payload(problem, engine_result) -> Dict[str, Any]:
    plan = engine_result.plan
    flow_ctx = _canonical_energy_flow_context(problem, plan)
    timestep_h = float(flow_ctx["timestep_h"] or 1.0)
    breakdown = dict(engine_result.cost_breakdown or {})
    plan_metadata = dict(getattr(plan, "metadata", {}) or {})
    solver_metadata = dict(engine_result.solver_metadata or {})
    # The source-flow variables are exact at depot × time-slot scope.  Unless
    # the solver carried vehicle/source/slot variables, a vehicle ledger is a
    # transparent post-allocation rather than another exact decision trace.
    depot_source_provenance_exact = bool(flow_ctx["source_provenance_exact"])
    vehicle_source_provenance_exact = bool(
        depot_source_provenance_exact
        and plan_metadata.get("vehicle_source_provenance_exact", False)
    )
    vehicle_source_allocation_method = (
        "solver_native"
        if vehicle_source_provenance_exact
        else "proportional_by_depot_timestep"
    )
    vehicle_source_precision_note = (
        "Vehicle source split is an exact MILP vehicle/source/slot decision trace."
        if vehicle_source_provenance_exact
        else (
            "Vehicle source split is inferred by proportional allocation of exact "
            "depot/time-slot source totals; it is not solver-native."
        )
    )
    penalty_enabled = bool(
        plan_metadata.get(
            "enable_contract_overage_penalty",
            solver_metadata.get("enable_contract_overage_penalty", True),
        )
    )
    raw_penalty_yen_per_kwh = plan_metadata.get(
        "contract_overage_penalty_yen_per_kwh",
        solver_metadata.get("contract_overage_penalty_yen_per_kwh", 0.0),
    )
    penalty_yen_per_kwh = float(raw_penalty_yen_per_kwh or 0.0)
    contract_overage_policy = "soft_penalty" if penalty_enabled and penalty_yen_per_kwh > 0.0 else "warning_only"

    rows: List[Dict[str, Any]] = []
    per_depot: List[Dict[str, Any]] = []
    all_slot_indices: set[int] = set()
    overall_peak_grid_kw = 0.0
    overall_peak_total_charge_kw = 0.0

    for depot_id in list(flow_ctx["depot_ids"]):
        depot_slots = set()
        for key in (
            "grid_to_bus_kwh_by_depot_slot",
            "pv_to_bus_kwh_by_depot_slot",
            "bess_to_bus_kwh_by_depot_slot",
            "pv_to_bess_kwh_by_depot_slot",
            "grid_to_bess_kwh_by_depot_slot",
            "pv_curtail_kwh_by_depot_slot",
            "bess_soc_kwh_by_depot_slot",
            "bess_soc_start_kwh_by_depot_slot",
            "bess_soc_end_kwh_by_depot_slot",
            "contract_over_limit_kwh_by_depot_slot",
            "pv_generation_kwh_by_depot_slot",
        ):
            depot_slots.update(dict(flow_ctx.get(key) or {}).get(depot_id, {}).keys())
        peak_grid_kw = 0.0
        peak_total_charge_kw = 0.0
        peak_contract_over_kw = 0.0
        grid_to_bus_total = 0.0
        pv_to_bus_total = 0.0
        bess_to_bus_total = 0.0
        pv_to_bess_total = 0.0
        grid_to_bess_total = 0.0
        pv_curtail_total = 0.0
        contract_over_total = 0.0
        contract_slot_count = 0

        for slot_idx in sorted(int(idx) for idx in depot_slots):
            all_slot_indices.add(slot_idx)
            grid_to_bus = float((flow_ctx["grid_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_to_bus = float((flow_ctx["pv_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            bess_to_bus = float((flow_ctx["bess_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_to_bess = float((flow_ctx["pv_to_bess_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            grid_to_bess = float((flow_ctx["grid_to_bess_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            raw_pv_curtail = float((flow_ctx["pv_curtail_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            bess_soc = float((flow_ctx["bess_soc_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            bess_soc_start = float((flow_ctx["bess_soc_start_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            bess_soc_end = float((flow_ctx["bess_soc_end_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, bess_soc) or 0.0)
            if bess_soc <= 0.0 and bess_soc_end > 0.0:
                bess_soc = bess_soc_end
            contract_over_limit_kwh = float((flow_ctx["contract_over_limit_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_generation_kwh = float((flow_ctx["pv_generation_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            if pv_generation_kwh > 0.0:
                pv_curtail = compute_pv_curtail_kwh(pv_generation_kwh, pv_to_bus, pv_to_bess)
            else:
                pv_curtail = max(raw_pv_curtail, 0.0)
            pv_balance_residual_kwh = pv_generation_kwh - pv_to_bus - pv_to_bess - pv_curtail
            contract_limit_kw = float((flow_ctx["depot_limit_kw"].get(depot_id, 0.0)) or 0.0)
            grid_import_total_kwh = grid_to_bus + grid_to_bess
            grid_import_for_contract_kwh = grid_import_total_kwh
            total_bus_charge_kwh = grid_to_bus + pv_to_bus + bess_to_bus
            total_bess_charge_kwh = pv_to_bess + grid_to_bess
            grid_import_kw = grid_import_total_kwh / timestep_h if timestep_h > 0.0 else 0.0
            total_charge_kw = total_bus_charge_kwh / timestep_h if timestep_h > 0.0 else 0.0
            if contract_limit_kw > 0.0:
                contract_limit_kwh = contract_limit_kw * timestep_h
                contract_over_limit_kwh = max(
                    contract_over_limit_kwh,
                    grid_import_total_kwh - contract_limit_kwh,
                    0.0,
                )
            contract_over_limit_kw = contract_over_limit_kwh / timestep_h if timestep_h > 0.0 else 0.0
            peak_grid_kw = max(peak_grid_kw, grid_import_kw)
            peak_total_charge_kw = max(peak_total_charge_kw, total_charge_kw)
            peak_contract_over_kw = max(peak_contract_over_kw, contract_over_limit_kw)
            overall_peak_grid_kw = max(overall_peak_grid_kw, grid_import_kw)
            overall_peak_total_charge_kw = max(overall_peak_total_charge_kw, total_charge_kw)
            grid_to_bus_total += grid_to_bus
            pv_to_bus_total += pv_to_bus
            bess_to_bus_total += bess_to_bus
            pv_to_bess_total += pv_to_bess
            grid_to_bess_total += grid_to_bess
            pv_curtail_total += pv_curtail
            contract_over_total += contract_over_limit_kwh
            if contract_over_limit_kwh > 1.0e-9:
                contract_slot_count += 1
            rows.append(
                {
                    "depot_id": depot_id,
                    "slot_index": slot_idx,
                    "grid_to_bus_kwh": grid_to_bus,
                    "pv_to_bus_kwh": pv_to_bus,
                    "bess_to_bus_kwh": bess_to_bus,
                    "pv_to_bess_kwh": pv_to_bess,
                    "grid_to_bess_kwh": grid_to_bess,
                    "pv_curtail_kwh": pv_curtail,
                    "pv_curtail_raw_kwh": raw_pv_curtail,
                    "pv_generation_kwh": pv_generation_kwh,
                    "pv_balance_residual_kwh": pv_balance_residual_kwh,
                    "bess_soc_kwh": bess_soc,
                    "bess_soc_start_kwh": bess_soc_start,
                    "bess_soc_end_kwh": bess_soc_end,
                    "grid_import_total_kwh": grid_import_total_kwh,
                    "grid_import_for_contract_kwh": grid_import_for_contract_kwh,
                    "grid_import_kw": grid_import_kw,
                    "grid_import_for_contract_kw": grid_import_kw,
                    "bus_charge_from_grid_kwh": grid_to_bus,
                    "bus_charge_from_bess_kwh": bess_to_bus,
                    "total_bus_charge_kwh": total_bus_charge_kwh,
                    "total_bess_charge_kwh": total_bess_charge_kwh,
                    "total_charge_kw": total_charge_kw,
                    "contract_limit_kw": contract_limit_kw,
                    "contract_over_limit_kwh": contract_over_limit_kwh,
                    "contract_over_limit_kw": contract_over_limit_kw,
                    "contract_limit_exceeded": contract_over_limit_kwh > 1.0e-9,
                    "energy_price_yen_per_kwh": float((flow_ctx["price_by_slot"].get(slot_idx, 0.0)) or 0.0),
                    "demand_charge_window_flag": bool(flow_ctx["demand_flag_by_slot"].get(slot_idx, False)),
                    # Legacy field retains its depot/time-slot meaning.
                    "source_provenance_exact": depot_source_provenance_exact,
                    "depot_source_provenance_exact": depot_source_provenance_exact,
                }
            )

        contract_overage_cost = contract_over_total * penalty_yen_per_kwh if penalty_enabled else 0.0
        asset = dict(getattr(problem, "depot_energy_assets", {}) or {}).get(depot_id)
        bess_initial_soc = float(getattr(asset, "bess_initial_soc_kwh", 0.0) or 0.0) if asset is not None else 0.0
        bess_terminal_min = float(getattr(asset, "bess_terminal_soc_min_kwh", 0.0) or 0.0) if asset is not None else 0.0
        raw_terminal_target = (flow_ctx.get("bess_terminal_soc_target_kwh_by_depot", {}) or {}).get(depot_id)
        if raw_terminal_target is None and asset is not None:
            raw_terminal_target = _resolved_bess_terminal_target_kwh(asset)
        bess_terminal_target = float(raw_terminal_target or 0.0) if asset is not None else 0.0
        depot_bess_soc_map = dict(flow_ctx["bess_soc_kwh_by_depot_slot"].get(depot_id, {}) or {})
        depot_bess_soc_end_map = dict(flow_ctx["bess_soc_end_kwh_by_depot_slot"].get(depot_id, {}) or {})
        bess_final_soc = (
            float(depot_bess_soc_end_map[max(depot_bess_soc_end_map.keys())])
            if depot_bess_soc_end_map
            else float(depot_bess_soc_map[max(depot_bess_soc_map.keys())])
            if depot_bess_soc_map
            else bess_initial_soc
        )
        bess_terminal_deviation = abs(bess_final_soc - bess_terminal_target) if asset is not None and bess_terminal_target > 0.0 else 0.0
        pv_generation_total = sum(
            float(value or 0.0)
            for value in dict(flow_ctx["pv_generation_kwh_by_depot_slot"].get(depot_id, {}) or {}).values()
        )
        pv_used_total = pv_to_bus_total + pv_to_bess_total
        per_depot.append(
            {
                "depot_id": depot_id,
                # Legacy field retains its depot/time-slot meaning.
                "source_provenance_exact": depot_source_provenance_exact and depot_id not in set(flow_ctx["derived_depots"]),
                "depot_source_provenance_exact": depot_source_provenance_exact and depot_id not in set(flow_ctx["derived_depots"]),
                "grid_to_bus_kwh": grid_to_bus_total,
                "pv_to_bus_kwh": pv_to_bus_total,
                "bess_to_bus_kwh": bess_to_bus_total,
                "pv_to_bess_kwh": pv_to_bess_total,
                "grid_to_bess_kwh": grid_to_bess_total,
                "pv_curtail_kwh": pv_curtail_total,
                "pv_generation_kwh": pv_generation_total,
                "pv_utilization_rate": compute_pv_utilization_rate(
                    pv_generation_total,
                    pv_to_bus_total,
                    pv_to_bess_total,
                ),
                "grid_import_total_kwh": grid_to_bus_total + grid_to_bess_total,
                "grid_import_for_contract_kwh": grid_to_bus_total + grid_to_bess_total,
                "total_bus_charge_kwh": grid_to_bus_total + pv_to_bus_total + bess_to_bus_total,
                "bus_charge_from_grid_kwh": grid_to_bus_total,
                "bus_charge_from_bess_kwh": bess_to_bus_total,
                "total_bess_charge_kwh": pv_to_bess_total + grid_to_bess_total,
                "bess_initial_soc_kwh": bess_initial_soc,
                "bess_final_soc_kwh": bess_final_soc,
                "bess_terminal_soc_delta_kwh": bess_final_soc - bess_initial_soc,
                "bess_terminal_soc_min_kwh": bess_terminal_min,
                "bess_terminal_soc_target_kwh": bess_terminal_target,
                "bess_terminal_soc_deviation_kwh": bess_terminal_deviation,
                "bess_terminal_soc_violation_kwh": max(bess_terminal_min - bess_final_soc, 0.0),
                "peak_grid_import_kw": peak_grid_kw,
                "peak_total_charge_kw": peak_total_charge_kw,
                "contract_limit_kw": float((flow_ctx["depot_limit_kw"].get(depot_id, 0.0)) or 0.0),
                "contract_over_limit_kwh": contract_over_total,
                "contract_over_limit_kw_peak": peak_contract_over_kw,
                "contract_over_limit_slot_count": contract_slot_count,
                "contract_limit_exceeded": contract_over_total > 1.0e-9,
                "contract_overage_penalty_enabled": penalty_enabled,
                "contract_overage_penalty_yen_per_kwh": penalty_yen_per_kwh,
                "contract_overage_cost_jpy": contract_overage_cost,
                "contract_overage_policy": contract_overage_policy,
            }
        )

    overall_grid_import_total_kwh = sum(float(row["grid_import_total_kwh"]) for row in per_depot)
    overall_contract_over_kwh = sum(float(row["contract_over_limit_kwh"]) for row in per_depot)
    expected_contract_over_cost = overall_contract_over_kwh * penalty_yen_per_kwh if penalty_enabled else 0.0
    overall_contract_over_cost = float(breakdown.get("contract_overage_cost", expected_contract_over_cost) or 0.0)
    if abs(overall_contract_over_cost - expected_contract_over_cost) <= 1.0e-6:
        overall_contract_over_cost = expected_contract_over_cost
    fuel_cost_jpy = float(breakdown.get("fuel_cost", 0.0) or 0.0)
    aggregate_energy_cost_jpy = float(breakdown.get("energy_cost", 0.0) or 0.0)
    if breakdown.get("electricity_cost") is not None:
        electricity_cost_jpy = float(breakdown.get("electricity_cost") or 0.0)
    elif breakdown.get("electricity_cost_final") is not None:
        electricity_cost_jpy = float(breakdown.get("electricity_cost_final") or 0.0)
    else:
        electricity_cost_jpy = (
            max(aggregate_energy_cost_jpy - fuel_cost_jpy, 0.0)
            if fuel_cost_jpy > 0.0
            else aggregate_energy_cost_jpy
        )
    overall_by_slot_grid_peak = 0.0
    overall_by_slot_charge_peak = 0.0
    for slot_idx in sorted(all_slot_indices):
        total_grid_import_kwh = 0.0
        total_charge_kwh = 0.0
        for depot_id in list(flow_ctx["depot_ids"]):
            total_grid_import_kwh += float((flow_ctx["grid_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            total_grid_import_kwh += float((flow_ctx["grid_to_bess_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            total_charge_kwh += float((flow_ctx["grid_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            total_charge_kwh += float((flow_ctx["pv_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            total_charge_kwh += float((flow_ctx["bess_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
        overall_by_slot_grid_peak = max(overall_by_slot_grid_peak, total_grid_import_kwh / timestep_h if timestep_h > 0.0 else 0.0)
        overall_by_slot_charge_peak = max(overall_by_slot_charge_peak, total_charge_kwh / timestep_h if timestep_h > 0.0 else 0.0)

    return {
        "summary": {
            "timestep_min": int(flow_ctx["timestep_min"]),
            # ``source_provenance_exact`` is retained for older consumers and
            # explicitly means depot × time-slot source-flow exactness.
            "source_provenance_exact": depot_source_provenance_exact,
            "source_provenance_scope": "depot_timestep",
            "depot_source_provenance_exact": depot_source_provenance_exact,
            "vehicle_source_provenance_exact": vehicle_source_provenance_exact,
            "vehicle_source_allocation_method": vehicle_source_allocation_method,
            "vehicle_source_precision_note": vehicle_source_precision_note,
            "source_provenance_note": str(flow_ctx["source_provenance_note"] or ""),
            "depots": per_depot,
            "totals": {
                "source_provenance_exact": depot_source_provenance_exact,
                "source_provenance_scope": "depot_timestep",
                "depot_source_provenance_exact": depot_source_provenance_exact,
                "vehicle_source_provenance_exact": vehicle_source_provenance_exact,
                "vehicle_source_allocation_method": vehicle_source_allocation_method,
                "grid_to_bus_kwh": sum(float(row["grid_to_bus_kwh"]) for row in per_depot),
                "pv_to_bus_kwh": sum(float(row["pv_to_bus_kwh"]) for row in per_depot),
                "bess_to_bus_kwh": sum(float(row["bess_to_bus_kwh"]) for row in per_depot),
                "pv_to_bess_kwh": sum(float(row["pv_to_bess_kwh"]) for row in per_depot),
                "grid_to_bess_kwh": sum(float(row["grid_to_bess_kwh"]) for row in per_depot),
                "pv_curtail_kwh": sum(float(row["pv_curtail_kwh"]) for row in per_depot),
                "pv_generation_kwh": sum(float(row["pv_generation_kwh"]) for row in per_depot),
                "pv_utilization_rate": compute_pv_utilization_rate(
                    sum(float(row["pv_generation_kwh"]) for row in per_depot),
                    sum(float(row["pv_to_bus_kwh"]) for row in per_depot),
                    sum(float(row["pv_to_bess_kwh"]) for row in per_depot),
                ),
                "grid_import_total_kwh": overall_grid_import_total_kwh,
                "grid_import_for_contract_kwh": overall_grid_import_total_kwh,
                "total_bus_charge_kwh": sum(float(row["total_bus_charge_kwh"]) for row in per_depot),
                "bus_charge_from_grid_kwh": sum(float(row["bus_charge_from_grid_kwh"]) for row in per_depot),
                "bus_charge_from_bess_kwh": sum(float(row["bus_charge_from_bess_kwh"]) for row in per_depot),
                "total_bess_charge_kwh": sum(float(row["total_bess_charge_kwh"]) for row in per_depot),
                "bess_initial_soc_kwh": sum(float(row["bess_initial_soc_kwh"]) for row in per_depot),
                "bess_final_soc_kwh": sum(float(row["bess_final_soc_kwh"]) for row in per_depot),
                "bess_terminal_soc_delta_kwh": sum(float(row["bess_terminal_soc_delta_kwh"]) for row in per_depot),
                "bess_terminal_soc_target_kwh": sum(float(row["bess_terminal_soc_target_kwh"]) for row in per_depot),
                "peak_grid_import_kw_any_depot": overall_peak_grid_kw,
                "peak_grid_import_kw_all_depots": overall_by_slot_grid_peak,
                "peak_total_charge_kw_any_depot": overall_peak_total_charge_kw,
                "peak_total_charge_kw_all_depots": overall_by_slot_charge_peak,
                "contract_over_limit_kwh": overall_contract_over_kwh,
                "contract_limit_exceeded": overall_contract_over_kwh > 1.0e-9,
                "contract_overage_penalty_enabled": penalty_enabled,
                "contract_overage_penalty_yen_per_kwh": penalty_yen_per_kwh,
                "contract_overage_cost_jpy": overall_contract_over_cost,
                "contract_overage_policy": contract_overage_policy,
                "contract_overage_warning": (
                    "Contract limit exceeded but no overage cost was applied; policy is warning_only."
                    if overall_contract_over_kwh > 1.0e-9 and contract_overage_policy == "warning_only"
                    else ""
                ),
                "bess_terminal_soc_violation_kwh": float(
                    solver_metadata.get(
                        "bess_terminal_soc_violation_kwh",
                        plan_metadata.get("bess_terminal_soc_violation_kwh", 0.0),
                    )
                    or 0.0
                ),
                "bess_terminal_soc_deviation_kwh": float(
                    solver_metadata.get(
                        "bess_terminal_soc_deviation_kwh",
                        plan_metadata.get(
                            "bess_terminal_soc_deviation_kwh",
                            sum(float(row.get("bess_terminal_soc_deviation_kwh", 0.0) or 0.0) for row in per_depot),
                        ),
                    )
                    or 0.0
                ),
                "readiness_topup_policy": {
                    "status": "metadata_only",
                    "warning": False,
                    "bess_terminal_soc_min_enforced": True,
                },
                "demand_charge_cost_jpy": float(breakdown.get("demand_cost", 0.0) or 0.0),
                "grid_purchase_cost_jpy": float(breakdown.get("grid_purchase_cost", 0.0) or 0.0),
                "bess_discharge_cost_jpy": float(breakdown.get("bess_discharge_cost", 0.0) or 0.0),
                "pv_self_consumption_cost_jpy": float(breakdown.get("pv_self_consumption_cost_jpy", 0.0) or 0.0),
                "electricity_cost_jpy": electricity_cost_jpy,
                "objective_value_jpy": float(engine_result.objective_value or 0.0),
                "objective_is_actual_cost": bool(breakdown.get("objective_is_actual_cost", False)),
                "supports_exact_milp": bool((engine_result.solver_metadata or {}).get("supports_exact_milp", False)),
            },
        },
        "rows": rows,
    }


def _canonical_simulation_condition_tables(
    canonical_problem: Optional[Any],
) -> Optional[tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]]:
    """Build condition tables from the canonical model actually solved.

    ``simulation_config`` is a frontend-facing editable document and can omit
    values inherited from a tariff CSV or the selected depot. These output
    tables are evidence artifacts, so their prices and import limits must be
    taken from the already-built canonical problem instead of that stale UI
    payload.
    """

    if canonical_problem is None:
        return None
    price_slots = sorted(
        list(getattr(canonical_problem, "price_slots", ()) or ()),
        key=lambda slot: int(getattr(slot, "slot_index", 0) or 0),
    )
    depots = list(getattr(canonical_problem, "depots", ()) or ())
    if not price_slots or not depots:
        return None

    tou_rows: List[Dict[str, Any]] = []
    contract_rows: List[Dict[str, Any]] = []
    depot_ids: List[str] = []
    metadata = dict(getattr(canonical_problem, "metadata", {}) or {})
    base_load_by_depot_slot = metadata.get("base_load_kw_by_depot_slot")
    base_load_by_slot = metadata.get("base_load_kw_by_slot")

    def canonical_base_load_kw(depot_id: str, slot_index: int) -> Any:
        """Return an explicitly represented base load, never infer one."""

        if isinstance(base_load_by_depot_slot, dict):
            per_depot = base_load_by_depot_slot.get(depot_id)
            if isinstance(per_depot, dict):
                if slot_index in per_depot:
                    return per_depot[slot_index]
                if str(slot_index) in per_depot:
                    return per_depot[str(slot_index)]
        if isinstance(base_load_by_slot, dict):
            if slot_index in base_load_by_slot:
                return base_load_by_slot[slot_index]
            if str(slot_index) in base_load_by_slot:
                return base_load_by_slot[str(slot_index)]
        if isinstance(base_load_by_slot, (list, tuple)) and 0 <= slot_index < len(
            base_load_by_slot
        ):
            return base_load_by_slot[slot_index]
        return None

    for depot in depots:
        depot_id = str(getattr(depot, "depot_id", "") or "").strip()
        if not depot_id:
            continue
        depot_ids.append(depot_id)
        import_limit_kw = float(getattr(depot, "import_limit_kw", 0.0) or 0.0)
        for slot in price_slots:
            slot_index = int(getattr(slot, "slot_index", 0) or 0)
            tou_rows.append(
                {
                    "site_id": depot_id,
                    "time_idx": slot_index,
                    "grid_energy_price_yen_per_kwh": float(
                        getattr(slot, "grid_buy_yen_per_kwh", 0.0) or 0.0
                    ),
                    "sell_back_price_yen_per_kwh": float(
                        getattr(slot, "grid_sell_yen_per_kwh", 0.0) or 0.0
                    ),
                    # A demand-charge weight is not generally a physical base
                    # load. Export a base load only when the canonical model
                    # represents one explicitly; otherwise leave it blank.
                    "base_load_kw": canonical_base_load_kw(depot_id, slot_index),
                    "demand_charge_weight": float(
                        getattr(slot, "demand_charge_weight", 0.0) or 0.0
                    ),
                    "grid_co2_factor_kg_per_kwh": float(
                        getattr(slot, "co2_factor", 0.0) or 0.0
                    ),
                }
            )
        contract_rows.append(
            {
                "site_id": depot_id,
                "site_type": "depot",
                # The canonical model represents one grid-import limit. It is
                # the operative contract limit; no separate transformer value
                # is invented when the input did not provide one.
                "contract_demand_limit_kw": import_limit_kw,
                "grid_import_limit_kw": import_limit_kw,
                "site_transformer_limit_kw": None,
            }
        )
    if not depot_ids:
        return None
    return (
        tou_rows,
        contract_rows,
        {
            "schema_version": "simulation_conditions_provenance_v1",
            "source": "canonical_problem",
            "price_source": "canonical_problem.price_slots",
            "contract_limit_source": "canonical_problem.depots[].import_limit_kw",
            "base_load_source": (
                "canonical_problem.metadata.base_load_kw_by_depot_slot_or_by_slot; "
                "blank_when_not_represented"
            ),
            "demand_charge_weight_source": (
                "canonical_problem.price_slots[].demand_charge_weight"
            ),
            "site_transformer_limit_semantics": "not_modeled_blank_in_csv",
            "depot_ids": sorted(depot_ids),
            "price_slot_count": len(price_slots),
        },
    )


def _rolling_execution_evidence(
    *,
    run_dir: Path,
    solver_metadata: Dict[str, Any],
) -> Dict[str, Any]:
    """Classify rolling from its persisted chain, never day-ahead metadata.

    A Phase 1 solve can carry rolling-shaped metadata without any hour being
    executed.  The accepted state therefore requires a complete, accepted
    chain summary produced by the in-process rolling service.
    """

    rolling_dir = run_dir / "rolling_hourly_chain"
    summary_path = rolling_dir / "rolling_chain_summary.json"
    failure_path = rolling_dir / "rolling_execution_failure.json"
    policy = str(solver_metadata.get("rolling_horizon_policy") or "").strip()
    minutes = solver_metadata.get("rolling_execution_minutes")
    if not summary_path.is_file():
        if failure_path.is_file():
            try:
                failure = json.loads(failure_path.read_text(encoding="utf-8"))
            except Exception as exc:
                failure = {
                    "status": "failed",
                    "reason": "rolling_failure_artifact_unreadable",
                    "error": f"{type(exc).__name__}: {exc}",
                }
            return {
                "status": "executed_not_accepted",
                "rolling_horizon_policy": policy or None,
                "rolling_execution_minutes": minutes,
                "chain_summary_path": None,
                "failure_artifact_path": str(failure_path),
                "rejection_reasons": [
                    str(failure.get("reason") or "rolling_service_failed")
                ],
                "failure": failure,
                "semantics": (
                    "The requested rolling chain did not produce an auditable "
                    "complete chain. The optimization job must fail."
                ),
            }
        return {
            "status": "not_executed",
            "rolling_horizon_policy": policy or None,
            "rolling_execution_minutes": minutes,
            "chain_summary_path": None,
            "semantics": (
                "No persisted hourly chain exists. Day-ahead metadata alone is "
                "not evidence that rolling was executed."
            ),
        }
    try:
        chain = json.loads(summary_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return {
            "status": "executed_not_accepted",
            "rolling_horizon_policy": policy or None,
            "rolling_execution_minutes": minutes,
            "chain_summary_path": str(summary_path),
            "chain_read_error": f"{type(exc).__name__}: {exc}",
            "semantics": "A rolling chain artifact exists but cannot be verified.",
        }
    acceptance_audit = rolling_chain_acceptance_audit(chain)
    checks = dict(acceptance_audit["acceptance_checks"])
    accepted = bool(acceptance_audit["accepted"])
    rejection_reasons = list(chain.get("rejection_reasons") or [])
    rejection_reasons.extend(
        f"missing_required_check:{name}"
        for name in acceptance_audit["missing_required_checks"]
    )
    rejection_reasons.extend(
        f"failed_acceptance_check:{name}"
        for name in acceptance_audit["failing_checks"]
    )
    return {
        "status": "executed_and_accepted" if accepted else "executed_not_accepted",
        "rolling_horizon_policy": (
            chain.get("remaining_day_charging_only_fixed_assignment")
            and "remaining_day_charging_only_fixed_assignment"
            or policy
            or None
        ),
        "rolling_execution_minutes": chain.get(
            "execution_minutes", minutes
        ),
        "chain_summary_path": str(summary_path),
        "chain_accepted": bool(chain.get("chain_accepted")),
        "acceptance_checks": checks,
        "rejection_reasons": sorted(set(rejection_reasons)),
        "semantics": (
            "Rolling evidence is accepted only when the persisted chain summary "
            "and every recorded acceptance check pass."
        ),
    }


def _results_workbook_cell_value(value: Any) -> Any:
    """Return an openpyxl-safe cell value without changing numeric costs.

    Cost breakdowns also carry structured provenance and diagnostic metadata.
    Excel cells cannot represent mappings or sequences directly, so preserve
    those values as deterministic JSON text while keeping scalar accounting
    components numeric for the reconciliation contract.
    """

    if value is None or isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    if isinstance(value, Mapping):
        value = dict(value)
    elif isinstance(value, tuple):
        value = list(value)
    if isinstance(value, (dict, list)):
        return json.dumps(
            value,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
            allow_nan=False,
        )
    raise TypeError(
        "results.xlsx cannot serialize non-scalar report value "
        f"of type {type(value).__name__}"
    )


def _write_results_workbook(
    *,
    run_dir: Path,
    summary: Dict[str, Any],
    cost_rows: List[Dict[str, Any]],
) -> None:
    """Write the standard workbook with scalar-safe report cells."""

    from openpyxl import Workbook

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["key", "value", "unit"])
    summary_sheet.append(
        ["objective_value", _results_workbook_cell_value(summary.get("objective_value")), "JPY"]
    )
    summary_sheet.append(
        [
            "solve_time_seconds",
            _results_workbook_cell_value(summary.get("solve_time_seconds")),
            "s",
        ]
    )
    summary_sheet.append(
        [
            "trip_count_served",
            _results_workbook_cell_value(summary.get("trip_count_served")),
            "trips",
        ]
    )
    summary_sheet.append(
        [
            "trip_count_unserved",
            _results_workbook_cell_value(summary.get("trip_count_unserved")),
            "trips",
        ]
    )

    cost_sheet = workbook.create_sheet("cost_breakdown")
    cost_sheet.append(["key", "value", "unit"])
    for row in cost_rows:
        cost_sheet.append(
            [
                _results_workbook_cell_value(row.get("key")),
                _results_workbook_cell_value(row.get("value")),
                _results_workbook_cell_value(row.get("unit")),
            ]
        )

    workbook.save(run_dir / "results.xlsx")


def _assignment_economic_audit_payload(
    *,
    canonical_problem: Optional[Any],
    optimization_result: Mapping[str, Any],
) -> Dict[str, Any]:
    """Expose Stage 1 energy semantics without inventing vehicle-source flows.

    The two-stage model has solver-native depot/slot source flows, not a
    solver-native attribution of PV or BESS energy to an individual bus.  This
    audit intentionally reports those two evidence levels separately so a
    teacher-facing report cannot turn a proportional allocation into source
    provenance.
    """

    solver_metadata = dict(optimization_result.get("solver_metadata") or {})
    problem_metadata = dict(
        getattr(canonical_problem, "metadata", {}) or {}
    )
    result_summary = dict(optimization_result.get("summary") or {})
    recourse_configuration = dict(
        solver_metadata.get("stage1_time_indexed_energy_recourse_configuration")
        or {}
    )
    recourse_weather = dict(
        solver_metadata.get("stage1_time_indexed_energy_recourse_weather_input")
        or {}
    )
    recourse_result = dict(
        solver_metadata.get("stage1_time_indexed_energy_recourse_result")
        or {}
    )
    cost_breakdown = dict(optimization_result.get("cost_breakdown") or {})
    assignment_energy_coupling_mode = str(
        solver_metadata.get("assignment_energy_coupling_mode")
        or recourse_configuration.get("semantics")
        or "not_available_from_solver_metadata"
    )

    def _finite_nonnegative(value: Any) -> Optional[float]:
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            return None
        return parsed if math.isfinite(parsed) and parsed >= 0.0 else None

    def _sum_numbers(value: Any) -> float:
        if isinstance(value, Mapping):
            return sum(_sum_numbers(item) for item in value.values())
        if isinstance(value, (list, tuple)):
            return sum(_sum_numbers(item) for item in value)
        parsed = _finite_nonnegative(value)
        return float(parsed or 0.0)

    def _single_value(values: list[float]) -> Optional[float]:
        normalized = sorted({round(value, 12) for value in values})
        return normalized[0] if len(normalized) == 1 else None

    electric_types = {"BEV", "PHEV", "FCEV"}
    price_values: list[float] = []
    electric_energy_rates: list[float] = []
    combustion_fuel_rates: list[float] = []
    diesel_price: Optional[float] = None
    if canonical_problem is not None:
        price_values = [
            value
            for slot in tuple(
                getattr(canonical_problem, "price_slots", ()) or ()
            )
            if (
                value := _finite_nonnegative(
                    getattr(slot, "grid_buy_yen_per_kwh", None)
                )
            )
            is not None
        ]
        vehicle_types = {
            str(getattr(vehicle_type, "vehicle_type_id", "") or ""): vehicle_type
            for vehicle_type in tuple(
                getattr(canonical_problem, "vehicle_types", ()) or ()
            )
        }
        for vehicle in tuple(
            getattr(canonical_problem, "vehicles", ()) or ()
        ):
            vehicle_type = vehicle_types.get(
                str(getattr(vehicle, "vehicle_type", "") or "")
            )
            # A vehicle's ``vehicle_type`` is a canonical type *identifier*,
            # not necessarily a powertrain label (for example, ``BYD_K8``).
            # Prefer the mapped canonical powertrain and retain the direct
            # label fallback for small test/diagnostic problems that omit a
            # vehicle-type table.
            powertrain_type = str(
                getattr(vehicle_type, "powertrain_type", "")
                or getattr(vehicle, "vehicle_type", "")
                or ""
            ).upper()
            is_electric = powertrain_type in electric_types
            for source in (vehicle, vehicle_type):
                if source is None:
                    continue
                if is_electric:
                    rate = _finite_nonnegative(
                        getattr(source, "energy_consumption_kwh_per_km", None)
                    )
                    if rate is not None and rate > 0.0:
                        electric_energy_rates.append(rate)
                        break
                else:
                    rate = _finite_nonnegative(
                        getattr(source, "fuel_consumption_l_per_km", None)
                    )
                    if rate is not None and rate > 0.0:
                        combustion_fuel_rates.append(rate)
                        break
        diesel_price = _finite_nonnegative(
            getattr(
                getattr(canonical_problem, "scenario", None),
                "diesel_price_yen_per_l",
                None,
            )
        )

    uniform_grid_price = _single_value(price_values)
    uniform_electric_energy_rate = _single_value(electric_energy_rates)
    uniform_combustion_fuel_rate = _single_value(combustion_fuel_rates)
    charge_efficiency = 0.95
    bev_grid_marginal_cost = (
        uniform_electric_energy_rate
        / charge_efficiency
        * uniform_grid_price
        if uniform_electric_energy_rate is not None
        and uniform_grid_price is not None
        else None
    )
    ice_marginal_cost = (
        uniform_combustion_fuel_rate * diesel_price
        if uniform_combustion_fuel_rate is not None and diesel_price is not None
        else None
    )
    grid_break_even = (
        ice_marginal_cost
        * charge_efficiency
        / uniform_electric_energy_rate
        if ice_marginal_cost is not None
        and uniform_electric_energy_rate is not None
        and uniform_electric_energy_rate > 0.0
        else None
    )

    gross_pv_available = _sum_numbers(
        recourse_weather.get("pv_available_kwh_by_depot")
    )
    if gross_pv_available <= 0.0:
        gross_pv_available = _sum_numbers(
            recourse_weather.get("pv_generation_kwh_by_depot_slot")
        )
    if gross_pv_available <= 0.0 and canonical_problem is not None:
        # Phase 4 has no Stage-1 recourse-weather metadata, and a failed solve
        # has no source-flow result to inspect.  The canonical depot assets are
        # nevertheless the authoritative input-side PV series, so preserve
        # that evidence instead of reporting a misleading zero-PV day.
        gross_pv_available = sum(
            _sum_numbers(
                getattr(asset, "pv_generation_kwh_by_slot", ()) or ()
            )
            for asset in dict(
                getattr(canonical_problem, "depot_energy_assets", {}) or {}
            ).values()
            if bool(getattr(asset, "pv_enabled", False))
        )
    pv_to_bus = _finite_nonnegative(
        recourse_result.get("pv_to_bus_kwh", cost_breakdown.get("pv_to_bus_kwh"))
    ) or 0.0
    pv_to_bess = _finite_nonnegative(
        recourse_result.get(
            "pv_to_bess_kwh", cost_breakdown.get("pv_to_bess_kwh")
        )
    ) or 0.0
    bess_to_bus = _finite_nonnegative(
        recourse_result.get(
            "bess_to_bus_kwh", cost_breakdown.get("bess_to_bus_kwh")
        )
    ) or 0.0
    grid_to_bus = _finite_nonnegative(
        recourse_result.get(
            "grid_to_bus_kwh", cost_breakdown.get("grid_to_bus_kwh")
        )
    ) or 0.0
    grid_to_bess = _finite_nonnegative(
        recourse_result.get(
            "grid_to_bess_kwh", cost_breakdown.get("grid_to_bess_kwh")
        )
    ) or 0.0

    candidate_rows = list(
        solver_metadata.get("stage1_stage2_candidate_evaluation") or []
    )
    selected_candidate_index = solver_metadata.get(
        "stage1_stage2_selected_candidate_index"
    )
    selected_candidate: Mapping[str, Any] = {}
    try:
        candidate_index = int(selected_candidate_index)
    except (TypeError, ValueError):
        candidate_index = 0
    if 1 <= candidate_index <= len(candidate_rows):
        raw_selected_candidate = candidate_rows[candidate_index - 1]
        if isinstance(raw_selected_candidate, Mapping):
            selected_candidate = raw_selected_candidate
    stage1_bev_trip_count = selected_candidate.get("bev_trips")
    trip_count_by_type = dict(result_summary.get("trip_count_by_type") or {})
    stage2_bev_trip_count = sum(
        int(value or 0)
        for vehicle_type, value in trip_count_by_type.items()
        if str(vehicle_type).upper() in electric_types
    )
    composition_certificate = dict(
        solver_metadata.get("stage1_used_powertrain_composition_search") or {}
    )

    return {
        "schema_version": "assignment_economic_audit_v1",
        "assignment_energy_coupling_mode": assignment_energy_coupling_mode,
        "trip_energy_model": (
            (getattr(canonical_problem, "metadata", {}) or {}).get(
                "trip_energy_model", "distance_average_v0"
            )
            if canonical_problem is not None
            else solver_metadata.get("trip_energy_model", "not_recorded")
        ),
        "trip_energy_model_provenance": (
            dict(
                (getattr(canonical_problem, "metadata", {}) or {}).get(
                    "trip_energy_model_provenance", {}
                )
            )
            if canonical_problem is not None
            else {}
        ),
        "pv_input_semantics_by_depot": (
            {
                str(depot_id): str(
                    getattr(asset, "pv_input_semantics", "not_recorded")
                )
                for depot_id, asset in dict(
                    getattr(canonical_problem, "depot_energy_assets", {}) or {}
                ).items()
            }
            if canonical_problem is not None
            else {}
        ),
        "charging_power_model": (
            (getattr(canonical_problem, "metadata", {}) or {}).get(
                "charging_power_model", "constant_power_v0"
            )
            if canonical_problem is not None
            else solver_metadata.get("charging_power_model", "not_recorded")
        ),
        # The objective preset is an input-side model contract.  Phase 4 does
        # not duplicate every canonical input into solver metadata, so fall
        # back to the canonical problem instead of exporting a misleading
        # null that makes otherwise identical comparison cases look different.
        "objective_preset": (
            solver_metadata.get("objective_preset")
            or problem_metadata.get("objective_preset")
        ),
        "co2_emissions_cap_kg": solver_metadata.get("co2_emissions_cap_kg"),
        "pv_supply_scale_by_depot": (
            {
                str(depot_id): float(
                    getattr(asset, "pv_supply_scale", 1.0)
                )
                for depot_id, asset in dict(
                    getattr(canonical_problem, "depot_energy_assets", {}) or {}
                ).items()
            }
            if canonical_problem is not None
            else {}
        ),
        "assignment_energy_coupling_stage2_authority": str(
            recourse_configuration.get("stage2_authority") or "not_recorded"
        ),
        "bev_grid_marginal_cost_jpy_per_km": bev_grid_marginal_cost,
        "ice_marginal_cost_jpy_per_km": ice_marginal_cost,
        "grid_energy_break_even_jpy_per_kwh": grid_break_even,
        "marginal_cost_assumptions": {
            "uniform_grid_price_jpy_per_kwh": uniform_grid_price,
            "electric_drive_energy_kwh_per_km": uniform_electric_energy_rate,
            "combustion_fuel_l_per_km": uniform_combustion_fuel_rate,
            "diesel_price_jpy_per_l": diesel_price,
            "charge_efficiency": charge_efficiency,
            "semantics": (
                "Grid BEV marginal cost is charger-input energy after the "
                "Stage 1/2 0.95 charge-efficiency contract.  A scalar is "
                "reported only when the selected scope has uniform price and "
                "powertrain coefficients."
            ),
        },
        # The Stage 1 model is slot-constrained, including PV availability,
        # charge timing, BESS dynamics, and terminal SOC. A single daily
        # renewable budget would overstate what can actually reach a BEV, so
        # preserve the requested field without inventing a scalar value.
        "renewable_budget_kwh": None,
        "gross_pv_available_kwh": gross_pv_available,
        "renewable_budget_semantics": (
            "not a scalar in slot-level assignment-coupled recourse; use "
            "gross_pv_available_kwh only as an input-side diagnostic, while "
            "the solver enforces PV, BESS, charging-time, and terminal-SOC "
            "constraints per slot. Initial BESS inventory is not credited as "
            "free renewable energy"
        ),
        "initial_bess_inventory_counted_as_free_kwh": 0.0,
        "renewable_energy_allocated_in_stage1_kwh": pv_to_bus + pv_to_bess,
        "renewable_energy_allocated_in_stage1_semantics": (
            "PV input allocated to bus charging or BESS charging; BESS output "
            "is reported separately because vehicle-level source attribution is "
            "not solver-native"
        ),
        "grid_energy_allocated_in_stage1_kwh": grid_to_bus + grid_to_bess,
        "stage1_source_flows_kwh": {
            "pv_to_bus_kwh": pv_to_bus,
            "pv_to_bess_kwh": pv_to_bess,
            "bess_to_bus_kwh": bess_to_bus,
            "grid_to_bus_kwh": grid_to_bus,
            "grid_to_bess_kwh": grid_to_bess,
        },
        "stage1_bev_trip_count": stage1_bev_trip_count,
        "stage2_bev_trip_count": stage2_bev_trip_count,
        "weather_response_expected": (
            "slot-limited PV/grid/BESS recourse changes the Stage 1 economic "
            "signal but imposes no directional BEV-use policy"
        ),
        "weather_response_observed": "not_assessable_from_single_case",
        "used_powertrain_composition_search": composition_certificate,
        "used_powertrain_composition_search_accepted": bool(
            solver_metadata.get(
                "stage1_used_powertrain_composition_search_accepted", False
            )
        ),
        "solver_objective_matches_accounting_total": bool(
            solver_metadata.get("solver_objective_matches_accounting_total", False)
        ),
        "vehicle_day_activation_semantics": (
            "Stage 1 used_vehicle_day binaries are linked to assignments and "
            "vehicle-day cost is charged once; canonical accounting remains "
            "the cost authority after accepted rolling."
        ),
        "vehicle_usage_cost_jpy_per_used_bus": (
            (getattr(canonical_problem, "metadata", {}) or {}).get(
                "vehicle_usage_cost_jpy_per_used_bus"
            )
            if canonical_problem is not None
            else None
        ),
        "vehicle_usage_cost_semantics": (
            (getattr(canonical_problem, "metadata", {}) or {}).get(
                "vehicle_usage_cost_semantics", "unclassified"
            )
            if canonical_problem is not None
            else "unclassified"
        ),
        "vehicle_usage_cost_semantics_classified": bool(
            canonical_problem is not None
            and (getattr(canonical_problem, "metadata", {}) or {}).get(
                "vehicle_usage_cost_semantics_classified", False
            )
        ),
        "vehicle_usage_cost_semantics_research_eligible": bool(
            canonical_problem is not None
            and (getattr(canonical_problem, "metadata", {}) or {}).get(
                "vehicle_usage_cost_semantics_research_eligible", False
            )
        ),
        "research_economic_claim_blocked_by_vehicle_usage_cost_semantics": bool(
            canonical_problem is not None
            and (getattr(canonical_problem, "metadata", {}) or {}).get(
                "vehicle_usage_cost_jpy_per_used_bus", 0.0
            )
            and not (getattr(canonical_problem, "metadata", {}) or {}).get(
                "vehicle_usage_cost_semantics_research_eligible", False
            )
        ),
        "source_provenance_level": (
            "depot_slot_solver_native_flows; no inferred vehicle-level PV/BESS "
            "allocation is presented as solver-native"
        ),
    }


def _powertrain_marginal_cost_audit_payload(
    canonical_problem: Optional[Any],
) -> Dict[str, Any]:
    """Compute source-specific marginal costs from canonical coefficients."""

    def _nonnegative(value: Any) -> Optional[float]:
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            return None
        return parsed if math.isfinite(parsed) and parsed >= 0.0 else None

    def _uniform(values: List[float]) -> Optional[float]:
        distinct = sorted({round(float(value), 12) for value in values})
        return distinct[0] if len(distinct) == 1 else None

    if canonical_problem is None:
        return {
            "schema_version": "powertrain_marginal_cost_audit_v1",
            "status": "UNAVAILABLE_CANONICAL_PROBLEM_MISSING",
        }

    vehicle_types = {
        str(getattr(item, "vehicle_type_id", "") or ""): item
        for item in tuple(
            getattr(canonical_problem, "vehicle_types", ()) or ()
        )
    }
    electric_rates: List[float] = []
    combustion_rates: List[float] = []
    combustion_co2_factors: List[float] = []
    for vehicle in tuple(getattr(canonical_problem, "vehicles", ()) or ()):
        vehicle_type = vehicle_types.get(
            str(getattr(vehicle, "vehicle_type", "") or "")
        )
        powertrain = str(
            getattr(vehicle_type, "powertrain_type", "")
            or getattr(vehicle, "vehicle_type", "")
            or ""
        ).upper()
        if powertrain in {"BEV", "PHEV", "FCEV"}:
            rate = _nonnegative(
                getattr(vehicle, "energy_consumption_kwh_per_km", None)
            ) or _nonnegative(
                getattr(
                    vehicle_type,
                    "energy_consumption_kwh_per_km",
                    None,
                )
            )
            if rate is not None and rate > 0.0:
                electric_rates.append(rate)
        else:
            rate = _nonnegative(
                getattr(vehicle, "fuel_consumption_l_per_km", None)
            ) or _nonnegative(
                getattr(vehicle_type, "fuel_consumption_l_per_km", None)
            )
            if rate is not None and rate > 0.0:
                combustion_rates.append(rate)
            factor = _nonnegative(
                getattr(vehicle_type, "co2_emission_kg_per_l", None)
            )
            if factor is not None and factor > 0.0:
                combustion_co2_factors.append(factor)

    price_values = [
        value
        for slot in tuple(getattr(canonical_problem, "price_slots", ()) or ())
        if (
            value := _nonnegative(
                getattr(slot, "grid_buy_yen_per_kwh", None)
            )
        )
        is not None
    ]
    grid_co2_values = [
        value
        for slot in tuple(getattr(canonical_problem, "price_slots", ()) or ())
        if (value := _nonnegative(getattr(slot, "co2_factor", None)))
        is not None
    ]
    assets = [
        asset
        for asset in dict(
            getattr(canonical_problem, "depot_energy_assets", {}) or {}
        ).values()
        if bool(getattr(asset, "bess_enabled", False))
    ]
    bess_charge_efficiencies = [
        value
        for asset in assets
        if (
            value := _nonnegative(
                getattr(asset, "bess_charge_efficiency", None)
            )
        )
        is not None
        and value > 0.0
    ]
    bess_discharge_efficiencies = [
        value
        for asset in assets
        if (
            value := _nonnegative(
                getattr(asset, "bess_discharge_efficiency", None)
            )
        )
        is not None
        and value > 0.0
    ]
    bess_cycle_costs = [
        value
        for asset in assets
        if (
            value := _nonnegative(
                getattr(asset, "bess_cycle_cost_yen_per_kwh", None)
            )
        )
        is not None
    ]

    bev_rate = _uniform(electric_rates)
    ice_rate = _uniform(combustion_rates)
    grid_price = _uniform(price_values)
    grid_co2 = _uniform(grid_co2_values)
    bess_charge_efficiency = (
        _uniform(bess_charge_efficiencies) if assets else 0.95
    )
    bess_discharge_efficiency = (
        _uniform(bess_discharge_efficiencies) if assets else 0.95
    )
    bess_cycle_cost = _uniform(bess_cycle_costs) if assets else 0.0
    scenario = getattr(canonical_problem, "scenario", None)
    diesel_price = _nonnegative(
        getattr(scenario, "diesel_price_yen_per_l", None)
    )
    co2_price = _nonnegative(getattr(scenario, "co2_price_per_kg", None))
    ice_co2_factor = _uniform(combustion_co2_factors) or _nonnegative(
        getattr(scenario, "ice_co2_kg_per_l", None)
    )
    pv_marginal_cost = _nonnegative(
        (getattr(canonical_problem, "metadata", {}) or {}).get(
            "pv_marginal_charge_cost_yen_per_kwh"
        )
    )
    if pv_marginal_cost is None:
        pv_marginal_cost = 0.0
    vehicle_charge_efficiency = 0.95

    required_scalars = {
        "bev_energy_consumption_kwh_per_km": bev_rate,
        "ice_fuel_consumption_l_per_km": ice_rate,
        "grid_price_jpy_per_kwh": grid_price,
        "diesel_price_jpy_per_l": diesel_price,
        "ice_co2_kg_per_l": ice_co2_factor,
        "co2_price_jpy_per_kg": co2_price,
        "bess_charge_efficiency": bess_charge_efficiency,
        "bess_discharge_efficiency": bess_discharge_efficiency,
        "bess_cycle_cost_jpy_per_kwh": bess_cycle_cost,
    }
    missing = [key for key, value in required_scalars.items() if value is None]
    status = "OK" if not missing else "UNRESOLVED_NONUNIFORM_OR_MISSING_COEFFICIENTS"

    bev_grid_cost = None
    bev_pv_direct_cost = None
    bev_pv_bess_cost = None
    ice_fuel_cost = None
    ice_co2_cost = None
    ice_total_cost = None
    grid_break_even = None
    if not missing:
        assert bev_rate is not None
        assert ice_rate is not None
        assert grid_price is not None
        assert diesel_price is not None
        assert ice_co2_factor is not None
        assert co2_price is not None
        assert bess_charge_efficiency is not None
        assert bess_discharge_efficiency is not None
        assert bess_cycle_cost is not None
        bev_grid_cost = (
            grid_price / vehicle_charge_efficiency * bev_rate
        )
        bev_pv_direct_cost = (
            pv_marginal_cost / vehicle_charge_efficiency * bev_rate
        )
        bev_pv_bess_cost = (
            (
                pv_marginal_cost / bess_charge_efficiency
                + bess_cycle_cost
            )
            / bess_discharge_efficiency
            / vehicle_charge_efficiency
            * bev_rate
        )
        ice_fuel_cost = ice_rate * diesel_price
        ice_co2_cost = ice_rate * ice_co2_factor * co2_price
        ice_total_cost = ice_fuel_cost + ice_co2_cost
        grid_break_even = (
            ice_total_cost * vehicle_charge_efficiency / bev_rate
            if bev_rate > 0.0
            else None
        )

    metadata = dict(getattr(canonical_problem, "metadata", {}) or {})
    vehicle_usage_cost = _nonnegative(
        metadata.get("vehicle_usage_cost_jpy_per_used_bus")
    ) or 0.0
    vehicle_usage_semantics = str(
        metadata.get("vehicle_usage_cost_semantics") or "unclassified"
    )
    return {
        "schema_version": "powertrain_marginal_cost_audit_v1",
        "status": status,
        "blocking_reasons": [
            f"scalar_unavailable:{key}" for key in missing
        ],
        "coefficients": {
            **required_scalars,
            "pv_marginal_cost_jpy_per_kwh": pv_marginal_cost,
            "vehicle_charge_efficiency": vehicle_charge_efficiency,
            "grid_co2_kg_per_kwh": grid_co2,
        },
        "marginal_costs_jpy_per_km": {
            "bev_grid": bev_grid_cost,
            "bev_pv_direct": bev_pv_direct_cost,
            "bev_pv_bess": bev_pv_bess_cost,
            "ice_fuel": ice_fuel_cost,
            "ice_co2": ice_co2_cost,
            "ice_total": ice_total_cost,
        },
        "grid_break_even_price_jpy_per_kwh_including_ice_co2_cost": (
            grid_break_even
        ),
        "grid_only_cheapest_powertrain": (
            "BEV"
            if bev_grid_cost is not None
            and ice_total_cost is not None
            and bev_grid_cost < ice_total_cost
            else "ICE"
            if bev_grid_cost is not None and ice_total_cost is not None
            else "UNRESOLVED"
        ),
        "vehicle_usage_cost_jpy_per_used_bus": vehicle_usage_cost,
        "vehicle_usage_cost_semantics": vehicle_usage_semantics,
        "vehicle_usage_cost_semantics_classified": bool(
            metadata.get("vehicle_usage_cost_semantics_classified", False)
        ),
        "vehicle_usage_cost_semantics_research_eligible": bool(
            metadata.get(
                "vehicle_usage_cost_semantics_research_eligible", False
            )
        ),
        "economic_claim_blocked": bool(
            vehicle_usage_cost > 0.0
            and not metadata.get(
                "vehicle_usage_cost_semantics_research_eligible", False
            )
        ),
        "formula_semantics": {
            "bev_grid": (
                "grid_price / vehicle_charge_efficiency * "
                "bev_energy_consumption"
            ),
            "bev_pv_direct": (
                "pv_marginal_cost / vehicle_charge_efficiency * "
                "bev_energy_consumption"
            ),
            "bev_pv_bess": (
                "(pv_marginal_cost / bess_charge_efficiency + "
                "bess_cycle_cost) / bess_discharge_efficiency / "
                "vehicle_charge_efficiency * bev_energy_consumption"
            ),
            "ice": (
                "ice_fuel_consumption * diesel_price + ice_co2_per_km * "
                "co2_price"
            ),
        },
    }


def _trip_powertrain_cost_comparison_rows(
    canonical_problem: Optional[Any],
    marginal_audit: Mapping[str, Any],
) -> List[Dict[str, Any]]:
    if canonical_problem is None:
        return []
    marginal_costs = dict(
        marginal_audit.get("marginal_costs_jpy_per_km") or {}
    )
    coefficients = dict(marginal_audit.get("coefficients") or {})
    bev_rate = coefficients.get("bev_energy_consumption_kwh_per_km")
    ice_rate = coefficients.get("ice_fuel_consumption_l_per_km")
    diesel_price = coefficients.get("diesel_price_jpy_per_l")
    ice_co2_factor = coefficients.get("ice_co2_kg_per_l")
    co2_price = coefficients.get("co2_price_jpy_per_kg")

    rows: List[Dict[str, Any]] = []
    for trip in sorted(
        tuple(getattr(canonical_problem, "trips", ()) or ()),
        key=lambda item: (
            int(getattr(item, "departure_min", 0) or 0),
            str(getattr(item, "trip_id", "") or ""),
        ),
    ):
        distance_km = max(float(getattr(trip, "distance_km", 0.0) or 0.0), 0.0)
        departure_min = int(getattr(trip, "departure_min", 0) or 0)
        hh, mm = divmod(departure_min % (24 * 60), 60)
        ice_fuel_cost = (
            distance_km * float(ice_rate) * float(diesel_price)
            if ice_rate is not None and diesel_price is not None
            else None
        )
        ice_co2_cost = (
            distance_km
            * float(ice_rate)
            * float(ice_co2_factor)
            * float(co2_price)
            if ice_rate is not None
            and ice_co2_factor is not None
            and co2_price is not None
            else None
        )
        ice_total_cost = (
            float(ice_fuel_cost or 0.0) + float(ice_co2_cost or 0.0)
            if ice_fuel_cost is not None and ice_co2_cost is not None
            else None
        )
        bev_grid_cost = (
            distance_km * float(marginal_costs["bev_grid"])
            if marginal_costs.get("bev_grid") is not None
            else None
        )
        bev_pv_direct_cost = (
            distance_km * float(marginal_costs["bev_pv_direct"])
            if marginal_costs.get("bev_pv_direct") is not None
            else None
        )
        bev_pv_bess_cost = (
            distance_km * float(marginal_costs["bev_pv_bess"])
            if marginal_costs.get("bev_pv_bess") is not None
            else None
        )
        cost_difference = (
            float(bev_grid_cost) - float(ice_total_cost)
            if bev_grid_cost is not None and ice_total_cost is not None
            else None
        )
        rows.append(
            {
                "trip_id": str(getattr(trip, "trip_id", "") or ""),
                "distance_km": round(distance_km, 9),
                "departure_time": f"{hh:02d}:{mm:02d}",
                "ice_fuel_cost_jpy": ice_fuel_cost,
                "ice_co2_cost_jpy": ice_co2_cost,
                "bev_energy_kwh": (
                    distance_km * float(bev_rate)
                    if bev_rate is not None
                    else None
                ),
                "bev_grid_cost_jpy": bev_grid_cost,
                "bev_pv_direct_cost_jpy": bev_pv_direct_cost,
                "bev_pv_bess_cost_jpy": bev_pv_bess_cost,
                "available_pv_bess_energy_at_relevant_slots": None,
                "charging_feasible": (
                    "UNRESOLVED_REQUIRES_DUTY_CHARGER_SOC_CONTEXT"
                ),
                "cheapest_powertrain": (
                    "BEV_GRID_ONLY"
                    if cost_difference is not None and cost_difference < 0.0
                    else "ICE_GRID_ONLY"
                    if cost_difference is not None
                    else "UNRESOLVED"
                ),
                "cost_difference_jpy": cost_difference,
                "comparison_semantics": (
                    "Grid-only BEV versus ICE operating marginal cost. "
                    "PV/BESS alternatives are conditional and are not "
                    "claimed feasible without a solved duty/charger/SOC path."
                ),
            }
        )
    return rows


def _persist_rich_run_outputs(
    *,
    run_dir: Path,
    scenario: Dict[str, Any],
    optimization_result: Dict[str, Any],
    optimization_audit: Dict[str, Any],
    result_payload: Dict[str, Any],
    sim_payload: Optional[Dict[str, Any]],
    canonical_solver_result: Optional[Dict[str, Any]],
    canonical_problem: Optional[Any] = None,
    graph_source_dir: Optional[Path] = None,
    charging_summary: Optional[Dict[str, Any]] = None,
    charging_flow_payload: Optional[Dict[str, Any]] = None,
    finalize_reporting: bool = False,
) -> Optional[Dict[str, Any]]:
    run_dir.mkdir(parents=True, exist_ok=True)

    unit_map = {
        "objective_value": "JPY",
        "solve_time_seconds": "s",
        "energy_cost": "JPY",
        "electricity_cost": "JPY",
        "pv_self_consumption_cost_jpy": "JPY",
        "pv_marginal_charge_cost_yen_per_kwh": "JPY/kWh",
        "pv_curtail_penalty_yen_per_kwh": "JPY/kWh",
        "demand_charge": "JPY",
        "vehicle_cost": "JPY",
        "vehicle_usage_cost": "JPY",
        "vehicle_usage_cost_jpy": "JPY",
        "vehicle_usage_cost_jpy_per_used_bus": "JPY/vehicle-day",
        "used_vehicle_day_count": "vehicle-day",
        "driver_cost": "JPY",
        "fuel_cost": "JPY",
        "fuel_cost_final": "JPY",
        "fuel_cost_provisional": "JPY",
        "fuel_cost_refueled": "JPY",
        "fuel_cost_realized": "JPY",
        "fuel_cost_provisional_leftover": "JPY",
        "penalty_unserved": "JPY",
        "return_leg_bonus": "JPY",
        "weather_strategy_objective_term_jpy_equivalent": "JPY-equivalent",
        "total_cost": "JPY",
        "accounting_total_cost_jpy": "JPY",
        "solver_objective_value": "JPY",
        "validated_operating_cost_jpy": "JPY",
        "co2_cost": "JPY",
        "total_co2_kg": "kg-CO2",
        "grid_to_bus_kwh": "kWh",
        "pv_to_bus_kwh": "kWh",
        "grid_to_bess_kwh": "kWh",
        "bess_to_bus_kwh": "kWh",
        "pv_to_bess_kwh": "kWh",
        "pv_curtail_kwh": "kWh",
        "grid_import_total_kwh": "kWh",
        "grid_import_for_contract_kwh": "kWh",
        "bus_charge_from_grid_kwh": "kWh",
        "bus_charge_from_bess_kwh": "kWh",
        "contract_over_limit_kwh": "kWh",
        "contract_overage_cost": "JPY",
        "peak_grid_kw": "kW",
    }

    (run_dir / "optimization_result.json").write_text(
        json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (run_dir / "optimization_audit.json").write_text(
        json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (run_dir / "solver_result.json").write_text(
        json.dumps(result_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if canonical_solver_result is not None:
        (run_dir / "canonical_solver_result.json").write_text(
            json.dumps(canonical_solver_result, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    accounting_summary = dict((optimization_result.get("graph_artifacts") or {}).get("accounting_summary") or {})
    solution_validity = dict(
        optimization_result.get("solution_validity")
        or (optimization_result.get("summary") or {}).get("solution_validity")
        or {}
    )
    if "validated_feasible" in solution_validity:
        evaluation_valid = bool(solution_validity.get("validated_feasible"))
    elif isinstance(canonical_solver_result, dict) and "feasible" in canonical_solver_result:
        evaluation_valid = bool(canonical_solver_result.get("feasible"))
    else:
        evaluation_valid = str(optimization_result.get("solver_status") or "").lower() in {
            "optimal",
            "feasible",
            "solved_feasible",
        }

    def _evaluation_float(value: Any) -> Optional[float]:
        if not evaluation_valid or value in (None, ""):
            return None
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            return None
        return parsed if math.isfinite(parsed) else None

    result_summary = dict(optimization_result.get("summary") or {})
    # Older/alternative solver paths do not necessarily provide settings.  Keep
    # rich-output persistence backward compatible and emit null telemetry in
    # that case rather than failing after a feasible solve.
    solver_settings = dict(optimization_result.get("solver_settings") or {})
    solver_metadata = dict(optimization_result.get("solver_metadata") or {})
    summary = {
        "scenario_id": optimization_result.get("scenario_id"),
        "mode": optimization_result.get("mode"),
        "solver_status": optimization_result.get("solver_status"),
        "objective_mode": optimization_result.get("objective_mode"),
        "objective_value": optimization_result.get("objective_value"),
        "objective_value_unit": "JPY",
        "objective_value_jpy": _evaluation_float(
            accounting_summary.get("objective_value_jpy", optimization_result.get("objective_value"))
        ),
        "total_cost_jpy": _evaluation_float(
            accounting_summary.get(
                "total_cost_jpy",
                (optimization_result.get("cost_breakdown") or {}).get(
                    "total_cost",
                    optimization_result.get("objective_value"),
                ),
            )
        ),
        "accounting_total_cost_jpy": _evaluation_float(
            accounting_summary.get(
                "accounting_total_cost_jpy",
                accounting_summary.get("total_cost_jpy"),
            )
        ),
        "solver_objective_value": _evaluation_float(
            accounting_summary.get(
                "solver_objective_value",
                accounting_summary.get(
                    "objective_value_jpy", optimization_result.get("objective_value")
                ),
            )
        ),
        "validated_operating_cost_jpy": _evaluation_float(
            accounting_summary.get("validated_operating_cost_jpy")
        ),
        "objective_is_actual_cost": bool(
            evaluation_valid
            and (optimization_result.get("cost_breakdown") or {}).get(
                "objective_is_actual_cost", False
            )
        ),
        "solver_objective_matches_accounting_total": bool(
            solver_metadata.get(
                "solver_objective_matches_accounting_total", False
            )
        ),
        "used_powertrain_composition_search_accepted": bool(
            solver_metadata.get(
                "stage1_used_powertrain_composition_search_accepted", False
            )
        ),
        "supports_exact_milp": bool((optimization_result.get("solver_metadata") or {}).get("supports_exact_milp", False)),
        "stage1_solver_status": solver_settings.get("stage1_solver_status"),
        "stage1_termination_reason": solver_settings.get(
            "stage1_termination_reason"
        ),
        "stage1_best_obj_stop_enabled": solver_settings.get(
            "stage1_best_obj_stop_enabled"
        ),
        "stage1_best_obj_stop_applied": solver_settings.get(
            "stage1_best_obj_stop_applied"
        ),
        "stage1_best_obj_stop_threshold": solver_settings.get(
            "stage1_best_obj_stop_threshold"
        ),
        "stage1_gurobi_raw_mip_gap_ratio": solver_settings.get(
            "stage1_gurobi_raw_mip_gap_ratio"
        ),
        "stage1_certified_mip_gap_ratio": solver_settings.get(
            "stage1_certified_mip_gap_ratio"
        ),
        "runtime_comparison_eligible": solver_settings.get(
            "runtime_comparison_eligible"
        ),
        "mip_gap_target_met": solver_settings.get("mip_gap_target_met"),
        "gurobi_threads": solver_settings.get("gurobi_threads"),
        "interactive_runtime_controls": dict(
            solver_settings.get("interactive_runtime_controls") or {}
        ),
        "interactive_operation_time_window_controls": dict(
            solver_settings.get("interactive_operation_time_window_controls") or {}
        ),
        "interactive_terminal_soc_controls": dict(
            solver_settings.get("interactive_terminal_soc_controls") or {}
        ),
        "bev_terminal_soc_policy": solver_metadata.get("bev_terminal_soc_policy"),
        "bev_terminal_soc_balance_satisfied": solver_metadata.get(
            "bev_terminal_soc_balance_satisfied"
        ),
        "bev_terminal_soc_total_drawdown_kwh": solver_metadata.get(
            "bev_terminal_soc_total_drawdown_kwh"
        ),
        "solve_time_seconds": optimization_result.get("solve_time_seconds"),
        "solve_time_unit": "s",
        "trip_count_served": (
            accounting_summary.get("served_trip_count", result_summary.get("trip_count_served"))
            if evaluation_valid
            else result_summary.get("trip_count_served")
        ),
        "trip_count_unserved": (
            accounting_summary.get("unserved_trip_count", result_summary.get("trip_count_unserved"))
            if evaluation_valid
            else result_summary.get("trip_count_unserved")
        ),
        "vehicle_count_used": accounting_summary.get("used_vehicle_count", (optimization_result.get("summary") or {}).get("vehicle_count_used")),
        "canonical_cost_components_jpy": dict(
            accounting_summary.get("canonical_cost_components_jpy") or {}
        ),
        "canonical_cost_component_status": dict(
            accounting_summary.get("canonical_cost_component_status") or {}
        ),
        "same_day_depot_cycles_enabled": (optimization_result.get("summary") or {}).get("same_day_depot_cycles_enabled"),
        "max_depot_cycles_per_vehicle_per_day": (optimization_result.get("summary") or {}).get("max_depot_cycles_per_vehicle_per_day"),
        "vehicle_fragment_counts": (optimization_result.get("summary") or {}).get("vehicle_fragment_counts"),
        "vehicles_with_multiple_fragments": (optimization_result.get("summary") or {}).get("vehicles_with_multiple_fragments"),
        "max_fragments_observed": (optimization_result.get("summary") or {}).get("max_fragments_observed"),
        "solution_validity": solution_validity,
        "result_status": optimization_result.get("result_status"),
        "failure_stage": optimization_result.get("failure_stage"),
        "research_kpi_eligible": bool(
            optimization_result.get("research_kpi_eligible", False)
        ),
    }
    (run_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (run_dir / "solver_settings.json").write_text(
        json.dumps(solver_settings, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    phase4_iis_guidance_audit = {
        "schema_version": "phase4_iis_assignment_guidance_audit_v1",
        "seed_stage2_iis_pattern_count": solver_settings.get(
            "phase4_phase3_seed_stage2_iis_assignment_guidance_pattern_count"
        ),
        "seed_stage2_iis_source_candidate_hashes": list(
            solver_settings.get(
                "phase4_phase3_seed_stage2_iis_assignment_guidance_source_candidate_hashes"
            )
            or ()
        ),
        "integrated_guidance_pattern_count": solver_settings.get(
            "integrated_phase3_iis_assignment_guidance_pattern_count"
        ),
        "integrated_guidance_pattern_hashes": list(
            solver_settings.get(
                "integrated_phase3_iis_assignment_guidance_pattern_hashes"
            )
            or ()
        ),
        "integrated_source_candidate_hashes": list(
            solver_settings.get(
                "integrated_phase3_iis_assignment_guidance_source_candidate_hashes"
            )
            or ()
        ),
        "promoted_assignment_variable_count": solver_settings.get(
            "integrated_phase3_iis_assignment_guidance_variable_count"
        ),
        "branch_priority": solver_settings.get(
            "integrated_phase3_iis_assignment_guidance_branch_priority"
        ),
        "semantics": solver_settings.get(
            "integrated_phase3_iis_assignment_guidance_semantics"
        ),
        "objective_changed": False,
        "feasible_set_changed": False,
        "preferred_assignment_value": None,
        "phase4_hard_cut_applied": False,
        "research_interpretation": (
            "Stage-2 IIS patterns receive non-directional assignment branch "
            "priorities; they are not treated as Phase-4 infeasibility proofs"
        ),
    }
    (run_dir / "phase4_iis_assignment_guidance_audit.json").write_text(
        json.dumps(phase4_iis_guidance_audit, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    assignment_economic_audit = _assignment_economic_audit_payload(
        canonical_problem=canonical_problem,
        optimization_result=optimization_result,
    )
    (run_dir / "assignment_economic_audit.json").write_text(
        json.dumps(assignment_economic_audit, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_csv_rows(
        run_dir / "assignment_economic_audit.csv",
        [
            {
                "objective_preset": assignment_economic_audit.get(
                    "objective_preset"
                ),
                "bev_grid_marginal_cost_jpy_per_km": (
                    assignment_economic_audit.get(
                        "bev_grid_marginal_cost_jpy_per_km"
                    )
                ),
                "ice_marginal_cost_jpy_per_km": (
                    assignment_economic_audit.get(
                        "ice_marginal_cost_jpy_per_km"
                    )
                ),
                "renewable_budget_kwh": assignment_economic_audit.get(
                    "renewable_budget_kwh"
                ),
                "renewable_energy_allocated_in_stage1_kwh": (
                    assignment_economic_audit.get(
                        "renewable_energy_allocated_in_stage1_kwh"
                    )
                ),
                "grid_energy_allocated_in_stage1_kwh": (
                    assignment_economic_audit.get(
                        "grid_energy_allocated_in_stage1_kwh"
                    )
                ),
                "stage1_bev_trip_count": assignment_economic_audit.get(
                    "stage1_bev_trip_count"
                ),
                "stage2_bev_trip_count": assignment_economic_audit.get(
                    "stage2_bev_trip_count"
                ),
                "assignment_energy_coupling_mode": (
                    assignment_economic_audit.get(
                        "assignment_energy_coupling_mode"
                    )
                ),
                "weather_response_expected": assignment_economic_audit.get(
                    "weather_response_expected"
                ),
                "weather_response_observed": assignment_economic_audit.get(
                    "weather_response_observed"
                ),
                "vehicle_usage_cost_jpy_per_used_bus": (
                    assignment_economic_audit.get(
                        "vehicle_usage_cost_jpy_per_used_bus"
                    )
                ),
                "vehicle_usage_cost_semantics": (
                    assignment_economic_audit.get(
                        "vehicle_usage_cost_semantics"
                    )
                ),
                "vehicle_usage_cost_semantics_classified": (
                    assignment_economic_audit.get(
                        "vehicle_usage_cost_semantics_classified"
                    )
                ),
                "vehicle_usage_cost_semantics_research_eligible": (
                    assignment_economic_audit.get(
                        "vehicle_usage_cost_semantics_research_eligible"
                    )
                ),
            }
        ],
        [
            "objective_preset",
            "bev_grid_marginal_cost_jpy_per_km",
            "ice_marginal_cost_jpy_per_km",
            "renewable_budget_kwh",
            "renewable_energy_allocated_in_stage1_kwh",
            "grid_energy_allocated_in_stage1_kwh",
            "stage1_bev_trip_count",
            "stage2_bev_trip_count",
            "assignment_energy_coupling_mode",
            "weather_response_expected",
            "weather_response_observed",
            "vehicle_usage_cost_jpy_per_used_bus",
            "vehicle_usage_cost_semantics",
            "vehicle_usage_cost_semantics_classified",
            "vehicle_usage_cost_semantics_research_eligible",
        ],
    )
    powertrain_marginal_cost_audit = (
        _powertrain_marginal_cost_audit_payload(canonical_problem)
    )
    (run_dir / "powertrain_marginal_cost_audit.json").write_text(
        json.dumps(
            powertrain_marginal_cost_audit,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    marginal_costs = dict(
        powertrain_marginal_cost_audit.get(
            "marginal_costs_jpy_per_km"
        )
        or {}
    )
    _write_csv_rows(
        run_dir / "powertrain_marginal_cost_audit.csv",
        [
            {
                "status": powertrain_marginal_cost_audit.get("status"),
                "bev_grid_marginal_cost_jpy_per_km": marginal_costs.get(
                    "bev_grid"
                ),
                "bev_pv_direct_marginal_cost_jpy_per_km": (
                    marginal_costs.get("bev_pv_direct")
                ),
                "bev_pv_bess_marginal_cost_jpy_per_km": (
                    marginal_costs.get("bev_pv_bess")
                ),
                "ice_fuel_marginal_cost_jpy_per_km": marginal_costs.get(
                    "ice_fuel"
                ),
                "ice_co2_marginal_cost_jpy_per_km": marginal_costs.get(
                    "ice_co2"
                ),
                "ice_total_marginal_cost_jpy_per_km": marginal_costs.get(
                    "ice_total"
                ),
                "grid_break_even_price_jpy_per_kwh": (
                    powertrain_marginal_cost_audit.get(
                        "grid_break_even_price_jpy_per_kwh_including_ice_co2_cost"
                    )
                ),
                "vehicle_usage_cost_jpy_per_used_bus": (
                    powertrain_marginal_cost_audit.get(
                        "vehicle_usage_cost_jpy_per_used_bus"
                    )
                ),
                "vehicle_usage_cost_semantics": (
                    powertrain_marginal_cost_audit.get(
                        "vehicle_usage_cost_semantics"
                    )
                ),
                "vehicle_usage_cost_semantics_research_eligible": (
                    powertrain_marginal_cost_audit.get(
                        "vehicle_usage_cost_semantics_research_eligible"
                    )
                ),
                "economic_claim_blocked": (
                    powertrain_marginal_cost_audit.get(
                        "economic_claim_blocked"
                    )
                ),
            }
        ],
        [
            "status",
            "bev_grid_marginal_cost_jpy_per_km",
            "bev_pv_direct_marginal_cost_jpy_per_km",
            "bev_pv_bess_marginal_cost_jpy_per_km",
            "ice_fuel_marginal_cost_jpy_per_km",
            "ice_co2_marginal_cost_jpy_per_km",
            "ice_total_marginal_cost_jpy_per_km",
            "grid_break_even_price_jpy_per_kwh",
            "vehicle_usage_cost_jpy_per_used_bus",
            "vehicle_usage_cost_semantics",
            "vehicle_usage_cost_semantics_research_eligible",
            "economic_claim_blocked",
        ],
    )
    trip_powertrain_rows = _trip_powertrain_cost_comparison_rows(
        canonical_problem,
        powertrain_marginal_cost_audit,
    )
    _write_csv_rows(
        run_dir / "trip_powertrain_cost_comparison.csv",
        trip_powertrain_rows,
        [
            "trip_id",
            "distance_km",
            "departure_time",
            "ice_fuel_cost_jpy",
            "ice_co2_cost_jpy",
            "bev_energy_kwh",
            "bev_grid_cost_jpy",
            "bev_pv_direct_cost_jpy",
            "bev_pv_bess_cost_jpy",
            "available_pv_bess_energy_at_relevant_slots",
            "charging_feasible",
            "cheapest_powertrain",
            "cost_difference_jpy",
            "comparison_semantics",
        ],
    )
    phase_name = str(solver_metadata.get("phase") or "")
    integrated_actual_cost = None
    if (
        phase_name == "phase4_integrated"
        and solver_metadata.get(
            "actual_cost_objective_numeric_reconciliation_passed"
        )
        is True
    ):
        integrated_actual_cost = (
            optimization_result.get("cost_breakdown") or {}
        ).get("total_cost")
    baseline_actual_cost = solver_metadata.get(
        "paired_baseline_actual_cost_jpy"
    )
    baseline_comparison_ready = bool(
        baseline_actual_cost is not None
        and integrated_actual_cost is not None
    )
    _write_csv_rows(
        run_dir / "baseline_vs_integrated_actual_cost.csv",
        [
            {
                "status": (
                    "READY"
                    if baseline_comparison_ready
                    else "UNAVAILABLE_REQUIRES_PAIRED_BASELINE_AND_PHASE4_RUN"
                ),
                "phase": phase_name,
                "baseline_actual_cost_jpy": baseline_actual_cost,
                "integrated_actual_cost_jpy": integrated_actual_cost,
                "cost_difference_jpy": (
                    float(integrated_actual_cost)
                    - float(baseline_actual_cost)
                    if baseline_comparison_ready
                    else None
                ),
                "solver_objective_matches_accounting_total": (
                    solver_metadata.get(
                        "solver_objective_matches_accounting_total"
                    )
                ),
                "semantics": (
                    "No baseline value is inferred from a Phase 3 proxy. "
                    "READY requires an explicitly paired canonical baseline "
                    "and a numerically reconciled Phase 4 actual-cost run."
                ),
            }
        ],
        [
            "status",
            "phase",
            "baseline_actual_cost_jpy",
            "integrated_actual_cost_jpy",
            "cost_difference_jpy",
            "solver_objective_matches_accounting_total",
            "semantics",
        ],
    )

    cost_breakdown = dict(optimization_result.get("cost_breakdown") or {})
    operating_cost = cost_breakdown.get("total_cost")
    operating_plus_modeled_assets = cost_breakdown.get(
        "total_cost_with_assets"
    )
    modeled_asset_cost = sum(
        float(cost_breakdown.get(key, 0.0) or 0.0)
        for key in ("pv_asset_cost", "bess_asset_cost")
    )
    lifecycle_scope = {
        "schema_version": "operating_and_lifecycle_cost_scope_v1",
        "daily_operating_cost_jpy": operating_cost,
        "daily_operating_cost_status": (
            "AVAILABLE" if operating_cost is not None else "UNAVAILABLE"
        ),
        "daily_operating_cost_semantics": (
            "Existing-asset dispatch cost: energy, fuel, demand, contract "
            "overage, CO2, enabled vehicle-day/driver, degradation, and other "
            "canonical operating components."
        ),
        "modeled_pv_bess_asset_cost_jpy_per_day": modeled_asset_cost,
        "operating_plus_modeled_assets_jpy_per_day": (
            operating_plus_modeled_assets
        ),
        "lifecycle_cost_status": "PARTIAL_MODEL_SCOPE",
        "lifecycle_cost_missing_components": [
            "charger_capex_and_maintenance",
            "explicit_vehicle_type_replacement_plan_when_acquisition_disabled",
            "financing_or_discount_rate_contract",
        ],
        "lifecycle_cost_semantics": (
            "PV/BESS and enabled vehicle dailyized asset terms are reported "
            "separately from operating cost. This is not a complete lifecycle "
            "cost unless every listed missing component is explicitly supplied."
        ),
    }
    (run_dir / "operating_and_lifecycle_cost_scope.json").write_text(
        json.dumps(lifecycle_scope, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_csv_rows(
        run_dir / "operating_and_lifecycle_cost_scope.csv",
        [lifecycle_scope],
        [
            "daily_operating_cost_jpy",
            "daily_operating_cost_status",
            "modeled_pv_bess_asset_cost_jpy_per_day",
            "operating_plus_modeled_assets_jpy_per_day",
            "lifecycle_cost_status",
            "lifecycle_cost_missing_components",
            "daily_operating_cost_semantics",
            "lifecycle_cost_semantics",
        ],
    )
    cost_rows = [
        {
            "key": key,
            "value": value,
            "unit": unit_map.get(key, ""),
        }
        for key, value in cost_breakdown.items()
    ]
    (run_dir / "cost_breakdown_detail.json").write_text(
        json.dumps({"rows": cost_rows}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_csv_rows(
        run_dir / "cost_breakdown_detail.csv",
        cost_rows,
        ["key", "value", "unit"],
    )

    objective_rows = [
        {"key": key, "value": value, "unit": unit_map.get(key, "")}
        for key, value in dict(result_payload.get("obj_breakdown") or {}).items()
    ]
    solution_validity = dict(
        optimization_result.get("solution_validity")
        or (optimization_result.get("summary") or {}).get("solution_validity")
        or {}
    )
    if solution_validity:
        objective_rows = [
            row for row in objective_rows if row.get("key") != "evaluation_feasible"
        ]
        objective_rows.append(
            {
                "key": "evaluation_feasible",
                "value": 1.0 if bool(solution_validity.get("validated_feasible")) else 0.0,
                "unit": "",
            }
        )
    (run_dir / "objective_breakdown.json").write_text(
        json.dumps({"rows": objective_rows}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_csv_rows(
        run_dir / "objective_breakdown.csv",
        objective_rows,
        ["key", "value", "unit"],
    )

    assignment = dict(result_payload.get("assignment") or {})
    vehicle_schedule_rows: List[Dict[str, Any]] = []
    for vehicle_id, trip_ids in assignment.items():
        for order, trip_id in enumerate(list(trip_ids or []), start=1):
            vehicle_schedule_rows.append(
                {
                    "vehicle_id": vehicle_id,
                    "sequence": order,
                    "trip_id": trip_id,
                }
            )
    _write_csv_rows(
        run_dir / "vehicle_schedule.csv",
        vehicle_schedule_rows,
        ["vehicle_id", "sequence", "trip_id"],
    )

    summary_payload = dict(optimization_result.get("summary") or {})
    trip_type_rows = [
        {
            "vehicle_type": vehicle_type,
            "trip_count": trip_count,
            "unit": "trips",
        }
        for vehicle_type, trip_count in dict(summary_payload.get("trip_count_by_type") or {}).items()
    ]
    (run_dir / "trip_type_counts.json").write_text(
        json.dumps({"rows": trip_type_rows}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_csv_rows(
        run_dir / "trip_type_counts.csv",
        trip_type_rows,
        ["vehicle_type", "trip_count", "unit"],
    )

    targeted_rows = [
        {"trip_id": trip_id, "status": "unserved"}
        for trip_id in list(result_payload.get("unserved_tasks") or [])
    ]
    (run_dir / "targeted_trips.json").write_text(
        json.dumps({"rows": targeted_rows}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_csv_rows(
        run_dir / "targeted_trips.csv",
        targeted_rows,
        ["trip_id", "status"],
    )

    sim_cfg = dict(scenario.get("simulation_config") or {})
    cost_cfg = dict(((scenario.get("scenario_overlay") or {}).get("cost_coefficients") or {}))
    (run_dir / "simulation_conditions.json").write_text(
        json.dumps(sim_cfg, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    weather_policy = dict(optimization_result.get("weather_policy") or {})
    if weather_policy.get("enabled"):
        forecast_payload = dict(weather_policy.get("forecast") or {})
        profile_payload = dict(weather_policy.get("operation_profile") or {})
        audit_payload = dict(weather_policy.get("audit") or {})
        representative_curve_payload = dict(weather_policy.get("representative_curve") or {})
        (run_dir / "weather_proxy_forecast.json").write_text(
            json.dumps(forecast_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (run_dir / "weather_operation_policy.json").write_text(
            json.dumps(profile_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (run_dir / "weather_policy_audit.json").write_text(
            json.dumps(audit_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        if representative_curve_payload:
            (run_dir / "weather_pv_representative_curve.json").write_text(
                json.dumps(representative_curve_payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

    # Legacy-compatible simulation condition tables.
    vehicles = list(scenario.get("vehicles") or [])
    vehicle_cost_rows: List[Dict[str, Any]] = []
    for vehicle in vehicles:
        if not isinstance(vehicle, dict):
            continue
        vehicle_type = str(vehicle.get("vehicle_type") or vehicle.get("type") or "").upper()
        fuel_unit_price = (
            vehicle.get("fuel_cost_coeff_yen_per_liter")
            or vehicle.get("fuelCostPerL")
            or vehicle.get("fuel_price_yen_per_liter")
            or vehicle.get("fuelPriceYenPerLiter")
        )
        if fuel_unit_price in (None, "") and vehicle_type not in {"BEV", "PHEV", "FCEV"}:
            fuel_unit_price = cost_cfg.get("diesel_price_per_l") or sim_cfg.get("diesel_price_per_l") or 0.0
        vehicle_cost_rows.append(
            {
                "vehicle_id": vehicle.get("vehicle_id") or vehicle.get("id") or vehicle.get("name") or "",
                "vehicle_type": vehicle_type,
                "fixed_use_cost_yen": vehicle.get("fixed_use_cost_yen") or vehicle.get("fixed_cost_yen") or 0.0,
                "fuel_cost_coeff_yen_per_liter": fuel_unit_price or 0.0,
                "battery_degradation_cost_coeff_yen_per_kwh": vehicle.get("battery_degradation_cost_coeff_yen_per_kwh") or 0.0,
                "co2_emission_coeff_kg_per_liter": vehicle.get("co2_emission_coeff_kg_per_liter") or vehicle.get("co2EmissionKgPerL") or 0.0,
            }
        )
    _write_csv_rows(
        run_dir / "simulation_conditions_vehicle_costs.csv",
        vehicle_cost_rows,
        [
            "vehicle_id",
            "vehicle_type",
            "fixed_use_cost_yen",
            "fuel_cost_coeff_yen_per_liter",
            "battery_degradation_cost_coeff_yen_per_kwh",
            "co2_emission_coeff_kg_per_liter",
        ],
    )

    canonical_condition_tables = _canonical_simulation_condition_tables(
        canonical_problem
    )
    if canonical_condition_tables is not None:
        tou_rows, contract_rows, condition_provenance = canonical_condition_tables
    else:
        # Retain a clearly-labelled legacy fallback for direct callers that do
        # not provide a canonical problem. The interactive BFF execution path
        # always supplies one, so formal evidence is generated from the model.
        slot_count = int(sim_cfg.get("planning_horizon_hours") or 24)
        timestep = int(
            sim_cfg.get("time_step_min") or sim_cfg.get("timestep_min") or 60
        )
        if timestep > 0:
            slot_count = max(slot_count * 60 // timestep, 1)
        tou_price_series = list(sim_cfg.get("tou_prices_yen_per_kwh") or [])
        default_price = float(sim_cfg.get("grid_energy_price_yen_per_kwh") or 0.0)
        grid_co2_series = list(sim_cfg.get("grid_co2_factor_kg_per_kwh") or [])
        base_load_series = list(sim_cfg.get("base_load_kw") or [])
        site_id = str(sim_cfg.get("depot_id") or "depot_A")
        tou_rows = []
        for time_idx in range(slot_count):
            tou_rows.append(
                {
                    "site_id": site_id,
                    "time_idx": time_idx,
                    "grid_energy_price_yen_per_kwh": (
                        tou_price_series[time_idx]
                        if time_idx < len(tou_price_series)
                        else default_price
                    ),
                    "sell_back_price_yen_per_kwh": 0.0,
                    "base_load_kw": (
                        base_load_series[time_idx]
                        if time_idx < len(base_load_series)
                        else 0.0
                    ),
                    "demand_charge_weight": None,
                    "grid_co2_factor_kg_per_kwh": (
                        grid_co2_series[time_idx]
                        if time_idx < len(grid_co2_series)
                        else 0.0
                    ),
                }
            )
        contract_rows = [
            {
                "site_id": site_id,
                "site_type": "depot",
                "contract_demand_limit_kw": float(
                    sim_cfg.get("contract_demand_limit_kw") or 0.0
                ),
                "grid_import_limit_kw": float(
                    sim_cfg.get("grid_import_limit_kw") or 0.0
                ),
                "site_transformer_limit_kw": float(
                    sim_cfg.get("site_transformer_limit_kw") or 0.0
                ),
            }
        ]
        condition_provenance = {
            "schema_version": "simulation_conditions_provenance_v1",
            "source": "legacy_scenario_payload_fallback",
            "reason": "canonical_problem_not_supplied_to_rich_output_writer",
            "price_source": "scenario.simulation_config",
            "contract_limit_source": "scenario.simulation_config",
        }
    _write_csv_rows(
        run_dir / "simulation_conditions_tou_prices.csv",
        tou_rows,
        [
            "site_id",
            "time_idx",
            "grid_energy_price_yen_per_kwh",
            "sell_back_price_yen_per_kwh",
            "base_load_kw",
            "demand_charge_weight",
            "grid_co2_factor_kg_per_kwh",
        ],
    )

    _write_csv_rows(
        run_dir / "simulation_conditions_contract_limits.csv",
        contract_rows,
        [
            "site_id",
            "site_type",
            "contract_demand_limit_kw",
            "grid_import_limit_kw",
            "site_transformer_limit_kw",
        ],
    )
    (run_dir / "simulation_conditions_provenance.json").write_text(
        json.dumps(condition_provenance, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    ice_co2_kg = float(
        cost_breakdown.get("ice_bus_co2_kg", cost_breakdown.get("ice_co2_kg", cost_breakdown.get("engine_bus_co2_kg", 0.0))) or 0.0
    )
    grid_co2_kg = float(
        cost_breakdown.get("grid_electricity_co2_kg", cost_breakdown.get("power_generation_co2_kg", 0.0)) or 0.0
    )
    pv_co2_kg = float(cost_breakdown.get("pv_operational_co2_kg", cost_breakdown.get("pv_co2_kg", 0.0)) or 0.0)
    bess_storage_co2_kg = float(cost_breakdown.get("bess_storage_operational_co2_kg", 0.0) or 0.0)
    total_components_co2_kg = ice_co2_kg + grid_co2_kg + pv_co2_kg + bess_storage_co2_kg
    total_co2_kg = float(cost_breakdown.get("total_co2_kg", total_components_co2_kg) or 0.0)
    if abs(total_co2_kg - total_components_co2_kg) > 1.0e-6 and ice_co2_kg == 0.0 and grid_co2_kg == 0.0 and pv_co2_kg == 0.0 and bess_storage_co2_kg == 0.0:
        grid_co2_kg = total_co2_kg
        total_components_co2_kg = total_co2_kg
    co2_rows = [
        {"component": "ice_bus_co2_kg", "value": ice_co2_kg},
        {"component": "grid_electricity_co2_kg", "value": grid_co2_kg},
        {"component": "pv_operational_co2_kg", "value": pv_co2_kg},
        {"component": "bess_storage_operational_co2_kg", "value": bess_storage_co2_kg},
        {"component": "total_co2_kg", "value": total_components_co2_kg},
    ]
    (run_dir / "co2_breakdown.json").write_text(
        json.dumps({"rows": co2_rows}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_csv_rows(
        run_dir / "co2_breakdown.csv",
        co2_rows,
        ["component", "value"],
    )

    raw_dir = run_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    (raw_dir / "optimization_result.json").write_text(
        json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (raw_dir / "optimization_audit.json").write_text(
        json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (raw_dir / "solver_result.json").write_text(
        json.dumps(result_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if canonical_solver_result is not None:
        (raw_dir / "canonical_solver_result.json").write_text(
            json.dumps(canonical_solver_result, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    raw_assignment_rows: List[Dict[str, Any]] = []
    for vehicle_id, trip_ids in assignment.items():
        for order, trip_id in enumerate(list(trip_ids or []), start=1):
            raw_assignment_rows.append(
                {
                    "vehicle_id": vehicle_id,
                    "sequence": order,
                    "trip_id": trip_id,
                }
            )
    _write_csv_rows(
        raw_dir / "assignment.csv",
        raw_assignment_rows,
        ["vehicle_id", "sequence", "trip_id"],
    )
    raw_unserved_rows = [
        {"trip_id": trip_id, "status": "unserved"}
        for trip_id in list(result_payload.get("unserved_tasks") or [])
    ]
    _write_csv_rows(
        raw_dir / "unserved_trips.csv",
        raw_unserved_rows,
        ["trip_id", "status"],
    )

    graph_artifacts = dict(optimization_result.get("graph_artifacts") or {})
    timeline_candidates: List[Path] = []
    if graph_artifacts.get("vehicle_timeline_path"):
        rel = Path(str(graph_artifacts.get("vehicle_timeline_path")))
        timeline_candidates.append(run_dir / rel)
        if graph_source_dir is not None:
            timeline_candidates.append(graph_source_dir / rel.name)
            timeline_candidates.append(graph_source_dir / rel)
    timeline_candidates.append(run_dir / "graph" / "vehicle_timeline.csv")
    if graph_source_dir is not None:
        timeline_candidates.append(graph_source_dir / "vehicle_timeline.csv")
    copied_timeline_src: Optional[Path] = None
    for src in timeline_candidates:
        if src.exists():
            shutil.copy2(src, run_dir / "vehicle_timeline_gantt.csv")
            shutil.copy2(src, run_dir / "vehicle_timelines.csv")
            copied_timeline_src = src
            break
    if copied_timeline_src is not None:
        try:
            with copied_timeline_src.open("r", encoding="utf-8", newline="") as handle:
                timeline_rows = list(csv.DictReader(handle))
            (run_dir / "vehicle_timelines.json").write_text(
                json.dumps(_canonical_vehicle_timelines_payload(timeline_rows), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except (OSError, csv.Error, TypeError, ValueError) as exc:
            logger.warning(
                "Failed to build vehicle_timelines.json from %s: %s",
                copied_timeline_src,
                exc,
                exc_info=True,
            )

    refuel_rows = []
    charging_rows = []
    if canonical_solver_result is not None:
        for item in list((canonical_solver_result.get("charging_schedule") or [])):
            if not isinstance(item, dict):
                continue
            charge_kw = float(item.get("charge_kw") or 0.0)
            discharge_kw = float(item.get("discharge_kw") or 0.0)
            charging_rows.append(
                {
                    "vehicle_id": item.get("vehicle_id"),
                    "charger_id": item.get("charger_id"),
                    "time_idx": item.get("slot_index"),
                    "z_charge": 1 if charge_kw > 1.0e-9 else 0,
                    "p_charge_kw": charge_kw,
                    "p_discharge_kw": discharge_kw,
                    "soc_kwh": "",
                    "charging_depot_id": item.get("charging_depot_id"),
                }
            )
        for item in list((canonical_solver_result.get("refueling_schedule") or [])):
            if not isinstance(item, dict):
                continue
            refuel_rows.append(
                {
                    "vehicle_id": item.get("vehicle_id"),
                    "slot_index": item.get("slot_index"),
                    "time_hhmm": item.get("time_hhmm"),
                    "refuel_liters": item.get("refuel_liters"),
                    "unit": "L",
                }
            )
    _write_csv_rows(
        run_dir / "charging_schedule.csv",
        charging_rows,
        [
            "vehicle_id",
            "charger_id",
            "time_idx",
            "z_charge",
            "p_charge_kw",
            "p_discharge_kw",
            "soc_kwh",
            "charging_depot_id",
        ],
    )
    _write_csv_rows(
        run_dir / "refuel_events.csv",
        refuel_rows,
        ["vehicle_id", "slot_index", "time_hhmm", "refuel_liters", "unit"],
    )

    charging_summary_payload = dict(charging_summary or optimization_result.get("charging_summary") or {})
    charging_flow_rows = list((charging_flow_payload or {}).get("rows") or [])
    if charging_summary_payload:
        (run_dir / "charging_summary.json").write_text(
            json.dumps(charging_summary_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        charging_source_provenance = {
            "schema_version": "charging_source_provenance_v1",
            "site_depot_timestep": {
                "exact": bool(
                    charging_summary_payload.get(
                        "depot_source_provenance_exact",
                        charging_summary_payload.get("source_provenance_exact", False),
                    )
                ),
                "scope": str(
                    charging_summary_payload.get(
                        "source_provenance_scope", "depot_timestep"
                    )
                ),
                "note": charging_summary_payload.get("source_provenance_note"),
            },
            "vehicle_timestep": {
                "exact": bool(
                    charging_summary_payload.get(
                        "vehicle_source_provenance_exact", False
                    )
                ),
                "allocation_method": str(
                    charging_summary_payload.get(
                        "vehicle_source_allocation_method",
                        "proportional_by_depot_timestep",
                    )
                ),
                "note": charging_summary_payload.get(
                    "vehicle_source_precision_note"
                ),
            },
            "interpretation": (
                "Site/depot source totals and vehicle source allocations have "
                "different precision scopes and must not be conflated."
            ),
        }
        (run_dir / "charging_source_provenance.json").write_text(
            json.dumps(charging_source_provenance, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        charging_summary_rows: List[Dict[str, Any]] = []
        for depot_row in list(charging_summary_payload.get("depots") or []):
            if isinstance(depot_row, dict):
                charging_summary_rows.append({"scope": "depot", **dict(depot_row)})
        totals_row = dict(charging_summary_payload.get("totals") or {})
        if totals_row:
            charging_summary_rows.append({"scope": "total", "depot_id": "", **totals_row})
        _write_csv_rows(
            run_dir / "charging_summary.csv",
            charging_summary_rows,
            [
                "scope",
                "depot_id",
                "source_provenance_exact",
                "source_provenance_scope",
                "depot_source_provenance_exact",
                "vehicle_source_provenance_exact",
                "vehicle_source_allocation_method",
                "grid_to_bus_kwh",
                "pv_to_bus_kwh",
                "bess_to_bus_kwh",
                "pv_to_bess_kwh",
                "grid_to_bess_kwh",
                "pv_curtail_kwh",
                "pv_generation_kwh",
                "pv_utilization_rate",
                "grid_import_total_kwh",
                "total_bus_charge_kwh",
                "total_bess_charge_kwh",
                "bess_initial_soc_kwh",
                "bess_final_soc_kwh",
                "bess_terminal_soc_min_kwh",
                "bess_terminal_soc_violation_kwh",
                "peak_grid_import_kw",
                "peak_grid_import_kw_any_depot",
                "peak_grid_import_kw_all_depots",
                "peak_total_charge_kw",
                "peak_total_charge_kw_any_depot",
                "peak_total_charge_kw_all_depots",
                "contract_limit_kw",
                "contract_over_limit_kwh",
                "contract_over_limit_kw_peak",
                "contract_over_limit_slot_count",
                "contract_limit_exceeded",
                "contract_overage_penalty_enabled",
                "contract_overage_penalty_yen_per_kwh",
                "contract_overage_cost_jpy",
                "demand_charge_cost_jpy",
                "grid_purchase_cost_jpy",
                "bess_discharge_cost_jpy",
                "electricity_cost_jpy",
            ],
        )

    if charging_flow_rows:
        (run_dir / "depot_energy_flows.json").write_text(
            json.dumps(
                {
                    "rows": charging_flow_rows,
                    "summary": charging_summary_payload,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        _write_csv_rows(
            run_dir / "depot_energy_flows.csv",
            charging_flow_rows,
            [
                "depot_id",
                "slot_index",
                "grid_to_bus_kwh",
                "pv_to_bus_kwh",
                "bess_to_bus_kwh",
                "pv_to_bess_kwh",
                "grid_to_bess_kwh",
                "pv_curtail_kwh",
                "pv_curtail_raw_kwh",
                "pv_generation_kwh",
                "pv_balance_residual_kwh",
                "bess_soc_kwh",
                "bess_soc_start_kwh",
                "bess_soc_end_kwh",
                "bess_terminal_soc_target_kwh",
                "bess_terminal_soc_deviation_kwh",
                "bess_terminal_soc_delta_kwh",
                "grid_import_total_kwh",
                "grid_import_for_contract_kwh",
                "grid_import_kw",
                "grid_import_for_contract_kw",
                "bus_charge_from_grid_kwh",
                "bus_charge_from_bess_kwh",
                "total_bus_charge_kwh",
                "total_bess_charge_kwh",
                "total_charge_kw",
                "contract_limit_kw",
                "contract_over_limit_kwh",
                "contract_over_limit_kw",
                "contract_limit_exceeded",
                "energy_price_yen_per_kwh",
                "demand_charge_window_flag",
                "source_provenance_exact",
                "depot_source_provenance_exact",
            ],
        )
    else:
        grid_to_bus_kwh = _evaluation_float(cost_breakdown.get("grid_to_bus_kwh"))
        bess_to_bus_kwh = _evaluation_float(cost_breakdown.get("bess_to_bus_kwh"))
        grid_to_bess_kwh = _evaluation_float(cost_breakdown.get("grid_to_bess_kwh"))
        grid_import_total_kwh = (
            grid_to_bus_kwh + grid_to_bess_kwh
            if grid_to_bus_kwh is not None and grid_to_bess_kwh is not None
            else None
        )
        fallback_rows = [
            {"metric": "grid_to_bus_kwh", "value": grid_to_bus_kwh, "unit": "kWh"},
            {"metric": "grid_to_bess_kwh", "value": grid_to_bess_kwh, "unit": "kWh"},
            {"metric": "grid_import_total_kwh", "value": grid_import_total_kwh, "unit": "kWh"},
            {"metric": "grid_import_for_contract_kwh", "value": grid_import_total_kwh, "unit": "kWh"},
            {"metric": "bus_charge_from_grid_kwh", "value": grid_to_bus_kwh, "unit": "kWh"},
            {"metric": "bus_charge_from_bess_kwh", "value": bess_to_bus_kwh, "unit": "kWh"},
        ]
        (run_dir / "depot_energy_flows.json").write_text(
            json.dumps({"rows": fallback_rows}, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        _write_csv_rows(
            run_dir / "depot_energy_flows.csv",
            fallback_rows,
            ["metric", "value", "unit"],
        )

    if charging_summary_payload:
        totals = dict(charging_summary_payload.get("totals") or {})

        def _site_result_value(key: str, *fallback_keys: str) -> Optional[float]:
            for candidate in (key, *fallback_keys):
                if candidate in totals:
                    return _evaluation_float(totals.get(candidate))
            return None

        site_rows = [
            {"metric": "grid_to_bus_kwh", "value": _site_result_value("grid_to_bus_kwh"), "unit": "kWh"},
            {"metric": "pv_to_bus_kwh", "value": _site_result_value("pv_to_bus_kwh"), "unit": "kWh"},
            {"metric": "bess_to_bus_kwh", "value": _site_result_value("bess_to_bus_kwh"), "unit": "kWh"},
            {"metric": "pv_to_bess_kwh", "value": _site_result_value("pv_to_bess_kwh"), "unit": "kWh"},
            {"metric": "grid_to_bess_kwh", "value": _site_result_value("grid_to_bess_kwh"), "unit": "kWh"},
            {"metric": "pv_curtail_kwh", "value": _site_result_value("pv_curtail_kwh"), "unit": "kWh"},
            {"metric": "pv_generation_kwh", "value": _site_result_value("pv_generation_kwh"), "unit": "kWh"},
            {"metric": "pv_utilization_rate", "value": _site_result_value("pv_utilization_rate"), "unit": "ratio"},
            {"metric": "bess_terminal_soc_violation_kwh", "value": _site_result_value("bess_terminal_soc_violation_kwh"), "unit": "kWh"},
            {"metric": "grid_import_total_kwh", "value": _site_result_value("grid_import_total_kwh"), "unit": "kWh"},
            {"metric": "grid_import_for_contract_kwh", "value": _site_result_value("grid_import_for_contract_kwh", "grid_import_total_kwh"), "unit": "kWh"},
            {"metric": "bus_charge_from_grid_kwh", "value": _site_result_value("bus_charge_from_grid_kwh", "grid_to_bus_kwh"), "unit": "kWh"},
            {"metric": "bus_charge_from_bess_kwh", "value": _site_result_value("bus_charge_from_bess_kwh", "bess_to_bus_kwh"), "unit": "kWh"},
            {"metric": "peak_grid_import_kw_all_depots", "value": _site_result_value("peak_grid_import_kw_all_depots"), "unit": "kW"},
            {"metric": "contract_over_limit_kwh", "value": _site_result_value("contract_over_limit_kwh"), "unit": "kWh"},
            {"metric": "contract_overage_cost_jpy", "value": _site_result_value("contract_overage_cost_jpy"), "unit": "JPY"},
            {"metric": "demand_charge_cost_jpy", "value": _site_result_value("demand_charge_cost_jpy"), "unit": "JPY"},
            {"metric": "grid_purchase_cost_jpy", "value": _site_result_value("grid_purchase_cost_jpy"), "unit": "JPY"},
            {"metric": "bess_discharge_cost_jpy", "value": _site_result_value("bess_discharge_cost_jpy"), "unit": "JPY"},
            {"metric": "electricity_cost_jpy", "value": _site_result_value("electricity_cost_jpy"), "unit": "JPY"},
            {"metric": "contract_limit_exceeded", "value": bool(totals.get("contract_limit_exceeded")) if evaluation_valid else None, "unit": "flag"},
        ]
    else:
        grid_to_bus_kwh = _evaluation_float(cost_breakdown.get("grid_to_bus_kwh"))
        bess_to_bus_kwh = _evaluation_float(cost_breakdown.get("bess_to_bus_kwh"))
        grid_to_bess_kwh = _evaluation_float(cost_breakdown.get("grid_to_bess_kwh"))
        grid_import_total_kwh = (
            grid_to_bus_kwh + grid_to_bess_kwh
            if grid_to_bus_kwh is not None and grid_to_bess_kwh is not None
            else None
        )
        site_rows = [
            {"metric": "grid_to_bus_kwh", "value": grid_to_bus_kwh, "unit": "kWh"},
            {"metric": "grid_to_bess_kwh", "value": grid_to_bess_kwh, "unit": "kWh"},
            {"metric": "grid_import_total_kwh", "value": grid_import_total_kwh, "unit": "kWh"},
            {"metric": "grid_import_for_contract_kwh", "value": grid_import_total_kwh, "unit": "kWh"},
            {"metric": "bus_charge_from_grid_kwh", "value": grid_to_bus_kwh, "unit": "kWh"},
            {"metric": "bus_charge_from_bess_kwh", "value": bess_to_bus_kwh, "unit": "kWh"},
        ]
    _write_csv_rows(run_dir / "site_power_balance.csv", site_rows, ["metric", "value", "unit"])

    kpi_summary = {
        "total_cost_jpy": float(cost_breakdown.get("total_cost", 0.0) or 0.0),
        "objective_value_jpy": float(optimization_result.get("objective_value", 0.0) or 0.0),
        "objective_is_actual_cost": bool(cost_breakdown.get("objective_is_actual_cost", False)),
        "supports_exact_milp": bool((optimization_result.get("solver_metadata") or {}).get("supports_exact_milp", False)),
        "electricity_cost_jpy": float(
            cost_breakdown.get("electricity_cost", cost_breakdown.get("electricity_cost_final", 0.0))
            or 0.0
        ),
        "fuel_cost_jpy": float(cost_breakdown.get("fuel_cost", 0.0) or 0.0),
        "fuel_cost_final_jpy": float(cost_breakdown.get("fuel_cost_final", cost_breakdown.get("fuel_cost", 0.0)) or 0.0),
        "fuel_cost_final_source": str(cost_breakdown.get("fuel_cost_final_source", "provisional_distance_based") or "provisional_distance_based"),
        "fuel_cost_provisional_jpy": float(cost_breakdown.get("fuel_cost_provisional", cost_breakdown.get("provisional_ice_drive_cost", 0.0)) or 0.0),
        "fuel_cost_refueled_jpy": float(cost_breakdown.get("fuel_cost_refueled", cost_breakdown.get("realized_ice_refuel_cost", 0.0)) or 0.0),
        "fuel_cost_provisional_leftover_jpy": float(cost_breakdown.get("fuel_cost_provisional_leftover", cost_breakdown.get("leftover_ice_provisional_cost", 0.0)) or 0.0),
        "propulsion_energy_cost_jpy": float(cost_breakdown.get("energy_cost", 0.0) or 0.0),
        "pv_self_consumption_cost_jpy": float(cost_breakdown.get("pv_self_consumption_cost_jpy", 0.0) or 0.0),
        "pv_marginal_charge_cost_yen_per_kwh": float(
            cost_breakdown.get("pv_marginal_charge_cost_yen_per_kwh", 0.0) or 0.0
        ),
        "weather_strategy_objective_term_jpy_equivalent": float(
            cost_breakdown.get("weather_strategy_objective_term_jpy_equivalent", 0.0) or 0.0
        ),
        "grid_import_total_kwh": float(
            dict((charging_summary_payload or {}).get("totals") or {}).get("grid_import_total_kwh", 0.0)
            or (float(cost_breakdown.get("grid_to_bus_kwh", 0.0) or 0.0) + float(cost_breakdown.get("grid_to_bess_kwh", 0.0) or 0.0))
        ),
        "grid_import_for_contract_kwh": float(
            dict((charging_summary_payload or {}).get("totals") or {}).get("grid_import_for_contract_kwh", 0.0)
            or dict((charging_summary_payload or {}).get("totals") or {}).get("grid_import_total_kwh", 0.0)
            or (float(cost_breakdown.get("grid_to_bus_kwh", 0.0) or 0.0) + float(cost_breakdown.get("grid_to_bess_kwh", 0.0) or 0.0))
        ),
        "bus_charge_from_grid_kwh": float(
            dict((charging_summary_payload or {}).get("totals") or {}).get("bus_charge_from_grid_kwh", 0.0)
            or float(cost_breakdown.get("grid_to_bus_kwh", 0.0) or 0.0)
        ),
        "bus_charge_from_bess_kwh": float(
            dict((charging_summary_payload or {}).get("totals") or {}).get("bus_charge_from_bess_kwh", 0.0)
            or float(cost_breakdown.get("bess_to_bus_kwh", 0.0) or 0.0)
        ),
        "served_trip_count": int(summary_payload.get("trip_count_served") or 0),
        "unserved_trip_count": int(summary_payload.get("trip_count_unserved") or 0),
        "solver_runtime_sec": float(optimization_result.get("solve_time_seconds") or 0.0),
        "time_limit_seconds_requested": solver_settings.get("time_limit_seconds_requested"),
        "time_limit_seconds_effective": solver_settings.get("time_limit_seconds_effective"),
        "mip_gap_requested_ratio": solver_settings.get("mip_gap_requested_ratio"),
        "mip_gap_requested_percent": solver_settings.get("mip_gap_requested_percent"),
        "mip_gap_achieved_ratio": solver_settings.get("mip_gap_achieved_ratio"),
        "mip_gap_achieved_percent": solver_settings.get("mip_gap_achieved_percent"),
        "stage1_solver_status": solver_settings.get("stage1_solver_status"),
        "stage1_termination_reason": solver_settings.get(
            "stage1_termination_reason"
        ),
        "stage1_best_obj_stop_enabled": solver_settings.get(
            "stage1_best_obj_stop_enabled"
        ),
        "stage1_best_obj_stop_applied": solver_settings.get(
            "stage1_best_obj_stop_applied"
        ),
        "stage1_best_obj_stop_threshold": solver_settings.get(
            "stage1_best_obj_stop_threshold"
        ),
        "stage1_gurobi_raw_mip_gap_ratio": solver_settings.get(
            "stage1_gurobi_raw_mip_gap_ratio"
        ),
        "stage1_certified_mip_gap_ratio": solver_settings.get(
            "stage1_certified_mip_gap_ratio"
        ),
        "runtime_comparison_eligible": solver_settings.get(
            "runtime_comparison_eligible"
        ),
        "gurobi_threads": solver_settings.get("gurobi_threads"),
    }
    if accounting_summary:
        def _prefer_accounting_value(key: str, fallback: Any) -> Any:
            value = accounting_summary.get(key, None)
            if value in (None, "", 0, 0.0):
                return fallback
            return value

        kpi_summary.update(
            {
                "total_cost_jpy": float(accounting_summary.get("total_cost_jpy", kpi_summary["total_cost_jpy"]) or 0.0),
                "accounting_total_cost_jpy": accounting_summary.get(
                    "accounting_total_cost_jpy", accounting_summary.get("total_cost_jpy")
                ),
                "objective_value_jpy": float(accounting_summary.get("objective_value_jpy", kpi_summary["objective_value_jpy"]) or 0.0),
                "solver_objective_value": accounting_summary.get(
                    "solver_objective_value", accounting_summary.get("objective_value_jpy")
                ),
                "validated_operating_cost_jpy": accounting_summary.get(
                    "validated_operating_cost_jpy"
                ),
                "energy_cost_jpy": float(accounting_summary.get("energy_cost_jpy", kpi_summary["electricity_cost_jpy"]) or 0.0),
                "demand_cost_jpy": float(
                    accounting_summary.get(
                        "demand_cost_jpy",
                        kpi_summary.get("demand_charge_cost_jpy", kpi_summary.get("demand_cost_jpy", 0.0)),
                    )
                    or 0.0
                ),
                "fuel_cost_jpy": float(_prefer_accounting_value("fuel_cost_jpy", kpi_summary["fuel_cost_jpy"]) or 0.0),
                "co2_cost_jpy": float(_prefer_accounting_value("co2_cost_jpy", kpi_summary.get("co2_cost_jpy", 0.0)) or 0.0),
                "battery_degradation_cost_jpy": float(_prefer_accounting_value("battery_degradation_cost_jpy", kpi_summary.get("battery_degradation_cost_jpy", 0.0)) or 0.0),
                "contract_overage_cost_jpy": float(_prefer_accounting_value("contract_overage_cost_jpy", kpi_summary.get("contract_overage_cost_jpy", 0.0)) or 0.0),
                "served_trip_count": int(accounting_summary.get("served_trip_count", kpi_summary["served_trip_count"]) or 0),
                "unserved_trip_count": int(accounting_summary.get("unserved_trip_count", kpi_summary["unserved_trip_count"]) or 0),
                "bev_trip_count": int(accounting_summary.get("bev_trip_count", kpi_summary.get("bev_trip_count", 0)) or 0),
                "ice_trip_count": int(accounting_summary.get("ice_trip_count", kpi_summary.get("ice_trip_count", 0)) or 0),
                "used_vehicle_count": int(accounting_summary.get("used_vehicle_count", kpi_summary.get("used_vehicle_count", 0)) or 0),
                "available_vehicle_count": int(accounting_summary.get("available_vehicle_count", kpi_summary.get("available_vehicle_count", 0)) or 0),
                "vehicle_utilization_ratio": float(accounting_summary.get("vehicle_utilization_ratio", kpi_summary.get("vehicle_utilization_ratio", 0.0)) or 0.0),
                "service_km": float(accounting_summary.get("service_km", kpi_summary.get("service_km", 0.0)) or 0.0),
                "deadhead_before_km": float(accounting_summary.get("deadhead_before_km", kpi_summary.get("deadhead_before_km", 0.0)) or 0.0),
                "deadhead_after_km": float(accounting_summary.get("deadhead_after_km", kpi_summary.get("deadhead_after_km", 0.0)) or 0.0),
                "deadhead_total_km": float(accounting_summary.get("deadhead_total_km", kpi_summary.get("deadhead_total_km", 0.0)) or 0.0),
                "pv_generation_kwh": float(accounting_summary.get("pv_generation_kwh", kpi_summary.get("pv_generation_total_kwh", 0.0)) or 0.0),
                "pv_to_bus_kwh": float(accounting_summary.get("pv_to_bus_kwh", kpi_summary.get("pv_to_bus_kwh", 0.0)) or 0.0),
                "pv_to_bess_kwh": float(accounting_summary.get("pv_to_bess_kwh", kpi_summary.get("pv_to_bess_kwh", 0.0)) or 0.0),
                "bess_to_bus_kwh": float(accounting_summary.get("bess_to_bus_kwh", kpi_summary.get("bess_to_bus_kwh", 0.0)) or 0.0),
                "pv_curtailed_kwh": float(accounting_summary.get("pv_curtailed_kwh", kpi_summary.get("pv_curtail_kwh", 0.0)) or 0.0),
                "pv_utilization_ratio": float(accounting_summary.get("pv_utilization_ratio", kpi_summary.get("pv_utilization_ratio", 0.0)) or 0.0),
                "grid_to_bus_kwh": float(accounting_summary.get("grid_to_bus_kwh", kpi_summary.get("grid_to_bus_kwh", 0.0)) or 0.0),
                "grid_to_bess_kwh": float(accounting_summary.get("grid_to_bess_kwh", kpi_summary.get("grid_to_bess_kwh", 0.0)) or 0.0),
                "grid_total_kwh": float(accounting_summary.get("grid_total_kwh", kpi_summary.get("grid_import_total_kwh", 0.0)) or 0.0),
                "peak_grid_kw": float(accounting_summary.get("peak_grid_kw", kpi_summary.get("peak_grid_import_kw_all_depots", 0.0)) or 0.0),
                "total_charge_input_kwh": float(accounting_summary.get("total_charge_input_kwh", kpi_summary.get("total_charge_input_kwh", 0.0)) or 0.0),
                "min_soc_ratio": float(accounting_summary.get("min_soc_ratio", kpi_summary.get("min_soc_pct", 0.0)) or 0.0),
                "mean_soc_ratio": float(accounting_summary.get("mean_soc_ratio", kpi_summary.get("average_soc_pct", 0.0)) or 0.0),
                "final_min_soc_ratio": float(accounting_summary.get("final_min_soc_ratio", kpi_summary.get("min_soc_pct", 0.0)) or 0.0),
                "final_mean_soc_ratio": float(accounting_summary.get("final_mean_soc_ratio", kpi_summary.get("average_soc_pct", 0.0)) or 0.0),
                "objective_is_actual_cost": bool(accounting_summary.get("objective_is_actual_cost", kpi_summary.get("objective_is_actual_cost", False))),
                "supports_exact_milp": bool(accounting_summary.get("supports_exact_milp", kpi_summary.get("supports_exact_milp", False))),
                "fallback_applied": bool(accounting_summary.get("fallback_applied", kpi_summary.get("fallback_applied", False))),
                # Accounting's legacy flag is vehicle-level: it is false for
                # proportional post-allocation even when the site flow is exact.
                "vehicle_source_provenance_exact": bool(
                    accounting_summary.get(
                        "vehicle_source_provenance_exact",
                        accounting_summary.get(
                            "charging_source_provenance_exact",
                            kpi_summary.get("vehicle_source_provenance_exact", False),
                        ),
                    )
                ),
                "vehicle_source_allocation_method": str(
                    accounting_summary.get(
                        "vehicle_charging_source_allocation_method",
                        kpi_summary.get(
                            "vehicle_source_allocation_method",
                            "proportional_by_depot_timestep",
                        ),
                    )
                    or "proportional_by_depot_timestep"
                ),
                "contract_power_kw": float(accounting_summary.get("contract_power_kw", kpi_summary.get("contract_power_kw", 0.0)) or 0.0),
                "contract_power_exceeded": bool(accounting_summary.get("contract_power_exceeded", kpi_summary.get("contract_power_exceeded", False))),
                "contract_overage_kw": float(accounting_summary.get("contract_overage_kw", kpi_summary.get("contract_overage_kw", 0.0)) or 0.0),
                "contract_power_mode": str(accounting_summary.get("contract_power_mode", kpi_summary.get("contract_power_mode", "report_only")) or "report_only"),
            }
        )
    if charging_summary_payload:
        totals = dict(charging_summary_payload.get("totals") or {})
        depot_source_provenance_exact = bool(
            charging_summary_payload.get(
                "depot_source_provenance_exact",
                charging_summary_payload.get("source_provenance_exact", False),
            )
        )
        vehicle_source_provenance_exact = bool(
            charging_summary_payload.get(
                "vehicle_source_provenance_exact",
                kpi_summary.get("vehicle_source_provenance_exact", False),
            )
        )
        kpi_summary.update(
            {
                "pv_to_bus_kwh": float(totals.get("pv_to_bus_kwh", 0.0) or 0.0),
                "bess_to_bus_kwh": float(totals.get("bess_to_bus_kwh", 0.0) or 0.0),
                "grid_to_bess_kwh": float(totals.get("grid_to_bess_kwh", 0.0) or 0.0),
                "grid_import_for_contract_kwh": float(totals.get("grid_import_for_contract_kwh", totals.get("grid_import_total_kwh", 0.0)) or 0.0),
                "bus_charge_from_grid_kwh": float(totals.get("bus_charge_from_grid_kwh", totals.get("grid_to_bus_kwh", 0.0)) or 0.0),
                "bus_charge_from_bess_kwh": float(totals.get("bus_charge_from_bess_kwh", totals.get("bess_to_bus_kwh", 0.0)) or 0.0),
                "pv_to_bess_kwh": float(totals.get("pv_to_bess_kwh", 0.0) or 0.0),
                "contract_over_limit_kwh": float(totals.get("contract_over_limit_kwh", 0.0) or 0.0),
                "contract_overage_cost_jpy": float(totals.get("contract_overage_cost_jpy", 0.0) or 0.0),
                "demand_charge_cost_jpy": float(totals.get("demand_charge_cost_jpy", 0.0) or 0.0),
                "peak_grid_import_kw_all_depots": float(totals.get("peak_grid_import_kw_all_depots", 0.0) or 0.0),
                "contract_limit_exceeded": bool(totals.get("contract_limit_exceeded", False)),
                "depot_source_provenance_exact": depot_source_provenance_exact,
                "vehicle_source_provenance_exact": vehicle_source_provenance_exact,
                "vehicle_source_allocation_method": str(
                    charging_summary_payload.get(
                        "vehicle_source_allocation_method",
                        kpi_summary.get(
                            "vehicle_source_allocation_method",
                            "proportional_by_depot_timestep",
                        ),
                    )
                    or "proportional_by_depot_timestep"
                ),
                # This combined flag is only true when neither the site nor
                # per-vehicle source breakdown is inferred.
                "charging_source_provenance_exact": bool(
                    depot_source_provenance_exact
                    and vehicle_source_provenance_exact
                ),
                "charging_source_provenance_scope": "site_and_vehicle",
            }
        )
    if not evaluation_valid:
        _invalidate_mapping_metrics(kpi_summary)
        kpi_summary.update(
            {
                "served_trip_count": int(result_summary.get("trip_count_served") or 0),
                "unserved_trip_count": int(result_summary.get("trip_count_unserved") or 0),
                "result_status": optimization_result.get("result_status") or "INVALID",
                "failure_stage": optimization_result.get("failure_stage")
                or "result_validation",
                "research_kpi_eligible": False,
            }
        )
    (run_dir / "kpi_summary.json").write_text(
        json.dumps(kpi_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    reporting_finalizer_result: Optional[Dict[str, Any]] = None
    optimization_audit["experiment_report"] = {
        "status": "not_requested",
        "reason": "Reporting finalization was not requested for this run.",
    }
    if finalize_reporting:
        reporting_finalizer_result = _run_reporting_finalizer(run_dir)
        graph_artifacts = dict(optimization_result.get("graph_artifacts") or {})
        graph_artifacts["reporting_finalizer"] = reporting_finalizer_result
        optimization_result["graph_artifacts"] = graph_artifacts
        if reporting_finalizer_result.get("status") != "completed":
            warning = (
                "Reporting artifact finalization failed after optimization outputs were written: "
                f"{reporting_finalizer_result.get('error') or 'unknown error'}"
            )
            optimization_result["warnings"] = list(
                dict.fromkeys([*list(optimization_result.get("warnings") or []), warning])
            )
            optimization_audit["warnings"] = list(
                dict.fromkeys([*list(optimization_audit.get("warnings") or []), warning])
            )
        (run_dir / "optimization_result.json").write_text(
            json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (run_dir / "optimization_audit.json").write_text(
            json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (raw_dir / "optimization_result.json").write_text(
            json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (raw_dir / "optimization_audit.json").write_text(
            json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        if reporting_finalizer_result.get("status") == "completed":
            try:
                finalized_accounting = _finalized_accounting_summary_for_experiment_report(
                    run_dir
                )
                summary = _synchronize_finalized_accounting_summary(
                    run_dir=run_dir,
                    summary=summary,
                    finalized_accounting=finalized_accounting,
                )
                experiment_report = log_optimization_experiment(
                    scenario_id=str(optimization_result.get("scenario_id") or ""),
                    scenario_doc=scenario,
                    optimization_result=optimization_result,
                    accounting_summary_override=finalized_accounting,
                    git_commit_override=str(optimization_audit.get("git_sha") or "")
                    or None,
                )
                optimization_result["experiment_report"] = experiment_report
                source_json_path = experiment_report.get("json_path")
                source_md_path = experiment_report.get("md_path")
                local_json_path = None
                local_md_path = None
                if isinstance(source_json_path, str) and source_json_path.strip():
                    src_json = Path(source_json_path)
                    if src_json.exists():
                        local_json_path = "experiment_report.json"
                        shutil.copy2(src_json, run_dir / local_json_path)
                if isinstance(source_md_path, str) and source_md_path.strip():
                    src_md = Path(source_md_path)
                    if src_md.exists():
                        local_md_path = "experiment_report.md"
                        shutil.copy2(src_md, run_dir / local_md_path)
                optimization_audit["experiment_report"] = {
                    "status": "generated",
                    "experiment_id": experiment_report.get("experiment_id"),
                    "source_json_path": source_json_path,
                    "source_md_path": source_md_path,
                    "run_json_path": local_json_path,
                    "run_md_path": local_md_path,
                    "accounting_source": finalized_accounting.get(
                        "experiment_report_accounting_source"
                    ),
                    "accounting_reconciled": bool(
                        finalized_accounting.get("experiment_report_accounting_reconciled")
                    ),
                    "accounting_residual_jpy": finalized_accounting.get(
                        "experiment_report_accounting_residual_jpy"
                    ),
                }
            except Exception as exc:
                warning = f"Canonical experiment report generation failed: {exc}"
                optimization_result["warnings"] = list(
                    dict.fromkeys([*list(optimization_result.get("warnings") or []), warning])
                )
                optimization_audit["warnings"] = list(
                    dict.fromkeys([*list(optimization_audit.get("warnings") or []), warning])
                )
                optimization_audit["experiment_report"] = {
                    "status": "failed",
                    "error": str(exc),
                }
                raise RuntimeError(warning) from exc
        else:
            optimization_audit["experiment_report"] = {
                "status": "not_generated",
                "reason": "Canonical reporting finalization failed.",
                "reporting_finalizer_status": reporting_finalizer_result.get("status"),
            }

        (run_dir / "optimization_result.json").write_text(
            json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (run_dir / "optimization_audit.json").write_text(
            json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (raw_dir / "optimization_result.json").write_text(
            json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (raw_dir / "optimization_audit.json").write_text(
            json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    try:
        _write_results_workbook(
            run_dir=run_dir,
            summary=summary,
            cost_rows=cost_rows,
        )
    except Exception as exc:
        if finalize_reporting:
            raise RuntimeError(
                f"Required results.xlsx generation failed: {exc}"
            ) from exc

    run_input_manifest_path = run_dir / RUN_INPUT_MANIFEST_FILE
    run_input_manifest = {}
    if run_input_manifest_path.is_file():
        run_input_manifest = json.loads(
            run_input_manifest_path.read_text(encoding="utf-8")
        )
    solver_metadata = dict(optimization_result.get("solver_metadata") or {})
    input_git_provenance = dict(run_input_manifest.get("code_provenance") or {})
    git_provenance = {
        "git_sha": solver_metadata.get("git_sha", input_git_provenance.get("git_sha")),
        "git_dirty": solver_metadata.get("git_dirty", input_git_provenance.get("git_dirty")),
        "git_state_available": bool(
            solver_metadata.get(
                "git_state_available",
                input_git_provenance.get("git_state_available", False),
            )
        ),
        "git_state_error": solver_metadata.get(
            "git_state_error",
            input_git_provenance.get("git_state_error"),
        ),
    }
    rolling_execution = _rolling_execution_evidence(
        run_dir=run_dir,
        solver_metadata=solver_metadata,
    )
    objective_accounting_reconciliation = (
        _solver_objective_accounting_reconciliation_payload(
            run_dir=run_dir,
            optimization_result=optimization_result,
        )
    )
    (run_dir / SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE).write_text(
        json.dumps(
            objective_accounting_reconciliation,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    composition_search_certificate = _read_optional_json_mapping(
        run_dir / STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE
    )
    research_claim_scope = _research_claim_scope_payload(
        optimization_result=optimization_result,
        solver_settings=solver_settings,
        weather_policy=weather_policy,
        rolling_execution=rolling_execution,
        objective_accounting_reconciliation=objective_accounting_reconciliation,
        composition_search_certificate=composition_search_certificate,
    )
    claim_status = {
        "research_submission_ready": research_claim_scope.get(
            "research_submission_ready"
        ),
        "teacher_release_status": research_claim_scope.get(
            "teacher_release_status"
        ),
    }
    for key in ("diagnostic_only", "blocking_reason"):
        if key in research_claim_scope:
            claim_status[key] = research_claim_scope[key]
    optimization_result.update(claim_status)
    optimization_audit.update(claim_status)
    optimization_result["research_claim_scope"] = research_claim_scope
    optimization_audit["research_claim_scope"] = research_claim_scope
    (run_dir / "research_claim_scope.json").write_text(
        json.dumps(research_claim_scope, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    summary_path = run_dir / "summary.json"
    if summary_path.is_file():
        finalized_summary = json.loads(
            summary_path.read_text(encoding="utf-8")
        )
        if isinstance(finalized_summary, dict):
            finalized_summary.update(
                {
                    "run_profile": research_claim_scope.get("run_profile"),
                    "rolling_execution": rolling_execution,
                    "research_submission_ready": research_claim_scope.get(
                        "research_submission_ready"
                    ),
                    "teacher_release_status": research_claim_scope.get(
                        "teacher_release_status"
                    ),
                    "teacher_release_failed_checks": research_claim_scope.get(
                        "teacher_release_failed_checks"
                    ),
                    "mip_gap_target_met": solver_settings.get(
                        "mip_gap_target_met"
                    ),
                    **{
                        key: research_claim_scope[key]
                        for key in ("diagnostic_only", "blocking_reason")
                        if key in research_claim_scope
                    },
                }
            )
            summary_path.write_text(
                json.dumps(finalized_summary, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
    if finalize_reporting:
        _prepend_experiment_release_header(
            run_dir=run_dir,
            research_claim_scope=research_claim_scope,
            rolling_execution=rolling_execution,
            solver_settings=solver_settings,
            optimization_result=optimization_result,
        )
        _write_results_workbook_release_status(
            run_dir=run_dir,
            research_claim_scope=research_claim_scope,
            rolling_execution=rolling_execution,
            solver_settings=solver_settings,
            result_claim_classification=dict(
                optimization_result.get("result_claim_classification") or {}
            ),
        )
        final_cost_reconciliation = _assert_final_cost_artifact_consistency(
            run_dir=run_dir,
            optimization_result=optimization_result,
        )
        optimization_result[
            "final_cost_reconciliation"
        ] = final_cost_reconciliation
        optimization_audit[
            "final_cost_reconciliation"
        ] = final_cost_reconciliation
        if rolling_execution.get("status") == "executed_and_accepted":
            from bff.services.optimization_run.literature_figures import (
                generate_literature_figure_bundle,
            )

            literature_figure_bundle = generate_literature_figure_bundle(
                run_dir
            )
            literature_figure_summary = {
                "status": literature_figure_bundle.get("status"),
                "schema_version": literature_figure_bundle.get(
                    "schema_version"
                ),
                "figure_count": literature_figure_bundle.get("figure_count"),
                "diagnostic_only": literature_figure_bundle.get(
                    "diagnostic_only"
                ),
                "manifest": "graph/literature_figures/manifest.json",
            }
            optimization_result[
                "literature_figure_bundle"
            ] = literature_figure_summary
            optimization_audit[
                "literature_figure_bundle"
            ] = literature_figure_summary
        else:
            literature_figure_summary = {
                "status": "NOT_GENERATED",
                "reason": (
                    "literature figures require an executed and accepted "
                    "hourly rolling chain"
                ),
                "diagnostic_only": True,
            }
            optimization_result[
                "literature_figure_bundle"
            ] = literature_figure_summary
            optimization_audit[
                "literature_figure_bundle"
            ] = literature_figure_summary
    (run_dir / "optimization_result.json").write_text(
        json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (run_dir / "optimization_audit.json").write_text(
        json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (raw_dir / "optimization_result.json").write_text(
        json.dumps(optimization_result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (raw_dir / "optimization_audit.json").write_text(
        json.dumps(optimization_audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    run_manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "scenario_id": optimization_result.get("scenario_id"),
        "service_date": (
            (optimization_result.get("graph_artifacts") or {}).get("accounting_summary") or {}
        ).get("service_date")
        or ((scenario.get("simulation_config") or {}).get("service_date") or ""),
        **git_provenance,
        "research_run": bool(solver_metadata.get("research_run", False)),
        "research_run_accepted": bool(solver_metadata.get("research_run_accepted", False)),
        "research_submission_git_provenance_eligible": bool(
            solver_metadata.get("research_submission_git_provenance_eligible", False)
        ),
        "requested_phase": solver_metadata.get("requested_phase"),
        "resolved_phase": solver_metadata.get("resolved_phase"),
        "executed_phase": solver_metadata.get("executed_phase"),
        "files": sorted(
            [p.relative_to(run_dir).as_posix() for p in run_dir.rglob("*") if p.is_file()]
        ),
        "units": unit_map,
        "weather_proxy_enabled": bool(weather_policy.get("enabled")),
        "weather_proxy_version": (
            (weather_policy.get("forecast") or {}).get("version")
            if weather_policy.get("enabled")
            else None
        ),
        "weather_operation_mode": (
            (weather_policy.get("operation_profile") or {}).get("operation_mode")
            if weather_policy.get("enabled")
            else None
        ),
        "weather_analog_date": (
            (weather_policy.get("forecast") or {}).get("analog_date")
            if weather_policy.get("enabled")
            else None
        ),
        "weather_typical_class": (
            (weather_policy.get("representative_curve") or {}).get("typical_weather_class")
            if weather_policy.get("enabled")
            else None
        ),
        "weather_pv_curve_applied": (
            bool((weather_policy.get("audit") or {}).get("weather_pv_forecast_applied"))
            if weather_policy.get("enabled")
            else False
        ),
        "weather_pv_source_dates": (
            list((weather_policy.get("representative_curve") or {}).get("source_dates") or ())
            if weather_policy.get("enabled")
            else []
        ),
        "solver_settings": solver_settings,
        "simulation_conditions": condition_provenance,
        "run_input_provenance": {
            "status": "OK" if run_input_manifest else "MISSING",
            "manifest_path": (
                RUN_INPUT_MANIFEST_FILE if run_input_manifest else None
            ),
            "schema_version": run_input_manifest.get("schema_version"),
            "prepared_input_id": run_input_manifest.get("prepared_input_id"),
            "prepared_source_sha256": run_input_manifest.get(
                "prepared_source_sha256"
            ),
            "artifacts": dict(run_input_manifest.get("artifacts") or {}),
            "git_state_available": bool(
                input_git_provenance.get("git_state_available", False)
            ),
            "git_sha": input_git_provenance.get("git_sha"),
        },
        "rolling_execution": rolling_execution,
        "research_claim_scope": research_claim_scope,
        **claim_status,
        "teacher_release_failed_checks": research_claim_scope.get(
            "teacher_release_failed_checks"
        ),
        "experiment_report": dict(optimization_audit.get("experiment_report") or {}),
        "formal_phase3_weather_submission_readiness": {
            # A single run can clear its own day-ahead/rolling release gate, but
            # it cannot by itself establish the paired same-service-date
            # weather comparison contract.
            "ready": False,
            "reason": (
                "An accepted individual frontend run is not yet a formal "
                "Phase 3 weather comparison. A matched same-service-date pair "
                "and comparison audit are still required."
            ),
            "day_ahead_research_run_accepted": bool(
                solver_metadata.get("research_run_accepted", False)
            ),
            "git_provenance_eligible": bool(
                solver_metadata.get("research_submission_git_provenance_eligible", False)
            ),
            "rolling_execution_status": rolling_execution["status"],
        },
        "graph": {
            "manifest_path": "graph/manifest.json",
            "route_band_diagrams_manifest": str(
                graph_artifacts.get("manifest_path") or "graph/route_band_diagrams/manifest.json"
            ),
            "route_band_diagram_count": int(graph_artifacts.get("diagram_count") or 0),
            "vehicle_operation_diagrams_manifest": str(
                graph_artifacts.get("vehicle_operation_diagram_manifest_path") or ""
            ),
            "vehicle_operation_diagram_count": int(
                graph_artifacts.get("vehicle_operation_diagram_count") or 0
            ),
            "reporting_finalizer": reporting_finalizer_result
            or graph_artifacts.get("reporting_finalizer")
            or {},
        },
    }
    (run_dir / "run_manifest.json").write_text(
        json.dumps(run_manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return reporting_finalizer_result


def _enforce_frontend_run_artifact_contract(
    *,
    run_dir: Path,
    optimization_result: Dict[str, Any],
    optimization_audit: Dict[str, Any],
    reporting_finalizer_result: Optional[Dict[str, Any]],
    research_run: bool,
    require_rolling: bool,
) -> Dict[str, Any]:
    """Persist and enforce the final frontend output bundle contract.

    This must run after every other run-directory finalizer, including the
    Rolling manifest refresh. No run artifact may be changed after the second
    audit because the audit stores hashes of the final files.
    """

    run_dir = Path(run_dir)
    raw_dir = run_dir / "raw"
    solver_metadata = dict(optimization_result.get("solver_metadata") or {})
    require_two_stage_composition_certificate = bool(
        research_run
        and str(solver_metadata.get("optimization_structure") or "").lower()
        == "two_stage"
    )
    artifact_audit = persist_frontend_run_artifact_audit(
        run_dir,
        research_run=research_run,
        require_rolling=require_rolling,
        require_two_stage_composition_certificate=(
            require_two_stage_composition_certificate
        ),
    )
    artifact_audit_summary = {
        "status": artifact_audit.get("status"),
        "accepted": artifact_audit.get("accepted"),
        "artifact": "artifact_completeness.json",
        "required_artifact_count": artifact_audit.get(
            "required_artifact_count"
        ),
        "verified_artifact_count": artifact_audit.get(
            "verified_artifact_count"
        ),
        "missing_artifacts": artifact_audit.get("missing_artifacts"),
        "empty_artifacts": artifact_audit.get("empty_artifacts"),
        "content_errors": artifact_audit.get("content_errors"),
    }
    optimization_result["artifact_completeness"] = artifact_audit_summary
    optimization_audit["artifact_completeness"] = artifact_audit_summary
    for path, payload in (
        (run_dir / "optimization_result.json", optimization_result),
        (run_dir / "optimization_audit.json", optimization_audit),
        (raw_dir / "optimization_result.json", optimization_result),
        (raw_dir / "optimization_audit.json", optimization_audit),
    ):
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    reporting_result = dict(reporting_finalizer_result or {})
    reporting_result.update(
        {
            "artifact_completeness_status": artifact_audit.get("status"),
            "artifact_completeness_artifact": "artifact_completeness.json",
            "required_artifact_count": artifact_audit.get(
                "required_artifact_count"
            ),
            "verified_artifact_count": artifact_audit.get(
                "verified_artifact_count"
            ),
        }
    )
    run_manifest_path = run_dir / "run_manifest.json"
    run_manifest = json.loads(run_manifest_path.read_text(encoding="utf-8"))
    if not isinstance(run_manifest, dict):
        raise RuntimeError("run_manifest.json is not a JSON object")
    run_manifest["artifact_completeness"] = artifact_audit_summary
    graph_manifest = dict(run_manifest.get("graph") or {})
    graph_manifest["reporting_finalizer"] = reporting_result
    run_manifest["graph"] = graph_manifest
    run_manifest["files"] = sorted(
        path.relative_to(run_dir).as_posix()
        for path in run_dir.rglob("*")
        if path.is_file()
    )
    run_manifest_path.write_text(
        json.dumps(run_manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Re-audit after the result and run manifest record the gate. This second
    # pass is the immutable hash snapshot used for retrospective verification.
    artifact_audit = persist_frontend_run_artifact_audit(
        run_dir,
        research_run=research_run,
        require_rolling=require_rolling,
        require_two_stage_composition_certificate=(
            require_two_stage_composition_certificate
        ),
    )
    if artifact_audit.get("accepted") is not True:
        raise RuntimeError(
            "Required frontend run artifacts are incomplete: "
            + json.dumps(
                {
                    "missing_artifacts": artifact_audit.get(
                        "missing_artifacts"
                    ),
                    "empty_artifacts": artifact_audit.get(
                        "empty_artifacts"
                    ),
                    "invalid_json_artifacts": artifact_audit.get(
                        "invalid_json_artifacts"
                    ),
                    "content_errors": artifact_audit.get("content_errors"),
                    "workbook_errors": artifact_audit.get(
                        "workbook_errors"
                    ),
                    "run_manifest_errors": artifact_audit.get(
                        "run_manifest_errors"
                    ),
                },
                ensure_ascii=False,
                sort_keys=True,
            )
        )
    return reporting_result


def _run_reporting_finalizer(run_dir: Path) -> Dict[str, Any]:
    try:
        from src.reporting import rebuild_reporting_artifacts_in_place

        result = rebuild_reporting_artifacts_in_place(run_dir)
        return {
            "status": "completed",
            "updated_files": list(result.updated_files),
            "validation_status": dict(result.validation_status),
            "warnings": list(result.warnings),
        }
    except Exception as exc:
        return {
            "status": "failed",
            "error": str(exc),
            "traceback": traceback.format_exc(),
        }


def _finalized_accounting_summary_for_experiment_report(run_dir: Path) -> Dict[str, Any]:
    """Load the final accounting ledger used by the human experiment report.

    The reporting finalizer can rebuild the canonical ledger after the solver
    result is first assembled.  Reading its sidecar artifacts here prevents an
    early, provisional objective breakdown from being presented as the final
    accounting cost.
    """

    def _read_mapping(path: Path) -> Dict[str, Any]:
        if not path.is_file():
            return {}
        loaded = json.loads(path.read_text(encoding="utf-8"))
        return dict(loaded) if isinstance(loaded, dict) else {}

    kpi_summary = _read_mapping(run_dir / "kpi_summary.json")
    summary = _read_mapping(run_dir / "summary.json")
    ledger = _read_mapping(run_dir / "graph" / "canonical_cost_ledger.json")
    if not summary:
        raise ValueError("Finalized summary.json is missing")
    if not ledger:
        raise ValueError("Finalized graph/canonical_cost_ledger.json is missing")

    canonical = dict(kpi_summary)
    # summary.json is the finalizer's canonical source for the values that a
    # reviewer compares across runs.  It intentionally overrides the older
    # provisional values in kpi_summary.json.
    for key in (
        "objective_value_jpy",
        "total_co2_kg",
        "served_trip_count",
        "unserved_trip_count",
        "bev_trip_count",
        "ice_trip_count",
        "used_vehicle_count",
        "bus_charging_total_kwh",
        "peak_grid_import_kw",
        "objective_is_actual_cost",
    ):
        if key in summary:
            canonical[key] = summary[key]

    components = dict(ledger.get("components") or {})
    details = dict(ledger.get("details") or {})
    co2 = dict(ledger.get("co2") or {})
    required_component_keys = (
        "electricity_cost_jpy",
        "fuel_cost_jpy",
        "demand_charge_cost_jpy",
        "vehicle_usage_cost_jpy",
        "co2_cost_jpy",
    )
    missing_components = [
        key for key in required_component_keys if components.get(key) is None
    ]
    if missing_components:
        raise ValueError(
            "Finalized canonical cost ledger is missing components: "
            + ", ".join(missing_components)
        )
    required_ledger_keys = (
        "accounting_total_cost_jpy",
        "accounting_residual_jpy",
    )
    missing_ledger_keys = [
        key for key in required_ledger_keys if ledger.get(key) is None
    ]
    if missing_ledger_keys:
        raise ValueError(
            "Finalized canonical cost ledger is missing fields: "
            + ", ".join(missing_ledger_keys)
        )
    residual = float(ledger["accounting_residual_jpy"])
    tolerance = float(
        ledger.get("accounting_residual_tolerance_jpy", 1.0e-6) or 1.0e-6
    )
    if not math.isclose(residual, 0.0, abs_tol=tolerance):
        raise ValueError(
            "Finalized canonical cost ledger does not reconcile: "
            f"residual={residual}, tolerance={tolerance}"
        )

    accounting_total = float(ledger["accounting_total_cost_jpy"])
    canonical.update(
        {
            "total_cost_jpy": accounting_total,
            "accounting_total_cost_jpy": accounting_total,
            # ``summary.energy_cost_jpy`` is the established electricity-only
            # contract. Keep ICE fuel separate; callers that need combined
            # propulsion energy must opt into the explicitly named aggregate.
            "energy_cost_jpy": float(components["electricity_cost_jpy"]),
            "electricity_cost_jpy": float(
                components["electricity_cost_jpy"]
            ),
            "propulsion_energy_cost_jpy": (
                float(components["electricity_cost_jpy"])
                + float(components["fuel_cost_jpy"])
            ),
            "grid_purchase_cost_jpy": float(
                details.get("grid_purchase_cost_jpy", 0.0) or 0.0
            ),
            "bess_total_flow_cost_jpy": float(
                details.get("bess_total_flow_cost_jpy", 0.0) or 0.0
            ),
            "demand_charge_cost_jpy": float(
                components["demand_charge_cost_jpy"]
            ),
            "fuel_cost_jpy": float(components["fuel_cost_jpy"]),
            "co2_cost_jpy": float(components["co2_cost_jpy"]),
            "vehicle_usage_cost_jpy": float(
                components["vehicle_usage_cost_jpy"]
            ),
            "total_co2_kg": co2.get(
                "total_co2_kg", canonical.get("total_co2_kg")
            ),
            "canonical_cost_components_jpy": {
                str(key): float(value)
                for key, value in components.items()
            },
            "canonical_cost_component_status": dict(
                ledger.get("component_status") or {}
            ),
        }
    )
    canonical["experiment_report_accounting_reconciled"] = True
    canonical["experiment_report_accounting_residual_jpy"] = residual
    canonical["experiment_report_accounting_source"] = (
        "graph/canonical_cost_ledger.json"
    )
    return canonical


def _synchronize_finalized_accounting_summary(
    *,
    run_dir: Path,
    summary: Dict[str, Any],
    finalized_accounting: Dict[str, Any],
) -> Dict[str, Any]:
    """Make summary.json expose the same finalized ledger used by reports.

    The reporting finalizer may replace provisional day-ahead accounting after
    the initial summary has been written.  Keep the machine-readable summary
    aligned with the canonical ledger before final cross-artifact comparison.
    """

    required_keys = (
        "total_cost_jpy",
        "accounting_total_cost_jpy",
        "energy_cost_jpy",
        "electricity_cost_jpy",
        "propulsion_energy_cost_jpy",
        "fuel_cost_jpy",
        "demand_charge_cost_jpy",
        "vehicle_usage_cost_jpy",
        "co2_cost_jpy",
        "canonical_cost_components_jpy",
        "canonical_cost_component_status",
    )
    missing_keys = [
        key for key in required_keys if finalized_accounting.get(key) is None
    ]
    if missing_keys:
        raise ValueError(
            "Finalized accounting is missing summary fields: "
            + ", ".join(missing_keys)
        )

    summary_path = run_dir / "summary.json"
    persisted = json.loads(summary_path.read_text(encoding="utf-8"))
    if not isinstance(persisted, dict):
        raise ValueError("Finalized summary.json is not a JSON object")
    updates = {key: finalized_accounting[key] for key in required_keys}
    summary.update(updates)
    persisted.update(updates)
    summary_path.write_text(
        json.dumps(persisted, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return summary


def _should_finalize_reporting_after_rolling(
    rolling_technical_failure: Optional[BaseException],
) -> bool:
    """Return whether canonical reports have a complete rolling cost source.

    A failed rolling chain must surface its primary step error. Running the
    final accounting reconciler on an incomplete executed day would replace
    that actionable error with a secondary ``accounting not eligible`` error.
    """

    return rolling_technical_failure is None


def _assert_final_cost_artifact_consistency(
    *,
    run_dir: Path,
    optimization_result: Dict[str, Any],
    tolerance_jpy: float = 1.0e-6,
) -> Dict[str, Any]:
    """Fail the frontend job unless every final cost artifact agrees.

    Accepted rolling execution makes ``executed_day_accounting.json`` the only
    final cost source.  Day-ahead objective values remain available as solver
    telemetry but are not substituted for the executed accounting total.
    """

    executed_path = (
        run_dir / "rolling_hourly_chain" / "executed_day_accounting.json"
    )
    if not executed_path.is_file():
        return {
            "status": "SKIPPED",
            "reason": "no_executed_rolling_day",
        }

    def _json_object(path: Path) -> Dict[str, Any]:
        loaded = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(loaded, dict):
            raise RuntimeError(f"Final cost artifact is not a JSON object: {path}")
        return dict(loaded)

    executed = _json_object(executed_path)
    if executed.get("eligible") is not True:
        raise RuntimeError(
            "Executed-day accounting is not eligible: "
            f"reason={executed.get('reason')!r}, "
            f"rejection_reasons={list(executed.get('rejection_reasons') or ())!r}"
        )
    ledger = _json_object(run_dir / "graph" / "canonical_cost_ledger.json")
    if ledger.get("source") != (
        "rolling_hourly_chain/executed_day_accounting.json"
    ):
        raise RuntimeError(
            "Final canonical cost ledger does not identify executed rolling "
            f"accounting as its source: {ledger.get('source')!r}"
        )
    summary = _json_object(run_dir / "summary.json")
    experiment = _json_object(run_dir / "experiment_report.json")
    report_results = dict(experiment.get("results") or {})

    with (run_dir / "cost_breakdown_detail.csv").open(
        "r", encoding="utf-8", newline=""
    ) as handle:
        detail_rows = list(csv.DictReader(handle))
    detail_by_key = {
        str(row.get("key") or ""): row.get("value") for row in detail_rows
    }

    from openpyxl import load_workbook

    workbook = load_workbook(run_dir / "results.xlsx", data_only=True)
    if "cost_breakdown" not in workbook.sheetnames:
        raise RuntimeError("results.xlsx is missing cost_breakdown sheet")
    workbook_costs = {
        str(row[0].value or ""): row[1].value
        for row in workbook["cost_breakdown"].iter_rows(min_row=2)
        if row and row[0].value not in (None, "")
    }
    workbook.close()
    executed_breakdown = dict(executed.get("cost_breakdown") or {})
    ledger_components = dict(ledger.get("components") or {})
    pre_failures: Dict[str, Any] = {}

    def _required_component_value(
        mapping: Dict[str, Any],
        key: str,
        *,
        artifact: str,
        metric: str,
    ) -> Optional[float]:
        """Read a required monetary value without fabricating a fallback.

        A reconciliation artifact is diagnostic evidence.  In particular, a
        missing value must remain ``null`` in the persisted observation and
        residual maps instead of being replaced with ``0.0`` merely to make
        downstream arithmetic convenient.
        """
        if key not in mapping or mapping.get(key) is None:
            pre_failures[f"{metric}:{artifact}:missing"] = key
            return None
        try:
            value = float(mapping[key])
        except (TypeError, ValueError):
            pre_failures[f"{metric}:{artifact}:invalid"] = mapping[key]
            return None
        if not math.isfinite(value):
            pre_failures[f"{metric}:{artifact}:nonfinite"] = value
            return None
        return value

    executed_total = _required_component_value(
        executed_breakdown,
        "total_cost",
        artifact="executed_day_accounting",
        metric="total_cost_jpy",
    )
    expected_by_metric = {
        "total_cost_jpy": executed_total,
        "electricity_cost_jpy": _required_component_value(
            executed_breakdown,
            "electricity_cost",
            artifact="executed_day_accounting",
            metric="electricity_cost_jpy",
        ),
        "fuel_cost_jpy": _required_component_value(
            executed_breakdown,
            "fuel_cost",
            artifact="executed_day_accounting",
            metric="fuel_cost_jpy",
        ),
        "demand_charge_cost_jpy": _required_component_value(
            executed_breakdown,
            "demand_cost",
            artifact="executed_day_accounting",
            metric="demand_charge_cost_jpy",
        ),
        "vehicle_usage_cost_jpy": _required_component_value(
            executed_breakdown,
            "vehicle_usage_cost",
            artifact="executed_day_accounting",
            metric="vehicle_usage_cost_jpy",
        ),
        "co2_cost_jpy": _required_component_value(
            executed_breakdown,
            "co2_cost",
            artifact="executed_day_accounting",
            metric="co2_cost_jpy",
        ),
    }
    observed_by_metric = {
        "total_cost_jpy": {
            "executed_day_accounting": executed_total,
            "canonical_cost_ledger": _required_component_value(
                ledger,
                "accounting_total_cost_jpy",
                artifact="canonical_cost_ledger",
                metric="total_cost_jpy",
            ),
            "summary": _required_component_value(
                summary,
                "accounting_total_cost_jpy",
                artifact="summary",
                metric="total_cost_jpy",
            ),
            "experiment_report_json": _required_component_value(
                report_results,
                "total_cost_jpy",
                artifact="experiment_report_json",
                metric="total_cost_jpy",
            ),
            "cost_breakdown_detail": _required_component_value(
                detail_by_key,
                "total_cost",
                artifact="cost_breakdown_detail",
                metric="total_cost_jpy",
            ),
            "results_xlsx": _required_component_value(
                workbook_costs,
                "total_cost",
                artifact="results_xlsx",
                metric="total_cost_jpy",
            ),
            "optimization_result": _required_component_value(
                optimization_result,
                "final_accounting_total_cost_jpy",
                artifact="optimization_result",
                metric="total_cost_jpy",
            ),
        },
        "electricity_cost_jpy": {
            "canonical_cost_ledger": _required_component_value(
                ledger_components,
                "electricity_cost_jpy",
                artifact="canonical_cost_ledger",
                metric="electricity_cost_jpy",
            ),
            "summary": _required_component_value(
                summary,
                (
                    "electricity_cost_jpy"
                    if "electricity_cost_jpy" in summary
                    else "energy_cost_jpy"
                ),
                artifact="summary",
                metric="electricity_cost_jpy",
            ),
            "experiment_report_json": _required_component_value(
                report_results,
                "electricity_cost_jpy",
                artifact="experiment_report_json",
                metric="electricity_cost_jpy",
            ),
            "cost_breakdown_detail": _required_component_value(
                detail_by_key,
                "electricity_cost",
                artifact="cost_breakdown_detail",
                metric="electricity_cost_jpy",
            ),
            "results_xlsx": _required_component_value(
                workbook_costs,
                "electricity_cost",
                artifact="results_xlsx",
                metric="electricity_cost_jpy",
            ),
        },
        "fuel_cost_jpy": {
            "canonical_cost_ledger": _required_component_value(
                ledger_components,
                "fuel_cost_jpy",
                artifact="canonical_cost_ledger",
                metric="fuel_cost_jpy",
            ),
            "summary": _required_component_value(
                summary,
                "fuel_cost_jpy",
                artifact="summary",
                metric="fuel_cost_jpy",
            ),
            "experiment_report_json": _required_component_value(
                report_results,
                "diesel_cost_jpy",
                artifact="experiment_report_json",
                metric="fuel_cost_jpy",
            ),
            "cost_breakdown_detail": _required_component_value(
                detail_by_key,
                "fuel_cost",
                artifact="cost_breakdown_detail",
                metric="fuel_cost_jpy",
            ),
            "results_xlsx": _required_component_value(
                workbook_costs,
                "fuel_cost",
                artifact="results_xlsx",
                metric="fuel_cost_jpy",
            ),
        },
        "demand_charge_cost_jpy": {
            "canonical_cost_ledger": _required_component_value(
                ledger_components,
                "demand_charge_cost_jpy",
                artifact="canonical_cost_ledger",
                metric="demand_charge_cost_jpy",
            ),
            "summary": _required_component_value(
                summary,
                "demand_charge_cost_jpy",
                artifact="summary",
                metric="demand_charge_cost_jpy",
            ),
            "experiment_report_json": _required_component_value(
                report_results,
                "demand_charge_jpy",
                artifact="experiment_report_json",
                metric="demand_charge_cost_jpy",
            ),
            "cost_breakdown_detail": _required_component_value(
                detail_by_key,
                "demand_charge",
                artifact="cost_breakdown_detail",
                metric="demand_charge_cost_jpy",
            ),
            "results_xlsx": _required_component_value(
                workbook_costs,
                "demand_charge",
                artifact="results_xlsx",
                metric="demand_charge_cost_jpy",
            ),
        },
        "vehicle_usage_cost_jpy": {
            "canonical_cost_ledger": _required_component_value(
                ledger_components,
                "vehicle_usage_cost_jpy",
                artifact="canonical_cost_ledger",
                metric="vehicle_usage_cost_jpy",
            ),
            "experiment_report_json": _required_component_value(
                report_results,
                "vehicle_usage_cost_jpy",
                artifact="experiment_report_json",
                metric="vehicle_usage_cost_jpy",
            ),
            "cost_breakdown_detail": _required_component_value(
                detail_by_key,
                "vehicle_usage_cost",
                artifact="cost_breakdown_detail",
                metric="vehicle_usage_cost_jpy",
            ),
            "results_xlsx": _required_component_value(
                workbook_costs,
                "vehicle_usage_cost",
                artifact="results_xlsx",
                metric="vehicle_usage_cost_jpy",
            ),
        },
        "co2_cost_jpy": {
            "canonical_cost_ledger": _required_component_value(
                ledger_components,
                "co2_cost_jpy",
                artifact="canonical_cost_ledger",
                metric="co2_cost_jpy",
            ),
            "summary": _required_component_value(
                summary,
                "co2_cost_jpy",
                artifact="summary",
                metric="co2_cost_jpy",
            ),
            "experiment_report_json": _required_component_value(
                report_results,
                "co2_cost_jpy",
                artifact="experiment_report_json",
                metric="co2_cost_jpy",
            ),
            "cost_breakdown_detail": _required_component_value(
                detail_by_key,
                "co2_cost",
                artifact="cost_breakdown_detail",
                metric="co2_cost_jpy",
            ),
            "results_xlsx": _required_component_value(
                workbook_costs,
                "co2_cost",
                artifact="results_xlsx",
                metric="co2_cost_jpy",
            ),
        },
    }
    component_status = dict(ledger.get("component_status") or {})
    if not component_status:
        pre_failures["canonical_cost_ledger:component_status:missing"] = True
    else:
        summary_components = dict(
            summary.get("canonical_cost_components_jpy") or {}
        )
        report_components = dict(
            report_results.get("canonical_cost_components_jpy") or {}
        )
        optimization_components = dict(
            optimization_result.get("cost_breakdown") or {}
        )

        for component_key, (default_source_key, _flag_key) in (
            CANONICAL_LEDGER_COMPONENT_SOURCES.items()
        ):
            status = dict(component_status.get(component_key) or {})
            source_key = str(
                status.get("source_key") or default_source_key
            )
            enabled = status.get("enabled") is True
            status_label = str(status.get("status") or "")
            source_present = status.get("source_present") is True
            ledger_value = _required_component_value(
                ledger_components,
                component_key,
                artifact="canonical_cost_ledger",
                metric=component_key,
            )
            if not enabled:
                if status_label != "SKIPPED":
                    pre_failures[
                        f"{component_key}:component_status"
                    ] = status_label
                if ledger_value is not None and not math.isclose(
                    ledger_value,
                    0.0,
                    rel_tol=0.0,
                    abs_tol=float(tolerance_jpy),
                ):
                    pre_failures[
                        f"{component_key}:disabled_nonzero"
                    ] = ledger_value
            elif not source_present:
                pre_failures[
                    f"{component_key}:enabled_source_missing"
                ] = source_key

            # Disabled components have an expected ledger value of exactly
            # zero, but they still must be represented consistently by every
            # final artifact. Do not let ``SKIPPED`` bypass those comparisons.
            expected_by_metric[component_key] = ledger_value
            existing_observations = observed_by_metric.get(component_key, {})

            def _component_artifact_name(artifact: str) -> str:
                """Avoid overwriting a direct report observation with its map."""

                return (
                    f"{artifact}_canonical_component"
                    if artifact in existing_observations
                    else artifact
                )

            component_observations = {
                _component_artifact_name(
                    "executed_day_accounting"
                ): _required_component_value(
                    executed_breakdown,
                    source_key,
                    artifact=_component_artifact_name(
                        "executed_day_accounting"
                    ),
                    metric=component_key,
                ),
                _component_artifact_name("canonical_cost_ledger"): ledger_value,
                _component_artifact_name("summary"): _required_component_value(
                    summary_components,
                    component_key,
                    artifact=_component_artifact_name("summary"),
                    metric=component_key,
                ),
                _component_artifact_name(
                    "experiment_report_json"
                ): _required_component_value(
                    report_components,
                    component_key,
                    artifact=_component_artifact_name(
                        "experiment_report_json"
                    ),
                    metric=component_key,
                ),
                _component_artifact_name(
                    "cost_breakdown_detail"
                ): _required_component_value(
                    detail_by_key,
                    source_key,
                    artifact=_component_artifact_name(
                        "cost_breakdown_detail"
                    ),
                    metric=component_key,
                ),
                _component_artifact_name("results_xlsx"): _required_component_value(
                    workbook_costs,
                    source_key,
                    artifact=_component_artifact_name("results_xlsx"),
                    metric=component_key,
                ),
                _component_artifact_name(
                    "optimization_result"
                ): _required_component_value(
                    optimization_components,
                    source_key,
                    artifact=_component_artifact_name(
                        "optimization_result"
                    ),
                    metric=component_key,
                ),
            }
            if existing_observations:
                existing_observations.update(component_observations)
            else:
                observed_by_metric[component_key] = component_observations
    residuals_by_metric = {
        metric: {
            artifact: (
                value - expected_by_metric[metric]
                if value is not None and expected_by_metric[metric] is not None
                else None
            )
            for artifact, value in observations.items()
        }
        for metric, observations in observed_by_metric.items()
    }
    failures = dict(pre_failures)
    for metric, residuals in residuals_by_metric.items():
        for artifact, residual in residuals.items():
            if residual is not None and not math.isclose(
                residual,
                0.0,
                rel_tol=0.0,
                abs_tol=float(tolerance_jpy),
            ):
                failures[f"{metric}:{artifact}"] = residual
    report_md_path = run_dir / "experiment_report.md"
    report_md = report_md_path.read_text(encoding="utf-8")
    marker_prefix = "canonical_executed_total_cost_jpy: `"
    marker_values: List[str] = []
    for line in report_md.splitlines():
        normalized_line = line[2:] if line.startswith("- ") else line
        if (
            normalized_line.startswith(marker_prefix)
            and normalized_line.endswith("`")
        ):
            marker_values.append(normalized_line[len(marker_prefix) : -1])
    if len(marker_values) != 1:
        failures["experiment_report_md_canonical_total"] = {
            "reason": "missing_or_ambiguous_marker",
            "marker_count": len(marker_values),
        }
    else:
        try:
            markdown_total = float(marker_values[0])
        except (TypeError, ValueError):
            failures["experiment_report_md_canonical_total"] = {
                "reason": "invalid_marker_value",
                "value": marker_values[0],
            }
        else:
            if not math.isfinite(markdown_total):
                failures["experiment_report_md_canonical_total"] = {
                    "reason": "nonfinite_marker_value",
                    "value": markdown_total,
                }
            elif executed_total is not None and not math.isclose(
                markdown_total,
                executed_total,
                rel_tol=0.0,
                abs_tol=float(tolerance_jpy),
            ):
                failures["experiment_report_md_canonical_total"] = {
                    "expected": executed_total,
                    "observed": markdown_total,
                    "residual": markdown_total - executed_total,
                }
    payload = {
        "schema_version": "final_cost_reconciliation_v1",
        "status": "OK" if not failures else "ERROR",
        "source": (
            "rolling_hourly_chain/executed_day_accounting.json"
        ),
        "tolerance_jpy": float(tolerance_jpy),
        "expected_by_metric_jpy": expected_by_metric,
        "observed_by_metric_jpy": observed_by_metric,
        "residual_to_executed_day_by_metric_jpy": residuals_by_metric,
        "failed_artifacts": failures,
    }
    (run_dir / "final_cost_reconciliation.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    if failures:
        raise RuntimeError(
            "Final cost artifacts disagree with executed rolling accounting: "
            + json.dumps(failures, ensure_ascii=False, sort_keys=True)
        )
    return payload


def _prepend_experiment_release_header(
    *,
    run_dir: Path,
    research_claim_scope: Dict[str, Any],
    rolling_execution: Dict[str, Any],
    solver_settings: Dict[str, Any],
    optimization_result: Dict[str, Any],
) -> None:
    """Put the release/claim gate before any human-facing result narrative."""

    report_path = run_dir / "experiment_report.md"
    if not report_path.is_file():
        return
    marker = "<!-- frontend-run-release-header-v1 -->"
    body = report_path.read_text(encoding="utf-8")
    if marker in body:
        body = body.split("<!-- /frontend-run-release-header-v1 -->", 1)[-1].lstrip()
    failed_checks = list(
        research_claim_scope.get("teacher_release_failed_checks") or ()
    )
    objective_is_actual_cost = bool(
        dict(optimization_result.get("cost_breakdown") or {}).get(
            "objective_is_actual_cost", False
        )
    )
    canonical_ledger_path = run_dir / "graph" / "canonical_cost_ledger.json"
    canonical_ledger = (
        json.loads(canonical_ledger_path.read_text(encoding="utf-8"))
        if canonical_ledger_path.is_file()
        else {}
    )
    canonical_total = canonical_ledger.get("accounting_total_cost_jpy")
    canonical_source = canonical_ledger.get("source")
    result_claim = dict(
        optimization_result.get("result_claim_classification") or {}
    )
    header = [
        marker,
        "# Run and release status",
        "",
        f"- run_profile: `{research_claim_scope.get('run_profile')}`",
        f"- rolling_execution.status: `{rolling_execution.get('status')}`",
        (
            "- rolling_execution_minutes: "
            f"`{rolling_execution.get('rolling_execution_minutes')}`"
        ),
        (
            "- research_submission_ready: "
            f"`{str(bool(research_claim_scope.get('research_submission_ready'))).lower()}`"
        ),
        (
            "- teacher_release_status: "
            f"`{research_claim_scope.get('teacher_release_status')}`"
        ),
        (
            "- failed_checks: "
            + (
                ", ".join(f"`{value}`" for value in failed_checks)
                if failed_checks
                else "none"
            )
        ),
        (
            "- requested_mip_gap: "
            f"`{solver_settings.get('mip_gap_requested_ratio')}`"
        ),
        (
            "- stage1_gurobi_raw_mip_gap: "
            f"`{solver_settings.get('stage1_gurobi_raw_mip_gap_ratio')}`"
        ),
        (
            "- stage1_certified_mip_gap: "
            f"`{solver_settings.get('stage1_certified_mip_gap_ratio')}`"
        ),
        (
            "- mip_gap_target_met: "
            f"`{str(bool(solver_settings.get('mip_gap_target_met'))).lower()}`"
        ),
        (
            "- solver_termination_reason: "
            f"`{solver_settings.get('solver_termination_reason')}`"
        ),
        f"- result_claim_label: `{result_claim.get('label')}`",
        (
            "- optimality_claim_eligible: "
            f"`{str(bool(result_claim.get('optimality_claim_eligible'))).lower()}`"
        ),
        (
            "- optimality_blocking_reasons: "
            + (
                ", ".join(
                    f"`{value}`"
                    for value in list(
                        result_claim.get("optimality_blocking_reasons") or ()
                    )
                )
                or "none"
            )
        ),
        (
            "- objective_is_actual_cost: "
            f"`{str(objective_is_actual_cost).lower()}`"
        ),
        (
            "- cost_semantics: `canonical accounting total is the cost KPI; "
            "a non-cost or proxy objective is reported separately`"
        ),
        (
            "- canonical_cost_source: "
            f"`{canonical_source}`"
        ),
        (
            "- canonical_executed_total_cost_jpy: "
            f"`{canonical_total!r}`"
        ),
    ]
    if research_claim_scope.get("run_profile") == DAY_AHEAD_EXPLORATORY_PROFILE:
        header.extend(
            [
                "",
                "> **DAY-AHEAD ONLY - NOT A ROLLING RESULT**",
            ]
        )
    header.extend(["", "<!-- /frontend-run-release-header-v1 -->", ""])
    report_path.write_text("\n".join(header) + body, encoding="utf-8")


def _write_results_workbook_release_status(
    *,
    run_dir: Path,
    research_claim_scope: Dict[str, Any],
    rolling_execution: Dict[str, Any],
    solver_settings: Dict[str, Any],
    result_claim_classification: Optional[Dict[str, Any]] = None,
) -> None:
    """Synchronize results.xlsx with the final rolling/release decision."""

    workbook_path = run_dir / "results.xlsx"
    if not workbook_path.is_file():
        return
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path)
    if "release_status" in workbook.sheetnames:
        del workbook["release_status"]
    sheet = workbook.create_sheet("release_status", 0)
    sheet.append(["field", "value"])
    rows = (
        ("run_profile", research_claim_scope.get("run_profile")),
        ("rolling_execution_status", rolling_execution.get("status")),
        (
            "rolling_execution_minutes",
            rolling_execution.get("rolling_execution_minutes"),
        ),
        (
            "research_submission_ready",
            research_claim_scope.get("research_submission_ready"),
        ),
        ("teacher_release_status", research_claim_scope.get("teacher_release_status")),
        (
            "teacher_release_failed_checks",
            ";".join(
                research_claim_scope.get("teacher_release_failed_checks") or ()
            ),
        ),
        ("requested_mip_gap", solver_settings.get("mip_gap_requested_ratio")),
        (
            "stage1_gurobi_raw_mip_gap",
            solver_settings.get("stage1_gurobi_raw_mip_gap_ratio"),
        ),
        (
            "stage1_certified_mip_gap",
            solver_settings.get("stage1_certified_mip_gap_ratio"),
        ),
        ("mip_gap_target_met", solver_settings.get("mip_gap_target_met")),
        (
            "result_claim_label",
            dict(result_claim_classification or {}).get("label"),
        ),
        (
            "optimality_claim_eligible",
            dict(result_claim_classification or {}).get(
                "optimality_claim_eligible"
            ),
        ),
    )
    for row in rows:
        sheet.append(list(row))
    workbook.save(workbook_path)


def _mark_frontend_run_claims_failed(
    *,
    run_dir: Path,
    error: BaseException,
) -> None:
    """Downgrade partially written human-facing artifacts after a job failure.

    Reporting finalization writes several artifacts before all later gates have
    completed. If one of those gates raises, preserve diagnostics but never
    leave a release-capable label in the run directory.
    """

    scope_path = run_dir / "research_claim_scope.json"
    if not scope_path.is_file():
        return
    loaded_scope = json.loads(scope_path.read_text(encoding="utf-8"))
    if not isinstance(loaded_scope, dict):
        return
    scope = dict(loaded_scope)
    failed_checks = set(scope.get("teacher_release_failed_checks") or ())
    failed_checks.update(
        {
            "frontend_run_failed",
            "final_reporting_or_artifact_contract_failed",
        }
    )
    failure = {
        "type": type(error).__name__,
        "message": str(error),
    }
    failed_artifact_audit_summary: Optional[Dict[str, Any]] = None
    artifact_audit_path = run_dir / "artifact_completeness.json"
    if artifact_audit_path.is_file():
        terminal_failure_reason = "frontend_run_failed_after_finalization"
        try:
            loaded_artifact_audit = json.loads(
                artifact_audit_path.read_text(encoding="utf-8")
            )
        except (OSError, json.JSONDecodeError) as exc:
            artifact_audit = {
                "prior_audit_read_error": f"{type(exc).__name__}: {exc}"
            }
        else:
            artifact_audit = (
                dict(loaded_artifact_audit)
                if isinstance(loaded_artifact_audit, dict)
                else {"prior_audit_invalid_type": type(loaded_artifact_audit).__name__}
            )
        content_errors = list(artifact_audit.get("content_errors") or ())
        if terminal_failure_reason not in content_errors:
            content_errors.append(terminal_failure_reason)
        artifact_audit.update(
            {
                "status": "ERROR",
                "accepted": False,
                "content_errors": content_errors,
                "terminal_failure": failure,
            }
        )
        artifact_audit_path.write_text(
            json.dumps(artifact_audit, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        failed_artifact_audit_summary = {
            "status": "ERROR",
            "accepted": False,
            "artifact": artifact_audit_path.name,
            "reason": terminal_failure_reason,
        }
    scope.update(
        {
            "research_submission_ready": False,
            "teacher_release_status": "BLOCKED",
            "teacher_release_failed_checks": sorted(map(str, failed_checks)),
            "result_label": "diagnostic_run_finalization_failed",
            "diagnostic_only": True,
            "diagnostic_label": "NOT USED FOR RESEARCH CONCLUSIONS",
            "finalization_failure": failure,
        }
    )
    scope_path.write_text(
        json.dumps(scope, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    summary_path = run_dir / "summary.json"
    if summary_path.is_file():
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        if isinstance(summary, dict):
            summary.update(
                {
                    "research_submission_ready": False,
                    "teacher_release_status": "BLOCKED",
                    "teacher_release_failed_checks": scope[
                        "teacher_release_failed_checks"
                    ],
                    "diagnostic_only": True,
                    "diagnostic_label": scope["diagnostic_label"],
                    "finalization_failure": failure,
                }
            )
            summary_path.write_text(
                json.dumps(summary, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

    optimization_result: Dict[str, Any] = {}
    optimization_result_path = run_dir / "optimization_result.json"
    if optimization_result_path.is_file():
        loaded_result = json.loads(
            optimization_result_path.read_text(encoding="utf-8")
        )
        if isinstance(loaded_result, dict):
            optimization_result = dict(loaded_result)
            optimization_result["research_claim_scope"] = scope
            optimization_result["diagnostic_only"] = True
            optimization_result["finalization_failure"] = failure
            if failed_artifact_audit_summary is not None:
                optimization_result["artifact_completeness"] = (
                    failed_artifact_audit_summary
                )
            for path in (
                optimization_result_path,
                run_dir / "raw" / "optimization_result.json",
            ):
                if path.parent.is_dir():
                    path.write_text(
                        json.dumps(optimization_result, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )

    optimization_audit_path = run_dir / "optimization_audit.json"
    if optimization_audit_path.is_file():
        loaded_audit = json.loads(optimization_audit_path.read_text(encoding="utf-8"))
        if isinstance(loaded_audit, dict):
            audit = dict(loaded_audit)
            audit["research_claim_scope"] = scope
            audit["finalization_failure"] = failure
            if failed_artifact_audit_summary is not None:
                audit["artifact_completeness"] = failed_artifact_audit_summary
            for path in (
                optimization_audit_path,
                run_dir / "raw" / "optimization_audit.json",
            ):
                if path.parent.is_dir():
                    path.write_text(
                        json.dumps(audit, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )

    solver_settings = dict(optimization_result.get("solver_settings") or {})
    solver_metadata = dict(optimization_result.get("solver_metadata") or {})
    rolling_execution = _rolling_execution_evidence(
        run_dir=run_dir,
        solver_metadata=solver_metadata,
    )
    _prepend_experiment_release_header(
        run_dir=run_dir,
        research_claim_scope=scope,
        rolling_execution=rolling_execution,
        solver_settings=solver_settings,
        optimization_result=optimization_result,
    )
    _write_results_workbook_release_status(
        run_dir=run_dir,
        research_claim_scope=scope,
        rolling_execution=rolling_execution,
        solver_settings=solver_settings,
        result_claim_classification=dict(
            optimization_result.get("result_claim_classification") or {}
        ),
    )

    run_manifest_path = run_dir / "run_manifest.json"
    if run_manifest_path.is_file():
        run_manifest = json.loads(run_manifest_path.read_text(encoding="utf-8"))
        if isinstance(run_manifest, dict):
            run_manifest_updates = {
                "run_state": "failed",
                "research_claim_scope": scope,
                "teacher_release_status": "BLOCKED",
                "teacher_release_failed_checks": scope[
                    "teacher_release_failed_checks"
                ],
                "finalization_failure": failure,
            }
            if failed_artifact_audit_summary is not None:
                run_manifest_updates["artifact_completeness"] = (
                    failed_artifact_audit_summary
                )
            run_manifest.update(run_manifest_updates)
            run_manifest_path.write_text(
                json.dumps(run_manifest, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
    # The rolling provenance manifest is distinct from run_manifest.json. A
    # report-finalization exception must not leave its previous `complete`
    # state behind as contradictory provenance.
    if (run_dir / "input_audit.json").is_file():
        refresh_frontend_rolling_manifest(
            run_dir=run_dir,
            run_state="reporting_finalization_failed",
        )


def _read_optional_json_mapping(path: Path) -> Dict[str, Any]:
    """Load a local JSON object without turning diagnostic persistence into a crash."""

    if not path.is_file():
        return {}
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        return {}
    return dict(loaded) if isinstance(loaded, Mapping) else {}


def _finite_release_currency(value: Any) -> Optional[float]:
    """Return a finite numeric currency value without accepting booleans."""

    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    numeric = float(value)
    return numeric if math.isfinite(numeric) else None


def _solver_objective_accounting_reconciliation_payload(
    *,
    run_dir: Path,
    optimization_result: Mapping[str, Any],
) -> Dict[str, Any]:
    """Persist auditable solver-objective versus final-accounting evidence.

    The final rolling ledger is authoritative for cost.  This artifact compares
    it numerically with the solver objective while preserving the distinct
    two-stage semantics; matching numbers do not by themselves upgrade a proxy
    objective into an actual-cost objective.
    """

    ledger = _read_optional_json_mapping(
        Path(run_dir) / "graph" / "canonical_cost_ledger.json"
    )
    solver_metadata = dict(optimization_result.get("solver_metadata") or {})
    solver_value = _finite_release_currency(
        optimization_result.get("objective_value")
    )
    solver_source = "optimization_result.objective_value"
    if solver_value is None:
        solver_value = _finite_release_currency(
            ledger.get("solver_objective_value")
        )
        solver_source = (
            "graph/canonical_cost_ledger.json.solver_objective_value"
            if solver_value is not None
            else "unavailable"
        )
    accounting_total = _finite_release_currency(
        ledger.get("accounting_total_cost_jpy")
    )
    accounting_source = str(ledger.get("source") or "unavailable")
    tolerance = _finite_release_currency(
        ledger.get("accounting_residual_tolerance_jpy")
    )
    if tolerance is None or tolerance < 0.0:
        tolerance = 1.0e-6
    numeric_values_available = (
        solver_value is not None and accounting_total is not None
    )
    difference = (
        float(solver_value - accounting_total)
        if numeric_values_available
        else None
    )
    absolute_difference = abs(difference) if difference is not None else None
    residual_within_tolerance = bool(
        absolute_difference is not None and absolute_difference <= tolerance
    )
    objective_is_actual_cost = ledger.get("objective_is_actual_cost") is True
    matches_canonical_accounting_total = bool(
        numeric_values_available
        and residual_within_tolerance
        and objective_is_actual_cost
        and accounting_source
        == "rolling_hourly_chain/executed_day_accounting.json"
    )
    canonical_cost_value = _finite_release_currency(
        solver_metadata.get("integrated_lexicographic_cost_objective_jpy")
    )
    canonical_cost_source = (
        "solver_metadata.integrated_lexicographic_cost_objective_jpy"
        if canonical_cost_value is not None
        else solver_source if objective_is_actual_cost else "unavailable"
    )
    if canonical_cost_value is None and objective_is_actual_cost:
        canonical_cost_value = solver_value
    canonical_cost_values_available = bool(
        canonical_cost_value is not None and accounting_total is not None
    )
    canonical_cost_difference = (
        float(canonical_cost_value - accounting_total)
        if canonical_cost_values_available
        else None
    )
    canonical_cost_absolute_difference = (
        abs(canonical_cost_difference)
        if canonical_cost_difference is not None
        else None
    )
    canonical_cost_residual_within_tolerance = bool(
        canonical_cost_absolute_difference is not None
        and canonical_cost_absolute_difference <= tolerance
    )
    canonical_cost_contract_applied = bool(
        objective_is_actual_cost
        or solver_metadata.get("integrated_actual_cost_contract_applied")
        is True
    )
    canonical_cost_matches_accounting_total = bool(
        canonical_cost_values_available
        and canonical_cost_residual_within_tolerance
        and canonical_cost_contract_applied
        and accounting_source
        == "rolling_hourly_chain/executed_day_accounting.json"
    )
    return {
        "schema_version": SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_SCHEMA_VERSION,
        "solver_objective_value_jpy": solver_value,
        "solver_objective_source": solver_source,
        "canonical_accounting_total_jpy": accounting_total,
        "canonical_accounting_source": accounting_source,
        "canonical_accounting_artifact": "graph/canonical_cost_ledger.json",
        "difference_jpy": difference,
        "absolute_difference_jpy": absolute_difference,
        "tolerance_jpy": tolerance,
        "numeric_values_available": numeric_values_available,
        "numeric_residual_within_tolerance": residual_within_tolerance,
        "objective_is_actual_cost": objective_is_actual_cost,
        "matches_canonical_accounting_total": (
            matches_canonical_accounting_total
        ),
        "canonical_cost_objective_value_jpy": canonical_cost_value,
        "canonical_cost_objective_source": canonical_cost_source,
        "canonical_cost_difference_jpy": canonical_cost_difference,
        "canonical_cost_absolute_difference_jpy": (
            canonical_cost_absolute_difference
        ),
        "canonical_cost_numeric_values_available": (
            canonical_cost_values_available
        ),
        "canonical_cost_residual_within_tolerance": (
            canonical_cost_residual_within_tolerance
        ),
        "canonical_cost_contract_applied": canonical_cost_contract_applied,
        "canonical_cost_matches_accounting_total": (
            canonical_cost_matches_accounting_total
        ),
        "legacy_solver_metadata_matches_accounting_total": (
            solver_metadata.get("solver_objective_matches_accounting_total")
        ),
        "objective_semantics": str(
            solver_metadata.get("objective_semantics")
            or ledger.get("objective_mode")
            or "unavailable"
        ),
        "semantics": (
            "Numeric comparison of the solver objective with the canonical "
            "accounting total. The canonical_cost_* fields separately audit "
            "the certified cost level of a lexicographic solve. A formal "
            "match requires the accepted executed-day accounting source and "
            "never follows from a legacy boolean alone."
        ),
    }


def _research_claim_scope_payload(
    *,
    optimization_result: Dict[str, Any],
    solver_settings: Dict[str, Any],
    weather_policy: Dict[str, Any],
    rolling_execution: Dict[str, Any],
    objective_accounting_reconciliation: Mapping[str, Any] | None = None,
    composition_search_certificate: Mapping[str, Any] | None = None,
) -> Dict[str, Any]:
    """State what a manual frontend run may, and may not, support.

    This is deliberately a claim gate rather than a score.  It prevents a
    PV-only manual day-ahead artifact from being relabelled later as a formal
    weather-adaptive dispatch result, a runtime benchmark, or an integrated
    global optimum.
    """

    metadata = dict(optimization_result.get("solver_metadata") or {})
    solution_validity = dict(optimization_result.get("solution_validity") or {})
    decision_policy = dict(
        (weather_policy.get("audit") or {}).get("decision_policy")
        or weather_policy.get("decision_policy")
        or {}
    )
    weather_enabled = bool(weather_policy.get("enabled", False))
    policy_scope = str(decision_policy.get("policy_scope") or "not_enabled")
    bev_utilization_policy = dict(
        optimization_result.get("bev_utilization_policy") or {}
    )
    all_available_bevs_required = bool(
        bev_utilization_policy.get("enabled", False)
    )
    physically_feasible = bool(solution_validity.get("validated_feasible", False))
    is_integrated_exact = bool(metadata.get("supports_integrated_exact_milp", False))
    mip_gap_target_met = solver_settings.get("mip_gap_target_met") is True
    optimality_claim_eligible = bool(
        physically_feasible and is_integrated_exact and mip_gap_target_met
    )
    is_manual_unaccepted = not bool(metadata.get("research_run_accepted", False))
    run_profile = str(
        optimization_result.get("run_profile")
        or metadata.get("run_profile")
        or DEFAULT_FRONTEND_RUN_PROFILE
    )
    research_failed_checks = [
        str(value)
        for value in list(
            metadata.get("research_acceptance_failed_checks")
            or metadata.get("research_run_failed_checks")
            or ()
        )
    ]
    if not research_failed_checks:
        research_failed_checks = sorted(
            str(name)
            for name, passed in dict(
                metadata.get("research_acceptance_checks") or {}
            ).items()
            if passed is not True
        )
    teacher_release_failed_checks = list(research_failed_checks)
    if not physically_feasible:
        teacher_release_failed_checks.append("physical_schedule_not_validated")
    if not bool(metadata.get("research_run", False)):
        teacher_release_failed_checks.append("research_run_not_requested")
    if not bool(metadata.get("research_run_accepted", False)):
        teacher_release_failed_checks.append("day_ahead_research_acceptance_failed")
    if rolling_execution.get("status") != "executed_and_accepted":
        teacher_release_failed_checks.append("hourly_rolling_chain_not_accepted")
    if not bool(
        metadata.get("research_submission_git_provenance_eligible", False)
    ):
        teacher_release_failed_checks.append("git_provenance_not_research_eligible")
    if run_profile == DAY_AHEAD_EXPLORATORY_PROFILE:
        teacher_release_failed_checks.append("day_ahead_only_exploratory_profile")
    prepared_scope_audit = dict(
        optimization_result.get("prepared_scope_audit") or {}
    )
    if (
        bool(metadata.get("research_run", False))
        and "formal_transition_network_ready" in prepared_scope_audit
        and prepared_scope_audit.get("formal_transition_network_ready") is not True
    ):
        teacher_release_failed_checks.append(
            "route_band_off_deadhead_matrix_incomplete"
        )
    if (
        bool(metadata.get("research_run", False))
        and "formal_vehicle_trip_compatibility_ready" in prepared_scope_audit
        and prepared_scope_audit.get("formal_vehicle_trip_compatibility_ready")
        is not True
    ):
        teacher_release_failed_checks.append(
            "vehicle_trip_compatibility_contract_incomplete"
        )
    is_two_stage = (
        str(metadata.get("optimization_structure") or "").lower()
        == "two_stage"
    )
    objective_accounting_evidence_errors: list[str] = []
    composition_certificate_evidence_errors: list[str] = []
    if is_two_stage and bool(metadata.get("research_run", False)):
        # Do not substitute a summary/metadata boolean for proof. The
        # reconciliation must carry the numeric residual and the certificate
        # must carry the actual composition search result.
        objective_accounting_evidence_errors = (
            validate_solver_objective_accounting_reconciliation(
                objective_accounting_reconciliation,
                require_match=True,
            )
        )
        if objective_accounting_evidence_errors:
            teacher_release_failed_checks.append(
                "solver_objective_canonical_accounting_mismatch"
            )
        composition_certificate_evidence_errors = (
            validate_stage1_used_powertrain_composition_search(
                composition_search_certificate,
                require_accepted=True,
            )
        )
        if composition_certificate_evidence_errors:
            teacher_release_failed_checks.append(
                "used_powertrain_composition_search_not_certified"
            )
    # A single frontend run cannot establish the required same-service-date,
    # fixed-control counterfactual evidence. Keep release claims blocked until
    # the separately verified pair manifest exists; this also prevents an
    # interrupted reporting finalizer from leaving a misleading READY label.
    teacher_release_failed_checks.append(
        "controlled_counterfactual_pair_not_verified"
    )
    teacher_release_failed_checks = sorted(set(teacher_release_failed_checks))
    research_submission_ready = not teacher_release_failed_checks
    if weather_enabled and policy_scope == "pv_curve_only":
        result_label = "exploratory_pv_supply_sensitivity_not_weather_adaptive_dispatch"
    elif weather_enabled:
        result_label = "manual_weather_policy_day_ahead_result_not_formal_comparison"
    else:
        result_label = "manual_day_ahead_feasibility_result"

    allowed_claims: list[str] = []
    if physically_feasible:
        allowed_claims.append("physical_schedule_feasibility_under_recorded_inputs")
    if weather_enabled:
        allowed_claims.append("recorded_pv_supply_effect_on_depot_energy_flows")
    if all_available_bevs_required and physically_feasible:
        allowed_claims.append(
            "all_available_bev_minimum_use_policy_feasibility"
        )
    disallowed_claims = [
        "integrated_global_total_cost_optimum",
        "actual_monthly_demand_charge_savings",
        "pv_or_bess_investment_economics",
    ]
    if policy_scope != "weather_dispatch_policy":
        disallowed_claims.append("weather_adaptive_dispatch_or_charging_policy")
    if is_manual_unaccepted:
        disallowed_claims.append("formal_research_weather_comparison")
    if all_available_bevs_required:
        disallowed_claims.append(
            "unconstrained_total_cost_minimum_without_policy_constraint"
        )
    if rolling_execution.get("status") != "executed_and_accepted":
        disallowed_claims.append("hourly_rolling_reoptimization_performance")
    # A single manually initiated run can record whether BestObjStop was out of
    # the way, but it cannot itself establish a runtime comparison.  That also
    # needs matched cross-case controls and repeated measurements.
    disallowed_claims.append("wall_clock_runtime_comparison")
    nonformal_markers = (
        {
            "diagnostic_only": True,
            "blocking_reason": "dirty_or_nonformal_run",
        }
        if not bool(metadata.get("research_run", False))
        else {}
    )

    return {
        "schema_version": "research_claim_scope_v1",
        "run_profile": run_profile,
        "research_submission_ready": research_submission_ready,
        "teacher_release_status": (
            "READY" if research_submission_ready else "BLOCKED"
        ),
        "teacher_release_failed_checks": teacher_release_failed_checks,
        "result_label": result_label,
        "physical_feasibility_claim_eligible": physically_feasible,
        "weather_policy_scope": policy_scope,
        "weather_adaptive_dispatch_claim_eligible": (
            weather_enabled
            and policy_scope == "weather_dispatch_policy"
            and not is_manual_unaccepted
        ),
        "formal_weather_comparison_claim_eligible": False,
        "integrated_global_optimality_claim_eligible": optimality_claim_eligible,
        "runtime_comparison_claim_eligible": False,
        "demand_charge_claim_scope": (
            "planning_horizon_allocation_proxy_not_actual_monthly_bill_savings"
        ),
        "asset_economics_claim_eligible": False,
        **nonformal_markers,
        "bev_utilization_policy": bev_utilization_policy,
        "allowed_claims": allowed_claims,
        "disallowed_claims": sorted(set(disallowed_claims)),
        "evidence": {
            "research_run": bool(metadata.get("research_run", False)),
            "research_run_accepted": bool(
                metadata.get("research_run_accepted", False)
            ),
            "solver_objective_matches_accounting_total": metadata.get(
                "solver_objective_matches_accounting_total"
            ),
            "used_powertrain_composition_search_accepted": metadata.get(
                "stage1_used_powertrain_composition_search_accepted"
            ),
            "solver_objective_accounting_reconciliation_artifact": (
                SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE
            ),
            "solver_objective_accounting_reconciliation_errors": (
                objective_accounting_evidence_errors
            ),
            "stage1_used_powertrain_composition_search_artifact": (
                STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE
            ),
            "stage1_used_powertrain_composition_search_errors": (
                composition_certificate_evidence_errors
            ),
            "optimization_structure": metadata.get("optimization_structure"),
            "supports_integrated_exact_milp": is_integrated_exact,
            "mip_gap_target_met": mip_gap_target_met,
            "weather_policy_scope": policy_scope,
            "rolling_execution_status": rolling_execution.get("status"),
            "rolling_execution_minutes": rolling_execution.get(
                "rolling_execution_minutes"
            ),
            "research_submission_git_provenance_eligible": bool(
                metadata.get("research_submission_git_provenance_eligible", False)
            ),
            "research_acceptance_failed_checks": research_failed_checks,
            "stage1_best_obj_stop_applied": solver_settings.get(
                "stage1_best_obj_stop_applied"
            ),
            "stage1_stop_rule_runtime_control_eligible": solver_settings.get(
                "runtime_comparison_eligible"
            ),
            "runtime_comparison_claim_requires": [
                "matched_cross_case_solver_controls",
                "fixed_explicit_gurobi_threads",
                "multiple_repetitions",
            ],
        },
    }


def _canonical_vehicle_timeline_rows(
    *,
    problem,
    engine_result,
    scenario_id: str,
    graph_context: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    from src.result_exporter import _route_band_key

    rows: List[Dict[str, Any]] = []
    problem_trip_by_id = problem.trip_by_id()
    vehicle_by_id = {str(vehicle.vehicle_id): vehicle for vehicle in problem.vehicles}
    band_labels_by_band_id = dict((graph_context or {}).get("band_labels_by_band_id") or {})
    base_date = _canonical_output_base_date(problem, graph_context)
    depot_name_by_id = {
        str(depot.depot_id): str(getattr(depot, "name", "") or getattr(depot, "depot_id", "") or "")
        for depot in problem.depots
    }
    duties_by_vehicle = engine_result.plan.duties_by_vehicle()
    charge_slots_by_vehicle: Dict[str, List[Any]] = defaultdict(list)
    for slot in engine_result.plan.charging_slots:
        charge_slots_by_vehicle[str(slot.vehicle_id)].append(slot)
    refuel_slots_by_vehicle: Dict[str, List[Any]] = defaultdict(list)
    for slot in engine_result.plan.refuel_slots:
        refuel_slots_by_vehicle[str(slot.vehicle_id)].append(slot)

    active_vehicle_ids = vehicle_ids_with_timeline_activity(
        duties_by_vehicle,
        engine_result.plan.charging_slots,
        engine_result.plan.refuel_slots,
    )
    for vehicle_id in active_vehicle_ids:
        duties = list(duties_by_vehicle.get(str(vehicle_id), []))
        vehicle = vehicle_by_id.get(str(vehicle_id))
        depot_id = str(getattr(vehicle, "home_depot_id", "") or "")
        depot_label = depot_name_by_id.get(depot_id) or depot_id
        vehicle_type = str(getattr(vehicle, "vehicle_type", "") or (duties[0].vehicle_type if duties else ""))
        band_counter: Counter[str] = Counter()
        for duty in duties:
            for leg in duty.legs:
                trip_id = str(getattr(leg.trip, "trip_id", "") or "")
                problem_trip = problem_trip_by_id.get(trip_id)
                if problem_trip is None:
                    continue
                route_family_code = str(getattr(leg.trip, "route_family_code", "") or "")
                route_id = str(getattr(leg.trip, "route_id", problem_trip.route_id) or problem_trip.route_id)
                band_id = _route_band_key(route_family_code, route_id)
                if band_id:
                    band_counter[band_id] += 1
        primary_band_id = ""
        primary_band_label = ""
        if band_counter:
            primary_band_id = sorted(
                band_counter.items(),
                key=lambda item: (-item[1], item[0]),
            )[0][0]
            primary_band_label = str(band_labels_by_band_id.get(primary_band_id) or primary_band_id)

        for duty in duties:
            prev_trip = None
            prev_band_id = ""
            for leg in duty.legs:
                dispatch_trip = leg.trip
                trip_id = str(dispatch_trip.trip_id or "")
                problem_trip = problem_trip_by_id.get(trip_id)
                if problem_trip is None:
                    continue
                route_family_code = str(getattr(dispatch_trip, "route_family_code", "") or "")
                route_id = str(getattr(dispatch_trip, "route_id", problem_trip.route_id) or problem_trip.route_id)
                band_id = _route_band_key(route_family_code, route_id)
                band_label = str(band_labels_by_band_id.get(band_id) or band_id)
                start_dt = _canonical_datetime_from_min(base_date, int(dispatch_trip.departure_min))
                end_min = int(dispatch_trip.arrival_min)
                if end_min <= int(dispatch_trip.departure_min):
                    end_min += 24 * 60
                end_dt = _canonical_datetime_from_min(base_date, end_min)
                variant = str(getattr(dispatch_trip, "route_variant_type", "") or "")
                deadhead_min = max(int(getattr(leg, "deadhead_from_prev_min", 0) or 0), 0)
                if deadhead_min > 0:
                    deadhead_start = _canonical_datetime_from_min(
                        base_date,
                        int(dispatch_trip.departure_min) - deadhead_min,
                    )
                    deadhead_band_id = band_id if prev_trip is None or prev_band_id == band_id else ""
                    rows.append(
                        {
                            "scenario_id": scenario_id,
                            "depot_id": depot_id,
                            "vehicle_id": str(vehicle_id),
                            "vehicle_type": vehicle_type,
                            "band_id": deadhead_band_id,
                            "band_label": str(band_labels_by_band_id.get(deadhead_band_id) or deadhead_band_id),
                            "vehicle_primary_band_id": primary_band_id,
                            "vehicle_primary_band_label": primary_band_label,
                            "start_time": deadhead_start.isoformat(),
                            "end_time": start_dt.isoformat(),
                            "state": "deadhead",
                            "route_id": "",
                            "route_family_code": route_family_code,
                            "route_series_code": deadhead_band_id,
                            "event_route_band_id": deadhead_band_id,
                            "trip_id": "",
                            "from_location_id": (
                                depot_label
                                if prev_trip is None
                                else str(getattr(prev_trip, "destination", "") or "")
                            ),
                            "to_location_id": str(dispatch_trip.origin or ""),
                            "from_location_type": "depot" if prev_trip is None else "terminal",
                            "to_location_type": "terminal",
                            "direction": "",
                            "route_variant_type": "",
                            "energy_delta_kwh": -_canonical_estimated_deadhead_energy_kwh(
                                problem,
                                deadhead_min=deadhead_min,
                                trip_energy_kwh=float(getattr(problem_trip, "energy_kwh", 0.0) or 0.0),
                                trip_distance_km=float(getattr(problem_trip, "distance_km", 0.0) or 0.0),
                            ),
                            "distance_km": _canonical_deadhead_distance_km(problem, deadhead_min),
                            "duration_min": float(deadhead_min),
                            "is_deadhead": True,
                            "is_charge": False,
                            "is_service": False,
                            "is_idle": False,
                            "is_depot_move": prev_trip is None,
                            "is_short_turn": False,
                            "charger_id": "",
                            "charge_power_kw": "",
                            "refuel_liters": "",
                        }
                    )

                rows.append(
                    {
                        "scenario_id": scenario_id,
                        "depot_id": depot_id,
                        "vehicle_id": str(vehicle_id),
                        "vehicle_type": vehicle_type,
                        "band_id": band_id,
                        "band_label": band_label,
                        "vehicle_primary_band_id": primary_band_id,
                        "vehicle_primary_band_label": primary_band_label,
                        "start_time": start_dt.isoformat(),
                        "end_time": end_dt.isoformat(),
                        "state": "service",
                        "route_id": route_id,
                        "route_family_code": route_family_code,
                        "route_series_code": band_id,
                        "event_route_band_id": band_id,
                        "trip_id": trip_id,
                        "from_location_id": str(dispatch_trip.origin or ""),
                        "to_location_id": str(dispatch_trip.destination or ""),
                        "from_location_type": "terminal",
                        "to_location_type": "terminal",
                        "direction": str(getattr(dispatch_trip, "direction", "") or ""),
                        "route_variant_type": variant,
                        "energy_delta_kwh": -max(float(getattr(problem_trip, "energy_kwh", 0.0) or 0.0), 0.0),
                        "distance_km": max(float(getattr(problem_trip, "distance_km", 0.0) or 0.0), 0.0),
                        "duration_min": max((end_dt - start_dt).total_seconds() / 60.0, 0.0),
                        "is_deadhead": False,
                        "is_charge": False,
                        "is_service": True,
                        "is_idle": False,
                        "is_depot_move": variant in {"depot_move", "depot_in", "depot_out"},
                        "is_short_turn": variant == "short_turn",
                        "charger_id": "",
                        "charge_power_kw": "",
                        "refuel_liters": "",
                    }
                )
                prev_trip = dispatch_trip
                prev_band_id = band_id

        for start_slot, end_slot, charger_id, avg_charge_kw, avg_discharge_kw, location_id in _canonical_charge_segments(
            problem,
            charge_slots_by_vehicle.get(str(vehicle_id), []),
            fallback_location_id=depot_id,
        ):
            charge_start = _canonical_slot_datetime(problem, base_date, start_slot)
            charge_end = _canonical_slot_datetime(problem, base_date, end_slot)
            duration_min = max((charge_end - charge_start).total_seconds() / 60.0, 0.0)
            net_power_kw = avg_charge_kw - avg_discharge_kw
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "depot_id": depot_id,
                    "vehicle_id": str(vehicle_id),
                    "vehicle_type": vehicle_type,
                    "band_id": "",
                    "band_label": "",
                    "vehicle_primary_band_id": primary_band_id,
                    "vehicle_primary_band_label": primary_band_label,
                    "start_time": charge_start.isoformat(),
                    "end_time": charge_end.isoformat(),
                    "state": "charge",
                    "route_id": "",
                    "route_family_code": "",
                    "route_series_code": "",
                    "event_route_band_id": "",
                    "trip_id": "",
                    "from_location_id": location_id,
                    "to_location_id": location_id,
                    "from_location_type": "charger",
                    "to_location_type": "charger",
                    "direction": "",
                    "route_variant_type": "",
                    "energy_delta_kwh": net_power_kw * duration_min / 60.0,
                    "distance_km": 0.0,
                    "duration_min": duration_min,
                    "is_deadhead": False,
                    "is_charge": True,
                    "is_service": False,
                    "is_idle": False,
                    "is_depot_move": False,
                    "is_short_turn": False,
                    "charger_id": charger_id,
                    "charge_power_kw": net_power_kw,
                    "refuel_liters": "",
                }
            )

        for refuel_slot in sorted(
            refuel_slots_by_vehicle.get(str(vehicle_id), []),
            key=lambda slot: (int(getattr(slot, "slot_index", 0) or 0), str(getattr(slot, "vehicle_id", "") or "")),
        ):
            liters = max(float(getattr(refuel_slot, "refuel_liters", 0.0) or 0.0), 0.0)
            if liters <= 0.0:
                continue
            slot_index = int(getattr(refuel_slot, "slot_index", 0) or 0)
            refuel_start = _canonical_slot_datetime(problem, base_date, slot_index)
            refuel_end = _canonical_slot_datetime(problem, base_date, slot_index + 1)
            location_id = str(getattr(refuel_slot, "location_id", "") or depot_id or depot_label)
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "depot_id": depot_id,
                    "vehicle_id": str(vehicle_id),
                    "vehicle_type": vehicle_type,
                    "band_id": "",
                    "band_label": "",
                    "vehicle_primary_band_id": primary_band_id,
                    "vehicle_primary_band_label": primary_band_label,
                    "start_time": refuel_start.isoformat(),
                    "end_time": refuel_end.isoformat(),
                    "state": "refuel",
                    "route_id": "",
                    "route_family_code": "",
                    "route_series_code": "",
                    "event_route_band_id": "",
                    "trip_id": "",
                    "from_location_id": location_id,
                    "to_location_id": location_id,
                    "from_location_type": "depot",
                    "to_location_type": "depot",
                    "direction": "",
                    "route_variant_type": "depot_refuel",
                    "energy_delta_kwh": "",
                    "distance_km": 0.0,
                    "duration_min": max((refuel_end - refuel_start).total_seconds() / 60.0, 0.0),
                    "is_deadhead": False,
                    "is_charge": False,
                    "is_service": False,
                    "is_idle": False,
                    "is_depot_move": True,
                    "is_short_turn": False,
                    "charger_id": "",
                    "charge_power_kw": "",
                    "refuel_liters": round(liters, 4),
                }
            )
    rows.sort(key=lambda row: (str(row.get("vehicle_id") or ""), str(row.get("start_time") or ""), str(row.get("trip_id") or "")))
    return rows


def _canonical_charge_segments(
    problem,
    charging_slots: List[Any],
    *,
    fallback_location_id: str,
) -> List[tuple[int, int, str, float, float, str]]:
    del problem
    grouped: Dict[tuple[str, str], Dict[int, tuple[float, float, str]]] = defaultdict(dict)
    for slot in charging_slots:
        slot_index = int(getattr(slot, "slot_index", 0) or 0)
        charger_id = str(getattr(slot, "charger_id", "") or "")
        location_id = str(getattr(slot, "charging_depot_id", "") or fallback_location_id)
        grouped[(charger_id, location_id)][slot_index] = (
            max(float(getattr(slot, "charge_kw", 0.0) or 0.0), 0.0),
            max(float(getattr(slot, "discharge_kw", 0.0) or 0.0), 0.0),
            location_id,
        )

    segments: List[tuple[int, int, str, float, float, str]] = []
    for (charger_id, location_id), slot_map in grouped.items():
        ordered_slots = sorted(slot_map)
        if not ordered_slots:
            continue
        seg_start = ordered_slots[0]
        seg_end = seg_start + 1
        charge_values = [slot_map[seg_start][0]]
        discharge_values = [slot_map[seg_start][1]]
        for slot_index in ordered_slots[1:]:
            if slot_index == seg_end:
                seg_end += 1
                charge_values.append(slot_map[slot_index][0])
                discharge_values.append(slot_map[slot_index][1])
                continue
            segments.append(
                (
                    seg_start,
                    seg_end,
                    charger_id,
                    sum(charge_values) / len(charge_values),
                    sum(discharge_values) / len(discharge_values),
                    location_id,
                )
            )
            seg_start = slot_index
            seg_end = slot_index + 1
            charge_values = [slot_map[slot_index][0]]
            discharge_values = [slot_map[slot_index][1]]
        segments.append(
            (
                seg_start,
                seg_end,
                charger_id,
                sum(charge_values) / len(charge_values),
                sum(discharge_values) / len(discharge_values),
                location_id,
            )
        )
    return sorted(segments, key=lambda item: (item[0], item[2], item[5]))


def _canonical_trip_assignment_rows(
    *,
    problem,
    engine_result,
    scenario_id: str,
    base_date: date,
    timeline_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    primary_band_by_vehicle = {
        str(row.get("vehicle_id") or ""): {
            "band_id": str(row.get("vehicle_primary_band_id") or ""),
            "band_label": str(row.get("vehicle_primary_band_label") or ""),
        }
        for row in timeline_rows
        if str(row.get("vehicle_id") or "").strip() and str(row.get("vehicle_primary_band_id") or "").strip()
    }
    problem_trip_by_id = problem.trip_by_id()
    vehicle_by_id = {str(vehicle.vehicle_id): vehicle for vehicle in problem.vehicles}
    rows: List[Dict[str, Any]] = []
    sequence_by_vehicle: Dict[str, int] = defaultdict(int)
    for vehicle_id, duties in sorted(engine_result.plan.duties_by_vehicle().items()):
        for duty in duties:
            vehicle_id = str(vehicle_id)
            # ``duties_by_vehicle`` is the canonical chronological order.  Do
            # not sort by trip_id: lexical IDs can reverse the actual service
            # order and make downstream timeline validation meaningless.
            vehicle = vehicle_by_id.get(vehicle_id)
            duty_legs = list(duty.legs)
            for index, leg in enumerate(duty_legs):
                sequence_by_vehicle[vehicle_id] += 1
                sequence = sequence_by_vehicle[vehicle_id]
                dispatch_trip = leg.trip
                trip_id = str(dispatch_trip.trip_id or "")
                problem_trip = problem_trip_by_id.get(trip_id)
                if problem_trip is None:
                    continue
                route_family_code = str(getattr(dispatch_trip, "route_family_code", "") or "")
                departure_dt = _canonical_datetime_from_min(base_date, int(getattr(dispatch_trip, "departure_min", 0) or 0))
                arrival_dt = _canonical_datetime_from_min(base_date, int(getattr(dispatch_trip, "arrival_min", 0) or 0))
                fuel_rate, co2_rate = _canonical_vehicle_fuel_and_co2_rates(
                    problem,
                    vehicle,
                )
                is_electric = (
                    vehicle is not None
                    and is_electric_vehicle(problem, vehicle)
                )
                service_fuel_l = (
                    0.0
                    if is_electric
                    else max(float(problem_trip.distance_km or 0.0), 0.0)
                    * fuel_rate
                )
                rows.append(
                    {
                        "scenario_id": scenario_id,
                        "trip_id": trip_id,
                        "route_id": str(getattr(dispatch_trip, "route_id", problem_trip.route_id) or problem_trip.route_id),
                        "route_family_code": route_family_code,
                        "route_series_code": route_family_code or str(getattr(dispatch_trip, "route_id", problem_trip.route_id) or problem_trip.route_id),
                        "band_id": route_family_code or str(getattr(dispatch_trip, "route_id", problem_trip.route_id) or problem_trip.route_id),
                        "direction": str(getattr(dispatch_trip, "direction", "") or ""),
                        "route_variant_type": str(getattr(dispatch_trip, "route_variant_type", "unknown") or "unknown"),
                        "scheduled_departure": departure_dt.isoformat(),
                        "scheduled_arrival": arrival_dt.isoformat(),
                        "actual_departure": departure_dt.isoformat(),
                        "actual_arrival": arrival_dt.isoformat(),
                        "assigned_vehicle_id": vehicle_id,
                        "vehicle_sequence": sequence,
                        "assigned_vehicle_type": str(getattr(vehicle, "vehicle_type", "") or ""),
                        "assigned_depot_id": str(getattr(vehicle, "home_depot_id", "") or ""),
                        "assigned_vehicle_band_id": str((primary_band_by_vehicle.get(vehicle_id) or {}).get("band_id") or ""),
                        "served_flag": True,
                        "unserved_reason": "",
                        "energy_used_kwh": float(getattr(problem_trip, "energy_kwh", 0.0) or 0.0),
                        "distance_km": float(getattr(problem_trip, "distance_km", 0.0) or 0.0),
                        "fuel_used_l": float(service_fuel_l),
                        "ice_co2_kg": float(service_fuel_l * co2_rate),
                        "delay_departure_min": 0.0,
                        "delay_arrival_min": 0.0,
                        "deadhead_before_km": _canonical_deadhead_distance_km(problem, int(getattr(leg, "deadhead_from_prev_min", 0) or 0)),
                        # Connection ownership is exclusively the next trip's
                        # before-event.  Terminal return is exported in the
                        # movement event ledger, never inferred here.
                        "deadhead_after_km": 0.0,
                        "swap_type": "none",
                    }
                )
    rows.sort(key=lambda row: (str(row.get("assigned_vehicle_id") or ""), int(row.get("vehicle_sequence", 0) or 0), str(row.get("scheduled_departure") or ""), str(row.get("trip_id") or "")))
    return rows


def _canonical_vehicle_fuel_and_co2_rates(problem, vehicle) -> tuple[float, float]:
    vehicle_type_id = str(getattr(vehicle, "vehicle_type", "") or "")
    vehicle_type = next(
        (
            item
            for item in tuple(getattr(problem, "vehicle_types", ()) or ())
            if str(getattr(item, "vehicle_type_id", "") or "") == vehicle_type_id
        ),
        None,
    )
    fuel_rate = max(
        float(
            getattr(vehicle, "fuel_consumption_l_per_km", None)
            or getattr(vehicle_type, "fuel_consumption_l_per_km", 0.0)
            or 0.0
        ),
        0.0,
    )
    co2_rate = max(
        float(
            getattr(vehicle_type, "co2_emission_kg_per_l", 0.0)
            or getattr(getattr(problem, "scenario", None), "ice_co2_kg_per_l", 0.0)
            or 0.0
        ),
        0.0,
    )
    return fuel_rate, co2_rate


def _canonical_movement_event_rows(
    *,
    problem,
    engine_result,
    scenario_id: str,
    base_date: date,
    operator_id: str,
) -> List[Dict[str, Any]]:
    """Export every modeled non-service movement exactly once.

    A connection belongs to the next trip.  It is not copied to the previous
    trip's ``deadhead_after`` field.  Terminal return is emitted only when the
    canonical return-to-home helper confirms that the model contains it.
    """

    problem_trip_by_id = problem.trip_by_id()
    vehicle_by_id = {
        str(vehicle.vehicle_id): vehicle
        for vehicle in tuple(getattr(problem, "vehicles", ()) or ())
    }
    rows: List[Dict[str, Any]] = []
    seen_event_ids: set[str] = set()
    for vehicle_id, duties in sorted(engine_result.plan.duties_by_vehicle().items()):
        vehicle_id = str(vehicle_id)
        vehicle = vehicle_by_id.get(vehicle_id)
        if vehicle is None:
            raise ValueError(
                f"canonical movement export cannot resolve vehicle {vehicle_id!r}"
            )
        fuel_rate, co2_rate = _canonical_vehicle_fuel_and_co2_rates(
            problem, vehicle
        )
        vehicle_type = str(getattr(vehicle, "vehicle_type", "") or "")
        is_electric = is_electric_vehicle(problem, vehicle)

        def append_event(
            *,
            duty_id: str,
            event_type: str,
            sequence: int,
            start_min: int,
            end_min: int,
            from_location_id: str,
            to_location_id: str,
            previous_trip_id: str,
            next_trip_id: str,
            reference_trip,
        ) -> None:
            duration_min = max(int(end_min) - int(start_min), 0)
            if duration_min <= 0:
                return
            event_id = (
                f"{vehicle_id}:{duty_id}:{sequence}:{event_type}:"
                f"{previous_trip_id or 'depot'}:{next_trip_id or 'depot'}"
            )
            if event_id in seen_event_ids:
                raise ValueError(
                    f"duplicate canonical movement event generated: {event_id}"
                )
            seen_event_ids.add(event_id)
            distance_km = deadhead_distance_km(problem, duration_min)
            energy_rate = vehicle_energy_rate_kwh_per_km(
                problem,
                vehicle,
                reference_trip,
            )
            bev_energy_kwh = distance_km * energy_rate if is_electric else 0.0
            ice_fuel_liter = distance_km * fuel_rate if not is_electric else 0.0
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "operator_id": operator_id,
                    "event_id": event_id,
                    "event_type": event_type,
                    "vehicle_id": vehicle_id,
                    "vehicle_type": vehicle_type,
                    "duty_id": duty_id,
                    "event_start": _canonical_datetime_from_min(
                        base_date, start_min
                    ).isoformat(),
                    "event_end": _canonical_datetime_from_min(
                        base_date, end_min
                    ).isoformat(),
                    "duration_min": float(duration_min),
                    "distance_km": float(distance_km),
                    "from_location_id": from_location_id,
                    "to_location_id": to_location_id,
                    "previous_trip_id": previous_trip_id,
                    "next_trip_id": next_trip_id,
                    "bev_energy_kwh": float(bev_energy_kwh),
                    "ice_fuel_liter": float(ice_fuel_liter),
                    "ice_co2_kg": float(ice_fuel_liter * co2_rate),
                    "distance_method": "deadhead_minutes_x_problem_deadhead_speed",
                    "energy_method": "distance_x_canonical_vehicle_energy_rate",
                    "fuel_method": "distance_x_canonical_vehicle_fuel_rate",
                    "provenance_mode": "solver_plan_exact",
                    "created_by_stage": "canonical_bff_export",
                }
            )

        for duty in duties:
            duty_id = str(getattr(duty, "duty_id", "") or "")
            duty_legs = list(getattr(duty, "legs", ()) or ())
            for index, leg in enumerate(duty_legs):
                trip_id = str(getattr(leg.trip, "trip_id", "") or "")
                trip = problem_trip_by_id.get(trip_id)
                if trip is None:
                    raise ValueError(
                        f"canonical movement export cannot resolve trip {trip_id!r}"
                    )
                deadhead_min = max(
                    int(getattr(leg, "deadhead_from_prev_min", 0) or 0),
                    0,
                )
                if deadhead_min <= 0:
                    continue
                previous_trip = (
                    problem_trip_by_id.get(
                        str(getattr(duty_legs[index - 1].trip, "trip_id", "") or "")
                    )
                    if index > 0
                    else None
                )
                append_event(
                    duty_id=duty_id,
                    event_type="startup" if index == 0 else "connection",
                    sequence=index,
                    start_min=int(trip.departure_min) - deadhead_min,
                    end_min=int(trip.departure_min),
                    from_location_id=(
                        str(getattr(vehicle, "home_depot_id", "") or "")
                        if previous_trip is None
                        else str(previous_trip.destination or "")
                    ),
                    to_location_id=str(trip.origin or ""),
                    previous_trip_id=(
                        str(previous_trip.trip_id) if previous_trip is not None else ""
                    ),
                    next_trip_id=str(trip.trip_id),
                    reference_trip=trip,
                )
            if not duty_legs:
                continue
            last_trip_id = str(
                getattr(duty_legs[-1].trip, "trip_id", "") or ""
            )
            last_trip = problem_trip_by_id.get(last_trip_id)
            if last_trip is None:
                raise ValueError(
                    f"canonical movement export cannot resolve final trip "
                    f"{last_trip_id!r}"
                )
            return_exists, return_min = return_deadhead_min_to_home(
                problem,
                vehicle,
                last_trip,
            )
            if return_exists and return_min > 0:
                append_event(
                    duty_id=duty_id,
                    event_type="terminal_return",
                    sequence=len(duty_legs),
                    start_min=int(last_trip.arrival_min),
                    end_min=int(last_trip.arrival_min) + int(return_min),
                    from_location_id=str(last_trip.destination or ""),
                    to_location_id=str(
                        getattr(vehicle, "home_depot_id", "") or ""
                    ),
                    previous_trip_id=str(last_trip.trip_id),
                    next_trip_id="",
                    reference_trip=last_trip,
                )
    rows.sort(
        key=lambda row: (
            str(row.get("vehicle_id") or ""),
            str(row.get("event_start") or ""),
            str(row.get("event_id") or ""),
        )
    )
    return rows


def _canonical_soc_event_rows(
    *,
    problem,
    engine_result,
    scenario_id: str,
    base_date: date,
) -> List[Dict[str, Any]]:
    problem_trip_by_id = problem.trip_by_id()
    vehicle_by_id = {str(vehicle.vehicle_id): vehicle for vehicle in problem.vehicles}
    charge_slots_by_vehicle: Dict[str, List[Any]] = defaultdict(list)
    for slot in engine_result.plan.charging_slots:
        charge_slots_by_vehicle[str(slot.vehicle_id)].append(slot)

    rows: List[Dict[str, Any]] = []
    duties_by_vehicle = engine_result.plan.duties_by_vehicle()
    for vehicle_id, duties in duties_by_vehicle.items():
        vehicle = vehicle_by_id.get(vehicle_id)
        if vehicle is None or str(getattr(vehicle, "vehicle_type", "") or "").upper() not in {"BEV", "PHEV", "FCEV"}:
            continue
        battery_kwh = max(float(getattr(vehicle, "battery_capacity_kwh", 0.0) or 0.0), 0.0)
        min_soc = max(float(getattr(vehicle, "reserve_soc", 0.0) or 0.0), 0.0)
        max_soc = battery_kwh if battery_kwh > 0.0 else 0.0
        current_soc = _canonical_vehicle_initial_soc_kwh(vehicle)
        events: List[tuple[int, int, Dict[str, Any]]] = []
        for duty in duties:
            prev_trip = None
            for leg in duty.legs:
                dispatch_trip = leg.trip
                trip_id = str(dispatch_trip.trip_id or "")
                problem_trip = problem_trip_by_id.get(trip_id)
                if problem_trip is None:
                    continue
                deadhead_min = max(int(getattr(leg, "deadhead_from_prev_min", 0) or 0), 0)
                if deadhead_min > 0:
                    events.append(
                        (
                            int(dispatch_trip.departure_min) - deadhead_min,
                            0,
                            {
                                "event_type": "deadhead",
                                "trip_id": "",
                                "route_id": "",
                                "location_id": str(getattr(prev_trip, "destination", "") or getattr(vehicle, "home_depot_id", "") or ""),
                                "delta_kwh": -_canonical_estimated_deadhead_energy_kwh(
                                    problem,
                                    deadhead_min=deadhead_min,
                                    trip_energy_kwh=float(getattr(problem_trip, "energy_kwh", 0.0) or 0.0),
                                    trip_distance_km=float(getattr(problem_trip, "distance_km", 0.0) or 0.0),
                                ),
                            },
                        )
                    )
                events.append(
                    (
                        int(dispatch_trip.departure_min),
                        1,
                        {
                            "event_type": "service_trip",
                            "trip_id": trip_id,
                            "route_id": str(getattr(dispatch_trip, "route_id", problem_trip.route_id) or problem_trip.route_id),
                            "location_id": str(getattr(dispatch_trip, "origin_stop_id", "") or dispatch_trip.origin or ""),
                            "delta_kwh": -max(float(getattr(problem_trip, "energy_kwh", 0.0) or 0.0), 0.0),
                        },
                    )
                )
                prev_trip = dispatch_trip
        for start_slot, end_slot, _charger_id, avg_charge_kw, avg_discharge_kw, location_id in _canonical_charge_segments(
            problem,
            charge_slots_by_vehicle.get(vehicle_id, []),
            fallback_location_id=str(getattr(vehicle, "home_depot_id", "") or ""),
        ):
            duration_h = max(end_slot - start_slot, 0) * max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1) / 60.0
            events.append(
                (
                    _canonical_horizon_start_min(problem) + start_slot * max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1),
                    2,
                    {
                        "event_type": "charge_segment",
                        "trip_id": "",
                        "route_id": "",
                        "location_id": location_id,
                        "delta_kwh": (avg_charge_kw - avg_discharge_kw) * duration_h,
                    },
                )
            )
        events.sort(key=lambda item: (item[0], item[1]))
        for minute_from_midnight, _order, payload in events:
            before = current_soc
            delta_kwh = float(payload.get("delta_kwh", 0.0) or 0.0)
            after = current_soc + delta_kwh
            current_soc = after
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "vehicle_id": vehicle_id,
                    "event_time": _canonical_datetime_from_min(base_date, minute_from_midnight).isoformat(),
                    "event_type": str(payload.get("event_type") or ""),
                    "trip_id": str(payload.get("trip_id") or ""),
                    "route_id": str(payload.get("route_id") or ""),
                    "location_id": str(payload.get("location_id") or ""),
                    "soc_kwh_before": before,
                    "soc_kwh_after": after,
                    "soc_pct_before": (before / battery_kwh * 100.0) if battery_kwh > 0.0 else 0.0,
                    "soc_pct_after": (after / battery_kwh * 100.0) if battery_kwh > 0.0 else 0.0,
                    "delta_kwh": delta_kwh,
                    "battery_capacity_kwh": battery_kwh,
                    "energy_consumed_kwh": max(-delta_kwh, 0.0),
                    "energy_charged_kwh": max(delta_kwh, 0.0),
                    "reserve_margin_kwh": after - min_soc,
                    "min_soc_constraint_kwh": min_soc,
                    "max_soc_constraint_kwh": max_soc,
                }
            )
    rows.sort(key=lambda row: (str(row.get("vehicle_id", "")), str(row.get("event_time", ""))))
    return rows


def _canonical_depot_power_rows_5min(
    *,
    problem,
    engine_result,
    scenario_id: str,
    base_date: date,
    operator_id: str = "UNKNOWN_OPERATOR",
) -> List[Dict[str, Any]]:
    plan = engine_result.plan
    flow_ctx = _canonical_energy_flow_context(problem, plan)
    slot_count = max(
        len(problem.price_slots),
        max((int(getattr(slot, "slot_index", 0) or 0) + 1 for slot in plan.charging_slots), default=0),
        max(
            (
                int(slot_idx) + 1
                for key in (
                    "grid_to_bus_kwh_by_depot_slot",
                    "pv_to_bus_kwh_by_depot_slot",
                    "bess_to_bus_kwh_by_depot_slot",
                    "pv_to_bess_kwh_by_depot_slot",
                    "grid_to_bess_kwh_by_depot_slot",
                    "pv_curtail_kwh_by_depot_slot",
                    "bess_soc_kwh_by_depot_slot",
                    "bess_soc_start_kwh_by_depot_slot",
                    "bess_soc_end_kwh_by_depot_slot",
                    "contract_over_limit_kwh_by_depot_slot",
                )
                for slot_map in dict(flow_ctx.get(key) or {}).values()
                for slot_idx in dict(slot_map or {}).keys()
            ),
            default=0,
        ),
    )
    if slot_count <= 0:
        return []
    timestep_min = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
    timestep_h = timestep_min / 60.0
    output_slot_min = float(timestep_min)
    output_slot_h = output_slot_min / 60.0
    slot_scale = output_slot_min / float(timestep_min)

    slot_values_by_depot: Dict[str, Dict[int, Dict[str, float]]] = defaultdict(dict)
    for depot_id in list(flow_ctx["depot_ids"]):
        for slot_idx in range(slot_count):
            grid_to_bus = float((flow_ctx["grid_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_to_bus = float((flow_ctx["pv_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            bess_to_bus = float((flow_ctx["bess_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_to_bess = float((flow_ctx["pv_to_bess_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            grid_to_bess = float((flow_ctx["grid_to_bess_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            raw_pv_curtail = float((flow_ctx["pv_curtail_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_generation = float((flow_ctx["pv_generation_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_curtail = compute_pv_curtail_kwh(pv_generation, pv_to_bus, pv_to_bess) if pv_generation > 0.0 else max(raw_pv_curtail, 0.0)
            bess_soc_kwh = float((flow_ctx["bess_soc_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            bess_soc_start_kwh = float((flow_ctx["bess_soc_start_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, bess_soc_kwh) or 0.0)
            bess_soc_end_kwh = float((flow_ctx["bess_soc_end_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, bess_soc_kwh) or 0.0)
            contract_over_limit_kwh = float((flow_ctx["contract_over_limit_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            contract_limit_kw = float((flow_ctx["depot_limit_kw"].get(depot_id, 0.0)) or 0.0)
            grid_import_for_contract_kwh = grid_to_bus + grid_to_bess
            if contract_limit_kw > 0.0:
                contract_over_limit_kwh = max(
                    contract_over_limit_kwh,
                    grid_import_for_contract_kwh - (contract_limit_kw * timestep_h),
                    0.0,
                )
            slot_values_by_depot[depot_id][slot_idx] = {
                "grid_import_kw": grid_import_for_contract_kwh / timestep_h,
                "grid_import_for_contract_kwh": grid_import_for_contract_kwh,
                "grid_import_for_contract_kw": grid_import_for_contract_kwh / timestep_h,
                "pv_generation_kw": pv_generation / timestep_h,
                "pv_used_for_charging_kw": pv_to_bus / timestep_h,
                "pv_used_for_building_kw": 0.0,
                "pv_curtailed_kw": pv_curtail / timestep_h,
                "building_load_kw": 0.0,
                "battery_storage_charge_kw": (pv_to_bess + grid_to_bess) / timestep_h,
                "battery_storage_discharge_kw": bess_to_bus / timestep_h,
                "total_charge_kw": (grid_to_bus + pv_to_bus + bess_to_bus) / timestep_h,
                "net_load_kw": (grid_to_bus + grid_to_bess) / timestep_h,
                "grid_to_bus_kwh": grid_to_bus,
                "pv_to_bus_kwh": pv_to_bus,
                "bess_to_bus_kwh": bess_to_bus,
                "pv_to_bess_kwh": pv_to_bess,
                "grid_to_bess_kwh": grid_to_bess,
                "bus_charge_from_grid_kwh": grid_to_bus,
                "bus_charge_from_bess_kwh": bess_to_bus,
                "bess_soc_kwh": bess_soc_kwh,
                "bess_soc_start_kwh": bess_soc_start_kwh,
                "bess_soc_end_kwh": bess_soc_end_kwh,
                "contract_limit_kw": contract_limit_kw,
                "contract_over_limit_kwh": contract_over_limit_kwh,
                "contract_over_limit_kw": contract_over_limit_kwh / timestep_h if timestep_h > 0.0 else 0.0,
            }

    rows: List[Dict[str, Any]] = []
    for depot_id, slot_map in slot_values_by_depot.items():
        peak_grid = max((values.get("grid_import_kw", 0.0) for values in slot_map.values()), default=0.0)
        cumulative_grid_kwh = 0.0
        cumulative_grid_cost = 0.0
        for slot_idx in range(slot_count):
            minute = slot_idx * timestep_min
            slot_idx = min(int(minute // timestep_min), max(slot_count - 1, 0))
            values = slot_map.get(slot_idx, {})
            grid_import_kw = float(values.get("grid_import_kw", 0.0) or 0.0)
            pv_generation_kw = float(values.get("pv_generation_kw", 0.0) or 0.0)
            pv_curtailed_kw = float(values.get("pv_curtailed_kw", 0.0) or 0.0)
            grid_to_bus_hourly_source_kwh = float(values.get("grid_to_bus_kwh", 0.0) or 0.0)
            pv_to_bus_hourly_source_kwh = float(values.get("pv_to_bus_kwh", 0.0) or 0.0)
            bess_to_bus_hourly_source_kwh = float(values.get("bess_to_bus_kwh", 0.0) or 0.0)
            pv_to_bess_hourly_source_kwh = float(values.get("pv_to_bess_kwh", 0.0) or 0.0)
            grid_to_bess_hourly_source_kwh = float(values.get("grid_to_bess_kwh", 0.0) or 0.0)
            grid_import_for_contract_hourly_source_kwh = float(values.get("grid_import_for_contract_kwh", 0.0) or 0.0)
            bus_charge_from_grid_hourly_source_kwh = float(values.get("bus_charge_from_grid_kwh", 0.0) or 0.0)
            bus_charge_from_bess_hourly_source_kwh = float(values.get("bus_charge_from_bess_kwh", 0.0) or 0.0)
            contract_over_limit_hourly_source_kwh = float(values.get("contract_over_limit_kwh", 0.0) or 0.0)
            bess_soc_kwh = float(values.get("bess_soc_kwh", 0.0) or 0.0)
            bess_soc_start_kwh = float(values.get("bess_soc_start_kwh", bess_soc_kwh) or 0.0)
            bess_soc_end_kwh = float(values.get("bess_soc_end_kwh", bess_soc_kwh) or 0.0)
            grid_to_bus_slot_kwh = grid_to_bus_hourly_source_kwh * slot_scale
            pv_to_bus_slot_kwh = pv_to_bus_hourly_source_kwh * slot_scale
            bess_to_bus_slot_kwh = bess_to_bus_hourly_source_kwh * slot_scale
            pv_to_bess_slot_kwh = pv_to_bess_hourly_source_kwh * slot_scale
            grid_to_bess_slot_kwh = grid_to_bess_hourly_source_kwh * slot_scale
            grid_import_for_contract_slot_kwh = grid_import_for_contract_hourly_source_kwh * slot_scale
            bus_charge_from_grid_slot_kwh = bus_charge_from_grid_hourly_source_kwh * slot_scale
            bus_charge_from_bess_slot_kwh = bus_charge_from_bess_hourly_source_kwh * slot_scale
            contract_over_limit_slot_kwh = contract_over_limit_hourly_source_kwh * slot_scale
            grid_import_kwh = grid_to_bus_slot_kwh + grid_to_bess_slot_kwh
            pv_generation_kwh = pv_generation_kw * output_slot_h
            pv_curtail_kwh = pv_curtailed_kw * output_slot_h
            bus_charge_total_kwh = grid_to_bus_slot_kwh + pv_to_bus_slot_kwh + bess_to_bus_slot_kwh
            price = float(flow_ctx["price_by_slot"].get(slot_idx, 0.0) or 0.0)
            grid_purchase_cost = grid_import_kwh * price
            cumulative_grid_kwh += grid_import_kwh
            cumulative_grid_cost += grid_purchase_cost
            power_balance_error = grid_import_kwh - grid_to_bus_slot_kwh - grid_to_bess_slot_kwh
            pv_balance_error = pv_generation_kwh - pv_to_bus_slot_kwh - pv_to_bess_slot_kwh - pv_curtail_kwh
            timestamp = (
                datetime.combine(base_date, datetime.min.time(), timezone(timedelta(hours=9)))
                + timedelta(minutes=_canonical_horizon_start_min(problem) + minute)
            ).isoformat()
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "run_id": str(getattr(engine_result, "run_id", "") or ""),
                    "operator_id": str(operator_id or "UNKNOWN_OPERATOR"),
                    "timestamp": timestamp,
                    "depot_id": depot_id,
                    "slot_index": slot_idx,
                    "slot_minutes": output_slot_min,
                    "total_charge_kw": float(values.get("total_charge_kw", 0.0) or 0.0),
                    "grid_import_kw": grid_import_kw,
                    "grid_import_kwh": grid_import_kwh,
                    "grid_import_cumulative_kwh": cumulative_grid_kwh,
                    "grid_import_for_contract_kw": float(values.get("grid_import_for_contract_kw", grid_import_kw) or 0.0),
                    "grid_import_slot_kwh": grid_import_kw * output_slot_h,
                    "grid_import_for_contract_slot_kwh": grid_import_for_contract_slot_kwh,
                    "grid_import_hourly_source_kwh": grid_to_bus_hourly_source_kwh + grid_to_bess_hourly_source_kwh,
                    "grid_import_for_contract_hourly_source_kwh": grid_import_for_contract_hourly_source_kwh,
                    "bus_charge_from_grid_kwh": bus_charge_from_grid_slot_kwh,
                    "bus_charge_from_grid_slot_kwh": bus_charge_from_grid_slot_kwh,
                    "bus_charge_from_grid_hourly_source_kwh": bus_charge_from_grid_hourly_source_kwh,
                    "bus_charge_from_bess_kwh": bus_charge_from_bess_slot_kwh,
                    "bus_charge_from_bess_slot_kwh": bus_charge_from_bess_slot_kwh,
                    "bus_charge_from_bess_hourly_source_kwh": bus_charge_from_bess_hourly_source_kwh,
                    "grid_to_bus_kwh": grid_to_bus_slot_kwh,
                    "grid_to_bus_slot_kwh": grid_to_bus_slot_kwh,
                    "grid_to_bus_hourly_source_kwh": grid_to_bus_hourly_source_kwh,
                    "pv_to_bus_kwh": pv_to_bus_slot_kwh,
                    "pv_to_bus_slot_kwh": pv_to_bus_slot_kwh,
                    "pv_to_bus_hourly_source_kwh": pv_to_bus_hourly_source_kwh,
                    "bess_to_bus_kwh": bess_to_bus_slot_kwh,
                    "bess_to_bus_slot_kwh": bess_to_bus_slot_kwh,
                    "bess_to_bus_hourly_source_kwh": bess_to_bus_hourly_source_kwh,
                    "pv_to_bess_kwh": pv_to_bess_slot_kwh,
                    "pv_to_bess_slot_kwh": pv_to_bess_slot_kwh,
                    "pv_to_bess_hourly_source_kwh": pv_to_bess_hourly_source_kwh,
                    "grid_to_bess_kwh": grid_to_bess_slot_kwh,
                    "grid_import_total_kwh": grid_import_kwh,
                    "grid_to_bess_slot_kwh": grid_to_bess_slot_kwh,
                    "grid_to_bess_hourly_source_kwh": grid_to_bess_hourly_source_kwh,
                    "bess_soc_kwh": bess_soc_end_kwh,
                    "bess_soc_start_kwh": bess_soc_start_kwh,
                    "bess_soc_end_kwh": bess_soc_end_kwh,
                    "pv_generation_kw": pv_generation_kw,
                    "pv_generation_kwh": pv_generation_kwh,
                    "pv_generation_slot_kwh": pv_generation_kw * output_slot_h,
                    "pv_used_for_charging_kw": float(values.get("pv_used_for_charging_kw", 0.0) or 0.0),
                    "pv_used_for_building_kw": float(values.get("pv_used_for_building_kw", 0.0) or 0.0),
                    "pv_curtailed_kw": pv_curtailed_kw,
                    "pv_curtail_kwh": pv_curtail_kwh,
                    "pv_curtailed_slot_kwh": pv_curtailed_kw * output_slot_h,
                    "building_load_kw": float(values.get("building_load_kw", 0.0) or 0.0),
                    "battery_storage_charge_kw": float(values.get("battery_storage_charge_kw", 0.0) or 0.0),
                    "battery_storage_discharge_kw": float(values.get("battery_storage_discharge_kw", 0.0) or 0.0),
                    "net_load_kw": float(values.get("net_load_kw", 0.0) or 0.0),
                    "contract_limit_kw": float(values.get("contract_limit_kw", 0.0) or 0.0),
                    "contract_over_limit_kwh": contract_over_limit_slot_kwh,
                    "contract_over_limit_slot_kwh": contract_over_limit_slot_kwh,
                    "contract_over_limit_hourly_source_kwh": contract_over_limit_hourly_source_kwh,
                    "contract_over_limit_kw": float(values.get("contract_over_limit_kw", 0.0) or 0.0),
                    "contract_limit_exceeded": float(values.get("contract_over_limit_kwh", 0.0) or 0.0) > 1.0e-9,
                    "demand_peak_candidate": abs(float(values.get("grid_import_kw", 0.0) or 0.0) - peak_grid) <= 1.0e-9,
                    "bus_charge_total_kwh": bus_charge_total_kwh,
                    "energy_price_yen_per_kwh": price,
                    "grid_purchase_cost_jpy": grid_purchase_cost,
                    "grid_purchase_cumulative_cost_jpy": cumulative_grid_cost,
                    "power_balance_error_kwh": power_balance_error,
                    "pv_balance_error_kwh": pv_balance_error,
                    "demand_charge_window_flag": bool(flow_ctx["demand_flag_by_slot"].get(slot_idx, False)),
                    "source_provenance_exact": bool(flow_ctx["source_provenance_exact"]),
                }
            )
    rows.sort(key=lambda row: (str(row.get("depot_id", "")), str(row.get("timestamp", ""))))
    return rows


def _research_timestamp_parts(timestamp: Any) -> tuple[str, str]:
    try:
        parsed = datetime.fromisoformat(str(timestamp))
        return parsed.date().isoformat(), parsed.strftime("%H:%M")
    except ValueError:
        text = str(timestamp or "")
        return text[:10], text[11:16]


def _research_rows_from_depot_power(
    depot_power_rows: List[Dict[str, Any]],
    *,
    base_date: date,
) -> Dict[str, List[Dict[str, Any]]]:
    rows_by_depot_time: Dict[tuple[str, str], Dict[str, Any]] = {}
    depot_ids = sorted({str(row.get("depot_id") or "") for row in depot_power_rows if str(row.get("depot_id") or "")})
    slot_minutes = int(float(depot_power_rows[0].get("slot_minutes", 30) or 30)) if depot_power_rows else 30
    for row in depot_power_rows:
        _out_date, out_time = _research_timestamp_parts(row.get("timestamp"))
        depot_id = str(row.get("depot_id") or "")
        if depot_id:
            rows_by_depot_time[(depot_id, out_time)] = row
    initial_bess_soc_by_depot: Dict[str, float] = {}
    for row in depot_power_rows:
        depot_id = str(row.get("depot_id") or "")
        if depot_id and depot_id not in initial_bess_soc_by_depot:
            initial_bess_soc_by_depot[depot_id] = float(
                row.get("bess_soc_start_kwh", row.get("bess_soc_end_kwh", row.get("bess_soc_kwh", 0.0))) or 0.0
            )
    grid_rows: List[Dict[str, Any]] = []
    pv_rows: List[Dict[str, Any]] = []
    flow_rows: List[Dict[str, Any]] = []
    charge_rows: List[Dict[str, Any]] = []
    for depot_id in depot_ids:
        bess_soc_carry = float(initial_bess_soc_by_depot.get(depot_id, 0.0) or 0.0)
        for minute in range(0, 24 * 60, slot_minutes):
            out_time = f"{minute // 60:02d}:{minute % 60:02d}"
            row = rows_by_depot_time.get((depot_id, out_time), {})
            if row:
                bess_soc_start_kwh = float(row.get("bess_soc_start_kwh", row.get("bess_soc_kwh", bess_soc_carry)) or 0.0)
                bess_soc_end_kwh = float(row.get("bess_soc_end_kwh", row.get("bess_soc_kwh", bess_soc_start_kwh)) or 0.0)
            else:
                bess_soc_start_kwh = bess_soc_carry
                bess_soc_end_kwh = bess_soc_carry
            bess_soc_kwh = bess_soc_end_kwh
            bess_soc_carry = bess_soc_end_kwh
            out_date = base_date.isoformat()
            energy_price = float(
                row.get(
                    "energy_price_yen_per_kwh",
                    row.get("grid_energy_price_yen_per_kwh", 0.0),
                )
                or 0.0
            )
            base = {
                "date": out_date,
                "time": out_time,
                "depot_id": depot_id,
            }
            grid_import_kw = float(row.get("grid_import_kw", 0.0) or 0.0)
            grid_import_for_contract_kw = float(row.get("grid_import_for_contract_kw", grid_import_kw) or 0.0)
            contract_limit_kw = float(row.get("contract_limit_kw", 0.0) or 0.0)
            grid_rows.append(
                {
                    **base,
                    "grid_import_kw": grid_import_kw,
                    "grid_import_for_contract_kw": grid_import_for_contract_kw,
                    "grid_import_slot_kwh": float(row.get("grid_import_slot_kwh", 0.0) or 0.0),
                    "grid_import_for_contract_slot_kwh": float(row.get("grid_import_for_contract_slot_kwh", row.get("grid_import_slot_kwh", 0.0)) or 0.0),
                    "contract_limit_kw": contract_limit_kw,
                    "contract_over_limit_slot_kwh": float(row.get("contract_over_limit_slot_kwh", 0.0) or 0.0),
                }
            )
            pv_rows.append(
                {
                    **base,
                    "pv_generation_kw": float(row.get("pv_generation_kw", 0.0) or 0.0),
                    "pv_generation_slot_kwh": float(row.get("pv_generation_slot_kwh", 0.0) or 0.0),
                    "pv_curtailed_slot_kwh": float(row.get("pv_curtailed_slot_kwh", 0.0) or 0.0),
                }
            )
            flow_rows.append(
                {
                    **base,
                    "pv_generation_kwh": float(row.get("pv_generation_slot_kwh", row.get("pv_generation_kwh", 0.0)) or 0.0),
                    "pv_generation_slot_kwh": float(row.get("pv_generation_slot_kwh", row.get("pv_generation_kwh", 0.0)) or 0.0),
                    "pv_curtailed_kwh": float(row.get("pv_curtailed_slot_kwh", row.get("pv_curtail_kwh", 0.0)) or 0.0),
                    "pv_curtailed_slot_kwh": float(row.get("pv_curtailed_slot_kwh", row.get("pv_curtail_kwh", 0.0)) or 0.0),
                    "grid_to_bus_slot_kwh": float(row.get("grid_to_bus_slot_kwh", 0.0) or 0.0),
                    "pv_to_bus_slot_kwh": float(row.get("pv_to_bus_slot_kwh", 0.0) or 0.0),
                    "pv_to_bess_slot_kwh": float(row.get("pv_to_bess_slot_kwh", 0.0) or 0.0),
                    "bess_to_bus_slot_kwh": float(row.get("bess_to_bus_slot_kwh", 0.0) or 0.0),
                    "grid_to_bess_slot_kwh": float(row.get("grid_to_bess_slot_kwh", 0.0) or 0.0),
                    "grid_import_for_contract_slot_kwh": float(row.get("grid_import_for_contract_slot_kwh", row.get("grid_import_slot_kwh", 0.0)) or 0.0),
                    "bus_charge_from_grid_slot_kwh": float(row.get("bus_charge_from_grid_slot_kwh", row.get("grid_to_bus_slot_kwh", 0.0)) or 0.0),
                    "bus_charge_from_bess_slot_kwh": float(row.get("bus_charge_from_bess_slot_kwh", row.get("bess_to_bus_slot_kwh", 0.0)) or 0.0),
                    "bess_soc_kwh": bess_soc_kwh,
                    "bess_soc_start_kwh": bess_soc_start_kwh,
                    "bess_soc_end_kwh": bess_soc_end_kwh,
                    "energy_price_yen_per_kwh": energy_price,
                    "grid_purchase_cost_jpy": float(row.get("grid_purchase_cost_jpy", 0.0) or 0.0),
                    "contract_limit_kw": float(row.get("contract_limit_kw", 0.0) or 0.0),
                    "contract_over_limit_kwh": float(row.get("contract_over_limit_kwh", 0.0) or 0.0),
                    "contract_over_limit_slot_kwh": float(row.get("contract_over_limit_slot_kwh", 0.0) or 0.0),
                    "contract_over_limit_kw": float(row.get("contract_over_limit_kw", 0.0) or 0.0),
                    "contract_limit_exceeded": bool(row.get("contract_limit_exceeded", False)),
                    "demand_charge_window_flag": bool(row.get("demand_charge_window_flag", False)),
                }
            )
            total_charge_kw = float(row.get("total_charge_kw", 0.0) or 0.0)
            charge_rows.append(
                {
                    **base,
                    "total_bus_charge_kw": total_charge_kw,
                    "total_bus_charge_slot_kwh": total_charge_kw * (float(row.get("slot_minutes", slot_minutes) or slot_minutes) / 60.0),
                }
            )
    return {
        "grid_import_timeseries.csv": grid_rows,
        "pv_generation_timeseries.csv": pv_rows,
        "energy_flow_timeseries.csv": flow_rows,
        "bus_charging_total_timeseries.csv": charge_rows,
    }


def _hhmm_to_minute(value: str) -> int:
    hour, minute = str(value or "00:00").split(":", 1)
    return (int(hour) % 24) * 60 + (int(minute) % 60)


def _research_depot_time_index(
    depot_power_rows: List[Dict[str, Any]],
) -> tuple[List[str], Dict[tuple[str, str], Dict[str, Any]]]:
    rows_by_depot_time: Dict[tuple[str, str], Dict[str, Any]] = {}
    depot_ids = sorted({str(row.get("depot_id") or "") for row in depot_power_rows if str(row.get("depot_id") or "")})
    for row in depot_power_rows:
        _out_date, out_time = _research_timestamp_parts(row.get("timestamp"))
        depot_id = str(row.get("depot_id") or "")
        if depot_id:
            rows_by_depot_time[(depot_id, out_time)] = row
    return depot_ids, rows_by_depot_time


def _research_slot_index_for_time(problem, time_hhmm: str) -> int:
    timestep_min = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
    minute = _hhmm_to_minute(time_hhmm)
    horizon_start = _canonical_horizon_start_min(problem)
    if minute < horizon_start:
        minute += 24 * 60
    return max((minute - horizon_start) // timestep_min, 0)


def _research_vehicle_maps(problem) -> tuple[Dict[str, Any], Dict[str, Any]]:
    vehicle_by_id = {str(getattr(vehicle, "vehicle_id", "") or ""): vehicle for vehicle in list(getattr(problem, "vehicles", ()) or ())}
    vehicle_type_by_id = {
        str(getattr(vehicle_type, "vehicle_type_id", "") or ""): vehicle_type
        for vehicle_type in list(getattr(problem, "vehicle_types", ()) or ())
    }
    return vehicle_by_id, vehicle_type_by_id


def _research_vehicle_fuel_rate_l_per_km(vehicle: Any, vehicle_type: Any) -> float:
    for source in (vehicle, vehicle_type):
        value = getattr(source, "fuel_consumption_l_per_km", None) if source is not None else None
        if value is not None:
            parsed = max(float(value or 0.0), 0.0)
            if parsed > 0.0:
                return parsed
    return 0.0


def _research_vehicle_co2_kg_per_l(problem, vehicle_type: Any) -> float:
    value = getattr(vehicle_type, "co2_emission_kg_per_l", None) if vehicle_type is not None else None
    parsed = max(float(value or 0.0), 0.0) if value is not None else 0.0
    return parsed if parsed > 0.0 else max(float(getattr(problem.scenario, "ice_co2_kg_per_l", 0.0) or 0.0), 0.0)


def _research_is_electric_vehicle(vehicle: Any, vehicle_type: Any) -> bool:
    powertrain = str(getattr(vehicle_type, "powertrain_type", "") or getattr(vehicle, "vehicle_type", "") or "").upper()
    return powertrain in {"BEV", "PHEV", "FCEV"}


def _research_spread_event_to_5min(
    bucket: Dict[tuple[str, str], Dict[str, float]],
    *,
    base_date: date,
    depot_id: str,
    start_time: Any,
    end_time: Any,
    values: Dict[str, float],
    slot_minutes: int = 30,
) -> None:
    try:
        start = datetime.fromisoformat(str(start_time))
        end = datetime.fromisoformat(str(end_time))
    except ValueError:
        return
    if end <= start:
        return
    base_midnight = datetime.combine(base_date, datetime.min.time(), tzinfo=start.tzinfo)
    start_min = (start - base_midnight).total_seconds() / 60.0
    end_min = (end - base_midnight).total_seconds() / 60.0
    duration = max(end_min - start_min, 1.0e-9)
    bucket_start = int(start_min // slot_minutes) * slot_minutes
    while bucket_start < end_min:
        bucket_end = bucket_start + slot_minutes
        overlap = max(min(bucket_end, end_min) - max(bucket_start, start_min), 0.0)
        if overlap > 0.0:
            out_minute = int(bucket_start % (24 * 60))
            out_time = f"{out_minute // 60:02d}:{out_minute % 60:02d}"
            target = bucket.setdefault((str(depot_id), out_time), defaultdict(float))
            for key, value in values.items():
                target[key] += float(value or 0.0) * overlap / duration
        bucket_start += slot_minutes


def _research_fuel_time_bucket(
    *,
    problem,
    timeline_rows: List[Dict[str, Any]],
    refuel_rows: List[Dict[str, Any]],
    base_date: date,
    slot_minutes: int,
) -> Dict[tuple[str, str], Dict[str, float]]:
    vehicle_by_id, vehicle_type_by_id = _research_vehicle_maps(problem)
    bucket: Dict[tuple[str, str], Dict[str, float]] = {}
    for row in timeline_rows:
        vehicle_id = str(row.get("vehicle_id") or "")
        vehicle = vehicle_by_id.get(vehicle_id)
        vehicle_type = vehicle_type_by_id.get(str(getattr(vehicle, "vehicle_type", "") or row.get("vehicle_type") or ""))
        if vehicle is None or _research_is_electric_vehicle(vehicle, vehicle_type):
            continue
        fuel_rate = _research_vehicle_fuel_rate_l_per_km(vehicle, vehicle_type)
        if fuel_rate <= 0.0:
            continue
        distance_km = max(float(row.get("distance_km", 0.0) or 0.0), 0.0)
        fuel_l = distance_km * fuel_rate
        if fuel_l <= 0.0:
            continue
        co2 = fuel_l * _research_vehicle_co2_kg_per_l(problem, vehicle_type)
        state = str(row.get("state") or "")
        trip_fuel = fuel_l if state == "service" or bool(row.get("is_service")) else 0.0
        deadhead_fuel = fuel_l if state == "deadhead" or bool(row.get("is_deadhead")) else 0.0
        _research_spread_event_to_5min(
            bucket,
            base_date=base_date,
            depot_id=str(row.get("depot_id") or getattr(vehicle, "home_depot_id", "") or "depot_default"),
            start_time=row.get("start_time"),
            end_time=row.get("end_time"),
            values={
                "trip_fuel_liters": trip_fuel,
                "deadhead_fuel_liters": deadhead_fuel,
                "fuel_liters": fuel_l,
                "ice_bus_co2_kg": co2,
            },
            slot_minutes=slot_minutes,
        )
    vehicle_home = _canonical_vehicle_home_depot_map(problem)
    for row in refuel_rows:
        liters = max(float(row.get("refuel_liters", 0.0) or 0.0), 0.0)
        if liters <= 0.0:
            continue
        vehicle_id = str(row.get("vehicle_id") or "")
        depot_id = str(row.get("location_id") or vehicle_home.get(vehicle_id) or "depot_default")
        time_hhmm = str(row.get("time_hhmm") or "00:00")[:5]
        target = bucket.setdefault((depot_id, time_hhmm), defaultdict(float))
        target["refuel_liters"] += liters
    return bucket


def _research_extended_timeseries_exports(
    *,
    problem,
    engine_result,
    depot_power_rows: List[Dict[str, Any]],
    timeline_rows: List[Dict[str, Any]],
    refuel_rows: List[Dict[str, Any]],
    base_date: date,
) -> Dict[str, List[Dict[str, Any]]]:
    depot_ids, rows_by_depot_time = _research_depot_time_index(depot_power_rows)
    depot_ids = sorted(set(depot_ids) | {str(getattr(depot, "depot_id", "") or "") for depot in list(getattr(problem, "depots", ()) or ()) if str(getattr(depot, "depot_id", "") or "")})
    if not depot_ids:
        depot_ids = ["depot_default"]
    price_by_slot = {
        int(getattr(slot, "slot_index", 0) or 0): float(getattr(slot, "grid_buy_yen_per_kwh", 0.0) or 0.0)
        for slot in list(getattr(problem, "price_slots", ()) or ())
    }
    co2_by_slot = {
        int(getattr(slot, "slot_index", 0) or 0): float(getattr(slot, "co2_factor", 0.0) or 0.0)
        for slot in list(getattr(problem, "price_slots", ()) or ())
    }
    demand_flag_by_slot = {
        int(getattr(slot, "slot_index", 0) or 0): bool(float(getattr(slot, "demand_charge_weight", 0.0) or 0.0) > 0.0)
        for slot in list(getattr(problem, "price_slots", ()) or ())
    }
    assets = dict(getattr(problem, "depot_energy_assets", {}) or {})
    breakdown = dict(getattr(engine_result, "cost_breakdown", {}) or {})
    plan_metadata = dict(getattr(engine_result.plan, "metadata", {}) or {})
    solver_metadata = dict(getattr(engine_result, "solver_metadata", {}) or {})
    timestep_min = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
    pv_unit = max(float(breakdown.get("pv_marginal_charge_cost_yen_per_kwh", getattr(problem, "metadata", {}).get("pv_marginal_charge_cost_yen_per_kwh", 0.0)) or 0.0), 0.0)
    contract_penalty = max(float(plan_metadata.get("contract_overage_penalty_yen_per_kwh", solver_metadata.get("contract_overage_penalty_yen_per_kwh", getattr(problem, "metadata", {}).get("contract_overage_penalty_yen_per_kwh", 0.0))) or 0.0), 0.0)
    fuel_bucket = _research_fuel_time_bucket(
        problem=problem,
        timeline_rows=timeline_rows,
        refuel_rows=refuel_rows,
        base_date=base_date,
        slot_minutes=timestep_min,
    )
    flow_ctx = _canonical_energy_flow_context(problem, engine_result.plan)
    depot_source_exact = bool(flow_ctx.get("source_provenance_exact", False))
    vehicle_source_exact = bool(
        depot_source_exact
        and dict(getattr(engine_result.plan, "metadata", {}) or {}).get("vehicle_source_provenance_exact", False)
    )

    co2_rows: List[Dict[str, Any]] = []
    cost_rows: List[Dict[str, Any]] = []
    contract_rows: List[Dict[str, Any]] = []
    bess_rows: List[Dict[str, Any]] = []
    fuel_rows: List[Dict[str, Any]] = []
    for depot_id in depot_ids:
        asset = assets.get(depot_id)
        bess_capacity = max(float(getattr(asset, "bess_energy_kwh", 0.0) or 0.0), 0.0) if asset is not None else 0.0
        bess_bounds = _bess_soc_bounds_for_asset(asset) if asset is not None else None
        if bess_bounds is not None:
            bess_min, bess_max, bess_initial_soc = bess_bounds
        else:
            bess_min = 0.0
            bess_max = bess_capacity
            bess_initial_soc = max(float(getattr(asset, "bess_initial_soc_kwh", 0.0) or 0.0), 0.0) if asset is not None else 0.0
        bess_terminal_min = min(max(float(getattr(asset, "bess_terminal_soc_min_kwh", 0.0) or 0.0), bess_min), bess_max) if asset is not None else 0.0
        resolved_terminal_target = (
            _resolved_bess_terminal_target_kwh(asset) if asset is not None else None
        )
        has_terminal_target = resolved_terminal_target is not None
        bess_terminal_target = float(resolved_terminal_target or 0.0)
        bess_cycle_unit = max(float(getattr(asset, "bess_cycle_cost_yen_per_kwh", 0.0) or 0.0), 0.0) if asset is not None else 0.0
        carried_bess_soc = bess_initial_soc
        for minute in range(0, 24 * 60, timestep_min):
            time_hhmm = f"{minute // 60:02d}:{minute % 60:02d}"
            base = {"date": base_date.isoformat(), "time": time_hhmm, "depot_id": depot_id}
            row = rows_by_depot_time.get((depot_id, time_hhmm), {})
            slot_idx = _research_slot_index_for_time(problem, time_hhmm)
            price = float(row.get("energy_price_yen_per_kwh", price_by_slot.get(slot_idx, 0.0)) or 0.0)
            co2_factor = float(co2_by_slot.get(slot_idx, 0.0) or 0.0)
            grid_contract_kwh = float(row.get("grid_import_for_contract_slot_kwh", row.get("grid_import_slot_kwh", 0.0)) or 0.0)
            grid_to_bus_kwh = float(row.get("grid_to_bus_slot_kwh", row.get("grid_to_bus_kwh", 0.0)) or 0.0)
            grid_to_bess_kwh = float(row.get("grid_to_bess_slot_kwh", row.get("grid_to_bess_kwh", 0.0)) or 0.0)
            pv_to_bus_kwh = float(row.get("pv_to_bus_slot_kwh", row.get("pv_to_bus_kwh", 0.0)) or 0.0)
            pv_to_bess_kwh = float(row.get("pv_to_bess_slot_kwh", row.get("pv_to_bess_kwh", 0.0)) or 0.0)
            bess_to_bus_kwh = float(row.get("bess_to_bus_slot_kwh", row.get("bess_to_bus_kwh", 0.0)) or 0.0)
            contract_limit_kw = float(row.get("contract_limit_kw", 0.0) or 0.0)
            grid_contract_kw = float(row.get("grid_import_for_contract_kw", row.get("grid_import_kw", 0.0)) or 0.0)
            contract_over_kwh = float(row.get("contract_over_limit_slot_kwh", row.get("contract_over_limit_kwh", 0.0)) or 0.0)
            contract_over_cost = contract_over_kwh * contract_penalty
            pv_used_kwh = pv_to_bus_kwh + pv_to_bess_kwh
            pv_generation_kwh = float(row.get("pv_generation_slot_kwh", 0.0) or 0.0)
            pv_curtail_kwh = float(row.get("pv_curtailed_slot_kwh", 0.0) or 0.0)
            grid_co2 = grid_contract_kwh * co2_factor
            fuel_values = fuel_bucket.get((depot_id, time_hhmm), {})
            ice_co2 = float(fuel_values.get("ice_bus_co2_kg", 0.0) or 0.0)
            co2_rows.append(
                {
                    **base,
                    "grid_import_for_contract_kwh": grid_contract_kwh,
                    "grid_co2_factor_kg_per_kwh": co2_factor,
                    "grid_electricity_co2_kg": grid_co2,
                    "ice_bus_co2_kg": ice_co2,
                    "pv_operational_co2_kg": 0.0,
                    "bess_storage_operational_co2_kg": 0.0,
                    "total_co2_kg": grid_co2 + ice_co2,
                    "source_provenance_exact": depot_source_exact,
                }
            )
            cost_rows.append(
                {
                    **base,
                    "grid_energy_price_yen_per_kwh": price,
                    "grid_purchase_cost_jpy": (grid_to_bus_kwh + grid_to_bess_kwh) * price,
                    "bess_discharge_cost_jpy": bess_to_bus_kwh * bess_cycle_unit,
                    "pv_self_consumption_cost_jpy": pv_used_kwh * pv_unit,
                    "contract_overage_cost_jpy": contract_over_cost,
                    "electricity_energy_cost_jpy": ((grid_to_bus_kwh + grid_to_bess_kwh) * price) + (bess_to_bus_kwh * bess_cycle_unit) + (pv_used_kwh * pv_unit),
                    "demand_charge_window_flag": bool(row.get("demand_charge_window_flag", demand_flag_by_slot.get(slot_idx, False))),
                    "demand_peak_candidate": bool(row.get("demand_peak_candidate", False)),
                    "demand_charge_is_peak_based_not_time_additive": True,
                    "source_provenance_exact": depot_source_exact,
                }
            )
            contract_rows.append(
                {
                    **base,
                    "grid_import_for_contract_kw": grid_contract_kw,
                    "grid_import_for_contract_kwh": grid_contract_kwh,
                    "contract_limit_kw": contract_limit_kw,
                    "contract_margin_kw": contract_limit_kw - grid_contract_kw if contract_limit_kw > 0.0 else 0.0,
                    "contract_over_limit_kw": float(row.get("contract_over_limit_kw", 0.0) or 0.0),
                    "contract_over_limit_kwh": contract_over_kwh,
                    "contract_limit_exceeded": bool(row.get("contract_limit_exceeded", contract_over_kwh > 1.0e-9)),
                    "contract_overage_cost_jpy": contract_over_cost,
                    "bess_to_bus_excluded_from_contract_kwh": bess_to_bus_kwh,
                    "source_provenance_exact": depot_source_exact,
                }
            )
            if row:
                bess_soc_start = float(row.get("bess_soc_start_kwh", row.get("bess_soc_kwh", carried_bess_soc)) or 0.0)
                bess_soc_end = float(row.get("bess_soc_end_kwh", row.get("bess_soc_kwh", bess_soc_start)) or 0.0)
            else:
                bess_soc_start = carried_bess_soc
                bess_soc_end = carried_bess_soc
            if bess_bounds is not None:
                bess_soc_start = min(max(bess_soc_start, bess_min), bess_max)
                bess_soc_end = min(max(bess_soc_end, bess_min), bess_max)
            bess_soc = bess_soc_end
            carried_bess_soc = bess_soc_end
            bess_charge_kwh = pv_to_bess_kwh + grid_to_bess_kwh
            bess_rows.append(
                {
                    **base,
                    "bess_charge_kwh": bess_charge_kwh,
                    "bess_charge_kw": bess_charge_kwh / (timestep_min / 60.0),
                    "bess_discharge_kwh": bess_to_bus_kwh,
                    "bess_discharge_kw": bess_to_bus_kwh / (timestep_min / 60.0),
                    "bess_net_charge_kwh": bess_charge_kwh - bess_to_bus_kwh,
                    "pv_to_bess_kwh": pv_to_bess_kwh,
                    "grid_to_bess_kwh": grid_to_bess_kwh,
                    "bess_to_bus_kwh": bess_to_bus_kwh,
                    "bess_soc_kwh": bess_soc,
                    "bess_soc_start_kwh": bess_soc_start,
                    "bess_soc_end_kwh": bess_soc_end,
                    "bess_soc_percent": (bess_soc / bess_capacity * 100.0) if bess_capacity > 0.0 else 0.0,
                    "bess_soc_min_kwh": bess_min,
                    "bess_soc_max_kwh": bess_max,
                    "bess_terminal_soc_min_kwh": bess_terminal_min,
                    "bess_terminal_soc_target_kwh": bess_terminal_target,
                    "bess_terminal_soc_deviation_kwh": abs(bess_soc - bess_terminal_target) if has_terminal_target and minute >= 24 * 60 - timestep_min else 0.0,
                    "bess_terminal_soc_delta_kwh": bess_soc - bess_initial_soc if minute >= 24 * 60 - timestep_min else 0.0,
                    "bess_terminal_soc_violation_kwh": max(bess_terminal_min - bess_soc, 0.0) if minute >= 24 * 60 - timestep_min else 0.0,
                    "bess_cycle_cost_jpy": bess_to_bus_kwh * bess_cycle_unit,
                    "source_provenance_exact": depot_source_exact,
                }
            )
            fuel_rows.append(
                {
                    **base,
                    "trip_fuel_liters": float(fuel_values.get("trip_fuel_liters", 0.0) or 0.0),
                    "deadhead_fuel_liters": float(fuel_values.get("deadhead_fuel_liters", 0.0) or 0.0),
                    "fuel_liters": float(fuel_values.get("fuel_liters", 0.0) or 0.0),
                    "refuel_liters": float(fuel_values.get("refuel_liters", 0.0) or 0.0),
                    "net_fuel_inventory_delta_liters": float(fuel_values.get("refuel_liters", 0.0) or 0.0) - float(fuel_values.get("fuel_liters", 0.0) or 0.0),
                    "ice_bus_co2_kg": ice_co2,
                }
            )
    return {
        "co2_timeseries.csv": co2_rows,
        "cost_timeseries.csv": cost_rows,
        "contract_limit_timeseries.csv": contract_rows,
        "bess_timeseries.csv": bess_rows,
        "fuel_timeseries.csv": fuel_rows,
    }


def _research_vehicle_charging_source_timeseries_rows(
    *,
    problem,
    engine_result,
    base_date: date,
    operator_id: str = "UNKNOWN_OPERATOR",
) -> List[Dict[str, Any]]:
    timestep_min = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
    timestep_h = timestep_min / 60.0
    horizon_start = _canonical_horizon_start_min(problem)
    vehicle_by_id, _vehicle_type_by_id = _research_vehicle_maps(problem)
    active_vehicle_ids = sorted({str(getattr(slot, "vehicle_id", "") or "") for slot in list(getattr(engine_result.plan, "charging_slots", ()) or ()) if str(getattr(slot, "vehicle_id", "") or "")})
    if not active_vehicle_ids:
        return []
    flow_ctx = _canonical_energy_flow_context(problem, engine_result.plan)
    depot_source_exact = bool(flow_ctx.get("source_provenance_exact", False))
    vehicle_source_exact = bool(
        depot_source_exact
        and dict(getattr(engine_result.plan, "metadata", {}) or {}).get("vehicle_source_provenance_exact", False)
    )
    allocation_method = "solver_native" if vehicle_source_exact else "proportional_by_timestep"
    values: Dict[tuple[str, str], Dict[str, Any]] = {}
    charge_rows_by_depot_slot: Dict[tuple[str, int], List[Dict[str, Any]]] = defaultdict(list)
    for slot in list(getattr(engine_result.plan, "charging_slots", ()) or ()):
        vehicle_id = str(getattr(slot, "vehicle_id", "") or "")
        if not vehicle_id:
            continue
        source, depot_id = _canonical_charging_source_and_depot(problem, slot)
        charge_kw = max(float(getattr(slot, "charge_kw", 0.0) or 0.0), 0.0)
        discharge_kw = max(float(getattr(slot, "discharge_kw", 0.0) or 0.0), 0.0)
        net_kw = max(charge_kw - discharge_kw, 0.0)
        if net_kw <= 0.0:
            continue
        slot_idx = int(getattr(slot, "slot_index", 0) or 0)
        start_minute = horizon_start + int(getattr(slot, "slot_index", 0) or 0) * timestep_min
        out_minute = start_minute % (24 * 60)
        out_time = f"{out_minute // 60:02d}:{out_minute % 60:02d}"
        charge_kwh = net_kw * timestep_h
        if vehicle_source_exact:
            target = values.setdefault((vehicle_id, out_time), defaultdict(float))
            target[f"{source}_to_vehicle_kwh"] += charge_kwh
            target["total_charge_kwh"] += charge_kwh
            target["charge_kw"] += net_kw
            target[f"{source}_charge_kw"] += net_kw
            target["_depot_id"] = depot_id
            target["_charger_id"] = str(getattr(slot, "charger_id", "") or "")
        else:
            charge_rows_by_depot_slot[(depot_id, slot_idx)].append(
                {
                    "vehicle_id": vehicle_id,
                    "time": out_time,
                    "depot_id": depot_id,
                    "charger_id": str(getattr(slot, "charger_id", "") or ""),
                    "source": source,
                    "charge_kwh": charge_kwh,
                    "charge_kw": net_kw,
                }
            )

    if not vehicle_source_exact:
        for (depot_id, slot_idx), slot_rows in sorted(charge_rows_by_depot_slot.items(), key=lambda item: (item[0][0], item[0][1])):
            total_charge_kwh = sum(float(row.get("charge_kwh", 0.0) or 0.0) for row in slot_rows)
            if total_charge_kwh <= 0.0:
                continue
            grid_total = float((flow_ctx["grid_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            pv_total = float((flow_ctx["pv_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            bess_total = float((flow_ctx["bess_to_bus_kwh_by_depot_slot"].get(depot_id, {}) or {}).get(slot_idx, 0.0) or 0.0)
            if grid_total + pv_total + bess_total <= 1.0e-9:
                source = str(slot_rows[0].get("source") or "grid") if slot_rows else "grid"
                if source == "pv":
                    pv_total = total_charge_kwh
                elif source == "bess":
                    bess_total = total_charge_kwh
                else:
                    grid_total = total_charge_kwh
            remaining_grid = grid_total
            remaining_pv = pv_total
            remaining_bess = bess_total
            ordered = sorted(slot_rows, key=lambda row: (str(row.get("vehicle_id") or ""), str(row.get("charger_id") or "")))
            for index, row in enumerate(ordered):
                charge_kwh = float(row.get("charge_kwh", 0.0) or 0.0)
                share = charge_kwh / total_charge_kwh if total_charge_kwh > 0.0 else 0.0
                is_last = index == len(ordered) - 1
                if is_last:
                    grid_kwh = remaining_grid
                    pv_kwh = remaining_pv
                    bess_kwh = remaining_bess
                else:
                    grid_kwh = grid_total * share
                    pv_kwh = pv_total * share
                    bess_kwh = bess_total * share
                    remaining_grid -= grid_kwh
                    remaining_pv -= pv_kwh
                    remaining_bess -= bess_kwh
                vehicle_id = str(row.get("vehicle_id") or "")
                out_time = str(row.get("time") or "00:00")
                target = values.setdefault((vehicle_id, out_time), defaultdict(float))
                target["grid_to_vehicle_kwh"] += grid_kwh
                target["pv_to_vehicle_kwh"] += pv_kwh
                target["bess_to_vehicle_kwh"] += bess_kwh
                target["total_charge_kwh"] += grid_kwh + pv_kwh + bess_kwh
                target["charge_kw"] += float(row.get("charge_kw", 0.0) or 0.0)
                target["grid_charge_kw"] += grid_kwh / timestep_h if timestep_h > 0.0 else 0.0
                target["pv_charge_kw"] += pv_kwh / timestep_h if timestep_h > 0.0 else 0.0
                target["bess_charge_kw"] += bess_kwh / timestep_h if timestep_h > 0.0 else 0.0
                target["_depot_id"] = depot_id
                target["_charger_id"] = str(row.get("charger_id") or "")
    rows: List[Dict[str, Any]] = []
    for vehicle_id in active_vehicle_ids:
        vehicle = vehicle_by_id.get(vehicle_id)
        fallback_depot = str(getattr(vehicle, "home_depot_id", "") or "") if vehicle is not None else ""
        for minute in range(0, 24 * 60, timestep_min):
            time_hhmm = f"{minute // 60:02d}:{minute % 60:02d}"
            item = values.get((vehicle_id, time_hhmm), {})
            depot_id = str(item.get("_depot_id") or fallback_depot)
            grid_kwh = float(item.get("grid_to_vehicle_kwh", 0.0) or 0.0)
            pv_kwh = float(item.get("pv_to_vehicle_kwh", 0.0) or 0.0)
            bess_kwh = float(item.get("bess_to_vehicle_kwh", 0.0) or 0.0)
            rows.append(
                {
                    "date": base_date.isoformat(),
                    "time": time_hhmm,
                    "timestamp": datetime.combine(base_date, datetime.min.time()).replace(hour=minute // 60, minute=minute % 60).isoformat(),
                    "slot_minutes": timestep_min,
                    "vehicle_id": vehicle_id,
                    "operator_id": operator_id,
                    "depot_id": depot_id,
                    "route_id": "",
                    "charger_id": str(item.get("_charger_id") or ""),
                    "grid_to_vehicle_kwh": grid_kwh,
                    "pv_to_vehicle_kwh": pv_kwh,
                    "bess_to_vehicle_kwh": bess_kwh,
                    "total_charge_kwh": grid_kwh + pv_kwh + bess_kwh,
                    "charge_kw": float(item.get("charge_kw", 0.0) or 0.0),
                    "grid_charge_kw": float(item.get("grid_charge_kw", 0.0) or 0.0),
                    "pv_charge_kw": float(item.get("pv_charge_kw", 0.0) or 0.0),
                    "bess_charge_kw": float(item.get("bess_charge_kw", 0.0) or 0.0),
                    "source_provenance_exact": vehicle_source_exact,
                    "depot_source_provenance_exact": depot_source_exact,
                    "vehicle_charging_source_allocation_method": allocation_method,
                    "vehicle_charging_source_is_solver_native": vehicle_source_exact,
                    "vehicle_source_split_note": (
                        "Vehicle source split is an exact MILP vehicle/source/slot decision trace."
                        if vehicle_source_exact
                        else "Vehicle source split is post-allocated by depot/timestep site source ratios because per-vehicle source decision variables are not present."
                    ),
                }
            )
    return rows


def _research_vehicle_soc_timeseries_rows(
    *,
    problem,
    engine_result,
    base_date: date,
    timeline_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    soc_by_vehicle = _normalize_depot_slot_mapping(
        getattr(engine_result.plan, "vehicle_soc_kwh_by_vehicle_slot", {})
    )
    if not soc_by_vehicle:
        return []
    vehicle_by_id = {str(vehicle.vehicle_id): vehicle for vehicle in problem.vehicles}
    timeline_by_vehicle: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for item in timeline_rows:
        timeline_by_vehicle[str(item.get("vehicle_id") or "")].append(item)

    timestep_min = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
    points = list(range(0, 24 * 60, timestep_min))
    rows: List[Dict[str, Any]] = []
    for vehicle_id, slot_map in sorted(soc_by_vehicle.items()):
        vehicle = vehicle_by_id.get(vehicle_id)
        capacity = max(float(getattr(vehicle, "battery_capacity_kwh", 0.0) or 0.0), 0.0) if vehicle is not None else 0.0
        depot_id = str(getattr(vehicle, "home_depot_id", "") or "") if vehicle is not None else ""
        max_slot = max(slot_map.keys(), default=0)
        for minute in points:
            slot_idx = min(int(minute // timestep_min), max_slot)
            soc_kwh = float(slot_map.get(slot_idx, 0.0) or 0.0)
            timestamp = datetime.combine(base_date, datetime.min.time()) + timedelta(minutes=minute)
            state = "idle"
            for event in timeline_by_vehicle.get(vehicle_id, []):
                try:
                    start = datetime.fromisoformat(str(event.get("start_time")))
                    end = datetime.fromisoformat(str(event.get("end_time")))
                except ValueError:
                    continue
                if start <= timestamp < end:
                    state = str(event.get("state") or "idle")
                    break
            rows.append(
                {
                    "date": base_date.isoformat(),
                    "time": timestamp.strftime("%H:%M"),
                    "vehicle_id": vehicle_id,
                    "soc_kwh": soc_kwh,
                    "soc_percent": (soc_kwh / capacity * 100.0) if capacity > 0.0 else 0.0,
                    "state": state,
                    "depot_id": depot_id,
                }
            )
    return rows


def _research_fuel_summary_rows(
    *,
    problem,
    timeline_rows: List[Dict[str, Any]],
    refuel_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    vehicle_by_id = {str(vehicle.vehicle_id): vehicle for vehicle in problem.vehicles}
    refuel_by_vehicle: Dict[str, float] = defaultdict(float)
    for row in refuel_rows:
        refuel_by_vehicle[str(row.get("vehicle_id") or "")] += float(row.get("refuel_liters", 0.0) or 0.0)

    accum: Dict[str, Dict[str, float]] = defaultdict(lambda: {"trip_fuel_liters": 0.0, "deadhead_fuel_liters": 0.0})
    for row in timeline_rows:
        vehicle_id = str(row.get("vehicle_id") or "")
        vehicle = vehicle_by_id.get(vehicle_id)
        if vehicle is None or str(getattr(vehicle, "vehicle_type", "") or "").upper() in {"BEV", "PHEV", "FCEV"}:
            continue
        fuel_rate = max(float(getattr(vehicle, "fuel_consumption_l_per_km", 0.0) or 0.0), 0.0)
        if fuel_rate <= 0.0:
            continue
        liters = max(float(row.get("distance_km", 0.0) or 0.0), 0.0) * fuel_rate
        if bool(row.get("is_service")):
            accum[vehicle_id]["trip_fuel_liters"] += liters
        elif bool(row.get("is_deadhead")):
            accum[vehicle_id]["deadhead_fuel_liters"] += liters

    rows: List[Dict[str, Any]] = []
    totals = {"trip_fuel_liters": 0.0, "deadhead_fuel_liters": 0.0, "refuel_liters": 0.0}
    for vehicle_id in sorted(set(accum.keys()) | set(refuel_by_vehicle.keys())):
        vehicle = vehicle_by_id.get(vehicle_id)
        vehicle_type = str(getattr(vehicle, "vehicle_type", "UNKNOWN") or "UNKNOWN") if vehicle is not None else "UNKNOWN"
        trip_l = float(accum.get(vehicle_id, {}).get("trip_fuel_liters", 0.0) or 0.0)
        deadhead_l = float(accum.get(vehicle_id, {}).get("deadhead_fuel_liters", 0.0) or 0.0)
        refuel_l = float(refuel_by_vehicle.get(vehicle_id, 0.0) or 0.0)
        fuel_l = trip_l + deadhead_l
        rows.append(
            {
                "vehicle_id": vehicle_id,
                "vehicle_type": vehicle_type,
                "fuel_liters": fuel_l,
                "trip_fuel_liters": trip_l,
                "deadhead_fuel_liters": deadhead_l,
                "refuel_liters": refuel_l,
                "unit": "L",
            }
        )
        totals["trip_fuel_liters"] += trip_l
        totals["deadhead_fuel_liters"] += deadhead_l
        totals["refuel_liters"] += refuel_l
    if rows:
        rows.append(
            {
                "vehicle_id": "TOTAL",
                "vehicle_type": "ALL_ICE",
                "fuel_liters": totals["trip_fuel_liters"] + totals["deadhead_fuel_liters"],
                "trip_fuel_liters": totals["trip_fuel_liters"],
                "deadhead_fuel_liters": totals["deadhead_fuel_liters"],
                "refuel_liters": totals["refuel_liters"],
                "unit": "L",
            }
        )
    return rows


def _canonical_vehicle_timelines_payload(timeline_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in timeline_rows:
        grouped[str(row.get("vehicle_id") or "")].append(row)
    for vehicle_id in list(grouped):
        grouped[vehicle_id] = sorted(
            grouped[vehicle_id],
            key=lambda item: (str(item.get("start_time") or ""), str(item.get("state") or "")),
        )
    return {
        "timeline_schema_version": "canonical_v1",
        "vehicle_timelines": dict(grouped),
        "vehicle_gantt_rows": list(timeline_rows),
    }


def _canonical_kpi_summary_json(
    *,
    problem,
    engine_result,
    scenario_id: str,
    soc_rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    plan = engine_result.plan
    served_trip_count = len(plan.served_trip_ids)
    total_trip_count = len(problem.trips)
    served_distance = 0.0
    served_energy = 0.0
    deadhead_distance = 0.0
    for duty in plan.duties:
        for leg in duty.legs:
            trip_info = problem.trip_by_id().get(str(leg.trip.trip_id))
            if trip_info is None:
                continue
            served_distance += float(getattr(trip_info, "distance_km", 0.0) or 0.0)
            served_energy += float(getattr(trip_info, "energy_kwh", 0.0) or 0.0)
            deadhead_distance += _canonical_deadhead_distance_km(problem, int(getattr(leg, "deadhead_from_prev_min", 0) or 0))
    total_charge_energy = sum(
        max(float(getattr(slot, "charge_kw", 0.0) or 0.0) - float(getattr(slot, "discharge_kw", 0.0) or 0.0), 0.0)
        * max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1)
        / 60.0
        for slot in plan.charging_slots
    )
    grid_import_total_kwh = sum(
        float(value or 0.0)
        for depot_map in (plan.grid_to_bus_kwh_by_depot_slot or {}).values()
        for value in (depot_map or {}).values()
    ) + sum(
        float(value or 0.0)
        for depot_map in (plan.grid_to_bess_kwh_by_depot_slot or {}).values()
        for value in (depot_map or {}).values()
    )
    timestep_h = max(int(getattr(problem.scenario, "timestep_min", 0) or 0), 1) / 60.0
    peak_grid_import_kw = 0.0
    for depot_id, depot_map in (plan.grid_to_bus_kwh_by_depot_slot or {}).items():
        bess_map = (plan.grid_to_bess_kwh_by_depot_slot or {}).get(depot_id, {})
        for slot_idx, value in (depot_map or {}).items():
            peak_grid_import_kw = max(
                peak_grid_import_kw,
                (float(value or 0.0) + float((bess_map or {}).get(slot_idx, 0.0) or 0.0)) / timestep_h,
            )
    soc_pct_values = [float(row.get("soc_pct_after", 0.0) or 0.0) for row in soc_rows]
    charger_usage: Dict[str, set[int]] = defaultdict(set)
    for slot in plan.charging_slots:
        charger_id = str(getattr(slot, "charger_id", "") or "")
        if charger_id:
            charger_usage[charger_id].add(int(getattr(slot, "slot_index", 0) or 0))
    slot_count = max(len(problem.price_slots), 1)
    utilization_values = [len(slot_indices) / slot_count for slot_indices in charger_usage.values()]
    pv_generated_total_kwh = sum(
        float(value or 0.0)
        for asset in (problem.depot_energy_assets or {}).values()
        for value in getattr(asset, "pv_generation_kwh_by_slot", ())
    )
    pv_self_consumption_kwh = sum(
        float(value or 0.0)
        for depot_map in (plan.pv_to_bus_kwh_by_depot_slot or {}).values()
        for value in (depot_map or {}).values()
    ) + sum(
        float(value or 0.0)
        for depot_map in (plan.pv_to_bess_kwh_by_depot_slot or {}).values()
        for value in (depot_map or {}).values()
    )
    breakdown = dict(engine_result.cost_breakdown or {})
    return {
        "scenario_id": scenario_id,
        "fleet_size": len(problem.vehicles),
        "served_trip_count": served_trip_count,
        "unserved_trip_count": len(plan.unserved_trip_ids),
        "served_trip_rate": float(served_trip_count / total_trip_count) if total_trip_count > 0 else 0.0,
        "total_distance_km": served_distance,
        "total_deadhead_km": deadhead_distance,
        "deadhead_ratio": float(deadhead_distance / served_distance) if served_distance > 0 else 0.0,
        "total_energy_consumption_kwh": served_energy,
        "total_charging_energy_kwh": total_charge_energy,
        "peak_grid_import_kw": peak_grid_import_kw,
        "peak_charge_kw": max((float(getattr(slot, "charge_kw", 0.0) or 0.0) for slot in plan.charging_slots), default=0.0),
        "pv_generation_total_kwh": pv_generated_total_kwh,
        "pv_self_consumption_kwh": pv_self_consumption_kwh,
        "pv_utilization_ratio": float(pv_self_consumption_kwh / pv_generated_total_kwh) if pv_generated_total_kwh > 0 else 0.0,
        "min_soc_pct": min(soc_pct_values) if soc_pct_values else 0.0,
        "average_soc_pct": (sum(soc_pct_values) / len(soc_pct_values)) if soc_pct_values else 0.0,
        "charger_utilization_avg": (sum(utilization_values) / len(utilization_values)) if utilization_values else 0.0,
        "charger_utilization_max": max(utilization_values) if utilization_values else 0.0,
        "total_cost_jpy": float(
            breakdown.get("total_cost")
            if breakdown.get("total_cost") is not None
            else engine_result.objective_value
            or 0.0
        ),
        "accounting_total_cost_jpy": float(
            breakdown.get("total_cost")
            if breakdown.get("total_cost") is not None
            else 0.0
        ),
        "objective_value_jpy": float(engine_result.objective_value or 0.0),
        "solver_objective_value": float(engine_result.objective_value or 0.0),
        "validated_operating_cost_jpy": (
            float(breakdown.get("total_cost") or 0.0)
            if bool(engine_result.feasible)
            and str(engine_result.solver_status or "").upper()
            in {"OPTIMAL", "FEASIBLE", "SOLVED_FEASIBLE", "PHASE2_ASSIGNMENT_FEASIBLE"}
            and (
                not bool((engine_result.solver_metadata or {}).get("research_run", False))
                or bool(
                    (engine_result.solver_metadata or {}).get(
                        "research_accounting_cost_eligible", False
                    )
                )
            )
            else None
        ),
        "objective_is_actual_cost": bool(breakdown.get("objective_is_actual_cost", False)),
        "supports_exact_milp": bool((engine_result.solver_metadata or {}).get("supports_exact_milp", False)),
        "electricity_cost_jpy": float(
            breakdown.get("electricity_cost", breakdown.get("electricity_cost_final", 0.0))
            or 0.0
        ),
        "fuel_cost_jpy": float(breakdown.get("fuel_cost", 0.0) or 0.0),
        "vehicle_usage_cost_jpy": float(breakdown.get("vehicle_usage_cost", breakdown.get("vehicle_usage_cost_jpy", 0.0)) or 0.0),
        "used_vehicle_day_count": int(breakdown.get("used_vehicle_day_count", 0) or 0),
        "vehicle_usage_cost_jpy_per_used_bus": float(breakdown.get("vehicle_usage_cost_jpy_per_used_bus", 0.0) or 0.0),
        "fuel_cost_final_jpy": float(breakdown.get("fuel_cost_final", breakdown.get("fuel_cost", 0.0)) or 0.0),
        "fuel_cost_final_source": str(breakdown.get("fuel_cost_final_source", "provisional_distance_based") or "provisional_distance_based"),
        "fuel_cost_provisional_jpy": float(breakdown.get("fuel_cost_provisional", breakdown.get("provisional_ice_drive_cost", 0.0)) or 0.0),
        "fuel_cost_refueled_jpy": float(breakdown.get("fuel_cost_refueled", breakdown.get("realized_ice_refuel_cost", 0.0)) or 0.0),
        "fuel_cost_provisional_leftover_jpy": float(breakdown.get("fuel_cost_provisional_leftover", breakdown.get("leftover_ice_provisional_cost", 0.0)) or 0.0),
        "pv_self_consumption_cost_jpy": float(breakdown.get("pv_self_consumption_cost_jpy", 0.0) or 0.0),
        "pv_marginal_charge_cost_yen_per_kwh": float(
            breakdown.get("pv_marginal_charge_cost_yen_per_kwh", 0.0)
            or getattr(problem, "metadata", {}).get("pv_marginal_charge_cost_yen_per_kwh", 0.0)
            or 0.0
        ),
        "propulsion_energy_cost_jpy": float(breakdown.get("energy_cost", 0.0) or 0.0),
        "electricity_cost_basis": str(
            breakdown.get("energy_cost_basis")
            or "realized_supply_plus_inventory_valuation"
        ),
        "energy_cash_purchase_cost_jpy": float(
            breakdown.get("energy_cash_purchase_cost_jpy", 0.0) or 0.0
        ),
        "energy_inventory_valuation_cost_jpy": float(
            breakdown.get("energy_inventory_valuation_cost_jpy", 0.0) or 0.0
        ),
        "ev_unreplenished_drive_energy_kwh": float(
            breakdown.get("ev_unreplenished_drive_energy_kwh", 0.0) or 0.0
        ),
        "ev_energy_inventory_balanced": bool(
            breakdown.get("ev_energy_inventory_balanced", False)
        ),
        "bev_terminal_soc_policy": str(
            (engine_result.solver_metadata or {}).get("bev_terminal_soc_policy")
            or getattr(problem, "metadata", {}).get("bev_terminal_soc_policy")
            or "minimum_only"
        ),
        "bev_terminal_soc_total_drawdown_kwh": float(
            (engine_result.solver_metadata or {}).get(
                "bev_terminal_soc_total_drawdown_kwh", 0.0
            )
            or 0.0
        ),
        "bev_terminal_soc_balance_satisfied": bool(
            (engine_result.solver_metadata or {}).get(
                "bev_terminal_soc_balance_satisfied", False
            )
        ),
        "research_accounting_cost_eligible": bool(
            (engine_result.solver_metadata or {}).get(
                "research_accounting_cost_eligible", False
            )
        ),
        "research_cost_optimality_eligible": bool(
            (engine_result.solver_metadata or {}).get(
                "research_cost_optimality_eligible", False
            )
        ),
        "electricity_cost_provisional_jpy": float(breakdown.get("provisional_ev_drive_cost", 0.0) or 0.0),
        "electricity_cost_charged_jpy": float(
            breakdown.get(
                "realized_ev_charge_cost",
                breakdown.get("electricity_cost", breakdown.get("electricity_cost_final", 0.0)),
            )
            or 0.0
        ),
        "grid_energy_provisional_kwh": float(sum(float(value or 0.0) for depot_map in (plan.grid_to_bus_kwh_by_depot_slot or {}).values() for value in (depot_map or {}).values())),
        "grid_energy_charged_kwh": grid_import_total_kwh,
        "pv_to_bus_kwh": float(breakdown.get("pv_to_bus_kwh", 0.0) or 0.0),
        "bess_to_bus_kwh": float(breakdown.get("bess_to_bus_kwh", 0.0) or 0.0),
        "pv_to_bess_kwh": float(breakdown.get("pv_to_bess_kwh", 0.0) or 0.0),
        "grid_to_bess_kwh": float(breakdown.get("grid_to_bess_kwh", 0.0) or 0.0),
        "contract_over_limit_kwh": float(breakdown.get("contract_over_limit_kwh", 0.0) or 0.0),
        "contract_overage_cost_jpy": float(breakdown.get("contract_overage_cost", 0.0) or 0.0),
        "demand_charge_cost_jpy": float(breakdown.get("demand_cost", 0.0) or 0.0),
        "co2_kg": float(breakdown.get("total_co2_kg", 0.0) or 0.0),
        "solver_runtime_sec": float((engine_result.solver_metadata or {}).get("solve_time_sec", 0.0) or 0.0),
        "solution_status": str(getattr(engine_result, "solver_status", "") or "").lower(),
    }


def _canonical_deadhead_ratio_by_band(timeline_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    accum: Dict[str, Dict[str, float]] = {}
    for row in timeline_rows:
        band_id = str(row.get("band_id") or row.get("event_route_band_id") or "").strip()
        if not band_id:
            continue
        item = accum.setdefault(
            band_id,
            {
                "service_events": 0.0,
                "deadhead_events": 0.0,
                "service_km": 0.0,
                "deadhead_km": 0.0,
                "service_min": 0.0,
                "deadhead_min": 0.0,
            },
        )
        distance_km = float(row.get("distance_km") or 0.0)
        duration_min = float(row.get("duration_min") or 0.0)
        if bool(row.get("is_service")) or str(row.get("state") or "") == "service":
            item["service_events"] += 1.0
            item["service_km"] += distance_km
            item["service_min"] += duration_min
        if bool(row.get("is_deadhead")) or str(row.get("state") or "") == "deadhead":
            item["deadhead_events"] += 1.0
            item["deadhead_km"] += distance_km
            item["deadhead_min"] += duration_min
    rows: List[Dict[str, Any]] = []
    for band_id, values in sorted(accum.items()):
        service_km = float(values.get("service_km") or 0.0)
        deadhead_km = float(values.get("deadhead_km") or 0.0)
        rows.append(
            {
                "band_id": band_id,
                "service_events": int(values.get("service_events") or 0),
                "deadhead_events": int(values.get("deadhead_events") or 0),
                "service_km": service_km,
                "deadhead_km": deadhead_km,
                "service_min": float(values.get("service_min") or 0.0),
                "deadhead_min": float(values.get("deadhead_min") or 0.0),
                "deadhead_service_km_ratio": deadhead_km / service_km if service_km > 0 else 0.0,
            }
        )
    return rows


def _persist_canonical_graph_exports(
    *,
    scenario: Dict[str, Any],
    problem,
    engine_result,
    scenario_id: str,
    output_dir: str,
) -> Dict[str, Any]:
    from bff.mappers.scenario_to_problemdata import _build_graph_export_context
    from src.optimization.accounting.export import (
        FUEL_TIMESERIES_FIELDNAMES,
        export_accounting_outputs,
    )
    from src.result_exporter import (
        _build_route_band_diagram_assets,
        _build_vehicle_operation_diagram_assets,
        _write_csv,
        _write_route_band_diagram_assets,
        _write_vehicle_operation_diagram_assets,
        _filter_timeline_rows_for_day,
    )

    trips = [
        dict(item)
        for item in list(scenario.get("trips") or scenario.get("timetable_rows") or [])
        if isinstance(item, dict)
    ]
    tasks = [SimpleNamespace(task_id=str(trip.get("trip_id") or "")) for trip in trips]
    graph_context = _build_graph_export_context(scenario, trips, tasks)
    base_date = _canonical_output_base_date(problem, graph_context)
    operator_id = str(
        scenario.get("operator_id")
        or scenario.get("operatorId")
        or (problem.metadata or {}).get("operator_id")
        or ""
    ).strip()
    if not operator_id:
        observed_operator_ids = sorted(
            {
                str(getattr(trip, "operator_id", "") or "").strip()
                for trip in tuple(getattr(getattr(problem, "dispatch_context", None), "trips", ()) or ())
                if str(getattr(trip, "operator_id", "") or "").strip()
            }
        )
        operator_id = observed_operator_ids[0] if len(observed_operator_ids) == 1 else "UNKNOWN_OPERATOR"
    timeline_rows = _canonical_vehicle_timeline_rows(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
        graph_context=graph_context,
    )
    for row in timeline_rows:
        row["operator_id"] = operator_id
    trip_assignment_rows = _canonical_trip_assignment_rows(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
        base_date=base_date,
        timeline_rows=timeline_rows,
    )
    for row in trip_assignment_rows:
        row["operator_id"] = operator_id
    movement_event_rows = _canonical_movement_event_rows(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
        base_date=base_date,
        operator_id=operator_id,
    )
    soc_rows = _canonical_soc_event_rows(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
        base_date=base_date,
    )
    depot_power_rows = _canonical_depot_power_rows_5min(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
        base_date=base_date,
        operator_id=operator_id,
    )
    cost_breakdown = _canonical_cost_breakdown_json(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
    )
    canonical_cost_ledger = _canonical_cost_ledger_json(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
    )
    kpi_summary = _canonical_kpi_summary_json(
        problem=problem,
        engine_result=engine_result,
        scenario_id=scenario_id,
        soc_rows=soc_rows,
    )
    source_flow_context = _canonical_energy_flow_context(problem, engine_result.plan)
    depot_source_provenance_exact = bool(
        source_flow_context.get("source_provenance_exact", False)
    )
    vehicle_source_provenance_exact = bool(
        depot_source_provenance_exact
        and dict(getattr(engine_result.plan, "metadata", {}) or {}).get(
            "vehicle_source_provenance_exact", False
        )
    )
    vehicle_source_allocation_method = (
        "solver_native"
        if vehicle_source_provenance_exact
        else "proportional_by_depot_timestep"
    )
    try:
        from src.optimization.accounting import build_accounting_artifacts
    except Exception:
        build_accounting_artifacts = None
    accounting_artifacts = None
    if build_accounting_artifacts is not None:
        vehicle_charging_source_rows = _research_vehicle_charging_source_timeseries_rows(
            problem=problem,
            engine_result=engine_result,
            base_date=base_date,
            operator_id=operator_id,
        )
        vehicle_soc_timeseries_rows = _research_vehicle_soc_timeseries_rows(
            problem=problem,
            engine_result=engine_result,
            base_date=base_date,
            timeline_rows=timeline_rows,
        )
        research_energy_exports = _research_rows_from_depot_power(
            depot_power_rows,
            base_date=base_date,
        )
        energy_flow_export_rows = list(research_energy_exports.get("energy_flow_timeseries.csv") or [])
        pv_generation_timeseries_total = sum(
            float(row.get("pv_generation_slot_kwh", row.get("pv_generation_kwh", 0.0)) or 0.0)
            for row in list(research_energy_exports.get("pv_generation_timeseries.csv") or [])
        )
        depot_energy_flows_pv_total = sum(
            float(row.get("pv_generation_slot_kwh", row.get("pv_generation_kwh", 0.0)) or 0.0)
            for row in energy_flow_export_rows
        )
        grid_co2_factor_by_slot = {
            int(getattr(slot, "slot_index", 0) or 0): float(getattr(slot, "co2_factor", 0.0) or 0.0)
            for slot in list(getattr(problem, "price_slots", ()) or ())
        }
        problem_metadata = dict(getattr(problem, "metadata", {}) or {})
        raw_initial_soc_policy = problem_metadata.get("weather_initial_soc_policy") or {}
        if isinstance(raw_initial_soc_policy, dict):
            initial_soc_policy = str(problem_metadata.get("initial_soc_policy") or raw_initial_soc_policy.get("mode") or "scenario_file")
        else:
            initial_soc_policy = str(problem_metadata.get("initial_soc_policy") or raw_initial_soc_policy or "scenario_file")
        available_vehicle_count = sum(
            1 for vehicle in list(getattr(problem, "vehicles", ()) or []) if bool(getattr(vehicle, "available", True))
        )
        evaluated_cost_breakdown = dict(engine_result.cost_breakdown or {})
        demand_rate = 0.0
        peak_grid_kw = float(
            evaluated_cost_breakdown.get("peak_grid_kw", 0.0) or 0.0
        )
        demand_cost = float(
            evaluated_cost_breakdown.get("demand_cost", 0.0) or 0.0
        )
        if peak_grid_kw > 0.0:
            demand_rate = demand_cost / peak_grid_kw
        scenario_cost_coeffs = dict(((scenario.get("scenario_overlay") or {}).get("cost_coefficients") or {}))
        generated_at = datetime.now(timezone.utc).isoformat()
        weather_reference_date = str(
            problem_metadata.get("weather_observation_date")
            or (
                (scenario.get("simulation_config") or {}).get(
                    "weather_observation_date"
                )
            )
            or ((scenario.get("simulation_config") or {}).get("weather_reference_date"))
            or ((scenario.get("simulation_config") or {}).get("service_date"))
            or base_date.isoformat()
        )[:10]
        weather_date = date.fromisoformat(weather_reference_date)
        accounting_artifacts = build_accounting_artifacts(
            problem=problem,
            scenario_id=scenario_id,
            run_id=str(Path(output_dir).name),
            service_date=base_date,
            weather_date=weather_date,
            operator_id=operator_id,
            trip_assignment_rows=trip_assignment_rows,
            movement_event_rows=movement_event_rows,
            vehicle_soc_timeseries_rows=vehicle_soc_timeseries_rows,
            vehicle_charging_source_rows=vehicle_charging_source_rows,
            energy_flow_rows=energy_flow_export_rows,
            metadata={
                "scenario_id": scenario_id,
                "run_id": str(Path(output_dir).name),
                "service_date": base_date.isoformat(),
                "weather_date": weather_date.isoformat(),
                "weather_reference_date": weather_reference_date,
                "weather_profile": str(((scenario.get("simulation_config") or {}).get("weather_profile") or "") or ""),
                "operation_mode": str(((scenario.get("simulation_config") or {}).get("operation_mode") or getattr(problem.scenario, "objective_mode", "")) or ""),
                "run_created_at": str((engine_result.solver_metadata or {}).get("started_at") or generated_at),
                "output_generated_at": generated_at,
                "pv_generation_timeseries_total_kwh": pv_generation_timeseries_total,
                "depot_energy_flows_pv_generation_total_kwh": depot_energy_flows_pv_total,
                "grid_co2_factor_by_slot": grid_co2_factor_by_slot,
                "initial_soc_policy": initial_soc_policy,
                "initial_soc_min_ratio": problem_metadata.get("initial_soc_min_ratio"),
                "initial_soc_max_ratio": problem_metadata.get("initial_soc_max_ratio"),
                "initial_soc_random_seed": problem_metadata.get("initial_soc_random_seed", (engine_result.solver_metadata or {}).get("random_seed", "")),
                "operator_id": operator_id,
                "slot_minutes": int(getattr(problem.scenario, "timestep_min", 30) or 30),
                "num_periods": len(getattr(problem, "price_slots", ()) or ()),
                "planning_horizon_hours": float(getattr(problem.scenario, "planning_horizon_hours", 0.0) or 0.0),
                "vehicle_usage_cost_jpy_per_used_bus": float((engine_result.cost_breakdown or {}).get("vehicle_usage_cost_jpy_per_used_bus", getattr(problem, "metadata", {}).get("vehicle_usage_cost_jpy_per_used_bus", 0.0)) or 0.0),
                "cost_component_flags": dict(getattr(problem, "metadata", {}).get("cost_component_flags", {}) or {}),
                "available_vehicle_count": available_vehicle_count,
                "objective_value": float(engine_result.objective_value or 0.0),
                "objective_is_actual_cost": bool(kpi_summary.get("objective_is_actual_cost", False)),
                "solver_objective_matches_accounting_total": bool(
                    (engine_result.solver_metadata or {}).get(
                        "solver_objective_matches_accounting_total", False
                    )
                ),
                "objective_semantics": str(
                    (engine_result.solver_metadata or {}).get("objective_semantics")
                    or "single_solver_objective"
                ),
                "supports_exact_milp": bool((engine_result.solver_metadata or {}).get("supports_exact_milp", False)),
                "fallback_applied": bool((engine_result.solver_metadata or {}).get("fallback_applied", False)),
                "charging_source_provenance_exact": bool(
                    vehicle_charging_source_rows
                    and all(bool(row.get("source_provenance_exact", False)) for row in vehicle_charging_source_rows)
                ),
                "vehicle_source_provenance_exact": vehicle_source_provenance_exact,
                "depot_source_provenance_exact": depot_source_provenance_exact,
                "vehicle_charging_source_allocation_method": str(
                    (
                        vehicle_charging_source_rows[0].get(
                            "vehicle_charging_source_allocation_method"
                        )
                        if vehicle_charging_source_rows
                        else vehicle_source_allocation_method
                    )
                    or vehicle_source_allocation_method
                ),
                "vehicle_charging_source_is_solver_native": bool(
                    vehicle_charging_source_rows
                    and all(bool(row.get("vehicle_charging_source_is_solver_native", False)) for row in vehicle_charging_source_rows)
                ),
                "fuel_price_jpy_per_liter": float(
                    scenario_cost_coeffs.get("diesel_price_per_l", scenario_cost_coeffs.get("fuel_price_yen_per_liter", 0.0)) or 0.0
                ),
                "canonical_fuel_cost_jpy": float(
                    evaluated_cost_breakdown.get("fuel_cost", 0.0) or 0.0
                ),
                "canonical_ice_co2_kg": float(
                    evaluated_cost_breakdown.get("ice_co2_kg", 0.0) or 0.0
                ),
                "canonical_total_co2_kg": float(
                    evaluated_cost_breakdown.get("total_co2_kg", 0.0) or 0.0
                ),
                "co2_price_jpy_per_kg": float(
                    scenario_cost_coeffs.get("co2_price_per_kg", scenario_cost_coeffs.get("carbon_price_jpy_per_kg", 0.0)) or 0.0
                ),
                "battery_degradation_price_jpy_per_kwh": float(
                    scenario_cost_coeffs.get("battery_degradation_cost_coeff_yen_per_kwh", 0.0) or 0.0
                ),
                "demand_rate_jpy_per_kw": demand_rate,
                "contract_power_kw": float(
                    max(
                        (
                            float(getattr(depot, "import_limit_kw", 0.0) or 0.0)
                            for depot in list(getattr(problem, "depots", ()) or [])
                        ),
                        default=0.0,
                    )
                ),
                "contract_power_exceeded": bool(kpi_summary.get("contract_limit_exceeded", False)),
                "contract_overage_kw": float(
                    kpi_summary.get("contract_over_limit_kwh", 0.0) or 0.0
                )
                / max(float(getattr(problem.scenario, "timestep_min", 5) or 5) / 60.0, 1.0e-9),
                "contract_power_mode": str(
                    kpi_summary.get("contract_overage_policy", "report_only") or "report_only"
                ),
                "solver_status": str(engine_result.solver_status or ""),
                "phase": (engine_result.solver_metadata or {}).get("phase"),
                "research_run": bool((engine_result.solver_metadata or {}).get("research_run", False)),
                "research_run_accepted": bool((engine_result.solver_metadata or {}).get("research_run_accepted", False)),
                "full_operational_validation": bool(
                    str((engine_result.solver_metadata or {}).get("phase") or "")
                    != "phase2_assignment_only"
                    and bool(engine_result.feasible)
                ),
                "validated_feasible": bool(engine_result.feasible),
                "requested_phase": (engine_result.solver_metadata or {}).get("requested_phase"),
                "resolved_phase": (engine_result.solver_metadata or {}).get("resolved_phase"),
                "executed_phase": (engine_result.solver_metadata or {}).get("executed_phase"),
                "mip_gap_requested_ratio": float((engine_result.solver_metadata or {}).get("requested_mip_gap", (engine_result.solver_metadata or {}).get("mip_gap", 0.0)) or 0.0),
                "mip_gap_requested_percent": float((engine_result.solver_metadata or {}).get("requested_mip_gap", (engine_result.solver_metadata or {}).get("mip_gap", 0.0)) or 0.0) * 100.0,
                "mip_gap_achieved_ratio": (engine_result.solver_metadata or {}).get("achieved_mip_gap", (engine_result.solver_metadata or {}).get("final_gap")),
"mip_gap_achieved_percent": (
                    None
                    if (engine_result.solver_metadata or {}).get("achieved_mip_gap", (engine_result.solver_metadata or {}).get("final_gap")) is None
                    else float((engine_result.solver_metadata or {}).get("achieved_mip_gap", (engine_result.solver_metadata or {}).get("final_gap")) or 0.0) * 100.0
                ),
                "stage1_mip_gap": (engine_result.solver_metadata or {}).get("stage1_mip_gap"),
                "stage2_mip_gap": (engine_result.solver_metadata or {}).get("stage2_mip_gap"),
                "supports_two_stage_milp": (engine_result.solver_metadata or {}).get("supports_two_stage_milp"),
                "supports_integrated_exact_milp": (engine_result.solver_metadata or {}).get("supports_integrated_exact_milp"),
                "optimization_structure": (engine_result.solver_metadata or {}).get("optimization_structure"),
            },
        )
        kpi_summary = dict(accounting_artifacts.summary)
    refuel_rows = [
        {
            "vehicle_id": str(getattr(slot, "vehicle_id", "") or ""),
            "slot_index": int(getattr(slot, "slot_index", 0) or 0),
            "time_hhmm": _canonical_slot_datetime(problem, base_date, int(getattr(slot, "slot_index", 0) or 0)).strftime("%H:%M"),
            "refuel_liters": float(getattr(slot, "refuel_liters", 0.0) or 0.0),
            "location_id": str(getattr(slot, "location_id", "") or ""),
        }
        for slot in engine_result.plan.refuel_slots
    ]
    research_energy_exports = _research_rows_from_depot_power(
        depot_power_rows,
        base_date=base_date,
    )
    research_extended_exports = _research_extended_timeseries_exports(
        problem=problem,
        engine_result=engine_result,
        depot_power_rows=depot_power_rows,
        timeline_rows=timeline_rows,
        refuel_rows=refuel_rows,
        base_date=base_date,
    )
    vehicle_charging_source_rows = _research_vehicle_charging_source_timeseries_rows(
        problem=problem,
        engine_result=engine_result,
        base_date=base_date,
        operator_id=operator_id,
    )
    vehicle_soc_timeseries_rows = _research_vehicle_soc_timeseries_rows(
        problem=problem,
        engine_result=engine_result,
        base_date=base_date,
        timeline_rows=timeline_rows,
    )
    fuel_summary_rows = _research_fuel_summary_rows(
        problem=problem,
        timeline_rows=timeline_rows,
        refuel_rows=refuel_rows,
    )
    graph_dir = Path(output_dir) / "graph"
    graph_dir.mkdir(parents=True, exist_ok=True)
    if accounting_artifacts is not None:
        try:
            accounting_paths = export_accounting_outputs(graph_dir, accounting_artifacts)
            kpi_summary = dict(accounting_artifacts.summary)
            research_extended_exports["co2_timeseries.csv"] = [dict(row) for row in accounting_artifacts.co2_timeseries]
            research_extended_exports["fuel_timeseries.csv"] = [dict(row) for row in accounting_artifacts.fuel_timeseries]
        except Exception:
            accounting_paths = {}
    else:
        accounting_paths = {}
    _write_csv(graph_dir / "vehicle_timeline.csv", timeline_rows)
    _write_csv(graph_dir / "soc_events.csv", soc_rows)
    _write_csv(graph_dir / "depot_power_timeseries.csv", depot_power_rows)
    for filename, rows in research_energy_exports.items():
        _write_csv(graph_dir / filename, rows)
    for filename, rows in research_extended_exports.items():
        if filename == "fuel_timeseries.csv":
            # An all-BEV solution has no fuel rows, but it is still a complete
            # canonical result. Preserve the schema so the final artifact gate
            # can distinguish a valid empty relation from a truncated file.
            _write_csv_rows(
                graph_dir / filename,
                rows,
                list(FUEL_TIMESERIES_FIELDNAMES),
            )
        else:
            _write_csv(graph_dir / filename, rows)
    _write_csv(graph_dir / "vehicle_charging_source_timeseries.csv", vehicle_charging_source_rows)
    _write_csv(graph_dir / "vehicle_soc_timeseries.csv", vehicle_soc_timeseries_rows)
    _write_csv_rows(
        graph_dir / "fuel_summary.csv",
        fuel_summary_rows,
        list(_FUEL_SUMMARY_FIELDS),
    )
    _write_csv(graph_dir / "trip_assignment.csv", trip_assignment_rows)
    # A zero-refuel day is valid evidence.  Preserve its schema rather than
    # emitting a zero-byte file, because the graph manifest declares this
    # artifact and the final artifact audit must be able to distinguish an
    # empty event set from a missing or truncated export.
    _write_csv_rows(
        graph_dir / "refuel_events.csv",
        refuel_rows,
        ["vehicle_id", "slot_index", "time_hhmm", "refuel_liters", "location_id"],
    )
    deadhead_ratio_rows = _canonical_deadhead_ratio_by_band(timeline_rows)
    _write_csv(graph_dir / "deadhead_ratio_by_band.csv", deadhead_ratio_rows)
    (graph_dir / "deadhead_ratio_by_band.json").write_text(
        json.dumps({"rows": deadhead_ratio_rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    accounting_summary = dict(getattr(accounting_artifacts, "summary", {}) or {})
    if accounting_summary:
        # Keep the three cost meanings explicit in every canonical graph
        # export.  Infeasible results retain diagnostic ledger totals, but the
        # validated operating cost is deliberately null.
        cost_breakdown["accounting_total_cost_jpy"] = accounting_summary.get(
            "accounting_total_cost_jpy", accounting_summary.get("total_cost_jpy")
        )
        cost_breakdown["solver_objective_value"] = accounting_summary.get(
            "solver_objective_value", accounting_summary.get("objective_value_jpy")
        )
        cost_breakdown["validated_operating_cost_jpy"] = accounting_summary.get(
            "validated_operating_cost_jpy"
        )
    kpi_summary.update(
        {
            "depot_source_provenance_exact": depot_source_provenance_exact,
            "vehicle_source_provenance_exact": vehicle_source_provenance_exact,
            "vehicle_source_allocation_method": vehicle_source_allocation_method,
            "charging_source_provenance_exact": bool(
                depot_source_provenance_exact and vehicle_source_provenance_exact
            ),
            "charging_source_provenance_scope": "site_and_vehicle",
        }
    )
    charging_source_provenance = {
        "schema_version": "charging_source_provenance_v1",
        "site_depot_timestep": {
            "exact": depot_source_provenance_exact,
            "scope": "depot_timestep",
            "note": source_flow_context.get("source_provenance_note"),
        },
        "vehicle_timestep": {
            "exact": vehicle_source_provenance_exact,
            "allocation_method": vehicle_source_allocation_method,
            "note": (
                "Vehicle source split is solver-native."
                if vehicle_source_provenance_exact
                else (
                    "Vehicle source split is inferred by proportional allocation "
                    "of exact depot/time-slot source totals."
                )
            ),
        },
    }
    (graph_dir / "charging_source_provenance.json").write_text(
        json.dumps(charging_source_provenance, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    calendar_weather_validation = dict(
        getattr(problem, "metadata", {}).get("service_calendar_validation") or {}
    )
    research_fleet_validation = dict(
        getattr(problem, "metadata", {}).get("research_fleet_validation") or {}
    )
    (graph_dir / "calendar_weather_validation.json").write_text(
        json.dumps(calendar_weather_validation, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (graph_dir / "research_fleet_validation.json").write_text(
        json.dumps(research_fleet_validation, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (graph_dir / "cost_breakdown.json").write_text(
        json.dumps(cost_breakdown, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (graph_dir / "canonical_cost_ledger.json").write_text(
        json.dumps(canonical_cost_ledger, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (graph_dir / "kpi_summary.json").write_text(
        json.dumps(kpi_summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Multi-day diagram support
    planning_days = int(problem.scenario.planning_days or 1)
    simulation_cfg = scenario.get("simulation_config") or {}
    solver_cfg = ((scenario.get("scenario_overlay") or {}).get("solver_config") or {})
    assets: Dict[str, Any] = {"entries": [], "svgs": {}}
    vehicle_operation_assets: Dict[str, Any] = {"entries": [], "svgs": {}}
    if planning_days > 1:
        all_vehicle_operation_assets: Dict[str, Any] = {"entries": [], "svgs": {}}
        all_route_band_assets: Dict[str, Any] = {"entries": [], "svgs": {}}
        timestep_min = int(problem.scenario.timestep_min or 30)
        for day_idx in range(planning_days):
            day_rows = _filter_timeline_rows_for_day(timeline_rows, day_idx, timestep_min)
            day_vehicle_operation_assets = _build_vehicle_operation_diagram_assets(
                day_rows,
                f"{scenario_id}_d{day_idx}",
            )
            day_route_band_assets = _build_route_band_diagram_assets(
                day_rows,
                f"{scenario_id}_d{day_idx}",
                graph_context=graph_context,
                trip_assignment_rows=trip_assignment_rows,
            )
            for entry in day_vehicle_operation_assets.get("entries", []):
                entry["day_index"] = day_idx
                entry["diagram_file"] = f"day_{day_idx}/{entry.get('diagram_file', '')}"
                all_vehicle_operation_assets["entries"].append(entry)
            for svg_key, svg_content in (day_vehicle_operation_assets.get("svg_payloads") or day_vehicle_operation_assets.get("svgs") or {}).items():
                all_vehicle_operation_assets["svgs"][f"day_{day_idx}/{svg_key}"] = svg_content
            for entry in day_route_band_assets.get("entries", []):
                entry["day_index"] = day_idx
                entry["diagram_file"] = f"day_{day_idx}/{entry.get('diagram_file', '')}"
                all_route_band_assets["entries"].append(entry)
            for svg_key, svg_content in (day_route_band_assets.get("svg_payloads") or day_route_band_assets.get("svgs") or {}).items():
                all_route_band_assets["svgs"][f"day_{day_idx}/{svg_key}"] = svg_content
        vehicle_operation_assets = all_vehicle_operation_assets
        assets = all_route_band_assets
    else:
        vehicle_operation_assets = _build_vehicle_operation_diagram_assets(
            timeline_rows,
            scenario_id,
        )
        assets = _build_route_band_diagram_assets(
            timeline_rows,
            scenario_id,
            graph_context=graph_context,
            trip_assignment_rows=trip_assignment_rows,
        )
    route_band_dir = graph_dir / "route_band_diagrams"
    if route_band_dir.exists():
        shutil.rmtree(route_band_dir)
    _write_route_band_diagram_assets(graph_dir, assets, planning_days=planning_days)
    _write_vehicle_operation_diagram_assets(
        graph_dir,
        vehicle_operation_assets,
        planning_days=planning_days,
    )

    graph_manifest = {
        "schema_version": "canonical_graph_v1",
        "scenario_id": scenario_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "time_resolution_minutes": int(getattr(problem.scenario, "timestep_min", 30) or 30),
        "timezone": "Asia/Tokyo",
        "source": "canonical_assignment_plan",
        "bess_soc_kwh_semantics": "slot_end",
        "files": [
            "vehicle_timeline.csv",
            "soc_events.csv",
            "depot_power_timeseries.csv",
            "grid_import_timeseries.csv",
            "pv_generation_timeseries.csv",
            "energy_flow_timeseries.csv",
            "bus_charging_total_timeseries.csv",
            "co2_timeseries.csv",
            "cost_timeseries.csv",
            "contract_limit_timeseries.csv",
            "bess_timeseries.csv",
            "charging_source_provenance.json",
            "calendar_weather_validation.json",
            "research_fleet_validation.json",
            "vehicle_charging_source_timeseries.csv",
            "fuel_timeseries.csv",
            "vehicle_soc_timeseries.csv",
            "vehicle_slot_ledger.csv",
            "vehicle_slot_ledger.json",
            "movement_event_ledger.csv",
            "movement_event_ledger.json",
            "vehicle_energy_ledger.csv",
            "vehicle_energy_ledger.json",
            "energy_flow_ledger.csv",
            "energy_flow_ledger.json",
            "fuel_canonical_ledger.csv",
            "initial_soc_ledger.csv",
            "initial_soc_precheck.csv",
            "data_flow_validation.csv",
            "fuel_summary.csv",
            "trip_assignment.csv",
            "refuel_events.csv",
            "deadhead_ratio_by_band.csv",
            "deadhead_ratio_by_band.json",
            "cost_breakdown.json",
            "canonical_cost_ledger.json",
            "kpi_summary.json",
        ],
        "optional_exports": {
            "route_band_diagrams": {
                "enabled": bool(assets.get("entries")),
                "grouping_key": "band_id",
                "diagram_format": "svg",
                "manifest_file": "route_band_diagrams/manifest.json",
                "diagram_count": len(list(assets.get("entries") or [])),
            },
            "vehicle_operation_diagrams": {
                "enabled": bool(vehicle_operation_assets.get("entries")),
                "grouping_key": "vehicle_id",
                "diagram_format": "svg",
                "manifest_file": (
                    "vehicle_operation_diagrams/manifest.json"
                    if vehicle_operation_assets.get("entries")
                    else ""
                ),
                "diagram_count": len(list(vehicle_operation_assets.get("entries") or [])),
            }
        },
    }
    (graph_dir / "manifest.json").write_text(
        json.dumps(graph_manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    manifest_relpath = None
    manifest_relpath = "graph/route_band_diagrams/manifest.json"
    return {
        "enabled": bool(assets.get("entries")),
        "diagram_count": len(list(assets.get("entries") or [])),
        "manifest_path": manifest_relpath,
        "vehicle_operation_diagram_manifest_path": (
            "graph/vehicle_operation_diagrams/manifest.json"
            if vehicle_operation_assets.get("entries")
            else None
        ),
        "vehicle_timeline_path": "graph/vehicle_timeline.csv",
        "graph_manifest_path": "graph/manifest.json",
        "trip_assignment_path": "graph/trip_assignment.csv",
        "deadhead_ratio_by_band_path": "graph/deadhead_ratio_by_band.csv",
        "soc_events_path": "graph/soc_events.csv",
        "depot_power_timeseries_path": "graph/depot_power_timeseries.csv",
        "grid_import_timeseries_path": "graph/grid_import_timeseries.csv",
        "pv_generation_timeseries_path": "graph/pv_generation_timeseries.csv",
        "energy_flow_timeseries_path": "graph/energy_flow_timeseries.csv",
        "bus_charging_total_timeseries_path": "graph/bus_charging_total_timeseries.csv",
        "co2_timeseries_path": "graph/co2_timeseries.csv",
        "cost_timeseries_path": "graph/cost_timeseries.csv",
        "contract_limit_timeseries_path": "graph/contract_limit_timeseries.csv",
        "bess_timeseries_path": "graph/bess_timeseries.csv",
        "charging_source_provenance_path": "graph/charging_source_provenance.json",
        "vehicle_charging_source_timeseries_path": "graph/vehicle_charging_source_timeseries.csv",
        "fuel_timeseries_path": "graph/fuel_timeseries.csv",
        "vehicle_soc_timeseries_path": "graph/vehicle_soc_timeseries.csv",
        "fuel_summary_path": "graph/fuel_summary.csv",
        "cost_breakdown_path": "graph/cost_breakdown.json",
        "canonical_cost_ledger_path": "graph/canonical_cost_ledger.json",
        "kpi_summary_path": "graph/kpi_summary.json",
        "vehicle_slot_ledger_path": accounting_paths.get("vehicle_slot_ledger_csv", "graph/vehicle_slot_ledger.csv"),
        "movement_event_ledger_path": accounting_paths.get(
            "movement_event_ledger_csv", "graph/movement_event_ledger.csv"
        ),
        "vehicle_energy_ledger_path": accounting_paths.get("vehicle_energy_ledger_csv", "graph/vehicle_energy_ledger.csv"),
        "energy_flow_ledger_path": accounting_paths.get("energy_flow_ledger_csv", "graph/energy_flow_ledger.csv"),
        "fuel_canonical_ledger_path": accounting_paths.get("fuel_canonical_ledger_csv", "graph/fuel_canonical_ledger.csv"),
        "initial_soc_ledger_path": accounting_paths.get("initial_soc_ledger_csv", "graph/initial_soc_ledger.csv"),
        "initial_soc_precheck_path": accounting_paths.get("initial_soc_precheck_csv", "graph/initial_soc_precheck.csv"),
        "data_flow_validation_path": accounting_paths.get("data_flow_validation_csv", "graph/data_flow_validation.csv"),
        "calendar_weather_validation_path": "graph/calendar_weather_validation.json",
        "research_fleet_validation_path": "graph/research_fleet_validation.json",
        "accounting_summary": getattr(accounting_artifacts, "summary", {}),
        "reporting_finalizer": {
            "status": "deferred",
            "reason": "requires top-level rich run outputs before canonical reporting rebuild",
        },
        "refuel_events_path": "graph/refuel_events.csv",
        "planning_days": planning_days,
    }


def _solution_validity_payload(
    *,
    solver_status: Any,
    feasible: Any,
    trip_count_unserved: Any,
    infeasibility_reasons: List[Any],
    solver_metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    status = str(solver_status or "").strip()
    status_upper = status.upper()
    reasons = [str(item) for item in list(infeasibility_reasons or []) if str(item).strip()]
    # Physical schedule validity and research acceptance are deliberately
    # separate.  A fleet-contract or exactness rejection must not turn a
    # physically valid schedule into an infeasible one.
    blocking_reasons: List[str] = []
    research_blocking_reasons: List[str] = []
    meta = dict(solver_metadata or {})

    is_fallback_status = (
        "FALLBACK" in status_upper
        or "BASELINE" in status_upper
        or "UNAVAILABLE" in status_upper
        or "GUARDRAIL" in status_upper
    )
    if is_fallback_status:
        blocking_reasons.append("baseline_fallback")
    if status_upper in {"REPAIRED_HEURISTIC", "POSTSOLVE_REPAIRED"}:
        blocking_reasons.append("repaired_heuristic")
    if status_upper == "DEBUG_RESULT" or meta.get("debug_mode") or meta.get("result_class") == "debug_result":
        blocking_reasons.append("debug_result")
    if meta.get("result_class") == "assignment_only_result" or meta.get("phase") == "phase2_assignment_only":
        blocking_reasons.append("assignment_only_result")
    if status_upper == "PARTIAL_BASELINE_FALLBACK":
        blocking_reasons.append("partial_baseline_fallback")
    if status_upper == "TRUTHFUL_BASELINE_GUARDRAIL":
        blocking_reasons.append("truthful_baseline_guardrail")
    if status_upper == "POSTSOLVE_REPAIRED":
        blocking_reasons.append("postsolve_repaired")
    if meta.get("postsolve_soc_repair_applied"):
        blocking_reasons.append("repaired_heuristic")
    if meta.get("postsolve_charging_recomputed"):
        blocking_reasons.append("repaired_heuristic")
    if meta.get("postsolve_modified_solution"):
        blocking_reasons.append("repaired_heuristic")
    terminal_policy = str(
        meta.get("bev_terminal_soc_policy") or ""
    ).strip().lower()
    if (
        terminal_policy == "return_to_initial"
        and not bool(meta.get("bev_terminal_soc_balance_satisfied", False))
    ):
        blocking_reasons.append("bev_terminal_soc_balance_failed")
    bess_terminal_deviation_kwh = abs(
        float(meta.get("bess_terminal_soc_deviation_kwh", 0.0) or 0.0)
    )
    bess_terminal_tolerance_kwh = max(
        float(meta.get("bess_terminal_soc_tolerance_kwh", 1.0e-6) or 1.0e-6),
        0.0,
    )
    if bess_terminal_deviation_kwh > bess_terminal_tolerance_kwh:
        blocking_reasons.append("bess_terminal_soc_balance_failed")
    if bool(meta.get("research_run", False)) and not bool(
        meta.get("research_run_accepted", False)
    ):
        research_blocking_reasons.append("research_acceptance_failed")
    if not bool(feasible):
        blocking_reasons.append("postsolve_infeasible")
    if reasons:
        blocking_reasons.append("infeasibility_reasons_present")
    try:
        unserved = int(trip_count_unserved or 0)
    except (TypeError, ValueError):
        unserved = 0
    if unserved > 0:
        blocking_reasons.append("unserved_trips_present")
    blocking_reasons = sorted(set(blocking_reasons))
    research_blocking_reasons = sorted(set(research_blocking_reasons))
    supports_exact = bool(meta.get("supports_exact_milp", False))
    fallback_applied = bool(meta.get("fallback_applied", False) or meta.get("fallback_reason"))
    ordinary_feasible_status = status_upper in {
        "SOLVED_FEASIBLE",
        "OPTIMAL",
        "FEASIBLE",
    }
    incumbent_limit_status = status_upper in {
        "TIME_LIMIT",
        "OBJECTIVE_LIMIT",
        "SOLUTION_LIMIT",
    }
    has_validated_incumbent_status = bool(
        ordinary_feasible_status
        or (
            incumbent_limit_status
            and bool(meta.get("has_feasible_incumbent", False))
        )
    )
    result_class: str
    status_reason: str
    if not blocking_reasons and has_validated_incumbent_status:
        if incumbent_limit_status:
            # A time/objective/solution-limit incumbent can be physically
            # valid without being an optimality result.  Keep those claims
            # separate instead of mislabelling an empty-error result as
            # infeasible.
            status_reason = "validated_feasible_limit_incumbent"
            result_class = "validated_non_exact"
        elif supports_exact and not fallback_applied:
            status_reason = "validated_feasible_no_cancellation" if unserved == 0 else "validated_feasible"
            result_class = "exact_or_validated"
        else:
            status_reason = "validated_feasible_non_exact"
            result_class = "validated_non_exact"
    elif "partial_baseline_fallback" in blocking_reasons:
        status_reason = "partial_baseline_fallback"
        result_class = "baseline_fallback"
    elif "truthful_baseline_guardrail" in blocking_reasons:
        status_reason = "truthful_baseline_guardrail"
        result_class = "baseline_fallback"
    elif "debug_result" in blocking_reasons:
        status_reason = "debug_result_not_research_kpi"
        result_class = "debug_result"
    elif "assignment_only_result" in blocking_reasons:
        status_reason = "assignment_only_no_charging_soc_validation"
        result_class = "assignment_only_result"
    elif (
        "bev_terminal_soc_balance_failed" in blocking_reasons
        or "bess_terminal_soc_balance_failed" in blocking_reasons
    ):
        status_reason = "terminal_soc_balance_failed"
        result_class = "postsolve_infeasible"
    elif "repaired_heuristic" in blocking_reasons or "postsolve_repaired" in blocking_reasons:
        status_reason = "repaired_heuristic"
        result_class = "repaired_heuristic"
    elif "baseline_fallback" in blocking_reasons:
        status_reason = "baseline_fallback_or_postsolve_infeasible"
        result_class = "baseline_fallback"
    elif "postsolve_infeasible" in blocking_reasons:
        status_reason = "postsolve_infeasible"
        result_class = "postsolve_infeasible"
    elif "unserved_trips_present" in blocking_reasons:
        status_reason = "unserved_trips_present"
        result_class = "postsolve_infeasible"
    else:
        status_reason = "infeasibility_reasons_present"
        result_class = "postsolve_infeasible"
    validated_feasible = bool(
        not blocking_reasons and has_validated_incumbent_status
    )
    validated_no_cancellation = bool(validated_feasible and unserved == 0)
    research_assignment_eligible = bool(
        str(meta.get("phase") or "") == "phase2_assignment_only"
        and bool(meta.get("research_run", False))
        and bool(meta.get("research_run_accepted", False))
        and bool(meta.get("research_feasibility_eligible", False))
    )
    return {
        "validated_no_cancellation": validated_no_cancellation,
        "validated_feasible": bool(validated_feasible),
        "status_reason": status_reason,
        "result_class": result_class,
        "blocking_reasons": blocking_reasons,
        "physical_blocking_reasons": blocking_reasons,
        "physical_validation_status": (
            "VALID" if validated_feasible else "INVALID"
        ),
        "research_acceptance_status": (
            "ACCEPTED"
            if bool(meta.get("research_run_accepted", False))
            else "REJECTED"
            if bool(meta.get("research_run", False))
            else "NOT_REQUESTED"
        ),
        "research_blocking_reasons": research_blocking_reasons,
        "research_acceptance_failed_checks": sorted(
            str(name)
            for name, passed in dict(
                meta.get("research_acceptance_checks") or {}
            ).items()
            if passed is not True
        ),
        "research_kpi_eligible": bool(
            validated_feasible
            and result_class == "exact_or_validated"
            and bool(meta.get("research_run", False))
            and bool(meta.get("research_run_accepted", False))
            and bool(meta.get("research_cost_kpi_eligible", False))
        ),
        "research_feasibility_eligible": bool(
            validated_feasible
            and bool(meta.get("research_run", False))
            and bool(meta.get("research_run_accepted", False))
            and bool(meta.get("research_feasibility_eligible", False))
        ),
        # Phase 2 validates vehicle-trip assignment only.  Keep it distinct
        # from full operational/SOC feasibility so callers cannot infer that
        # charging feasibility was evaluated.
        "research_assignment_eligible": research_assignment_eligible,
        "research_cost_kpi_eligible": bool(
            meta.get("research_cost_kpi_eligible", False)
        ),
        "validation_metrics": dict(meta.get("validation_metrics") or {}),
        "research_acceptance_checks": dict(meta.get("research_acceptance_checks") or {}),
        "terminal_soc_validation": {
            "bev_terminal_soc_policy": terminal_policy or None,
            "bev_terminal_soc_balance_satisfied": meta.get(
                "bev_terminal_soc_balance_satisfied"
            ),
            "bess_terminal_soc_deviation_kwh": bess_terminal_deviation_kwh,
            "bess_terminal_soc_tolerance_kwh": bess_terminal_tolerance_kwh,
        },
    }


_INVALID_RESULT_METRIC_KEYS = {
    "total_cost",
    "total_cost_jpy",
    "objective_value",
    "objective_value_jpy",
    "accounting_total_cost_jpy",
    "gross_operating_cost_jpy",
    "reported_total_cost_jpy",
    "solver_objective_value",
    "validated_operating_cost_jpy",
    "energy_cost",
    "energy_cost_jpy",
    "electricity_cost",
    "electricity_cost_jpy",
    "electricity_cost_final",
    "demand_charge",
    "demand_charge_cost_jpy",
    "demand_cost",
    "demand_cost_jpy",
    "fuel_cost",
    "fuel_cost_jpy",
    "fuel_cost_final",
    "fuel_cost_final_jpy",
    "fuel_cost_provisional",
    "fuel_cost_provisional_jpy",
    "fuel_cost_refueled",
    "fuel_cost_refueled_jpy",
    "fuel_cost_provisional_leftover",
    "fuel_cost_provisional_leftover_jpy",
    "co2_cost",
    "co2_cost_jpy",
    "battery_degradation_cost",
    "battery_degradation_cost_jpy",
    "degradation_cost",
    "deviation_cost",
    "vehicle_cost",
    "vehicle_cost_jpy",
    "vehicle_usage_cost",
    "vehicle_usage_cost_jpy",
    "driver_cost",
    "driver_cost_jpy",
    "penalty_unserved",
    "unserved_penalty",
    "switch_cost",
    "return_leg_bonus",
    "weather_strategy_objective_term_jpy_equivalent",
    "contract_overage_cost",
    "contract_overage_cost_jpy",
    "propulsion_energy_cost_jpy",
    "pv_self_consumption_cost_jpy",
    "grid_to_bus_kwh",
    "grid_to_bess_kwh",
    "grid_import_kwh",
    "grid_import_total_kwh",
    "grid_import_for_contract_kwh",
    "bus_charge_from_grid_kwh",
    "bus_charge_from_bess_kwh",
    "total_bus_charge_kwh",
    "total_bess_charge_kwh",
    "total_charge_input_kwh",
    "grid_total_kwh",
    "pv_to_bus_kwh",
    "pv_to_bess_kwh",
    "pv_curtail_kwh",
    "pv_curtailed_kwh",
    "pv_utilization_ratio",
    "pv_utilization_rate",
    "bess_to_bus_kwh",
    "bess_charge_kwh",
    "bess_discharge_kwh",
    "peak_grid_kw",
    "peak_grid_import_kw",
    "peak_grid_import_kw_all_depots",
    "peak_grid_import_kw_any_depot",
    "peak_total_charge_kw",
    "peak_total_charge_kw_all_depots",
    "peak_total_charge_kw_any_depot",
    "contract_over_limit_kwh",
    "contract_over_limit_kw_peak",
    "contract_over_limit_slot_count",
    "contract_limit_exceeded",
    "grid_purchase_cost_jpy",
    "bess_discharge_cost_jpy",
    "total_co2_kg",
}


def _canonical_trip_counts(
    canonical_solver_result: Optional[Dict[str, Any]],
    *,
    fallback_served: int,
    fallback_unserved: int,
) -> tuple[int, int]:
    canonical = dict(canonical_solver_result or {})

    def _count(explicit_key: str, ids_key: str, fallback: int) -> int:
        explicit = canonical.get(explicit_key)
        if explicit not in (None, ""):
            try:
                return max(int(explicit), 0)
            except (TypeError, ValueError):
                pass
        if ids_key in canonical:
            return len(list(canonical.get(ids_key) or []))
        return max(int(fallback), 0)

    return (
        _count("trip_count_served", "served_trip_ids", fallback_served),
        _count("trip_count_unserved", "unserved_trip_ids", fallback_unserved),
    )


def _invalid_result_failure_stage(optimization_result: Dict[str, Any]) -> str:
    settings = dict(optimization_result.get("solver_settings") or {})
    stage2_status = str(settings.get("stage2_solver_status") or "").lower()
    if stage2_status and stage2_status not in {"optimal", "feasible", "solved_feasible"}:
        return "stage2_energy_dispatch"
    stage1_status = str(settings.get("stage1_solver_status") or "").lower()
    if stage1_status in {"infeasible", "inf_or_unbd", "unbounded", "no_valid_incumbent"}:
        return "stage1_assignment"
    return "postsolve_validation"


def _invalidate_mapping_metrics(payload: Dict[str, Any]) -> None:
    for key in _INVALID_RESULT_METRIC_KEYS:
        if key in payload:
            payload[key] = None
    payload["objective_is_actual_cost"] = False
    payload["solver_objective_matches_accounting_total"] = False


def _apply_invalid_result_kpi_gate(
    optimization_result: Dict[str, Any],
    canonical_solver_result: Optional[Dict[str, Any]],
) -> None:
    validity = dict(optimization_result.get("solution_validity") or {})
    if bool(validity.get("validated_feasible", False)):
        return

    solver_status = str(optimization_result.get("solver_status") or "")
    result_status = "INFEASIBLE" if "INFEASIBLE" in solver_status.upper() else "INVALID"
    failure_stage = _invalid_result_failure_stage(optimization_result)
    summary = dict(optimization_result.get("summary") or {})
    served, unserved = _canonical_trip_counts(
        canonical_solver_result,
        fallback_served=int(summary.get("trip_count_served") or 0),
        fallback_unserved=int(summary.get("trip_count_unserved") or 0),
    )

    optimization_result.update(
        {
            "result_status": result_status,
            "failure_stage": failure_stage,
            "research_kpi_eligible": False,
            "objective_value": None,
            "kpi_eligibility_reason": str(
                validity.get("status_reason") or "canonical_result_not_validated_feasible"
            ),
        }
    )
    summary.update(
        {
            "trip_count_served": served,
            "trip_count_unserved": unserved,
            "coverage_rank_primary": unserved,
            "result_status": result_status,
            "failure_stage": failure_stage,
            "research_kpi_eligible": False,
        }
    )
    optimization_result["summary"] = summary

    cost_breakdown = dict(optimization_result.get("cost_breakdown") or {})
    _invalidate_mapping_metrics(cost_breakdown)
    cost_breakdown["evaluation_feasible"] = 0.0
    optimization_result["cost_breakdown"] = cost_breakdown

    graph_artifacts = dict(optimization_result.get("graph_artifacts") or {})
    accounting_summary = dict(graph_artifacts.get("accounting_summary") or {})
    _invalidate_mapping_metrics(accounting_summary)
    accounting_summary.update(
        {
            "served_trip_count": served,
            "unserved_trip_count": unserved,
            "research_kpi_eligible": False,
            "result_status": result_status,
            "failure_stage": failure_stage,
        }
    )
    graph_artifacts["accounting_summary"] = accounting_summary
    optimization_result["graph_artifacts"] = graph_artifacts

    charging_summary = optimization_result.get("charging_summary")
    if isinstance(charging_summary, dict):
        charging_summary.update(
            {
                "result_status": result_status,
                "failure_stage": failure_stage,
                "research_kpi_eligible": False,
            }
        )
        totals = dict(charging_summary.get("totals") or {})
        _invalidate_mapping_metrics(totals)
        charging_summary["totals"] = totals
        gated_depots = []
        for depot in list(charging_summary.get("depots") or []):
            if not isinstance(depot, dict):
                continue
            gated_depot = dict(depot)
            _invalidate_mapping_metrics(gated_depot)
            gated_depots.append(gated_depot)
        charging_summary["depots"] = gated_depots


def _feasible_candidate_job_message(
    classification: Dict[str, Any],
) -> str:
    """Describe passed and unresolved gates without conflating gap and scope."""

    blockers = set(classification.get("optimality_blocking_reasons") or [])
    gap_target_met = classification.get("mip_gap_target_met") is True
    passed_gates = ["physical checks"]
    if gap_target_met:
        gap_gate = (
            "the certified Stage 1 MIP gap target"
            if classification.get("certified_mip_gap") is not None
            else "the requested MIP gap target"
        )
        passed_gates.append(gap_gate)

    unresolved_claims: List[str] = []
    if "requested_mip_gap_not_met" in blockers:
        unresolved_claims.append("the requested MIP gap")
    if "not_an_integrated_global_assignment_and_charging_milp" in blockers:
        unresolved_claims.append("integrated global optimality")
    if not unresolved_claims:
        unresolved_claims.append("the remaining optimality claim")

    if len(unresolved_claims) == 1:
        unresolved_text = unresolved_claims[0]
        verb = "is"
    else:
        unresolved_text = (
            ", ".join(unresolved_claims[:-1])
            + f" and {unresolved_claims[-1]}"
        )
        verb = "are"
    return (
        f"Feasible candidate complete; {' and '.join(passed_gates)} passed, "
        f"but {unresolved_text} {verb} not established."
    )


def _apply_result_claim_classification(
    optimization_result: Dict[str, Any],
) -> Dict[str, Any]:
    """Label a feasible incumbent without overstating global optimality."""

    validity = dict(optimization_result.get("solution_validity") or {})
    settings = dict(optimization_result.get("solver_settings") or {})
    metadata = dict(optimization_result.get("solver_metadata") or {})
    physically_feasible = bool(validity.get("validated_feasible", False))
    optimality_blockers: List[str] = []
    if physically_feasible:
        if settings.get("mip_gap_target_met") is not True:
            optimality_blockers.append("requested_mip_gap_not_met")
        if not bool(metadata.get("supports_integrated_exact_milp", False)):
            optimality_blockers.append(
                "not_an_integrated_global_assignment_and_charging_milp"
            )
    optimality_claim_eligible = bool(
        physically_feasible and not optimality_blockers
    )
    if physically_feasible and not optimality_claim_eligible:
        label = "feasible_candidate"
        display_name = "Feasible candidate"
        optimization_result["result_status"] = "FEASIBLE_CANDIDATE"
    elif physically_feasible:
        label = "validated_optimality_claim_candidate"
        display_name = "Validated result"
    else:
        label = "invalid_or_infeasible_result"
        display_name = "Invalid or infeasible result"
    mip_gap_target_met = settings.get("mip_gap_target_met") is True
    certified_mip_gap = settings.get(
        "certified_mip_gap_ratio",
        settings.get("stage1_certified_mip_gap_ratio"),
    )
    if label == "feasible_candidate" and mip_gap_target_met:
        interpretation = (
            (
                "A physically feasible incumbent meeting the certified Stage 1 "
                "MIP gap target; remaining optimality blockers are listed "
                "separately and still forbid an integrated global-optimum claim."
                if settings.get("certified_mip_gap_ratio") is None
                and settings.get("stage1_certified_mip_gap_ratio") is not None
                else "A physically feasible incumbent meeting a separately "
                "certified integrated MIP gap target; remaining optimality "
                "blockers are listed separately."
            )
            if certified_mip_gap is not None
            else "A physically feasible incumbent meeting the requested MIP "
            "gap target; remaining optimality blockers are listed separately."
        )
    elif label == "feasible_candidate":
        interpretation = (
            "A physically feasible incumbent; do not describe it as a global "
            "optimum or as meeting the requested MIP gap."
        )
    else:
        interpretation = (
            "See physical and research acceptance artifacts for scope."
        )
    payload = {
        "schema_version": "result_claim_classification_v1",
        "label": label,
        "display_name": display_name,
        "physical_feasibility_claim_eligible": physically_feasible,
        "optimality_claim_eligible": optimality_claim_eligible,
        "optimality_blocking_reasons": optimality_blockers,
        "requested_mip_gap": settings.get("mip_gap_requested_ratio"),
        "mip_gap_target_met": mip_gap_target_met,
        "certified_mip_gap": certified_mip_gap,
        "gurobi_raw_mip_gap": settings.get(
            "stage1_gurobi_raw_mip_gap_ratio"
        ),
        "solver_objective_matches_accounting_total": bool(
            metadata.get("solver_objective_matches_accounting_total", False)
        ),
        "used_powertrain_composition_search_accepted": bool(
            metadata.get(
                "stage1_used_powertrain_composition_search_accepted", False
            )
        ),
        "interpretation": interpretation,
    }
    optimization_result["result_claim_classification"] = payload
    summary = optimization_result.get("summary")
    if isinstance(summary, dict):
        summary["result_claim_classification"] = payload
        summary["result_status"] = optimization_result.get("result_status")
    return payload


def _solver_settings_payload(
    *,
    time_limit_seconds_requested: Any,
    mip_gap_requested: Any,
    solver_metadata: Dict[str, Any],
    random_seed_requested: Any = None,
    stage1_time_limit_seconds_requested: Any = None,
    stage2_time_limit_seconds_requested: Any = None,
) -> Dict[str, Any]:
    def _float_or_none(value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _int_or_none(value: Any) -> Optional[int]:
        parsed = _float_or_none(value)
        return None if parsed is None else int(parsed)

    metadata = dict(solver_metadata or {})
    effective_limits = dict(metadata.get("effective_limits") or {})
    phase4_seed_audit = dict(
        metadata.get("phase4_phase3_seed_audit") or {}
    )
    integrated_start_audit = dict(
        metadata.get("integrated_warm_start_audit") or {}
    )
    random_seed = _int_or_none(metadata.get("random_seed"))
    if random_seed is None:
        random_seed = _int_or_none(random_seed_requested)
    requested_gap = _float_or_none(mip_gap_requested)
    has_feasible_incumbent = bool(metadata.get("has_feasible_incumbent", False))
    achieved_gap = (
        _float_or_none(metadata.get("achieved_mip_gap", metadata.get("final_gap")))
        if has_feasible_incumbent
        else None
    )
    effective_time_limit = _int_or_none(
        effective_limits.get("time_limit_sec", metadata.get("time_limit_sec", time_limit_seconds_requested))
    )
    stage1_best_obj_stop_enabled = metadata.get("stage1_best_obj_stop_enabled")
    stage1_best_obj_stop_applied = metadata.get("stage1_best_obj_stop_applied")
    stage1_certified_gap = _float_or_none(
        metadata.get("stage1_certified_mip_gap_ratio")
    )
    certified_gap = _float_or_none(
        metadata.get("certified_mip_gap_ratio")
    )
    if certified_gap is None:
        certified_gap = stage1_certified_gap
    certified_bound = _float_or_none(
        metadata.get("certified_best_bound")
    )
    if certified_bound is None:
        certified_bound = _float_or_none(
            metadata.get("stage1_certified_best_bound")
        )
    gap_for_target = certified_gap if certified_gap is not None else achieved_gap
    mip_gap_target_met = bool(
        requested_gap is not None
        and gap_for_target is not None
        and gap_for_target <= requested_gap
    )
    return {
        "solve_time_sec": _float_or_none(metadata.get("solve_time_sec")),
        "time_limit_seconds_requested": _int_or_none(time_limit_seconds_requested),
        "time_limit_seconds_effective": effective_time_limit,
        "stage1_time_limit_seconds_requested": _int_or_none(
            stage1_time_limit_seconds_requested
        ),
        "stage2_time_limit_seconds_requested": _int_or_none(
            stage2_time_limit_seconds_requested
        ),
        "mip_gap_requested_ratio": requested_gap,
        "mip_gap_requested_percent": None if requested_gap is None else requested_gap * 100.0,
        "mip_gap_achieved_ratio": achieved_gap,
        "mip_gap_achieved_percent": None if achieved_gap is None else achieved_gap * 100.0,
        "gurobi_raw_best_bound": _float_or_none(
            metadata.get("raw_best_bound", metadata.get("best_bound"))
        ),
        "gurobi_raw_mip_gap_ratio": _float_or_none(
            metadata.get("raw_mip_gap_ratio", metadata.get("final_gap"))
        ),
        "certified_best_bound": certified_bound,
        "certified_mip_gap_ratio": certified_gap,
        "certified_mip_gap_percent": (
            None if certified_gap is None else certified_gap * 100.0
        ),
        "certified_mip_gap_semantics": str(
            metadata.get("certified_mip_gap_semantics")
            or metadata.get("stage1_certified_mip_gap_semantics")
            or ""
        ),
        "gurobi_mip_gap_is_ratio": True,
        "has_feasible_incumbent": has_feasible_incumbent,
        "incumbent_count": _int_or_none(
            metadata.get("incumbent_count")
        ),
        "first_feasible_sec": _float_or_none(
            metadata.get("first_feasible_sec")
        ),
        "nodes_explored": _int_or_none(
            metadata.get("nodes_explored")
        ),
        "solver_termination_reason": metadata.get("termination_reason"),
        "supports_exact_milp": bool(metadata.get("supports_exact_milp", False)),
        "fallback_applied": bool(metadata.get("fallback_applied", False)),
        "fallback_reason": str(metadata.get("fallback_reason") or ""),
        "research_run": bool(metadata.get("research_run", False)),
        "research_run_accepted": bool(metadata.get("research_run_accepted", False)),
        "research_acceptance_checks": dict(metadata.get("research_acceptance_checks") or {}),
        "research_feasibility_eligible": bool(metadata.get("research_feasibility_eligible", False)),
        "research_assignment_eligible": bool(
            str(metadata.get("phase") or "") == "phase2_assignment_only"
            and bool(metadata.get("research_run", False))
            and bool(metadata.get("research_run_accepted", False))
            and bool(metadata.get("research_feasibility_eligible", False))
        ),
        "research_cost_kpi_eligible": bool(metadata.get("research_cost_kpi_eligible", False)),
        "git_sha": metadata.get("git_sha"),
        "git_dirty": metadata.get("git_dirty"),
        "git_sha_after_solve": metadata.get("git_sha_after_solve"),
        "git_dirty_after_solve": metadata.get("git_dirty_after_solve"),
        "git_state_unchanged_during_solve": metadata.get(
            "git_state_unchanged_during_solve"
        ),
        "git_state_available": bool(metadata.get("git_state_available", False)),
        "git_state_error": metadata.get("git_state_error"),
        "research_submission_git_provenance_eligible": bool(
            metadata.get("research_submission_git_provenance_eligible", False)
        ),
        "bff_runtime_git_attestation": dict(
            metadata.get("bff_runtime_git_attestation") or {}
        ),
        "source_provenance_exact": bool(metadata.get("source_provenance_exact", False)),
        "derived_source_split": bool(metadata.get("derived_source_split", False)),
        "synthetic_pv_fallback_allowed": bool(metadata.get("synthetic_pv_fallback_allowed", False)),
        "synthetic_pv_fallback_applied": bool(metadata.get("synthetic_pv_fallback_applied", False)),
        "successor_pruning_enabled": bool(metadata.get("successor_pruning_enabled", False)),
        "arc_pruning_summary": dict(metadata.get("arc_pruning_summary") or {}),
        "requested_mip_gap": requested_gap,
        "achieved_mip_gap": achieved_gap,
        "mip_gap_target_met": mip_gap_target_met,
        "requested_phase": str(metadata.get("requested_phase") or ""),
        "requested_phase_token": str(metadata.get("requested_phase_token") or ""),
        "resolved_phase": str(metadata.get("resolved_phase") or ""),
        "executed_phase": str(metadata.get("executed_phase") or ""),
        "stage1_solver_status": metadata.get("stage1_solver_status"),
        "stage1_termination_reason": metadata.get("stage1_termination_reason"),
        "stage1_best_obj_stop_enabled": stage1_best_obj_stop_enabled,
        "stage1_best_obj_stop_applied": stage1_best_obj_stop_applied,
        "stage1_best_obj_stop_threshold": metadata.get(
            "stage1_certified_gap_stop_threshold"
        ),
        "stage1_best_obj_stop_triggered": bool(
            metadata.get("stage1_certified_gap_stop_triggered", False)
        ),
        "stage1_gurobi_raw_best_bound": _float_or_none(
            metadata.get("stage1_gurobi_raw_best_bound")
        ),
        "stage1_gurobi_raw_mip_gap_ratio": _float_or_none(
            metadata.get("stage1_gurobi_raw_mip_gap_ratio")
        ),
        "stage1_gurobi_raw_mip_gap_percent": (
            None
            if _float_or_none(metadata.get("stage1_gurobi_raw_mip_gap_ratio")) is None
            else _float_or_none(metadata.get("stage1_gurobi_raw_mip_gap_ratio"))
            * 100.0
        ),
        "stage1_certified_best_bound": _float_or_none(
            metadata.get("stage1_certified_best_bound")
        ),
        "stage1_certified_mip_gap_ratio": stage1_certified_gap,
        "stage1_certified_mip_gap_percent": (
            None
            if stage1_certified_gap is None
            else stage1_certified_gap * 100.0
        ),
        "stage1_certified_mip_gap_semantics": metadata.get(
            "stage1_certified_mip_gap_semantics"
        ),
        "stage1_weather_aware_lower_bound": _float_or_none(
            metadata.get("stage1_weather_aware_lower_bound")
        ),
        "stage1_weather_aware_lower_bound_semantics": metadata.get(
            "stage1_weather_aware_lower_bound_semantics"
        ),
        "stage1_analytical_objective_lower_bound": _float_or_none(
            metadata.get("stage1_analytical_objective_lower_bound")
        ),
        "stage1_vehicle_usage_analytical_lower_bound": _float_or_none(
            metadata.get(
                "stage1_vehicle_usage_analytical_lower_bound"
            )
        ),
        "stage1_analytical_weather_energy_fuel_lower_bound": _float_or_none(
            metadata.get(
                "stage1_analytical_weather_energy_fuel_lower_bound"
            )
        ),
        "stage1_analytical_weather_energy_fuel_lower_bound_details": dict(
            metadata.get(
                "stage1_analytical_weather_energy_fuel_lower_bound_details"
            )
            or {}
        ),
        "stage1_analytical_total_objective_certificate_eligible": bool(
            metadata.get(
                "stage1_analytical_total_objective_certificate_eligible",
                False,
            )
        ),
        "stage1_analytical_total_objective_certificate_blockers": list(
            metadata.get(
                "stage1_analytical_total_objective_certificate_blockers"
            )
            or []
        ),
        "stage1_analytical_objective_lower_bound_semantics": metadata.get(
            "stage1_analytical_objective_lower_bound_semantics"
        ),
        "stage1_runtime_seconds": _float_or_none(
            metadata.get("stage1_runtime_seconds")
        ),
        "stage1_primary_runtime_seconds": _float_or_none(
            metadata.get("stage1_primary_runtime_seconds")
        ),
        "stage1_primary_search_time_limit_seconds": _float_or_none(
            metadata.get("stage1_primary_search_time_limit_seconds")
        ),
        "stage1_candidate_enumeration_reserve_seconds": _float_or_none(
            metadata.get("stage1_candidate_enumeration_reserve_seconds")
        ),
        "stage1_cost_ranked_composition_budget_enabled": bool(
            metadata.get(
                "stage1_cost_ranked_composition_budget_enabled",
                False,
            )
        ),
        "stage1_cost_ranked_composition_budget_semantics": metadata.get(
            "stage1_cost_ranked_composition_budget_semantics"
        ),
        "stage1_candidate_enumeration_runtime_seconds": _float_or_none(
            metadata.get("stage1_candidate_enumeration_runtime_seconds")
        ),
        "stage1_candidate_powertrain_pattern_no_good_cut_count": _int_or_none(
            metadata.get(
                "stage1_candidate_powertrain_pattern_no_good_cut_count"
            )
        ),
        "stage1_candidate_enumeration_events": list(
            metadata.get("stage1_candidate_enumeration_events") or []
        ),
        "stage1_time_limit_seconds_effective": _int_or_none(
            metadata.get("stage1_time_limit_sec_effective")
        ),
        "stage2_runtime_seconds": _float_or_none(
            metadata.get("stage2_runtime_seconds")
        ),
        "stage2_time_limit_seconds_effective": _int_or_none(
            metadata.get("stage2_time_limit_sec_effective")
        ),
        "stage1_search_telemetry": dict(
            metadata.get("stage1_search_telemetry") or {}
        ),
        "stage1_model_variable_count": _int_or_none(
            metadata.get("stage1_model_variable_count")
        ),
        "stage1_model_constraint_count": _int_or_none(
            metadata.get("stage1_model_constraint_count")
        ),
        "fragment_temporal_occupancy_constraint_count": _int_or_none(
            metadata.get("fragment_temporal_occupancy_constraint_count")
        ),
        "fragment_pairwise_depot_reset_constraint_count": _int_or_none(
            metadata.get("fragment_pairwise_depot_reset_constraint_count")
        ),
        "fragment_pairwise_depot_reset_constraint_mode": metadata.get(
            "fragment_pairwise_depot_reset_constraint_mode"
        ),
        "fragment_transition_lazy_separator": dict(
            metadata.get("fragment_transition_lazy_separator") or {}
        ),
        "overlap_clique_constraint_count": _int_or_none(
            metadata.get("overlap_clique_constraint_count")
        ),
        "stage1_vehicle_count_lower_bound": _int_or_none(
            metadata.get("stage1_vehicle_count_lower_bound")
        ),
        "stage1_vehicle_count_lower_bound_constraint_count": _int_or_none(
            metadata.get(
                "stage1_vehicle_count_lower_bound_constraint_count"
            )
        ),
        "stage1_vehicle_count_lower_bound_semantics": metadata.get(
            "stage1_vehicle_count_lower_bound_semantics"
        ),
        "stage1_identical_vehicle_groups": list(
            metadata.get("stage1_identical_vehicle_groups") or ()
        ),
        "stage1_identical_vehicle_group_count": _int_or_none(
            metadata.get("stage1_identical_vehicle_group_count")
        ),
        "stage1_identical_vehicle_activation_prefix_constraint_count": _int_or_none(
            metadata.get(
                "stage1_identical_vehicle_activation_prefix_constraint_count"
            )
        ),
        "stage1_energy_cost_proxy_used_in_objective": bool(
            metadata.get("stage1_energy_cost_proxy_used_in_objective", False)
        ),
        "stage1_time_indexed_energy_recourse_configuration": dict(
            metadata.get(
                "stage1_time_indexed_energy_recourse_configuration"
            )
            or {}
        ),
        "stage1_time_indexed_energy_recourse_weather_input": dict(
            metadata.get(
                "stage1_time_indexed_energy_recourse_weather_input"
            )
            or {}
        ),
        "stage1_time_indexed_energy_recourse_result": dict(
            metadata.get("stage1_time_indexed_energy_recourse_result")
            or {}
        ),
        "stage1_accounting_objective_components": dict(
            metadata.get("stage1_accounting_objective_components")
            or {}
        ),
        "stage1_driver_cost_constraint_count": _int_or_none(
            metadata.get("stage1_driver_cost_constraint_count")
        ),
        "stage1_degradation_cost_term_count": _int_or_none(
            metadata.get("stage1_degradation_cost_term_count")
        ),
        "stage1_switch_cost_term_count": _int_or_none(
            metadata.get("stage1_switch_cost_term_count")
        ),
        "stage1_stage2_candidate_limit_requested": _int_or_none(
            metadata.get("stage1_stage2_candidate_limit_requested")
        ),
        "stage1_composition_search_radius_requested": _int_or_none(
            metadata.get("stage1_composition_search_radius_requested")
        ),
        "stage1_bev_frontier_enabled": bool(
            metadata.get("stage1_bev_frontier_enabled", False)
        ),
        "stage1_composition_search_runtime_seconds": _float_or_none(
            metadata.get("stage1_composition_search_runtime_seconds")
        ),
        "stage1_composition_search_certificate_evidence_wall_seconds": (
            _float_or_none(
                metadata.get(
                    "stage1_composition_search_certificate_evidence_wall_seconds"
                )
            )
        ),
        "stage1_used_powertrain_composition_search": dict(
            metadata.get("stage1_used_powertrain_composition_search") or {}
        ),
        "stage1_used_powertrain_composition_search_accepted": bool(
            metadata.get(
                "stage1_used_powertrain_composition_search_accepted", False
            )
        ),
        "bev_cost_frontier": dict(
            metadata.get("bev_cost_frontier") or {}
        ),
        "integrated_actual_cost_objective_requested": bool(
            metadata.get(
                "integrated_actual_cost_objective_requested", False
            )
        ),
        "integrated_actual_cost_contract_applied": bool(
            metadata.get("integrated_actual_cost_contract_applied", False)
        ),
        "integrated_mip_focus": _int_or_none(
            metadata.get("integrated_mip_focus")
        ),
        "integrated_heuristics": _float_or_none(
            metadata.get("integrated_heuristics")
        ),
        "integrated_symmetry": _int_or_none(
            metadata.get("integrated_symmetry")
        ),
        "integrated_search_profile": dict(
            metadata.get("integrated_search_profile") or {}
        ),
        "integrated_analytical_objective_lower_bound": _float_or_none(
            metadata.get("integrated_analytical_objective_lower_bound")
        ),
        "integrated_vehicle_usage_analytical_lower_bound": _float_or_none(
            metadata.get(
                "integrated_vehicle_usage_analytical_lower_bound"
            )
        ),
        "integrated_analytical_weather_energy_fuel_lower_bound": _float_or_none(
            metadata.get(
                "integrated_analytical_weather_energy_fuel_lower_bound"
            )
        ),
        "integrated_analytical_weather_energy_fuel_lower_bound_details": dict(
            metadata.get(
                "integrated_analytical_weather_energy_fuel_lower_bound_details"
            )
            or {}
        ),
        "integrated_analytical_objective_floor_constraint_count": _int_or_none(
            metadata.get(
                "integrated_analytical_objective_floor_constraint_count"
            )
        ),
        "integrated_analytical_objective_floor_certificate_eligible": bool(
            metadata.get(
                "integrated_analytical_objective_floor_certificate_eligible",
                False,
            )
        ),
        "integrated_analytical_objective_floor_blockers": list(
            metadata.get("integrated_analytical_objective_floor_blockers")
            or ()
        ),
        "integrated_analytical_objective_lower_bound_semantics": metadata.get(
            "integrated_analytical_objective_lower_bound_semantics"
        ),
        "integrated_verified_start_search_bounds": dict(
            metadata.get("integrated_verified_start_search_bounds") or {}
        ),
        "integrated_verified_start_objective_cap_constraint_count": _int_or_none(
            metadata.get(
                "integrated_verified_start_objective_cap_constraint_count"
            )
        ),
        "integrated_verified_start_vehicle_day_cap_constraint_count": _int_or_none(
            metadata.get(
                "integrated_verified_start_vehicle_day_cap_constraint_count"
            )
        ),
        "integrated_verified_start_search_bound_semantics": metadata.get(
            "integrated_verified_start_search_bound_semantics"
        ),
        "integrated_identical_vehicle_groups": list(
            metadata.get("integrated_identical_vehicle_groups") or ()
        ),
        "integrated_identical_vehicle_group_count": _int_or_none(
            metadata.get("integrated_identical_vehicle_group_count")
        ),
        "integrated_identical_vehicle_activation_prefix_constraint_count": _int_or_none(
            metadata.get(
                "integrated_identical_vehicle_activation_prefix_constraint_count"
            )
        ),
        "integrated_identical_vehicle_symmetry_semantics": metadata.get(
            "integrated_identical_vehicle_symmetry_semantics"
        ),
        "integrated_phase3_iis_assignment_guidance_pattern_count": _int_or_none(
            metadata.get(
                "integrated_phase3_iis_assignment_guidance_pattern_count"
            )
        ),
        "integrated_phase3_iis_assignment_guidance_pattern_hashes": list(
            metadata.get(
                "integrated_phase3_iis_assignment_guidance_pattern_hashes"
            )
            or ()
        ),
        "integrated_phase3_iis_assignment_guidance_source_candidate_hashes": list(
            metadata.get(
                "integrated_phase3_iis_assignment_guidance_source_candidate_hashes"
            )
            or ()
        ),
        "integrated_phase3_iis_assignment_guidance_variable_count": _int_or_none(
            metadata.get(
                "integrated_phase3_iis_assignment_guidance_variable_count"
            )
        ),
        "integrated_phase3_iis_assignment_guidance_branch_priority": _int_or_none(
            metadata.get(
                "integrated_phase3_iis_assignment_guidance_branch_priority"
            )
        ),
        "integrated_phase3_iis_assignment_guidance_semantics": metadata.get(
            "integrated_phase3_iis_assignment_guidance_semantics"
        ),
        "integrated_root_method": _int_or_none(
            metadata.get("integrated_root_method")
        ),
        "integrated_node_method": _int_or_none(
            metadata.get("integrated_node_method")
        ),
        "integrated_soft_mem_limit_gb": _float_or_none(
            metadata.get("integrated_soft_mem_limit_gb")
        ),
        "integrated_redundant_arc_link_constraints_omitted": _int_or_none(
            metadata.get("integrated_redundant_arc_link_constraints_omitted")
        ),
        "integrated_fragment_pairwise_constraint_count": _int_or_none(
            metadata.get("integrated_fragment_pairwise_constraint_count")
        ),
        "integrated_fragment_pairwise_constraint_mode": metadata.get(
            "integrated_fragment_pairwise_constraint_mode"
        ),
        "integrated_fragment_transition_lazy_separator": dict(
            metadata.get("integrated_fragment_transition_lazy_separator") or {}
        ),
        "integrated_fragment_occupancy_constraint_count": _int_or_none(
            metadata.get("integrated_fragment_occupancy_constraint_count")
        ),
        "integrated_overlap_clique_constraint_count": _int_or_none(
            metadata.get("integrated_overlap_clique_constraint_count")
        ),
        "integrated_activity_blocking_formulation": metadata.get(
            "integrated_activity_blocking_formulation"
        ),
        "integrated_activity_blocking_constraint_count": _int_or_none(
            metadata.get("integrated_activity_blocking_constraint_count")
        ),
        "integrated_redundant_endpoint_away_blocking_terms_omitted": (
            _int_or_none(
                metadata.get(
                    "integrated_redundant_endpoint_away_blocking_terms_omitted"
                )
            )
        ),
        "integrated_redundant_endpoint_away_blocking_semantics": metadata.get(
            "integrated_redundant_endpoint_away_blocking_semantics"
        ),
        "integrated_certified_gap_stop_threshold": _float_or_none(
            metadata.get("integrated_certified_gap_stop_threshold")
        ),
        "integrated_certified_gap_at_verified_start": _float_or_none(
            metadata.get("integrated_certified_gap_at_verified_start")
        ),
        "integrated_certified_gap_stop_applied": bool(
            metadata.get("integrated_certified_gap_stop_applied", False)
        ),
        "integrated_certified_gap_stop_semantics": metadata.get(
            "integrated_certified_gap_stop_semantics"
        ),
        "phase4_phase3_seed_audit": phase4_seed_audit,
        "phase4_phase3_seed_enabled": bool(
            phase4_seed_audit.get("requested", False)
        ),
        "phase4_phase3_seed_time_limit_sec": _int_or_none(
            phase4_seed_audit.get("seed_time_limit_sec")
        ),
        "phase4_phase3_seed_wall_clock_budget_sec": _int_or_none(
            phase4_seed_audit.get("seed_wall_clock_budget_sec")
        ),
        "phase4_phase3_seed_wall_runtime_sec": _float_or_none(
            phase4_seed_audit.get("seed_wall_runtime_sec")
        ),
        "phase4_phase3_seed_model_build_overhead_allowance_sec": (
            _int_or_none(
                phase4_seed_audit.get(
                    "seed_model_build_overhead_allowance_sec"
                )
            )
        ),
        "phase4_phase3_seed_stage1_time_limit_sec": _int_or_none(
            phase4_seed_audit.get("seed_stage1_time_limit_sec")
        ),
        "phase4_phase3_seed_stage2_time_limit_sec": _int_or_none(
            phase4_seed_audit.get("seed_stage2_time_limit_sec")
        ),
        "phase4_phase3_seed_candidate_limit": _int_or_none(
            phase4_seed_audit.get("seed_stage1_stage2_candidate_limit")
        ),
        "phase4_phase3_seed_candidate_evaluation_order": (
            phase4_seed_audit.get(
                "seed_stage1_stage2_candidate_evaluation_order"
            )
        ),
        "phase4_phase3_seed_candidate_evaluation_initial_budget_sec": (
            _float_or_none(
                phase4_seed_audit.get(
                    "seed_stage1_stage2_candidate_evaluation_initial_budget_sec"
                )
            )
        ),
        "phase4_phase3_seed_stage2_iis_assignment_guidance_pattern_count": _int_or_none(
            phase4_seed_audit.get(
                "seed_stage2_iis_assignment_guidance_pattern_count"
            )
        ),
        "phase4_phase3_seed_stage2_iis_assignment_guidance_source_candidate_hashes": list(
            phase4_seed_audit.get(
                "seed_stage2_iis_assignment_guidance_source_candidate_hashes"
            )
            or ()
        ),
        "phase4_phase3_seed_stage2_iis_assignment_guidance_semantics": (
            phase4_seed_audit.get(
                "seed_stage2_iis_assignment_guidance_semantics"
            )
        ),
        "phase4_phase3_seed_cost_ranked_composition_budget_enabled": bool(
            phase4_seed_audit.get(
                "seed_cost_ranked_composition_budget_enabled",
                False,
            )
        ),
        "phase4_phase3_seed_cost_ranked_composition_budget_semantics": (
            phase4_seed_audit.get(
                "seed_cost_ranked_composition_budget_semantics"
            )
        ),
        "phase4_phase3_seed_composition_search_radius": _int_or_none(
            phase4_seed_audit.get(
                "seed_stage1_composition_search_radius"
            )
        ),
        "phase4_phase3_seed_available_vehicle_count": _int_or_none(
            phase4_seed_audit.get("seed_available_vehicle_count")
        ),
        "phase4_phase3_seed_required_candidate_limit": _int_or_none(
            phase4_seed_audit.get(
                "seed_composition_search_required_candidate_limit"
            )
        ),
        "phase4_phase3_seed_required_composition_search_radius": (
            _int_or_none(
                phase4_seed_audit.get(
                    "seed_composition_search_required_radius"
                )
            )
        ),
        "phase4_phase3_seed_composition_search_scope": (
            phase4_seed_audit.get("seed_composition_search_scope")
        ),
        "phase4_phase3_seed_inventory_span_truncated": bool(
            phase4_seed_audit.get(
                "seed_composition_search_inventory_span_truncated",
                True,
            )
        ),
        "phase4_phase3_seed_search_directionality": (
            phase4_seed_audit.get("seed_search_directionality")
        ),
        "phase4_phase3_seed_bev_frontier_enabled": bool(
            phase4_seed_audit.get("seed_bev_frontier_enabled", False)
        ),
        "phase4_phase3_seed_unused_bev_neighborhood_enabled": bool(
            phase4_seed_audit.get(
                "unused_bev_activation_neighborhood_enabled",
                False,
            )
        ),
        "phase4_phase3_seed_unused_bev_neighborhood_time_limit_sec": (
            _int_or_none(
                phase4_seed_audit.get(
                    "unused_bev_activation_neighborhood_time_limit_sec"
                )
            )
        ),
        "phase4_phase3_seed_unused_bev_neighborhood_per_solve_sec": (
            _int_or_none(
                phase4_seed_audit.get(
                    "unused_bev_activation_neighborhood_per_solve_sec"
                )
            )
        ),
        "phase4_phase3_seed_unused_bev_neighborhood_max_evaluations": (
            _int_or_none(
                phase4_seed_audit.get(
                    "unused_bev_activation_neighborhood_max_evaluations"
                )
            )
        ),
        "phase4_phase3_seed_route_band_repartition_time_limit_sec": (
            _int_or_none(
                phase4_seed_audit.get(
                    "route_band_repartition_time_limit_sec"
                )
            )
        ),
        "phase4_phase3_seed_unused_bev_identity_exchange_rounds": (
            _int_or_none(
                phase4_seed_audit.get(
                    "unused_bev_identity_exchange_rounds"
                )
            )
        ),
        "phase4_phase3_seed_powertrain_duty_swap_rounds": (
            _int_or_none(
                phase4_seed_audit.get("powertrain_duty_swap_rounds")
            )
        ),
        "phase4_phase3_seed_unused_bev_neighborhood": dict(
            phase4_seed_audit.get(
                "unused_bev_activation_neighborhood"
            )
            or {}
        ),
        "phase4_integrated_seed_recourse_preflight_enabled": bool(
            phase4_seed_audit.get(
                "integrated_seed_recourse_preflight_enabled",
                False,
            )
        ),
        "phase4_integrated_seed_recourse_time_limit_sec": _int_or_none(
            phase4_seed_audit.get(
                "integrated_seed_recourse_time_limit_sec"
            )
        ),
        "phase4_integrated_seed_recourse_preflight_requested": bool(
            integrated_start_audit.get(
                "dispatch_fixed_recourse_requested",
                False,
            )
        ),
        "phase4_integrated_seed_recourse_preflight_feasible": bool(
            integrated_start_audit.get(
                "integrated_dispatch_fixed_recourse_feasible",
                False,
            )
        ),
        "phase4_total_solver_time_budget_sec": _int_or_none(
            phase4_seed_audit.get("total_solver_time_budget_sec")
        ),
        "integrated_warm_start_audit": integrated_start_audit,
        "integrated_primary_objective_kind": metadata.get(
            "integrated_primary_objective_kind"
        ),
        "integrated_lexicographic_solve_mode": metadata.get(
            "integrated_lexicographic_solve_mode"
        ),
        "integrated_lexicographic_primary_value": _float_or_none(
            metadata.get("integrated_lexicographic_primary_value")
        ),
        "integrated_lexicographic_primary_best_bound": _float_or_none(
            metadata.get("integrated_lexicographic_primary_best_bound")
        ),
        "integrated_lexicographic_primary_certified": bool(
            metadata.get(
                "integrated_lexicographic_primary_certified",
                False,
            )
        ),
        "integrated_lexicographic_primary_certificate": metadata.get(
            "integrated_lexicographic_primary_certificate"
        ),
        "integrated_lexicographic_cost_status": metadata.get(
            "integrated_lexicographic_cost_status"
        ),
        "integrated_lexicographic_cost_objective_jpy": _float_or_none(
            metadata.get("integrated_lexicographic_cost_objective_jpy")
        ),
        "integrated_lexicographic_cost_best_bound_jpy": _float_or_none(
            metadata.get("integrated_lexicographic_cost_best_bound_jpy")
        ),
        "integrated_lexicographic_cost_raw_mip_gap_ratio": _float_or_none(
            metadata.get(
                "integrated_lexicographic_cost_raw_mip_gap_ratio"
            )
        ),
        "integrated_lexicographic_completed_objectives": list(
            metadata.get("integrated_lexicographic_completed_objectives")
            or ()
        ),
        "integrated_ev_utilization_mode": metadata.get(
            "integrated_ev_utilization_mode", "disabled"
        ),
        "integrated_actual_cost_upper_bound_jpy": _float_or_none(
            metadata.get("integrated_actual_cost_upper_bound_jpy")
        ),
        "integrated_actual_cost_upper_bound_delta_ratio": _float_or_none(
            metadata.get(
                "integrated_actual_cost_upper_bound_delta_ratio"
            )
        ),
        "integrated_actual_cost_upper_bound_verified": bool(
            metadata.get("integrated_actual_cost_upper_bound_verified", False)
        ),
        "integrated_primary_ice_fuel_l": _float_or_none(
            metadata.get("integrated_primary_ice_fuel_l")
        ),
        "actual_cost_objective_structural_contract_passed": bool(
            metadata.get(
                "actual_cost_objective_structural_contract_passed", False
            )
        ),
        "actual_cost_objective_numeric_reconciliation_passed": bool(
            metadata.get(
                "actual_cost_objective_numeric_reconciliation_passed",
                False,
            )
        ),
        "actual_cost_objective_residual_jpy": _float_or_none(
            metadata.get("actual_cost_objective_residual_jpy")
        ),
        "vehicle_usage_cost_semantics": metadata.get(
            "vehicle_usage_cost_semantics"
        ),
        "vehicle_usage_cost_semantics_classified": bool(
            metadata.get("vehicle_usage_cost_semantics_classified", False)
        ),
        "stage1_pool_solution_count": _int_or_none(
            metadata.get("stage1_pool_solution_count")
        ),
        "stage1_distinct_candidate_count": _int_or_none(
            metadata.get("stage1_distinct_candidate_count")
        ),
        "stage1_stage2_candidate_count_evaluated": _int_or_none(
            metadata.get("stage1_stage2_candidate_count_evaluated")
        ),
        "stage1_stage2_feasible_candidate_count": _int_or_none(
            metadata.get("stage1_stage2_feasible_candidate_count")
        ),
        "stage1_stage2_selected_candidate_index": _int_or_none(
            metadata.get("stage1_stage2_selected_candidate_index")
        ),
        "stage1_stage2_selected_candidate_hash": metadata.get(
            "stage1_stage2_selected_candidate_hash"
        ),
        "stage1_stage2_selected_canonical_actual_cost_jpy": _float_or_none(
            metadata.get(
                "stage1_stage2_selected_canonical_actual_cost_jpy"
            )
        ),
        "stage1_primary_incumbent_objective_jpy": _float_or_none(
            metadata.get("stage1_primary_incumbent_objective_jpy")
        ),
        "stage1_selected_candidate_relaxed_objective_jpy": _float_or_none(
            metadata.get(
                "stage1_selected_candidate_relaxed_objective_jpy"
            )
        ),
        "stage2_feedback_iteration": _int_or_none(
            metadata.get("stage2_feedback_iteration")
        ),
        "stage2_feedback_history": list(
            metadata.get("stage2_feedback_history") or ()
        ),
        # A single frontend run is never a runtime comparison. Matched solver
        # controls and repeated executions are a separate experiment contract.
        "runtime_comparison_eligible": False,
        "runtime_comparison_eligibility_reason": (
            "Stage 1 BestObjStop was active; compare wall-clock time only after "
            "disabling it in every case."
            if stage1_best_obj_stop_applied is True
            else "Single frontend execution: repeated matched cases are required "
            "before any wall-clock comparison claim."
        ),
        "gurobi_threads": _int_or_none(metadata.get("gurobi_threads")),
        "random_seed": random_seed,
        "stage1_gurobi_feasibility_tol": _float_or_none(
            metadata.get("stage1_gurobi_feasibility_tol")
        ),
        "stage2_gurobi_feasibility_tol": _float_or_none(
            metadata.get("stage2_gurobi_feasibility_tol")
        ),
        "stage2_gurobi_integrality_tol": _float_or_none(
            metadata.get("stage2_gurobi_integrality_tol")
        ),
        "stage1_numeric_diagnostics": dict(
            metadata.get("stage1_numeric_diagnostics") or {}
        ),
        "stage2_numeric_diagnostics": dict(
            metadata.get("stage2_numeric_diagnostics") or {}
        ),
        "interactive_runtime_controls": dict(
            metadata.get("interactive_runtime_controls") or {}
        ),
        "interactive_operation_time_window_controls": dict(
            metadata.get("interactive_operation_time_window_controls") or {}
        ),
        "interactive_terminal_soc_controls": dict(
            metadata.get("interactive_terminal_soc_controls") or {}
        ),
        "stage2_solver_status": metadata.get("stage2_solver_status"),
        "stage1_feasible": metadata.get("stage1_feasible"),
        "stage2_feasible": metadata.get("stage2_feasible"),
        "supports_two_stage_milp": metadata.get("supports_two_stage_milp"),
        "supports_integrated_exact_milp": metadata.get("supports_integrated_exact_milp"),
        "assignment_candidate_available": bool(metadata.get("assignment_candidate_available", False)),
    }


def _configure_assignment_energy_diagnostics(
    problem: Any,
    *,
    phase_token: str,
    output_dir: str | Path,
    research_run: bool,
) -> None:
    """Attach candidate diagnostics without broadening Phase-3 feedback.

    Phase 4 obtains its incumbent from an internal Phase-3 candidate search,
    so failed fixed-assignment Stage-2 candidates need the same IIS/path
    evidence as a directly requested Phase-3 run.  Recursive no-good feedback
    remains limited to a direct Phase-3 solve; enabling diagnostics must not
    silently change the Phase-4 search problem.
    """

    metadata = getattr(problem, "metadata", None)
    if not isinstance(metadata, dict):
        return
    if phase_token in {"phase3_two_stage", "phase4_integrated"}:
        metadata["phase3_diagnostics_dir"] = str(
            Path(output_dir) / "diagnostics"
        )
    if phase_token != "phase3_two_stage":
        return
    metadata["stage2_feedback_max_iterations"] = (
        2 if research_run else 1
    )
    metadata["stage2_feedback_policy"] = (
        "retry_only_after_gurobi_infeasible_certificate_with_"
        "full_assignment_no_good_cut"
    )


def _run_optimization(
    scenario_id: str,
    job_id: str,
    prepared_input_id: str,
    requested_prepared_input_id: Optional[str],
    mode: str,
    time_limit_seconds: int,
    mip_gap: float,
    random_seed: int,
    service_id: str,
    depot_id: Optional[str],
    rebuild_dispatch: bool,
    use_existing_duties: bool,
    alns_iterations: int,
    no_improvement_limit: int,
    destroy_fraction: float,
    timestep_min: Optional[int] = None,
    enable_weather_operation_policy: Optional[bool] = None,
    weather_proxy_forecast_path: Optional[str] = None,
    research_run: bool = False,
    stage1_time_limit_seconds: Optional[int] = None,
    stage2_time_limit_seconds: Optional[int] = None,
    stage1_best_obj_stop_enabled: bool = INTERACTIVE_STAGE1_BEST_OBJ_STOP_ENABLED,
    gurobi_threads: Optional[int] = INTERACTIVE_GUROBI_THREADS,
    run_profile: str = DEFAULT_FRONTEND_RUN_PROFILE,
    run_hourly_rolling: bool = True,
    rolling_execution_minutes: int = 60,
    frontend_request_payload: Optional[Dict[str, Any]] = None,
    stage1_stage2_candidate_limit: int = 1,
    stage1_composition_search_radius: int = 0,
    stage1_bev_frontier_enabled: bool = False,
    stage1_bev_frontier_min_count: int = 15,
    stage1_bev_frontier_max_count: int = 35,
    stage1_bev_frontier_target_time_limit_seconds: int = 120,
    integrated_actual_cost_objective: bool = False,
    integrated_ev_utilization_mode: str = "disabled",
    integrated_actual_cost_upper_bound_jpy: Optional[float] = None,
    integrated_actual_cost_upper_bound_delta_ratio: Optional[float] = None,
    co2_emissions_cap_kg: Optional[float] = None,
) -> None:
    output_dir: Optional[str] = None
    raw_frontend_request_payload = dict(frontend_request_payload or {})
    require_all_available_bevs = bool(
        raw_frontend_request_payload.get("require_all_available_bevs", False)
    )
    interactive_bev_utilization_policy: Dict[str, Any] = {
        "enabled": require_all_available_bevs,
        "status": "not_applied",
    }
    interactive_research_contract: Dict[str, Any] = {
        "enabled": bool(research_run),
        "status": "not_applied",
    }
    run_profile = normalize_frontend_run_profile(run_profile)
    rolling_required = frontend_rolling_is_required(run_profile)
    requested_rolling_controls = {
        "run_hourly_rolling": bool(run_hourly_rolling),
        "rolling_execution_minutes": int(rolling_execution_minutes),
    }
    # The ordinary frontend profile is server authoritative: stale clients or
    # hand-written requests cannot silently downgrade it to day-ahead only.
    run_hourly_rolling = bool(rolling_required)
    rolling_execution_minutes = 60
    effective_rolling_controls = {
        "run_profile": run_profile,
        "run_hourly_rolling": run_hourly_rolling,
        "rolling_execution_minutes": rolling_execution_minutes,
        "server_enforced": bool(rolling_required),
        "requested": requested_rolling_controls,
    }
    interactive_runtime_controls = _interactive_runtime_controls_payload(
        requested_stage1_best_obj_stop_enabled=stage1_best_obj_stop_enabled,
        requested_gurobi_threads=gurobi_threads,
    )
    # This worker is only entered by the BFF interactive endpoint.  Enforce at
    # the last boundary before OptimizationConfig construction rather than
    # trusting a UI/default request field.
    stage1_best_obj_stop_enabled = INTERACTIVE_STAGE1_BEST_OBJ_STOP_ENABLED
    gurobi_threads = INTERACTIVE_GUROBI_THREADS
    try:
        solver_mode = _normalize_solver_mode(mode)
        job_store.update_job(
            job_id,
            status="running",
            progress=5,
            message="Preparing optimization inputs...",
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="prepare",
                mode=mode,
            ),
        )

        if not depot_id:
            raise ValueError("No depot selected. Configure dispatch scope first.")

        base_scenario = store.get_scenario_document_shallow(scenario_id)
        prepared_input_path = _prepared_inputs_root() / scenario_id / f"{prepared_input_id}.json"
        prepared_payload = load_prepared_input(
            scenario_id=scenario_id,
            prepared_input_id=prepared_input_id,
            scenarios_dir=_prepared_inputs_root(),
        )
        scenario = materialize_scenario_from_prepared_input(
            base_scenario,
            prepared_payload,
        )
        _apply_timestep_min_to_scenario(scenario, timestep_min)
        interactive_operation_time_window_controls = (
            _apply_interactive_operation_time_window_controls(scenario)
        )
        scenario, weather_forecast, weather_profile = _prepare_weather_policy_for_scenario(
            scenario,
            enable_weather_operation_policy=enable_weather_operation_policy,
            weather_proxy_forecast_path=weather_proxy_forecast_path,
        )
        interactive_terminal_soc_controls = _apply_interactive_bev_terminal_soc_policy(
            scenario
        )
        interactive_research_contract = _apply_interactive_research_contract(
            scenario,
            research_run=bool(research_run),
            depot_id=depot_id,
        )

        if rebuild_dispatch:
            _persist_prepared_scope_artifacts(
                scenario_id,
                scenario,
                clear_stale_dispatch=False,
            )
            _rebuild_dispatch_artifacts(scenario_id, service_id, depot_id)
            scenario["duties"] = store.get_field(scenario_id, "duties") or []
            scenario["blocks"] = store.get_field(scenario_id, "blocks") or []
            graph_meta = store.get_field(scenario_id, "graph")
            if isinstance(graph_meta, dict):
                scenario["graph"] = {k: v for k, v in graph_meta.items() if k != "arcs"}
                scenario["graph"]["arcs"] = []
            else:
                scenario["graph"] = {
                    "source": "prepared_scope",
                    "total_arcs": 0,
                    "feasible_arcs": 0,
                    "infeasible_arcs": 0,
                }
        elif use_existing_duties:
            scenario["duties"] = store.get_field(scenario_id, "duties") or []
            scenario["blocks"] = store.get_field(scenario_id, "blocks") or []
            graph_meta = store.get_field(scenario_id, "graph")
            if isinstance(graph_meta, dict):
                scenario["graph"] = {k: v for k, v in graph_meta.items() if k != "arcs"}
                scenario["graph"]["arcs"] = []
            else:
                scenario["graph"] = {
                    "source": "prepared_scope",
                    "total_arcs": 0,
                    "feasible_arcs": 0,
                    "infeasible_arcs": 0,
                }
        else:
            scenario["duties"] = []
            scenario["blocks"] = []
            scenario["graph"] = {
                "source": "prepared_scope",
                "total_arcs": 0,
                "feasible_arcs": 0,
                "infeasible_arcs": 0,
            }
        feed_context = _scenario_feed_context(scenario_id)
        output_dir = _scoped_output_dir(
            root=str(output_paths.outputs_root()),
            feed_context=feed_context,
            scenario_id=scenario_id,
            stage="optimization",
            service_id=service_id,
            depot_id=depot_id,
        )
        # Capture the code state before the solver starts.  This same record is
        # written into both the input bundle and the result/audit artifacts so
        # a later reviewer cannot be left with an empty Git SHA.
        run_git_state = collect_git_state()
        _require_clean_research_git_state(
            research_run=bool(research_run),
            git_state=run_git_state,
        )
        runtime_git_attestation = _require_matching_research_runtime_git_state(
            research_run=bool(research_run),
            current_git_state=run_git_state,
        )

        charging_summary_payload: Optional[Dict[str, Any]] = None
        charging_flow_payload: Optional[Dict[str, Any]] = None
        charging_payload_warning: Optional[str] = None
        run_input_provenance: Dict[str, Any] = {
            "status": "not_captured",
            "reason": "solver_path_did_not_build_canonical_input_provenance",
        }

        if solver_mode in {
            "thesis_mode", "debug_mode", "mode_milp_only", "mode_alns_only",
            "mode_ga_only", "mode_abc_only", "mode_hybrid",
            "phase1_charging_only", "phase2_assignment_only", "phase3_two_stage",
            "phase4_integrated", "diagnostic",
        }:
            # CANONICAL PATH: Uses src/optimization/ engine stack
            opt_mode = _parse_optimization_mode(solver_mode)
            engine_label = str(opt_mode.value or "optimization").upper()
            job_store.update_job(
                job_id,
                status="running",
                progress=25,
                message="Building canonical problem...",
                metadata=_job_metadata(
                    scenario_id=scenario_id,
                    service_id=service_id,
                    depot_id=depot_id,
                    stage="build_canonical",
                    mode=mode,
                    extra={
                        "rebuild_dispatch": rebuild_dispatch,
                        "use_existing_duties": use_existing_duties,
                        "prepared_input_id": prepared_input_id,
                        "prepared_input_path": str(prepared_input_path),
                    },
                ),
            )
            phase_token = _phase_from_solver_mode(solver_mode)
            requested_phase = phase_token or solver_mode
            is_diagnostic_mode = solver_mode == "diagnostic" or solver_mode == "debug_mode"
            prepared_cost_cfg = dict(
                ((scenario.get("scenario_overlay") or {}).get("cost_coefficients") or {})
            )
            prepared_simulation_cfg = dict(
                scenario.get("simulation_config") or {}
            )
            effective_co2_emissions_cap_kg = (
                co2_emissions_cap_kg
                if co2_emissions_cap_kg is not None
                else prepared_cost_cfg.get(
                    "co2_emissions_cap_kg",
                    prepared_simulation_cfg.get("co2_emissions_cap_kg"),
                )
            )
            opt_config = OptimizationConfig(
                mode=opt_mode,
                time_limit_sec=time_limit_seconds,
                stage1_time_limit_sec=stage1_time_limit_seconds,
                stage2_time_limit_sec=stage2_time_limit_seconds,
                stage1_best_obj_stop_enabled=bool(stage1_best_obj_stop_enabled),
                stage1_stage2_candidate_limit=(
                    max(int(stage1_stage2_candidate_limit), 10)
                    if bool(research_run)
                    else max(int(stage1_stage2_candidate_limit), 1)
                ),
                stage1_composition_search_radius=(
                    max(int(stage1_composition_search_radius), 2)
                    if bool(research_run)
                    else max(int(stage1_composition_search_radius), 0)
                ),
                stage1_bev_frontier_enabled=bool(
                    stage1_bev_frontier_enabled
                ),
                stage1_bev_frontier_min_count=max(
                    int(stage1_bev_frontier_min_count),
                    0,
                ),
                stage1_bev_frontier_max_count=max(
                    int(stage1_bev_frontier_max_count),
                    0,
                ),
                stage1_bev_frontier_target_time_limit_sec=max(
                    float(stage1_bev_frontier_target_time_limit_seconds),
                    1.0,
                ),
                integrated_actual_cost_objective=bool(
                    integrated_actual_cost_objective
                ),
                integrated_ev_utilization_mode=str(
                    integrated_ev_utilization_mode or "disabled"
                ),
                integrated_actual_cost_upper_bound_jpy=(
                    None
                    if integrated_actual_cost_upper_bound_jpy is None
                    else float(integrated_actual_cost_upper_bound_jpy)
                ),
                integrated_actual_cost_upper_bound_delta_ratio=(
                    None
                    if integrated_actual_cost_upper_bound_delta_ratio is None
                    else float(
                        integrated_actual_cost_upper_bound_delta_ratio
                    )
                ),
                co2_emissions_cap_kg=(
                    None
                    if effective_co2_emissions_cap_kg is None
                    else float(effective_co2_emissions_cap_kg)
                ),
                phase4_phase3_seed_enabled=(
                    phase_token == "phase4_integrated"
                ),
                phase4_phase3_seed_time_limit_sec=(
                    min(
                        max(int(time_limit_seconds) // 6, 60),
                        600,
                    )
                    if phase_token == "phase4_integrated"
                    else 600
                ),
                # Formal actual-cost Phase 4 uses a neutral feasible start.
                # A one-sided ``used BEV >= K`` frontier would be a directed
                # search policy and could survive as the final incumbent at a
                # time limit.  Explicit frontier sensitivities remain Phase 3
                # experiments and are never injected here.
                phase4_phase3_seed_bev_frontier_enabled=False,
                phase4_phase3_seed_unused_bev_neighborhood_enabled=(
                    phase_token == "phase4_integrated"
                ),
                phase4_phase3_seed_unused_bev_neighborhood_time_limit_sec=120,
                phase4_phase3_seed_unused_bev_neighborhood_per_solve_sec=5,
                phase4_phase3_seed_unused_bev_neighborhood_max_evaluations=(
                    512
                ),
                phase4_phase3_seed_route_band_repartition_time_limit_sec=90,
                phase4_phase3_seed_powertrain_duty_swap_rounds=2,
                phase4_phase3_seed_unused_bev_identity_exchange_rounds=2,
                phase4_integrated_seed_recourse_preflight_enabled=(
                    phase_token == "phase4_integrated"
                ),
                phase4_integrated_seed_recourse_time_limit_sec=(
                    min(
                        max(int(time_limit_seconds) // 12, 60),
                        300,
                    )
                    if phase_token == "phase4_integrated"
                    else 300
                ),
                gurobi_threads=gurobi_threads,
                mip_gap=mip_gap,
                random_seed=random_seed,
                alns_iterations=alns_iterations,
                no_improvement_limit=no_improvement_limit,
                destroy_fraction=destroy_fraction,
                warm_start=True,
                thesis_mode=solver_mode in {"thesis_mode", "mode_milp_only"} or phase_token == "phase3_two_stage",
                debug_mode=solver_mode == "debug_mode" or phase_token == "diagnostic",
                research_run=bool(research_run),
                allow_postsolve_repair=(
                    solver_mode == "debug_mode" and not bool(research_run)
                ),
                phase=phase_token,
                requested_phase_token=solver_mode,
                requested_phase=requested_phase,
                resolved_phase=phase_token,
                executed_phase=phase_token,
                diagnostic_mode=is_diagnostic_mode,
            )
            problem = ProblemBuilder().build_from_scenario(
                scenario,
                depot_id=depot_id,
                service_id=service_id,
                config=opt_config,
                planning_days=max(
                    int(((scenario.get("simulation_config") or {}).get("planning_days") or 1)),
                    1,
                ),
            )
            if weather_forecast is not None and weather_profile is not None:
                problem = apply_weather_policy_to_problem(
                    problem,
                    weather_forecast,
                    weather_profile,
                    random_seed=random_seed,
                )
            interactive_bev_utilization_policy = (
                _apply_interactive_bev_utilization_policy(
                    problem,
                    require_all_available_bevs=require_all_available_bevs,
                )
            )
            _configure_assignment_energy_diagnostics(
                problem,
                phase_token=phase_token,
                output_dir=output_dir,
                research_run=bool(research_run),
            )
            run_input_provenance = persist_run_input_provenance(
                run_dir=Path(output_dir),
                base_scenario=base_scenario,
                effective_scenario=scenario,
                prepared_input=prepared_payload,
                prepared_input_path=prepared_input_path,
                requested_prepared_input_id=requested_prepared_input_id,
                frontend_request={
                    "raw_frontend_body": raw_frontend_request_payload,
                    "interactive_runtime_controls": interactive_runtime_controls,
                    "interactive_operation_time_window_controls": (
                        interactive_operation_time_window_controls
                    ),
                    "interactive_terminal_soc_controls": interactive_terminal_soc_controls,
                    "interactive_bev_utilization_policy": (
                        interactive_bev_utilization_policy
                    ),
                    "interactive_research_contract": (
                        interactive_research_contract
                    ),
                    "scenario_id": scenario_id,
                    "prepared_input_id": prepared_input_id,
                    "requested_prepared_input_id": requested_prepared_input_id,
                    "mode": mode,
                    "solver_mode_effective": solver_mode,
                    "time_limit_seconds": time_limit_seconds,
                    "stage1_time_limit_seconds": stage1_time_limit_seconds,
                    "stage2_time_limit_seconds": stage2_time_limit_seconds,
                    "stage1_best_obj_stop_enabled": bool(
                        stage1_best_obj_stop_enabled
                    ),
                    "stage1_stage2_candidate_limit": (
                        opt_config.stage1_stage2_candidate_limit
                    ),
                    "stage1_composition_search_radius": (
                        opt_config.stage1_composition_search_radius
                    ),
                    "stage1_bev_frontier_enabled": (
                        opt_config.stage1_bev_frontier_enabled
                    ),
                    "stage1_bev_frontier_min_count": (
                        opt_config.stage1_bev_frontier_min_count
                    ),
                    "stage1_bev_frontier_max_count": (
                        opt_config.stage1_bev_frontier_max_count
                    ),
                    "stage1_bev_frontier_target_time_limit_seconds": (
                        opt_config.stage1_bev_frontier_target_time_limit_sec
                    ),
                    "integrated_actual_cost_objective": (
                        opt_config.integrated_actual_cost_objective
                    ),
                    "gurobi_threads": gurobi_threads,
                    "run_profile": run_profile,
                    "run_hourly_rolling": run_hourly_rolling,
                    "rolling_execution_minutes": rolling_execution_minutes,
                    "effective_rolling_controls": effective_rolling_controls,
                    "mip_gap": mip_gap,
                    "random_seed": random_seed,
                    "service_id": service_id,
                    "depot_id": depot_id,
                    "rebuild_dispatch": rebuild_dispatch,
                    "use_existing_duties": use_existing_duties,
                    "alns_iterations": alns_iterations,
                    "no_improvement_limit": no_improvement_limit,
                    "destroy_fraction": destroy_fraction,
                    "timestep_min": timestep_min,
                    "enable_weather_operation_policy": (
                        enable_weather_operation_policy
                    ),
                    "weather_proxy_forecast_path": weather_proxy_forecast_path,
                    "research_run": bool(research_run),
                },
                optimization_config=opt_config,
                canonical_problem=problem,
                code_provenance=run_git_state,
            )
            feasible_arc_count = sum(
                len(v) for v in (problem.feasible_connections or {}).values()
            )
            build_report = ScenarioBuildReport(
                scenario_id=scenario_id,
                depot_id=depot_id or "",
                service_id=service_id,
                trip_count=len(problem.trips),
                task_count=len(problem.trips),
                vehicle_count=len(problem.vehicles),
                charger_count=len(problem.chargers),
                travel_connection_count=feasible_arc_count,
            )
            store.set_field(scenario_id, "problemdata_build_audit", build_report.to_dict())

            price_slots = list(problem.price_slots or [])
            pv_slots = list(problem.pv_slots or [])

            job_store.update_job(
                job_id,
                status="running",
                progress=55,
                message=f"Running {engine_label} optimizer ({mode})...",
                metadata=_job_metadata(
                    scenario_id=scenario_id,
                    service_id=service_id,
                    depot_id=depot_id,
                    stage="solve",
                    mode=mode,
                    extra={
                        "prepared_input_id": prepared_input_id,
                        "problem_summary": {
                            "trips": len(problem.trips),
                            "vehicles": len(problem.vehicles),
                            "chargers": len(problem.chargers),
                            "feasible_arcs": feasible_arc_count,
                            "price_slots": len(price_slots),
                            "pv_slots": len(pv_slots),
                            "time_limit_seconds_requested": time_limit_seconds,
                        },
                    },
                ),
            )
            solve_started_at = time.perf_counter()
            engine_result = OptimizationEngine().solve(problem, opt_config)
            solve_elapsed = time.perf_counter() - solve_started_at
            run_git_state_after_solve = collect_git_state()
            git_state_unchanged_during_solve = _validate_git_state_after_solve(
                research_run=bool(research_run),
                before=run_git_state,
                after=run_git_state_after_solve,
            )
            engine_solver_metadata = dict(engine_result.solver_metadata or {})
            engine_solver_metadata.update(
                {
                    "git_sha": run_git_state.get("git_sha"),
                    "git_dirty": run_git_state.get("git_dirty"),
                    "git_state_available": bool(
                        run_git_state.get("git_state_available", False)
                    ),
                    "git_state_error": run_git_state.get("git_state_error"),
                    "git_provenance_captured_before_solve": True,
                    "git_sha_after_solve": run_git_state_after_solve.get(
                        "git_sha"
                    ),
                    "git_dirty_after_solve": run_git_state_after_solve.get(
                        "git_dirty"
                    ),
                    "git_state_unchanged_during_solve": (
                        git_state_unchanged_during_solve
                    ),
                    "research_submission_git_provenance_eligible": bool(
                        run_git_state.get("git_state_available", False)
                        and run_git_state.get("git_dirty") is False
                        and git_state_unchanged_during_solve
                        and runtime_git_attestation.get(
                            "runtime_git_state_matches_current"
                        )
                    ),
                    "bff_runtime_git_attestation": runtime_git_attestation,
                    "interactive_runtime_controls": interactive_runtime_controls,
                    "interactive_operation_time_window_controls": (
                        interactive_operation_time_window_controls
                    ),
                    "interactive_terminal_soc_controls": interactive_terminal_soc_controls,
                    "interactive_research_contract": (
                        interactive_research_contract
                    ),
                }
            )
            thesis_ablation_dir = Path(output_dir) / "thesis_ablation"
            try:
                thesis_ablation_payload = build_day_ahead_ablation_candidates(
                    problem=problem,
                    optimized_plan=engine_result.plan,
                    optimized_solver_status=str(engine_result.solver_status or ""),
                    primary_optimization_structure=str(
                        engine_solver_metadata.get("optimization_structure")
                        or ""
                    ),
                )
                _persist_json_outputs(
                    str(thesis_ablation_dir),
                    {
                        "day_ahead_method_candidates.json": (
                            thesis_ablation_payload
                        )
                    },
                )
                _write_csv_rows(
                    thesis_ablation_dir / "day_ahead_method_candidates.csv",
                    ablation_candidate_csv_rows(thesis_ablation_payload),
                    list(THESIS_ABLATION_CSV_COLUMNS),
                )
                engine_solver_metadata["thesis_ablation_candidates"] = {
                    "status": thesis_ablation_payload.get("status"),
                    "artifact": (
                        "thesis_ablation/day_ahead_method_candidates.json"
                    ),
                    "csv_artifact": (
                        "thesis_ablation/day_ahead_method_candidates.csv"
                    ),
                    "payload_sha256": thesis_ablation_payload.get(
                        "payload_sha256"
                    ),
                    "available_method_ids": thesis_ablation_payload.get(
                        "available_method_ids"
                    ),
                    "missing_method_ids": thesis_ablation_payload.get(
                        "missing_method_ids"
                    ),
                    "research_conclusion_eligible": False,
                }
            except Exception as exc:
                thesis_ablation_failure = {
                    "schema_version": (
                        "thesis_day_ahead_ablation_candidates_failure_v1"
                    ),
                    "status": "FAILED",
                    "error": f"{type(exc).__name__}: {exc}",
                    "research_conclusion_eligible": False,
                }
                _persist_json_outputs(
                    str(thesis_ablation_dir),
                    {
                        "day_ahead_method_candidates_failure.json": (
                            thesis_ablation_failure
                        )
                    },
                )
                engine_solver_metadata["thesis_ablation_candidates"] = {
                    "status": "FAILED",
                    "artifact": (
                        "thesis_ablation/"
                        "day_ahead_method_candidates_failure.json"
                    ),
                    "error": thesis_ablation_failure["error"],
                    "research_conclusion_eligible": False,
                }
            # Production results are immutable dataclasses.  Lightweight
            # compatibility/test adapters may expose the same contract as a
            # mutable object, so do not make provenance capture itself prevent
            # result persistence for those adapters.
            if is_dataclass(engine_result):
                engine_result = replace(
                    engine_result,
                    solver_metadata=engine_solver_metadata,
                )
            else:
                engine_result.solver_metadata = engine_solver_metadata
            candidate_rows = list(
                engine_solver_metadata.get(
                    "stage1_stage2_candidate_evaluation"
                )
                or []
            )
            if candidate_rows:
                candidate_payload = {
                    "selection_semantics": engine_solver_metadata.get(
                        "stage1_stage2_candidate_selection_semantics"
                    ),
                    "integrated_global_optimality_claimed": bool(
                        engine_solver_metadata.get(
                            "stage1_stage2_candidate_global_optimality_claimed",
                            False,
                        )
                    ),
                    "requested_candidate_limit": (
                        engine_solver_metadata.get(
                            "stage1_stage2_candidate_limit_requested"
                        )
                    ),
                    "composition_search_radius_requested": (
                        engine_solver_metadata.get(
                            "stage1_composition_search_radius_requested"
                        )
                    ),
                    "composition_search_runtime_seconds": (
                        engine_solver_metadata.get(
                            "stage1_composition_search_runtime_seconds"
                        )
                    ),
                    "used_powertrain_composition_search": dict(
                        engine_solver_metadata.get(
                            "stage1_used_powertrain_composition_search"
                        )
                        or {}
                    ),
                    "pool_solution_count": engine_solver_metadata.get(
                        "stage1_pool_solution_count"
                    ),
                    "distinct_candidate_count": engine_solver_metadata.get(
                        "stage1_distinct_candidate_count"
                    ),
                    "candidate_count_evaluated": (
                        engine_solver_metadata.get(
                            "stage1_stage2_candidate_count_evaluated"
                        )
                    ),
                    "feasible_candidate_count": (
                        engine_solver_metadata.get(
                            "stage1_stage2_feasible_candidate_count"
                        )
                    ),
                    "selected_candidate_index": (
                        engine_solver_metadata.get(
                            "stage1_stage2_selected_candidate_index"
                        )
                    ),
                    "selected_candidate_hash": (
                        engine_solver_metadata.get(
                            "stage1_stage2_selected_candidate_hash"
                        )
                    ),
                    "selected_canonical_actual_cost_jpy": (
                        engine_solver_metadata.get(
                            "stage1_stage2_selected_canonical_actual_cost_jpy"
                        )
                    ),
                    "primary_incumbent_objective_jpy": (
                        engine_solver_metadata.get(
                            "stage1_primary_incumbent_objective_jpy"
                        )
                    ),
                    "selected_candidate_relaxed_objective_jpy": (
                        engine_solver_metadata.get(
                            "stage1_selected_candidate_relaxed_objective_jpy"
                        )
                    ),
                    "primary_runtime_seconds": engine_solver_metadata.get(
                        "stage1_primary_runtime_seconds"
                    ),
                    "primary_search_time_limit_seconds": (
                        engine_solver_metadata.get(
                            "stage1_primary_search_time_limit_seconds"
                        )
                    ),
                    "enumeration_reserve_seconds": engine_solver_metadata.get(
                        "stage1_candidate_enumeration_reserve_seconds"
                    ),
                    "enumeration_runtime_seconds": engine_solver_metadata.get(
                        "stage1_candidate_enumeration_runtime_seconds"
                    ),
                    "powertrain_pattern_no_good_cut_count": (
                        engine_solver_metadata.get(
                            "stage1_candidate_powertrain_pattern_no_good_cut_count"
                        )
                    ),
                    "enumeration_events": list(
                        engine_solver_metadata.get(
                            "stage1_candidate_enumeration_events"
                        )
                        or []
                    ),
                    "candidates": candidate_rows,
                }
                candidate_json_path = (
                    Path(output_dir)
                    / "stage1_stage2_candidate_evaluation.json"
                )
                candidate_json_path.write_text(
                    json.dumps(
                        candidate_payload,
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                candidate_csv_path = (
                    Path(output_dir)
                    / "stage1_stage2_candidate_evaluation.csv"
                )
                _write_csv_rows(
                    candidate_csv_path,
                    candidate_rows,
                    [
                        "candidate_index",
                        "stage1_pool_solution_index",
                        "stage1_candidate_source",
                        "stage1_composition_target_used_bev",
                        "stage1_composition_target_used_ice",
                        "minimum_used_bev_count",
                        "candidate_hash",
                        "assignment_hash",
                        "stage1_relaxed_objective_jpy",
                        "stage2_exact_objective_jpy",
                        "stage2_actual_canonical_cost_jpy",
                        "feasible",
                        "stage2_feasible",
                        "canonical_evaluation_feasible",
                        "physical_validation_feasible",
                        "physical_validation_error_count",
                        "physical_validation_error_hash",
                        "iis_hash",
                        "used_bev",
                        "used_ice",
                        "bev_trips",
                        "ice_trips",
                        "bev_service_distance_km",
                        "bev_deadhead_distance_km",
                        "ice_service_distance_km",
                        "ice_deadhead_distance_km",
                        "bev_total_movement_distance_km",
                        "ice_total_movement_distance_km",
                        "total_movement_distance_km",
                        "bev_movement_distance_share",
                        "ice_fuel_l",
                        "runtime_sec",
                        "stage2_runtime_sec",
                        "stage2_time_limit_sec_effective",
                        "stage2_solver_status",
                    ],
                )
            composition_search_payload = dict(
                engine_solver_metadata.get(
                    "stage1_used_powertrain_composition_search"
                )
                or {}
            )
            if composition_search_payload:
                composition_search_json_path = (
                    Path(output_dir)
                    / "stage1_used_powertrain_composition_search.json"
                )
                composition_search_json_path.write_text(
                    json.dumps(
                        composition_search_payload,
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                _write_csv_rows(
                    Path(output_dir)
                    / "stage1_used_powertrain_composition_search.csv",
                    list(
                        composition_search_payload.get("target_records") or []
                    ),
                    [
                        "target_used_bev",
                        "minimum_used_bev_count",
                        "target_used_ice",
                        "delta_used_bev_from_primary",
                        "delta_used_ice_from_primary",
                        "target_total_used_vehicle_count",
                        "target_within_selected_inventory",
                        "search_status",
                        "solver_status",
                        "frontier_status",
                        "solution_count",
                        "best_bound",
                        "mip_gap_ratio",
                        "time_limit_sec",
                        "solver_runtime_sec",
                        "candidate_hash",
                        "frontier_target_candidate_physical_validation_feasible",
                        "frontier_resolution_source",
                        "frontier_resolution_candidate_hash",
                        "frontier_resolution_actual_used_bev",
                        "frontier_resolution_actual_used_ice",
                        "frontier_resolution_canonical_cost_jpy",
                        "frontier_resolution_candidate_source_target_used_bev",
                        "actual_used_bev",
                        "actual_used_ice",
                        "candidate_accepted_for_stage2_evaluation",
                        "final_disposition",
                    ],
                )
            bev_cost_frontier_payload = dict(
                engine_solver_metadata.get("bev_cost_frontier") or {}
            )
            if bev_cost_frontier_payload:
                frontier_rows = list(
                    bev_cost_frontier_payload.get("rows") or []
                )
                (Path(output_dir) / "bev_cost_frontier.json").write_text(
                    json.dumps(
                        bev_cost_frontier_payload,
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )
                _write_csv_rows(
                    Path(output_dir) / "bev_cost_frontier.csv",
                    frontier_rows,
                    [
                        "case",
                        "minimum_bev_count",
                        "minimum_used_bev_count",
                        "status",
                        "raw_solver_status",
                        "solution_count",
                        "target_stage1_relaxed_objective_jpy",
                        "stage1_relaxed_objective_jpy",
                        "stage2_actual_canonical_cost_jpy",
                        "target_candidate_hash",
                        "candidate_hash",
                        "frontier_resolution_source",
                        "frontier_resolution_candidate_source_target_used_bev",
                        "target_candidate_physical_validation_feasible",
                        "actual_used_bev",
                        "actual_used_ice",
                        "used_bev",
                        "used_ice",
                        "actual_total_used_vehicle_count",
                        "total_used_vehicles",
                        "total_used_vehicle_count_fixed",
                        "stage2_feasible",
                        "canonical_evaluation_feasible",
                        "physical_validation_feasible",
                        "physical_validation",
                        "final_disposition",
                        "best_bound",
                        "mip_gap_ratio",
                        "mip_gap",
                        "time_limit_sec",
                        "solver_runtime_sec",
                        "runtime_sec",
                        "bev_trip_count",
                        "bev_service_distance_km",
                        "bev_deadhead_distance_km",
                        "ice_service_distance_km",
                        "ice_deadhead_distance_km",
                        "bev_total_movement_distance_km",
                        "ice_total_movement_distance_km",
                        "bev_movement_distance_share",
                        "bev_distance_share",
                        "ice_fuel_l",
                        "grid_import_kwh",
                        "pv_to_bus_kwh",
                        "pv_to_bess_kwh",
                        "bess_to_bus_kwh",
                        "total_cost_jpy",
                        "cost_increase_percent",
                        "objective_is_actual_cost",
                        "rolling_24_of_24",
                    ],
                )
                frontier_markdown_lines = [
                    "# BEV cost frontier",
                    "",
                    (
                        "Constraint: `sum(used electric vehicles) >= K`; "
                        "ICE and total fleet size remain endogenous."
                    ),
                    (
                        "Rows use the lowest canonical-cost physically feasible "
                        "evaluated candidate with actual used BEV >= K; this is "
                        "a candidate-pool envelope, not a global-optimum claim."
                    ),
                    "",
                    "| K | status | used BEV | used ICE | actual cost JPY | physical |",
                    "|---:|---|---:|---:|---:|---|",
                ]
                for row in frontier_rows:
                    frontier_markdown_lines.append(
                        "| {k} | {status} | {bev} | {ice} | {cost} | {physical} |".format(
                            k=row.get("minimum_used_bev_count"),
                            status=row.get("status"),
                            bev=row.get("actual_used_bev"),
                            ice=row.get("actual_used_ice"),
                            cost=row.get("stage2_actual_canonical_cost_jpy"),
                            physical=row.get("physical_validation_feasible"),
                        )
                    )
                frontier_markdown_lines.extend(
                    [
                        "",
                        "Monotonicity violations: "
                        + str(
                            bev_cost_frontier_payload.get(
                                "monotonicity_violation_count", 0
                            )
                        ),
                        "",
                        str(
                            bev_cost_frontier_payload.get(
                                "monotonicity_semantics", ""
                            )
                        ),
                    ]
                )
                (Path(output_dir) / "bev_cost_frontier.md").write_text(
                    "\n".join(frontier_markdown_lines) + "\n",
                    encoding="utf-8",
                )
                _write_csv_rows(
                    Path(output_dir)
                    / "maximum_bev_feasibility_search.csv",
                    frontier_rows,
                    [
                        "minimum_used_bev_count",
                        "status",
                        "raw_solver_status",
                        "actual_used_bev",
                        "actual_used_ice",
                        "physical_validation_feasible",
                        "final_disposition",
                        "time_limit_sec",
                        "solver_runtime_sec",
                        "best_bound",
                        "mip_gap_ratio",
                    ],
                )
            elif (
                engine_solver_metadata.get("integrated_ev_utilization_mode")
                == "minimum_ice_fuel_lexicographic"
            ):
                duty_vehicle_map = engine_result.plan.duty_vehicle_map()
                used_vehicle_ids = set(duty_vehicle_map.values())
                vehicle_type_by_id = {
                    str(vehicle.vehicle_id): str(vehicle.vehicle_type).upper()
                    for vehicle in problem.vehicles
                }
                used_bev_ids = {
                    vehicle_id
                    for vehicle_id in used_vehicle_ids
                    if vehicle_type_by_id.get(vehicle_id)
                    in {"BEV", "PHEV", "FCEV"}
                }
                policy_row = {
                    "case": (
                        "cost_constrained_maximum_ev_utilization"
                        if engine_solver_metadata.get(
                            "integrated_actual_cost_upper_bound_jpy"
                        )
                        is not None
                        else "maximum_ev_utilization"
                    ),
                    "status": engine_result.solver_status,
                    "actual_used_bev": len(used_bev_ids),
                    "actual_used_ice": len(used_vehicle_ids - used_bev_ids),
                    "actual_total_used_vehicle_count": len(used_vehicle_ids),
                    "ice_fuel_l": engine_solver_metadata.get(
                        "integrated_primary_ice_fuel_l"
                    ),
                    "total_cost_jpy": (
                        engine_result.cost_breakdown or {}
                    ).get("total_cost"),
                    "actual_cost_upper_bound_jpy": (
                        engine_solver_metadata.get(
                            "integrated_actual_cost_upper_bound_jpy"
                        )
                    ),
                    "actual_cost_upper_bound_delta_ratio": (
                        engine_solver_metadata.get(
                            "integrated_actual_cost_upper_bound_delta_ratio"
                        )
                    ),
                    "actual_cost_upper_bound_verified": (
                        engine_solver_metadata.get(
                            "integrated_actual_cost_upper_bound_verified"
                        )
                    ),
                    "objective_is_actual_cost": False,
                    "primary_objective": "minimum_ice_fuel_l",
                    "secondary_objective": "minimum_canonical_actual_cost",
                    "physical_validation_feasible": engine_result.feasible,
                    "mip_gap_ratio": engine_solver_metadata.get(
                        "achieved_mip_gap"
                    ),
                    "solver_runtime_sec": engine_solver_metadata.get(
                        "runtime_sec"
                    ),
                }
                _write_csv_rows(
                    Path(output_dir)
                    / "maximum_bev_feasibility_search.csv",
                    [policy_row],
                    [
                        "case",
                        "status",
                        "actual_used_bev",
                        "actual_used_ice",
                        "actual_total_used_vehicle_count",
                        "ice_fuel_l",
                        "total_cost_jpy",
                        "actual_cost_upper_bound_jpy",
                        "actual_cost_upper_bound_delta_ratio",
                        "actual_cost_upper_bound_verified",
                        "objective_is_actual_cost",
                        "primary_objective",
                        "secondary_objective",
                        "physical_validation_feasible",
                        "mip_gap_ratio",
                        "solver_runtime_sec",
                    ],
                )
            graph_artifacts = _persist_canonical_graph_exports(
                scenario=scenario,
                problem=problem,
                engine_result=engine_result,
                scenario_id=scenario_id,
                output_dir=output_dir,
            )
            try:
                charging_flow_payload = _canonical_charging_output_payload(problem, engine_result)
                charging_summary_payload = dict(charging_flow_payload.get("summary") or {})
            except Exception as exc:
                charging_flow_payload = None
                charging_summary_payload = None
                charging_payload_warning = (
                    "Charging breakdown export was skipped because the canonical energy-flow payload "
                    f"could not be constructed: {exc}"
                )
            smeta = dict(engine_result.solver_metadata or {})
            smeta.setdefault("solve_time_sec", float(solve_elapsed))
            if charging_payload_warning:
                warnings_list = list(smeta.get("warnings") or [])
                warnings_list.append(charging_payload_warning)
                smeta["warnings"] = warnings_list
            _cb = dict(engine_result.cost_breakdown or {})
            # Alias keys so _cost_breakdown() can read both naming conventions
            _cb.setdefault("electricity_cost", _cb.get("electricity_cost_final", _cb.get("energy_cost", 0.0)))
            _cb.setdefault("fuel_cost", _cb.get("fuel_cost", 0.0))
            _cb.setdefault("demand_charge_cost", _cb.get("demand_cost", 0.0))
            _cb.setdefault("battery_degradation_cost", _cb.get("degradation_cost", 0.0))
            _cb.setdefault("emission_cost", _cb.get("co2_cost", 0.0))
            _cb.update(normalize_pv_energy_breakdown(_cb))
            achieved_mip_gap = smeta.get("achieved_mip_gap", smeta.get("final_gap"))
            result_payload = {
                "status": engine_result.solver_status,
                "objective_value": engine_result.objective_value,
                "solve_time_seconds": float(smeta.get("solve_time_sec", 0.0) or solve_elapsed),
                "mip_gap": None if achieved_mip_gap is None else float(achieved_mip_gap),
                "requested_mip_gap": float(smeta.get("requested_mip_gap", smeta.get("mip_gap", mip_gap)) or 0.0),
                "achieved_mip_gap": None if achieved_mip_gap is None else float(achieved_mip_gap),
                "assignment": {
                    k: list(v)
                    for k, v in engine_result.plan.vehicle_paths().items()
                },
                "unserved_tasks": list(engine_result.plan.unserved_trip_ids),
                "obj_breakdown": _cb,
                "solver_metadata": smeta,
            }
            sim_payload = None
            vehicle_type_by_id = {
                v.vehicle_id: v.vehicle_type for v in problem.vehicles
            }
            # Expose full new-system result for solver_result field
            _full_new_result = ResultSerializer.serialize_result(engine_result)
            if charging_summary_payload is not None:
                _full_new_result["charging_summary"] = charging_summary_payload
            result_payload["warnings"] = list(_full_new_result.get("warnings") or [])
            result_payload["infeasibility_reasons"] = list(
                _full_new_result.get("infeasibility_reasons") or []
            )
            result_payload["strict_coverage_precheck"] = dict(
                _full_new_result.get("strict_coverage_precheck") or {}
            )
        else:
            # ── LEGACY PATH: Should not reach here due to _normalize_solver_mode gating ──
            # This branch is kept temporarily for backward compatibility during migration.
            import warnings
            warnings.warn(
                f"Legacy solver path triggered for mode '{mode}' (normalized: '{solver_mode}'). "
                f"This path is deprecated and will be removed. "
                f"Please update to canonical modes: mode_milp_only, mode_alns_only, mode_hybrid, etc.",
                DeprecationWarning,
                stacklevel=2,
            )
            job_store.update_job(
                job_id,
                status="running",
                progress=25,
                message="Building ProblemData from scenario...",
                metadata=_job_metadata(
                    scenario_id=scenario_id,
                    service_id=service_id,
                    depot_id=depot_id,
                    stage="build_problemdata",
                    mode=mode,
                    extra={
                        "rebuild_dispatch": rebuild_dispatch,
                        "use_existing_duties": use_existing_duties,
                        "prepared_input_id": prepared_input_id,
                        "requested_prepared_input_id": requested_prepared_input_id,
                        "prepared_input_path": str(prepared_input_path),
                    },
                ),
            )
            data, build_report = build_problem_data_from_scenario(
                scenario,
                depot_id=depot_id,
                service_id=service_id,
                mode=solver_mode,
                use_existing_duties=use_existing_duties,
                analysis_scope=scenario.get("dispatch_scope") or store.get_dispatch_scope(scenario_id),
            )
            store.set_field(scenario_id, "problemdata_build_audit", build_report.to_dict())

            if (
                int(build_report.travel_connection_count or 0) <= 0
                and int(build_report.task_count or 0) > int(build_report.vehicle_count or 0)
                and not bool(getattr(data, "allow_partial_service", False))
            ):
                setattr(data, "allow_partial_service", True)
                setattr(data, "service_coverage_mode", "penalized")
                auto_relax_msg = (
                    "No travel connections generated while allow_partial_service is OFF. "
                    "Auto-relaxed allow_partial_service=True for this run to avoid hard infeasible stop. "
                    f"tasks={build_report.task_count}, vehicles={build_report.vehicle_count}, "
                    f"travel_connections={build_report.travel_connection_count}, "
                    f"prepared_input_id={prepared_input_id}, "
                    f"requested_prepared_input_id={requested_prepared_input_id or '-'}, "
                    f"prepared_input_path={prepared_input_path}."
                )
                if hasattr(build_report, "warnings"):
                    build_report.warnings.append(auto_relax_msg)

            price_slots = list(getattr(data, "electricity_prices", []) or [])
            pv_slots = list(getattr(data, "pv_profiles", []) or [])

            job_store.update_job(
                job_id,
                status="running",
                progress=55,
                message=f"Running optimizer ({mode})...",
                metadata=_job_metadata(
                    scenario_id=scenario_id,
                    service_id=service_id,
                    depot_id=depot_id,
                    stage="solve",
                    mode=mode,
                    extra={
                        "prepared_input_id": prepared_input_id,
                        "requested_prepared_input_id": requested_prepared_input_id,
                        "prepared_input_path": str(prepared_input_path),
                        "problem_summary": {
                            "trips": len(getattr(data, "tasks", []) or []),
                            "vehicles": len(getattr(data, "vehicles", []) or []),
                            "chargers": len(getattr(data, "chargers", []) or []),
                            "travel_connections": build_report.travel_connection_count,
                            "allow_partial_service_effective": bool(getattr(data, "allow_partial_service", False)),
                            "price_slots": len(price_slots),
                            "pv_slots": len(pv_slots),
                            "time_limit_seconds_requested": time_limit_seconds,
                            "time_limit_seconds_effective": min(time_limit_seconds, 86400),
                        },
                    },
                ),
            )
            solve_output = solve_problem_data(
                data,
                mode=solver_mode,
                time_limit_seconds=time_limit_seconds,
                mip_gap=mip_gap,
                random_seed=random_seed,
                output_dir=_scoped_output_dir(
                    root=str(output_paths.outputs_root()),
                    feed_context=feed_context,
                    scenario_id=scenario_id,
                    stage="optimization",
                    service_id=service_id,
                    depot_id=depot_id,
                ),
                alns_iterations=alns_iterations,
                no_improvement_limit=no_improvement_limit,
                destroy_fraction=destroy_fraction,
            )
            result_payload = serialize_milp_result(solve_output["result"])
            sim_payload = (
                serialize_simulation_result(solve_output["sim_result"])
                if solve_output.get("sim_result") is not None
                else None
            )
            vehicle_type_by_id = {
                vehicle.vehicle_id: vehicle.vehicle_type
                for vehicle in data.vehicles
            }
            _full_new_result = None
            graph_artifacts = {"enabled": False, "diagram_count": 0}
        vehicle_count_by_type: Dict[str, int] = {}
        trip_count_by_type: Dict[str, int] = {}
        for vehicle_id, task_ids in (result_payload.get("assignment") or {}).items():
            if not task_ids:
                continue
            vehicle_type = str(vehicle_type_by_id.get(vehicle_id) or "UNKNOWN")
            vehicle_count_by_type[vehicle_type] = vehicle_count_by_type.get(vehicle_type, 0) + 1
            trip_count_by_type[vehicle_type] = trip_count_by_type.get(vehicle_type, 0) + len(task_ids)
        available_vehicle_count_by_type: Dict[str, int] = {}
        unused_available_vehicle_ids_by_type: Dict[str, List[str]] = {}
        used_vehicle_ids = {
            str(vehicle_id)
            for vehicle_id, task_ids in (result_payload.get("assignment") or {}).items()
            if task_ids
        }
        for vehicle in problem.vehicles:
            if not bool(getattr(vehicle, "available", True)):
                continue
            vehicle_id = str(getattr(vehicle, "vehicle_id", "") or "")
            vehicle_type = str(getattr(vehicle, "vehicle_type", "UNKNOWN") or "UNKNOWN").upper()
            available_vehicle_count_by_type[vehicle_type] = available_vehicle_count_by_type.get(vehicle_type, 0) + 1
            if vehicle_id and vehicle_id not in used_vehicle_ids:
                unused_available_vehicle_ids_by_type.setdefault(vehicle_type, []).append(vehicle_id)
        for vehicle_ids in unused_available_vehicle_ids_by_type.values():
            vehicle_ids.sort()
        electric_types = {"BEV", "PHEV", "FCEV"}
        ev_available_count = sum(
            count for vehicle_type, count in available_vehicle_count_by_type.items() if vehicle_type.upper() in electric_types
        )
        ev_used_count = sum(
            count for vehicle_type, count in vehicle_count_by_type.items() if vehicle_type.upper() in electric_types
        )
        ice_available_count = sum(
            count for vehicle_type, count in available_vehicle_count_by_type.items() if vehicle_type.upper() not in electric_types
        )
        ice_used_count = sum(
            count for vehicle_type, count in vehicle_count_by_type.items() if vehicle_type.upper() not in electric_types
        )
        objective_mode = str(
            (
                ((scenario.get("scenario_overlay") or {}).get("solver_config") or {}).get("objective_mode")
                or (scenario.get("simulation_config") or {}).get("objective_mode")
                or "total_cost"
            )
        )
        prepared_scope_summary = dict(scenario.get("prepared_scope_summary") or {})
        prepared_scenario_hash = str(
            prepared_scope_summary.get("scenario_hash")
            or (prepared_payload.get("scenario_hash") if isinstance(prepared_payload, dict) else "")
            or ""
        )
        prepared_scope_hash = str(
            prepared_scope_summary.get("scope_hash")
            or (prepared_payload.get("scope_hash") if isinstance(prepared_payload, dict) else "")
            or ""
        )
        prepared_scope_audit = dict(
            prepared_scope_summary.get("prepared_scope_audit")
            or (prepared_payload.get("prepared_scope_audit") if isinstance(prepared_payload, dict) else {})
            or {}
        )
        if prepared_scope_audit:
            result_payload["prepared_scope_audit"] = prepared_scope_audit
            if isinstance(_full_new_result, dict):
                _full_new_result["prepared_scope_audit"] = prepared_scope_audit
        service_coverage_mode = str(getattr(problem.scenario, "service_coverage_mode", "strict") or "strict")
        fixed_route_band_mode = bool((problem.metadata or {}).get("fixed_route_band_mode", False))
        daily_fragment_limit = int((problem.metadata or {}).get("daily_fragment_limit") or 1)
        available_vehicle_count_total = sum(
            1 for vehicle in problem.vehicles if bool(getattr(vehicle, "available", True))
        )
        unused_available_vehicle_ids = list(engine_result.plan.unused_available_vehicle_ids(problem))
        # Preserve the post-solve wall-clock fallback already inserted into
        # ``result_payload`` instead of rebuilding from stale engine metadata.
        solver_metadata = dict(result_payload.get("solver_metadata") or {})
        phase4_seed_audit_for_summary = dict(
            solver_metadata.get("phase4_phase3_seed_audit") or {}
        )
        stage1_primary_candidate_composition = dict(
            phase4_seed_audit_for_summary.get(
                "seed_stage1_primary_candidate_used_powertrain_composition"
            )
            or {}
        )
        strict_coverage_precheck = dict(
            solver_metadata.get("strict_coverage_precheck")
            or (result_payload.get("strict_coverage_precheck") if isinstance(result_payload, dict) else {})
            or (_full_new_result.get("strict_coverage_precheck") if isinstance(_full_new_result, dict) else {})
            or {}
        )
        result_warnings = list(
            (_full_new_result.get("warnings") if isinstance(_full_new_result, dict) else None)
            or result_payload.get("warnings")
            or []
        )
        result_infeasibility_reasons = list(
            (_full_new_result.get("infeasibility_reasons") if isinstance(_full_new_result, dict) else None)
            or result_payload.get("infeasibility_reasons")
            or []
        )
        startup_rejected_raw = (
            solver_metadata.get("startup_rejected_vehicle_ids_by_duty")
            or (engine_result.plan.metadata or {}).get("startup_rejected_vehicle_ids_by_duty")
            or {}
        )
        startup_rejected_vehicle_ids_by_duty: Dict[str, List[str]] = {}
        if isinstance(startup_rejected_raw, dict):
            for duty_id, vehicle_ids in startup_rejected_raw.items():
                normalized_vehicle_ids = sorted(
                    {
                        str(vehicle_id).strip()
                        for vehicle_id in list(vehicle_ids or [])
                        if str(vehicle_id).strip()
                    }
                )
                if normalized_vehicle_ids:
                    startup_rejected_vehicle_ids_by_duty[str(duty_id)] = normalized_vehicle_ids
        startup_rejected_duty_count = len(startup_rejected_vehicle_ids_by_duty)
        startup_rejected_vehicle_candidate_count = sum(
            len(vehicle_ids) for vehicle_ids in startup_rejected_vehicle_ids_by_duty.values()
        )
        startup_rejected_vehicle_count = len(
            {
                vehicle_id
                for vehicle_ids in startup_rejected_vehicle_ids_by_duty.values()
                for vehicle_id in vehicle_ids
            }
        )
        legacy_trip_count_served = sum(
            len(task_ids)
            for task_ids in (result_payload.get("assignment") or {}).values()
        )
        legacy_trip_count_unserved = len(result_payload.get("unserved_tasks") or [])
        trip_count_served, trip_count_unserved = _canonical_trip_counts(
            _full_new_result if isinstance(_full_new_result, dict) else None,
            fallback_served=legacy_trip_count_served,
            fallback_unserved=legacy_trip_count_unserved,
        )
        canonical_feasible = (
            bool(_full_new_result.get("feasible"))
            if isinstance(_full_new_result, dict) and "feasible" in _full_new_result
            else bool(engine_result.feasible)
        )
        solution_validity = _solution_validity_payload(
            solver_status=result_payload["status"],
            feasible=canonical_feasible,
            trip_count_unserved=trip_count_unserved,
            infeasibility_reasons=result_infeasibility_reasons,
            solver_metadata=solver_metadata,
        )
        if isinstance(_full_new_result, dict):
            _full_new_result["solution_validity"] = solution_validity
        result_payload["solution_validity"] = solution_validity
        weather_policy_payload = _weather_policy_payload_from_problem_metadata(
            dict(problem.metadata or {})
        )
        solver_settings = _solver_settings_payload(
            time_limit_seconds_requested=time_limit_seconds,
            mip_gap_requested=mip_gap,
            solver_metadata=solver_metadata,
            random_seed_requested=random_seed,
            stage1_time_limit_seconds_requested=(
                stage1_time_limit_seconds
            ),
            stage2_time_limit_seconds_requested=(
                stage2_time_limit_seconds
            ),
        )
        result_cost_breakdown = _cost_breakdown(result_payload, sim_payload)
        accounting_summary_for_result = dict((graph_artifacts or {}).get("accounting_summary") or {})
        objective_value_for_result = result_payload.get("objective_value")
        if (
            bool(result_cost_breakdown.get("objective_is_actual_cost", False))
            and accounting_summary_for_result.get("reported_total_cost_jpy") is not None
        ):
            objective_value_for_result = float(accounting_summary_for_result.get("reported_total_cost_jpy") or 0.0)
            result_payload["objective_value"] = objective_value_for_result
            result_cost_breakdown["total_cost"] = objective_value_for_result
            result_cost_breakdown["objective_value"] = objective_value_for_result
            if isinstance(result_payload.get("obj_breakdown"), dict):
                result_payload["obj_breakdown"]["total_cost"] = objective_value_for_result
                result_payload["obj_breakdown"]["objective_value"] = objective_value_for_result
            if isinstance(_full_new_result, dict):
                _full_new_result["objective_value"] = objective_value_for_result
                if isinstance(_full_new_result.get("cost_breakdown"), dict):
                    _full_new_result["cost_breakdown"]["total_cost"] = objective_value_for_result
                    _full_new_result["cost_breakdown"]["objective_value"] = objective_value_for_result

        optimization_result: Dict[str, Any] = {
            "scenario_id": scenario_id,
            "feed_context": feed_context,
            "scope": {"serviceId": service_id, "depotId": depot_id},
            "prepared_input_id": prepared_input_id,
            "prepared_scope_summary": prepared_scope_summary,
            "scenario_hash": prepared_scenario_hash,
            "scope_hash": prepared_scope_hash,
            "solver_status": result_payload["status"],
            "mode": mode,
            "solver_mode": solver_mode,
            "objective_mode": objective_mode,
            "objective_value": objective_value_for_result,
            "solve_time_seconds": result_payload.get("solve_time_seconds", 0.0),
            "mip_gap": result_payload.get("mip_gap"),
            "solver_metadata": solver_metadata,
            "solver_settings": solver_settings,
            "warnings": result_warnings,
            "infeasibility_reasons": result_infeasibility_reasons,
            "strict_coverage_precheck": strict_coverage_precheck,
            "prepared_scope_audit": prepared_scope_audit,
            "solution_validity": solution_validity,
            "result_class": solution_validity.get("result_class"),
            "research_kpi_eligible": bool(solution_validity.get("research_kpi_eligible", False)),
            "electricity_cost_basis": str(
                (sim_payload or {}).get("electricity_cost_basis") or "provisional_drive"
            ),
            "cost_breakdown": result_cost_breakdown,
            "dispatch_report": scenario.get("graph") or store.get_field(scenario_id, "graph") or {},
            "build_report": build_report.to_dict(),
            "summary": {
                "same_day_depot_cycles_enabled": bool(
                    dict(engine_result.solver_metadata or {}).get(
                        "same_day_depot_cycles_enabled",
                        getattr(problem.scenario, "allow_same_day_depot_cycles", True),
                    )
                ),
                "service_coverage_mode": service_coverage_mode,
                "fixed_route_band_mode": fixed_route_band_mode,
                "daily_fragment_limit": daily_fragment_limit,
                "prepared_input_id": prepared_input_id,
                "scenario_hash": prepared_scenario_hash,
                "scope_hash": prepared_scope_hash,
                "strict_coverage_precheck": strict_coverage_precheck,
                "prepared_scope_audit": prepared_scope_audit,
                "available_vehicle_count_total": available_vehicle_count_total,
                "unused_available_vehicle_ids": unused_available_vehicle_ids,
                "startup_infeasible_assignment_count": int(
                    solver_metadata.get("startup_infeasible_assignment_count")
                    or (engine_result.plan.metadata or {}).get("startup_infeasible_assignment_count")
                    or 0
                ),
                "startup_infeasible_trip_ids": list(
                    solver_metadata.get("startup_infeasible_trip_ids")
                    or (engine_result.plan.metadata or {}).get("startup_infeasible_trip_ids")
                    or []
                ),
                "startup_infeasible_vehicle_ids": list(
                    solver_metadata.get("startup_infeasible_vehicle_ids")
                    or (engine_result.plan.metadata or {}).get("startup_infeasible_vehicle_ids")
                    or []
                ),
                "startup_rejected_duty_count": int(startup_rejected_duty_count),
                "startup_rejected_vehicle_candidate_count": int(startup_rejected_vehicle_candidate_count),
                "startup_rejected_vehicle_count": int(startup_rejected_vehicle_count),
                "startup_rejected_vehicle_ids_by_duty": startup_rejected_vehicle_ids_by_duty,
                "max_depot_cycles_per_vehicle_per_day": int(
                    solver_metadata.get(
                        "max_depot_cycles_per_vehicle_per_day",
                        getattr(problem.scenario, "max_depot_cycles_per_vehicle_per_day", 1),
                    )
                    or 1
                ),
                "vehicle_count_used": sum(
                    1
                    for _vehicle_id, task_ids in (
                        result_payload.get("assignment") or {}
                    ).items()
                    if task_ids
                ),
                "vehicle_count_by_type": vehicle_count_by_type,
                "available_vehicle_count_by_type": available_vehicle_count_by_type,
                "used_vehicle_count_by_type": vehicle_count_by_type,
                "powertrain_composition_semantics": (
                    "final_published_solver_incumbent"
                ),
                "final_used_powertrain_composition": {
                    "used_bev": ev_used_count,
                    "used_ice": ice_used_count,
                },
                "stage1_primary_candidate_used_powertrain_composition": (
                    stage1_primary_candidate_composition
                ),
                "stage1_primary_candidate_composition_semantics": (
                    "phase3_initial_candidate_not_final_phase3_selection_"
                    "and_not_phase4_result"
                    if stage1_primary_candidate_composition
                    else None
                ),
                "unused_available_vehicle_ids_by_type": unused_available_vehicle_ids_by_type,
                "ev_available_count": ev_available_count,
                "ev_used_count": ev_used_count,
                "ev_unused_count": max(ev_available_count - ev_used_count, 0),
                "ice_available_count": ice_available_count,
                "ice_used_count": ice_used_count,
                "ice_unused_count": max(ice_available_count - ice_used_count, 0),
                "trip_count_by_type": trip_count_by_type,
                "trip_count_served": trip_count_served,
                "trip_count_unserved": trip_count_unserved,
                "coverage_rank_primary": trip_count_unserved,
                "solution_validity": solution_validity,
                "secondary_objective_value": result_payload.get("secondary_objective_value"),
                "vehicle_fragment_counts": dict(
                    engine_result.plan.vehicle_fragment_counts()
                ),
                "vehicles_with_multiple_fragments": list(
                    engine_result.plan.vehicles_with_multiple_fragments()
                ),
                "max_fragments_observed": int(engine_result.plan.max_fragments_observed()),
            },
            "solver_result": result_payload,
            "canonical_solver_result": _full_new_result,
            "canonical_problem_summary": {
                "trip_count": build_report.task_count,
                "vehicle_count": build_report.vehicle_count,
                "available_vehicle_count_total": available_vehicle_count_total,
                "charger_count": build_report.charger_count,
                "price_slot_count": len(price_slots),
                "pv_slot_count": len(pv_slots),
            },
            "graph_artifacts": graph_artifacts,
        }
        if weather_policy_payload is not None:
            optimization_result["weather_policy"] = weather_policy_payload
        optimization_result[
            "bev_utilization_policy"
        ] = interactive_bev_utilization_policy
        optimization_result[
            "formal_research_contract"
        ] = interactive_research_contract
        if charging_summary_payload is not None:
            optimization_result["charging_summary"] = charging_summary_payload
        if charging_payload_warning:
            optimization_result["charging_summary_warning"] = charging_payload_warning
        if sim_payload is not None:
            optimization_result["simulation_summary"] = sim_payload
        if isinstance(optimization_result.get("cost_breakdown"), dict):
            optimization_result["cost_breakdown"]["evaluation_feasible"] = (
                1.0 if bool(solution_validity.get("validated_feasible")) else 0.0
            )
        if isinstance(result_payload.get("obj_breakdown"), dict):
            result_payload["obj_breakdown"]["evaluation_feasible"] = (
                1.0 if bool(solution_validity.get("validated_feasible")) else 0.0
            )
        if isinstance(_full_new_result, dict) and isinstance(_full_new_result.get("cost_breakdown"), dict):
            _full_new_result["cost_breakdown"]["evaluation_feasible"] = (
                1.0 if bool(solution_validity.get("validated_feasible")) else 0.0
            )
        _apply_invalid_result_kpi_gate(
            optimization_result,
            _full_new_result if isinstance(_full_new_result, dict) else None,
        )
        if not bool(solution_validity.get("validated_feasible")) and isinstance(
            optimization_result.get("charging_summary"), dict
        ):
            charging_summary_payload = dict(optimization_result["charging_summary"])

        optimization_audit = {
            "scenario_id": scenario_id,
            "feed_context": feed_context,
            "depot_id": depot_id,
            "service_id": service_id,
            "prepared_input_id": prepared_input_id,
            "prepared_scope_summary": prepared_scope_summary,
            "scenario_hash": prepared_scenario_hash,
            "scope_hash": prepared_scope_hash,
            "run_input_provenance": run_input_provenance,
            "case_type": scenario.get("experiment_case_type"),
            "input_counts": {
                "vehicles": build_report.vehicle_count,
                "tasks": build_report.task_count,
                "travel_connections": build_report.travel_connection_count,
            },
            "output_counts": {
                "assigned_vehicles": optimization_result["summary"][
                    "vehicle_count_used"
                ],
                "served_trips": optimization_result["summary"]["trip_count_served"],
                "unserved_trips": optimization_result["summary"]["trip_count_unserved"],
            },
            "warnings": list(
                dict.fromkeys(
                    [
                        *list(build_report.warnings or []),
                        *result_warnings,
                        *list(prepared_scope_audit.get("warnings") or []),
                    ]
                )
            ),
            "errors": list(
                dict.fromkeys(
                    [
                        *list(build_report.errors or []),
                        *result_infeasibility_reasons,
                    ]
                )
            ),
            "solver_mode": mode,
            "solver_mode_effective": solver_mode,
            "service_coverage_mode": service_coverage_mode,
            "fixed_route_band_mode": fixed_route_band_mode,
            "daily_fragment_limit": daily_fragment_limit,
            "strict_coverage_precheck": strict_coverage_precheck,
            "prepared_scope_audit": prepared_scope_audit,
            "available_vehicle_count_total": available_vehicle_count_total,
            "unused_available_vehicle_ids": unused_available_vehicle_ids,
            "startup_rejected_duty_count": int(startup_rejected_duty_count),
            "startup_rejected_vehicle_candidate_count": int(startup_rejected_vehicle_candidate_count),
            "startup_rejected_vehicle_count": int(startup_rejected_vehicle_count),
            "startup_rejected_vehicle_ids_by_duty": startup_rejected_vehicle_ids_by_duty,
            "time_limit": time_limit_seconds,
            "mip_gap": mip_gap,
            "solver_settings": solver_settings,
            **solver_settings,
            "random_seed": random_seed,
            "gurobi_seed": random_seed,
            "alns_iterations": alns_iterations,
            "no_improvement_limit": no_improvement_limit,
            "destroy_fraction": destroy_fraction,
            "git_sha": run_git_state.get("git_sha"),
            "git_dirty": run_git_state.get("git_dirty"),
            "git_state_available": bool(
                run_git_state.get("git_state_available", False)
            ),
            "git_state_error": run_git_state.get("git_state_error"),
            "bff_runtime_git_attestation": runtime_git_attestation,
            "source_snapshot": store.get_field(scenario_id, "source_snapshot"),
            "output_dir": output_dir,
            "executed_at": datetime.now(timezone.utc).isoformat(),
        }
        if weather_policy_payload is not None:
            optimization_audit["weather_policy"] = weather_policy_payload.get("audit") or {}
        optimization_audit[
            "bev_utilization_policy"
        ] = interactive_bev_utilization_policy
        optimization_audit[
            "formal_research_contract"
        ] = interactive_research_contract
        optimization_result["audit"] = optimization_audit

        store.set_field(scenario_id, "optimization_result", optimization_result)
        store.set_field(scenario_id, "optimization_audit", optimization_audit)
        _persist_json_outputs(
            output_dir,
            {
                "optimization_result.json": optimization_result,
                "optimization_audit.json": optimization_audit,
                "assignment_validation_diagnostics.json": {
                    "research_run": bool(
                        (optimization_result.get("solver_metadata") or {}).get("research_run", False)
                    ),
                    "diagnostics": list(
                        (optimization_result.get("solver_metadata") or {}).get(
                            "assignment_validation_diagnostics", []
                        )
                        or []
                    ),
                },
            },
        )
        optimization_result["run_profile"] = run_profile
        optimization_result["rolling_execution_minutes"] = (
            rolling_execution_minutes if run_hourly_rolling else None
        )
        optimization_audit["run_profile"] = run_profile
        optimization_audit["effective_rolling_controls"] = effective_rolling_controls
        result_solver_metadata = dict(
            optimization_result.get("solver_metadata") or {}
        )
        result_solver_metadata.update(
            {
                "run_profile": run_profile,
                "rolling_horizon_policy": (
                    "remaining_day_charging_only_fixed_assignment"
                    if run_hourly_rolling
                    else None
                ),
                "rolling_execution_minutes": (
                    rolling_execution_minutes if run_hourly_rolling else None
                ),
            }
        )
        optimization_result["solver_metadata"] = result_solver_metadata

        # First persist only the completed day-ahead contract. Human-facing
        # reporting is intentionally deferred until rolling and its independent
        # acceptance audit have finished.
        _persist_rich_run_outputs(
            run_dir=Path(output_dir),
            scenario=scenario,
            optimization_result=optimization_result,
            optimization_audit=optimization_audit,
            result_payload=result_payload,
            sim_payload=sim_payload,
            canonical_solver_result=_full_new_result,
            canonical_problem=problem,
            graph_source_dir=Path(output_dir) / "graph",
            charging_summary=charging_summary_payload,
            charging_flow_payload=charging_flow_payload,
            finalize_reporting=False,
        )
        rolling_technical_failure: Optional[RollingChainExecutionError] = None
        if run_hourly_rolling:
            day_ahead_feasible = bool(
                getattr(engine_result, "feasible", False)
                and not list(
                    getattr(getattr(engine_result, "plan", None), "unserved_trip_ids", ())
                    or ()
                )
            )
            if not day_ahead_feasible:
                rolling_technical_failure = RollingChainExecutionError(
                    "Hourly rolling was not started because the day-ahead "
                    "canonical result is not fully feasible"
                )
                failure_payload = {
                    "status": "not_started",
                    "reason": "day_ahead_not_fully_feasible",
                    "run_profile": run_profile,
                }
                _persist_json_outputs(
                    str(Path(output_dir) / "rolling_hourly_chain"),
                    {"rolling_execution_failure.json": failure_payload},
                )
            else:
                job_store.update_job(
                    job_id,
                    status="running",
                    progress=70,
                    message="Day-ahead solved; preparing hourly rolling contract...",
                    metadata=_job_metadata(
                        scenario_id=scenario_id,
                        service_id=service_id,
                        depot_id=depot_id,
                        stage="day_ahead_solved",
                        mode=mode,
                        extra={
                            "run_dir": output_dir,
                            "prepared_input_id": prepared_input_id,
                            "run_profile": run_profile,
                        },
                    ),
                )
                try:
                    rolling_input_audit = persist_frontend_day_ahead_rolling_contract(
                        run_dir=Path(output_dir),
                        scenario=scenario,
                        problem=problem,
                        prepared_input_path=prepared_input_path,
                        scenario_id=scenario_id,
                        prepared_input_id=prepared_input_id,
                        service_id=service_id,
                        git_state=run_git_state,
                    )
                    if (
                        rolling_input_audit.get("calendar_validation_status")
                        == "ERROR"
                    ):
                        result_solver_metadata = dict(
                            optimization_result.get("solver_metadata") or {}
                        )
                        failed_checks = list(
                            result_solver_metadata.get(
                                "research_acceptance_failed_checks"
                            )
                            or ()
                        )
                        failed_checks.append("service_calendar_contract")
                        result_solver_metadata[
                            "research_acceptance_failed_checks"
                        ] = sorted(set(map(str, failed_checks)))
                        result_solver_metadata["research_run_accepted"] = False
                        optimization_result[
                            "solver_metadata"
                        ] = result_solver_metadata
                    job_store.update_job(
                        job_id,
                        status="running",
                        progress=75,
                        message="Running 60-minute rolling chain...",
                        metadata=_job_metadata(
                            scenario_id=scenario_id,
                            service_id=service_id,
                            depot_id=depot_id,
                            stage="rolling_running",
                            mode=mode,
                            extra={
                                "run_dir": output_dir,
                                "rolling_execution_minutes": 60,
                            },
                        ),
                    )
                    rolling_result = execute_frontend_rolling_chain(
                        run_dir=Path(output_dir),
                        problem=problem,
                        scenario_id=scenario_id,
                        prepared_input_id=prepared_input_id,
                        service_id=service_id,
                        depot_id=str(depot_id),
                        execution_minutes=rolling_execution_minutes,
                        # Rolling is a charging-only conditional solve. Preserve
                        # an explicit user override, otherwise use the audited
                        # service default rather than the long Stage 1 limit.
                        time_limit_sec=int(stage2_time_limit_seconds or 30),
                        mip_gap=mip_gap,
                        random_seed=random_seed,
                        gurobi_threads=gurobi_threads,
                    )
                    optimization_result["rolling_execution"] = {
                        "status": rolling_result.status,
                        "chain_summary_path": rolling_result.chain_summary_path,
                        "chain_accepted": rolling_result.chain_accepted,
                        "technical_failure_reasons": list(
                            rolling_result.technical_failure_reasons
                        ),
                    }
                    if rolling_result.technical_failure_reasons:
                        rolling_technical_failure = RollingChainExecutionError(
                            "Hourly rolling chain failed required execution "
                            "checks: "
                            + ", ".join(rolling_result.technical_failure_reasons)
                        )
                    else:
                        final_rolling_evidence = (
                            finalize_frontend_rolling_evidence(
                                run_dir=Path(output_dir),
                                scenario=scenario,
                                problem=problem,
                                optimization_result=optimization_result,
                            )
                        )
                        optimization_audit["final_rolling_evidence"] = {
                            "physical_schedule_validation_status": dict(
                                final_rolling_evidence.get(
                                    "physical_schedule_validation"
                                )
                                or {}
                            ).get("status"),
                            "final_accounting_source": optimization_result.get(
                                "final_accounting_source"
                            ),
                            "final_accounting_total_cost_jpy": (
                                optimization_result.get(
                                    "final_accounting_total_cost_jpy"
                                )
                            ),
                            "comparison_control_hash": dict(
                                final_rolling_evidence.get(
                                    "comparison_case_manifest"
                                )
                                or {}
                            ).get("comparison_control_hash"),
                        }
                except Exception as exc:
                    failure_payload = {
                        "status": "failed",
                        "reason": "rolling_service_exception",
                        "error": f"{type(exc).__name__}: {exc}",
                        "traceback": traceback.format_exc(),
                        "run_profile": run_profile,
                    }
                    _persist_json_outputs(
                        str(Path(output_dir) / "rolling_hourly_chain"),
                        {"rolling_execution_failure.json": failure_payload},
                    )
                    rolling_technical_failure = RollingChainExecutionError(
                        f"Hourly rolling service failed: {type(exc).__name__}: {exc}"
                    )
        else:
            optimization_result["rolling_execution"] = {
                "status": "not_executed",
                "reason": "explicit_day_ahead_exploratory_profile",
            }

        result_claim_classification = _apply_result_claim_classification(
            optimization_result
        )
        optimization_audit[
            "result_claim_classification"
        ] = result_claim_classification
        job_store.update_job(
            job_id,
            status="running",
            progress=95,
            message="Validating rolling evidence and finalizing reports...",
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="rolling_validating",
                mode=mode,
                extra={"run_dir": output_dir, "run_profile": run_profile},
            ),
        )
        reporting_finalizer_result = _persist_rich_run_outputs(
            run_dir=Path(output_dir),
            scenario=scenario,
            optimization_result=optimization_result,
            optimization_audit=optimization_audit,
            result_payload=result_payload,
            sim_payload=sim_payload,
            canonical_solver_result=_full_new_result,
            canonical_problem=problem,
            graph_source_dir=Path(output_dir) / "graph",
            charging_summary=charging_summary_payload,
            charging_flow_payload=charging_flow_payload,
            # Preserve the primary rolling failure. Final reporting requires a
            # complete eligible executed day and would otherwise replace the
            # actionable step error with a secondary accounting exception.
            finalize_reporting=_should_finalize_reporting_after_rolling(
                rolling_technical_failure
            ),
        )
        if (Path(output_dir) / "input_audit.json").is_file():
            refresh_frontend_rolling_manifest(
                run_dir=Path(output_dir),
                run_state=(
                    "complete"
                    if rolling_technical_failure is None
                    else "rolling_execution_failed"
                ),
            )
        if rolling_technical_failure is None:
            reporting_finalizer_result = (
                _enforce_frontend_run_artifact_contract(
                    run_dir=Path(output_dir),
                    optimization_result=optimization_result,
                    optimization_audit=optimization_audit,
                    reporting_finalizer_result=reporting_finalizer_result,
                    research_run=research_run,
                    require_rolling=run_hourly_rolling,
                )
            )
        if reporting_finalizer_result is not None:
            store.set_field(scenario_id, "optimization_result", optimization_result)
            store.set_field(scenario_id, "optimization_audit", optimization_audit)
        if rolling_technical_failure is not None:
            raise rolling_technical_failure
        is_fallback = bool(solution_validity.get("result_class") in {"baseline_fallback", "postsolve_infeasible", "postsolve_repaired", "repaired_heuristic", "debug_result"})
        is_feasible_candidate = (
            dict(
                optimization_result.get("result_claim_classification") or {}
            ).get("label")
            == "feasible_candidate"
        )
        final_status = (
            "optimized"
            if not is_fallback and not is_feasible_candidate
            else "optimized_provisional"
        )
        store.update_scenario(scenario_id, status=final_status)
        if is_feasible_candidate:
            job_message = _feasible_candidate_job_message(
                dict(
                    optimization_result.get("result_claim_classification")
                    or {}
                )
            )
        elif not is_fallback:
            job_message = "Optimization complete."
        else:
            job_message = (
                "Optimization complete "
                f"({solution_validity.get('status_reason', 'provisional')})."
            )
        job_store.update_job(
            job_id,
            status="completed",
            progress=100,
            message=job_message,
            result_key="optimization_result",
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="completed",
                mode=mode,
                extra={
                    "objective_value": optimization_result.get("objective_value"),
                    "solver_status": optimization_result.get("solver_status"),
                    "feed_context": feed_context,
                    "run_dir": output_dir,
                    "reporting_finalizer_status": dict(
                        (optimization_result.get("graph_artifacts") or {}).get("reporting_finalizer")
                        or {}
                    ).get("status"),
                    "artifact_completeness_status": dict(
                        reporting_finalizer_result or {}
                    ).get("artifact_completeness_status"),
                    "artifact_completeness_artifact": dict(
                        reporting_finalizer_result or {}
                    ).get("artifact_completeness_artifact"),
                    "required_artifact_count": dict(
                        reporting_finalizer_result or {}
                    ).get("required_artifact_count"),
                    "verified_artifact_count": dict(
                        reporting_finalizer_result or {}
                    ).get("verified_artifact_count"),
                },
            ),
        )
    except Exception as exc:
        if output_dir:
            try:
                _mark_frontend_run_claims_failed(
                    run_dir=Path(output_dir),
                    error=exc,
                )
            except Exception:
                # Preserve the primary optimization/reporting error even if a
                # best-effort diagnostic label cannot be written.
                pass
        job_store.update_job(
            job_id,
            status="failed",
            progress=100,
            message=(
                "Hourly rolling failed; day-ahead diagnostics were preserved."
                if isinstance(exc, RollingChainExecutionError)
                else "Optimization failed."
            ),
            error=traceback.format_exc(),
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="failed",
                mode=mode,
                extra={
                    "run_dir": output_dir,
                    "run_profile": run_profile,
                    "failure_type": type(exc).__name__,
                },
            ),
        )


def _apply_reoptimization_inputs(
    scenario: Dict[str, Any],
    body: ReoptimizeBody,
) -> Dict[str, Any]:
    updated = dict(scenario)
    if body.updated_pv_profile:
        updated["pv_profiles"] = body.updated_pv_profile
    updated["reoptimization_request"] = {
        "current_time": body.current_time,
        "actual_soc": dict(body.actual_soc),
        "actual_bess_soc_kwh": dict(body.actual_bess_soc_kwh),
        "observed_on_peak_kw_by_depot": dict(
            body.observed_on_peak_kw_by_depot
        ),
        "observed_off_peak_kw_by_depot": dict(
            body.observed_off_peak_kw_by_depot
        ),
        "actual_location_node_id": dict(body.actual_location_node_id),
        "delays": [item.model_dump() for item in body.delays],
        "reoptimization_strategy": body.reoptimization_strategy,
        "execution_minutes": body.execution_minutes,
        "bess_terminal_policy": body.bess_terminal_policy,
    }
    return updated


def _validate_day_ahead_result_contract(
    prior_result: Dict[str, Any],
    *,
    scenario_id: str,
    prepared_input_id: str,
    service_id: str,
    depot_id: str,
) -> None:
    """Reject a persisted result from a different scenario input or scope."""

    prior_scenario_id = str(prior_result.get("scenario_id") or "").strip()
    if prior_scenario_id != str(scenario_id):
        raise ValueError(
            "Persisted day-ahead result scenario mismatch: "
            f"expected {scenario_id!r}, got {prior_scenario_id!r}"
        )
    prior_prepared_input_id = str(
        prior_result.get("prepared_input_id") or ""
    ).strip()
    if prior_prepared_input_id != str(prepared_input_id):
        raise ValueError(
            "Persisted day-ahead result prepared input mismatch: "
            f"expected {prepared_input_id!r}, got {prior_prepared_input_id!r}"
        )

    prior_scope = prior_result.get("scope")
    if not isinstance(prior_scope, dict):
        raise ValueError("Persisted day-ahead result has no dispatch scope")
    expected_scope = {
        "serviceId": str(service_id),
        "depotId": str(depot_id),
    }
    actual_scope = {
        "serviceId": str(prior_scope.get("serviceId") or ""),
        "depotId": str(prior_scope.get("depotId") or ""),
    }
    if actual_scope != expected_scope:
        raise ValueError(
            "Persisted day-ahead result scope mismatch: "
            f"expected {expected_scope}, got {actual_scope}"
        )


def _run_reoptimization(
    scenario_id: str,
    job_id: str,
    body_payload: Dict[str, Any],
    prepared_input_id: str,
    service_id: str,
    depot_id: Optional[str],
) -> None:
    body = ReoptimizeBody(**body_payload)
    mode = body.mode
    timestep_min = _request_timestep_min(body.timestep_min, body.time_step_min)
    try:
        if not depot_id:
            raise ValueError("No depot selected. Configure dispatch scope first.")

        base_scenario = store.get_scenario_document_shallow(scenario_id)
        prior_optimization_result = store.get_field(
            scenario_id, "optimization_result"
        )
        prepared_payload = load_prepared_input(
            scenario_id=scenario_id,
            prepared_input_id=prepared_input_id,
            scenarios_dir=_prepared_inputs_root(),
        )
        scenario = _apply_reoptimization_inputs(
            materialize_scenario_from_prepared_input(base_scenario, prepared_payload),
            body,
        )
        _apply_timestep_min_to_scenario(scenario, timestep_min)
        _apply_interactive_operation_time_window_controls(scenario)
        job_store.update_job(
            job_id,
            status="running",
            progress=15,
            message="Building canonical problem for re-optimization...",
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="reopt_build",
                mode=mode,
                extra={"current_time": body.current_time},
            ),
        )
        solver_mode = _normalize_solver_mode(mode)
        phase_token = _phase_from_solver_mode(solver_mode)
        requested_phase = phase_token or solver_mode
        is_diagnostic_mode = solver_mode in {"debug_mode", "diagnostic"}
        config = OptimizationConfig(
            mode=_parse_optimization_mode(solver_mode),
            time_limit_sec=body.time_limit_seconds,
            mip_gap=body.mip_gap,
            random_seed=body.random_seed,
            alns_iterations=body.alns_iterations,
            no_improvement_limit=body.no_improvement_limit,
            destroy_fraction=body.destroy_fraction,
            rolling_current_min=hhmm_to_min(body.current_time),
            thesis_mode=(
                solver_mode in {"thesis_mode", "mode_milp_only"}
                or phase_token == "phase3_two_stage"
            ),
            debug_mode=is_diagnostic_mode,
            diagnostic_mode=solver_mode == "diagnostic",
            research_run=bool(body.research_run),
            allow_postsolve_repair=solver_mode == "debug_mode",
            phase=phase_token,
            requested_phase_token=solver_mode,
            requested_phase=requested_phase,
            resolved_phase=phase_token,
            executed_phase=phase_token,
        )
        problem = ProblemBuilder().build_from_scenario(
            scenario,
            depot_id=depot_id,
            service_id=service_id,
            config=config,
            planning_days=max(
                int(((scenario.get("simulation_config") or {}).get("planning_days") or 1)),
                1,
            ),
        )
        job_store.update_job(
            job_id,
            status="running",
            progress=55,
            message="Running rolling-horizon re-optimization...",
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="reopt_solve",
                mode=mode,
                extra={
                    "delay_count": len(body.delays),
                    "soc_updates": len(body.actual_soc),
                },
            ),
        )
        rolling_reoptimizer = RollingReoptimizer()
        strategy = str(body.reoptimization_strategy or "generic").strip().lower()
        serialized_day_ahead: Optional[Dict[str, Any]] = None
        if strategy == "day_ahead_hourly":
            if not isinstance(prior_optimization_result, dict):
                raise ValueError(
                    "day_ahead_hourly requires a persisted day-ahead optimization result"
                )
            _validate_day_ahead_result_contract(
                prior_optimization_result,
                scenario_id=scenario_id,
                prepared_input_id=prepared_input_id,
                service_id=service_id,
                depot_id=depot_id,
            )
            serialized_day_ahead = prior_optimization_result.get(
                "canonical_solver_result"
            )
            if not isinstance(serialized_day_ahead, dict):
                raise ValueError(
                    "Persisted optimization result has no canonical day-ahead assignment"
                )
            day_ahead_plan = assignment_plan_from_serialized_result(
                problem,
                serialized_day_ahead,
            )
            result = rolling_reoptimizer.reoptimize_charging_hour(
                problem,
                day_ahead_plan,
                config=config,
                current_min=hhmm_to_min(body.current_time),
                actual_soc=body.actual_soc,
                actual_bess_soc_kwh=body.actual_bess_soc_kwh,
                observed_on_peak_kw_by_depot=(
                    body.observed_on_peak_kw_by_depot
                ),
                observed_off_peak_kw_by_depot=(
                    body.observed_off_peak_kw_by_depot
                ),
                execution_minutes=body.execution_minutes,
                bess_terminal_policy=body.bess_terminal_policy,
            )
        elif strategy == "generic":
            result = rolling_reoptimizer.reoptimize(
                problem,
                config=config,
                current_min=hhmm_to_min(body.current_time),
                actual_soc=body.actual_soc,
                actual_bess_soc_kwh=body.actual_bess_soc_kwh,
            )
        else:
            raise ValueError(
                "reoptimization_strategy must be 'generic' or 'day_ahead_hourly'"
            )
        payload = {
            "scenario_id": scenario_id,
            "scope": {"serviceId": service_id, "depotId": depot_id},
            "prepared_input_id": prepared_input_id,
            "reoptimized": True,
            "reoptimization_request": {
                "current_time": body.current_time,
                "actual_soc": dict(body.actual_soc),
                "actual_bess_soc_kwh": dict(body.actual_bess_soc_kwh),
                "observed_on_peak_kw_by_depot": dict(
                    body.observed_on_peak_kw_by_depot
                ),
                "observed_off_peak_kw_by_depot": dict(
                    body.observed_off_peak_kw_by_depot
                ),
                "actual_location_node_id": dict(body.actual_location_node_id),
                "delays": [item.model_dump() for item in body.delays],
                "reoptimization_strategy": strategy,
                "execution_minutes": body.execution_minutes,
                "bess_terminal_policy": body.bess_terminal_policy,
            },
            **ResultSerializer.serialize_result(result),
        }
        if serialized_day_ahead is not None:
            # Keep the verified day-ahead assignment available for the next
            # hourly update; the newly optimized charging schedule must not
            # become the source of vehicle-trip assignment truth.
            payload["canonical_solver_result"] = serialized_day_ahead
            payload["day_ahead_reference"] = {
                "scenario_id": scenario_id,
                "prepared_input_id": prepared_input_id,
                "service_id": service_id,
                "depot_id": depot_id,
            }
        store.set_field(scenario_id, "optimization_result", payload)
        store.set_field(
            scenario_id,
            "optimization_audit",
            {
                "scenario_id": scenario_id,
                "depot_id": depot_id,
                "service_id": service_id,
                "solver_mode": mode,
                "reoptimized": True,
                "current_time": body.current_time,
                "delay_count": len(body.delays),
                "actual_soc_count": len(body.actual_soc),
                "random_seed": body.random_seed,
                "gurobi_seed": body.random_seed,
                "executed_at": datetime.now(timezone.utc).isoformat(),
                "git_sha": _git_sha(),
                "source_snapshot": store.get_field(scenario_id, "source_snapshot"),
            },
        )
        job_store.update_job(
            job_id,
            status="completed",
            progress=100,
            message="Re-optimization complete.",
            result_key="optimization_result",
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="reopt_completed",
                mode=mode,
                extra={"objective_value": payload.get("objective_value")},
            ),
        )
    except Exception:
        job_store.update_job(
            job_id,
            status="failed",
            message="Re-optimization failed.",
            error=traceback.format_exc(),
            metadata=_job_metadata(
                scenario_id=scenario_id,
                service_id=service_id,
                depot_id=depot_id,
                stage="reopt_failed",
                mode=mode,
            ),
        )


@router.get("/scenarios/{scenario_id}/optimization")
def get_optimization_result(scenario_id: str) -> Dict[str, Any]:
    _require_scenario(scenario_id)
    result = store.get_field(scenario_id, "optimization_result")
    if result is None:
        raise HTTPException(
            status_code=404,
            detail="Optimization has not been run yet. POST to /run-optimization first.",
        )
    if isinstance(result, dict) and "audit" not in result:
        audit = store.get_field(scenario_id, "optimization_audit")
        if audit is not None:
            result = {**result, "audit": audit}
    if isinstance(result, dict) and "electricity_cost_basis" not in result:
        simulation_summary = dict(result.get("simulation_summary") or {})
        result = {
            **result,
            "electricity_cost_basis": str(
                simulation_summary.get("electricity_cost_basis") or "provisional_drive"
            ),
        }
    return result


@router.get("/scenarios/{scenario_id}/optimization/capabilities")
def get_optimization_capabilities(scenario_id: str) -> Dict[str, Any]:
    _require_scenario(scenario_id)
    return _optimization_capabilities()


@router.get("/research/git-preflight")
def get_research_git_preflight() -> Dict[str, Any]:
    """Expose the same Git contract used by formal worker-side validation."""

    return _research_git_preflight_payload()


@router.post("/scenarios/{scenario_id}/run-optimization")
def run_optimization(
    scenario_id: str,
    body: Optional[RunOptimizationBody] = None,
    _app_state: dict = Depends(require_built),
) -> Dict[str, Any]:
    _require_scenario(scenario_id)
    request = body or RunOptimizationBody()
    _require_research_git_preflight_before_job_creation(
        research_run=bool(request.research_run)
    )
    if (
        request.stage1_bev_frontier_enabled
        and request.stage1_bev_frontier_min_count
        > request.stage1_bev_frontier_max_count
    ):
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                "stage1_bev_frontier_min_count must be <= "
                "stage1_bev_frontier_max_count",
                field="stage1_bev_frontier_min_count",
            ),
        )
    normalized_requested_mode = _normalize_solver_mode(request.mode)
    if request.stage1_bev_frontier_enabled and normalized_requested_mode != (
        "phase3_two_stage"
    ):
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                "stage1_bev_frontier_enabled requires phase3_two_stage",
                field="stage1_bev_frontier_enabled",
            ),
        )
    if request.integrated_actual_cost_objective and normalized_requested_mode != (
        "phase4_integrated"
    ):
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                "integrated_actual_cost_objective requires phase4_integrated",
                field="integrated_actual_cost_objective",
            ),
        )
    if (
        request.integrated_actual_cost_objective
        and request.integrated_ev_utilization_mode != "disabled"
    ):
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                "integrated actual-cost minimization and EV-utilization "
                "policy objectives are mutually exclusive",
                field="integrated_ev_utilization_mode",
            ),
        )
    if (
        request.integrated_ev_utilization_mode != "disabled"
        and normalized_requested_mode != "phase4_integrated"
    ):
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                "integrated_ev_utilization_mode requires phase4_integrated",
                field="integrated_ev_utilization_mode",
            ),
        )
    if (
        request.integrated_actual_cost_upper_bound_jpy is not None
        and request.integrated_ev_utilization_mode == "disabled"
    ):
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                "integrated_actual_cost_upper_bound_jpy requires an EV "
                "utilization mode",
                field="integrated_actual_cost_upper_bound_jpy",
            ),
        )
    if (
        request.integrated_actual_cost_upper_bound_delta_ratio is not None
        and request.integrated_actual_cost_upper_bound_jpy is None
    ):
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                "integrated_actual_cost_upper_bound_delta_ratio requires "
                "the absolute cost upper bound",
                field="integrated_actual_cost_upper_bound_delta_ratio",
            ),
        )
    try:
        run_profile = normalize_frontend_run_profile(request.run_profile)
    except ValueError as exc:
        raise HTTPException(
            status_code=422,
            detail=make_error(
                AppErrorCode.SCHEMA_VALIDATION_ERROR,
                str(exc),
                field="run_profile",
            ),
        ) from exc
    rolling_required = frontend_rolling_is_required(run_profile)
    # Preserve the parsed request before applying server-authoritative rolling
    # controls.  Provenance must distinguish what the client requested from
    # what the BFF actually executed.
    requested_frontend_payload = request.model_dump()
    effective_run_hourly_rolling = bool(rolling_required)
    effective_rolling_execution_minutes = 60
    timestep_min = _request_timestep_min(request.timestep_min, request.time_step_min)
    # Apply the request's depot/service to the persisted scope BEFORE building the
    # run preparation, so that resolve_scope sees the correct depotId/serviceId
    # even when the scenario's dispatch_scope was left empty by a previous operation.
    # Skip if a prepared_input_id was supplied — the prepare step already fixed the scope,
    # and persisting again would change the scenario_hash and invalidate the prepared input.
    if (request.service_id or request.depot_id) and not request.prepared_input_id:
        _resolve_dispatch_scope(
            scenario_id,
            service_id=request.service_id,
            depot_id=request.depot_id,
            persist=True,
        )
    scenario = store.get_scenario_document_shallow(scenario_id)
    if timestep_min is not None and not request.prepared_input_id:
        _apply_timestep_min_to_scenario(scenario, timestep_min)
        store.set_field(scenario_id, "simulation_config", scenario["simulation_config"])
    prep = get_or_build_run_preparation(
        scenario=scenario,
        built_dir=Path(_app_state.get("built_dir") or "data/built/tokyu_core"),
        scenarios_dir=_prepared_inputs_root(),
        routes_df=_app_state.get("routes_df"),
        force_rebuild=bool(request.force_reprepare),
    )
    if not prep.is_valid:
        error_code = AppErrorCode(prep.error_code) if prep.error_code else AppErrorCode.SCENARIO_INCOMPLETE
        raise HTTPException(
            status_code=422 if prep.error_code else 500,
            detail=make_error(
                error_code,
                f"Run preparation failed: {prep.error}",
                preparedInputId=prep.prepared_input_id,
                scopeSummary=prep.scope_summary,
            ),
        )
    _require_nonempty_prepared_scope(prep, action="Optimization preflight")
    _preflight_weather_proxy_request(
        scenario=scenario,
        enable_weather_operation_policy=request.enableWeatherOperationPolicy,
        weather_proxy_forecast_path=request.weatherProxyForecastPath,
    )
    if request.prepared_input_id and prep.prepared_input_id != request.prepared_input_id:
        raise HTTPException(
            status_code=409,
            detail=make_error(
                AppErrorCode.SCENARIO_INCOMPLETE,
                "Prepared input is stale. Run prepare again before starting optimization.",
                preparedInputId=request.prepared_input_id,
                currentPreparedInputId=prep.prepared_input_id,
            ),
        )
    scope = _resolve_dispatch_scope(
        scenario_id,
        service_id=request.service_id,
        depot_id=request.depot_id,
        persist=True,
    )
    job = job_store.create_job(execution_model=_executor_mode())
    job_store.update_job(
        job.job_id,
        metadata=_job_metadata(
            scenario_id=scenario_id,
            service_id=scope.get("serviceId") or "WEEKDAY",
            depot_id=scope.get("depotId"),
            stage="queued",
            mode=request.mode,
            extra={"persistence": dict(job_store.JOB_PERSISTENCE_INFO)},
        ),
    )
    submitted = _submit_optimization_job(
        fn=_run_optimization,
        args=(
            scenario_id,
            job.job_id,
            prep.prepared_input_id or "",
            request.prepared_input_id,
            request.mode,
            request.time_limit_seconds,
            request.mip_gap,
            request.random_seed,
            scope.get("serviceId") or "WEEKDAY",
            scope.get("depotId"),
            request.rebuild_dispatch,
            request.use_existing_duties,
            request.alns_iterations,
            request.no_improvement_limit,
            request.destroy_fraction,
            timestep_min,
            request.enableWeatherOperationPolicy,
            request.weatherProxyForecastPath,
            request.research_run,
            request.stage1_time_limit_seconds,
            request.stage2_time_limit_seconds,
            request.stage1_best_obj_stop_enabled,
            request.gurobi_threads,
            run_profile,
            effective_run_hourly_rolling,
            effective_rolling_execution_minutes,
            requested_frontend_payload,
            request.stage1_stage2_candidate_limit,
            request.stage1_composition_search_radius,
            request.stage1_bev_frontier_enabled,
            request.stage1_bev_frontier_min_count,
            request.stage1_bev_frontier_max_count,
            request.stage1_bev_frontier_target_time_limit_seconds,
            request.integrated_actual_cost_objective,
            request.integrated_ev_utilization_mode,
            request.integrated_actual_cost_upper_bound_jpy,
            request.integrated_actual_cost_upper_bound_delta_ratio,
            request.co2_emissions_cap_kg,
        ),
        job_id=job.job_id,
        scenario_id=scenario_id,
        service_id=scope.get("serviceId") or "WEEKDAY",
        depot_id=scope.get("depotId"),
        mode=request.mode,
        stage="worker_crashed",
    )
    if not submitted:
        job_store.update_job(
            job.job_id,
            status="failed",
            progress=100,
            message="Rejected because another optimization job is already running.",
            error="job_already_running",
        )
        raise HTTPException(
            status_code=503,
            detail=make_error(
                AppErrorCode.EXECUTION_IN_PROGRESS,
                "An optimization job is already running. Please retry after it completes.",
            ),
        )
    return job_store.job_to_dict(job)


@router.post("/scenarios/{scenario_id}/reoptimize")
def reoptimize(
    scenario_id: str,
    body: ReoptimizeBody,
    _app_state: dict = Depends(require_built),
) -> Dict[str, Any]:
    _require_scenario(scenario_id)
    _require_research_git_preflight_before_job_creation(
        research_run=bool(body.research_run)
    )
    timestep_min = _request_timestep_min(body.timestep_min, body.time_step_min)
    scenario = store.get_scenario_document_shallow(scenario_id)
    if timestep_min is not None and not body.prepared_input_id:
        _apply_timestep_min_to_scenario(scenario, timestep_min)
        store.set_field(scenario_id, "simulation_config", scenario["simulation_config"])
    prep = get_or_build_run_preparation(
        scenario=scenario,
        built_dir=Path(_app_state.get("built_dir") or "data/built/tokyu_core"),
        scenarios_dir=_prepared_inputs_root(),
        routes_df=_app_state.get("routes_df"),
    )
    if not prep.is_valid:
        error_code = AppErrorCode(prep.error_code) if prep.error_code else AppErrorCode.SCENARIO_INCOMPLETE
        raise HTTPException(
            status_code=422 if prep.error_code else 500,
            detail=make_error(
                error_code,
                f"Run preparation failed: {prep.error}",
                preparedInputId=prep.prepared_input_id,
                scopeSummary=prep.scope_summary,
            ),
        )
    _require_nonempty_prepared_scope(prep, action="Re-optimization preflight")
    if body.prepared_input_id and prep.prepared_input_id != body.prepared_input_id:
        raise HTTPException(
            status_code=409,
            detail=make_error(
                AppErrorCode.SCENARIO_INCOMPLETE,
                "Prepared input is stale. Run prepare again before starting re-optimization.",
                preparedInputId=body.prepared_input_id,
                currentPreparedInputId=prep.prepared_input_id,
            ),
        )
    scope = _resolve_dispatch_scope(
        scenario_id,
        service_id=body.service_id,
        depot_id=body.depot_id,
        persist=True,
    )
    job = job_store.create_job(execution_model=_executor_mode())
    job_store.update_job(
        job.job_id,
        metadata=_job_metadata(
            scenario_id=scenario_id,
            service_id=scope.get("serviceId") or "WEEKDAY",
            depot_id=scope.get("depotId"),
            stage="queued",
            mode=body.mode,
            extra={"persistence": dict(job_store.JOB_PERSISTENCE_INFO)},
        ),
    )
    submitted = _submit_optimization_job(
        fn=_run_reoptimization,
        args=(
            scenario_id,
            job.job_id,
            body.model_dump(),
            prep.prepared_input_id or "",
            scope.get("serviceId") or "WEEKDAY",
            scope.get("depotId"),
        ),
        job_id=job.job_id,
        scenario_id=scenario_id,
        service_id=scope.get("serviceId") or "WEEKDAY",
        depot_id=scope.get("depotId"),
        mode=body.mode,
        stage="reopt_worker_crashed",
    )
    if not submitted:
        job_store.update_job(
            job.job_id,
            status="failed",
            progress=100,
            message="Rejected because another optimization job is already running.",
            error="job_already_running",
        )
        raise HTTPException(
            status_code=503,
            detail=make_error(
                AppErrorCode.EXECUTION_IN_PROGRESS,
                "An optimization job is already running. Please retry after it completes.",
            ),
        )
    return job_store.job_to_dict(job)
