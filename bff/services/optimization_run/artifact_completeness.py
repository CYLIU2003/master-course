"""Fail-closed completeness audit for interactive optimization run artifacts."""

from __future__ import annotations

from collections import Counter
from collections.abc import Mapping
import csv
from decimal import Decimal, InvalidOperation
import hashlib
import json
import math
from pathlib import Path
from typing import Any, Iterable


ARTIFACT_CONTRACT_VERSION = "frontend_run_artifacts_v2"
PHYSICAL_VALIDATION_INPUT_MANIFEST_SCHEMA_VERSION = (
    "physical_validation_input_manifest_v1"
)
SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE = (
    "solver_objective_accounting_reconciliation.json"
)
SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_SCHEMA_VERSION = (
    "solver_objective_accounting_reconciliation_v1"
)
STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE = (
    "stage1_used_powertrain_composition_search.json"
)
STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_CSV_FILE = (
    "stage1_used_powertrain_composition_search.csv"
)
STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_SCHEMA_VERSION = (
    "stage1_used_powertrain_composition_search_v1"
)
STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_SCHEMA_V2 = (
    "stage1_used_powertrain_composition_search_v2"
)

# These artifacts are generated for every successfully finalized interactive
# optimization run.  Optional visualizations are validated through
# graph/manifest.json instead of being hard-coded here.
BASE_REQUIRED_ARTIFACTS = (
    "assignment_economic_audit.csv",
    "assignment_economic_audit.json",
    "baseline_vs_integrated_actual_cost.csv",
    "assignment_validation_diagnostics.json",
    "canonical_solver_result.json",
    "charging_schedule.csv",
    "charging_source_provenance.json",
    "charging_summary.csv",
    "charging_summary.json",
    "co2_breakdown.csv",
    "co2_breakdown.json",
    "cost_breakdown_detail.csv",
    "cost_breakdown_detail.json",
    "depot_energy_flows.csv",
    "depot_energy_flows.json",
    "experiment_report.json",
    "experiment_report.md",
    "kpi_summary.json",
    "objective_breakdown.csv",
    "objective_breakdown.json",
    "operating_and_lifecycle_cost_scope.csv",
    "operating_and_lifecycle_cost_scope.json",
    "optimization_audit.json",
    "optimization_result.json",
    "powertrain_marginal_cost_audit.csv",
    "powertrain_marginal_cost_audit.json",
    "raw/assignment.csv",
    "raw/canonical_solver_result.json",
    "raw/optimization_audit.json",
    "raw/optimization_result.json",
    "raw/solver_result.json",
    "raw/unserved_trips.csv",
    "rebuild_reporting_log.json",
    "refuel_events.csv",
    "research_claim_scope.json",
    "results.xlsx",
    "run_manifest.json",
    "simulation_conditions.json",
    "simulation_conditions_contract_limits.csv",
    "simulation_conditions_provenance.json",
    "simulation_conditions_tou_prices.csv",
    "simulation_conditions_vehicle_costs.csv",
    "site_power_balance.csv",
    SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE,
    "solver_result.json",
    "solver_settings.json",
    "strict_reconciliation.csv",
    "strict_reconciliation.md",
    "strict_reconciliation_after_rebuild.csv",
    "strict_reconciliation_after_rebuild.md",
    "summary.json",
    "targeted_trips.csv",
    "targeted_trips.json",
    "thesis_ablation/day_ahead_method_candidates.csv",
    "thesis_ablation/day_ahead_method_candidates.json",
    "trip_type_counts.csv",
    "trip_type_counts.json",
    "trip_powertrain_cost_comparison.csv",
    "vehicle_schedule.csv",
    "vehicle_timeline_gantt.csv",
    "vehicle_timelines.csv",
    "vehicle_timelines.json",
    "graph/canonical_cost_ledger.json",
    "graph/data_flow_validation.csv",
    "graph/manifest.json",
    "graph/refuel_events.csv",
)

TWO_STAGE_FORMAL_REQUIRED_ARTIFACTS = (
    STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE,
    STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_CSV_FILE,
)

BEV_FRONTIER_REQUIRED_ARTIFACTS = (
    "bev_cost_frontier.csv",
    "bev_cost_frontier.json",
    "bev_cost_frontier.md",
    "maximum_bev_feasibility_search.csv",
)

RESEARCH_PROVENANCE_ARTIFACTS = (
    "code_provenance.json",
    "optimization_parameters.json",
    "prepare_input_audit.json",
    "run_input_manifest.json",
    "run_input_summary.md",
    "run_input_validation.json",
    "scenario_input_snapshot.json",
)

ROLLING_REQUIRED_ARTIFACTS = (
    "comparison_case_manifest.json",
    "effective_pv_profiles.json",
    "effective_scenario.json",
    "final_cost_reconciliation.json",
    "input_audit.json",
    "manifest.json",
    "physical_validation_input_manifest.json",
    "physical_schedule_validation.json",
    "scenario_fleet_contract.json",
    "rolling_hourly_chain/charging_schedule.csv",
    "rolling_hourly_chain/day_ahead_vs_rolling_summary.json",
    "rolling_hourly_chain/executed_day_accounting.json",
    "rolling_hourly_chain/hourly_energy_flow_chart.csv",
    "rolling_hourly_chain/rolling_chain_summary.json",
    "graph/charger_occupancy_timeline.csv",
    "graph/physical_schedule_validation.json",
    "graph/physical_schedule_violations.csv",
    "graph/vehicle_event_timeline.csv",
    "graph/vehicle_location_timeline.csv",
    "graph/vehicle_soc_event_timeline.csv",
    "graph/literature_figures/manifest.json",
)

REQUIRED_WORKBOOK_SHEETS = (
    "summary",
    "cost_breakdown",
    "release_status",
)

_ROOT_REFUEL_EVENT_FIELDS = (
    "vehicle_id",
    "slot_index",
    "time_hhmm",
    "refuel_liters",
    "unit",
)
_GRAPH_REFUEL_EVENT_FIELDS = (
    "vehicle_id",
    "slot_index",
    "time_hhmm",
    "refuel_liters",
    "location_id",
)


def _load_json_object(path: Path) -> dict[str, Any]:
    loaded = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(loaded, dict):
        raise ValueError("expected a JSON object")
    return dict(loaded)


def _finite_json_number(value: Any) -> float | None:
    """Return a JSON numeric value only when it is finite.

    Booleans are deliberately excluded even though Python treats them as ints.
    A release artifact must not turn a truth value into a currency amount.
    """

    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    numeric = float(value)
    return numeric if math.isfinite(numeric) else None


def _nonnegative_integer(value: Any) -> int | None:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        return None
    return int(value)


def _sha256_hex(value: Any) -> bool:
    """Return whether ``value`` is a lower/upper-case SHA-256 hex digest."""

    return bool(
        isinstance(value, str)
        and len(value) == 64
        and all(character in "0123456789abcdefABCDEF" for character in value)
    )


def validate_solver_objective_accounting_reconciliation(
    payload: Mapping[str, Any] | None,
    *,
    require_match: bool,
) -> list[str]:
    """Validate numeric solver-to-accounting reconciliation evidence.

    The legacy ``solver_objective_matches_accounting_total`` flag is useful
    telemetry, but it is not evidence by itself.  Formal release gates require
    the two finite values, a derived residual, an explicit tolerance, and an
    accounting source that is authoritative for the accepted rolling day.
    """

    artifact = SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE
    if not isinstance(payload, Mapping):
        return [f"{artifact} must be a JSON object"]
    payload = dict(payload)
    errors: list[str] = []
    if (
        payload.get("schema_version")
        != SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_SCHEMA_VERSION
    ):
        errors.append(f"{artifact} has an invalid schema_version")

    for key in (
        "solver_objective_source",
        "canonical_accounting_source",
        "objective_semantics",
    ):
        if not str(payload.get(key) or "").strip():
            errors.append(f"{artifact}: {key} must be a non-empty string")

    boolean_keys = (
        "numeric_values_available",
        "numeric_residual_within_tolerance",
        "objective_is_actual_cost",
        "matches_canonical_accounting_total",
    )
    for key in boolean_keys:
        if not isinstance(payload.get(key), bool):
            errors.append(f"{artifact}: {key} must be a boolean")

    tolerance = _finite_json_number(payload.get("tolerance_jpy"))
    if tolerance is None or tolerance < 0.0:
        errors.append(f"{artifact}: tolerance_jpy must be a finite non-negative number")

    numeric_values_available = payload.get("numeric_values_available") is True
    if numeric_values_available:
        solver_value = _finite_json_number(payload.get("solver_objective_value_jpy"))
        accounting_total = _finite_json_number(
            payload.get("canonical_accounting_total_jpy")
        )
        difference = _finite_json_number(payload.get("difference_jpy"))
        absolute_difference = _finite_json_number(
            payload.get("absolute_difference_jpy")
        )
        if solver_value is None:
            errors.append(
                f"{artifact}: solver_objective_value_jpy must be finite when numeric_values_available"
            )
        if accounting_total is None:
            errors.append(
                f"{artifact}: canonical_accounting_total_jpy must be finite when numeric_values_available"
            )
        if difference is None:
            errors.append(
                f"{artifact}: difference_jpy must be finite when numeric_values_available"
            )
        if absolute_difference is None or absolute_difference < 0.0:
            errors.append(
                f"{artifact}: absolute_difference_jpy must be finite and non-negative when numeric_values_available"
            )
        if (
            solver_value is not None
            and accounting_total is not None
            and difference is not None
            and absolute_difference is not None
        ):
            expected_difference = solver_value - accounting_total
            if not math.isclose(
                difference,
                expected_difference,
                rel_tol=1.0e-12,
                abs_tol=1.0e-9,
            ):
                errors.append(
                    f"{artifact}: difference_jpy is not solver_objective_value_jpy minus canonical_accounting_total_jpy"
                )
            if not math.isclose(
                absolute_difference,
                abs(expected_difference),
                rel_tol=1.0e-12,
                abs_tol=1.0e-9,
            ):
                errors.append(
                    f"{artifact}: absolute_difference_jpy is inconsistent with difference_jpy"
                )
            if tolerance is not None:
                residual_within_tolerance = (
                    abs(expected_difference) <= tolerance
                )
                if (
                    payload.get("numeric_residual_within_tolerance")
                    is not residual_within_tolerance
                ):
                    errors.append(
                        f"{artifact}: numeric_residual_within_tolerance is not derived from the numeric residual"
                    )
    else:
        if payload.get("numeric_residual_within_tolerance") is True:
            errors.append(
                f"{artifact}: numeric_residual_within_tolerance cannot be true without numeric values"
            )
        if payload.get("matches_canonical_accounting_total") is True:
            errors.append(
                f"{artifact}: matches_canonical_accounting_total cannot be true without numeric values"
            )

    if not require_match:
        return errors

    if not numeric_values_available:
        errors.append(f"{artifact}: formal release requires finite numeric values")
    if payload.get("numeric_residual_within_tolerance") is not True:
        errors.append(
            f"{artifact}: formal release requires residual within tolerance"
        )
    if payload.get("objective_is_actual_cost") is not True:
        errors.append(
            f"{artifact}: formal release requires an actual-cost solver objective"
        )
    if (
        payload.get("canonical_accounting_source")
        != "rolling_hourly_chain/executed_day_accounting.json"
    ):
        errors.append(
            f"{artifact}: formal release requires the executed-day accounting source"
        )
    if payload.get("matches_canonical_accounting_total") is not True:
        errors.append(
            f"{artifact}: formal release requires a derived accounting match"
        )
    return errors


def validate_stage1_used_powertrain_composition_search(
    payload: Mapping[str, Any] | None,
    *,
    require_accepted: bool,
) -> list[str]:
    """Validate the persisted Stage-1 used-powertrain search certificate.

    A formal two-stage release cannot treat a summary boolean as a certificate.
    It needs a versioned search artifact that records either multiple physically
    feasible activated compositions, every in-inventory adjacent target as
    certified infeasible, or the explicit no-adjacent-inventory boundary.
    """

    artifact = STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE
    if not isinstance(payload, Mapping):
        return [f"{artifact} must be a JSON object"]
    payload = dict(payload)
    errors: list[str] = []
    if payload.get("schema_version") not in {
        STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_SCHEMA_VERSION,
        STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_SCHEMA_V2,
    }:
        errors.append(f"{artifact} has an invalid schema_version")
    frontier_mode = payload.get("search_mode") == "minimum_used_bev_frontier"

    for key in (
        "enabled",
        "multiple_feasible_compositions_found",
        "all_adjacent_targets_certified_infeasible",
        "inventory_has_no_adjacent_composition",
        "accepted_for_formal_composition_evidence",
    ):
        if not isinstance(payload.get(key), bool):
            errors.append(f"{artifact}: {key} must be a boolean")

    if _nonnegative_integer(payload.get("radius_requested")) is None:
        errors.append(f"{artifact}: radius_requested must be a non-negative integer")

    def _validate_pair(value: Any, *, label: str) -> tuple[int, int] | None:
        if not isinstance(value, Mapping):
            errors.append(f"{artifact}: {label} must be an object")
            return None
        used_bev = _nonnegative_integer(value.get("used_bev"))
        used_ice = _nonnegative_integer(value.get("used_ice"))
        if used_bev is None or used_ice is None:
            errors.append(
                f"{artifact}: {label} must contain non-negative integer used_bev and used_ice"
            )
            return None
        return used_bev, used_ice

    _validate_pair(
        payload.get("primary_used_powertrain_composition"),
        label="primary_used_powertrain_composition",
    )
    inventory = payload.get("selected_inventory")
    if not isinstance(inventory, Mapping):
        errors.append(f"{artifact}: selected_inventory must be an object")
    else:
        for key in (
            "available_electric_vehicle_count",
            "available_combustion_vehicle_count",
        ):
            if _nonnegative_integer(inventory.get(key)) is None:
                errors.append(
                    f"{artifact}: selected_inventory.{key} must be a non-negative integer"
                )
        for key in ("electric_vehicle_ids", "combustion_vehicle_ids"):
            if not isinstance(inventory.get(key), list):
                errors.append(
                    f"{artifact}: selected_inventory.{key} must be a list"
                )

    target_records = payload.get("target_records")
    if not isinstance(target_records, list):
        errors.append(f"{artifact}: target_records must be a list")
        target_records = []
    feasible_pairs_payload = payload.get("feasible_used_powertrain_compositions")
    if not isinstance(feasible_pairs_payload, list):
        errors.append(
            f"{artifact}: feasible_used_powertrain_compositions must be a list"
        )
        feasible_pairs_payload = []
    unresolved_targets = payload.get("unresolved_targets")
    if not isinstance(unresolved_targets, list):
        errors.append(f"{artifact}: unresolved_targets must be a list")
        unresolved_targets = []
    blocking_reasons = payload.get("blocking_reasons")
    if not isinstance(blocking_reasons, list) or any(
        not isinstance(reason, str) or not reason.strip()
        for reason in (blocking_reasons or [])
    ):
        errors.append(f"{artifact}: blocking_reasons must be a list of non-empty strings")
        blocking_reasons = []
    if not str(payload.get("semantics") or "").strip():
        errors.append(f"{artifact}: semantics must be a non-empty string")

    feasible_pairs: set[tuple[int, int]] = set()
    for index, pair in enumerate(feasible_pairs_payload):
        parsed = _validate_pair(
            pair,
            label=f"feasible_used_powertrain_compositions[{index}]",
        )
        if parsed is not None:
            feasible_pairs.add(parsed)

    valid_target_records: list[dict[str, Any]] = []
    in_inventory_target_records: list[dict[str, Any]] = []
    for index, record in enumerate(target_records):
        if not isinstance(record, Mapping):
            errors.append(f"{artifact}: target_records[{index}] must be an object")
            continue
        record = dict(record)
        target_bev = _nonnegative_integer(record.get("target_used_bev"))
        target_ice = _nonnegative_integer(record.get("target_used_ice"))
        minimum_bev = _nonnegative_integer(
            record.get("minimum_used_bev_count")
        )
        invalid_target = bool(
            target_bev is None
            or (
                frontier_mode
                and (
                    minimum_bev is None
                    or record.get("target_used_ice") is not None
                    or record.get("target_total_used_vehicle_count")
                    is not None
                )
            )
            or (not frontier_mode and target_ice is None)
        )
        if invalid_target:
            errors.append(
                f"{artifact}: target_records[{index}] has invalid target-count semantics"
            )
            continue
        if not isinstance(record.get("target_within_selected_inventory"), bool):
            errors.append(
                f"{artifact}: target_records[{index}].target_within_selected_inventory must be a boolean"
            )
            continue
        valid_target_records.append(record)
        if record["target_within_selected_inventory"] is True:
            in_inventory_target_records.append(record)

    if not require_accepted:
        return errors

    if payload.get("enabled") is not True:
        errors.append(f"{artifact}: formal release requires enabled=true")
    if not valid_target_records:
        errors.append(f"{artifact}: formal release requires target records")
    if payload.get("accepted_for_formal_composition_evidence") is not True:
        errors.append(
            f"{artifact}: formal release requires accepted_for_formal_composition_evidence=true"
        )
    if blocking_reasons:
        errors.append(f"{artifact}: accepted evidence cannot retain blocking_reasons")

    multiple_feasible = payload.get("multiple_feasible_compositions_found") is True
    all_infeasible = (
        payload.get("all_adjacent_targets_certified_infeasible") is True
    )
    no_adjacent_inventory = (
        payload.get("inventory_has_no_adjacent_composition") is True
    )
    frontier_all_resolved = (
        payload.get("all_requested_targets_resolved") is True
    )
    if multiple_feasible and len(feasible_pairs) < 2:
        errors.append(
            f"{artifact}: multiple_feasible_compositions_found requires two distinct feasible compositions"
        )
    if all_infeasible or frontier_mode:
        if not in_inventory_target_records:
            errors.append(
                f"{artifact}: all_adjacent_targets_certified_infeasible requires in-inventory targets"
            )
        for index, record in enumerate(in_inventory_target_records):
            if (
                frontier_mode
                and record.get("final_disposition")
                == "physically_feasible_stage2_candidate"
            ):
                continue
            if record.get("final_disposition") != "stage1_infeasibility_certificate":
                errors.append(
                    f"{artifact}: in-inventory target {index} lacks a stage1_infeasibility_certificate disposition"
                )
                continue
            certificate = record.get("infeasibility_certificate")
            if not isinstance(certificate, Mapping):
                errors.append(
                    f"{artifact}: in-inventory target {index} lacks an infeasibility certificate object"
                )
                continue
            certificate = dict(certificate)
            if (
                certificate.get("accepted_for_formal_composition_evidence")
                is not True
            ):
                errors.append(
                    f"{artifact}: in-inventory target {index} lacks an accepted infeasibility certificate"
                )
            if certificate.get("solver_status") != "infeasible":
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate solver_status must be infeasible"
                )
            if certificate.get("iis_generated") is not True:
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate must record a successful IIS"
                )
            iis_constraint_names = certificate.get("iis_constraint_names")
            target_constraint_names = certificate.get(
                "target_count_constraint_names"
            )
            if (
                not isinstance(iis_constraint_names, list)
                or not iis_constraint_names
                or any(
                    not isinstance(name, str) or not name.strip()
                    for name in iis_constraint_names
                )
            ):
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate IIS must be a non-empty string list"
                )
            if (
                not isinstance(target_constraint_names, list)
                or not target_constraint_names
                or any(
                    not isinstance(name, str) or not name.strip()
                    for name in target_constraint_names
                )
            ):
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate target-count names must be a non-empty string list"
                )
            elif not set(target_constraint_names).intersection(
                iis_constraint_names
                if isinstance(iis_constraint_names, list)
                else []
            ):
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate IIS omits the target-count constraint"
                )
            if certificate.get("target_count_constraint_in_iis") is not True:
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate must flag target_count_constraint_in_iis"
                )
            if not _sha256_hex(certificate.get("stage1_model_lp_sha256")):
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate needs a Stage 1 LP SHA-256"
                )
            if not _sha256_hex(certificate.get("solver_controls_hash")):
                errors.append(
                    f"{artifact}: in-inventory target {index} certificate needs a solver-controls SHA-256"
                )
            failure_reasons = certificate.get("failure_reasons")
            if not isinstance(failure_reasons, list) or failure_reasons:
                errors.append(
                    f"{artifact}: accepted in-inventory target {index} certificate cannot retain failure_reasons"
                )
    if no_adjacent_inventory and in_inventory_target_records:
        errors.append(
            f"{artifact}: inventory_has_no_adjacent_composition conflicts with in-inventory targets"
        )
    frontier_has_certificate = any(
        record.get("final_disposition")
        == "stage1_infeasibility_certificate"
        for record in in_inventory_target_records
    )
    if frontier_mode:
        if payload.get("frontier_total_used_vehicle_count_fixed") is not False:
            errors.append(
                f"{artifact}: BEV frontier must not fix total used vehicle count"
            )
        if not frontier_all_resolved:
            errors.append(
                f"{artifact}: formal BEV frontier requires all targets resolved"
            )
        if any(
            record.get("final_disposition") not in {
                "physically_feasible_stage2_candidate",
                "stage1_infeasibility_certificate",
            }
            for record in in_inventory_target_records
        ):
            errors.append(
                f"{artifact}: formal BEV frontier contains an unresolved target"
            )
        if not (multiple_feasible or frontier_has_certificate):
            errors.append(
                f"{artifact}: formal BEV frontier requires two feasible compositions or an infeasibility certificate"
            )
    elif not (multiple_feasible or all_infeasible or no_adjacent_inventory):
        errors.append(
            f"{artifact}: formal release requires feasible alternatives or a complete adjacent-composition certificate"
        )
    return errors


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalized_paths(paths: Iterable[str]) -> list[str]:
    return sorted(
        {
            Path(str(path)).as_posix()
            for path in paths
            if str(path or "").strip()
        }
    )


def _safe_relative_path(value: Any) -> str | None:
    text = str(value or "").strip().replace("\\", "/")
    if not text:
        return None
    path = Path(text)
    if path.is_absolute() or any(part == ".." for part in path.parts):
        return None
    return path.as_posix()


def _required_manifest_count(
    payload: dict[str, Any],
    *,
    key: str,
    artifact: str,
    content_errors: list[str],
) -> int:
    value = payload.get(key)
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        content_errors.append(
            f"{artifact}: {key} must be a non-negative integer"
        )
        return 0
    return value


def _validate_artifact_record(
    *,
    root: Path,
    record: Any,
    artifact: str,
    content_errors: list[str],
) -> str | None:
    if not isinstance(record, dict):
        content_errors.append(f"{artifact}: artifact record must be an object")
        return None
    relative_path = _safe_relative_path(record.get("path"))
    if relative_path is None:
        content_errors.append(
            f"{artifact}: artifact record has an empty or unsafe path"
        )
        return None
    size_bytes = record.get("size_bytes")
    if (
        isinstance(size_bytes, bool)
        or not isinstance(size_bytes, int)
        or size_bytes < 0
    ):
        content_errors.append(
            f"{artifact}: {relative_path} has invalid size_bytes"
        )
        return relative_path
    expected_sha256 = str(record.get("sha256") or "").strip().lower()
    if len(expected_sha256) != 64 or any(
        character not in "0123456789abcdef" for character in expected_sha256
    ):
        content_errors.append(
            f"{artifact}: {relative_path} has invalid sha256"
        )
        return relative_path
    path = root / relative_path
    if not path.is_file():
        content_errors.append(
            f"{artifact}: recorded artifact is missing: {relative_path}"
        )
        return relative_path
    try:
        actual_size = path.stat().st_size
        actual_sha256 = _sha256(path)
    except OSError as exc:
        content_errors.append(
            f"{artifact}: cannot verify {relative_path}: "
            f"{type(exc).__name__}: {exc}"
        )
        return relative_path
    if actual_size != size_bytes:
        content_errors.append(
            f"{artifact}: size mismatch for {relative_path}: "
            f"{actual_size} != {size_bytes}"
        )
    if actual_sha256 != expected_sha256:
        content_errors.append(
            f"{artifact}: sha256 mismatch for {relative_path}: "
            f"{actual_sha256} != {expected_sha256}"
        )
    return relative_path


def _validate_literature_artifact_integrity(
    *,
    run_dir: Path,
    manifest_path: Path,
    manifest: dict[str, Any],
    content_errors: list[str],
) -> None:
    artifact_name = manifest_path.relative_to(run_dir).as_posix()
    manifest_parent = manifest_path.parent
    canonical_source_paths: list[str] = []
    for entry in list(manifest.get("entries") or ()):
        if not isinstance(entry, dict):
            continue
        figure_id = str(entry.get("figure_id") or "<missing>")
        entry_artifact = f"{artifact_name}:{figure_id}"
        raw_files = entry.get("artifact_files")
        artifact_files = (
            [
                relative_path
                for value in raw_files
                if (relative_path := _safe_relative_path(value)) is not None
            ]
            if isinstance(raw_files, (list, tuple))
            else []
        )
        records = entry.get("artifact_records")
        if not isinstance(records, (list, tuple)):
            content_errors.append(
                f"{entry_artifact}: artifact_records must be a list"
            )
            continue
        recorded_paths = [
            relative_path
            for record in records
            if (
                relative_path := _validate_artifact_record(
                    root=manifest_parent,
                    record=record,
                    artifact=entry_artifact,
                    content_errors=content_errors,
                )
            )
            is not None
        ]
        if len(recorded_paths) != len(set(recorded_paths)):
            content_errors.append(
                f"{entry_artifact}: duplicate artifact record paths"
            )
        if sorted(artifact_files) != sorted(recorded_paths):
            content_errors.append(
                f"{entry_artifact}: artifact_files and artifact_records "
                "do not declare the same paths"
            )
        for raw_source in list(entry.get("canonical_sources") or ()):
            source_path = _safe_relative_path(raw_source)
            if source_path is None:
                content_errors.append(
                    f"{entry_artifact}: canonical source has an empty or "
                    f"unsafe path {raw_source!r}"
                )
                continue
            canonical_source_paths.append(source_path)

    source_artifacts = manifest.get("source_artifacts")
    if not isinstance(source_artifacts, dict) or not source_artifacts:
        content_errors.append(
            f"{artifact_name}: source_artifacts must be a non-empty object"
        )
        return
    recorded_source_paths: list[str] = []
    for raw_path, record in source_artifacts.items():
        declared_path = _safe_relative_path(raw_path)
        record_path = _validate_artifact_record(
            root=run_dir,
            record=record,
            artifact=f"{artifact_name}:source_artifacts",
            content_errors=content_errors,
        )
        if declared_path is None:
            content_errors.append(
                f"{artifact_name}: source_artifacts has an unsafe key "
                f"{raw_path!r}"
            )
            continue
        recorded_source_paths.append(declared_path)
        if record_path != declared_path:
            content_errors.append(
                f"{artifact_name}: source_artifacts key/path mismatch for "
                f"{declared_path}"
            )
    if len(recorded_source_paths) != len(set(recorded_source_paths)):
        content_errors.append(
            f"{artifact_name}: duplicate source artifact paths"
        )
    if sorted(set(canonical_source_paths)) != sorted(recorded_source_paths):
        content_errors.append(
            f"{artifact_name}: canonical_sources and source_artifacts "
            "do not declare the same paths"
        )


def required_frontend_artifacts(
    *,
    research_run: bool,
    require_rolling: bool,
    require_two_stage_composition_certificate: bool = False,
) -> list[str]:
    """Return the semantic artifact contract for one finalized frontend run."""

    paths = list(BASE_REQUIRED_ARTIFACTS)
    if research_run:
        paths.extend(RESEARCH_PROVENANCE_ARTIFACTS)
    if require_rolling:
        paths.extend(ROLLING_REQUIRED_ARTIFACTS)
    if require_two_stage_composition_certificate:
        paths.extend(TWO_STAGE_FORMAL_REQUIRED_ARTIFACTS)
    return _normalized_paths(paths)


def _validate_thesis_ablation_candidates(
    *,
    run_dir: Path,
    content_errors: list[str],
) -> None:
    """Validate that the automatic postprocessor produced honest candidates."""

    relative_path = "thesis_ablation/day_ahead_method_candidates.json"
    path = run_dir / relative_path
    if not path.is_file():
        return
    try:
        payload = _load_json_object(path)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        content_errors.append(f"{relative_path}: {exc}")
        return
    if payload.get("schema_version") != (
        "thesis_day_ahead_ablation_candidates_v1"
    ):
        content_errors.append(f"{relative_path}: invalid schema_version")
    if payload.get("comparison_scope") != "same_canonical_problem_day_ahead":
        content_errors.append(f"{relative_path}: invalid comparison_scope")
    if payload.get("rolling_costs_mixed_into_comparison") is not False:
        content_errors.append(
            f"{relative_path}: rolling costs must not be mixed into day-ahead candidates"
        )
    if payload.get("additional_solver_invoked_by_postprocessor") is not False:
        content_errors.append(
            f"{relative_path}: postprocessor must not hide an additional solve"
        )
    if payload.get("research_conclusion_eligible") is not False:
        content_errors.append(
            f"{relative_path}: partial candidates cannot be research-conclusion eligible"
        )
    method_rows = list(payload.get("methods") or [])
    by_method = {
        str(row.get("method_id") or ""): row
        for row in method_rows
        if isinstance(row, Mapping)
    }
    if set(by_method) != {"M0", "M1", "M2", "M3"}:
        content_errors.append(
            f"{relative_path}: methods must contain exactly M0, M1, M2, M3"
        )
    primary_structure = str(
        payload.get("primary_optimization_structure") or ""
    ).strip().lower()
    if payload.get("primary_optimization_structure_source") not in {
        "engine_result.solver_metadata",
        "plan.metadata",
    }:
        content_errors.append(
            f"{relative_path}: invalid primary_optimization_structure_source"
        )
    expected_available = {
        "charging_only": {"M0", "M1"},
        "assignment_only": {"M0", "M2"},
        "two_stage": {"M0", "M2"},
        "integrated": {"M0", "M2", "M3"},
    }.get(primary_structure)
    if expected_available is None:
        content_errors.append(
            f"{relative_path}: unsupported primary_optimization_structure"
        )
        expected_available = {"M0"}
    for method_id in expected_available:
        if dict(by_method.get(method_id) or {}).get("candidate_available") is not True:
            content_errors.append(
                f"{relative_path}: {method_id} candidate is unavailable"
            )
    for method_id in set(by_method) - expected_available:
        if dict(by_method.get(method_id) or {}).get("candidate_available") is not False:
            content_errors.append(
                f"{relative_path}: {method_id} is mislabeled as available for "
                f"{primary_structure}"
            )
    if primary_structure != "charging_only":
        m1 = dict(by_method.get("M1") or {})
        if m1.get("construction_status") != "SEPARATE_PHASE1_RUN_REQUIRED":
            content_errors.append(
                f"{relative_path}: M1 must remain an explicit separate-run requirement"
            )
    if primary_structure != "integrated":
        m3 = dict(by_method.get("M3") or {})
        if m3.get("construction_status") != (
            "SEPARATE_PHASE4_INTEGRATED_RUN_REQUIRED"
        ):
            content_errors.append(
                f"{relative_path}: non-integrated primary must not be labeled M3"
            )
    declared_sha = str(payload.get("payload_sha256") or "").strip().lower()
    content_without_sha = dict(payload)
    content_without_sha.pop("payload_sha256", None)
    actual_sha = hashlib.sha256(
        json.dumps(
            content_without_sha,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            default=str,
        ).encode("utf-8")
    ).hexdigest()
    if declared_sha != actual_sha:
        content_errors.append(f"{relative_path}: payload_sha256 mismatch")


def _validate_assignment_economic_audit(
    *,
    run_dir: Path,
    content_errors: list[str],
) -> None:
    """Check the mandatory audit's schema without asserting its conclusions."""

    json_path = run_dir / "assignment_economic_audit.json"
    csv_path = run_dir / "assignment_economic_audit.csv"
    required_keys = {
        "schema_version",
        "bev_grid_marginal_cost_jpy_per_km",
        "ice_marginal_cost_jpy_per_km",
        "renewable_budget_kwh",
        "renewable_energy_allocated_in_stage1_kwh",
        "grid_energy_allocated_in_stage1_kwh",
        "stage1_bev_trip_count",
        "stage2_bev_trip_count",
        "assignment_energy_coupling_mode",
        "weather_response_expected",
        "weather_response_observed",
        "vehicle_usage_cost_jpy_per_used_bus",
        "vehicle_usage_cost_semantics",
        "vehicle_usage_cost_semantics_classified",
        "vehicle_usage_cost_semantics_research_eligible",
    }
    if json_path.is_file():
        try:
            payload = _load_json_object(json_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(f"assignment_economic_audit.json: {exc}")
        else:
            if payload.get("schema_version") != "assignment_economic_audit_v1":
                content_errors.append(
                    "assignment_economic_audit.json has an invalid schema_version"
                )
            missing_keys = sorted(required_keys.difference(payload))
            if missing_keys:
                content_errors.append(
                    "assignment_economic_audit.json is missing required keys: "
                    + ", ".join(missing_keys)
                )
    if csv_path.is_file():
        expected_header = (
            "bev_grid_marginal_cost_jpy_per_km",
            "ice_marginal_cost_jpy_per_km",
            "renewable_budget_kwh",
            "renewable_energy_allocated_in_stage1_kwh",
            "grid_energy_allocated_in_stage1_kwh",
            "stage1_bev_trip_count",
            "stage2_bev_trip_count",
            "assignment_energy_coupling_mode",
            "weather_response_expected",
            "weather_response_observed",
            "vehicle_usage_cost_jpy_per_used_bus",
            "vehicle_usage_cost_semantics",
            "vehicle_usage_cost_semantics_classified",
            "vehicle_usage_cost_semantics_research_eligible",
        )
        try:
            with csv_path.open("r", encoding="utf-8", newline="") as handle:
                reader = csv.DictReader(handle)
                if tuple(reader.fieldnames or ()) != expected_header:
                    content_errors.append(
                        "assignment_economic_audit.csv header does not match "
                        "the required schema"
                    )
                elif len(list(reader)) != 1:
                    content_errors.append(
                        "assignment_economic_audit.csv must contain one audit row"
                    )
        except (OSError, UnicodeError, csv.Error) as exc:
            content_errors.append(f"assignment_economic_audit.csv: {exc}")


def _validate_release_evidence_artifacts(
    *,
    run_dir: Path,
    require_two_stage_composition_certificate: bool,
    content_errors: list[str],
) -> None:
    """Validate the persisted evidence relied on by formal release gates."""

    reconciliation_path = run_dir / SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE
    if reconciliation_path.is_file():
        try:
            reconciliation = _load_json_object(reconciliation_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(
                f"{SOLVER_OBJECTIVE_ACCOUNTING_RECONCILIATION_FILE}: {exc}"
            )
        else:
            content_errors.extend(
                validate_solver_objective_accounting_reconciliation(
                    reconciliation,
                    require_match=False,
                )
            )

    if not require_two_stage_composition_certificate:
        return

    certificate_path = run_dir / STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE
    if certificate_path.is_file():
        try:
            certificate = _load_json_object(certificate_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(
                f"{STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE}: {exc}"
            )
        else:
            content_errors.extend(
                validate_stage1_used_powertrain_composition_search(
                    certificate,
                    require_accepted=True,
                )
            )

    csv_path = run_dir / STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_CSV_FILE
    expected_header = (
        "target_used_bev",
        "minimum_used_bev_count",
        "target_used_ice",
        "delta_used_bev_from_primary",
        "delta_used_ice_from_primary",
        "target_total_used_vehicle_count",
        "target_within_selected_inventory",
        "search_status",
        "solver_status",
        "frontier_status",
        "solution_count",
        "best_bound",
        "mip_gap_ratio",
        "time_limit_sec",
        "solver_runtime_sec",
        "candidate_hash",
        "frontier_target_candidate_physical_validation_feasible",
        "frontier_resolution_source",
        "frontier_resolution_candidate_hash",
        "frontier_resolution_actual_used_bev",
        "frontier_resolution_actual_used_ice",
        "frontier_resolution_canonical_cost_jpy",
        "frontier_resolution_candidate_source_target_used_bev",
        "actual_used_bev",
        "actual_used_ice",
        "candidate_accepted_for_stage2_evaluation",
        "final_disposition",
    )
    if csv_path.is_file():
        try:
            with csv_path.open("r", encoding="utf-8", newline="") as handle:
                reader = csv.DictReader(handle)
                if tuple(reader.fieldnames or ()) != expected_header:
                    content_errors.append(
                        "stage1_used_powertrain_composition_search.csv header "
                        "does not match the required schema"
                    )
        except (OSError, UnicodeError, csv.Error) as exc:
            content_errors.append(
                "stage1_used_powertrain_composition_search.csv: "
                f"{exc}"
            )


def _graph_manifest_artifacts(
    *,
    run_dir: Path,
    content_errors: list[str],
) -> list[str]:
    manifest_path = run_dir / "graph" / "manifest.json"
    if not manifest_path.is_file():
        return []
    try:
        manifest = _load_json_object(manifest_path)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        content_errors.append(f"graph/manifest.json: {exc}")
        return []

    declared: list[str] = []
    for raw_path in list(manifest.get("files") or ()):
        relative_path = _safe_relative_path(raw_path)
        if relative_path is None:
            content_errors.append(
                f"graph/manifest.json: unsafe file path {raw_path!r}"
            )
            continue
        declared.append(f"graph/{relative_path}")
    optional = dict(manifest.get("optional_exports") or {})
    for name, raw_config in optional.items():
        config = dict(raw_config or {})
        if config.get("enabled") is not True:
            continue
        raw_manifest_file = config.get("manifest_file")
        manifest_file = _safe_relative_path(raw_manifest_file)
        if manifest_file is None:
            content_errors.append(
                f"graph/manifest.json: enabled optional export {name!r} "
                f"has an empty or unsafe manifest_file {raw_manifest_file!r}"
            )
            continue
        optional_manifest_relative = (
            Path("graph") / Path(manifest_file)
        ).as_posix()
        declared.append(optional_manifest_relative)
        optional_manifest_path = run_dir / optional_manifest_relative
        if not optional_manifest_path.is_file():
            continue
        try:
            optional_manifest = _load_json_object(
                optional_manifest_path
            )
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(
                f"{optional_manifest_relative}: {exc}"
            )
            continue
        optional_parent = Path(optional_manifest_relative).parent
        for entry in list(optional_manifest.get("entries") or ()):
            if not isinstance(entry, dict):
                content_errors.append(
                    f"{optional_manifest_relative}: non-object entry"
                )
                continue
            raw_artifact_files = entry.get("artifact_files")
            if raw_artifact_files is not None:
                if not isinstance(raw_artifact_files, (list, tuple)):
                    content_errors.append(
                        f"{optional_manifest_relative}: artifact_files must "
                        "be a list"
                    )
                    continue
                if not raw_artifact_files:
                    content_errors.append(
                        f"{optional_manifest_relative}: artifact_files is "
                        "empty"
                    )
                    continue
                for raw_artifact_file in raw_artifact_files:
                    artifact_file = _safe_relative_path(raw_artifact_file)
                    if artifact_file is None:
                        content_errors.append(
                            f"{optional_manifest_relative}: entry has an "
                            "empty or unsafe artifact file "
                            f"{raw_artifact_file!r}"
                        )
                        continue
                    declared.append(
                        (optional_parent / Path(artifact_file)).as_posix()
                    )
                continue
            raw_diagram_file = entry.get("diagram_file")
            diagram_file = _safe_relative_path(raw_diagram_file)
            if diagram_file is None:
                content_errors.append(
                    f"{optional_manifest_relative}: entry has an empty or "
                    f"unsafe diagram_file {raw_diagram_file!r}"
                )
                continue
            declared.append(
                (optional_parent / Path(diagram_file)).as_posix()
            )
    return _normalized_paths(declared)


def _rolling_step_artifacts(
    *,
    run_dir: Path,
    content_errors: list[str],
) -> list[str]:
    summary_path = (
        run_dir / "rolling_hourly_chain" / "rolling_chain_summary.json"
    )
    if not summary_path.is_file():
        return []
    try:
        summary = _load_json_object(summary_path)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        content_errors.append(
            f"rolling_hourly_chain/rolling_chain_summary.json: {exc}"
        )
        return []

    expected_step_count = int(summary.get("expected_step_count") or 0)
    step_count = int(summary.get("step_count") or 0)
    step_dirs = sorted(
        path
        for path in (run_dir / "rolling_hourly_chain").glob("step_*")
        if path.is_dir()
    )
    if expected_step_count <= 0:
        content_errors.append(
            "rolling_chain_summary.expected_step_count must be positive"
        )
    if step_count != expected_step_count:
        content_errors.append(
            "rolling_chain_summary step_count does not equal "
            f"expected_step_count: {step_count} != {expected_step_count}"
        )
    if len(step_dirs) != expected_step_count:
        content_errors.append(
            "rolling step directory count does not equal expected_step_count: "
            f"{len(step_dirs)} != {expected_step_count}"
        )
    if summary.get("chain_accepted") is not True:
        content_errors.append("rolling_chain_summary.chain_accepted is not true")
    if summary.get("all_steps_feasible") is not True:
        content_errors.append(
            "rolling_chain_summary.all_steps_feasible is not true"
        )

    artifacts: list[str] = []
    for step_index, step_dir in enumerate(step_dirs):
        required_step_files = [
            "hourly_solver_result.json",
            "hourly_summary.json",
        ]
        # The final slot has no successor handoff. Every earlier step must
        # persist its state transition for the next solve.
        if step_index < len(step_dirs) - 1:
            required_step_files.append("state_for_next_hour.json")
        for filename in required_step_files:
            artifacts.append(
                (step_dir.relative_to(run_dir) / filename).as_posix()
            )
    return _normalized_paths(artifacts)


def _normalize_refueling_row(
    row: Any,
    *,
    context: str,
    require_location: bool,
) -> tuple[str, int, str, Decimal, str]:
    """Return one refueling event in the canonical comparison representation."""

    if not isinstance(row, dict):
        raise ValueError(f"{context} must be an object")

    vehicle_id = str(row.get("vehicle_id") or "").strip()
    time_hhmm = str(row.get("time_hhmm") or "").strip()
    if not vehicle_id:
        raise ValueError(f"{context}.vehicle_id is empty")
    if not time_hhmm:
        raise ValueError(f"{context}.time_hhmm is empty")

    raw_slot_index = row.get("slot_index")
    if isinstance(raw_slot_index, bool):
        raise ValueError(f"{context}.slot_index is not an integer")
    try:
        slot_index = int(str(raw_slot_index).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{context}.slot_index is not an integer") from exc
    if slot_index < 0:
        raise ValueError(f"{context}.slot_index is negative")

    raw_liters = row.get("refuel_liters")
    if isinstance(raw_liters, bool):
        raise ValueError(f"{context}.refuel_liters is not finite")
    try:
        refuel_liters = Decimal(str(raw_liters).strip())
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"{context}.refuel_liters is not finite") from exc
    if not refuel_liters.is_finite() or refuel_liters < Decimal("0"):
        raise ValueError(f"{context}.refuel_liters is not finite or is negative")

    location_id = str(row.get("location_id") or "").strip()
    if require_location and not location_id:
        raise ValueError(f"{context}.location_id is empty")
    return vehicle_id, slot_index, time_hhmm, refuel_liters, location_id


def _read_refueling_csv_rows(
    *,
    path: Path,
    relative_path: str,
    fieldnames: tuple[str, ...],
    require_location: bool,
    require_unit: bool,
    content_errors: list[str],
) -> list[tuple[str, int, str, Decimal, str]] | None:
    """Read a refueling CSV strictly enough to compare it to canonical rows."""

    try:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            if tuple(reader.fieldnames or ()) != fieldnames:
                content_errors.append(
                    f"{relative_path}: CSV header does not match the required schema"
                )
                return None
            normalized_rows: list[tuple[str, int, str, Decimal, str]] = []
            for row_index, row in enumerate(reader, start=2):
                if None in row or any(value is None for value in row.values()):
                    content_errors.append(
                        f"{relative_path}: row {row_index} has an unexpected column count"
                    )
                    return None
                if require_unit and str(row.get("unit") or "").strip() != "L":
                    content_errors.append(
                        f"{relative_path}: row {row_index} unit is not 'L'"
                    )
                    return None
                normalized_rows.append(
                    _normalize_refueling_row(
                        row,
                        context=f"{relative_path} row {row_index}",
                        require_location=require_location,
                    )
                )
    except (OSError, UnicodeError, csv.Error, ValueError) as exc:
        content_errors.append(f"{relative_path}: {exc}")
        return None
    return normalized_rows


def _validate_refueling_event_exports(
    *,
    run_dir: Path,
    canonical: dict[str, Any],
    content_errors: list[str],
) -> None:
    """Bind both report refueling CSVs to canonical solver-native events."""

    root_export_path = run_dir / "refuel_events.csv"
    graph_export_path = run_dir / "graph" / "refuel_events.csv"
    if not root_export_path.is_file() or not graph_export_path.is_file():
        # Required-artifact checks report absent paths. Avoid duplicate errors
        # here while retaining strict comparison whenever both sources exist.
        return
    if any(path.stat().st_size <= 0 for path in (root_export_path, graph_export_path)):
        # Required-artifact checks report zero-byte paths.
        return

    raw_canonical_rows = canonical.get("refueling_schedule")
    if not isinstance(raw_canonical_rows, list):
        content_errors.append(
            "canonical_solver_result.refueling_schedule must be a list"
        )
        return
    try:
        canonical_rows = [
            _normalize_refueling_row(
                row,
                context=f"canonical_solver_result.refueling_schedule[{index}]",
                require_location=True,
            )
            for index, row in enumerate(raw_canonical_rows)
        ]
    except ValueError as exc:
        content_errors.append(str(exc))
        return

    root_rows = _read_refueling_csv_rows(
        path=root_export_path,
        relative_path="refuel_events.csv",
        fieldnames=_ROOT_REFUEL_EVENT_FIELDS,
        require_location=False,
        require_unit=True,
        content_errors=content_errors,
    )
    graph_rows = _read_refueling_csv_rows(
        path=graph_export_path,
        relative_path="graph/refuel_events.csv",
        fieldnames=_GRAPH_REFUEL_EVENT_FIELDS,
        require_location=True,
        require_unit=False,
        content_errors=content_errors,
    )
    if root_rows is not None and Counter(row[:4] for row in root_rows) != Counter(
        row[:4] for row in canonical_rows
    ):
        content_errors.append(
            "refuel_events.csv rows do not exactly match "
            "canonical_solver_result.refueling_schedule"
        )
    if graph_rows is not None and Counter(graph_rows) != Counter(canonical_rows):
        content_errors.append(
            "graph/refuel_events.csv rows do not exactly match "
            "canonical_solver_result.refueling_schedule"
        )


def _validate_rolling_content(
    *,
    run_dir: Path,
    content_errors: list[str],
) -> None:
    executed_path = (
        run_dir / "rolling_hourly_chain" / "executed_day_accounting.json"
    )
    if executed_path.is_file():
        try:
            executed = _load_json_object(executed_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(
                f"rolling_hourly_chain/executed_day_accounting.json: {exc}"
            )
        else:
            if executed.get("eligible") is not True:
                content_errors.append(
                    "executed_day_accounting.eligible is not true"
                )
            expected_slots = int(executed.get("expected_slot_count") or 0)
            executed_slots = int(executed.get("executed_slot_count") or 0)
            if expected_slots <= 0 or executed_slots != expected_slots:
                content_errors.append(
                    "executed-day slot coverage is incomplete: "
                    f"{executed_slots} != {expected_slots}"
                )

    input_manifest_path = run_dir / "physical_validation_input_manifest.json"
    if input_manifest_path.is_file():
        try:
            input_manifest = _load_json_object(input_manifest_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(
                f"physical_validation_input_manifest.json: {exc}"
            )
        else:
            if (
                input_manifest.get("schema_version")
                != PHYSICAL_VALIDATION_INPUT_MANIFEST_SCHEMA_VERSION
            ):
                content_errors.append(
                    "physical_validation_input_manifest schema_version is invalid"
                )
            expected_sources = {
                "assignment_source": "canonical_solver_result.json",
                "charging_source": (
                    "rolling_hourly_chain/charging_schedule.csv"
                ),
                "refueling_source": "canonical_solver_result.json",
            }
            for key, expected_source in expected_sources.items():
                if input_manifest.get(key) != expected_source:
                    content_errors.append(
                        "physical_validation_input_manifest "
                        f"{key} is not {expected_source!r}"
                    )
            for key, source in (
                ("canonical_solver_result_sha256", "canonical_solver_result.json"),
                (
                    "executed_charging_schedule_sha256",
                    "rolling_hourly_chain/charging_schedule.csv",
                ),
            ):
                declared_sha = str(input_manifest.get(key) or "").strip().lower()
                source_path = run_dir / source
                if (
                    len(declared_sha) != 64
                    or not all(
                        character in "0123456789abcdef"
                        for character in declared_sha
                    )
                ):
                    content_errors.append(
                        "physical_validation_input_manifest "
                        f"{key} is not a SHA-256 digest"
                    )
                elif (
                    source_path.is_file()
                    and declared_sha != _sha256(source_path)
                ):
                    content_errors.append(
                        "physical_validation_input_manifest "
                        f"{key} does not match {source}"
                    )
            assignment_hash = str(
                input_manifest.get("day_ahead_assignment_hash") or ""
            ).strip().lower()
            if (
                len(assignment_hash) != 64
                or not all(
                    character in "0123456789abcdef"
                    for character in assignment_hash
                )
            ):
                content_errors.append(
                    "physical_validation_input_manifest "
                    "day_ahead_assignment_hash is not a SHA-256 digest"
                )
            chain_summary_path = (
                run_dir
                / "rolling_hourly_chain"
                / "rolling_chain_summary.json"
            )
            canonical_path = run_dir / "canonical_solver_result.json"
            try:
                chain_summary = _load_json_object(chain_summary_path)
                canonical = _load_json_object(canonical_path)
            except (OSError, ValueError, json.JSONDecodeError) as exc:
                content_errors.append(
                    "physical_validation_input_manifest provenance source "
                    f"is unreadable: {exc}"
                )
            else:
                chain_canonical_sha = str(
                    chain_summary.get("day_ahead_result_sha256") or ""
                ).strip().lower()
                chain_assignment_hash = str(
                    chain_summary.get("day_ahead_assignment_hash") or ""
                ).strip().lower()
                if (
                    len(chain_canonical_sha) != 64
                    or not all(
                        character in "0123456789abcdef"
                        for character in chain_canonical_sha
                    )
                ):
                    content_errors.append(
                        "rolling_chain_summary day_ahead_result_sha256 "
                        "is not a SHA-256 digest"
                    )
                elif chain_canonical_sha != _sha256(canonical_path):
                    content_errors.append(
                        "rolling_chain_summary day_ahead_result_sha256 does "
                        "not match canonical_solver_result.json"
                    )
                elif (
                    input_manifest.get("canonical_solver_result_sha256")
                    != chain_canonical_sha
                ):
                    content_errors.append(
                        "physical_validation_input_manifest "
                        "canonical_solver_result_sha256 does not match "
                        "rolling_chain_summary"
                    )
                if (
                    len(chain_assignment_hash) != 64
                    or not all(
                        character in "0123456789abcdef"
                        for character in chain_assignment_hash
                    )
                ):
                    content_errors.append(
                        "rolling_chain_summary day_ahead_assignment_hash "
                        "is not a SHA-256 digest"
                    )
                elif assignment_hash != chain_assignment_hash:
                    content_errors.append(
                        "physical_validation_input_manifest "
                        "day_ahead_assignment_hash does not match "
                        "rolling_chain_summary"
                    )

                raw_paths = canonical.get("vehicle_paths")
                served_trip_ids = canonical.get("served_trip_ids")
                unserved_trip_ids = canonical.get("unserved_trip_ids")
                if not isinstance(raw_paths, dict):
                    content_errors.append(
                        "canonical_solver_result.vehicle_paths must be an object"
                    )
                elif not isinstance(served_trip_ids, list):
                    content_errors.append(
                        "canonical_solver_result.served_trip_ids must be a list"
                    )
                elif not isinstance(unserved_trip_ids, list):
                    content_errors.append(
                        "canonical_solver_result.unserved_trip_ids must be a list"
                    )
                elif not all(
                    isinstance(trip_ids, list)
                    for trip_ids in raw_paths.values()
                ):
                    content_errors.append(
                        "canonical_solver_result.vehicle_paths values must be lists"
                    )
                else:
                    assigned_trip_ids = [
                        str(trip_id)
                        for trip_ids in raw_paths.values()
                        for trip_id in trip_ids
                    ]
                    normalized_served_trip_ids = [
                        str(trip_id) for trip_id in served_trip_ids
                    ]
                    actual_counts = {
                        "vehicle_path_count": len(raw_paths),
                        "assigned_trip_occurrence_count": len(
                            assigned_trip_ids
                        ),
                        "served_trip_occurrence_count": len(
                            normalized_served_trip_ids
                        ),
                        "problem_trip_count": len(normalized_served_trip_ids)
                        + len(unserved_trip_ids),
                        "unserved_trip_count": len(unserved_trip_ids),
                    }
                    for key, actual_count in actual_counts.items():
                        if input_manifest.get(key) != actual_count:
                            content_errors.append(
                                "physical_validation_input_manifest "
                                f"{key} does not match canonical_solver_result"
                            )
                    if Counter(assigned_trip_ids) != Counter(
                        normalized_served_trip_ids
                    ):
                        content_errors.append(
                            "canonical_solver_result vehicle_paths and "
                            "served_trip_ids disagree"
                        )
                    if unserved_trip_ids:
                        content_errors.append(
                            "canonical_solver_result unserved_trip_ids is not empty"
                        )
                _validate_refueling_event_exports(
                    run_dir=run_dir,
                    canonical=canonical,
                    content_errors=content_errors,
                )
            counts: dict[str, int] = {}
            for key in (
                "vehicle_path_count",
                "assigned_trip_occurrence_count",
                "served_trip_occurrence_count",
                "problem_trip_count",
                "unserved_trip_count",
            ):
                value = input_manifest.get(key)
                if (
                    isinstance(value, bool)
                    or not isinstance(value, int)
                    or value < 0
                ):
                    content_errors.append(
                        "physical_validation_input_manifest "
                        f"{key} must be a non-negative integer"
                    )
                    continue
                counts[key] = value
            if counts and (
                counts.get("assigned_trip_occurrence_count")
                != counts.get("served_trip_occurrence_count")
                or counts.get("assigned_trip_occurrence_count")
                != counts.get("problem_trip_count")
                or counts.get("unserved_trip_count") != 0
            ):
                content_errors.append(
                    "physical_validation_input_manifest trip coverage counts "
                    "are inconsistent"
                )
            validation_contract = input_manifest.get("validation_contract")
            required_contract_checks = (
                "canonical_sha_matches_rolling_chain",
                "vehicle_paths_match_served_trip_ids",
                "vehicle_paths_cover_problem_trips_exactly",
                "unserved_trip_ids_empty",
                "executed_charging_overlay_only",
            )
            if not isinstance(validation_contract, dict) or any(
                validation_contract.get(key) is not True
                for key in required_contract_checks
            ):
                content_errors.append(
                    "physical_validation_input_manifest validation_contract "
                    "is incomplete or not verified"
                )

    physical_path = run_dir / "physical_schedule_validation.json"
    if physical_path.is_file():
        try:
            physical = _load_json_object(physical_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(
                f"physical_schedule_validation.json: {exc}"
            )
        else:
            if physical.get("accepted") is not True:
                content_errors.append(
                    "physical_schedule_validation.accepted is not true"
                )

    reconciliation_path = run_dir / "final_cost_reconciliation.json"
    if reconciliation_path.is_file():
        try:
            reconciliation = _load_json_object(reconciliation_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(f"final_cost_reconciliation.json: {exc}")
        else:
            if reconciliation.get("status") != "OK":
                content_errors.append(
                    "final_cost_reconciliation.status is not OK"
                )

    literature_manifest_path = (
        run_dir / "graph" / "literature_figures" / "manifest.json"
    )
    if literature_manifest_path.is_file():
        try:
            literature_manifest = _load_json_object(
                literature_manifest_path
            )
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            content_errors.append(
                f"graph/literature_figures/manifest.json: {exc}"
            )
        else:
            if (
                literature_manifest.get("schema_version")
                != "literature_figure_bundle_v1"
            ):
                content_errors.append(
                    "literature figure manifest schema_version is invalid"
                )
            if literature_manifest.get("status") != "READY":
                content_errors.append(
                    "literature figure manifest status is not READY"
                )
            figure_count = _required_manifest_count(
                literature_manifest,
                key="figure_count",
                artifact="graph/literature_figures/manifest.json",
                content_errors=content_errors,
            )
            entries = [
                entry
                for entry in list(
                    literature_manifest.get("entries") or ()
                )
                if isinstance(entry, dict)
                and entry.get("kind") == "figure"
            ]
            if figure_count < 5 or len(entries) != figure_count:
                content_errors.append(
                    "literature figure manifest does not declare all five "
                    "required figures"
                )
            for entry in entries:
                artifact_files = list(entry.get("artifact_files") or ())
                if not {
                    Path(str(path)).suffix.lower()
                    for path in artifact_files
                }.issuperset({".png", ".svg", ".csv"}):
                    content_errors.append(
                        "literature figure entry is missing PNG, SVG, or "
                        f"source CSV: {entry.get('figure_id')!r}"
                    )
            raw_entries = [
                entry
                for entry in list(
                    literature_manifest.get("entries") or ()
                )
                if isinstance(entry, dict)
                and entry.get("kind") == "raw_data_bundle"
            ]
            raw_data_csv_count = _required_manifest_count(
                literature_manifest,
                key="raw_data_csv_count",
                artifact="graph/literature_figures/manifest.json",
                content_errors=content_errors,
            )
            if raw_data_csv_count < 16 or len(raw_entries) != 1:
                content_errors.append(
                    "literature figure manifest does not declare the "
                    "analysis-ready raw CSV bundle"
                )
            elif len(
                [
                    path
                    for path in list(
                        raw_entries[0].get("artifact_files") or ()
                    )
                    if Path(str(path)).suffix.lower() == ".csv"
                ]
            ) != raw_data_csv_count:
                content_errors.append(
                    "literature raw-data CSV count does not match its "
                    "declared artifact files"
                )
            if (
                literature_manifest.get("raw_data_catalog")
                != "raw_data/raw_data_catalog.csv"
            ):
                content_errors.append(
                    "literature raw-data catalog path is missing or invalid"
                )
            _validate_literature_artifact_integrity(
                run_dir=run_dir,
                manifest_path=literature_manifest_path,
                manifest=literature_manifest,
                content_errors=content_errors,
            )


def audit_frontend_run_artifacts(
    run_dir: Path,
    *,
    research_run: bool,
    require_rolling: bool,
    require_two_stage_composition_certificate: bool = False,
) -> dict[str, Any]:
    """Audit files and essential content without changing solver results."""

    run_dir = Path(run_dir).resolve()
    content_errors: list[str] = []
    _validate_assignment_economic_audit(
        run_dir=run_dir,
        content_errors=content_errors,
    )
    _validate_thesis_ablation_candidates(
        run_dir=run_dir,
        content_errors=content_errors,
    )
    _validate_release_evidence_artifacts(
        run_dir=run_dir,
        require_two_stage_composition_certificate=(
            require_two_stage_composition_certificate
        ),
        content_errors=content_errors,
    )
    required = required_frontend_artifacts(
        research_run=research_run,
        require_rolling=require_rolling,
        require_two_stage_composition_certificate=(
            require_two_stage_composition_certificate
        ),
    )
    composition_path = run_dir / STAGE1_USED_POWERTRAIN_COMPOSITION_SEARCH_FILE
    if require_two_stage_composition_certificate and composition_path.is_file():
        try:
            composition_payload = _load_json_object(composition_path)
        except (OSError, ValueError, json.JSONDecodeError):
            composition_payload = {}
        if (
            composition_payload.get("search_mode")
            == "minimum_used_bev_frontier"
        ):
            required.extend(BEV_FRONTIER_REQUIRED_ARTIFACTS)
    required.extend(
        _graph_manifest_artifacts(
            run_dir=run_dir,
            content_errors=content_errors,
        )
    )
    if require_rolling:
        required.extend(
            _rolling_step_artifacts(
                run_dir=run_dir,
                content_errors=content_errors,
            )
        )
        _validate_rolling_content(
            run_dir=run_dir,
            content_errors=content_errors,
        )
    required = _normalized_paths(required)

    missing: list[str] = []
    empty: list[str] = []
    invalid_json: dict[str, str] = {}
    artifacts: dict[str, dict[str, Any]] = {}
    for relative_path in required:
        path = run_dir / relative_path
        if not path.is_file():
            missing.append(relative_path)
            continue
        size_bytes = path.stat().st_size
        if size_bytes <= 0:
            empty.append(relative_path)
            continue
        if path.suffix.lower() == ".json":
            try:
                json.loads(path.read_text(encoding="utf-8"))
            except (OSError, ValueError, json.JSONDecodeError) as exc:
                invalid_json[relative_path] = str(exc)
                continue
        artifacts[relative_path] = {
            "size_bytes": size_bytes,
            "sha256": _sha256(path),
        }

    workbook_errors: list[str] = []
    workbook_path = run_dir / "results.xlsx"
    if workbook_path.is_file() and workbook_path.stat().st_size > 0:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                workbook_path,
                read_only=True,
                data_only=False,
            )
            try:
                missing_sheets = sorted(
                    set(REQUIRED_WORKBOOK_SHEETS)
                    - set(workbook.sheetnames)
                )
            finally:
                workbook.close()
            if missing_sheets:
                workbook_errors.append(
                    "results.xlsx missing sheets: "
                    + ", ".join(missing_sheets)
                )
        except Exception as exc:
            workbook_errors.append(f"results.xlsx unreadable: {exc}")

    run_manifest_errors: list[str] = []
    run_manifest_path = run_dir / "run_manifest.json"
    if run_manifest_path.is_file():
        try:
            run_manifest = _load_json_object(run_manifest_path)
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            run_manifest_errors.append(f"run_manifest.json: {exc}")
        else:
            declared_files = {
                Path(str(path)).as_posix()
                for path in list(run_manifest.get("files") or ())
            }
            undeclared = sorted(
                path
                for path in required
                if path != "run_manifest.json" and path not in declared_files
            )
            if undeclared:
                run_manifest_errors.append(
                    "run_manifest.files omits required artifacts: "
                    + ", ".join(undeclared)
                )

    accepted = not any(
        (
            missing,
            empty,
            invalid_json,
            content_errors,
            workbook_errors,
            run_manifest_errors,
        )
    )
    return {
        "schema_version": ARTIFACT_CONTRACT_VERSION,
        "status": "OK" if accepted else "ERROR",
        "accepted": accepted,
        "run_dir": str(run_dir),
        "research_run": bool(research_run),
        "rolling_required": bool(require_rolling),
        "two_stage_composition_certificate_required": bool(
            require_two_stage_composition_certificate
        ),
        "required_artifact_count": len(required),
        "verified_artifact_count": len(artifacts),
        "total_file_count": sum(
            1 for path in run_dir.rglob("*") if path.is_file()
        ),
        "required_artifacts": required,
        "artifacts": artifacts,
        "missing_artifacts": missing,
        "empty_artifacts": empty,
        "invalid_json_artifacts": invalid_json,
        "content_errors": content_errors,
        "workbook_errors": workbook_errors,
        "run_manifest_errors": run_manifest_errors,
        "semantics": (
            "This gate proves required frontend output files are present, "
            "readable, and internally eligible where specified. It does not "
            "upgrade research acceptance or global optimality."
        ),
    }


def persist_frontend_run_artifact_audit(
    run_dir: Path,
    *,
    research_run: bool,
    require_rolling: bool,
    require_two_stage_composition_certificate: bool = False,
) -> dict[str, Any]:
    """Audit and persist ``artifact_completeness.json`` in the run root."""

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=research_run,
        require_rolling=require_rolling,
        require_two_stage_composition_certificate=(
            require_two_stage_composition_certificate
        ),
    )
    (Path(run_dir) / "artifact_completeness.json").write_text(
        json.dumps(audit, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return audit
