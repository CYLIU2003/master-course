from __future__ import annotations

import csv
import json

from bff.routers.optimization import _apply_invalid_result_kpi_gate
from src.optimization.rolling.acceptance import (
    ROLLING_CHAIN_REQUIRED_ACCEPTANCE_CHECKS,
)
from src.reporting.canonical_reporting import apply_solution_validity_gate


def test_api_result_gate_uses_canonical_counts_and_nulls_evaluation_metrics() -> None:
    optimization_result = {
        "solver_status": "infeasible",
        "objective_value": 0.0,
        "solution_validity": {
            "validated_feasible": False,
            "status_reason": "postsolve_infeasible",
        },
        "solver_settings": {
            "stage1_solver_status": "time_limit",
            "stage2_solver_status": "infeasible",
        },
        "summary": {"trip_count_served": 0, "trip_count_unserved": 0},
        "cost_breakdown": {
            "total_cost": 0.0,
            "fuel_cost": 0.0,
            "vehicle_usage_cost": 640_000.0,
            "objective_is_actual_cost": True,
        },
        "graph_artifacts": {
            "accounting_summary": {
                "total_cost_jpy": 0.0,
                "served_trip_count": 0,
                "unserved_trip_count": 0,
            }
        },
    }
    canonical = {
        "feasible": False,
        "trip_count_served": 0,
        "trip_count_unserved": 264,
        "unserved_trip_ids": [f"trip-{index}" for index in range(264)],
    }

    _apply_invalid_result_kpi_gate(optimization_result, canonical)

    assert optimization_result["result_status"] == "INFEASIBLE"
    assert optimization_result["failure_stage"] == "stage2_energy_dispatch"
    assert optimization_result["objective_value"] is None
    assert optimization_result["cost_breakdown"]["total_cost"] is None
    assert optimization_result["cost_breakdown"]["vehicle_usage_cost"] is None
    assert optimization_result["cost_breakdown"]["objective_is_actual_cost"] is False
    assert optimization_result["summary"]["trip_count_unserved"] == 264
    accounting = optimization_result["graph_artifacts"]["accounting_summary"]
    assert accounting["total_cost_jpy"] is None
    assert accounting["unserved_trip_count"] == 264


def test_reporting_gate_marks_infeasible_zero_ledgers_as_invalid(tmp_path) -> None:
    graph_dir = tmp_path / "graph"
    graph_dir.mkdir()
    (tmp_path / "canonical_solver_result.json").write_text(
        json.dumps(
            {
                "solver_status": "infeasible",
                "feasible": False,
                "trip_count_served": 0,
                "trip_count_unserved": 264,
                "unserved_trip_ids": [f"trip-{index}" for index in range(264)],
                "solution_validity": {
                    "validated_feasible": False,
                    "status_reason": "postsolve_infeasible",
                },
            }
        ),
        encoding="utf-8",
    )
    (tmp_path / "solver_settings.json").write_text(
        json.dumps(
            {
                "stage1_solver_status": "time_limit",
                "stage2_solver_status": "infeasible",
            }
        ),
        encoding="utf-8",
    )
    summary = {
        "solver_status": "infeasible",
        "total_cost_jpy": 0.0,
        "objective_value_jpy": 0.0,
        "trip_count_served": 0,
        "trip_count_unserved": 0,
        "solution_validity": {
            "validated_feasible": False,
            "status_reason": "postsolve_infeasible",
        },
    }
    kpi = {
        "solver_status": "infeasible",
        "total_cost_jpy": 0.0,
        "pv_generation_kwh": 614.7,
        "pv_to_bus_kwh": 0.0,
        "grid_import_kwh": 0.0,
        "total_co2_kg": 0.0,
        "served_trip_count": 0,
        "unserved_trip_count": 0,
        "energy": {"pv_generation_kwh": 614.7, "pv_to_bus_kwh": 0.0},
        "cost": {"total_cost_jpy": 0.0},
    }
    (tmp_path / "summary.json").write_text(json.dumps(summary), encoding="utf-8")
    (graph_dir / "kpi_summary.json").write_text(json.dumps(kpi), encoding="utf-8")
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        Workbook = None
        load_workbook = None
    if Workbook is not None:
        workbook = Workbook()
        workbook.active.title = "summary"
        workbook.active.append(["key", "value", "unit"])
        workbook.active.append(["objective_value", 0.0, "JPY"])
        workbook.active.append(["trip_count_served", 0, "trips"])
        workbook.active.append(["trip_count_unserved", 0, "trips"])
        cost_sheet = workbook.create_sheet("cost_breakdown")
        cost_sheet.append(["key", "value", "unit"])
        cost_sheet.append(["total_cost", 0.0, "JPY"])
        cost_sheet.append(["objective_is_actual_cost", True, ""])
        workbook.save(tmp_path / "results.xlsx")
    (tmp_path / "experiment_report.md").write_text(
        "# Old report\n\nTotal cost: 0 JPY\n", encoding="utf-8"
    )
    validation_path = graph_dir / "data_flow_validation.csv"
    validation_fields = [
        "check_name",
        "status",
        "expected_value",
        "actual_value",
        "difference",
        "tolerance",
        "severity",
        "message",
        "source_files",
    ]
    with validation_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=validation_fields)
        writer.writeheader()
        writer.writerow(
            {
                "check_name": "zero_ledger_balance",
                "status": "OK",
                "severity": "INFO",
            }
        )

    gated_summary, gated_kpi, rows = apply_solution_validity_gate(
        tmp_path,
        summary,
        kpi,
        [{"check_name": "zero_ledger_balance", "status": "OK", "severity": "INFO"}],
    )

    assert gated_summary["result_status"] == "INFEASIBLE"
    assert gated_summary["failure_stage"] == "stage2_energy_dispatch"
    assert gated_summary["total_cost_jpy"] is None
    assert gated_summary["trip_count_unserved"] == 264
    assert gated_kpi["pv_generation_kwh"] == 614.7
    assert gated_kpi["pv_to_bus_kwh"] is None
    assert gated_kpi["grid_import_kwh"] is None
    assert gated_kpi["total_co2_kg"] is None
    assert gated_kpi["research_kpi_eligible"] is False
    assert any(
        row["check_name"] == "solution_validity_gate"
        and row["status"] == "NG"
        and row["severity"] == "ERROR"
        for row in rows
    )
    if load_workbook is not None:
        gated_workbook = load_workbook(tmp_path / "results.xlsx", data_only=True)
        assert gated_workbook["result_status"]["B2"].value == "INFEASIBLE"
        assert gated_workbook["summary"]["B2"].value is None
        assert gated_workbook["summary"]["B4"].value == 264
        assert gated_workbook["cost_breakdown"]["B2"].value is None
        assert gated_workbook["cost_breakdown"]["B3"].value is False
    report = (tmp_path / "experiment_report.md").read_text(encoding="utf-8")
    assert report.startswith("<!-- solution-validity-gate -->")
    assert "research_kpi_eligible: `false`" in report


def test_export_experiment_report_blocks_submission_without_provenance(tmp_path):
    from dataclasses import dataclass

    from src.result_exporter import export_experiment_report

    @dataclass
    class _ProblemData:
        delta_t_min: float = 15.0
        num_periods: int = 96
        enable_pv: bool = True
        enable_v2g: bool = False
        enable_demand_charge: bool = True

    @dataclass
    class _VehicleSet:
        K_BEV: list = ()
        K_ICE: list = ()
        R: list = ()
        C: list = ()

    @dataclass
    class _MILP:
        status: str = "INFEASIBLE"
        objective_value: object = None
        mip_gap: object = None
        solve_time_sec: float = 0.0
        infeasibility_info: str = "strict coverage infeasible"

    @dataclass
    class _Sim:
        provisional_energy_cost: float = 0.0
        charged_energy_cost: float = 0.0
        total_energy_cost: float = 0.0
        total_demand_charge: float = 0.0
        total_fuel_cost: float = 0.0
        total_degradation_cost: float = 0.0
        total_operating_cost: float = 0.0
        served_task_ratio: float = 0.0
        unserved_tasks: list = ("trip-1",)
        total_grid_kwh: float = 0.0
        total_pv_kwh: float = 0.0
        pv_self_consumption_ratio: float = 0.0
        peak_demand_kw: float = 0.0
        total_co2_kg: float = 0.0
        soc_min_kwh: float = 0.0
        soc_violations: list = ()

    export_experiment_report(
        tmp_path,
        _ProblemData(),
        _VehicleSet(),
        _MILP(),
        _Sim(),
        run_label="test",
    )
    report = (tmp_path / "experiment_report.md").read_text(encoding="utf-8")

    assert "> EXPLORATORY — RESEARCH SUBMISSION BLOCKED" in report
    assert "research_submission_ready: False" in report
    assert "input_provenance_ready: False" in report
    assert "solver_status_not_research_feasible" in report
    assert "unserved_trips_remain" in report


def test_export_experiment_report_blocks_submission_without_rolling_chain(tmp_path):
    from dataclasses import dataclass

    from src.result_exporter import export_experiment_report

    (tmp_path / "manifest.json").write_text("{}", encoding="utf-8")
    (tmp_path / "input_audit.json").write_text("{}", encoding="utf-8")

    @dataclass
    class _ProblemData:
        delta_t_min: float = 15.0
        num_periods: int = 96
        enable_pv: bool = True
        enable_v2g: bool = False
        enable_demand_charge: bool = True

    @dataclass
    class _VehicleSet:
        K_BEV: list = ("ev-1",)
        K_ICE: list = ()
        R: list = ("trip-1",)
        C: list = ("chgr-1",)

    @dataclass
    class _MILP:
        status: str = "OPTIMAL"
        objective_value: object = 1.0
        mip_gap: object = 0.0
        solve_time_sec: float = 1.0
        infeasibility_info: str = ""

    @dataclass
    class _Sim:
        provisional_energy_cost: float = 0.0
        charged_energy_cost: float = 0.0
        total_energy_cost: float = 1.0
        total_demand_charge: float = 0.0
        total_fuel_cost: float = 0.0
        total_degradation_cost: float = 0.0
        total_operating_cost: float = 1.0
        served_task_ratio: float = 1.0
        unserved_tasks: list = ()
        total_grid_kwh: float = 0.0
        total_pv_kwh: float = 0.0
        pv_self_consumption_ratio: float = 0.0
        peak_demand_kw: float = 0.0
        total_co2_kg: float = 0.0
        soc_min_kwh: float = 0.0
        soc_violations: list = ()

    export_experiment_report(
        tmp_path,
        _ProblemData(),
        _VehicleSet(),
        _MILP(),
        _Sim(),
        run_label="test",
    )
    report = (tmp_path / "experiment_report.md").read_text(encoding="utf-8")

    assert "EXPLORATORY — RESEARCH SUBMISSION BLOCKED" in report
    assert "research_submission_ready: False" in report
    assert "input_provenance_ready: True" in report
    assert "hourly_rolling_chain_missing" in report


def test_export_experiment_report_blocks_when_rolling_chain_is_not_accepted(tmp_path):
    from dataclasses import dataclass

    from src.result_exporter import export_experiment_report

    (tmp_path / "manifest.json").write_text("{}", encoding="utf-8")
    (tmp_path / "input_audit.json").write_text("{}", encoding="utf-8")
    rolling_dir = tmp_path / "rolling_hourly_chain"
    rolling_dir.mkdir()
    (rolling_dir / "rolling_chain_summary.json").write_text(
        json.dumps({"chain_accepted": False}), encoding="utf-8"
    )

    @dataclass
    class _ProblemData:
        delta_t_min: float = 15.0
        num_periods: int = 96
        enable_pv: bool = True
        enable_v2g: bool = False
        enable_demand_charge: bool = True

    @dataclass
    class _VehicleSet:
        K_BEV: list = ("ev-1",)
        K_ICE: list = ()
        R: list = ("trip-1",)
        C: list = ("chgr-1",)

    @dataclass
    class _MILP:
        status: str = "OPTIMAL"
        objective_value: object = 1.0
        mip_gap: object = 0.0
        solve_time_sec: float = 1.0
        infeasibility_info: str = ""

    @dataclass
    class _Sim:
        provisional_energy_cost: float = 0.0
        charged_energy_cost: float = 0.0
        total_energy_cost: float = 1.0
        total_demand_charge: float = 0.0
        total_fuel_cost: float = 0.0
        total_degradation_cost: float = 0.0
        total_operating_cost: float = 1.0
        served_task_ratio: float = 1.0
        unserved_tasks: list = ()
        total_grid_kwh: float = 0.0
        total_pv_kwh: float = 0.0
        pv_self_consumption_ratio: float = 0.0
        peak_demand_kw: float = 0.0
        total_co2_kg: float = 0.0
        soc_min_kwh: float = 0.0
        soc_violations: list = ()

    export_experiment_report(
        tmp_path,
        _ProblemData(),
        _VehicleSet(),
        _MILP(),
        _Sim(),
        run_label="test",
    )
    report = (tmp_path / "experiment_report.md").read_text(encoding="utf-8")

    assert "> EXPLORATORY — RESEARCH SUBMISSION BLOCKED" in report
    assert "hourly_rolling_chain_not_accepted" in report


def test_export_experiment_report_accepts_only_verified_rolling_chain(tmp_path):
    from dataclasses import dataclass

    from src.result_exporter import export_experiment_report

    (tmp_path / "manifest.json").write_text("{}", encoding="utf-8")
    (tmp_path / "input_audit.json").write_text("{}", encoding="utf-8")
    rolling_dir = tmp_path / "rolling_hourly_chain"
    rolling_dir.mkdir()
    (rolling_dir / "rolling_chain_summary.json").write_text(
        json.dumps(
            {
                "chain_accepted": True,
                "acceptance_checks": {
                    name: True
                    for name in ROLLING_CHAIN_REQUIRED_ACCEPTANCE_CHECKS
                },
            }
        ),
        encoding="utf-8",
    )

    @dataclass
    class _ProblemData:
        delta_t_min: float = 15.0
        num_periods: int = 96
        enable_pv: bool = True
        enable_v2g: bool = False
        enable_demand_charge: bool = True

    @dataclass
    class _VehicleSet:
        K_BEV: list = ("ev-1",)
        K_ICE: list = ()
        R: list = ("trip-1",)
        C: list = ("chgr-1",)

    @dataclass
    class _MILP:
        status: str = "OPTIMAL"
        objective_value: object = 1.0
        mip_gap: object = 0.0
        solve_time_sec: float = 1.0
        infeasibility_info: str = ""

    @dataclass
    class _Sim:
        provisional_energy_cost: float = 0.0
        charged_energy_cost: float = 0.0
        total_energy_cost: float = 1.0
        total_demand_charge: float = 0.0
        total_fuel_cost: float = 0.0
        total_degradation_cost: float = 0.0
        total_operating_cost: float = 1.0
        served_task_ratio: float = 1.0
        unserved_tasks: list = ()
        total_grid_kwh: float = 0.0
        total_pv_kwh: float = 0.0
        pv_self_consumption_ratio: float = 0.0
        peak_demand_kw: float = 0.0
        total_co2_kg: float = 0.0
        soc_min_kwh: float = 0.0
        soc_violations: list = ()

    export_experiment_report(
        tmp_path,
        _ProblemData(),
        _VehicleSet(),
        _MILP(),
        _Sim(),
        run_label="test",
    )
    report = (tmp_path / "experiment_report.md").read_text(encoding="utf-8")

    assert "EXPLORATORY — RESEARCH SUBMISSION BLOCKED" not in report
    assert "research_submission_ready: True" in report
