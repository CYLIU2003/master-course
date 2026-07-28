from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bff.routers.optimization import _mark_frontend_run_claims_failed


def _write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload), encoding="utf-8")


def test_failed_finalization_downgrades_human_facing_release_claims(
    tmp_path: Path,
) -> None:
    """A failed frontend job may retain diagnostics but never READY labels."""

    ready_scope = {
        "run_profile": "day_ahead_and_hourly_rolling",
        "research_submission_ready": True,
        "teacher_release_status": "READY",
        "teacher_release_failed_checks": [],
    }
    _write_json(tmp_path / "research_claim_scope.json", ready_scope)
    _write_json(
        tmp_path / "summary.json",
        {
            "research_submission_ready": True,
            "teacher_release_status": "READY",
        },
    )
    _write_json(
        tmp_path / "optimization_result.json",
        {
            "solver_settings": {},
            "solver_metadata": {},
            "cost_breakdown": {},
        },
    )
    _write_json(tmp_path / "optimization_audit.json", {})
    _write_json(tmp_path / "run_manifest.json", {})
    _write_json(tmp_path / "input_audit.json", {})
    _write_json(tmp_path / "manifest.json", {"run_state": "complete"})
    _write_json(
        tmp_path / "artifact_completeness.json",
        {"status": "OK", "accepted": True, "content_errors": []},
    )
    (tmp_path / "experiment_report.md").write_text("report body\n", encoding="utf-8")

    workbook = Workbook()
    workbook.active.title = "cost_breakdown"
    workbook.save(tmp_path / "results.xlsx")

    _mark_frontend_run_claims_failed(
        run_dir=tmp_path,
        error=RuntimeError("final reconciliation failed"),
    )

    scope = json.loads((tmp_path / "research_claim_scope.json").read_text())
    summary = json.loads((tmp_path / "summary.json").read_text())
    result = json.loads((tmp_path / "optimization_result.json").read_text())
    manifest = json.loads((tmp_path / "run_manifest.json").read_text())
    rolling_manifest = json.loads((tmp_path / "manifest.json").read_text())
    artifact_audit = json.loads(
        (tmp_path / "artifact_completeness.json").read_text()
    )

    assert scope["research_submission_ready"] is False
    assert scope["teacher_release_status"] == "BLOCKED"
    assert scope["diagnostic_only"] is True
    assert "frontend_run_failed" in scope["teacher_release_failed_checks"]
    assert summary["teacher_release_status"] == "BLOCKED"
    assert result["research_claim_scope"]["teacher_release_status"] == "BLOCKED"
    assert manifest["run_state"] == "failed"
    assert rolling_manifest["run_state"] == "reporting_finalization_failed"
    assert artifact_audit["status"] == "ERROR"
    assert artifact_audit["accepted"] is False
    assert "frontend_run_failed_after_finalization" in artifact_audit[
        "content_errors"
    ]

    workbook = load_workbook(tmp_path / "results.xlsx", data_only=True)
    try:
        release_status = {
            str(row[0].value): row[1].value
            for row in workbook["release_status"].iter_rows(min_row=2)
        }
    finally:
        workbook.close()
    assert release_status["teacher_release_status"] == "BLOCKED"
    assert "teacher_release_status: `BLOCKED`" in (
        tmp_path / "experiment_report.md"
    ).read_text(encoding="utf-8")
