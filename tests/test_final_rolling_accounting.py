from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from bff.routers.optimization import (
    _apply_result_claim_classification,
    _assert_final_cost_artifact_consistency,
    _should_finalize_reporting_after_rolling,
    _synchronize_finalized_accounting_summary,
)
from bff.services.optimization_run.cost_breakdown import (
    CANONICAL_LEDGER_COMPONENT_SOURCES,
)


def _write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload), encoding="utf-8")


def _write_cost_artifacts(run_dir: Path, *, total: float) -> dict:
    source_components = {
        "electricity_cost": 10.125,
        "fuel_cost": 20.25,
        "demand_cost": 3.5,
        "vehicle_usage_cost": total - 35.0,
        "co2_cost": 1.125,
    }
    assert sum(source_components.values()) == total
    source_values = {
        source_key: float(source_components.get(source_key, 0.0))
        for source_key, _flag_key in (
            CANONICAL_LEDGER_COMPONENT_SOURCES.values()
        )
    }
    ledger_components = {
        component_key: source_values[source_key]
        for component_key, (source_key, _flag_key) in (
            CANONICAL_LEDGER_COMPONENT_SOURCES.items()
        )
    }
    enabled_component_keys = {
        component_key
        for component_key, (source_key, _flag_key) in (
            CANONICAL_LEDGER_COMPONENT_SOURCES.items()
        )
        if source_key in source_components
    }
    _write_json(
        run_dir / "rolling_hourly_chain" / "executed_day_accounting.json",
        {
            "eligible": True,
            "cost_breakdown": {"total_cost": total, **source_values},
        },
    )
    _write_json(
        run_dir / "graph" / "canonical_cost_ledger.json",
        {
            "source": "rolling_hourly_chain/executed_day_accounting.json",
            "accounting_total_cost_jpy": total,
            "components": ledger_components,
            "component_status": {
                component_key: {
                    "enabled": component_key in enabled_component_keys,
                    "status": (
                        "ENABLED"
                        if component_key in enabled_component_keys
                        else "SKIPPED"
                    ),
                    "source_key": source_key,
                    "source_present": True,
                    "value_jpy": ledger_components[component_key],
                }
                for component_key, (source_key, _flag_key) in (
                    CANONICAL_LEDGER_COMPONENT_SOURCES.items()
                )
            },
        },
    )
    _write_json(
        run_dir / "summary.json",
        {
            "accounting_total_cost_jpy": total,
            "energy_cost_jpy": source_components["electricity_cost"],
            "fuel_cost_jpy": source_components["fuel_cost"],
            "demand_charge_cost_jpy": source_components["demand_cost"],
            "co2_cost_jpy": source_components["co2_cost"],
            "canonical_cost_components_jpy": ledger_components,
        },
    )
    _write_json(
        run_dir / "experiment_report.json",
        {
            "results": {
                "total_cost_jpy": total,
                "electricity_cost_jpy": source_components["electricity_cost"],
                "diesel_cost_jpy": source_components["fuel_cost"],
                "demand_charge_jpy": source_components["demand_cost"],
                "vehicle_usage_cost_jpy": source_components[
                    "vehicle_usage_cost"
                ],
                "vehicle_fixed_cost_jpy": source_components.get(
                    "vehicle_cost", 0.0
                ),
                "vehicle_acquisition_cost_jpy": source_components.get(
                    "vehicle_acquisition_cost", 0.0
                ),
                "co2_cost_jpy": source_components["co2_cost"],
                "canonical_cost_components_jpy": ledger_components,
            }
        },
    )
    with (run_dir / "cost_breakdown_detail.csv").open(
        "w", encoding="utf-8", newline=""
    ) as handle:
        writer = csv.DictWriter(handle, fieldnames=("key", "value", "unit"))
        writer.writeheader()
        writer.writerow({"key": "total_cost", "value": total, "unit": "JPY"})
        for key, value in source_values.items():
            writer.writerow({"key": key, "value": value, "unit": "JPY"})
        writer.writerow(
            {
                "key": "demand_charge",
                "value": source_components["demand_cost"],
                "unit": "JPY",
            }
        )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cost_breakdown"
    sheet.append(["key", "value", "unit"])
    sheet.append(["total_cost", total, "JPY"])
    for key, value in source_values.items():
        sheet.append([key, value, "JPY"])
    sheet.append(["demand_charge", source_components["demand_cost"], "JPY"])
    workbook.save(run_dir / "results.xlsx")
    (run_dir / "experiment_report.md").write_text(
        f"canonical_executed_total_cost_jpy: `{total!r}`\n",
        encoding="utf-8",
    )
    return {
        "final_accounting_total_cost_jpy": total,
        "cost_breakdown": {"total_cost": total, **source_values},
    }


def test_final_cost_artifacts_must_equal_executed_rolling_accounting(
    tmp_path: Path,
) -> None:
    total = 724_618.0043661146
    optimization_result = _write_cost_artifacts(tmp_path, total=total)

    reconciliation = _assert_final_cost_artifact_consistency(
        run_dir=tmp_path,
        optimization_result=optimization_result,
    )

    assert reconciliation["status"] == "OK"
    assert reconciliation["failed_artifacts"] == {}


def test_rolling_failure_prevents_secondary_final_reporting() -> None:
    primary_failure = RuntimeError(
        "Positive Stage 2 charging power has no selected physical charger"
    )

    assert _should_finalize_reporting_after_rolling(None) is True
    assert _should_finalize_reporting_after_rolling(primary_failure) is False


def test_ineligible_executed_accounting_reports_its_reason(tmp_path: Path) -> None:
    _write_json(
        tmp_path / "rolling_hourly_chain" / "executed_day_accounting.json",
        {
            "eligible": False,
            "reason": "executed_slot_coverage_incomplete",
            "rejection_reasons": ["missing_slots:11-23"],
        },
    )

    with pytest.raises(
        RuntimeError,
        match="executed_slot_coverage_incomplete.*missing_slots:11-23",
    ):
        _assert_final_cost_artifact_consistency(
            run_dir=tmp_path,
            optimization_result={},
        )


def test_final_cost_mismatch_fails_job_above_one_micro_yen(
    tmp_path: Path,
) -> None:
    total = 724_618.0043661146
    optimization_result = _write_cost_artifacts(tmp_path, total=total)
    experiment = json.loads(
        (tmp_path / "experiment_report.json").read_text(encoding="utf-8")
    )
    experiment["results"]["total_cost_jpy"] = total - 34.5918146095
    _write_json(tmp_path / "experiment_report.json", experiment)

    with pytest.raises(RuntimeError, match="experiment_report_json"):
        _assert_final_cost_artifact_consistency(
            run_dir=tmp_path,
            optimization_result=optimization_result,
        )

    reconciliation = json.loads(
        (tmp_path / "final_cost_reconciliation.json").read_text(
            encoding="utf-8"
        )
    )
    assert reconciliation["status"] == "ERROR"


def test_missing_zero_valued_report_component_fails_with_reconciliation_diagnostic(
    tmp_path: Path,
) -> None:
    """A null report component is missing evidence, never an implicit zero."""

    total = 724_618.0043661146
    optimization_result = _write_cost_artifacts(tmp_path, total=total)
    experiment = json.loads(
        (tmp_path / "experiment_report.json").read_text(encoding="utf-8")
    )
    experiment["results"]["demand_charge_jpy"] = None
    _write_json(tmp_path / "experiment_report.json", experiment)

    with pytest.raises(
        RuntimeError,
        match="demand_charge_cost_jpy:experiment_report_json:missing",
    ):
        _assert_final_cost_artifact_consistency(
            run_dir=tmp_path,
            optimization_result=optimization_result,
        )

    reconciliation = json.loads(
        (tmp_path / "final_cost_reconciliation.json").read_text(
            encoding="utf-8"
        )
    )
    assert reconciliation["status"] == "ERROR"
    assert reconciliation["failed_artifacts"][
        "demand_charge_cost_jpy:experiment_report_json:missing"
    ] == "demand_charge_jpy"
    assert (
        reconciliation["observed_by_metric_jpy"][
            "demand_charge_cost_jpy"
        ]["experiment_report_json"]
        is None
    )
    assert (
        reconciliation["residual_to_executed_day_by_metric_jpy"][
            "demand_charge_cost_jpy"
        ]["experiment_report_json"]
        is None
    )


def test_markdown_canonical_total_accepts_ledger_precision_within_tolerance(
    tmp_path: Path,
) -> None:
    """The Markdown marker is numeric evidence, not a float-repr byte match."""

    total = 724_618.0043661146
    optimization_result = _write_cost_artifacts(tmp_path, total=total)
    (tmp_path / "experiment_report.md").write_text(
        "- canonical_executed_total_cost_jpy: "
        f"`{total + 5.0e-10!r}`\n",
        encoding="utf-8",
    )

    reconciliation = _assert_final_cost_artifact_consistency(
        run_dir=tmp_path,
        optimization_result=optimization_result,
    )

    assert reconciliation["status"] == "OK"


def test_finalized_ledger_replaces_provisional_summary_components(
    tmp_path: Path,
) -> None:
    """Final summary component maps must come from the canonical ledger."""

    _write_json(
        tmp_path / "summary.json",
        {
            "energy_cost_jpy": -1.8189894035458565e-12,
            "canonical_cost_components_jpy": {},
            "canonical_cost_component_status": {},
        },
    )
    finalized = {
        "total_cost_jpy": 707_808.6603727042,
        "accounting_total_cost_jpy": 707_808.6603727042,
        "energy_cost_jpy": -1.8189894035458565e-12,
        "electricity_cost_jpy": -1.8189894035458565e-12,
        "propulsion_energy_cost_jpy": 66_659.49730088498,
        "fuel_cost_jpy": 66_659.49730088498,
        "demand_charge_cost_jpy": 0.0,
        "vehicle_usage_cost_jpy": 640_000.0,
        "co2_cost_jpy": 1_149.1630718191466,
        "canonical_cost_components_jpy": {
            "electricity_cost_jpy": -1.8189894035458565e-12,
            "vehicle_usage_cost_jpy": 640_000.0,
        },
        "canonical_cost_component_status": {
            "electricity_cost_jpy": {"status": "ENABLED"},
        },
    }

    synchronized = _synchronize_finalized_accounting_summary(
        run_dir=tmp_path,
        summary={"canonical_cost_components_jpy": {}},
        finalized_accounting=finalized,
    )
    persisted = json.loads((tmp_path / "summary.json").read_text(encoding="utf-8"))

    assert synchronized["electricity_cost_jpy"] == pytest.approx(
        -1.8189894035458565e-12
    )
    assert synchronized["energy_cost_jpy"] == pytest.approx(
        -1.8189894035458565e-12
    )
    assert synchronized["propulsion_energy_cost_jpy"] == pytest.approx(
        66_659.49730088498
    )
    assert persisted["canonical_cost_components_jpy"] == finalized[
        "canonical_cost_components_jpy"
    ]
    assert persisted["canonical_cost_component_status"] == finalized[
        "canonical_cost_component_status"
    ]


def _write_full_component_artifacts(run_dir: Path) -> dict:
    component_values = {
        component_key: float(index)
        for index, component_key in enumerate(
            CANONICAL_LEDGER_COMPONENT_SOURCES,
            start=1,
        )
    }
    source_values = {
        source_key: component_values[component_key]
        for component_key, (source_key, _flag_key) in (
            CANONICAL_LEDGER_COMPONENT_SOURCES.items()
        )
    }
    source_values["demand_charge"] = source_values["demand_cost"]
    total = sum(component_values.values())
    _write_cost_artifacts(run_dir, total=total)
    _write_json(
        run_dir / "rolling_hourly_chain" / "executed_day_accounting.json",
        {
            "eligible": True,
            "cost_breakdown": {"total_cost": total, **source_values},
        },
    )
    _write_json(
        run_dir / "graph" / "canonical_cost_ledger.json",
        {
            "source": "rolling_hourly_chain/executed_day_accounting.json",
            "accounting_total_cost_jpy": total,
            "components": component_values,
            "component_status": {
                component_key: {
                    "enabled": True,
                    "status": "ENABLED",
                    "source_key": source_key,
                    "source_present": True,
                    "value_jpy": component_values[component_key],
                }
                for component_key, (source_key, _flag_key) in (
                    CANONICAL_LEDGER_COMPONENT_SOURCES.items()
                )
            },
        },
    )
    summary = json.loads((run_dir / "summary.json").read_text())
    summary.update(
        {
            "accounting_total_cost_jpy": total,
            "energy_cost_jpy": source_values["electricity_cost"],
            "electricity_cost_jpy": source_values["electricity_cost"],
            "fuel_cost_jpy": source_values["fuel_cost"],
            "demand_charge_cost_jpy": source_values["demand_cost"],
            "vehicle_usage_cost_jpy": source_values["vehicle_usage_cost"],
            "co2_cost_jpy": source_values["co2_cost"],
            "canonical_cost_components_jpy": component_values,
        }
    )
    _write_json(run_dir / "summary.json", summary)
    report = json.loads((run_dir / "experiment_report.json").read_text())
    report["results"].update(
        {
            "total_cost_jpy": total,
            "electricity_cost_jpy": source_values["electricity_cost"],
            "diesel_cost_jpy": source_values["fuel_cost"],
            "demand_charge_jpy": source_values["demand_cost"],
            "vehicle_usage_cost_jpy": source_values["vehicle_usage_cost"],
            "co2_cost_jpy": source_values["co2_cost"],
            "canonical_cost_components_jpy": component_values,
        }
    )
    _write_json(run_dir / "experiment_report.json", report)
    with (run_dir / "cost_breakdown_detail.csv").open(
        "w", encoding="utf-8", newline=""
    ) as handle:
        writer = csv.DictWriter(handle, fieldnames=("key", "value", "unit"))
        writer.writeheader()
        writer.writerow({"key": "total_cost", "value": total, "unit": "JPY"})
        for key, value in source_values.items():
            writer.writerow({"key": key, "value": value, "unit": "JPY"})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "cost_breakdown"
    sheet.append(["key", "value", "unit"])
    sheet.append(["total_cost", total, "JPY"])
    for key, value in source_values.items():
        sheet.append([key, value, "JPY"])
    workbook.save(run_dir / "results.xlsx")
    (run_dir / "experiment_report.md").write_text(
        f"canonical_executed_total_cost_jpy: `{total!r}`\n",
        encoding="utf-8",
    )
    return {
        "final_accounting_total_cost_jpy": total,
        "cost_breakdown": {"total_cost": total, **source_values},
    }


def test_all_enabled_canonical_components_reconcile_across_final_artifacts(
    tmp_path: Path,
) -> None:
    optimization_result = _write_full_component_artifacts(tmp_path)

    reconciliation = _assert_final_cost_artifact_consistency(
        run_dir=tmp_path,
        optimization_result=optimization_result,
    )

    assert reconciliation["status"] == "OK"
    assert set(CANONICAL_LEDGER_COMPONENT_SOURCES).issubset(
        reconciliation["expected_by_metric_jpy"]
    )


def test_missing_enabled_component_in_human_report_fails(
    tmp_path: Path,
) -> None:
    optimization_result = _write_full_component_artifacts(tmp_path)
    report = json.loads((tmp_path / "experiment_report.json").read_text())
    report["results"]["canonical_cost_components_jpy"].pop(
        "driver_cost_jpy"
    )
    _write_json(tmp_path / "experiment_report.json", report)

    with pytest.raises(RuntimeError, match="driver_cost_jpy"):
        _assert_final_cost_artifact_consistency(
            run_dir=tmp_path,
            optimization_result=optimization_result,
        )


def test_disabled_component_must_remain_zero_in_every_final_artifact(
    tmp_path: Path,
) -> None:
    """SKIPPED ledger components cannot become nonzero in a report map."""

    optimization_result = _write_full_component_artifacts(tmp_path)
    component_key = "vehicle_fixed_cost_jpy"
    source_key = "vehicle_cost"

    executed_path = tmp_path / "rolling_hourly_chain" / "executed_day_accounting.json"
    executed = json.loads(executed_path.read_text())
    executed["cost_breakdown"][source_key] = 0.0
    _write_json(executed_path, executed)

    ledger_path = tmp_path / "graph" / "canonical_cost_ledger.json"
    ledger = json.loads(ledger_path.read_text())
    ledger["components"][component_key] = 0.0
    ledger["component_status"][component_key].update(
        {"enabled": False, "status": "SKIPPED", "value_jpy": 0.0}
    )
    _write_json(ledger_path, ledger)

    summary_path = tmp_path / "summary.json"
    summary = json.loads(summary_path.read_text())
    summary["canonical_cost_components_jpy"][component_key] = 0.0
    _write_json(summary_path, summary)

    report_path = tmp_path / "experiment_report.json"
    report = json.loads(report_path.read_text())
    report["results"]["canonical_cost_components_jpy"][component_key] = 100.0
    _write_json(report_path, report)

    detail_path = tmp_path / "cost_breakdown_detail.csv"
    rows = list(csv.DictReader(detail_path.read_text(encoding="utf-8").splitlines()))
    for row in rows:
        if row["key"] == source_key:
            row["value"] = "0.0"
    with detail_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=("key", "value", "unit"))
        writer.writeheader()
        writer.writerows(rows)

    workbook = load_workbook(tmp_path / "results.xlsx")
    try:
        for row in workbook["cost_breakdown"].iter_rows(min_row=2):
            if row[0].value == source_key:
                row[1].value = 0.0
        workbook.save(tmp_path / "results.xlsx")
    finally:
        workbook.close()

    optimization_result["cost_breakdown"][source_key] = 0.0
    with pytest.raises(RuntimeError, match=component_key):
        _assert_final_cost_artifact_consistency(
            run_dir=tmp_path,
            optimization_result=optimization_result,
        )


def test_gap_miss_is_labeled_feasible_candidate() -> None:
    result = {
        "solution_validity": {"validated_feasible": True},
        "solver_settings": {
            "mip_gap_requested_ratio": 0.1,
            "mip_gap_target_met": False,
            "stage1_certified_mip_gap_ratio": 0.10488,
            "stage1_gurobi_raw_mip_gap_ratio": 1.0,
        },
        "solver_metadata": {"supports_integrated_exact_milp": False},
        "summary": {},
    }

    classification = _apply_result_claim_classification(result)

    assert classification["label"] == "feasible_candidate"
    assert classification["optimality_claim_eligible"] is False
    assert result["result_status"] == "FEASIBLE_CANDIDATE"
    assert set(classification["optimality_blocking_reasons"]) == {
        "requested_mip_gap_not_met",
        "not_an_integrated_global_assignment_and_charging_milp",
    }
