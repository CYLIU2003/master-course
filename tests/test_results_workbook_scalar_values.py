from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from bff.routers import optimization


def test_results_workbook_serializes_structured_cost_metadata(
    tmp_path: Path,
) -> None:
    """Workbook output must preserve, rather than reject, cost metadata."""

    flags = {
        "electricity_cost": True,
        "vehicle_fixed_cost": False,
    }
    optimization._write_results_workbook(
        run_dir=tmp_path,
        summary={
            "objective_value": 123.0,
            "solve_time_seconds": 4.5,
            "trip_count_served": 2,
            "trip_count_unserved": 0,
        },
        cost_rows=[
            {"key": "total_cost", "value": 123.0, "unit": "JPY"},
            {
                "key": "cost_component_flags",
                "value": flags,
                "unit": "metadata",
            },
            {
                "key": "diagnostic_labels",
                "value": ["executed", "accepted"],
                "unit": "metadata",
            },
        ],
    )

    workbook = load_workbook(tmp_path / "results.xlsx", data_only=True)
    try:
        rows = {
            str(row[0].value): row[1].value
            for row in workbook["cost_breakdown"].iter_rows(min_row=2)
        }
    finally:
        workbook.close()

    assert rows["total_cost"] == 123.0
    assert rows["cost_component_flags"] == json.dumps(
        flags,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )
    assert rows["diagnostic_labels"] == '["executed","accepted"]'


def test_results_workbook_rejects_unsupported_structured_value(
    tmp_path: Path,
) -> None:
    """Unexpected objects must fail reporting rather than be stringified."""

    with pytest.raises(TypeError, match="cannot serialize non-scalar report value"):
        optimization._write_results_workbook(
            run_dir=tmp_path,
            summary={},
            cost_rows=[
                {
                    "key": "unsupported",
                    "value": object(),
                    "unit": "metadata",
                }
            ],
        )
