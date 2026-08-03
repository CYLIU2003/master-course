from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

from openpyxl import Workbook
import pytest

from bff.routers import optimization
from bff.services.optimization_run.artifact_completeness import (
    audit_frontend_run_artifacts,
    required_frontend_artifacts,
    validate_stage1_used_powertrain_composition_search,
)


_ROOT_REFUEL_EVENT_FIELDS = (
    "vehicle_id",
    "slot_index",
    "time_hhmm",
    "refuel_liters",
    "unit",
)
_GRAPH_REFUEL_EVENT_FIELDS = (
    "vehicle_id",
    "slot_index",
    "time_hhmm",
    "refuel_liters",
    "location_id",
)


def _write_artifact(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix == ".json":
        path.write_text("{}\n", encoding="utf-8")
    elif path.suffix == ".xlsx":
        workbook = Workbook()
        workbook.active.title = "summary"
        workbook.create_sheet("cost_breakdown")
        workbook.create_sheet("release_status")
        workbook.save(path)
    elif path.suffix == ".csv":
        path.write_text("key,value\n", encoding="utf-8")
    else:
        path.write_text("artifact\n", encoding="utf-8")


def _write_refuel_event_csv(
    path: Path,
    *,
    fieldnames: tuple[str, ...],
    rows: list[dict],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        writer.writerows(rows)


def _write_refueling_exports(run_dir: Path, rows: list[dict]) -> None:
    _write_refuel_event_csv(
        run_dir / "refuel_events.csv",
        fieldnames=_ROOT_REFUEL_EVENT_FIELDS,
        rows=[
            {
                "vehicle_id": row["vehicle_id"],
                "slot_index": row["slot_index"],
                "time_hhmm": row["time_hhmm"],
                "refuel_liters": row["refuel_liters"],
                "unit": "L",
            }
            for row in rows
        ],
    )
    _write_refuel_event_csv(
        run_dir / "graph" / "refuel_events.csv",
        fieldnames=_GRAPH_REFUEL_EVENT_FIELDS,
        rows=rows,
    )


def _set_canonical_refueling_schedule(
    run_dir: Path,
    rows: list[dict],
) -> None:
    canonical_path = run_dir / "canonical_solver_result.json"
    canonical = json.loads(canonical_path.read_text(encoding="utf-8"))
    canonical["refueling_schedule"] = rows
    canonical_path.write_text(json.dumps(canonical), encoding="utf-8")
    canonical_sha256 = hashlib.sha256(canonical_path.read_bytes()).hexdigest()

    rolling_summary_path = (
        run_dir / "rolling_hourly_chain" / "rolling_chain_summary.json"
    )
    rolling_summary = json.loads(
        rolling_summary_path.read_text(encoding="utf-8")
    )
    rolling_summary["day_ahead_result_sha256"] = canonical_sha256
    rolling_summary_path.write_text(
        json.dumps(rolling_summary), encoding="utf-8"
    )

    input_manifest_path = run_dir / "physical_validation_input_manifest.json"
    input_manifest = json.loads(input_manifest_path.read_text(encoding="utf-8"))
    input_manifest["canonical_solver_result_sha256"] = canonical_sha256
    input_manifest_path.write_text(
        json.dumps(input_manifest), encoding="utf-8"
    )


def _canonical_refueling_event() -> dict:
    return {
        "vehicle_id": "vehicle-1",
        "slot_index": 1,
        "time_hhmm": "00:30",
        "refuel_liters": 12.5,
        "location_id": "depot-1",
    }


def _artifact_record(path: Path, *, root: Path) -> dict:
    return {
        "path": path.relative_to(root).as_posix(),
        "size_bytes": path.stat().st_size,
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
    }


def _solver_accounting_reconciliation() -> dict:
    return {
        "schema_version": "solver_objective_accounting_reconciliation_v1",
        "solver_objective_value_jpy": 100.0,
        "solver_objective_source": "optimization_result.objective_value",
        "canonical_accounting_total_jpy": 100.0,
        "canonical_accounting_source": "rolling_hourly_chain/executed_day_accounting.json",
        "difference_jpy": 0.0,
        "absolute_difference_jpy": 0.0,
        "tolerance_jpy": 1.0e-6,
        "numeric_values_available": True,
        "numeric_residual_within_tolerance": True,
        "objective_is_actual_cost": True,
        "matches_canonical_accounting_total": True,
        "objective_semantics": "actual_cost",
    }


def _composition_search_certificate() -> dict:
    return {
        "schema_version": "stage1_used_powertrain_composition_search_v1",
        "enabled": True,
        "radius_requested": 2,
        "primary_used_powertrain_composition": {"used_bev": 1, "used_ice": 1},
        "selected_inventory": {
            "available_electric_vehicle_count": 2,
            "available_combustion_vehicle_count": 2,
            "electric_vehicle_ids": ["bev-1", "bev-2"],
            "combustion_vehicle_ids": ["ice-1", "ice-2"],
        },
        "target_records": [
            {
                "target_used_bev": 2,
                "target_used_ice": 0,
                "target_within_selected_inventory": True,
                "search_status": "optimal",
                "solver_status": "optimal",
                "final_disposition": "physically_feasible_stage2_candidate",
            }
        ],
        "feasible_used_powertrain_compositions": [
            {"used_bev": 1, "used_ice": 1},
            {"used_bev": 2, "used_ice": 0},
        ],
        "multiple_feasible_compositions_found": True,
        "all_adjacent_targets_certified_infeasible": False,
        "inventory_has_no_adjacent_composition": False,
        "unresolved_targets": [],
        "accepted_for_formal_composition_evidence": True,
        "blocking_reasons": [],
        "semantics": "fixture composition search evidence",
    }


def _write_composition_search_csv(path: Path) -> None:
    fieldnames = (
        "target_used_bev",
        "minimum_used_bev_count",
        "target_used_ice",
        "delta_used_bev_from_primary",
        "delta_used_ice_from_primary",
        "target_total_used_vehicle_count",
        "target_within_selected_inventory",
        "search_status",
        "solver_status",
        "frontier_status",
        "solution_count",
        "best_bound",
        "mip_gap_ratio",
        "time_limit_sec",
        "solver_runtime_sec",
        "candidate_hash",
        "actual_used_bev",
        "actual_used_ice",
        "candidate_accepted_for_stage2_evaluation",
        "final_disposition",
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        writer.writerow(
            {
                "target_used_bev": 2,
                "minimum_used_bev_count": 2,
                "target_used_ice": 0,
                "target_within_selected_inventory": True,
                "search_status": "optimal",
                "solver_status": "optimal",
                "frontier_status": "FEASIBLE_INCUMBENT",
                "final_disposition": "physically_feasible_stage2_candidate",
            }
        )


def _complete_run(tmp_path: Path) -> Path:
    run_dir = tmp_path / "run"
    required = required_frontend_artifacts(
        research_run=True,
        require_rolling=True,
    )
    for relative_path in required:
        if relative_path == "run_manifest.json":
            continue
        _write_artifact(run_dir / relative_path)

    assignment_economic_audit = {
        "schema_version": "assignment_economic_audit_v1",
        "bev_grid_marginal_cost_jpy_per_km": 41.5578947368421,
        "ice_marginal_cost_jpy_per_km": 33.18584,
        "renewable_budget_kwh": 0.0,
        "renewable_energy_allocated_in_stage1_kwh": 0.0,
        "grid_energy_allocated_in_stage1_kwh": 10.0,
        "stage1_bev_trip_count": 1,
        "stage2_bev_trip_count": 1,
        "assignment_energy_coupling_mode": (
            "slot_level_assignment_coupled_continuous_energy_recourse"
        ),
        "weather_response_expected": "no_directional_policy",
        "weather_response_observed": "not_assessable_from_single_case",
        "vehicle_usage_cost_jpy_per_used_bus": 20_000.0,
        "vehicle_usage_cost_semantics": "operator_incremental_vehicle_day_cost",
        "vehicle_usage_cost_semantics_classified": True,
        "vehicle_usage_cost_semantics_research_eligible": True,
    }
    (run_dir / "assignment_economic_audit.json").write_text(
        json.dumps(assignment_economic_audit), encoding="utf-8"
    )
    with (run_dir / "assignment_economic_audit.csv").open(
        "w", encoding="utf-8", newline=""
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=list(
                (
                    "bev_grid_marginal_cost_jpy_per_km",
                    "ice_marginal_cost_jpy_per_km",
                    "renewable_budget_kwh",
                    "renewable_energy_allocated_in_stage1_kwh",
                    "grid_energy_allocated_in_stage1_kwh",
                    "stage1_bev_trip_count",
                    "stage2_bev_trip_count",
                    "assignment_energy_coupling_mode",
                    "weather_response_expected",
                    "weather_response_observed",
                    "vehicle_usage_cost_jpy_per_used_bus",
                    "vehicle_usage_cost_semantics",
                    "vehicle_usage_cost_semantics_classified",
                    "vehicle_usage_cost_semantics_research_eligible",
                )
            ),
        )
        writer.writeheader()
        writer.writerow(
            {
                key: assignment_economic_audit[key]
                for key in writer.fieldnames or ()
            }
        )
    (run_dir / "solver_objective_accounting_reconciliation.json").write_text(
        json.dumps(_solver_accounting_reconciliation()),
        encoding="utf-8",
    )

    graph_manifest = {
        "files": ["declared_graph_artifact.csv", "refuel_events.csv"],
        "optional_exports": {},
    }
    (run_dir / "graph" / "manifest.json").write_text(
        json.dumps(graph_manifest),
        encoding="utf-8",
    )
    _write_artifact(run_dir / "graph" / "declared_graph_artifact.csv")
    literature_entries = []
    literature_dir = run_dir / "graph" / "literature_figures"
    for index in range(5):
        artifact_files = [
            f"{index:02d}_figure.png",
            f"{index:02d}_figure.svg",
            f"{index:02d}_figure_source.csv",
        ]
        for artifact_file in artifact_files:
            _write_artifact(literature_dir / artifact_file)
        literature_entries.append(
            {
                "kind": "figure",
                "figure_id": f"figure-{index}",
                "artifact_files": artifact_files,
                "canonical_sources": ["scenario_fleet_contract.json"],
                "artifact_records": [
                    _artifact_record(
                        literature_dir / artifact_file,
                        root=literature_dir,
                    )
                    for artifact_file in artifact_files
                ],
            }
        )
    raw_data_files = [
        f"raw_data/{index:02d}_dataset.csv" for index in range(15)
    ] + ["raw_data/raw_data_catalog.csv"]
    for artifact_file in raw_data_files:
        _write_artifact(literature_dir / artifact_file)
    literature_entries.append(
        {
            "kind": "raw_data_bundle",
            "figure_id": "analysis_ready_raw_data",
            "artifact_files": raw_data_files,
            "canonical_sources": ["scenario_fleet_contract.json"],
            "artifact_records": [
                _artifact_record(
                    literature_dir / artifact_file,
                    root=literature_dir,
                )
                for artifact_file in raw_data_files
            ],
        }
    )
    source_path = run_dir / "scenario_fleet_contract.json"
    (literature_dir / "manifest.json").write_text(
        json.dumps(
            {
                "schema_version": "literature_figure_bundle_v1",
                "status": "READY",
                "figure_count": 5,
                "raw_data_csv_count": 16,
                "raw_data_catalog": "raw_data/raw_data_catalog.csv",
                "entries": literature_entries,
                "source_artifacts": {
                    "scenario_fleet_contract.json": _artifact_record(
                        source_path,
                        root=run_dir,
                    )
                },
            }
        ),
        encoding="utf-8",
    )
    graph_manifest["optional_exports"] = {
        "literature_figures": {
            "enabled": True,
            "manifest_file": "literature_figures/manifest.json",
        }
    }
    (run_dir / "graph" / "manifest.json").write_text(
        json.dumps(graph_manifest),
        encoding="utf-8",
    )

    rolling_dir = run_dir / "rolling_hourly_chain"
    rolling_summary = {
        "expected_step_count": 2,
        "step_count": 2,
        "chain_accepted": True,
        "all_steps_feasible": True,
    }
    (rolling_dir / "rolling_chain_summary.json").write_text(
        json.dumps(rolling_summary),
        encoding="utf-8",
    )
    (rolling_dir / "executed_day_accounting.json").write_text(
        json.dumps(
            {
                "eligible": True,
                "expected_slot_count": 4,
                "executed_slot_count": 4,
            }
        ),
        encoding="utf-8",
    )
    canonical_path = run_dir / "canonical_solver_result.json"
    executed_charging_path = rolling_dir / "charging_schedule.csv"
    canonical_path.write_text(
        json.dumps(
            {
                "vehicle_paths": {"vehicle-1": ["trip-1", "trip-2"]},
                "served_trip_ids": ["trip-1", "trip-2"],
                "unserved_trip_ids": [],
                "trip_count_served": 2,
                "trip_count_unserved": 0,
                "refueling_schedule": [],
            }
        ),
        encoding="utf-8",
    )
    _write_refueling_exports(run_dir, [])
    rolling_summary.update(
        {
            "day_ahead_result_sha256": hashlib.sha256(
                canonical_path.read_bytes()
            ).hexdigest(),
            "day_ahead_assignment_hash": "a" * 64,
        }
    )
    (rolling_dir / "rolling_chain_summary.json").write_text(
        json.dumps(rolling_summary),
        encoding="utf-8",
    )
    (run_dir / "physical_validation_input_manifest.json").write_text(
        json.dumps(
            {
                "schema_version": "physical_validation_input_manifest_v1",
                "assignment_source": "canonical_solver_result.json",
                "charging_source": "rolling_hourly_chain/charging_schedule.csv",
                "refueling_source": "canonical_solver_result.json",
                "day_ahead_assignment_hash": "a" * 64,
                "canonical_solver_result_sha256": hashlib.sha256(
                    canonical_path.read_bytes()
                ).hexdigest(),
                "executed_charging_schedule_sha256": hashlib.sha256(
                    executed_charging_path.read_bytes()
                ).hexdigest(),
                "vehicle_path_count": 1,
                "assigned_trip_occurrence_count": 2,
                "served_trip_occurrence_count": 2,
                "problem_trip_count": 2,
                "unserved_trip_count": 0,
                "validation_contract": {
                    "canonical_sha_matches_rolling_chain": True,
                    "vehicle_paths_match_served_trip_ids": True,
                    "vehicle_paths_cover_problem_trips_exactly": True,
                    "unserved_trip_ids_empty": True,
                    "executed_charging_overlay_only": True,
                },
            }
        ),
        encoding="utf-8",
    )
    (run_dir / "physical_schedule_validation.json").write_text(
        json.dumps({"accepted": True}),
        encoding="utf-8",
    )
    (run_dir / "final_cost_reconciliation.json").write_text(
        json.dumps({"status": "OK"}),
        encoding="utf-8",
    )
    for index in range(2):
        step_dir = rolling_dir / f"step_{index:02d}_{index:02d}00"
        step_files = [
            "hourly_solver_result.json",
            "hourly_summary.json",
        ]
        if index == 0:
            step_files.append("state_for_next_hour.json")
        for filename in step_files:
            _write_artifact(step_dir / filename)

    declared_files = sorted(
        path.relative_to(run_dir).as_posix()
        for path in run_dir.rglob("*")
        if path.is_file()
    )
    (run_dir / "run_manifest.json").write_text(
        json.dumps(
            {
                "research_run": True,
                "research_claim_scope": {
                    "run_profile": "day_ahead_and_hourly_rolling"
                },
                "files": declared_files,
            }
        ),
        encoding="utf-8",
    )
    return run_dir


def test_complete_frontend_run_artifact_contract_passes(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "OK"
    assert audit["accepted"] is True
    assert audit["missing_artifacts"] == []
    assert audit["content_errors"] == []
    assert (
        audit["required_artifact_count"]
        == audit["verified_artifact_count"]
    )


def test_two_stage_formal_contract_requires_valid_composition_certificate(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)

    missing_certificate = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
        require_two_stage_composition_certificate=True,
    )
    assert missing_certificate["accepted"] is False
    assert "stage1_used_powertrain_composition_search.json" in (
        missing_certificate["missing_artifacts"]
    )

    certificate_path = run_dir / "stage1_used_powertrain_composition_search.json"
    certificate_path.write_text(
        json.dumps(_composition_search_certificate()),
        encoding="utf-8",
    )
    _write_composition_search_csv(
        run_dir / "stage1_used_powertrain_composition_search.csv"
    )
    manifest_path = run_dir / "run_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["files"] = sorted(
        path.relative_to(run_dir).as_posix()
        for path in run_dir.rglob("*")
        if path.is_file()
    )
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
        require_two_stage_composition_certificate=True,
    )

    assert audit["status"] == "OK"
    assert audit["accepted"] is True


def test_composition_certificate_rejects_status_only_infeasibility() -> None:
    """A true-looking summary cannot replace a target-bound IIS artifact."""

    certificate = _composition_search_certificate()
    certificate.update(
        {
            "feasible_used_powertrain_compositions": [
                {"used_bev": 1, "used_ice": 1}
            ],
            "multiple_feasible_compositions_found": False,
            "all_adjacent_targets_certified_infeasible": True,
        }
    )
    certificate["target_records"] = [
        {
            "target_used_bev": 2,
            "target_used_ice": 0,
            "target_within_selected_inventory": True,
            "search_status": "infeasible",
            "solver_status": "infeasible",
            "final_disposition": "stage1_infeasibility_certificate",
            # Deliberately omits successful-IIS and exact-LP-hash evidence.
            "infeasibility_certificate": {
                "solver_status": "infeasible",
                "accepted_for_formal_composition_evidence": True,
            },
        }
    ]

    errors = validate_stage1_used_powertrain_composition_search(
        certificate,
        require_accepted=True,
    )

    assert any("successful IIS" in error for error in errors)
    assert any("Stage 1 LP SHA-256" in error for error in errors)


def test_artifact_contract_accepts_matching_canonical_refueling_exports(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    event = _canonical_refueling_event()
    _set_canonical_refueling_schedule(run_dir, [event])
    _write_refueling_exports(run_dir, [event])

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "OK"
    assert audit["accepted"] is True


def test_artifact_contract_rejects_header_only_refueling_exports_with_canonical_event(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    _set_canonical_refueling_schedule(run_dir, [_canonical_refueling_event()])

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "refuel_events.csv rows do not exactly match" in error
        for error in audit["content_errors"]
    )
    assert any(
        "graph/refuel_events.csv rows do not exactly match" in error
        for error in audit["content_errors"]
    )


def test_artifact_contract_requires_graph_refuel_events_even_if_omitted_from_manifest(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    graph_manifest_path = run_dir / "graph" / "manifest.json"
    graph_manifest = json.loads(graph_manifest_path.read_text(encoding="utf-8"))
    graph_manifest["files"].remove("refuel_events.csv")
    graph_manifest_path.write_text(json.dumps(graph_manifest), encoding="utf-8")
    (run_dir / "graph" / "refuel_events.csv").unlink()

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert "graph/refuel_events.csv" in audit["missing_artifacts"]


@pytest.mark.parametrize(
    ("relative_path", "malformed_header"),
    [
        ("refuel_events.csv", "vehicle_id,slot_index,time_hhmm,refuel_liters\n"),
        ("graph/refuel_events.csv", "vehicle_id,slot_index,time_hhmm,refuel_liters\n"),
    ],
)
def test_artifact_contract_rejects_refueling_export_schema_mismatch(
    tmp_path: Path,
    relative_path: str,
    malformed_header: str,
) -> None:
    run_dir = _complete_run(tmp_path)
    (run_dir / relative_path).write_text(malformed_header, encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        error.startswith(f"{relative_path}: CSV header does not match")
        for error in audit["content_errors"]
    )


@pytest.mark.parametrize("relative_path", ["refuel_events.csv", "graph/refuel_events.csv"])
def test_artifact_contract_rejects_refueling_export_row_mismatch(
    tmp_path: Path,
    relative_path: str,
) -> None:
    run_dir = _complete_run(tmp_path)
    event = _canonical_refueling_event()
    _set_canonical_refueling_schedule(run_dir, [event])
    _write_refueling_exports(run_dir, [event])
    if relative_path == "refuel_events.csv":
        _write_refuel_event_csv(
            run_dir / relative_path,
            fieldnames=_ROOT_REFUEL_EVENT_FIELDS,
            rows=[
                {
                    "vehicle_id": event["vehicle_id"],
                    "slot_index": event["slot_index"],
                    "time_hhmm": event["time_hhmm"],
                    "refuel_liters": 13.0,
                    "unit": "L",
                }
            ],
        )
    else:
        _write_refuel_event_csv(
            run_dir / relative_path,
            fieldnames=_GRAPH_REFUEL_EVENT_FIELDS,
            rows=[{**event, "location_id": "other-depot"}],
        )

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        error.startswith(f"{relative_path} rows do not exactly match")
        for error in audit["content_errors"]
    )


def test_artifact_contract_rejects_tampered_physical_validation_input_manifest(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    manifest_path = run_dir / "physical_validation_input_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["canonical_solver_result_sha256"] = "0" * 64
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "canonical_solver_result_sha256 does not match" in error
        for error in audit["content_errors"]
    )


def test_artifact_contract_rejects_manifest_counts_that_do_not_match_canonical(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    manifest_path = run_dir / "physical_validation_input_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update(
        {
            "vehicle_path_count": 0,
            "assigned_trip_occurrence_count": 0,
            "served_trip_occurrence_count": 0,
            "problem_trip_count": 0,
        }
    )
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "vehicle_path_count does not match canonical_solver_result" in error
        for error in audit["content_errors"]
    )


def test_artifact_contract_rejects_manifest_assignment_hash_not_in_chain(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    manifest_path = run_dir / "physical_validation_input_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["day_ahead_assignment_hash"] = "f" * 64
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "day_ahead_assignment_hash does not match rolling_chain_summary"
        in error
        for error in audit["content_errors"]
    )


def test_literature_artifact_hash_mismatch_fails_contract(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    artifact = (
        run_dir
        / "graph"
        / "literature_figures"
        / "00_figure_source.csv"
    )
    artifact.write_text("key,value\ntampered,1\n", encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "sha256 mismatch for 00_figure_source.csv" in error
        for error in audit["content_errors"]
    )


def test_literature_source_hash_mismatch_fails_contract(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    (run_dir / "scenario_fleet_contract.json").write_text(
        '{"changed": true}\n',
        encoding="utf-8",
    )

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "sha256 mismatch for scenario_fleet_contract.json" in error
        for error in audit["content_errors"]
    )


def test_literature_manifest_cannot_omit_canonical_source_hash(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    manifest_path = (
        run_dir / "graph" / "literature_figures" / "manifest.json"
    )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["source_artifacts"] = {}
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "source_artifacts must be a non-empty object" in error
        for error in audit["content_errors"]
    )


def test_missing_results_workbook_fails_frontend_artifact_contract(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    (run_dir / "results.xlsx").unlink()

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert "results.xlsx" in audit["missing_artifacts"]


def test_missing_rolling_step_state_fails_frontend_artifact_contract(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    (
        run_dir
        / "rolling_hourly_chain"
        / "step_00_0000"
        / "state_for_next_hour.json"
    ).unlink()

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert (
        "rolling_hourly_chain/step_00_0000/state_for_next_hour.json"
        in audit["missing_artifacts"]
    )


def test_invalid_canonical_json_fails_frontend_artifact_contract(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    (run_dir / "canonical_solver_result.json").write_text(
        "{broken",
        encoding="utf-8",
    )

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert "canonical_solver_result.json" in audit[
        "invalid_json_artifacts"
    ]


def test_unsafe_graph_manifest_path_fails_without_leaving_run_dir(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    (run_dir / "graph" / "manifest.json").write_text(
        json.dumps(
            {
                "files": ["../../outside.csv"],
                "optional_exports": {},
            }
        ),
        encoding="utf-8",
    )

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "unsafe file path" in error
        for error in audit["content_errors"]
    )


def test_missing_literature_bundle_artifact_fails_contract(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    (
        run_dir
        / "graph"
        / "literature_figures"
        / "00_figure_source.csv"
    ).unlink()

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert (
        "graph/literature_figures/00_figure_source.csv"
        in audit["missing_artifacts"]
    )


def test_malformed_literature_bundle_counts_fail_without_crashing(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    manifest_path = (
        run_dir / "graph" / "literature_figures" / "manifest.json"
    )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["figure_count"] = "5"
    manifest["raw_data_csv_count"] = True
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    audit = audit_frontend_run_artifacts(
        run_dir,
        research_run=True,
        require_rolling=True,
    )

    assert audit["status"] == "ERROR"
    assert any(
        "figure_count must be a non-negative integer" in error
        for error in audit["content_errors"]
    )
    assert any(
        "raw_data_csv_count must be a non-negative integer" in error
        for error in audit["content_errors"]
    )


def test_frontend_gate_hashes_the_final_run_manifest(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    optimization_result: dict = {}
    optimization_audit: dict = {}

    reporting = optimization._enforce_frontend_run_artifact_contract(
        run_dir=run_dir,
        optimization_result=optimization_result,
        optimization_audit=optimization_audit,
        reporting_finalizer_result={"status": "completed"},
        research_run=True,
        require_rolling=True,
    )

    completeness = json.loads(
        (run_dir / "artifact_completeness.json").read_text(
            encoding="utf-8"
        )
    )
    manifest_bytes = (run_dir / "run_manifest.json").read_bytes()
    assert completeness["status"] == "OK"
    assert completeness["artifacts"]["run_manifest.json"][
        "sha256"
    ] == hashlib.sha256(manifest_bytes).hexdigest()
    assert reporting["artifact_completeness_status"] == "OK"
    assert optimization_result["artifact_completeness"]["accepted"] is True
