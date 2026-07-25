"""Finalize reporting ledgers from canonical optimization run artifacts.

The functions in this module do not run solvers, simulations, dispatch, or
vehicle assignment. They only reconcile reporting artifacts from existing run
CSV/JSON ledgers.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import shutil
from datetime import datetime, timezone
from dataclasses import dataclass
from pathlib import Path
from statistics import median
from typing import Any


TOL = 1e-6

REPORTING_FILES = [
    "graph/kpi_summary.json",
    "summary.json",
    "cost_breakdown_detail.csv",
    "graph/energy_flow_ledger.csv",
    "graph/data_flow_validation.csv",
    "results.xlsx",
    "experiment_report.md",
]

INVALID_SUMMARY_KPI_FIELDS = {
    "objective_value",
    "objective_value_jpy",
    "total_cost_jpy",
    "accounting_total_cost_jpy",
    "reported_total_cost_jpy",
    "gross_operating_cost_jpy",
    "solver_objective_value",
    "validated_operating_cost_jpy",
    "grid_purchase_cost_jpy",
    "demand_charge_cost_jpy",
    "fuel_cost_jpy",
    "co2_cost_jpy",
    "vehicle_cost_jpy",
    "vehicle_usage_cost_jpy",
    "driver_cost_jpy",
    "unserved_penalty_jpy",
    "switch_cost_jpy",
    "battery_degradation_cost_jpy",
    "contract_overage_cost_jpy",
}

INVALID_KPI_FIELDS = {
    "total_cost_jpy",
    "accounting_total_cost_jpy",
    "gross_operating_cost_jpy",
    "reported_total_cost_jpy",
    "objective_value",
    "objective_value_jpy",
    "solver_objective_value",
    "validated_operating_cost_jpy",
    "energy_cost_jpy",
    "demand_cost_jpy",
    "fuel_cost_jpy",
    "co2_cost_jpy",
    "battery_degradation_cost_jpy",
    "contract_overage_cost_jpy",
    "vehicle_usage_cost_jpy",
    "vehicle_cost_jpy",
    "driver_cost_jpy",
    "unserved_penalty_jpy",
    "switch_cost_jpy",
    "grid_purchase_cost_jpy",
    "demand_charge_cost_jpy",
    "pv_to_bus_cost_jpy",
    "pv_to_bess_cost_jpy",
    "bess_to_bus_cost_jpy",
    "bess_total_flow_cost_jpy",
    "pv_to_bus_kwh",
    "pv_to_bess_kwh",
    "pv_curtailed_kwh",
    "pv_curtailment_kwh",
    "pv_export_kwh",
    "pv_utilization_ratio",
    "bess_to_bus_kwh",
    "bess_charge_kwh",
    "bess_discharge_to_bus_kwh",
    "bess_discharge_kwh",
    "grid_to_bus_kwh",
    "grid_to_bess_kwh",
    "grid_total_kwh",
    "grid_import_kwh",
    "facility_load_kwh",
    "peak_grid_import_kw",
    "peak_grid_import_kw_all_depots",
    "peak_grid_import_kw_any_depot",
    "peak_grid_kw",
    "peak_total_charge_kw",
    "peak_total_charge_kw_all_depots",
    "peak_total_charge_kw_any_depot",
    "contract_over_limit_kwh",
    "contract_over_limit_kw_peak",
    "contract_over_limit_slot_count",
    "contract_limit_exceeded",
    "bus_charging_total_kwh",
    "total_bus_charge_kwh",
    "total_bess_charge_kwh",
    "total_charge_input_kwh",
    "bev_charge_input_kwh",
    "bev_charge_to_battery_kwh",
    "bev_charge_loss_kwh",
    "bev_drive_energy_kwh",
    "bev_drive_consumption_kwh",
    "ice_fuel_consumed_l",
    "ice_fuel_l",
    "ice_refueled_l",
    "fuel_consumption_l",
    "refuel_l",
    "grid_co2_kg",
    "electricity_co2_kg",
    "ice_co2_kg",
    "fuel_co2_kg",
    "total_co2_kg",
    "min_soc_ratio",
    "mean_soc_ratio",
    "final_min_soc_ratio",
    "final_mean_soc_ratio",
}


@dataclass(frozen=True)
class ReportingRebuildResult:
    run_dir: Path
    updated_files: list[str]
    ledger_totals: dict[str, float]
    cost_totals: dict[str, float]
    validation_status: dict[str, Any]
    warnings: list[str]


def require_file(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"required file not found: {path}")


def load_json(path: Path) -> dict[str, Any]:
    require_file(path)
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    require_file(path)
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def as_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(parsed) or math.isinf(parsed):
        return default
    return parsed


def sum_column(path: Path, column: str) -> float:
    _, rows = read_csv(path)
    return sum(as_float(row.get(column)) for row in rows)


def max_column(path: Path, column: str) -> float:
    _, rows = read_csv(path)
    values = [as_float(row.get(column)) for row in rows if row.get(column) not in (None, "")]
    return max(values) if values else 0.0


def key_value_rows(path: Path) -> tuple[list[str], list[dict[str, str]], dict[str, int]]:
    fieldnames, rows = read_csv(path)
    if "key" in fieldnames and "value" in fieldnames:
        index = {row["key"]: i for i, row in enumerate(rows)}
        return fieldnames, rows, index
    if "component" in fieldnames and "yen" in fieldnames:
        normalized = [
            {"key": row.get("component", ""), "value": row.get("yen", ""), "unit": "JPY"}
            for row in rows
        ]
        index = {row["key"]: i for i, row in enumerate(normalized)}
        return ["key", "value", "unit"], normalized, index
    if "term" in fieldnames and "value" in fieldnames:
        normalized = [
            {"key": row.get("term", ""), "value": row.get("value", ""), "unit": row.get("unit", "")}
            for row in rows
        ]
        index = {row["key"]: i for i, row in enumerate(normalized)}
        return ["key", "value", "unit"], normalized, index
    raise ValueError(f"expected key/value, component/yen, or term/value CSV: {path}")


def get_kv(rows: list[dict[str, str]], index: dict[str, int], key: str, default: float = 0.0) -> float:
    if key not in index:
        return default
    return as_float(rows[index[key]].get("value"), default)


def set_kv(
    rows: list[dict[str, str]],
    index: dict[str, int],
    key: str,
    value: Any,
    unit: str | None = None,
) -> None:
    text_value = repr(float(value)) if isinstance(value, (int, float)) else str(value)
    if key in index:
        row = rows[index[key]]
        row["value"] = text_value
        if unit is not None and "unit" in row:
            row["unit"] = unit
        return
    row = {"key": key, "value": text_value}
    if rows and "unit" in rows[0]:
        row["unit"] = unit or ""
    rows.append(row)
    index[key] = len(rows) - 1


def sha256_file(path: Path) -> str | None:
    if not path.is_file():
        return None
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_timestamp(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    return value.replace(" ", "T")


def bess_timestamp(row: dict[str, str]) -> str:
    if row.get("timestamp"):
        return normalize_timestamp(row["timestamp"])
    if row.get("date") and row.get("time"):
        return normalize_timestamp(f"{row['date']}T{row['time']}:00" if len(row["time"]) == 5 else f"{row['date']}T{row['time']}")
    if row.get("time"):
        return normalize_timestamp(row["time"])
    return ""


def infer_bess_capacity(bess_rows: list[dict[str, str]]) -> float:
    candidates: list[float] = []
    for row in bess_rows:
        soc = as_float(row.get("bess_soc_kwh"))
        pct = as_float(row.get("bess_soc_percent"))
        if soc > 0 and pct > 0:
            candidates.append(soc * 100.0 / pct)
    if candidates:
        return median(candidates)

    soc_max = max(as_float(row.get("bess_soc_max_kwh")) for row in bess_rows) if bess_rows else 0.0
    if soc_max > 0:
        return soc_max / 0.8
    return 0.0


def copy_run(input_dir: Path, output_dir: Path, overwrite: bool) -> None:
    input_dir = input_dir.resolve()
    output_dir = output_dir.resolve()
    if not input_dir.is_dir():
        raise FileNotFoundError(f"input-run-dir not found: {input_dir}")
    if input_dir == output_dir:
        raise ValueError("input-run-dir and output-run-dir must differ")
    if output_dir.exists():
        if not overwrite:
            raise FileExistsError(f"ERROR: output-run-dir already exists: {output_dir}")
        shutil.rmtree(output_dir)
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(input_dir, output_dir)


def update_energy_flow_bess_metadata(run_dir: Path) -> dict[str, Any]:
    energy_path = run_dir / "graph" / "energy_flow_ledger.csv"
    bess_path = run_dir / "graph" / "bess_timeseries.csv"
    energy_fields, energy_rows = read_csv(energy_path)
    bess_fields, bess_rows = read_csv(bess_path)
    has_explicit_soc_transition = {
        "bess_soc_start_kwh",
        "bess_soc_end_kwh",
    }.issubset(bess_fields)

    required_cols = [
        "bess_capacity_kwh",
        "bess_soc_min_kwh",
        "bess_soc_max_kwh",
        "bess_terminal_soc_min_kwh",
        "bess_soc_start_kwh",
        "bess_soc_end_kwh",
    ]
    for col in required_cols:
        if col not in energy_fields:
            energy_fields.append(col)

    capacity = infer_bess_capacity(bess_rows)
    conditions = _load_optional_json(run_dir / "simulation_conditions.json")
    raw_assets = conditions.get("depot_energy_assets") or []
    if isinstance(raw_assets, dict):
        raw_assets = [raw_assets]
    efficiency_by_depot: dict[str, dict[str, float]] = {}
    for asset in raw_assets:
        if not isinstance(asset, dict):
            continue
        depot_id = str(asset.get("depot_id") or "").strip()
        if not depot_id:
            continue
        efficiency_by_depot[depot_id] = {
            "charge": min(max(as_float(asset.get("bess_charge_efficiency"), 1.0), 1.0e-9), 1.0),
            "discharge": min(max(as_float(asset.get("bess_discharge_efficiency"), 1.0), 1.0e-9), 1.0),
        }
    by_timestamp = {bess_timestamp(row): row for row in bess_rows if bess_timestamp(row)}
    by_index = {str(i): row for i, row in enumerate(bess_rows)}

    join_key = "row_number"
    matched_rows: list[dict[str, str]] = []
    if energy_rows and all(normalize_timestamp(row.get("timestamp", "")) in by_timestamp for row in energy_rows):
        join_key = "timestamp"
        matched_rows = [by_timestamp[normalize_timestamp(row.get("timestamp", ""))] for row in energy_rows]
    elif energy_rows and all(row.get("time") in by_timestamp for row in energy_rows if row.get("time")):
        join_key = "time"
        matched_rows = [by_timestamp[row.get("time", "")] for row in energy_rows]
    elif energy_rows and all(row.get("slot_index") in by_index for row in energy_rows):
        join_key = "slot_index"
        matched_rows = [by_index[row.get("slot_index", "")] for row in energy_rows]
    else:
        if len(energy_rows) != len(bess_rows):
            raise ValueError(
                "cannot align energy_flow_ledger.csv and bess_timeseries.csv by timestamp, slot_index, or row number"
            )
        matched_rows = bess_rows

    for row, bess in zip(energy_rows, matched_rows):
        row["bess_capacity_kwh"] = repr(float(capacity))
        row["bess_soc_min_kwh"] = repr(as_float(bess.get("bess_soc_min_kwh")))
        row["bess_soc_max_kwh"] = repr(as_float(bess.get("bess_soc_max_kwh")))
        row["bess_terminal_soc_min_kwh"] = repr(as_float(bess.get("bess_terminal_soc_min_kwh")))
        if "bess_soc_start_kwh" in bess or "bess_soc_kwh" in bess:
            row["bess_soc_start_kwh"] = repr(
                as_float(bess.get("bess_soc_start_kwh"), as_float(bess.get("bess_soc_kwh")))
            )
        if "bess_soc_end_kwh" in bess or "bess_soc_kwh" in bess:
            row["bess_soc_end_kwh"] = repr(
                as_float(bess.get("bess_soc_end_kwh"), as_float(bess.get("bess_soc_kwh")))
            )

    write_csv(energy_path, energy_fields, energy_rows)
    return {
        "bess_metadata_source": "graph/bess_timeseries.csv",
        "bess_metadata_join_key": join_key,
        "bess_capacity_kwh": capacity,
        "bess_soc_min_kwh": max(as_float(row.get("bess_soc_min_kwh")) for row in bess_rows) if bess_rows else 0.0,
        "bess_soc_max_kwh": max(as_float(row.get("bess_soc_max_kwh")) for row in bess_rows) if bess_rows else 0.0,
        "bess_efficiency_by_depot": efficiency_by_depot,
        "bess_soc_transition_verifiable": has_explicit_soc_transition,
        "bess_soc_transition_source": (
            "explicit_start_end_columns"
            if has_explicit_soc_transition
            else "legacy_single_soc_column_not_transition_verifiable"
        ),
    }


def compute_ledger_totals(run_dir: Path) -> dict[str, float]:
    graph = run_dir / "graph"
    energy_path = graph / "energy_flow_ledger.csv"
    cost_path = graph / "cost_timeseries.csv"
    fuel_path = graph / "fuel_canonical_ledger.csv"
    fuel_ts_path = graph / "fuel_timeseries.csv"
    co2_path = graph / "co2_timeseries.csv"

    totals = {
        "pv_generation_kwh": sum_column(energy_path, "pv_generation_kwh"),
        "pv_to_bus_kwh": sum_column(energy_path, "pv_to_bus_kwh"),
        "pv_to_bess_kwh": sum_column(energy_path, "pv_to_bess_kwh"),
        "pv_curtailed_kwh": sum_column(energy_path, "pv_curtailed_kwh"),
        "pv_curtailment_kwh": sum_column(energy_path, "pv_curtailment_kwh"),
        "grid_to_bus_kwh": sum_column(energy_path, "grid_to_bus_kwh"),
        "grid_to_bess_kwh": sum_column(energy_path, "grid_to_bess_kwh"),
        "grid_import_kwh": sum_column(energy_path, "grid_import_kwh"),
        "bess_to_bus_kwh": sum_column(energy_path, "bess_to_bus_kwh"),
        "bus_charging_total_kwh": sum_column(energy_path, "bus_charging_total_kwh"),
        "bess_charge_kwh": sum_column(energy_path, "bess_charge_kwh"),
        "bess_discharge_kwh": sum_column(energy_path, "bess_discharge_kwh"),
        "pv_to_bus_cost_jpy": sum_column(energy_path, "pv_to_bus_cost_jpy"),
        "pv_to_bess_cost_jpy": sum_column(energy_path, "pv_to_bess_cost_jpy"),
        "bess_to_bus_cost_jpy": sum_column(energy_path, "bess_to_bus_cost_jpy"),
        "bess_total_flow_cost_jpy": sum_column(energy_path, "bess_total_flow_cost_jpy"),
        "peak_grid_import_kw": max_column(energy_path, "grid_import_kw"),
        "grid_purchase_cost_jpy": sum_column(cost_path, "grid_purchase_cost_jpy"),
        "fuel_consumption_l": sum_column(fuel_path, "fuel_consumption_l"),
        "fuel_cost_jpy": sum_column(fuel_path, "fuel_cost_jpy"),
        "fuel_timeseries_consumption_l": sum_column(fuel_ts_path, "fuel_consumption_l"),
        "fuel_timeseries_cost_jpy": sum_column(fuel_ts_path, "fuel_cost_jpy"),
        "grid_co2_kg": sum_column(co2_path, "grid_co2_kg"),
        "ice_co2_kg": sum_column(co2_path, "ice_co2_kg"),
        "total_co2_kg": sum_column(co2_path, "total_co2_kg"),
    }
    return totals


def compute_assignment_summary(run_dir: Path) -> dict[str, int]:
    path = run_dir / "graph" / "trip_assignment.csv"
    if not path.is_file():
        return {}
    _, rows = read_csv(path)

    def is_served(row: dict[str, str]) -> bool:
        return (
            str(row.get("served_flag", "true")).strip().lower()
            not in {"false", "0", "no"}
            and bool(str(row.get("assigned_vehicle_id") or "").strip())
        )

    served_rows = [row for row in rows if is_served(row)]
    unserved_rows = [row for row in rows if not is_served(row)]

    def vehicle_type(row: dict[str, str]) -> str:
        return str(row.get("assigned_vehicle_type") or row.get("vehicle_type") or "").strip().upper()

    def service_date(row: dict[str, str]) -> str:
        for key in ("actual_departure", "scheduled_departure", "service_date"):
            value = str(row.get(key) or "").strip()
            if value:
                return value[:10]
        return ""

    unique_trip_ids = {str(row.get("trip_id") or "").strip() for row in served_rows}
    unique_trip_ids.discard("")
    unserved_trip_ids = {str(row.get("trip_id") or "").strip() for row in unserved_rows}
    unserved_trip_ids.discard("")
    used_vehicle_ids = {str(row.get("assigned_vehicle_id") or "").strip() for row in served_rows}
    used_vehicle_ids.discard("")
    used_vehicle_days = {
        (str(row.get("assigned_vehicle_id") or "").strip(), service_date(row))
        for row in served_rows
    }
    return {
        "served_trip_count": len(unique_trip_ids),
        "unserved_trip_count": len(unserved_trip_ids),
        "bev_trip_count": len(
            {str(row.get("trip_id") or "") for row in served_rows if vehicle_type(row) == "BEV"}
        ),
        "ice_trip_count": len(
            {str(row.get("trip_id") or "") for row in served_rows if vehicle_type(row) == "ICE"}
        ),
        "used_vehicle_count": len(used_vehicle_ids),
        "used_vehicle_day_count": len(used_vehicle_days),
    }


def update_cost_breakdown(
    run_dir: Path,
    totals: dict[str, float],
    assignment: dict[str, int],
) -> dict[str, Any]:
    path = run_dir / "cost_breakdown_detail.csv"
    fields, rows, index = key_value_rows(path)

    ledger_path = run_dir / "graph" / "canonical_cost_ledger.json"
    if ledger_path.exists():
        ledger = load_json(ledger_path)
        components = dict(ledger.get("components") or {})
        details = dict(ledger.get("details") or {})
        co2_metadata = dict(ledger.get("co2") or {})
        usage = dict(ledger.get("usage") or {})
        ledger_source = str(ledger_path.relative_to(run_dir))
        accounting_total = as_float(ledger.get("accounting_total_cost_jpy"))
        accounting_residual = as_float(ledger.get("accounting_residual_jpy"))
    else:
        # Backward-compatible read path for pre-v1 runs. This still consumes
        # solver-evaluated graph components; it never infers a missing cost
        # from mutable reporting CSVs.
        legacy_path = run_dir / "graph" / "cost_breakdown.json"
        legacy = load_json(legacy_path)
        legacy_components = dict(legacy.get("components") or {})
        details = {
            "grid_purchase_cost_jpy": as_float(
                legacy_components.get("grid_energy_cost_jpy")
            ),
            "bess_total_flow_cost_jpy": (
                as_float(legacy_components.get("pv_self_consumption_cost_jpy"))
                + as_float(legacy_components.get("bess_discharge_cost_jpy"))
                + as_float(legacy_components.get("pv_curtail_cost_jpy"))
            ),
        }
        co2_metadata = {}
        usage = {
            "vehicle_usage_cost_jpy_per_used_bus": as_float(
                legacy_components.get("vehicle_usage_cost_jpy_per_used_bus")
            )
        }
        components = {
            "electricity_cost_jpy": as_float(
                legacy_components.get("electricity_energy_cost")
            ),
            "fuel_cost_jpy": as_float(legacy_components.get("fuel_cost_final")),
            "demand_charge_cost_jpy": as_float(
                legacy_components.get("demand_charge_cost")
            ),
            "contract_overage_cost_jpy": as_float(
                legacy_components.get("contract_overage_cost_jpy")
            ),
            "vehicle_fixed_cost_jpy": as_float(
                legacy_components.get("vehicle_fixed_cost")
            ),
            "vehicle_usage_cost_jpy": as_float(
                legacy_components.get("vehicle_usage_cost_jpy")
            ),
            "driver_cost_jpy": as_float(legacy_components.get("driver_cost")),
            "unserved_penalty_jpy": as_float(
                legacy_components.get("unserved_trip_penalty")
            ),
            "switch_cost_jpy": 0.0,
            "battery_degradation_cost_jpy": as_float(
                legacy_components.get("battery_degradation_cost")
            ),
            "deviation_cost_jpy": 0.0,
            "co2_cost_jpy": as_float(legacy_components.get("co2_cost")),
        }
        accounting_total = float(sum(components.values()))
        accounting_residual = as_float(legacy.get("total_cost")) - accounting_total
        ledger_source = str(legacy_path.relative_to(run_dir))

    electricity_cost = as_float(components.get("electricity_cost_jpy"))
    grid_purchase_cost = as_float(details.get("grid_purchase_cost_jpy"))
    bess_total_flow_cost = as_float(details.get("bess_total_flow_cost_jpy"))
    fuel_cost = as_float(components.get("fuel_cost_jpy"))
    energy_cost = electricity_cost + fuel_cost
    demand_charge = as_float(components.get("demand_charge_cost_jpy"))
    vehicle_cost = as_float(components.get("vehicle_fixed_cost_jpy"))
    vehicle_usage_cost = as_float(components.get("vehicle_usage_cost_jpy"))
    vehicle_usage_unit_cost = as_float(
        usage.get("vehicle_usage_cost_jpy_per_used_bus")
    )
    battery_degradation_cost = as_float(
        components.get("battery_degradation_cost_jpy")
    )
    co2_cost = as_float(components.get("co2_cost_jpy"))
    carbon_price = as_float(
        co2_metadata.get(
            "carbon_price_jpy_per_kg",
            (
                co2_cost / totals["total_co2_kg"]
                if totals["total_co2_kg"] > 0.0
                else 0.0
            ),
        )
    )

    set_kv(rows, index, "electricity_cost", electricity_cost, "JPY")
    set_kv(rows, index, "electricity_cost_final", electricity_cost, "")
    set_kv(rows, index, "grid_purchase_cost", grid_purchase_cost, "")
    set_kv(
        rows,
        index,
        "pv_self_consumption_cost_jpy",
        totals["pv_to_bus_cost_jpy"] + totals["pv_to_bess_cost_jpy"],
        "JPY",
    )
    set_kv(rows, index, "bess_discharge_cost", totals["bess_to_bus_cost_jpy"], "JPY")
    set_kv(rows, index, "energy_cost", energy_cost, "JPY")
    set_kv(rows, index, "fuel_cost", fuel_cost, "JPY")
    set_kv(rows, index, "fuel_cost_final", fuel_cost, "JPY")
    set_kv(rows, index, "fuel_cost_provisional", fuel_cost, "JPY")
    set_kv(rows, index, "fuel_cost_provisional_leftover", fuel_cost, "JPY")
    set_kv(rows, index, "total_fuel_cost", fuel_cost, "")
    set_kv(rows, index, "demand_charge", demand_charge, "JPY")
    set_kv(rows, index, "total_demand_charge", demand_charge, "JPY")
    set_kv(rows, index, "vehicle_cost", vehicle_cost, "JPY")
    set_kv(
        rows,
        index,
        "driver_cost",
        as_float(components.get("driver_cost_jpy")),
        "JPY",
    )
    set_kv(
        rows,
        index,
        "contract_overage_cost",
        as_float(components.get("contract_overage_cost_jpy")),
        "JPY",
    )
    set_kv(
        rows,
        index,
        "unserved_penalty",
        as_float(components.get("unserved_penalty_jpy")),
        "JPY",
    )
    set_kv(
        rows,
        index,
        "switch_cost",
        as_float(components.get("switch_cost_jpy")),
        "JPY",
    )
    set_kv(
        rows,
        index,
        "deviation_cost",
        as_float(components.get("deviation_cost_jpy")),
        "JPY",
    )
    set_kv(rows, index, "vehicle_usage_cost", vehicle_usage_cost, "JPY")
    set_kv(rows, index, "vehicle_usage_cost_jpy", vehicle_usage_cost, "JPY")
    set_kv(
        rows,
        index,
        "vehicle_usage_cost_jpy_per_used_bus",
        vehicle_usage_unit_cost,
        "JPY/vehicle-day",
    )
    if assignment:
        set_kv(rows, index, "used_vehicle_day_count", assignment["used_vehicle_day_count"], "vehicle-day")
    set_kv(rows, index, "battery_degradation_cost", battery_degradation_cost, "JPY")
    set_kv(rows, index, "degradation_cost", battery_degradation_cost, "JPY")
    set_kv(rows, index, "total_degradation_cost", battery_degradation_cost, "JPY")

    set_kv(rows, index, "grid_co2_kg", totals["grid_co2_kg"], "kg-CO2")
    set_kv(rows, index, "grid_electricity_co2_kg", totals["grid_co2_kg"], "")
    set_kv(rows, index, "power_generation_co2_kg", totals["grid_co2_kg"], "")
    set_kv(rows, index, "ice_co2_kg", totals["ice_co2_kg"], "")
    set_kv(rows, index, "ice_bus_co2_kg", totals["ice_co2_kg"], "")
    set_kv(rows, index, "engine_bus_co2_kg", totals["ice_co2_kg"], "")
    set_kv(rows, index, "total_co2_kg", totals["total_co2_kg"], "kg-CO2")
    set_kv(rows, index, "co2_cost", co2_cost, "JPY")

    real_total = accounting_total
    total_with_assets = real_total + get_kv(rows, index, "pv_asset_cost") + get_kv(rows, index, "bess_asset_cost")
    set_kv(rows, index, "total_cost", real_total, "JPY")
    set_kv(rows, index, "total_cost_with_assets", total_with_assets, "")
    set_kv(rows, index, "total_operating_cost", real_total, "JPY")
    set_kv(rows, index, "accounting_residual_jpy", accounting_residual, "JPY")

    write_csv(path, fields, rows)
    return {
        "electricity_cost": electricity_cost,
        "grid_purchase_cost": grid_purchase_cost,
        "bess_total_flow_cost": bess_total_flow_cost,
        "demand_charge": demand_charge,
        "fuel_cost": fuel_cost,
        "energy_cost": energy_cost,
        "co2_cost": co2_cost,
        "battery_degradation_cost": battery_degradation_cost,
        "vehicle_usage_cost": vehicle_usage_cost,
        "carbon_price_jpy_per_kg": carbon_price,
        "accounting_residual_jpy": accounting_residual,
        "canonical_cost_ledger_source": ledger_source,
        "total_cost": real_total,
        "total_cost_with_assets": total_with_assets,
    }


def objective_value_from_breakdown(run_dir: Path, fallback: float) -> float:
    path = run_dir / "objective_breakdown.csv"
    fields, rows, index = key_value_rows(path)
    return get_kv(rows, index, "objective_value", fallback)


def update_summary(
    run_dir: Path,
    cost: dict[str, float],
    assignment: dict[str, int],
) -> dict[str, Any]:
    path = run_dir / "summary.json"
    summary = load_json(path)
    objective_value = objective_value_from_breakdown(run_dir, as_float(summary.get("objective_value_jpy", summary.get("objective_value"))))
    total_cost = cost["total_cost"]
    solver_objective_matches_accounting_total = bool(
        summary.get("solver_objective_matches_accounting_total", False)
    ) and abs(objective_value - total_cost) <= 1.0e-6
    objective_is_actual_cost = solver_objective_matches_accounting_total

    summary["objective_value"] = objective_value
    summary["objective_value_jpy"] = objective_value
    summary["objective_value_unit"] = "JPY"
    summary["total_cost_jpy"] = total_cost
    summary["accounting_total_cost_jpy"] = total_cost
    summary["reported_total_cost_jpy"] = total_cost
    summary["gross_operating_cost_jpy"] = total_cost
    if summary.get("validated_operating_cost_jpy") is not None:
        summary["validated_operating_cost_jpy"] = total_cost
    summary["grid_purchase_cost_jpy"] = cost["grid_purchase_cost"]
    summary["energy_cost_jpy"] = cost["electricity_cost"]
    summary["demand_charge_cost_jpy"] = cost["demand_charge"]
    summary["fuel_cost_jpy"] = cost["fuel_cost"]
    summary["co2_cost_jpy"] = cost["co2_cost"]
    summary["accounting_residual_jpy"] = cost["accounting_residual_jpy"]
    summary["canonical_cost_ledger_source"] = cost[
        "canonical_cost_ledger_source"
    ]
    summary["objective_is_actual_cost"] = objective_is_actual_cost
    summary["solver_objective_matches_accounting_total"] = solver_objective_matches_accounting_total
    if assignment:
        summary["trip_count_served"] = assignment["served_trip_count"]
        summary["trip_count_unserved"] = assignment["unserved_trip_count"]
        summary["vehicle_count_used"] = assignment["used_vehicle_count"]
        summary["trip_count_by_type"] = {
            "BEV": assignment["bev_trip_count"],
            "ICE": assignment["ice_trip_count"],
        }
    summary["cost_definition"] = {
        "total_cost_jpy": "gross operating cost from graph/canonical_cost_ledger.json",
        "reported_total_cost_jpy": "same as gross_operating_cost_jpy",
        "gross_operating_cost_jpy": "sum of solver-evaluated canonical cost ledger terms",
        "objective_value_jpy": "solver objective value; equals reported_total_cost_jpy only when objective_is_actual_cost=true",
        "objective_is_actual_cost": objective_is_actual_cost,
        "solver_objective_matches_accounting_total": solver_objective_matches_accounting_total,
    }
    write_json(path, summary)
    return summary


def update_kpi_summary(
    run_dir: Path,
    totals: dict[str, float],
    cost: dict[str, float],
    bess_metadata: dict[str, Any],
    assignment: dict[str, int],
) -> dict[str, Any]:
    path = run_dir / "graph" / "kpi_summary.json"
    kpi = load_json(path)
    objective_value = objective_value_from_breakdown(run_dir, as_float(kpi.get("objective_value_jpy", kpi.get("objective_value"))))
    solver_objective_matches_accounting_total = bool(
        kpi.get("solver_objective_matches_accounting_total", False)
    ) and abs(objective_value - cost["total_cost"]) <= 1.0e-6
    objective_is_actual_cost = solver_objective_matches_accounting_total

    energy_keys = [
        "pv_generation_kwh",
        "pv_to_bus_kwh",
        "pv_to_bess_kwh",
        "pv_curtailed_kwh",
        "pv_curtailment_kwh",
        "grid_to_bus_kwh",
        "grid_to_bess_kwh",
        "grid_import_kwh",
        "bess_to_bus_kwh",
        "bus_charging_total_kwh",
        "bess_charge_kwh",
        "bess_discharge_kwh",
        "peak_grid_import_kw",
    ]
    for key in energy_keys:
        if key in totals:
            kpi[key] = totals[key]

    kpi["grid_total_kwh"] = totals["grid_import_kwh"]
    kpi["peak_grid_kw"] = totals["peak_grid_import_kw"]
    kpi["bess_discharge_to_bus_kwh"] = totals["bess_discharge_kwh"]

    cost_keys = {
        "grid_purchase_cost_jpy": cost["grid_purchase_cost"],
        "demand_charge_cost_jpy": cost["demand_charge"],
        "fuel_cost_jpy": cost["fuel_cost"],
        "co2_cost_jpy": cost["co2_cost"],
        "battery_degradation_cost_jpy": cost["battery_degradation_cost"],
        "pv_to_bus_cost_jpy": totals["pv_to_bus_cost_jpy"],
        "pv_to_bess_cost_jpy": totals["pv_to_bess_cost_jpy"],
        "bess_to_bus_cost_jpy": totals["bess_to_bus_cost_jpy"],
        "bess_total_flow_cost_jpy": totals["bess_total_flow_cost_jpy"],
    }
    for key, value in cost_keys.items():
        kpi[key] = value

    kpi["energy_cost_jpy"] = cost["energy_cost"]
    kpi["demand_cost_jpy"] = cost["demand_charge"]
    kpi["total_cost_jpy"] = cost["total_cost"]
    kpi["accounting_total_cost_jpy"] = cost["total_cost"]
    kpi["gross_operating_cost_jpy"] = cost["total_cost"]
    kpi["reported_total_cost_jpy"] = cost["total_cost"]
    kpi["vehicle_usage_cost_jpy"] = cost["vehicle_usage_cost"]
    kpi["accounting_residual_jpy"] = cost["accounting_residual_jpy"]
    kpi["canonical_cost_ledger_source"] = cost[
        "canonical_cost_ledger_source"
    ]
    if kpi.get("validated_operating_cost_jpy") is not None:
        kpi["validated_operating_cost_jpy"] = cost["total_cost"]
    kpi["objective_value"] = objective_value
    kpi["objective_value_jpy"] = objective_value
    kpi["objective_is_actual_cost"] = objective_is_actual_cost
    kpi["solver_objective_matches_accounting_total"] = solver_objective_matches_accounting_total
    if assignment:
        kpi.update(assignment)

    kpi["fuel_consumption_l"] = totals["fuel_consumption_l"]
    kpi["ice_fuel_consumed_l"] = totals["fuel_consumption_l"]
    kpi["ice_fuel_l"] = totals["fuel_consumption_l"]
    kpi["grid_co2_kg"] = totals["grid_co2_kg"]
    kpi["electricity_co2_kg"] = totals["grid_co2_kg"]
    kpi["ice_co2_kg"] = totals["ice_co2_kg"]
    kpi["fuel_co2_kg"] = totals["ice_co2_kg"]
    kpi["total_co2_kg"] = totals["total_co2_kg"]

    kpi.setdefault("energy", {})
    kpi["energy"].update({key: kpi[key] for key in energy_keys if key in kpi})
    kpi["energy"]["grid_import_kwh"] = totals["grid_import_kwh"]
    kpi["energy"]["bess_discharge_kwh"] = totals["bess_discharge_kwh"]

    kpi.setdefault("fuel", {})
    kpi["fuel"].update(
        {
            "fuel_consumption_l": totals["fuel_consumption_l"],
            "fuel_cost_jpy": cost["fuel_cost"],
            "fuel_source_of_truth": "fuel_canonical_ledger",
        }
    )

    kpi.setdefault("co2", {})
    kpi["co2"].update(
        {
            "total_co2_kg": totals["total_co2_kg"],
            "grid_co2_kg": totals["grid_co2_kg"],
            "ice_co2_kg": totals["ice_co2_kg"],
            "co2_boundary": "grid_plus_ice",
            "co2_accounting_method": "grid_import_based",
        }
    )

    kpi.setdefault("cost", {})
    kpi["cost"].update(cost_keys)
    kpi["cost"].update(
        {
            "energy_cost_jpy": cost["energy_cost"],
            "gross_operating_cost_jpy": cost["total_cost"],
            "reported_total_cost_jpy": cost["total_cost"],
            "total_cost_jpy": cost["total_cost"],
            "objective_value": objective_value,
            "objective_value_jpy": objective_value,
            "objective_is_actual_cost": objective_is_actual_cost,
            "accounting_residual_jpy": cost["accounting_residual_jpy"],
            "canonical_cost_ledger_source": cost[
                "canonical_cost_ledger_source"
            ],
        }
    )

    kpi.setdefault("bess", {})
    kpi["bess"].update(
        {
            "capacity_kwh": bess_metadata["bess_capacity_kwh"],
            "soc_min_kwh": bess_metadata["bess_soc_min_kwh"],
            "soc_max_kwh": bess_metadata["bess_soc_max_kwh"],
            "pv_to_bess_kwh": totals["pv_to_bess_kwh"],
            "grid_to_bess_kwh": totals["grid_to_bess_kwh"],
            "bess_to_bus_kwh": totals["bess_to_bus_kwh"],
            "bess_charge_kwh": totals["bess_charge_kwh"],
            "bess_discharge_kwh": totals["bess_discharge_kwh"],
            "pv_to_bess_cost_jpy": totals["pv_to_bess_cost_jpy"],
            "pv_to_bus_cost_jpy": totals["pv_to_bus_cost_jpy"],
            "bess_to_bus_cost_jpy": totals["bess_to_bus_cost_jpy"],
            "bess_total_flow_cost_jpy": totals["bess_total_flow_cost_jpy"],
        }
    )

    kpi.setdefault("metadata", {})
    kpi["metadata"].update(
        {
            "bess_metadata_source": bess_metadata["bess_metadata_source"],
            "bess_metadata_join_key": bess_metadata["bess_metadata_join_key"],
            "bess_soc_transition_verifiable": bess_metadata[
                "bess_soc_transition_verifiable"
            ],
            "bess_soc_transition_source": bess_metadata["bess_soc_transition_source"],
        }
    )
    kpi["bess_metadata_source"] = bess_metadata["bess_metadata_source"]
    kpi["bess_metadata_join_key"] = bess_metadata["bess_metadata_join_key"]
    kpi["bess_soc_transition_verifiable"] = bess_metadata[
        "bess_soc_transition_verifiable"
    ]
    kpi["bess_soc_transition_source"] = bess_metadata["bess_soc_transition_source"]
    kpi["cost_definition"] = {
        "total_cost_jpy": "gross operating cost based on canonical reporting ledgers",
        "reported_total_cost_jpy": "same as gross_operating_cost_jpy",
        "gross_operating_cost_jpy": "actual operating cost terms from reporting ledgers",
        "objective_value_jpy": "solver/fallback objective value; may include rewards and penalties",
        "objective_is_actual_cost": objective_is_actual_cost,
    }

    write_json(path, kpi)
    root_path = run_dir / "kpi_summary.json"
    if root_path.is_file():
        root_kpi = load_json(root_path)
        authoritative_keys = {
            *energy_keys,
            *cost_keys,
            "grid_total_kwh",
            "peak_grid_kw",
            "bess_discharge_to_bus_kwh",
            "energy_cost_jpy",
            "demand_cost_jpy",
            "total_cost_jpy",
            "accounting_total_cost_jpy",
            "gross_operating_cost_jpy",
            "reported_total_cost_jpy",
            "vehicle_usage_cost_jpy",
            "validated_operating_cost_jpy",
            "objective_value",
            "objective_value_jpy",
            "objective_is_actual_cost",
            "solver_objective_matches_accounting_total",
            "fuel_consumption_l",
            "ice_fuel_consumed_l",
            "total_co2_kg",
            "grid_co2_kg",
            "ice_co2_kg",
            "served_trip_count",
            "unserved_trip_count",
            "bev_trip_count",
            "ice_trip_count",
            "used_vehicle_count",
            "used_vehicle_day_count",
        }
        root_kpi.update({key: kpi.get(key) for key in authoritative_keys if key in kpi})
        write_json(root_path, root_kpi)
    return kpi


def _load_optional_json(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {}
    try:
        return load_json(path)
    except (OSError, json.JSONDecodeError):
        return {}


def _first_nonnegative_int(*values: Any) -> int | None:
    for value in values:
        if value in (None, ""):
            continue
        try:
            return max(int(value), 0)
        except (TypeError, ValueError):
            continue
    return None


def _infeasible_failure_stage(
    solver_status: str,
    solver_settings: dict[str, Any],
) -> str:
    stage2_status = str(solver_settings.get("stage2_solver_status") or "").lower()
    if stage2_status and stage2_status not in {"optimal", "feasible", "solved_feasible"}:
        return "stage2_energy_dispatch"
    stage1_status = str(solver_settings.get("stage1_solver_status") or "").lower()
    if stage1_status in {"infeasible", "inf_or_unbd", "unbounded", "no_valid_incumbent"}:
        return "stage1_assignment"
    if "infeasible" in solver_status.lower():
        return "postsolve_validation"
    return "result_validation"


def _null_fields(payload: dict[str, Any], fields: set[str]) -> None:
    for field in fields:
        if field in payload:
            payload[field] = None


def _gate_results_workbook(
    run_dir: Path,
    *,
    result_status: str,
    failure_stage: str,
    served_trip_count: int | None,
    unserved_trip_count: int | None,
) -> None:
    """Invalidate reader-facing workbook cells without deleting diagnostics."""

    workbook_path = run_dir / "results.xlsx"
    if not workbook_path.is_file():
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        return

    workbook = load_workbook(workbook_path)
    if "summary" in workbook.sheetnames:
        summary_sheet = workbook["summary"]
        summary_values = {
            "objective_value": None,
            "trip_count_served": served_trip_count,
            "trip_count_unserved": unserved_trip_count,
        }
        for row in summary_sheet.iter_rows(min_row=2):
            key = str(row[0].value or "")
            if key in summary_values:
                row[1].value = summary_values[key]

    if "cost_breakdown" in workbook.sheetnames:
        cost_sheet = workbook["cost_breakdown"]
        preserved_flags = {
            "evaluation_feasible": 0.0,
            "objective_is_actual_cost": False,
            "solver_objective_matches_accounting_total": False,
        }
        for row in cost_sheet.iter_rows(min_row=2):
            key = str(row[0].value or "")
            row[1].value = preserved_flags.get(key)

    if "result_status" in workbook.sheetnames:
        del workbook["result_status"]
    status_sheet = workbook.create_sheet("result_status", 0)
    status_sheet.append(["key", "value"])
    status_sheet.append(["result_status", result_status])
    status_sheet.append(["failure_stage", failure_stage])
    status_sheet.append(["research_kpi_eligible", False])
    workbook.save(workbook_path)


def _gate_experiment_report(
    run_dir: Path,
    *,
    result_status: str,
    failure_stage: str,
    unserved_trip_count: int | None,
) -> None:
    report_path = run_dir / "experiment_report.md"
    if not report_path.is_file():
        return
    original = report_path.read_text(encoding="utf-8")
    marker = "<!-- solution-validity-gate -->"
    if marker in original:
        return
    warning = (
        f"{marker}\n"
        "# INVALID RESULT — KPI使用禁止\n\n"
        f"- result_status: `{result_status}`\n"
        f"- failure_stage: `{failure_stage}`\n"
        f"- unserved_trip_count: `{unserved_trip_count}`\n"
        "- research_kpi_eligible: `false`\n"
        "- 費用・エネルギーフロー・CO₂は評価不能であり、下記の旧数値を研究結果として使用しない。\n\n"
        "---\n\n"
    )
    report_path.write_text(warning + original, encoding="utf-8")


def update_experiment_report_metrics(
    run_dir: Path,
    *,
    cost: dict[str, float],
    totals: dict[str, float],
) -> None:
    """Synchronize reader-facing report values with finalized ledgers."""

    report_path = run_dir / "experiment_report.md"
    if not report_path.is_file():
        return
    report = report_path.read_text(encoding="utf-8")
    solver_settings = _load_optional_json(run_dir / "solver_settings.json")

    replacements = {
        r"^\| (?:総コスト|会計総費用) \| .*? \|$": f"| 会計総費用 | {cost['total_cost']:,.2f} JPY |",
        r"^\| 　電気代 \| .*? \|$": f"| 　電気代 | {cost['electricity_cost']:,.2f} JPY |",
        r"^\| 　軽油代 \| .*? \|$": f"| 　軽油代 | {cost['fuel_cost']:,.2f} JPY |",
        r"^\| 　デマンド料金 \| .*? \|$": f"| 　デマンド料金 | {cost['demand_charge']:,.2f} JPY |",
        r"^\| CO₂排出量 \| .*? \|$": f"| CO₂排出量 | {totals['total_co2_kg']:,.4f} kg |",
    }
    for pattern, replacement in replacements.items():
        report = re.sub(pattern, replacement, report, flags=re.MULTILINE)

    requested_gap = solver_settings.get("mip_gap_requested_percent")
    achieved_gap = solver_settings.get("mip_gap_achieved_percent")
    if requested_gap is not None:
        report = re.sub(
            r"^\| MIP Gap 目標 \| .*? \|$",
            f"| MIP Gap 目標 | {float(requested_gap):.3f} % |",
            report,
            flags=re.MULTILINE,
        )
    if achieved_gap is not None:
        report = re.sub(
            r"^\| MIP Gap 実績 \| .*? \|$",
            f"| MIP Gap 実績 | {float(achieved_gap):.4f} % |",
            report,
            flags=re.MULTILINE,
        )

    marker = "<!-- canonical-accounting-definition -->"
    if marker not in report:
        note = (
            f"{marker}\n"
            "> 会計総費用は最終台帳の費用合計です。目的値は二段階最適化の評価値であり、"
            "両者が一致すると明示された場合を除き、同じ値として扱いません。\n\n"
        )
        report = report.replace("## 結果サマリ\n\n", "## 結果サマリ\n\n" + note, 1)
    report_path.write_text(report, encoding="utf-8")


def apply_solution_validity_gate(
    run_dir: Path,
    summary: dict[str, Any],
    kpi: dict[str, Any],
    validation_rows: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    """Invalidate evaluation KPIs when the canonical result is not usable.

    Raw solver artifacts and ledgers remain untouched for diagnosis.  Only
    reader-facing KPI summaries are gated so an infeasible zero-row ledger
    cannot be misread as a valid zero-cost solution.
    """

    canonical = _load_optional_json(run_dir / "canonical_solver_result.json")
    if not canonical:
        canonical = _load_optional_json(run_dir / "raw" / "canonical_solver_result.json")
    solution_validity = dict(
        summary.get("solution_validity")
        or canonical.get("solution_validity")
        or kpi.get("solution_validity")
        or {}
    )
    solver_status = str(
        summary.get("solver_status")
        or canonical.get("solver_status")
        or kpi.get("solver_status")
        or ""
    )
    has_explicit_validity = "validated_feasible" in solution_validity
    has_canonical_feasibility = "feasible" in canonical
    status_upper = solver_status.upper()
    status_explicitly_invalid = any(
        token in status_upper
        for token in (
            "INFEASIBLE",
            "NO_VALID_INCUMBENT",
            "BASELINE_FALLBACK",
            "POSTSOLVE_REPAIRED",
            "DEBUG_RESULT",
        )
    )
    if not (has_explicit_validity or has_canonical_feasibility or status_explicitly_invalid):
        return summary, kpi, validation_rows

    if has_explicit_validity:
        validated_feasible = bool(solution_validity.get("validated_feasible"))
    elif has_canonical_feasibility:
        validated_feasible = bool(canonical.get("feasible"))
    else:
        validated_feasible = status_upper in {
            "OPTIMAL",
            "FEASIBLE",
            "SOLVED_FEASIBLE",
        }

    if validated_feasible:
        return summary, kpi, validation_rows

    solver_settings = _load_optional_json(run_dir / "solver_settings.json")
    unserved_trip_count = _first_nonnegative_int(
        canonical.get("trip_count_unserved"),
        len(canonical.get("unserved_trip_ids") or [])
        if "unserved_trip_ids" in canonical
        else None,
        summary.get("trip_count_unserved"),
        kpi.get("unserved_trip_count"),
    )
    served_trip_count = _first_nonnegative_int(
        canonical.get("trip_count_served"),
        len(canonical.get("served_trip_ids") or [])
        if "served_trip_ids" in canonical
        else None,
        summary.get("trip_count_served"),
        kpi.get("served_trip_count"),
    )
    result_status = "INFEASIBLE" if "INFEASIBLE" in solver_status.upper() else "INVALID"
    failure_stage = _infeasible_failure_stage(solver_status, solver_settings)
    status_reason = str(
        solution_validity.get("status_reason") or "canonical_result_not_validated_feasible"
    )

    _null_fields(summary, INVALID_SUMMARY_KPI_FIELDS)
    summary.update(
        {
            "result_status": result_status,
            "failure_stage": failure_stage,
            "research_kpi_eligible": False,
            "objective_is_actual_cost": False,
            "solver_objective_matches_accounting_total": False,
            "kpi_eligibility_reason": status_reason,
        }
    )
    if served_trip_count is not None:
        summary["trip_count_served"] = served_trip_count
    if unserved_trip_count is not None:
        summary["trip_count_unserved"] = unserved_trip_count
    summary["cost_definition"] = {
        "availability": "unavailable because the canonical result is not validated feasible",
        "objective_is_actual_cost": False,
        "solver_objective_matches_accounting_total": False,
    }

    _null_fields(kpi, INVALID_KPI_FIELDS)
    kpi.update(
        {
            "result_status": result_status,
            "failure_stage": failure_stage,
            "research_kpi_eligible": False,
            "objective_is_actual_cost": False,
            "solver_objective_matches_accounting_total": False,
            "is_optimization_result": False,
            "result_interpretation": "invalid_result_kpis_unavailable",
            "optimization_status": result_status,
            "physical_feasibility_status": result_status,
            "is_physically_feasible": False,
            "kpi_eligibility_reason": status_reason,
        }
    )
    if served_trip_count is not None:
        kpi["served_trip_count"] = served_trip_count
    if unserved_trip_count is not None:
        kpi["unserved_trip_count"] = unserved_trip_count

    for section_name in ("cost", "fuel", "co2"):
        section = kpi.get(section_name)
        if isinstance(section, dict):
            for key in list(section):
                if key not in {
                    "objective_value_definition",
                    "gross_operating_cost_definition",
                    "fuel_source_of_truth",
                    "co2_boundary",
                    "co2_accounting_method",
                    "bess_co2_source_tracking",
                }:
                    section[key] = None
    energy = kpi.get("energy")
    if isinstance(energy, dict):
        for key in list(energy):
            if key != "pv_generation_kwh":
                energy[key] = None
    bess = kpi.get("bess")
    if isinstance(bess, dict):
        for key in list(bess):
            if key not in {"capacity_kwh", "soc_min_kwh", "soc_max_kwh"}:
                bess[key] = None
    kpi["cost_definition"] = dict(summary["cost_definition"])

    gate_row = validation_row(
        "solution_validity_gate",
        True,
        False,
        severity="ERROR",
        source_files="canonical_solver_result.json;summary.json;solver_settings.json",
    )
    gate_row["message"] = (
        f"{result_status}: evaluation KPIs are unavailable; failure_stage={failure_stage}; "
        f"reason={status_reason}"
    )
    validation_rows = [
        row for row in validation_rows if row.get("check_name") != "solution_validity_gate"
    ]
    validation_rows.append(gate_row)

    write_json(run_dir / "summary.json", summary)
    write_json(run_dir / "graph" / "kpi_summary.json", kpi)
    if (run_dir / "kpi_summary.json").is_file():
        write_json(run_dir / "kpi_summary.json", kpi)
    validation_path = run_dir / "graph" / "data_flow_validation.csv"
    fields, _ = read_csv(validation_path)
    write_csv(validation_path, fields, validation_rows)
    _gate_results_workbook(
        run_dir,
        result_status=result_status,
        failure_stage=failure_stage,
        served_trip_count=served_trip_count,
        unserved_trip_count=unserved_trip_count,
    )
    _gate_experiment_report(
        run_dir,
        result_status=result_status,
        failure_stage=failure_stage,
        unserved_trip_count=unserved_trip_count,
    )
    return summary, kpi, validation_rows


def validation_row(
    check_name: str,
    expected: Any,
    actual: Any,
    tolerance: float = TOL,
    severity: str = "INFO",
    source_files: str = "",
) -> dict[str, Any]:
    if isinstance(expected, bool) or isinstance(actual, bool):
        ok = expected is actual
        diff: Any = 0.0 if ok else 1.0
    else:
        expected_f = as_float(expected)
        actual_f = as_float(actual)
        diff = actual_f - expected_f
        ok = abs(diff) <= tolerance
    return {
        "check_name": check_name,
        "status": "OK" if ok else "NG",
        "expected_value": expected,
        "actual_value": actual,
        "difference": diff,
        "tolerance": tolerance,
        "severity": severity if not ok else "INFO",
        "message": "OK" if ok else f"{check_name} differs by {diff}",
        "source_files": source_files,
    }


def update_data_flow_validation(
    run_dir: Path,
    totals: dict[str, float],
    cost: dict[str, float],
    summary: dict[str, Any],
    kpi: dict[str, Any],
    bess_metadata: dict[str, Any],
) -> list[dict[str, Any]]:
    path = run_dir / "graph" / "data_flow_validation.csv"
    fields, rows = read_csv(path)
    _, cost_rows, cost_index = key_value_rows(run_dir / "cost_breakdown_detail.csv")
    cost_breakdown_total = get_kv(cost_rows, cost_index, "total_cost")
    cost_breakdown_fuel_cost = get_kv(cost_rows, cost_index, "fuel_cost")
    cost_breakdown_ice_co2 = get_kv(cost_rows, cost_index, "ice_co2_kg")
    cost_breakdown_total_co2 = get_kv(cost_rows, cost_index, "total_co2_kg")
    vehicle_source_path = run_dir / "graph" / "vehicle_charging_source_timeseries.csv"
    vehicle_grid_to_bus = sum_column(vehicle_source_path, "grid_to_vehicle_kwh") if vehicle_source_path.exists() else 0.0
    vehicle_pv_to_bus = sum_column(vehicle_source_path, "pv_to_vehicle_kwh") if vehicle_source_path.exists() else 0.0
    vehicle_bess_to_bus = sum_column(vehicle_source_path, "bess_to_vehicle_kwh") if vehicle_source_path.exists() else 0.0
    _, energy_rows = read_csv(run_dir / "graph" / "energy_flow_ledger.csv")
    efficiency_by_depot = dict(bess_metadata.get("bess_efficiency_by_depot") or {})
    bess_transition_error = 0.0
    for row in energy_rows:
        depot_efficiency = dict(efficiency_by_depot.get(str(row.get("depot_id") or "")) or {})
        charge_efficiency = min(max(as_float(depot_efficiency.get("charge"), 1.0), 1.0e-9), 1.0)
        discharge_efficiency = min(max(as_float(depot_efficiency.get("discharge"), 1.0), 1.0e-9), 1.0)
        expected_end = (
            as_float(row.get("bess_soc_start_kwh"))
            + as_float(row.get("bess_charge_kwh")) * charge_efficiency
            - as_float(row.get("bess_discharge_kwh")) / discharge_efficiency
        )
        bess_transition_error += abs(as_float(row.get("bess_soc_end_kwh")) - expected_end)
    required_fields = [
        "check_name",
        "status",
        "expected_value",
        "actual_value",
        "difference",
        "tolerance",
        "severity",
        "message",
        "source_files",
    ]
    if fields != required_fields:
        fields = required_fields

    if bool(bess_metadata.get("bess_soc_transition_verifiable")):
        bess_transition_check = validation_row(
            "bess_soc_transition_balance",
            0.0,
            bess_transition_error,
            tolerance=1.0e-6,
            source_files="energy_flow_ledger.csv;simulation_conditions.json",
        )
    else:
        bess_transition_check = {
            "check_name": "bess_soc_transition_balance",
            "status": "SKIPPED",
            "expected_value": "",
            "actual_value": "",
            "difference": "",
            "tolerance": 1.0e-6,
            "severity": "INFO",
            "message": (
                "BESS SOC transition was not checked because the source artifact has only "
                "bess_soc_kwh and no explicit start/end SOC columns."
            ),
            "source_files": "energy_flow_ledger.csv;graph/bess_timeseries.csv",
        }

    checks = [
        validation_row(
            "pv_generation_balance",
            totals["pv_generation_kwh"],
            totals["pv_to_bus_kwh"] + totals["pv_to_bess_kwh"] + totals["pv_curtailed_kwh"],
            tolerance=1.0e-3,
            source_files="energy_flow_ledger.csv",
        ),
        validation_row(
            "bus_charging_balance",
            totals["bus_charging_total_kwh"],
            totals["grid_to_bus_kwh"] + totals["pv_to_bus_kwh"] + totals["bess_to_bus_kwh"],
            tolerance=1.0e-3,
            source_files="energy_flow_ledger.csv",
        ),
        validation_row(
            "grid_import_balance",
            totals["grid_import_kwh"],
            totals["grid_to_bus_kwh"] + totals["grid_to_bess_kwh"],
            tolerance=1.0e-3,
            source_files="energy_flow_ledger.csv",
        ),
        validation_row(
            "bess_charge_balance",
            totals["bess_charge_kwh"],
            totals["pv_to_bess_kwh"] + totals["grid_to_bess_kwh"],
            tolerance=1.0e-3,
            source_files="energy_flow_ledger.csv",
        ),
        validation_row(
            "bess_discharge_balance",
            totals["bess_discharge_kwh"],
            totals["bess_to_bus_kwh"],
            tolerance=1.0e-3,
            source_files="energy_flow_ledger.csv",
        ),
        bess_transition_check,
        validation_row(
            "vehicle_grid_to_bus_allocation_matches_site_total",
            totals["grid_to_bus_kwh"],
            vehicle_grid_to_bus,
            tolerance=1.0e-3,
            severity="WARNING",
            source_files="vehicle_charging_source_timeseries.csv;energy_flow_ledger.csv",
        ),
        validation_row(
            "vehicle_pv_to_bus_allocation_matches_site_total",
            totals["pv_to_bus_kwh"],
            vehicle_pv_to_bus,
            tolerance=1.0e-3,
            severity="WARNING",
            source_files="vehicle_charging_source_timeseries.csv;energy_flow_ledger.csv",
        ),
        validation_row(
            "vehicle_bess_to_bus_allocation_matches_site_total",
            totals["bess_to_bus_kwh"],
            vehicle_bess_to_bus,
            tolerance=1.0e-3,
            severity="WARNING",
            source_files="vehicle_charging_source_timeseries.csv;energy_flow_ledger.csv",
        ),
        validation_row(
            "kpi_grid_purchase_cost_matches_cost_timeseries",
            totals["grid_purchase_cost_jpy"],
            as_float(kpi.get("grid_purchase_cost_jpy")),
            source_files="kpi_summary.json;cost_timeseries.csv",
        ),
        validation_row(
            "kpi_demand_charge_matches_cost_breakdown",
            cost["demand_charge"],
            as_float(kpi.get("demand_charge_cost_jpy")),
            source_files="kpi_summary.json;cost_breakdown_detail.csv",
        ),
        validation_row(
            "canonical_cost_ledger_accounting_residual",
            0.0,
            cost["accounting_residual_jpy"],
            tolerance=1.0e-6,
            severity="ERROR",
            source_files=cost["canonical_cost_ledger_source"],
        ),
        validation_row(
            "objective_value_matches_canonical_accounting_total",
            cost["total_cost"],
            objective_value_from_breakdown(run_dir, 0.0),
            tolerance=1.0e-6,
            severity="ERROR",
            source_files=(
                f"{cost['canonical_cost_ledger_source']};objective_breakdown.csv"
            ),
        ),
        validation_row(
            "kpi_fuel_cost_matches_fuel_canonical",
            totals["fuel_cost_jpy"],
            as_float(kpi.get("fuel_cost_jpy")),
            severity="ERROR",
            source_files=(
                "kpi_summary.json;fuel_canonical_ledger.csv;"
                f"{cost['canonical_cost_ledger_source']}"
            ),
        ),
        validation_row(
            "kpi_total_cost_matches_cost_breakdown",
            cost_breakdown_total,
            as_float(kpi.get("reported_total_cost_jpy")),
            source_files="kpi_summary.json;cost_breakdown_detail.csv",
        ),
        validation_row(
            "summary_total_cost_matches_cost_breakdown",
            cost_breakdown_total,
            as_float(summary.get("gross_operating_cost_jpy")),
            source_files="summary.json;cost_breakdown_detail.csv",
        ),
        validation_row(
            "summary_objective_value_matches_objective_breakdown",
            objective_value_from_breakdown(run_dir, 0.0),
            as_float(summary.get("objective_value_jpy")),
            source_files="summary.json;objective_breakdown.csv",
        ),
        validation_row(
            "summary_objective_actual_cost_flag_matches_values",
            abs(objective_value_from_breakdown(run_dir, 0.0) - cost_breakdown_total) <= 1.0e-6,
            bool(summary.get("objective_is_actual_cost")),
            source_files="summary.json",
        ),
        validation_row(
            "cost_breakdown_fuel_cost_matches_fuel_canonical",
            totals["fuel_cost_jpy"],
            cost_breakdown_fuel_cost,
            severity="ERROR",
            source_files=(
                "cost_breakdown_detail.csv;fuel_canonical_ledger.csv;"
                f"{cost['canonical_cost_ledger_source']}"
            ),
        ),
        validation_row(
            "cost_breakdown_ice_co2_matches_co2_timeseries",
            totals["ice_co2_kg"],
            cost_breakdown_ice_co2,
            source_files="cost_breakdown_detail.csv;co2_timeseries.csv",
        ),
        validation_row(
            "cost_breakdown_total_co2_matches_co2_timeseries",
            totals["total_co2_kg"],
            cost_breakdown_total_co2,
            source_files="cost_breakdown_detail.csv;co2_timeseries.csv",
        ),
        validation_row(
            "energy_flow_bess_capacity_matches_bess_timeseries",
            bess_metadata["bess_capacity_kwh"],
            max_column(run_dir / "graph" / "energy_flow_ledger.csv", "bess_capacity_kwh"),
            source_files="energy_flow_ledger.csv;bess_timeseries.csv",
        ),
        validation_row(
            "energy_flow_bess_soc_min_matches_bess_timeseries",
            bess_metadata["bess_soc_min_kwh"],
            max_column(run_dir / "graph" / "energy_flow_ledger.csv", "bess_soc_min_kwh"),
            source_files="energy_flow_ledger.csv;bess_timeseries.csv",
        ),
        validation_row(
            "energy_flow_bess_soc_max_matches_bess_timeseries",
            bess_metadata["bess_soc_max_kwh"],
            max_column(run_dir / "graph" / "energy_flow_ledger.csv", "bess_soc_max_kwh"),
            source_files="energy_flow_ledger.csv;bess_timeseries.csv",
        ),
    ]

    replace_names = {row["check_name"] for row in checks}
    replace_names.update(
        {
            "cost_breakdown_balance",
            "kpi_grid_purchase_cost_matches_cost_ledger",
            "kpi_demand_charge_cost_matches_cost_ledger",
            "pv_generation_balance",
            "bus_charging_balance",
            "grid_import_balance",
            "bess_charge_balance",
            "bess_discharge_balance",
            "bess_soc_transition_balance",
            "vehicle_grid_to_bus_allocation_matches_site_total",
            "vehicle_pv_to_bus_allocation_matches_site_total",
            "vehicle_bess_to_bus_allocation_matches_site_total",
        }
    )
    kept = [row for row in rows if row.get("check_name") not in replace_names]
    kept.append(
        validation_row(
            "cost_breakdown_balance",
            cost["total_cost"],
            as_float(kpi.get("reported_total_cost_jpy")),
            source_files="cost_breakdown_detail.csv;kpi_summary.json",
        )
    )
    kept.extend(checks)
    write_csv(path, fields, kept)
    return kept


def update_validation_counts(run_dir: Path, rows: list[dict[str, Any]], kpi: dict[str, Any]) -> None:
    errors = sum(1 for row in rows if row.get("status") == "NG" and row.get("severity") == "ERROR")
    warnings = sum(1 for row in rows if row.get("status") == "NG" and row.get("severity") == "WARNING")
    status = "ERROR" if errors else ("WARNING" if warnings else "OK")

    kpi["data_flow_validation_status"] = status
    kpi["data_flow_error_count"] = errors
    kpi["data_flow_warning_count"] = warnings
    kpi["validation_status"] = status
    kpi.setdefault("validation", {})
    kpi["validation"].update(
        {"data_flow_validation_status": status, "error_count": errors, "warning_count": warnings}
    )
    write_json(run_dir / "graph" / "kpi_summary.json", kpi)
    root_path = run_dir / "kpi_summary.json"
    if root_path.is_file():
        root_kpi = load_json(root_path)
        root_kpi.update(
            {
                "data_flow_validation_status": status,
                "data_flow_error_count": errors,
                "data_flow_warning_count": warnings,
                "validation_status": status,
            }
        )
        write_json(root_path, root_kpi)


def status_for_checks(rows: list[dict[str, Any]], check_names: set[str]) -> str:
    selected = [row for row in rows if row.get("check_name") in check_names]
    if not selected:
        return "SKIPPED"
    evaluated = [row for row in selected if row.get("status") != "SKIPPED"]
    if not evaluated:
        return "SKIPPED"
    return "OK" if all(row.get("status") == "OK" for row in evaluated) else "NG"


def write_strict_reconciliation(run_dir: Path, rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    def domain_row(domain: str, status: str, message: str, severity: str | None = None) -> dict[str, Any]:
        return {
            "domain": domain,
            "check_name": f"{domain}_strict_reconciliation",
            "source_a": "canonical reporting ledgers",
            "source_a_value": "",
            "source_b": "final reporting artifacts",
            "source_b_value": "",
            "difference": "",
            "tolerance": TOL,
            "status": status,
            "severity": severity or ("ERROR" if status == "NG" else "INFO"),
            "message": message,
        }

    strict_rows = [
        domain_row(
            "energy",
            status_for_checks(
                rows,
                {
                    "pv_generation_balance",
                    "bus_charging_balance",
                    "grid_import_balance",
                    "bess_charge_balance",
                    "bess_discharge_balance",
                    "bess_soc_transition_balance",
                },
            ),
            "Existing energy-flow checks preserved after reporting finalization.",
        ),
        domain_row(
            "identity",
            status_for_checks(rows, {"operator_id_empty_count", "service_date_consistency"}),
            "Identity checks were not modified by the finalizer.",
        ),
        domain_row(
            "vehicle-charge-allocation",
            status_for_checks(
                rows,
                {
                    "vehicle_grid_to_bus_allocation_matches_site_total",
                    "vehicle_pv_to_bus_allocation_matches_site_total",
                    "vehicle_bess_to_bus_allocation_matches_site_total",
                },
            ),
            "Vehicle charging source allocation ledgers were reconciled against site totals.",
        ),
        domain_row(
            "fuel",
            status_for_checks(
                rows,
                {
                    "fuel_cost_matches_fuel_consumption",
                    "cost_breakdown_fuel_cost_matches_fuel_canonical",
                    "kpi_fuel_cost_matches_fuel_canonical",
                },
            ),
            "Fuel cost follows fuel_canonical_ledger.csv.",
        ),
        domain_row(
            "co2",
            status_for_checks(
                rows,
                {
                    "co2_total_equals_grid_plus_ice",
                    "cost_breakdown_ice_co2_matches_co2_timeseries",
                    "cost_breakdown_total_co2_matches_co2_timeseries",
                },
            ),
            "CO2 totals follow co2_timeseries.csv.",
        ),
        domain_row(
            "cost",
            status_for_checks(
                rows,
                {
                    "kpi_grid_purchase_cost_matches_cost_timeseries",
                    "kpi_demand_charge_matches_cost_breakdown",
                    "kpi_total_cost_matches_cost_breakdown",
                    "summary_total_cost_matches_cost_breakdown",
                },
            ),
            "Cost KPI and summary totals are separated from objective value.",
        ),
        domain_row(
            "bess-metadata",
            status_for_checks(
                rows,
                {
                    "energy_flow_bess_capacity_matches_bess_timeseries",
                    "energy_flow_bess_soc_min_matches_bess_timeseries",
                    "energy_flow_bess_soc_max_matches_bess_timeseries",
                },
            ),
            "BESS capacity and SOC bounds come from bess_timeseries.csv.",
        ),
        domain_row(
            "vehicle-soc-violation",
            "OUT_OF_SCOPE_REMAINS",
            "Remains because this run was not re-optimized.",
            "WARNING",
        ),
        domain_row(
            "solver-status",
            "NG"
            if summary.get("result_status") in {"INFEASIBLE", "INVALID"}
            else (
                "OUT_OF_SCOPE_REMAINS"
                if summary.get("solver_status") == "BASELINE_FALLBACK"
                else "OK"
            ),
            "Canonical result is not validated feasible; evaluation KPIs were gated."
            if summary.get("result_status") in {"INFEASIBLE", "INVALID"}
            else (
                "BASELINE_FALLBACK remains because this run was not re-optimized."
                if summary.get("solver_status") == "BASELINE_FALLBACK"
                else "Solver status does not indicate BASELINE_FALLBACK."
            ),
            "ERROR"
            if summary.get("result_status") in {"INFEASIBLE", "INVALID"}
            else (
                "WARNING"
                if summary.get("solver_status") == "BASELINE_FALLBACK"
                else "INFO"
            ),
        ),
    ]
    legacy_rows = [
        {
            "category": "energy",
            "status": strict_rows[0]["status"],
            "message": strict_rows[0]["message"],
        },
        {
            "category": "identity",
            "status": strict_rows[1]["status"],
            "message": strict_rows[1]["message"],
        },
        {
            "category": "vehicle-charge-allocation",
            "status": strict_rows[2]["status"],
            "message": strict_rows[2]["message"],
        },
        {
            "category": "fuel",
            "status": strict_rows[3]["status"],
            "message": strict_rows[3]["message"],
        },
        {
            "category": "co2",
            "status": strict_rows[4]["status"],
            "message": strict_rows[4]["message"],
        },
        {
            "category": "cost",
            "status": strict_rows[5]["status"],
            "message": strict_rows[5]["message"],
        },
        {
            "category": "bess-metadata",
            "status": strict_rows[6]["status"],
            "message": strict_rows[6]["message"],
        },
        {
            "category": "vehicle-soc-violation",
            "status": strict_rows[7]["status"],
            "message": strict_rows[7]["message"],
        },
        {
            "category": "solver-status",
            "status": strict_rows[8]["status"],
            "message": strict_rows[8]["message"],
        },
    ]
    fields = [
        "domain",
        "check_name",
        "source_a",
        "source_a_value",
        "source_b",
        "source_b_value",
        "difference",
        "tolerance",
        "status",
        "severity",
        "message",
    ]
    write_csv(run_dir / "strict_reconciliation.csv", fields, strict_rows)
    write_csv(run_dir / "strict_reconciliation_after_rebuild.csv", ["category", "status", "message"], legacy_rows)

    md_lines = ["| Domain | Status | Severity | Message |", "|---|---|---|---|"]
    for row in strict_rows:
        md_lines.append(f"| {row['domain']} | {row['status']} | {row['severity']} | {row['message']} |")
    md_text = "\n".join(md_lines) + "\n"
    (run_dir / "strict_reconciliation.md").write_text(md_text, encoding="utf-8")
    (run_dir / "strict_reconciliation_after_rebuild.md").write_text(md_text, encoding="utf-8")


def write_manifest(input_dir: Path, output_dir: Path) -> None:
    entries = []
    for rel in REPORTING_FILES + [
        "rebuild_reporting_log.json",
        "strict_reconciliation.csv",
        "strict_reconciliation.md",
        "strict_reconciliation_after_rebuild.csv",
        "strict_reconciliation_after_rebuild.md",
    ]:
        input_path = input_dir / rel
        output_path = output_dir / rel
        input_hash = sha256_file(input_path)
        output_hash = sha256_file(output_path)
        entries.append(
            {
                "path": rel,
                "input_sha256": input_hash,
                "output_sha256": output_hash,
                "changed_from_input": input_hash != output_hash,
                "created_in_output": input_hash is None and output_hash is not None,
            }
        )
    manifest = {
        "input_run_dir": str(input_dir),
        "output_run_dir": str(output_dir),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_run_modified": False,
        "manifest_path": "changed_files_manifest.json",
        "files": entries,
    }
    write_json(output_dir / "changed_files_manifest.json", manifest)


def _reporting_log_payload(
    run_dir: Path,
    totals: dict[str, float],
    cost: dict[str, float],
    bess_metadata: dict[str, Any],
    kpi: dict[str, Any],
    mode: str,
    input_dir: Path | None = None,
) -> dict[str, Any]:
    log = {
        "mode": mode,
        "run_dir": str(run_dir),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "solver_rerun": False,
        "simulation_rerun": False,
        "vehicle_assignment_regenerated": False,
        "no_reoptimization_performed": True,
        "updated_files": REPORTING_FILES,
        "source_of_truth": {
            "energy": "graph/energy_flow_ledger.csv",
            "bess_metadata": "graph/bess_timeseries.csv",
            "cost": cost["canonical_cost_ledger_source"],
            "grid_purchase_cost": cost["canonical_cost_ledger_source"],
            "demand_charge": cost["canonical_cost_ledger_source"],
            "fuel": "graph/fuel_canonical_ledger.csv",
            "co2": "graph/co2_timeseries.csv",
            "objective": "objective_breakdown.csv",
            "gross_operating_cost": cost["canonical_cost_ledger_source"],
        },
        "totals": totals,
        "cost": cost,
        "bess_metadata": bess_metadata,
        "validation_status": kpi.get("data_flow_validation_status"),
        "validation_error_count": kpi.get("data_flow_error_count"),
        "out_of_scope_remaining": ["vehicle SOC violation", "BASELINE_FALLBACK"],
    }
    if input_dir is not None:
        log["input_run_dir"] = str(input_dir)
        log["output_run_dir"] = str(run_dir)
        log["source_run_modified"] = False
    return log


def rebuild_reporting_artifacts_in_place(run_dir: Path) -> ReportingRebuildResult:
    run_dir = run_dir.resolve()
    bess_metadata = update_energy_flow_bess_metadata(run_dir)
    totals = compute_ledger_totals(run_dir)
    assignment = compute_assignment_summary(run_dir)
    cost = update_cost_breakdown(run_dir, totals, assignment)
    summary = update_summary(run_dir, cost, assignment)
    kpi = update_kpi_summary(run_dir, totals, cost, bess_metadata, assignment)
    update_experiment_report_metrics(run_dir, cost=cost, totals=totals)
    validation_rows = update_data_flow_validation(run_dir, totals, cost, summary, kpi, bess_metadata)
    summary, kpi, validation_rows = apply_solution_validity_gate(
        run_dir,
        summary,
        kpi,
        validation_rows,
    )
    update_validation_counts(run_dir, validation_rows, kpi)
    kpi = load_json(run_dir / "graph" / "kpi_summary.json")
    write_strict_reconciliation(run_dir, validation_rows, summary)
    log = _reporting_log_payload(run_dir, totals, cost, bess_metadata, kpi, "in_place_after_optimization")
    write_json(run_dir / "rebuild_reporting_log.json", log)
    warnings = [
        str(row.get("check_name"))
        for row in validation_rows
        if row.get("status") == "NG" and row.get("severity") == "WARNING"
    ]
    return ReportingRebuildResult(
        run_dir=run_dir,
        updated_files=list(REPORTING_FILES),
        ledger_totals=totals,
        cost_totals=cost,
        validation_status={
            "status": kpi.get("data_flow_validation_status"),
            "error_count": kpi.get("data_flow_error_count"),
            "warning_count": kpi.get("data_flow_warning_count"),
        },
        warnings=warnings,
    )


def rebuild_reporting_artifacts_to_output_dir(
    input_dir: Path,
    output_dir: Path,
    overwrite: bool = False,
) -> dict[str, Any]:
    input_dir = input_dir.resolve()
    output_dir = output_dir.resolve()
    copy_run(input_dir, output_dir, overwrite)
    result = rebuild_reporting_artifacts_in_place(output_dir)
    kpi = load_json(output_dir / "graph" / "kpi_summary.json")
    log = _reporting_log_payload(
        output_dir,
        result.ledger_totals,
        result.cost_totals,
        {"source": "graph/bess_timeseries.csv"},
        kpi,
        "copy_backfill_existing_run",
        input_dir=input_dir,
    )
    write_json(output_dir / "rebuild_reporting_log.json", log)
    write_manifest(input_dir, output_dir)
    return log
