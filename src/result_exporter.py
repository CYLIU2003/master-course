"""
result_exporter.py — CSV / JSON / Markdown / Excel 出力

仕様書 §13, §14.7 担当:
  - summary.json            (§13.1.1)
  - vehicle_schedule.csv    (§13.1.2)
  - charging_schedule.csv   (§13.1.3)
  - site_power_balance.csv  (§13.1.4)
  - experiment_report.md    (§13.1.5)
  - results.xlsx            (Excel multi-sheet export)
"""

from __future__ import annotations

import csv
import hashlib
import html
import json
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
import re
from typing import Any, Callable, Dict, List, Optional

from .data_schema import ProblemData
from .milp_model import MILPResult
from .model_sets import ModelSets
from .parameter_builder import DerivedParams, get_grid_price
from .route_code_utils import extract_route_series_from_candidates
from .simulator import SimulationResult


def _to_scalar_metric(value: Any) -> float:
    if isinstance(value, dict):
        numeric_values = []
        for raw in value.values():
            try:
                numeric_values.append(float(raw))
            except (TypeError, ValueError):
                continue
        if not numeric_values:
            return 0.0
        return sum(numeric_values) / len(numeric_values)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _make_run_dir(output_root: str | Path) -> Path:
    """output/run_yyyymmdd_hhmm/ ディレクトリを作成して返す"""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    run_dir = Path(output_root) / f"run_{ts}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def export_all(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp_result: MILPResult,
    sim_result: SimulationResult,
    output_root: str | Path = "output",
    run_label: Optional[str] = None,
) -> Path:
    """
    全出力ファイルを一括生成する。

    Returns
    -------
    Path
        出力ディレクトリ
    """
    run_dir = _make_run_dir(output_root)

    export_summary_json(run_dir, milp_result, sim_result, run_label)
    export_vehicle_schedule(run_dir, data, ms, dp, milp_result)
    export_charging_schedule(run_dir, ms, milp_result)
    export_site_power_balance(run_dir, ms, milp_result, sim_result, data)
    export_experiment_report(run_dir, data, ms, milp_result, sim_result, run_label)
    export_targeted_trips(run_dir, data, milp_result)
    export_trip_type_counts(run_dir, data)
    export_cost_breakdown_detail(run_dir, sim_result)
    export_co2_breakdown(run_dir, data, ms, dp, milp_result, sim_result)
    export_vehicle_timelines(run_dir, data, ms, dp, milp_result)
    export_objective_breakdown(run_dir, milp_result, sim_result)
    export_simulation_conditions(run_dir, data, dp)
    export_graph_exports_phase1(run_dir, data, ms, dp, milp_result, sim_result, run_label)
    try:
        export_excel(data, ms, dp, milp_result, sim_result, run_dir, run_label)
    except ImportError:
        pass  # openpyxl 未インストール時はスキップ

    return run_dir


def export_graph_exports_phase1(
    run_dir: Path,
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
    sim: SimulationResult,
    run_label: Optional[str] = None,
) -> None:
    """
    Graph export specification Phase 1.

    Required outputs:
    - manifest.json
    - vehicle_timeline.csv
    - soc_events.csv
    - depot_power_timeseries_5min.csv
    - trip_assignment.csv
    - cost_breakdown.json
    - kpi_summary.json
    """
    scenario_id = _extract_scenario_id(run_dir, run_label)
    base_date = _extract_base_date(run_dir)
    graph_dir = run_dir / "graph"
    graph_dir.mkdir(parents=True, exist_ok=True)

    graph_export_context = (
        dict(getattr(data, "graph_export_context", {}) or {})
        if isinstance(getattr(data, "graph_export_context", None), dict)
        else None
    )
    planning_start_time = _planning_start_time_text(graph_export_context)
    vehicle_timeline_rows = _build_vehicle_timeline_rows(
        data,
        ms,
        dp,
        milp,
        scenario_id,
        base_date,
        planning_start_time=planning_start_time,
    )
    soc_event_rows = _build_soc_event_rows(
        data,
        ms,
        dp,
        milp,
        scenario_id,
        base_date,
        planning_start_time=planning_start_time,
    )
    depot_power_rows = _build_depot_power_rows_5min(
        data,
        ms,
        dp,
        milp,
        scenario_id,
        base_date,
        planning_start_time=planning_start_time,
    )
    trip_assignment_rows = _build_trip_assignment_rows(
        data,
        dp,
        milp,
        scenario_id,
        base_date,
        planning_start_time=planning_start_time,
    )
    cost_breakdown = _build_cost_breakdown_json(data, sim, scenario_id)
    kpi_summary = _build_kpi_summary_json(data, ms, dp, milp, sim, scenario_id)
    refuel_event_rows = _build_refuel_event_rows(
        data,
        ms,
        dp,
        milp,
        scenario_id,
        base_date,
        planning_start_time=planning_start_time,
    )
    route_band_diagrams = _build_route_band_diagram_assets(
        vehicle_timeline_rows,
        scenario_id,
        graph_context=graph_export_context,
    )

    _write_csv(graph_dir / "vehicle_timeline.csv", vehicle_timeline_rows)
    _write_csv(graph_dir / "soc_events.csv", soc_event_rows)
    _write_csv(graph_dir / "depot_power_timeseries_5min.csv", depot_power_rows)
    _write_csv(graph_dir / "trip_assignment.csv", trip_assignment_rows)
    _write_csv(graph_dir / "refuel_events.csv", refuel_event_rows)
    with open(graph_dir / "cost_breakdown.json", "w", encoding="utf-8") as f:
        json.dump(cost_breakdown, f, ensure_ascii=False, indent=2)
    with open(graph_dir / "kpi_summary.json", "w", encoding="utf-8") as f:
        json.dump(kpi_summary, f, ensure_ascii=False, indent=2)
    _write_route_band_diagram_assets(graph_dir, route_band_diagrams)

    files = [
        "vehicle_timeline.csv",
        "soc_events.csv",
        "depot_power_timeseries_5min.csv",
        "trip_assignment.csv",
        "refuel_events.csv",
        "cost_breakdown.json",
        "kpi_summary.json",
    ]
    manifest = {
        "schema_version": "1.0.0",
        "scenario_id": scenario_id,
        "generated_at": _tokyo_now().isoformat(),
        "time_resolution_minutes": 5,
        "timezone": "Asia/Tokyo",
        "has_pv": bool(data.enable_pv),
        "has_optimization_result": True,
        "has_dispatch_result": True,
        "has_simulation_result": True,
        "files": files,
        "optional_exports": {
            "route_band_diagrams": {
                "enabled": bool(route_band_diagrams["entries"]),
                "grouping_key": "band_id",
                "diagram_format": "svg",
                "manifest_file": (
                    "route_band_diagrams/manifest.json"
                    if route_band_diagrams["entries"]
                    else ""
                ),
                "diagram_count": len(route_band_diagrams["entries"]),
            }
        },
    }
    with open(graph_dir / "manifest.json", "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)


def _tokyo_now() -> datetime:
    return datetime.now(timezone(timedelta(hours=9)))


def _extract_scenario_id(run_dir: Path, run_label: Optional[str]) -> str:
    parts = [part for part in run_dir.parts if part]
    lowered = [part.lower() for part in parts]
    if "optimization" in lowered:
        idx = lowered.index("optimization")
        if idx + 1 < len(parts):
            value = str(parts[idx + 1]).strip()
            if value:
                return value
    if run_label and str(run_label).strip():
        return str(run_label).strip()
    return "unknown_scenario"


def _extract_base_date(run_dir: Path) -> date:
    date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    for part in run_dir.parts:
        text = str(part).strip()
        if date_pattern.match(text):
            try:
                return date.fromisoformat(text)
            except ValueError:
                continue
    return _tokyo_now().date()


def _planning_start_time_text(graph_context: Optional[Dict[str, Any]]) -> str:
    if isinstance(graph_context, dict):
        value = str(graph_context.get("planning_start_time") or "").strip()
        if value:
            return value
    return "00:00"


def _planning_start_components(planning_start_time: str) -> tuple[int, int]:
    parts = str(planning_start_time or "00:00").split(":")
    try:
        start_hour = int(parts[0])
        start_minute = int(parts[1])
    except (IndexError, ValueError):
        return (0, 0)
    return (start_hour % 24, start_minute % 60)


def _slot_to_iso(
    base_date: date,
    slot_idx: int,
    delta_t_min: float,
    *,
    planning_start_time: str = "00:00",
) -> str:
    tz = timezone(timedelta(hours=9))
    start_hour, start_minute = _planning_start_components(planning_start_time)
    dt0 = datetime.combine(base_date, time(start_hour % 24, start_minute % 60), tz)
    dt = dt0 + timedelta(minutes=float(slot_idx) * float(delta_t_min))
    return dt.isoformat()


def _slot_to_hhmm(
    slot_idx: int,
    delta_t_min: float,
    *,
    planning_start_time: str = "00:00",
) -> str:
    start_hour, start_minute = _planning_start_components(planning_start_time)
    total = int(round(start_hour * 60 + start_minute + float(slot_idx) * float(delta_t_min)))
    total %= 24 * 60
    return f"{total // 60:02d}:{total % 60:02d}"


def _build_refuel_event_rows(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
    scenario_id: str,
    base_date: date,
    *,
    planning_start_time: str = "00:00",
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    vehicle_band_lookup = _vehicle_primary_route_band_lookup(ms, dp, milp)
    for vehicle_id in ms.K_ALL:
        vehicle = dp.vehicle_lut.get(vehicle_id)
        depot_id = str(getattr(vehicle, "home_depot", "") or "")
        vehicle_type = str(getattr(vehicle, "vehicle_type", "") or "")
        primary_band = vehicle_band_lookup.get(vehicle_id, {})
        route_band_id = str(primary_band.get("band_id") or "")
        route_band_label = str(primary_band.get("band_label") or route_band_id)
        route_family_code = str(primary_band.get("route_family_code") or "")
        series = list(milp.refuel_schedule_l.get(vehicle_id, []))
        for slot_idx, liters in enumerate(series):
            refuel_liters = float(liters or 0.0)
            if refuel_liters <= 1.0e-9:
                continue
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "vehicle_id": vehicle_id,
                    "vehicle_type": vehicle_type,
                    "depot_id": depot_id,
                    "route_band_id": route_band_id,
                    "route_band_label": route_band_label,
                    "route_family_code": route_family_code,
                    "slot_index": int(slot_idx),
                    "event_time": _slot_to_iso(
                        base_date,
                        int(slot_idx),
                        data.delta_t_min,
                        planning_start_time=planning_start_time,
                    ),
                    "time_hhmm": _slot_to_hhmm(
                        int(slot_idx),
                        data.delta_t_min,
                        planning_start_time=planning_start_time,
                    ),
                    "refuel_liters": round(refuel_liters, 4),
                }
            )
    rows.sort(key=lambda row: (str(row.get("event_time") or ""), str(row.get("vehicle_id") or "")))
    return rows


def _state_from_event_type(event_type: str) -> str:
    value = str(event_type or "").strip().lower()
    if value == "service":
        return "service"
    if value == "deadhead":
        return "deadhead"
    if value in {"charging", "charge"}:
        return "charge"
    if value == "refuel":
        return "refuel"
    return "idle"


def _route_band_key(route_family_code: Any, route_id: Any) -> str:
    route_series_code, _route_series_prefix, _route_series_number, _series_source = (
        extract_route_series_from_candidates(route_family_code, route_id)
    )
    return str(route_series_code or route_family_code or route_id or "").strip()


def _vehicle_primary_route_band_lookup(
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
) -> Dict[str, Dict[str, str]]:
    lookup: Dict[str, Dict[str, str]] = {}
    for vehicle_id in ms.K_ALL:
        assigned = sorted(
            milp.assignment.get(vehicle_id, []),
            key=lambda task_id: dp.task_lut.get(task_id).start_time_idx if dp.task_lut.get(task_id) else 0,
        )
        primary_route_id = ""
        primary_route_family_code = ""
        primary_band_id = ""
        for task_id in assigned:
            task = dp.task_lut.get(task_id)
            if task is None:
                continue
            primary_route_id = str(getattr(task, "route_id", "") or "").strip()
            primary_route_family_code = str(getattr(task, "route_family_code", "") or "").strip()
            primary_band_id = _route_band_key(primary_route_family_code, primary_route_id)
            if primary_band_id or primary_route_id:
                break
        lookup[vehicle_id] = {
            "band_id": primary_band_id,
            "band_label": primary_band_id,
            "route_id": primary_route_id,
            "route_family_code": primary_route_family_code,
        }
    return lookup


def _build_vehicle_timeline_rows(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
    scenario_id: str,
    base_date: date,
    *,
    planning_start_time: str = "00:00",
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    vehicle_band_lookup = _vehicle_primary_route_band_lookup(ms, dp, milp)

    for vehicle_id in ms.K_ALL:
        vehicle = dp.vehicle_lut.get(vehicle_id)
        vehicle_type = str(getattr(vehicle, "vehicle_type", "") or "")
        depot_id = str(getattr(vehicle, "home_depot", "") or "")
        primary_band = vehicle_band_lookup.get(vehicle_id, {})
        primary_band_id = str(primary_band.get("band_id", "") or "")
        primary_band_label = str(primary_band.get("band_label", "") or "")
        primary_route_family_code = str(primary_band.get("route_family_code", "") or "")
        assigned = sorted(
            milp.assignment.get(vehicle_id, []),
            key=lambda task_id: dp.task_lut.get(task_id).start_time_idx if dp.task_lut.get(task_id) else 0,
        )
        previous_task_id: Optional[str] = None

        for task_id in assigned:
            task = dp.task_lut.get(task_id)
            if task is None:
                continue

            if previous_task_id is not None:
                deadhead_slots = int(dp.deadhead_time_slot.get(previous_task_id, {}).get(task_id, 0) or 0)
                if deadhead_slots > 0:
                    previous_task = dp.task_lut.get(previous_task_id)
                    previous_band_id = (
                        _route_band_key(
                            getattr(previous_task, "route_family_code", None) if previous_task else None,
                            getattr(previous_task, "route_id", None) if previous_task else None,
                        )
                        if previous_task is not None
                        else ""
                    )
                    next_band_id = _route_band_key(task.route_family_code, task.route_id)
                    deadhead_event_band_id = (
                        next_band_id if next_band_id and next_band_id == previous_band_id else ""
                    )
                    dh_start_idx = max(int(task.start_time_idx) - deadhead_slots, 0)
                    dh_end_idx = int(task.start_time_idx)
                    dh_start = _slot_to_iso(
                        base_date,
                        dh_start_idx,
                        data.delta_t_min,
                        planning_start_time=planning_start_time,
                    )
                    dh_end = _slot_to_iso(
                        base_date,
                        dh_end_idx,
                        data.delta_t_min,
                        planning_start_time=planning_start_time,
                    )
                    rows.append(
                        {
                            "scenario_id": scenario_id,
                            "depot_id": depot_id,
                            "vehicle_id": vehicle_id,
                            "vehicle_type": vehicle_type,
                            "band_id": deadhead_event_band_id,
                            "band_label": deadhead_event_band_id,
                            "vehicle_primary_band_id": primary_band_id,
                            "vehicle_primary_band_label": primary_band_label,
                            "start_time": dh_start,
                            "end_time": dh_end,
                            "state": "deadhead",
                            "route_id": "",
                            "route_family_code": primary_route_family_code,
                            "route_series_code": deadhead_event_band_id,
                            "event_route_band_id": deadhead_event_band_id,
                            "trip_id": "",
                            "from_location_id": str(getattr(dp.task_lut.get(previous_task_id), "destination", "") or ""),
                            "to_location_id": str(task.origin or ""),
                            "from_location_type": "terminal",
                            "to_location_type": "terminal",
                            "direction": "",
                            "route_variant_type": "",
                            "energy_delta_kwh": "",
                            "distance_km": float(dp.deadhead_distance_km.get(previous_task_id, {}).get(task_id, 0.0) or 0.0),
                            "duration_min": max((dh_end_idx - dh_start_idx) * float(data.delta_t_min), 0.0),
                            "is_deadhead": True,
                            "is_charge": False,
                            "is_service": False,
                            "is_idle": False,
                            "is_depot_move": False,
                            "is_short_turn": False,
                            "charger_id": "",
                            "charge_power_kw": "",
                            "refuel_liters": "",
                        }
                    )

            start_idx = int(task.start_time_idx)
            end_idx = int(task.end_time_idx)
            start_time = _slot_to_iso(
                base_date,
                start_idx,
                data.delta_t_min,
                planning_start_time=planning_start_time,
            )
            end_time = _slot_to_iso(
                base_date,
                end_idx,
                data.delta_t_min,
                planning_start_time=planning_start_time,
            )
            variant = str(task.route_variant_type or "unknown")
            is_short_turn = variant == "short_turn"
            is_depot_move = variant in {"depot_move", "depot_in", "depot_out"}
            energy_delta = -float(task.energy_required_kwh_bev or 0.0)
            route_family_code = str(task.route_family_code or "")
            route_series_code = _route_band_key(route_family_code, task.route_id)

            rows.append(
                {
                    "scenario_id": scenario_id,
                    "depot_id": depot_id,
                    "vehicle_id": vehicle_id,
                    "vehicle_type": vehicle_type,
                    "band_id": route_series_code,
                    "band_label": route_series_code,
                    "vehicle_primary_band_id": primary_band_id,
                    "vehicle_primary_band_label": primary_band_label,
                    "start_time": start_time,
                    "end_time": end_time,
                    "state": "service",
                    "route_id": str(task.route_id or ""),
                    "route_family_code": route_family_code,
                    "route_series_code": route_series_code,
                    "event_route_band_id": route_series_code,
                    "trip_id": str(task.task_id),
                    "from_location_id": str(task.origin or ""),
                    "to_location_id": str(task.destination or ""),
                    "from_location_type": "terminal",
                    "to_location_type": "terminal",
                    "direction": _normalize_direction(task.direction),
                    "route_variant_type": variant,
                    "energy_delta_kwh": energy_delta,
                    "distance_km": float(task.distance_km or 0.0),
                    "duration_min": max((end_idx - start_idx) * float(data.delta_t_min), 0.0),
                    "is_deadhead": False,
                    "is_charge": False,
                    "is_service": True,
                    "is_idle": False,
                    "is_depot_move": is_depot_move,
                    "is_short_turn": is_short_turn,
                    "charger_id": "",
                    "charge_power_kw": "",
                    "refuel_liters": "",
                }
            )
            previous_task_id = task_id

        charge_by_slot: Dict[int, float] = defaultdict(float)
        charge_by_slot_charger: Dict[int, str] = {}
        for charger_id, power_series in milp.charge_power_kw.get(vehicle_id, {}).items():
            for idx, raw in enumerate(power_series):
                kw = float(raw or 0.0)
                if kw <= 0.0:
                    continue
                charge_by_slot[idx] += kw
                if idx not in charge_by_slot_charger:
                    charge_by_slot_charger[idx] = charger_id

        if charge_by_slot:
            active_slots = sorted(charge_by_slot.keys())
            seg_start = active_slots[0]
            seg_values: List[float] = [charge_by_slot[seg_start]]
            seg_charger = charge_by_slot_charger.get(seg_start, "")

            def _append_charge_segment(start_slot: int, end_slot_exclusive: int, values: List[float], charger: str) -> None:
                if end_slot_exclusive <= start_slot:
                    return
                start_time = _slot_to_iso(
                    base_date,
                    start_slot,
                    data.delta_t_min,
                    planning_start_time=planning_start_time,
                )
                end_time = _slot_to_iso(
                    base_date,
                    end_slot_exclusive,
                    data.delta_t_min,
                    planning_start_time=planning_start_time,
                )
                avg_power = sum(values) / len(values) if values else 0.0
                duration_min = (end_slot_exclusive - start_slot) * float(data.delta_t_min)
                rows.append(
                    {
                        "scenario_id": scenario_id,
                        "depot_id": depot_id,
                        "vehicle_id": vehicle_id,
                        "vehicle_type": vehicle_type,
                        "band_id": "",
                        "band_label": "",
                        "vehicle_primary_band_id": primary_band_id,
                        "vehicle_primary_band_label": primary_band_label,
                        "start_time": start_time,
                        "end_time": end_time,
                        "state": "charge",
                        "route_id": "",
                        "route_family_code": primary_route_family_code,
                        "route_series_code": "",
                        "event_route_band_id": "",
                        "trip_id": "",
                        "from_location_id": depot_id,
                        "to_location_id": depot_id,
                        "from_location_type": "charger",
                        "to_location_type": "charger",
                        "direction": "",
                        "route_variant_type": "",
                        "energy_delta_kwh": avg_power * (duration_min / 60.0),
                        "distance_km": 0.0,
                        "duration_min": duration_min,
                        "is_deadhead": False,
                        "is_charge": True,
                        "is_service": False,
                        "is_idle": False,
                        "is_depot_move": False,
                        "is_short_turn": False,
                        "charger_id": charger,
                        "charge_power_kw": avg_power,
                        "refuel_liters": "",
                    }
                )

            previous_slot = seg_start
            for slot in active_slots[1:]:
                if slot == previous_slot + 1:
                    seg_values.append(charge_by_slot[slot])
                    previous_slot = slot
                    continue
                _append_charge_segment(seg_start, previous_slot + 1, seg_values, seg_charger)
                seg_start = slot
                previous_slot = slot
                seg_values = [charge_by_slot[slot]]
                seg_charger = charge_by_slot_charger.get(slot, "")
            _append_charge_segment(seg_start, previous_slot + 1, seg_values, seg_charger)

        refuel_series = list(milp.refuel_schedule_l.get(vehicle_id, []))
        for slot_idx, liters in enumerate(refuel_series):
            refuel_liters = float(liters or 0.0)
            if refuel_liters <= 1.0e-9:
                continue
            start_time = _slot_to_iso(
                base_date,
                slot_idx,
                data.delta_t_min,
                planning_start_time=planning_start_time,
            )
            end_time = _slot_to_iso(
                base_date,
                slot_idx + 1,
                data.delta_t_min,
                planning_start_time=planning_start_time,
            )
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "depot_id": depot_id,
                    "vehicle_id": vehicle_id,
                    "vehicle_type": vehicle_type,
                    "band_id": "",
                    "band_label": "",
                    "vehicle_primary_band_id": primary_band_id,
                    "vehicle_primary_band_label": primary_band_label,
                    "start_time": start_time,
                    "end_time": end_time,
                    "state": "refuel",
                    "route_id": "",
                    "route_family_code": primary_route_family_code,
                    "route_series_code": "",
                    "event_route_band_id": "",
                    "trip_id": "",
                    "from_location_id": depot_id,
                    "to_location_id": depot_id,
                    "from_location_type": "depot",
                    "to_location_type": "depot",
                    "direction": "",
                    "route_variant_type": "depot_refuel",
                    "energy_delta_kwh": "",
                    "distance_km": 0.0,
                    "duration_min": float(data.delta_t_min),
                    "is_deadhead": False,
                    "is_charge": False,
                    "is_service": False,
                    "is_idle": False,
                    "is_depot_move": True,
                    "is_short_turn": False,
                    "charger_id": "",
                    "charge_power_kw": "",
                    "refuel_liters": round(refuel_liters, 4),
                }
            )

    rows.sort(key=lambda row: (str(row.get("vehicle_id", "")), str(row.get("start_time", "")), str(row.get("state", ""))))
    return rows


def _build_soc_event_rows(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
    scenario_id: str,
    base_date: date,
    *,
    planning_start_time: str = "00:00",
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for vehicle_id in ms.K_BEV:
        series = list(milp.soc_series.get(vehicle_id, []))
        if len(series) < 2:
            continue
        vehicle = dp.vehicle_lut.get(vehicle_id)
        battery_kwh = float(getattr(vehicle, "battery_capacity", 0.0) or 0.0)
        min_soc = float(getattr(vehicle, "soc_min", 0.0) or 0.0)
        max_soc = float(getattr(vehicle, "soc_max", battery_kwh) or battery_kwh)
        location_id = str(getattr(vehicle, "home_depot", "") or "")
        for t_idx in range(1, len(series)):
            soc_before = float(series[t_idx - 1] or 0.0)
            soc_after = float(series[t_idx] or 0.0)
            delta = soc_after - soc_before
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "vehicle_id": vehicle_id,
                    "event_time": _slot_to_iso(
                        base_date,
                        t_idx,
                        data.delta_t_min,
                        planning_start_time=planning_start_time,
                    ),
                    "event_type": "simulation_tick",
                    "trip_id": "",
                    "route_id": "",
                    "location_id": location_id,
                    "soc_kwh_before": soc_before,
                    "soc_kwh_after": soc_after,
                    "soc_pct_before": (soc_before / battery_kwh * 100.0) if battery_kwh > 0 else 0.0,
                    "soc_pct_after": (soc_after / battery_kwh * 100.0) if battery_kwh > 0 else 0.0,
                    "delta_kwh": delta,
                    "battery_capacity_kwh": battery_kwh,
                    "energy_consumed_kwh": max(-delta, 0.0),
                    "energy_charged_kwh": max(delta, 0.0),
                    "reserve_margin_kwh": soc_after - min_soc,
                    "min_soc_constraint_kwh": min_soc,
                    "max_soc_constraint_kwh": max_soc,
                }
            )
    rows.sort(key=lambda row: (str(row.get("vehicle_id", "")), str(row.get("event_time", ""))))
    return rows


def _build_depot_power_rows_5min(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
    scenario_id: str,
    base_date: date,
    *,
    planning_start_time: str = "00:00",
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if data.num_periods <= 0:
        return rows

    horizon_min = int(round(float(data.num_periods) * float(data.delta_t_min)))
    five_min_points = list(range(0, max(horizon_min, 1), 5))

    charge_kw_by_site_slot: Dict[str, Dict[int, float]] = defaultdict(lambda: defaultdict(float))
    for vehicle_id, by_charger in milp.charge_power_kw.items():
        for charger_id, pwr_series in by_charger.items():
            charger = dp.charger_lut.get(charger_id)
            site_id = str(getattr(charger, "site_id", "") or "")
            if not site_id:
                continue
            for slot_idx, raw_kw in enumerate(pwr_series):
                charge_kw_by_site_slot[site_id][slot_idx] += float(raw_kw or 0.0)

    site_ids = sorted(set(ms.I_CHARGE) | set(dp.site_lut.keys()) | set(charge_kw_by_site_slot.keys()))
    for site_id in site_ids:
        grid_series = list(milp.grid_import_kw.get(site_id, []))
        pv_used_series = list(milp.pv_used_kw.get(site_id, []))
        charge_slot_map = charge_kw_by_site_slot.get(site_id, {})

        def _slot_value(series: List[float], slot_idx: int) -> float:
            if slot_idx < 0:
                return 0.0
            if slot_idx < len(series):
                return float(series[slot_idx] or 0.0)
            return 0.0

        peak_grid = 0.0
        for slot_idx in range(data.num_periods):
            peak_grid = max(peak_grid, _slot_value(grid_series, slot_idx))

        for minute in five_min_points:
            slot_idx = min(int(minute // max(float(data.delta_t_min), 1.0)), max(data.num_periods - 1, 0))
            grid_import_kw = _slot_value(grid_series, slot_idx)
            pv_used_kw = _slot_value(pv_used_series, slot_idx)
            pv_generation_kw = float(dp.pv_gen_kw.get(site_id, {}).get(slot_idx, 0.0) or 0.0)
            total_charge_kw = float(charge_slot_map.get(slot_idx, 0.0) or 0.0)
            building_load_kw = float(dp.base_load_kw.get(site_id, {}).get(slot_idx, 0.0) or 0.0)
            pv_curtailed_kw = max(pv_generation_kw - pv_used_kw, 0.0)
            net_load_kw = max(grid_import_kw + pv_used_kw, 0.0)
            rows.append(
                {
                    "scenario_id": scenario_id,
                    "timestamp": _slot_to_iso(
                        base_date,
                        int(minute / max(float(data.delta_t_min), 1e-6)),
                        data.delta_t_min,
                        planning_start_time=planning_start_time,
                    )
                    if float(data.delta_t_min) == 5.0
                    else (
                        lambda start_hour, start_minute: (
                            datetime.combine(
                                base_date,
                                time(start_hour, start_minute),
                                timezone(timedelta(hours=9)),
                            )
                            + timedelta(minutes=minute)
                        ).isoformat()
                    )(*_planning_start_components(planning_start_time)),
                    "depot_id": site_id,
                    "total_charge_kw": total_charge_kw,
                    "grid_import_kw": grid_import_kw,
                    "pv_generation_kw": pv_generation_kw,
                    "pv_used_for_charging_kw": pv_used_kw,
                    "pv_used_for_building_kw": 0.0,
                    "pv_curtailed_kw": pv_curtailed_kw,
                    "building_load_kw": building_load_kw,
                    "battery_storage_charge_kw": 0.0,
                    "battery_storage_discharge_kw": 0.0,
                    "net_load_kw": net_load_kw,
                    "demand_peak_candidate": abs(grid_import_kw - peak_grid) <= 1e-9,
                    "energy_price_yen_per_kwh": float(get_grid_price(dp, site_id, slot_idx, default=0.0) or 0.0),
                    "demand_charge_window_flag": bool(data.enable_demand_charge),
                }
            )
    rows.sort(key=lambda row: (str(row.get("depot_id", "")), str(row.get("timestamp", ""))))
    return rows


def _build_trip_assignment_rows(
    data: ProblemData,
    dp: DerivedParams,
    milp: MILPResult,
    scenario_id: str,
    base_date: date,
    *,
    planning_start_time: str = "00:00",
) -> List[Dict[str, Any]]:
    assigned_vehicle_by_task: Dict[str, str] = {}
    for vehicle_id, tasks in milp.assignment.items():
        for task_id in tasks:
            assigned_vehicle_by_task[str(task_id)] = str(vehicle_id)
    vehicle_band_lookup = _vehicle_primary_route_band_lookup(
        ModelSets(K_ALL=list(milp.assignment.keys())),
        dp,
        milp,
    )

    rows: List[Dict[str, Any]] = []
    for task in data.tasks:
        trip_id = str(task.task_id)
        vehicle_id = assigned_vehicle_by_task.get(trip_id, "")
        served = bool(vehicle_id)
        vehicle = dp.vehicle_lut.get(vehicle_id) if vehicle_id else None
        route_family_code = str(task.route_family_code or "")
        route_series_code = _route_band_key(route_family_code, task.route_id)
        vehicle_band = vehicle_band_lookup.get(vehicle_id, {}) if vehicle_id else {}
        start_iso = _slot_to_iso(
            base_date,
            int(task.start_time_idx),
            data.delta_t_min,
            planning_start_time=planning_start_time,
        )
        end_iso = _slot_to_iso(
            base_date,
            int(task.end_time_idx),
            data.delta_t_min,
            planning_start_time=planning_start_time,
        )
        rows.append(
            {
                "scenario_id": scenario_id,
                "trip_id": trip_id,
                "route_id": str(task.route_id or ""),
                "route_family_code": route_family_code,
                "route_series_code": route_series_code,
                "band_id": route_series_code,
                "direction": _normalize_direction(task.direction),
                "route_variant_type": str(task.route_variant_type or "unknown"),
                "scheduled_departure": start_iso,
                "scheduled_arrival": end_iso,
                "actual_departure": start_iso,
                "actual_arrival": end_iso,
                "assigned_vehicle_id": vehicle_id,
                "assigned_vehicle_type": str(getattr(vehicle, "vehicle_type", "") or "") if vehicle else "",
                "assigned_depot_id": str(getattr(vehicle, "home_depot", "") or "") if vehicle else "",
                "assigned_vehicle_band_id": str(vehicle_band.get("band_id", "") or ""),
                "served_flag": served,
                "unserved_reason": "" if served else "unassigned",
                "energy_used_kwh": float(task.energy_required_kwh_bev or 0.0),
                "distance_km": float(task.distance_km or 0.0),
                "delay_departure_min": 0.0,
                "delay_arrival_min": 0.0,
                "deadhead_before_km": 0.0,
                "deadhead_after_km": 0.0,
                "swap_type": "none",
            }
        )
    rows.sort(key=lambda row: str(row.get("trip_id", "")))
    return rows


def _safe_export_name(value: str, *, fallback: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r'[\\/:*?"<>|\s]+', "_", text)
    text = text.strip("._")
    return text or fallback


def _parse_iso_minute(value: Any) -> Optional[int]:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        dt = datetime.fromisoformat(text)
    except ValueError:
        parts = text.split(":")
        if len(parts) < 2:
            return None
        try:
            hour = int(parts[0])
            minute = int(parts[1])
        except ValueError:
            return None
        return hour * 60 + minute
    return int(dt.hour) * 60 + int(dt.minute)


def _clean_location_sequence(sequence: List[str]) -> List[str]:
    cleaned: List[str] = []
    for label in sequence:
        normalized = str(label or "").strip()
        if not normalized:
            continue
        if cleaned and cleaned[-1] == normalized:
            continue
        cleaned.append(normalized)
    return cleaned


def _sequence_alignment_score(sequence: List[str], axis: List[str]) -> int:
    axis_index = {label: idx for idx, label in enumerate(axis)}
    positions = [axis_index[label] for label in sequence if label in axis_index]
    if not positions:
        return -1
    if len(positions) == 1:
        return 100 + len(sequence)
    monotonic_pairs = sum(
        1
        for previous, current in zip(positions, positions[1:])
        if previous <= current
    )
    span = max(positions) - min(positions) + 1
    return monotonic_pairs * 1000 + len(positions) * 10 - span


def _orient_sequence_to_axis(sequence: List[str], axis: List[str]) -> List[str]:
    if not axis:
        return list(sequence)
    forward = list(sequence)
    backward = list(reversed(sequence))
    if _sequence_alignment_score(forward, axis) >= _sequence_alignment_score(backward, axis):
        return forward
    return backward


def _insert_location_segment(
    axis: List[str],
    segment: List[str],
    *,
    before_label: str = "",
    after_label: str = "",
) -> List[str]:
    insertable = [label for label in segment if label and label not in axis]
    if not insertable:
        return axis
    if before_label in axis and after_label in axis:
        before_idx = axis.index(before_label)
        after_idx = axis.index(after_label)
        insert_at = before_idx if after_idx < before_idx else after_idx + 1
        axis[insert_at:insert_at] = insertable
        return axis
    if before_label in axis:
        before_idx = axis.index(before_label)
        axis[before_idx:before_idx] = insertable
        return axis
    if after_label in axis:
        after_idx = axis.index(after_label)
        axis[after_idx + 1 : after_idx + 1] = insertable
        return axis
    axis.extend(insertable)
    return axis


def _merge_location_sequence_into_axis(axis: List[str], sequence: List[str]) -> List[str]:
    if not sequence:
        return axis
    merged = list(axis)
    aligned = _orient_sequence_to_axis(sequence, merged)
    axis_index = {label: idx for idx, label in enumerate(merged)}
    known_indices = [idx for idx, label in enumerate(aligned) if label in axis_index]
    if not known_indices:
        for label in aligned:
            if label not in merged:
                merged.append(label)
        return merged

    first_known_idx = known_indices[0]
    merged = _insert_location_segment(
        merged,
        aligned[:first_known_idx],
        before_label=aligned[first_known_idx],
    )
    for left_idx, right_idx in zip(known_indices, known_indices[1:]):
        merged = _insert_location_segment(
            merged,
            aligned[left_idx + 1 : right_idx],
            before_label=aligned[right_idx],
            after_label=aligned[left_idx],
        )
    merged = _insert_location_segment(
        merged,
        aligned[known_indices[-1] + 1 :],
        after_label=aligned[known_indices[-1]],
    )
    return merged


def _merge_location_sequences(sequences: Optional[List[List[str]]]) -> List[str]:
    cleaned_sequences = [
        _clean_location_sequence(sequence)
        for sequence in sequences or []
    ]
    cleaned_sequences = [sequence for sequence in cleaned_sequences if sequence]
    if not cleaned_sequences:
        return []

    def _seed_key(sequence: List[str]) -> tuple[int, int, tuple[str, ...]]:
        labels = set(sequence)
        overlap = sum(
            len(labels.intersection(other))
            for other in cleaned_sequences
        )
        return (len(sequence), overlap, tuple(sequence))

    axis = list(max(cleaned_sequences, key=_seed_key))
    for _ in range(max(2, len(cleaned_sequences) + 1)):
        changed = False
        for sequence in cleaned_sequences:
            merged = _merge_location_sequence_into_axis(axis, sequence)
            if merged != axis:
                axis = merged
                changed = True
        if not changed:
            break
    return axis


_SIDE_LOCATION_KEYWORDS = (
    "営業所",
    "車庫",
    "操車所",
    "折返",
    "折返所",
    "待機場",
    "事業所",
    "基地",
)


def _is_side_location_label(label: str) -> bool:
    normalized = str(label or "").strip()
    if not normalized:
        return False
    return any(keyword in normalized for keyword in _SIDE_LOCATION_KEYWORDS)


def _diagram_location_labels(
    rows: List[Dict[str, Any]],
    main_axis: List[str],
) -> List[str]:
    ordered_main_axis = [label for label in main_axis if str(label or "").strip()]
    if not ordered_main_axis:
        return []

    top_labels: List[str] = []
    bottom_labels: List[str] = []
    top_seen: set[str] = set()
    bottom_seen: set[str] = set()
    assigned_seen: set[str] = set()
    main_axis_set = set(ordered_main_axis)
    origin_counts: Dict[str, int] = defaultdict(int)
    destination_counts: Dict[str, int] = defaultdict(int)

    for row in rows:
        from_label = str(row.get("from_location_id") or "").strip()
        to_label = str(row.get("to_location_id") or "").strip()
        if from_label and from_label not in main_axis_set:
            origin_counts[from_label] += 1
        if to_label and to_label not in main_axis_set:
            destination_counts[to_label] += 1

    for row in rows:
        for label, is_origin in (
            (str(row.get("from_location_id") or "").strip(), True),
            (str(row.get("to_location_id") or "").strip(), False),
        ):
            if not label or label in main_axis_set or label in assigned_seen:
                continue
            prefer_top = origin_counts.get(label, 0) >= destination_counts.get(label, 0)
            if not _is_side_location_label(label):
                prefer_top = is_origin
            if prefer_top:
                if label not in top_seen:
                    top_labels.append(label)
                    top_seen.add(label)
                    assigned_seen.add(label)
                continue
            if label not in bottom_seen:
                bottom_labels.append(label)
                bottom_seen.add(label)
                assigned_seen.add(label)

    return [*top_labels, *ordered_main_axis, *bottom_labels]


def _ordered_location_labels_from_rows(rows: List[Dict[str, Any]]) -> List[str]:
    all_labels: set[str] = set()
    adjacency: Dict[str, Dict[str, int]] = defaultdict(dict)

    def _add_edge(from_label: str, to_label: str) -> None:
        if not from_label or not to_label or from_label == to_label:
            return
        adjacency[from_label][to_label] = adjacency[from_label].get(to_label, 0) + 1
        adjacency[to_label][from_label] = adjacency[to_label].get(from_label, 0) + 1

    for row in rows:
        from_label = str(row.get("from_location_id") or "").strip()
        to_label = str(row.get("to_location_id") or "").strip()
        if from_label:
            all_labels.add(from_label)
        if to_label:
            all_labels.add(to_label)
        if (
            str(row.get("state") or "") == "service"
            and from_label
            and to_label
        ):
            _add_edge(from_label, to_label)
    if not adjacency:
        return sorted(all_labels)

    def _degree(label: str) -> int:
        return len(adjacency.get(label, {}))

    starts = sorted(
        all_labels,
        key=lambda label: (
            0 if _degree(label) <= 1 else 1,
            -sum(adjacency.get(label, {}).values()),
            label,
        ),
    )
    ordered: List[str] = []
    visited: set[str] = set()
    for start in starts:
        if start in visited:
            continue
        current = start
        previous = ""
        while current and current not in visited:
            ordered.append(current)
            visited.add(current)
            candidates = [
                (neighbor, count)
                for neighbor, count in adjacency.get(current, {}).items()
                if neighbor not in visited and neighbor != previous
            ]
            if not candidates:
                break
            candidates.sort(key=lambda item: (-item[1], item[0]))
            previous, current = current, candidates[0][0]
    ordered.extend(sorted(label for label in all_labels if label not in visited))
    return ordered


def _ordered_location_labels(
    rows: List[Dict[str, Any]],
    *,
    sequences: Optional[List[List[str]]] = None,
) -> List[str]:
    sequence_axis = _merge_location_sequences(sequences)
    row_axis = _ordered_location_labels_from_rows(rows)
    if not sequence_axis:
        return row_axis
    merged_axis = list(sequence_axis)
    for label in row_axis:
        if label not in merged_axis:
            merged_axis.append(label)
    return merged_axis


def _vehicle_line_color(vehicle_id: str, vehicle_type: str) -> str:
    digest = hashlib.sha1(str(vehicle_id).encode("utf-8")).hexdigest()
    seed = int(digest[:8], 16)
    vehicle_type_upper = str(vehicle_type or "").upper()
    if vehicle_type_upper == "BEV":
        hue_min, hue_max = 145.0, 190.0
    elif vehicle_type_upper == "ICE":
        hue_min, hue_max = 12.0, 42.0
    else:
        hue_min, hue_max = 215.0, 265.0
    hue = hue_min + (seed % 1000) / 1000.0 * (hue_max - hue_min)
    sat = 62.0 + float((seed // 17) % 12)
    light = 38.0 + float((seed // 29) % 18)
    return f"hsl({hue:.1f} {sat:.1f}% {light:.1f}%)"


def _vehicle_type_legend_color(vehicle_type: str) -> str:
    vehicle_type_upper = str(vehicle_type or "").upper()
    if vehicle_type_upper == "BEV":
        return "hsl(165 70% 38%)"
    if vehicle_type_upper == "ICE":
        return "hsl(24 78% 46%)"
    return "hsl(240 18% 45%)"


def _graph_context_band_sequences(
    graph_context: Optional[Dict[str, Any]],
    band_id: str,
) -> List[List[str]]:
    if not isinstance(graph_context, dict):
        return []
    raw_sequences = (graph_context.get("band_stop_sequences") or {}).get(band_id) or []
    sequences: List[List[str]] = []
    for sequence in raw_sequences:
        if not isinstance(sequence, list):
            continue
        cleaned = [str(item or "").strip() for item in sequence if str(item or "").strip()]
        if cleaned:
            sequences.append(cleaned)
    return sequences


def _graph_context_task_stop_points(
    graph_context: Optional[Dict[str, Any]],
    task_id: str,
) -> List[Dict[str, Any]]:
    if not isinstance(graph_context, dict):
        return []
    raw_points = (graph_context.get("task_stop_sequences") or {}).get(task_id) or []
    return [dict(item) for item in raw_points if isinstance(item, dict)]


def _parse_iso_datetime(value: Any) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _graph_context_depot_label(
    graph_context: Optional[Dict[str, Any]],
    depot_id: str,
) -> str:
    normalized_depot_id = str(depot_id or "").strip()
    if not normalized_depot_id:
        return ""
    if isinstance(graph_context, dict):
        labels = graph_context.get("depot_labels_by_id") or {}
        if isinstance(labels, dict):
            label = str(labels.get(normalized_depot_id) or "").strip()
            if label:
                return label
    return normalized_depot_id


def _band_row_key(row: Dict[str, Any]) -> tuple[Any, ...]:
    return (
        str(row.get("vehicle_id") or ""),
        str(row.get("band_id") or ""),
        str(row.get("state") or ""),
        str(row.get("start_time") or ""),
        str(row.get("end_time") or ""),
        str(row.get("from_location_id") or ""),
        str(row.get("to_location_id") or ""),
        str(row.get("trip_id") or ""),
        str(row.get("charger_id") or ""),
    )


def _normalized_band_row(
    row: Dict[str, Any],
    *,
    band_id: str,
    band_label: str,
    graph_context: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    normalized = dict(row)
    depot_id = str(normalized.get("depot_id") or "").strip()
    depot_label = _graph_context_depot_label(graph_context, depot_id)
    if depot_id and depot_label:
        for key in ("from_location_id", "to_location_id"):
            value = str(normalized.get(key) or "").strip()
            if value == depot_id:
                normalized[key] = depot_label
    state = str(normalized.get("state") or "").strip()
    if state in {"charge", "idle", "refuel"} and depot_label:
        normalized["from_location_id"] = depot_label
        normalized["to_location_id"] = depot_label
    if not str(normalized.get("band_id") or "").strip():
        normalized["band_id"] = band_id
    if not str(normalized.get("band_label") or "").strip():
        normalized["band_label"] = band_label or band_id
    if not str(normalized.get("route_series_code") or "").strip():
        normalized["route_series_code"] = band_id
    if not str(normalized.get("event_route_band_id") or "").strip() and state in {
        "deadhead",
        "charge",
        "refuel",
        "idle",
    }:
        normalized["event_route_band_id"] = band_id
    if state == "charge":
        normalized["is_deadhead"] = False
        normalized["is_charge"] = True
        normalized["is_service"] = False
        normalized["is_idle"] = False
        normalized["is_depot_move"] = True
    elif state == "idle":
        normalized["is_deadhead"] = False
        normalized["is_charge"] = False
        normalized["is_service"] = False
        normalized["is_idle"] = True
        normalized["is_depot_move"] = True
    elif state == "refuel":
        normalized["is_deadhead"] = False
        normalized["is_charge"] = False
        normalized["is_service"] = False
        normalized["is_idle"] = False
        normalized["is_depot_move"] = True
    return normalized


def _inferred_depot_move_minutes(total_gap_min: float) -> int:
    gap = max(int(round(float(total_gap_min))), 0)
    if gap <= 0:
        return 0
    return max(10, min(25, max(gap // 4, 10)))


def _make_inferred_band_row(
    *,
    template_row: Dict[str, Any],
    band_id: str,
    band_label: str,
    state: str,
    start_dt: datetime,
    end_dt: datetime,
    from_label: str,
    to_label: str,
    route_variant_type: str = "",
    charge_power_kw: Any = "",
) -> Optional[Dict[str, Any]]:
    if end_dt <= start_dt:
        return None
    duration_min = max((end_dt - start_dt).total_seconds() / 60.0, 0.0)
    depot_label = str(template_row.get("depot_label") or "").strip()
    from_type = "depot" if from_label == depot_label and depot_label else "terminal"
    to_type = "depot" if to_label == depot_label and depot_label else "terminal"
    row = dict(template_row)
    row.update(
        {
            "band_id": band_id,
            "band_label": band_label or band_id,
            "start_time": start_dt.isoformat(),
            "end_time": end_dt.isoformat(),
            "state": state,
            "route_id": "",
            "route_series_code": band_id,
            "event_route_band_id": band_id,
            "trip_id": "",
            "from_location_id": from_label,
            "to_location_id": to_label,
            "from_location_type": from_type,
            "to_location_type": to_type,
            "direction": "",
            "route_variant_type": route_variant_type,
            "energy_delta_kwh": 0.0,
            "distance_km": 0.0,
            "duration_min": duration_min,
            "is_deadhead": state == "deadhead",
            "is_charge": state == "charge",
            "is_service": False,
            "is_idle": state == "idle",
            "is_depot_move": True,
            "is_short_turn": False,
            "charger_id": row.get("charger_id") if state == "charge" else "",
            "charge_power_kw": charge_power_kw if state == "charge" else "",
        }
    )
    return row


def _append_gap_as_depot_presence(
    *,
    append_row: Callable[[Dict[str, Any]], None],
    template_row: Dict[str, Any],
    band_id: str,
    band_label: str,
    start_dt: Optional[datetime],
    end_dt: Optional[datetime],
    origin_label: str,
    destination_label: str,
    depot_label: str,
) -> None:
    if start_dt is None or end_dt is None or end_dt <= start_dt:
        return
    total_gap_min = max((end_dt - start_dt).total_seconds() / 60.0, 0.0)
    if total_gap_min <= 0.0:
        return
    move_min = min(
        _inferred_depot_move_minutes(total_gap_min),
        max(int(total_gap_min // 2), 0),
    )
    arrival_end = start_dt
    departure_start = end_dt
    if origin_label != depot_label:
        arrival_end = min(end_dt, start_dt + timedelta(minutes=move_min))
        inferred = _make_inferred_band_row(
            template_row=template_row,
            band_id=band_id,
            band_label=band_label,
            state="deadhead",
            start_dt=start_dt,
            end_dt=arrival_end,
            from_label=origin_label,
            to_label=depot_label,
            route_variant_type="depot_in",
        )
        if inferred:
            append_row(inferred)
    if destination_label != depot_label:
        departure_start = max(start_dt, end_dt - timedelta(minutes=move_min))
    idle_start = arrival_end if origin_label != depot_label else start_dt
    idle_end = departure_start if destination_label != depot_label else end_dt
    inferred_idle = _make_inferred_band_row(
        template_row=template_row,
        band_id=band_id,
        band_label=band_label,
        state="idle",
        start_dt=idle_start,
        end_dt=idle_end,
        from_label=depot_label,
        to_label=depot_label,
        route_variant_type="depot_stay",
    )
    if inferred_idle:
        append_row(inferred_idle)
    if destination_label != depot_label:
        inferred = _make_inferred_band_row(
            template_row=template_row,
            band_id=band_id,
            band_label=band_label,
            state="deadhead",
            start_dt=departure_start,
            end_dt=end_dt,
            from_label=depot_label,
            to_label=destination_label,
            route_variant_type="depot_out",
        )
        if inferred:
            append_row(inferred)


def _row_is_charge(row: Dict[str, Any]) -> bool:
    return str(row.get("state") or "").strip() in {"charge", "refuel"}


def _row_is_explicit_band_deadhead(row: Dict[str, Any], band_id: str) -> bool:
    return (
        str(row.get("state") or "").strip() == "deadhead"
        and str(row.get("band_id") or "").strip() == band_id
    )


def _vehicle_band_rows(
    *,
    vehicle_rows: List[Dict[str, Any]],
    band_id: str,
    band_label: str,
    graph_context: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    normalized_full_rows = [
        _normalized_band_row(
            row,
            band_id=band_id,
            band_label=band_label,
            graph_context=graph_context,
        )
        for row in vehicle_rows
    ]
    normalized_full_rows.sort(
        key=lambda row: (str(row.get("start_time") or ""), str(row.get("state") or ""))
    )
    relevant_indices = [
        idx
        for idx, row in enumerate(normalized_full_rows)
        if str(row.get("state") or "").strip() == "service"
        and str(row.get("band_id") or "").strip() == band_id
    ]
    if not relevant_indices:
        return []

    sample_row = dict(normalized_full_rows[relevant_indices[0]])
    depot_label = _graph_context_depot_label(
        graph_context,
        str(sample_row.get("depot_id") or ""),
    )
    sample_row["depot_label"] = depot_label

    band_rows: List[Dict[str, Any]] = []
    seen_keys: set[tuple[Any, ...]] = set()

    def _append(row: Dict[str, Any]) -> None:
        normalized = _normalized_band_row(
            row,
            band_id=band_id,
            band_label=band_label,
            graph_context=graph_context,
        )
        key = _band_row_key(normalized)
        if key in seen_keys:
            return
        seen_keys.add(key)
        band_rows.append(normalized)

    for row in normalized_full_rows:
        if (
            str(row.get("band_id") or "").strip() == band_id
            and str(row.get("state") or "").strip() in {"service", "deadhead"}
        ):
            _append(row)

    day_datetimes = [
        dt
        for row in normalized_full_rows
        for dt in (
            _parse_iso_datetime(row.get("start_time")),
            _parse_iso_datetime(row.get("end_time")),
        )
        if dt is not None
    ]
    if not day_datetimes:
        return sorted(
            band_rows,
            key=lambda row: (str(row.get("vehicle_id") or ""), str(row.get("start_time") or "")),
        )
    first_dt = min(day_datetimes)
    day_start = datetime.combine(first_dt.date(), time(0, 0), first_dt.tzinfo)
    day_end = datetime.combine(first_dt.date(), time(23, 59), first_dt.tzinfo)

    first_relevant_idx = relevant_indices[0]
    first_relevant_row = normalized_full_rows[first_relevant_idx]
    first_relevant_start = _parse_iso_datetime(first_relevant_row.get("start_time"))
    first_relevant_origin = str(first_relevant_row.get("from_location_id") or depot_label).strip() or depot_label
    earlier_rows = normalized_full_rows[:first_relevant_idx]
    earlier_non_charge_rows = [row for row in earlier_rows if not _row_is_charge(row)]
    if not earlier_non_charge_rows:
        earlier_charge_rows = [row for row in earlier_rows if _row_is_charge(row)]
        previous_charge_end = day_start
        if earlier_charge_rows:
            first_charge_start = _parse_iso_datetime(earlier_charge_rows[0].get("start_time"))
            if first_charge_start is not None and first_charge_start > day_start:
                _append_gap_as_depot_presence(
                    append_row=_append,
                    template_row=sample_row,
                    band_id=band_id,
                    band_label=band_label,
                    start_dt=day_start,
                    end_dt=first_charge_start,
                    origin_label=depot_label,
                    destination_label=depot_label,
                    depot_label=depot_label,
                )
            for charge_row in earlier_charge_rows:
                _append(charge_row)
            previous_charge_end = _parse_iso_datetime(earlier_charge_rows[-1].get("end_time")) or day_start
        _append_gap_as_depot_presence(
            append_row=_append,
            template_row=sample_row,
            band_id=band_id,
            band_label=band_label,
            start_dt=previous_charge_end,
            end_dt=first_relevant_start,
            origin_label=depot_label,
            destination_label=first_relevant_origin,
            depot_label=depot_label,
        )

    for left_idx, right_idx in zip(relevant_indices, relevant_indices[1:]):
        left_row = normalized_full_rows[left_idx]
        right_row = normalized_full_rows[right_idx]
        between_rows = normalized_full_rows[left_idx + 1 : right_idx]
        blocking_rows = [
            row
            for row in between_rows
            if not _row_is_charge(row) and not _row_is_explicit_band_deadhead(row, band_id)
        ]
        if blocking_rows:
            continue

        charge_rows = [row for row in between_rows if _row_is_charge(row)]
        left_end = _parse_iso_datetime(left_row.get("end_time"))
        right_start = _parse_iso_datetime(right_row.get("start_time"))
        left_destination = str(left_row.get("to_location_id") or depot_label).strip() or depot_label
        right_origin = str(right_row.get("from_location_id") or depot_label).strip() or depot_label

        if charge_rows:
            first_charge_start = _parse_iso_datetime(charge_rows[0].get("start_time"))
            if first_charge_start is not None:
                _append_gap_as_depot_presence(
                    append_row=_append,
                    template_row=sample_row,
                    band_id=band_id,
                    band_label=band_label,
                    start_dt=left_end,
                    end_dt=first_charge_start,
                    origin_label=left_destination,
                    destination_label=depot_label,
                    depot_label=depot_label,
                )
            previous_charge_end: Optional[datetime] = None
            for charge_row in charge_rows:
                charge_start = _parse_iso_datetime(charge_row.get("start_time"))
                if previous_charge_end is not None and charge_start is not None:
                    _append_gap_as_depot_presence(
                        append_row=_append,
                        template_row=sample_row,
                        band_id=band_id,
                        band_label=band_label,
                        start_dt=previous_charge_end,
                        end_dt=charge_start,
                        origin_label=depot_label,
                        destination_label=depot_label,
                        depot_label=depot_label,
                    )
                _append(charge_row)
                previous_charge_end = _parse_iso_datetime(charge_row.get("end_time")) or previous_charge_end
            _append_gap_as_depot_presence(
                append_row=_append,
                template_row=sample_row,
                band_id=band_id,
                band_label=band_label,
                start_dt=previous_charge_end,
                end_dt=right_start,
                origin_label=depot_label,
                destination_label=right_origin,
                depot_label=depot_label,
            )
            continue

        gap_min = (
            (right_start - left_end).total_seconds() / 60.0
            if left_end is not None and right_start is not None
            else 0.0
        )
        if gap_min >= 60.0:
            _append_gap_as_depot_presence(
                append_row=_append,
                template_row=sample_row,
                band_id=band_id,
                band_label=band_label,
                start_dt=left_end,
                end_dt=right_start,
                origin_label=left_destination,
                destination_label=right_origin,
                depot_label=depot_label,
            )

    last_relevant_idx = relevant_indices[-1]
    last_relevant_row = normalized_full_rows[last_relevant_idx]
    last_relevant_end = _parse_iso_datetime(last_relevant_row.get("end_time"))
    last_relevant_destination = str(last_relevant_row.get("to_location_id") or depot_label).strip() or depot_label
    later_rows = normalized_full_rows[last_relevant_idx + 1 :]
    later_non_charge_rows = [row for row in later_rows if not _row_is_charge(row)]
    if not later_non_charge_rows:
        later_charge_rows = [row for row in later_rows if _row_is_charge(row)]
        if later_charge_rows:
            first_charge_start = _parse_iso_datetime(later_charge_rows[0].get("start_time"))
            _append_gap_as_depot_presence(
                append_row=_append,
                template_row=sample_row,
                band_id=band_id,
                band_label=band_label,
                start_dt=last_relevant_end,
                end_dt=first_charge_start,
                origin_label=last_relevant_destination,
                destination_label=depot_label,
                depot_label=depot_label,
            )
            previous_charge_end: Optional[datetime] = None
            for charge_row in later_charge_rows:
                charge_start = _parse_iso_datetime(charge_row.get("start_time"))
                if previous_charge_end is not None and charge_start is not None:
                    _append_gap_as_depot_presence(
                        append_row=_append,
                        template_row=sample_row,
                        band_id=band_id,
                        band_label=band_label,
                        start_dt=previous_charge_end,
                        end_dt=charge_start,
                        origin_label=depot_label,
                        destination_label=depot_label,
                        depot_label=depot_label,
                    )
                _append(charge_row)
                previous_charge_end = _parse_iso_datetime(charge_row.get("end_time")) or previous_charge_end
            _append_gap_as_depot_presence(
                append_row=_append,
                template_row=sample_row,
                band_id=band_id,
                band_label=band_label,
                start_dt=previous_charge_end,
                end_dt=day_end,
                origin_label=depot_label,
                destination_label=depot_label,
                depot_label=depot_label,
            )
        else:
            _append_gap_as_depot_presence(
                append_row=_append,
                template_row=sample_row,
                band_id=band_id,
                band_label=band_label,
                start_dt=last_relevant_end,
                end_dt=day_end,
                origin_label=last_relevant_destination,
                destination_label=depot_label,
                depot_label=depot_label,
            )

    return sorted(
        band_rows,
        key=lambda row: (
            str(row.get("vehicle_id") or ""),
            str(row.get("start_time") or ""),
            str(row.get("state") or ""),
            str(row.get("from_location_id") or ""),
        ),
    )


def _interpolated_band_path_pairs(
    *,
    graph_context: Optional[Dict[str, Any]],
    band_id: str,
    from_label: str,
    to_label: str,
    start_minute: int,
    end_minute: int,
    location_labels: List[str],
) -> List[tuple[float, str]]:
    sequences = _graph_context_band_sequences(graph_context, band_id)
    best_path: List[str] = []
    best_span: Optional[int] = None
    normalized_from = str(from_label or "").strip()
    normalized_to = str(to_label or "").strip()
    if not normalized_from or not normalized_to:
        return []

    for sequence in sequences:
        if normalized_from not in sequence or normalized_to not in sequence:
            continue
        for from_idx, label in enumerate(sequence):
            if label != normalized_from:
                continue
            for to_idx, target in enumerate(sequence):
                if target != normalized_to or to_idx == from_idx:
                    continue
                span = abs(to_idx - from_idx)
                if best_span is not None and span >= best_span:
                    continue
                segment = sequence[from_idx : to_idx + 1] if from_idx < to_idx else list(
                    reversed(sequence[to_idx : from_idx + 1])
                )
                if len(segment) < 2:
                    continue
                best_span = span
                best_path = segment

    if len(best_path) < 2:
        return []

    usable_path = [label for label in best_path if label in location_labels]
    if len(usable_path) < 2:
        return []

    if end_minute <= start_minute:
        end_minute = start_minute + 1
    total_steps = max(len(usable_path) - 1, 1)
    return [
        (
            start_minute + (end_minute - start_minute) * idx / total_steps,
            label,
        )
        for idx, label in enumerate(usable_path)
    ]


def _route_band_diagram_svg(
    *,
    scenario_id: str,
    band_id: str,
    rows: List[Dict[str, Any]],
    band_label: Optional[str] = None,
    location_labels: Optional[List[str]] = None,
    graph_context: Optional[Dict[str, Any]] = None,
) -> str:
    if not rows:
        return ""

    min_minute = 0
    max_minute = 24 * 60 - 1

    location_labels = list(
        location_labels
        or _diagram_location_labels(
            rows,
            _ordered_location_labels(
                rows,
                sequences=_graph_context_band_sequences(graph_context, band_id),
            ),
        )
    )
    if not location_labels:
        location_labels = ["depot"]

    top_margin = 90
    max_stop_label_len = max(len(str(label or "")) for label in location_labels)
    left_margin = max(220, min(480, 48 + max_stop_label_len * 14))
    right_margin = 420
    bottom_margin = 70
    hour_span = 24.0
    plot_width = max(2400, int(hour_span * 120.0))
    width = left_margin + plot_width + right_margin
    lane_height = 92 if len(location_labels) <= 18 else 84
    plot_height = max(260, (len(location_labels) - 1) * lane_height + 40)

    def _x(minute: int) -> float:
        if max_minute <= min_minute:
            return float(left_margin)
        raw_x = float(left_margin) + (float(minute - min_minute) / float(max_minute - min_minute)) * float(plot_width)
        return min(max(raw_x, float(left_margin)), float(left_margin + plot_width))

    def _y(label: str) -> float:
        if len(location_labels) <= 1:
            return float(top_margin + plot_height / 2.0)
        idx = location_labels.index(label)
        return float(top_margin) + (float(idx) / float(len(location_labels) - 1)) * float(plot_height)

    band_rows_by_vehicle: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        band_rows_by_vehicle[str(row.get("vehicle_id") or "")].append(row)
    vehicle_ids = sorted(vehicle_id for vehicle_id in band_rows_by_vehicle.keys() if vehicle_id)
    max_vehicle_label_len = max(
        (
            len(f"{vehicle_id} [{str((band_rows_by_vehicle.get(vehicle_id) or [{}])[0].get('vehicle_type') or '?')}]")
            for vehicle_id in vehicle_ids
        ),
        default=24,
    )
    right_margin = max(340, min(760, 120 + max_vehicle_label_len * 8))
    width = left_margin + plot_width + right_margin

    hour_marks = list(range(0, 24 * 60, 60))
    display_label = str(band_label or band_id or "").strip() or band_id
    legend_y = top_margin + 18
    legend_height = 58 + 90 + len(vehicle_ids) * 18 + 36
    height = max(top_margin + plot_height + bottom_margin, int(legend_y + legend_height + 20))
    clip_id = f"clip-{_safe_export_name(f'{scenario_id}-{band_id}', fallback='band')}"
    svg_parts: List[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        "<defs>",
        f'<clipPath id="{clip_id}"><rect x="{left_margin}" y="{top_margin}" width="{plot_width}" height="{plot_height}"/></clipPath>',
        "</defs>",
        '<rect width="100%" height="100%" fill="#fffdf8"/>',
        f'<text x="{left_margin}" y="34" font-size="24" font-family="Segoe UI, Meiryo, sans-serif" fill="#14213d">Route Band Diagram: {html.escape(display_label)}</text>',
        f'<text x="{left_margin}" y="60" font-size="13" font-family="Segoe UI, Meiryo, sans-serif" fill="#5c677d">scenario={html.escape(scenario_id)} / vehicles={len(vehicle_ids)} / route-only stop axis / full-day 00:00-23:59 / depot stay inferred</text>',
    ]

    for label in location_labels:
        y = _y(label)
        svg_parts.append(
            f'<line x1="{left_margin}" y1="{y:.1f}" x2="{left_margin + plot_width}" y2="{y:.1f}" stroke="#d9dde5" stroke-width="1"/>'
        )
        svg_parts.append(
            f'<text x="{left_margin - 12}" y="{y + 5:.1f}" text-anchor="end" font-size="13" font-family="Segoe UI, Meiryo, sans-serif" fill="#22304a">{html.escape(label)}</text>'
        )

    for minute in hour_marks:
        x = _x(minute)
        svg_parts.append(
            f'<line x1="{x:.1f}" y1="{top_margin}" x2="{x:.1f}" y2="{top_margin + plot_height}" stroke="#eceff4" stroke-width="1"/>'
        )
        svg_parts.append(
            f'<text x="{x:.1f}" y="{top_margin + plot_height + 26}" text-anchor="middle" font-size="12" font-family="Segoe UI, Meiryo, sans-serif" fill="#394867">{minute // 60:02d}:{minute % 60:02d}</text>'
        )
    svg_parts.append(
        f'<text x="{left_margin + plot_width:.1f}" y="{top_margin + plot_height + 26}" text-anchor="end" font-size="12" font-family="Segoe UI, Meiryo, sans-serif" fill="#394867">23:59</text>'
    )

    svg_parts.append(
        f'<rect x="{left_margin}" y="{top_margin}" width="{plot_width}" height="{plot_height}" fill="none" stroke="#98a2b3" stroke-width="1.2"/>'
    )

    legend_x = left_margin + plot_width + 28
    svg_parts.append(
        f'<text x="{legend_x}" y="{legend_y - 10}" font-size="14" font-family="Segoe UI, Meiryo, sans-serif" fill="#14213d">Vehicle Types</text>'
    )
    for type_idx, vehicle_type in enumerate(["BEV", "ICE"]):
        type_y = legend_y + type_idx * 18
        svg_parts.append(
            f'<line x1="{legend_x}" y1="{type_y:.1f}" x2="{legend_x + 18}" y2="{type_y:.1f}" stroke="{_vehicle_type_legend_color(vehicle_type)}" stroke-width="4"/>'
        )
        svg_parts.append(
            f'<text x="{legend_x + 24}" y="{type_y + 4:.1f}" font-size="11" font-family="Consolas, Meiryo, monospace" fill="#22304a">{html.escape(vehicle_type)}</text>'
        )
    legend_y += 58
    svg_parts.append(
        f'<text x="{legend_x}" y="{legend_y - 10}" font-size="14" font-family="Segoe UI, Meiryo, sans-serif" fill="#14213d">Line Styles</text>'
    )
    style_specs = [
        ("service", "", 3.6, "#22304a"),
        ("depot deadhead", ' stroke-dasharray="8 5"', 2.4, "#22304a"),
        ("depot stay", ' stroke-dasharray="2 6"', 2.2, "#22304a"),
        ("charge at depot", ' stroke-dasharray="4 4"', 3.0, "#22304a"),
        ("ICE refuel mark", "", 0.0, "#6cab2f"),
    ]
    for style_idx, (style_label, dash, stroke_width, style_color) in enumerate(style_specs):
        style_y = legend_y + style_idx * 18
        if style_label == "ICE refuel mark":
            cx = legend_x + 9
            cy = style_y
            points = f"{cx:.1f},{cy - 4.0:.1f} {cx + 4.0:.1f},{cy:.1f} {cx:.1f},{cy + 4.0:.1f} {cx - 4.0:.1f},{cy:.1f}"
            svg_parts.append(
                f'<polygon points="{points}" fill="{style_color}" opacity="0.95"/>'
            )
        else:
            svg_parts.append(
                f'<line x1="{legend_x}" y1="{style_y:.1f}" x2="{legend_x + 18}" y2="{style_y:.1f}" stroke="{style_color}" stroke-width="{stroke_width:.1f}"{dash}/>'
            )
        svg_parts.append(
            f'<text x="{legend_x + 24}" y="{style_y + 4:.1f}" font-size="11" font-family="Consolas, Meiryo, monospace" fill="#22304a">{html.escape(style_label)}</text>'
        )
    legend_y += 90
    svg_parts.append(
        f'<text x="{legend_x}" y="{legend_y - 10}" font-size="14" font-family="Segoe UI, Meiryo, sans-serif" fill="#14213d">Vehicles</text>'
    )

    for legend_idx, vehicle_id in enumerate(vehicle_ids):
        vehicle_rows = sorted(
            band_rows_by_vehicle[vehicle_id],
            key=lambda row: (str(row.get("start_time") or ""), str(row.get("state") or "")),
        )
        vehicle_type = str(vehicle_rows[0].get("vehicle_type") or "")
        color = _vehicle_line_color(vehicle_id, vehicle_type)
        legend_row_y = legend_y + legend_idx * 18
        svg_parts.append(
            f'<line x1="{legend_x}" y1="{legend_row_y:.1f}" x2="{legend_x + 18}" y2="{legend_row_y:.1f}" stroke="{color}" stroke-width="3"/>'
        )
        svg_parts.append(
            f'<text x="{legend_x + 24}" y="{legend_row_y + 4:.1f}" font-size="11" font-family="Consolas, Meiryo, monospace" fill="#22304a">{html.escape(vehicle_id)} [{html.escape(vehicle_type or "?")}]</text>'
        )

    first_label_done: set[str] = set()
    plot_elements: List[str] = [f'<g clip-path="url(#{clip_id})">']
    label_elements: List[str] = []
    for vehicle_id in vehicle_ids:
        vehicle_rows = sorted(
            band_rows_by_vehicle[vehicle_id],
            key=lambda row: (str(row.get("start_time") or ""), str(row.get("state") or "")),
        )
        vehicle_type = str(vehicle_rows[0].get("vehicle_type") or "")
        color = _vehicle_line_color(vehicle_id, vehicle_type)
        for row in vehicle_rows:
            start_minute = _parse_iso_minute(row.get("start_time"))
            end_minute = _parse_iso_minute(row.get("end_time"))
            if start_minute is None or end_minute is None:
                continue
            from_label = str(
                row.get("from_location_id")
                or row.get("to_location_id")
                or row.get("depot_id")
                or location_labels[0]
            ).strip() or location_labels[0]
            to_label = str(
                row.get("to_location_id")
                or row.get("from_location_id")
                or row.get("depot_id")
                or location_labels[0]
            ).strip() or location_labels[0]
            if from_label not in location_labels:
                from_label = location_labels[0]
            if to_label not in location_labels:
                to_label = location_labels[0]
            x1 = _x(start_minute)
            x2 = _x(max(end_minute, start_minute + 1))
            y1 = _y(from_label)
            y2 = _y(to_label)
            state = str(row.get("state") or "")
            if state == "service":
                dash = ""
                stroke_width = 3.6
                opacity = 0.96
            elif state == "deadhead":
                dash = ' stroke-dasharray="8 5"'
                stroke_width = 2.4
                opacity = 0.86
            elif state == "idle":
                dash = ' stroke-dasharray="2 6"'
                stroke_width = 2.2
                opacity = 0.60
            elif state == "charge":
                dash = ' stroke-dasharray="4 4"'
                stroke_width = 3.0
                opacity = 0.76
            elif state == "refuel":
                dash = ' stroke-dasharray="1 4"'
                stroke_width = 2.0
                opacity = 0.78
            else:
                dash = ' stroke-dasharray="3 4"'
                stroke_width = 2.2
                opacity = 0.82
            point_pairs = [(x1, y1), (x2, y2)]
            if state == "service":
                stop_points = _graph_context_task_stop_points(
                    graph_context,
                    str(row.get("trip_id") or ""),
                )
                candidate_pairs: List[tuple[float, float]] = []
                for stop_point in stop_points:
                    stop_label = str(stop_point.get("stop_label") or "").strip()
                    stop_minute = _parse_iso_minute(
                        stop_point.get("departure_time") or stop_point.get("arrival_time")
                    )
                    if (
                        not stop_label
                        or stop_label not in location_labels
                        or stop_minute is None
                    ):
                        continue
                    candidate_pairs.append((_x(stop_minute), _y(stop_label)))
                if len(candidate_pairs) >= 2:
                    point_pairs = candidate_pairs
                else:
                    interpolated_pairs = _interpolated_band_path_pairs(
                        graph_context=graph_context,
                        band_id=band_id,
                        from_label=from_label,
                        to_label=to_label,
                        start_minute=start_minute,
                        end_minute=end_minute,
                        location_labels=location_labels,
                    )
                    if len(interpolated_pairs) >= 2:
                        point_pairs = [
                            (_x(int(round(minute))), _y(label))
                            for minute, label in interpolated_pairs
                        ]
            path = " ".join(
                (
                    f"M {point_pairs[0][0]:.1f} {point_pairs[0][1]:.1f}",
                    *[
                        f"L {point_x:.1f} {point_y:.1f}"
                        for point_x, point_y in point_pairs[1:]
                    ],
                )
            )
            plot_elements.append(
                f'<path d="{path}" fill="none" stroke="{color}" stroke-width="{stroke_width:.1f}" stroke-linecap="round" opacity="{opacity:.2f}"{dash}/>'
            )
            for point_x, point_y in point_pairs:
                plot_elements.append(
                    f'<circle cx="{point_x:.1f}" cy="{point_y:.1f}" r="2.4" fill="{color}" opacity="0.95"/>'
                )
            if state == "refuel":
                marker_x = point_pairs[-1][0]
                marker_y = point_pairs[-1][1]
                marker_points = (
                    f"{marker_x:.1f},{marker_y - 4.8:.1f} "
                    f"{marker_x + 4.8:.1f},{marker_y:.1f} "
                    f"{marker_x:.1f},{marker_y + 4.8:.1f} "
                    f"{marker_x - 4.8:.1f},{marker_y:.1f}"
                )
                plot_elements.append(
                    f'<polygon points="{marker_points}" fill="#6cab2f" stroke="#3f7d1b" stroke-width="0.8" opacity="0.98"/>'
                )
            if vehicle_id not in first_label_done:
                first_label_done.add(vehicle_id)
                label_x = min(max(x1 + 6.0, float(left_margin + 6)), float(left_margin + plot_width - 140))
                label_y = min(max(y1 - 6.0, float(top_margin + 12)), float(top_margin + plot_height - 6))
                label_elements.append(
                    f'<text x="{label_x:.1f}" y="{label_y:.1f}" font-size="10.5" font-family="Consolas, Meiryo, monospace" fill="{color}">{html.escape(vehicle_id)}</text>'
                )

    plot_elements.append("</g>")
    svg_parts.extend(plot_elements)
    svg_parts.extend(label_elements)
    svg_parts.append("</svg>")
    return "".join(svg_parts)


def _build_route_band_diagram_assets(
    vehicle_timeline_rows: List[Dict[str, Any]],
    scenario_id: str,
    *,
    graph_context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    service_rows_all = [
        row
        for row in vehicle_timeline_rows
        if str(row.get("state") or "") == "service" and str(row.get("band_id") or "").strip()
    ]
    service_bands_by_vehicle: Dict[str, set[str]] = defaultdict(set)
    for row in service_rows_all:
        vehicle_id = str(row.get("vehicle_id") or "").strip()
        band_id = str(row.get("band_id") or "").strip()
        if vehicle_id and band_id:
            service_bands_by_vehicle[vehicle_id].add(band_id)

    full_rows_by_vehicle: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in vehicle_timeline_rows:
        vehicle_id = str(row.get("vehicle_id") or "").strip()
        if not vehicle_id:
            continue
        full_rows_by_vehicle[vehicle_id].append(dict(row))

    rows_by_band: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in service_rows_all:
        band_id = str(row.get("band_id") or "").strip()
        if not band_id:
            continue
        rows_by_band[band_id].append(row)

    entries: List[Dict[str, Any]] = []
    svg_payloads: Dict[str, str] = {}
    for band_id in sorted(rows_by_band.keys()):
        grouped_rows = rows_by_band[band_id]
        service_rows = [row for row in grouped_rows if str(row.get("state") or "") == "service"]
        if not service_rows:
            continue
        band_label = str(
            ((graph_context or {}).get("band_labels_by_band_id") or {}).get(band_id)
            or service_rows[0].get("band_label")
            or band_id
        )
        band_vehicle_ids = sorted(
            {
                str(row.get("vehicle_id") or "").strip()
                for row in service_rows
                if str(row.get("vehicle_id") or "").strip()
            }
        )
        band_rows: List[Dict[str, Any]] = []
        for vehicle_id in band_vehicle_ids:
            vehicle_band_rows = _vehicle_band_rows(
                vehicle_rows=full_rows_by_vehicle.get(vehicle_id, []),
                band_id=band_id,
                band_label=band_label,
                graph_context=graph_context,
            )
            band_rows.extend(vehicle_band_rows)
        band_sequences = _graph_context_band_sequences(graph_context, band_id)
        main_location_labels = _ordered_location_labels(service_rows, sequences=band_sequences)
        route_location_labels = _diagram_location_labels(band_rows, main_location_labels)
        vehicle_ids = sorted(
            {
                str(row.get("vehicle_id") or "")
                for row in band_rows
                if str(row.get("vehicle_id") or "").strip()
            }
        )
        vehicle_type_counts: Dict[str, int] = {}
        for vehicle_id in vehicle_ids:
            vehicle_rows = [
                row for row in band_rows if str(row.get("vehicle_id") or "").strip() == vehicle_id
            ]
            vehicle_type = str(vehicle_rows[0].get("vehicle_type") or "").strip() if vehicle_rows else ""
            if not vehicle_type:
                continue
            vehicle_type_counts[vehicle_type] = vehicle_type_counts.get(vehicle_type, 0) + 1
        event_route_band_ids = sorted(
            {
                str(row.get("event_route_band_id") or "").strip()
                for row in service_rows
                if str(row.get("event_route_band_id") or "").strip()
            }
        )
        shared_vehicle_ids = sorted(
            vehicle_id
            for vehicle_id in vehicle_ids
            if len(service_bands_by_vehicle.get(vehicle_id, set())) > 1
        )
        filename = f"{_safe_export_name(band_id, fallback='route_band')}.svg"
        svg_payloads[filename] = _route_band_diagram_svg(
            scenario_id=scenario_id,
            band_id=band_id,
            rows=band_rows,
            band_label=band_label,
            location_labels=route_location_labels,
            graph_context=graph_context,
        )
        entries.append(
            {
                "scenario_id": scenario_id,
                "band_id": band_id,
                "band_label": band_label,
                "vehicle_count": len(vehicle_ids),
                "vehicle_ids": vehicle_ids,
                "vehicle_type_counts": vehicle_type_counts,
                "service_trip_count": len(service_rows),
                "location_count": len(route_location_labels),
                "event_route_band_ids": event_route_band_ids,
                "shared_vehicle_ids": shared_vehicle_ids,
                "mixed_event_route_band_detected": bool(shared_vehicle_ids),
                "diagram_file": filename,
            }
        )
    return {"entries": entries, "svg_payloads": svg_payloads}


def _write_route_band_diagram_assets(target_root: Path, assets: Dict[str, Any]) -> None:
    entries = list(assets.get("entries") or [])
    svg_payloads = dict(assets.get("svg_payloads") or {})
    if not entries:
        return
    output_dir = target_root / "route_band_diagrams"
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest = {
        "schema_version": "1.0.0",
        "generated_at": _tokyo_now().isoformat(),
        "grouping_key": "band_id",
        "diagram_format": "svg",
        "entries": entries,
    }
    with open(output_dir / "manifest.json", "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    for filename, svg_text in svg_payloads.items():
        (output_dir / filename).write_text(svg_text, encoding="utf-8")


def _build_cost_breakdown_json(data: ProblemData, sim: SimulationResult, scenario_id: str) -> Dict[str, Any]:
    components = {
        "electricity_energy_cost": float(sim.total_energy_cost or 0.0),
        "demand_charge_cost": float(sim.total_demand_charge or 0.0),
        "diesel_cost": float(sim.total_fuel_cost or 0.0),
        "co2_cost": float(data.co2_price_per_kg or 0.0) * float(sim.total_co2_kg or 0.0),
        "battery_degradation_cost": float(sim.total_degradation_cost or 0.0),
        "charger_operation_cost": 0.0,
        "pv_capex_daily_equivalent": 0.0,
        "ess_cost": 0.0,
        "unserved_trip_penalty": float(len(sim.unserved_tasks or [])) * float(data.objective_weights.get("unserved_penalty", 0.0) or 0.0),
    }
    total_cost = float(sim.total_operating_cost or 0.0)
    return {
        "scenario_id": scenario_id,
        "currency": "JPY",
        "total_cost": total_cost,
        "components": components,
        "meta": {
            "objective_mode": str(getattr(data, "objective_mode", "total_cost") or "total_cost"),
            "solver_mode": "unknown",
            "includes_pv": bool(data.enable_pv),
        },
    }


def _build_kpi_summary_json(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
    sim: SimulationResult,
    scenario_id: str,
) -> Dict[str, Any]:
    served_count = max(len(data.tasks) - len(sim.unserved_tasks or []), 0)
    total_distance = sum(float(task.distance_km or 0.0) for task in data.tasks)
    deadhead_distance = 0.0
    for vehicle_id, tasks in milp.assignment.items():
        sorted_tasks = sorted(tasks, key=lambda task_id: dp.task_lut.get(task_id).start_time_idx if dp.task_lut.get(task_id) else 0)
        for i in range(len(sorted_tasks) - 1):
            deadhead_distance += float(dp.deadhead_distance_km.get(sorted_tasks[i], {}).get(sorted_tasks[i + 1], 0.0) or 0.0)

    total_energy = sum(float(task.energy_required_kwh_bev or 0.0) for task in data.tasks)
    total_charge_energy = 0.0
    for vehicle_id, by_charger in milp.charge_power_kw.items():
        for charger_id, series in by_charger.items():
            for kw in series:
                total_charge_energy += float(kw or 0.0) * float(data.delta_t_hour)

    charger_values = [float(v) for v in (sim.charger_utilization or {}).values()]
    min_soc_pct = 0.0
    if ms.K_BEV:
        pct_values: List[float] = []
        for vehicle_id in ms.K_BEV:
            vehicle = dp.vehicle_lut.get(vehicle_id)
            capacity = float(getattr(vehicle, "battery_capacity", 0.0) or 0.0)
            series = list(milp.soc_series.get(vehicle_id, []))
            if capacity <= 0 or not series:
                continue
            pct_values.extend([(float(v) / capacity) * 100.0 for v in series])
        if pct_values:
            min_soc_pct = min(pct_values)
            avg_soc_pct = sum(pct_values) / len(pct_values)
        else:
            avg_soc_pct = 0.0
    else:
        avg_soc_pct = 0.0

    return {
        "scenario_id": scenario_id,
        "fleet_size": len(ms.K_ALL),
        "served_trip_count": served_count,
        "unserved_trip_count": len(sim.unserved_tasks or []),
        "served_trip_rate": float(sim.served_task_ratio or 0.0),
        "total_distance_km": float(total_distance),
        "total_deadhead_km": float(deadhead_distance),
        "deadhead_ratio": float(deadhead_distance / total_distance) if total_distance > 0 else 0.0,
        "total_energy_consumption_kwh": float(total_energy),
        "total_charging_energy_kwh": float(total_charge_energy),
        "peak_grid_import_kw": float(sim.peak_demand_kw or 0.0),
        "peak_charge_kw": float(max((max(series) if series else 0.0) for by_charger in milp.charge_power_kw.values() for series in by_charger.values()) if milp.charge_power_kw else 0.0),
        "pv_generation_total_kwh": float(sim.total_pv_kwh or 0.0),
        "pv_self_consumption_kwh": float(sim.total_pv_kwh or 0.0),
        "pv_utilization_ratio": float(sim.pv_self_consumption_ratio or 0.0),
        "min_soc_pct": float(min_soc_pct),
        "average_soc_pct": float(avg_soc_pct),
        "charger_utilization_avg": (sum(charger_values) / len(charger_values)) if charger_values else 0.0,
        "charger_utilization_max": max(charger_values) if charger_values else 0.0,
        "total_cost_jpy": float(sim.total_operating_cost or 0.0),
        "electricity_cost_jpy": float(sim.total_energy_cost or 0.0),
        "electricity_cost_basis": str(sim.energy_cost_basis or "provisional_drive"),
        "electricity_cost_provisional_jpy": float(sim.provisional_energy_cost or 0.0),
        "electricity_cost_charged_jpy": float(sim.charged_energy_cost or 0.0),
        "grid_energy_provisional_kwh": float(sim.provisional_grid_kwh or 0.0),
        "grid_energy_charged_kwh": float(sim.charged_grid_kwh or 0.0),
        "co2_kg": float(sim.total_co2_kg or 0.0),
        "solver_runtime_sec": float(milp.solve_time_sec or 0.0),
        "solution_status": str(milp.status or "UNKNOWN").lower(),
    }


# ---------------------------------------------------------------------------
# §13.1.1 summary.json
# ---------------------------------------------------------------------------


def export_summary_json(
    run_dir: Path,
    milp: MILPResult,
    sim: SimulationResult,
    run_label: Optional[str] = None,
) -> None:
    summary = {
        "run_label": run_label or "",
        "timestamp": datetime.now().isoformat(),
        "status": milp.status,
        "objective_value": milp.objective_value,
        "mip_gap": milp.mip_gap,
        "solve_time_sec": milp.solve_time_sec,
        "infeasibility_info": milp.infeasibility_info,
        "cost_breakdown": {
            "total_operating_cost": sim.total_operating_cost,
            "electricity_cost": sim.total_energy_cost,
            "electricity_cost_basis": sim.energy_cost_basis,
            "electricity_cost_provisional": sim.provisional_energy_cost,
            "electricity_cost_charged": sim.charged_energy_cost,
            "demand_charge": sim.total_demand_charge,
            "fuel_cost": sim.total_fuel_cost,
            "degradation_cost": sim.total_degradation_cost,
        },
        "kpi": {
            "served_task_ratio": sim.served_task_ratio,
            "unserved_tasks": sim.unserved_tasks,
            "total_grid_kwh": sim.total_grid_kwh,
            "grid_kwh_provisional": sim.provisional_grid_kwh,
            "grid_kwh_charged": sim.charged_grid_kwh,
            "total_pv_kwh": sim.total_pv_kwh,
            "pv_self_consumption_ratio": sim.pv_self_consumption_ratio,
            "peak_demand_kw": sim.peak_demand_kw,
            "total_co2_kg": sim.total_co2_kg,
            "soc_min_kwh": sim.soc_min_kwh,
            "soc_violations": sim.soc_violations,
            "vehicle_utilization": sim.vehicle_utilization,
            "charger_utilization": sim.charger_utilization,
        },
    }
    with open(run_dir / "summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)


def _normalize_direction(direction: Optional[str]) -> str:
    value = str(direction or "").strip().lower()
    if value in {"outbound", "out", "up", "0", "上り", "上り便", "↗"}:
        return "outbound"
    if value in {"inbound", "in", "down", "1", "下り", "下り便", "↙"}:
        return "inbound"
    if value in {"circular", "loop", "循環", "循環線"}:
        return "circular"
    return "unknown"


def _variant_bucket(variant: Optional[str]) -> str:
    value = str(variant or "").strip().lower()
    if value in {"main", "main_outbound", "main_inbound"}:
        return "main"
    if value == "short_turn":
        return "short_turn"
    if value in {"depot", "depot_in", "depot_out"}:
        return "depot"
    return "unknown"


def export_targeted_trips(run_dir: Path, data: ProblemData, milp: MILPResult) -> None:
    served = {task_id for tasks in milp.assignment.values() for task_id in tasks}
    rows: List[Dict[str, Any]] = []
    for task in data.tasks:
        route_series_code, _route_series_prefix, _route_series_number, _series_source = extract_route_series_from_candidates(
            task.route_family_code,
            task.route_id,
        )
        rows.append(
            {
                "task_id": task.task_id,
                "route_id": task.route_id or "",
                "route_series_code": route_series_code,
                "service_id": task.service_id or "",
                "direction": _normalize_direction(task.direction),
                "route_variant_type": task.route_variant_type or "unknown",
                "origin": task.origin,
                "destination": task.destination,
                "start_time_idx": task.start_time_idx,
                "end_time_idx": task.end_time_idx,
                "distance_km": task.distance_km,
                "served": task.task_id in served,
            }
        )

    payload = {
        "targeted_task_count": len(data.tasks),
        "served_task_count": len(served),
        "unserved_task_count": max(len(data.tasks) - len(served), 0),
        "tasks": rows,
    }
    with open(run_dir / "targeted_trips.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    _write_csv(run_dir / "targeted_trips.csv", rows)


def export_trip_type_counts(run_dir: Path, data: ProblemData) -> None:
    counts = {
        "main_outbound": 0,
        "main_inbound": 0,
        "main_circular": 0,
        "short_turn_outbound": 0,
        "short_turn_inbound": 0,
        "short_turn_circular": 0,
        "depot": 0,
        "unknown": 0,
    }
    by_route: Dict[str, int] = defaultdict(int)

    for task in data.tasks:
        direction = _normalize_direction(task.direction)
        variant = _variant_bucket(task.route_variant_type)
        route_id = task.route_id or "(unknown_route)"
        by_route[route_id] += 1

        if variant == "main":
            key = f"main_{direction}" if direction in {"outbound", "inbound", "circular"} else "unknown"
        elif variant == "short_turn":
            key = (
                f"short_turn_{direction}"
                if direction in {"outbound", "inbound", "circular"}
                else "unknown"
            )
        elif variant == "depot":
            key = "depot"
        else:
            key = "unknown"
        counts[key] = counts.get(key, 0) + 1

    payload = {
        "total_task_count": len(data.tasks),
        "counts": counts,
        "counts_by_route": dict(sorted(by_route.items())),
    }
    with open(run_dir / "trip_type_counts.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    _write_csv(
        run_dir / "trip_type_counts.csv",
        [{"bucket": key, "count": value} for key, value in counts.items()],
    )


def export_cost_breakdown_detail(run_dir: Path, sim: SimulationResult) -> None:
    rows = [
        {"component": "total_operating_cost", "yen": sim.total_operating_cost},
        {"component": "electricity_cost", "yen": sim.total_energy_cost},
        {"component": "demand_charge", "yen": sim.total_demand_charge},
        {"component": "vehicle_fixed_cost", "yen": sim.total_vehicle_fixed_cost},
        {"component": "driver_cost", "yen": sim.total_driver_cost},
        {"component": "fuel_cost", "yen": sim.total_fuel_cost},
        {"component": "degradation_cost", "yen": sim.total_degradation_cost},
    ]
    with open(run_dir / "cost_breakdown_detail.json", "w", encoding="utf-8") as f:
        json.dump({"cost_breakdown": rows}, f, ensure_ascii=False, indent=2)
    _write_csv(run_dir / "cost_breakdown_detail.csv", rows)


def export_co2_breakdown(
    run_dir: Path,
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
    sim: SimulationResult,
) -> None:
    co2_ice = 0.0
    for vehicle_id in ms.K_ICE:
        vehicle = dp.vehicle_lut.get(vehicle_id)
        if vehicle is None:
            continue
        for task_id in milp.assignment.get(vehicle_id, []):
            co2_ice += vehicle.co2_emission_coeff * dp.task_fuel_ice.get(task_id, 0.0)

    has_grid_co2_factor = any(
        value > 0.0
        for site_map in dp.grid_co2_factor.values()
        for value in site_map.values()
    )
    co2_grid = max(sim.total_co2_kg - co2_ice, 0.0) if has_grid_co2_factor else None

    payload = {
        "total_co2_kg": sim.total_co2_kg,
        "engine_bus_co2_kg": round(co2_ice, 4),
        "power_generation_co2_kg": round(co2_grid, 4) if co2_grid is not None else None,
        "power_generation_co2_note": (
            "Calculated from grid CO2 factors"
            if co2_grid is not None
            else "Not implemented yet (grid CO2 factor unavailable)"
        ),
    }
    with open(run_dir / "co2_breakdown.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    _write_csv(
        run_dir / "co2_breakdown.csv",
        [
            {"component": "engine_bus_co2_kg", "value": payload["engine_bus_co2_kg"]},
            {"component": "power_generation_co2_kg", "value": payload["power_generation_co2_kg"]},
            {"component": "total_co2_kg", "value": payload["total_co2_kg"]},
        ],
    )


def export_vehicle_timelines(
    run_dir: Path,
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
) -> None:
    base_date = _tokyo_now().date()

    def _slot_to_minute(slot_idx: int) -> int:
        return int(round(float(slot_idx) * float(data.delta_t_min)))

    def _minute_to_hhmm(total_minute: int) -> str:
        minute = int(total_minute) % (24 * 60)
        return f"{minute // 60:02d}:{minute % 60:02d}"

    def _direction_label(direction: Optional[str]) -> str:
        normalized = _normalize_direction(direction)
        return normalized if normalized != "unknown" else ""

    def _service_label(task: Any) -> str:
        route_id = (getattr(task, "route_id", None) or "").strip()
        direction = _direction_label(getattr(task, "direction", None))
        variant = (getattr(task, "route_variant_type", None) or "").strip()
        parts = [part for part in [route_id, direction, variant] if part]
        if parts:
            return " | ".join(parts)
        origin = getattr(task, "origin", "")
        destination = getattr(task, "destination", "")
        return f"{origin} -> {destination}".strip()

    def _enrich_event(vehicle_id: str, event: Dict[str, Any], event_seq: int) -> Dict[str, Any]:
        start_idx = int(event.get("start_time_idx", 0) or 0)
        end_idx = int(event.get("end_time_idx", 0) or 0)
        start_min = _slot_to_minute(start_idx)
        end_min = _slot_to_minute(end_idx)
        duration_slots = max(end_idx - start_idx, 0)
        duration_min = max(end_min - start_min, 0)
        event_type = str(event.get("event_type", ""))
        if event_type == "service":
            activity_group = "operation"
        elif event_type == "deadhead":
            activity_group = "deadhead"
        elif event_type == "charging":
            activity_group = "charging"
        elif event_type == "refuel":
            activity_group = "refuel"
        else:
            activity_group = "other"

        row = {
            "vehicle_id": vehicle_id,
            "event_id": f"{vehicle_id}:{event_seq:04d}",
            "event_seq": event_seq,
            "event_type": event_type,
            "activity_group": activity_group,
            "gantt_lane": vehicle_id,
            "start_time_idx": start_idx,
            "end_time_idx": end_idx,
            "duration_slots": duration_slots,
            "duration_min": duration_min,
            "start_minute": start_min,
            "end_minute": end_min,
            "start_hhmm": _minute_to_hhmm(start_min),
            "end_hhmm": _minute_to_hhmm(end_min),
            "task_id": event.get("task_id") or "",
            "route_id": event.get("route_id") or "",
            "route_series_code": event.get("route_series_code") or "",
            "direction": _direction_label(event.get("direction")),
            "route_variant_type": event.get("route_variant_type") or "",
            "service_id": event.get("service_id") or "",
            "service_label": event.get("service_label") or "",
            "origin": event.get("origin") or "",
            "destination": event.get("destination") or "",
            "from_task_id": event.get("from_task_id") or "",
            "to_task_id": event.get("to_task_id") or "",
            "from_route_id": event.get("from_route_id") or "",
            "to_route_id": event.get("to_route_id") or "",
            "deadhead_time_slot": event.get("deadhead_time_slot") or 0,
            "deadhead_distance_km": event.get("deadhead_distance_km") or 0.0,
            "charger_id": event.get("charger_id") or "",
            "avg_power_kw": event.get("avg_power_kw") or 0.0,
            "refuel_liters": event.get("refuel_liters") or 0.0,
            "distance_km": event.get("distance_km") or 0.0,
            "timeline_note": event.get("timeline_note") or "",
        }
        return row

    timeline_by_vehicle: Dict[str, List[Dict[str, Any]]] = {}
    csv_rows: List[Dict[str, Any]] = []

    for vehicle_id in ms.K_ALL:
        events: List[Dict[str, Any]] = []
        assigned = sorted(
            milp.assignment.get(vehicle_id, []),
            key=lambda task_id: dp.task_lut.get(task_id).start_time_idx if dp.task_lut.get(task_id) else 0,
        )

        previous_task_id: Optional[str] = None
        for task_id in assigned:
            task = dp.task_lut.get(task_id)
            if task is None:
                continue

            if previous_task_id is not None:
                dh_slots = dp.deadhead_time_slot.get(previous_task_id, {}).get(task_id, 0)
                if dh_slots > 0:
                    prev_task = dp.task_lut.get(previous_task_id)
                    deadhead_event = {
                        "event_type": "deadhead",
                        "from_task_id": previous_task_id,
                        "to_task_id": task_id,
                        "from_route_id": getattr(prev_task, "route_id", None),
                        "to_route_id": task.route_id,
                        "timeline_note": "between services",
                        "start_time_idx": max(task.start_time_idx - dh_slots, 0),
                        "end_time_idx": task.start_time_idx,
                        "deadhead_time_slot": dh_slots,
                        "deadhead_distance_km": dp.deadhead_distance_km.get(previous_task_id, {}).get(task_id, 0.0),
                    }
                    events.append(deadhead_event)

            service_event = {
                "event_type": "service",
                "task_id": task_id,
                "route_id": task.route_id,
                "route_series_code": (
                    extract_route_series_from_candidates(task.route_family_code, task.route_id)[0]
                ),
                "direction": task.direction,
                "route_variant_type": task.route_variant_type,
                "service_id": task.service_id,
                "service_label": _service_label(task),
                "origin": task.origin,
                "destination": task.destination,
                "start_time_idx": task.start_time_idx,
                "end_time_idx": task.end_time_idx,
                "distance_km": task.distance_km,
            }
            events.append(service_event)
            previous_task_id = task_id

        for charger_id, flags in milp.charge_schedule.get(vehicle_id, {}).items():
            powers = milp.charge_power_kw.get(vehicle_id, {}).get(charger_id, [0.0] * len(flags))
            start_idx: Optional[int] = None
            power_sum = 0.0
            power_count = 0
            for idx, flag in enumerate(flags):
                active = idx < len(powers) and (flag > 0 or powers[idx] > 1e-6)
                if active and start_idx is None:
                    start_idx = idx
                if active:
                    power_sum += float(powers[idx]) if idx < len(powers) else 0.0
                    power_count += 1
                if not active and start_idx is not None:
                    events.append(
                        {
                            "event_type": "charging",
                            "charger_id": charger_id,
                            "start_time_idx": start_idx,
                            "end_time_idx": idx,
                            "avg_power_kw": (power_sum / power_count) if power_count else 0.0,
                            "timeline_note": "plug-in charging",
                        }
                    )
                    start_idx = None
                    power_sum = 0.0
                    power_count = 0
            if start_idx is not None:
                events.append(
                    {
                        "event_type": "charging",
                        "charger_id": charger_id,
                        "start_time_idx": start_idx,
                        "end_time_idx": len(flags),
                        "avg_power_kw": (power_sum / power_count) if power_count else 0.0,
                        "timeline_note": "plug-in charging",
                    }
                )

        for slot_idx, liters in enumerate(list(milp.refuel_schedule_l.get(vehicle_id, []))):
            refuel_liters = float(liters or 0.0)
            if refuel_liters <= 1.0e-9:
                continue
            events.append(
                {
                    "event_type": "refuel",
                    "start_time_idx": slot_idx,
                    "end_time_idx": slot_idx + 1,
                    "refuel_liters": round(refuel_liters, 4),
                    "origin": getattr(dp.vehicle_lut.get(vehicle_id), "home_depot", "") or "",
                    "destination": getattr(dp.vehicle_lut.get(vehicle_id), "home_depot", "") or "",
                    "timeline_note": "depot refuel",
                }
            )

        events.sort(key=lambda event: (event.get("start_time_idx", 0), event.get("event_type", "")))
        enriched_events: List[Dict[str, Any]] = []
        for seq, event in enumerate(events, start=1):
            enriched = _enrich_event(vehicle_id, event, seq)
            enriched_events.append(enriched)
            csv_rows.append(enriched)

        timeline_by_vehicle[vehicle_id] = enriched_events

    with open(run_dir / "vehicle_timelines.json", "w", encoding="utf-8") as f:
        json.dump(
            {
                "timeline_schema_version": "2.0",
                "delta_t_min": data.delta_t_min,
                "vehicle_timelines": timeline_by_vehicle,
                "vehicle_gantt_rows": csv_rows,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )
    refuel_event_rows = _build_refuel_event_rows(
        data,
        ms,
        dp,
        milp,
        scenario_id="",
        base_date=base_date,
        planning_start_time="00:00",
    )
    _write_csv(run_dir / "vehicle_timelines.csv", csv_rows)
    _write_csv(run_dir / "vehicle_timeline_gantt.csv", csv_rows)
    _write_csv(run_dir / "refuel_events.csv", refuel_event_rows)


def export_objective_breakdown(run_dir: Path, milp: MILPResult, sim: SimulationResult) -> None:
    breakdown = dict(milp.obj_breakdown or {})
    if not breakdown:
        breakdown = {
            "electricity_cost": sim.total_energy_cost,
            "demand_charge": sim.total_demand_charge,
            "fuel_cost": sim.total_fuel_cost,
            "degradation_cost": sim.total_degradation_cost,
            "vehicle_fixed_cost": sim.total_vehicle_fixed_cost,
            "driver_cost": sim.total_driver_cost,
        }
    with open(run_dir / "objective_breakdown.json", "w", encoding="utf-8") as f:
        json.dump({"objective_breakdown": breakdown}, f, ensure_ascii=False, indent=2)
    _write_csv(
        run_dir / "objective_breakdown.csv",
        [{"term": key, "value": value} for key, value in breakdown.items()],
    )


def export_simulation_conditions(run_dir: Path, data: ProblemData, dp: DerivedParams) -> None:
    vehicle_rows: List[Dict[str, Any]] = []
    for vehicle in data.vehicles:
        vehicle_rows.append(
            {
                "vehicle_id": vehicle.vehicle_id,
                "vehicle_type": vehicle.vehicle_type,
                "fixed_use_cost_yen": vehicle.fixed_use_cost,
                "fuel_cost_coeff_yen_per_liter": vehicle.fuel_cost_coeff,
                "battery_degradation_cost_coeff_yen_per_kwh": vehicle.battery_degradation_cost_coeff,
                "co2_emission_coeff_kg_per_liter": vehicle.co2_emission_coeff,
            }
        )

    vehicle_cost_summary: Dict[str, Dict[str, float]] = {}
    for row in vehicle_rows:
        vtype = str(row["vehicle_type"])
        bucket = vehicle_cost_summary.setdefault(
            vtype,
            {
                "count": 0.0,
                "avg_fixed_use_cost_yen": 0.0,
                "avg_fuel_cost_coeff_yen_per_liter": 0.0,
                "avg_battery_degradation_cost_coeff_yen_per_kwh": 0.0,
            },
        )
        bucket["count"] += 1.0
        bucket["avg_fixed_use_cost_yen"] += float(row["fixed_use_cost_yen"])
        bucket["avg_fuel_cost_coeff_yen_per_liter"] += float(row["fuel_cost_coeff_yen_per_liter"])
        bucket["avg_battery_degradation_cost_coeff_yen_per_kwh"] += float(
            row["battery_degradation_cost_coeff_yen_per_kwh"]
        )

    for bucket in vehicle_cost_summary.values():
        count = bucket["count"] if bucket["count"] > 0 else 1.0
        bucket["avg_fixed_use_cost_yen"] = bucket["avg_fixed_use_cost_yen"] / count
        bucket["avg_fuel_cost_coeff_yen_per_liter"] = (
            bucket["avg_fuel_cost_coeff_yen_per_liter"] / count
        )
        bucket["avg_battery_degradation_cost_coeff_yen_per_kwh"] = (
            bucket["avg_battery_degradation_cost_coeff_yen_per_kwh"] / count
        )

    tou_rows: List[Dict[str, Any]] = []
    for site_id, by_time in dp.grid_price.items():
        for time_idx in sorted(by_time.keys()):
            tou_rows.append(
                {
                    "site_id": site_id,
                    "time_idx": time_idx,
                    "grid_energy_price_yen_per_kwh": by_time.get(time_idx, 0.0),
                    "sell_back_price_yen_per_kwh": dp.sell_back_price.get(site_id, {}).get(time_idx, 0.0),
                    "base_load_kw": dp.base_load_kw.get(site_id, {}).get(time_idx, 0.0),
                    "grid_co2_factor_kg_per_kwh": dp.grid_co2_factor.get(site_id, {}).get(time_idx, 0.0),
                }
            )

    price_values = [float(row["grid_energy_price_yen_per_kwh"]) for row in tou_rows]
    electricity_price_summary = {
        "has_tou_price_table": len(tou_rows) > 0,
        "time_slot_count": len(price_values),
        "grid_energy_price_min_yen_per_kwh": min(price_values) if price_values else None,
        "grid_energy_price_max_yen_per_kwh": max(price_values) if price_values else None,
        "grid_energy_price_avg_yen_per_kwh": (
            sum(price_values) / len(price_values) if price_values else None
        ),
    }

    contract_rows: List[Dict[str, Any]] = []
    for site in data.sites:
        contract_rows.append(
            {
                "site_id": site.site_id,
                "site_type": site.site_type,
                "contract_demand_limit_kw": site.contract_demand_limit_kw,
                "grid_import_limit_kw": site.grid_import_limit_kw,
                "site_transformer_limit_kw": site.site_transformer_limit_kw,
            }
        )

    payload = {
        "timestamp": datetime.now().isoformat(),
        "time_settings": {
            "delta_t_min": data.delta_t_min,
            "num_periods": data.num_periods,
            "planning_horizon_hours": data.planning_horizon_hours,
        },
        "timeSettings": {
            "deltaTMin": data.delta_t_min,
            "numPeriods": data.num_periods,
            "planningHorizonHours": data.planning_horizon_hours,
        },
        "flags": {
            "enable_pv": data.enable_pv,
            "enable_v2g": data.enable_v2g,
            "enable_demand_charge": data.enable_demand_charge,
            "enable_battery_degradation": data.enable_battery_degradation,
            "allow_partial_service": data.allow_partial_service,
        },
        "flagsCamel": {
            "enablePv": data.enable_pv,
            "enableV2g": data.enable_v2g,
            "enableDemandCharge": data.enable_demand_charge,
            "enableBatteryDegradation": data.enable_battery_degradation,
            "allowPartialService": data.allow_partial_service,
        },
        "unit_prices_and_costs": {
            "vehicle_introduction_cost_source": "vehicle.fixed_use_cost",
            "fuel_unit_price_source": "vehicle.fuel_cost_coeff",
            "battery_degradation_unit_price_source": "vehicle.battery_degradation_cost_coeff",
            "co2_price_per_kg": data.co2_price_per_kg,
            "demand_charge_rate_per_kw": data.demand_charge_rate_per_kw,
            "electricity_price_summary": electricity_price_summary,
        },
        "unitPricesAndCosts": {
            "vehicleIntroductionCostSource": "vehicle.fixed_use_cost",
            "fuelUnitPriceSource": "vehicle.fuel_cost_coeff",
            "batteryDegradationUnitPriceSource": "vehicle.battery_degradation_cost_coeff",
            "co2PricePerKg": data.co2_price_per_kg,
            "demandChargeRatePerKw": data.demand_charge_rate_per_kw,
            "electricityPriceSummary": {
                "hasTouPriceTable": electricity_price_summary["has_tou_price_table"],
                "timeSlotCount": electricity_price_summary["time_slot_count"],
                "gridEnergyPriceMinYenPerKwh": electricity_price_summary[
                    "grid_energy_price_min_yen_per_kwh"
                ],
                "gridEnergyPriceMaxYenPerKwh": electricity_price_summary[
                    "grid_energy_price_max_yen_per_kwh"
                ],
                "gridEnergyPriceAvgYenPerKwh": electricity_price_summary[
                    "grid_energy_price_avg_yen_per_kwh"
                ],
            },
        },
        "demand_and_contract_conditions": {
            "demand_charge_rate_per_kw": data.demand_charge_rate_per_kw,
            "objective_weight_demand_charge_cost": data.objective_weights.get(
                "demand_charge_cost", 0.0
            ),
            "contract_limit_penalty_multiplier": data.objective_weights.get(
                "contract_limit_penalty_multiplier"
            ),
            "contract_limit_penalty_note": (
                "If null, no explicit penalty multiplier is configured in objective_weights"
            ),
            "contract_limits_by_site": contract_rows,
        },
        "demandAndContractConditions": {
            "demandChargeRatePerKw": data.demand_charge_rate_per_kw,
            "objectiveWeightDemandChargeCost": data.objective_weights.get(
                "demand_charge_cost", 0.0
            ),
            "contractLimitPenaltyMultiplier": data.objective_weights.get(
                "contract_limit_penalty_multiplier"
            ),
            "contractLimitsBySite": contract_rows,
        },
        "extensible_coefficients": {
            "objective_weights": dict(data.objective_weights),
            "big_m": {
                "BIG_M_ASSIGN": data.BIG_M_ASSIGN,
                "BIG_M_CHARGE": data.BIG_M_CHARGE,
                "BIG_M_SOC": data.BIG_M_SOC,
                "EPSILON": data.EPSILON,
            },
        },
        "extensibleCoefficients": {
            "objectiveWeights": dict(data.objective_weights),
            "bigM": {
                "bigMAssign": data.BIG_M_ASSIGN,
                "bigMCharge": data.BIG_M_CHARGE,
                "bigMSoc": data.BIG_M_SOC,
                "epsilon": data.EPSILON,
            },
        },
        "vehicle_costs": vehicle_rows,
        "vehicleCosts": vehicle_rows,
        "vehicle_cost_summary_by_type": vehicle_cost_summary,
        "vehicleCostSummaryByType": vehicle_cost_summary,
        "tou_prices": tou_rows,
        "touPrices": tou_rows,
    }

    with open(run_dir / "simulation_conditions.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    _write_csv(run_dir / "simulation_conditions_vehicle_costs.csv", vehicle_rows)
    _write_csv(run_dir / "simulation_conditions_tou_prices.csv", tou_rows)
    _write_csv(run_dir / "simulation_conditions_contract_limits.csv", contract_rows)


# ---------------------------------------------------------------------------
# §13.1.2 vehicle_schedule.csv
# ---------------------------------------------------------------------------


def export_vehicle_schedule(
    run_dir: Path,
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp: MILPResult,
) -> None:
    rows = []
    for k in ms.K_ALL:
        assigned = milp.assignment.get(k, [])
        if assigned:
            for r_id in assigned:
                task = dp.task_lut.get(r_id)
                rows.append(
                    {
                        "vehicle_id": k,
                        "vehicle_type": dp.vehicle_lut[k].vehicle_type,
                        "task_id": r_id,
                        "route_series_code": (
                            extract_route_series_from_candidates(task.route_family_code if task else None, task.route_id if task else None)[0]
                            if task
                            else ""
                        ),
                        "start_time_idx": task.start_time_idx if task else "",
                        "end_time_idx": task.end_time_idx if task else "",
                        "origin": task.origin if task else "",
                        "destination": task.destination if task else "",
                        "distance_km": task.distance_km if task else "",
                        "energy_kwh_bev": task.energy_required_kwh_bev if task else "",
                    }
                )
        else:
            rows.append(
                {
                    "vehicle_id": k,
                    "vehicle_type": dp.vehicle_lut[k].vehicle_type,
                    "task_id": "(unassigned)",
                    "route_series_code": "",
                    "start_time_idx": "",
                    "end_time_idx": "",
                    "origin": "",
                    "destination": "",
                    "distance_km": "",
                    "energy_kwh_bev": "",
                }
            )

    _write_csv(run_dir / "vehicle_schedule.csv", rows)


# ---------------------------------------------------------------------------
# §13.1.3 charging_schedule.csv
# ---------------------------------------------------------------------------


def export_charging_schedule(
    run_dir: Path,
    ms: ModelSets,
    milp: MILPResult,
) -> None:
    rows = []
    for k in ms.K_BEV:
        soc_series = milp.soc_series.get(k, [])
        for c in ms.C:
            pwr_series = milp.charge_power_kw.get(k, {}).get(c, [0.0] * len(ms.T))
            z_series = milp.charge_schedule.get(k, {}).get(c, [0] * len(ms.T))
            for t_idx in ms.T:
                soc_val = soc_series[t_idx] if t_idx < len(soc_series) else ""
                rows.append(
                    {
                        "vehicle_id": k,
                        "charger_id": c,
                        "time_idx": t_idx,
                        "z_charge": z_series[t_idx] if t_idx < len(z_series) else 0,
                        "p_charge_kw": pwr_series[t_idx]
                        if t_idx < len(pwr_series)
                        else 0.0,
                        "soc_kwh": soc_val,
                    }
                )

    _write_csv(run_dir / "charging_schedule.csv", rows)


# ---------------------------------------------------------------------------
# §13.1.4 site_power_balance.csv
# ---------------------------------------------------------------------------


def export_site_power_balance(
    run_dir: Path,
    ms: ModelSets,
    milp: MILPResult,
    sim: SimulationResult,
    data: ProblemData,
) -> None:
    rows = []
    for site_id in ms.I_CHARGE:
        grid_series = milp.grid_import_kw.get(site_id, [0.0] * len(ms.T))
        pv_series = milp.pv_used_kw.get(site_id, [0.0] * len(ms.T))
        for t_idx in ms.T:
            rows.append(
                {
                    "site_id": site_id,
                    "time_idx": t_idx,
                    "grid_import_kw": grid_series[t_idx]
                    if t_idx < len(grid_series)
                    else 0.0,
                    "pv_used_kw": pv_series[t_idx] if t_idx < len(pv_series) else 0.0,
                    "peak_demand_kw": milp.peak_demand_kw.get(site_id, 0.0),
                }
            )

    _write_csv(run_dir / "site_power_balance.csv", rows)


# ---------------------------------------------------------------------------
# §13.1.5 experiment_report.md
# ---------------------------------------------------------------------------


def export_experiment_report(
    run_dir: Path,
    data: ProblemData,
    ms: ModelSets,
    milp: MILPResult,
    sim: SimulationResult,
    run_label: Optional[str] = None,
) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines = [
        f"# 実験レポート — {run_label or ts}",
        "",
        "## 条件一覧",
        f"- 実行日時: {ts}",
        f"- 車両数 BEV: {len(ms.K_BEV)}, ICE: {len(ms.K_ICE)}",
        f"- タスク数: {len(ms.R)}",
        f"- 充電器数: {len(ms.C)}",
        f"- 時間刻み: {data.delta_t_min:.0f} 分 ({data.num_periods} スロット)",
        f"- PV 有効: {data.enable_pv}",
        f"- V2G 有効: {data.enable_v2g}",
        f"- デマンド料金有効: {data.enable_demand_charge}",
        "",
        "## ソルバー結果",
        f"- ステータス: **{milp.status}**",
        f"- 目的関数値: {milp.objective_value}",
        f"- MIP ギャップ: {milp.mip_gap}",
        f"- 計算時間: {milp.solve_time_sec:.2f} 秒",
        "",
        "## 目的関数内訳",
        f"| 項目 | 値 [円] |",
        f"|------|---------|",
        f"| 電力量料金 | {sim.total_energy_cost:,.0f} |",
        f"| デマンド料金 | {sim.total_demand_charge:,.0f} |",
        f"| 燃料費 | {sim.total_fuel_cost:,.0f} |",
        f"| 電池劣化 | {sim.total_degradation_cost:,.0f} |",
        f"| **合計** | **{sim.total_operating_cost:,.0f}** |",
        "",
        "## 主要 KPI",
        f"- タスク担当率: {sim.served_task_ratio * 100:.1f} %",
        f"- 未担当タスク: {sim.unserved_tasks or 'なし'}",
        f"- 系統受電量: {sim.total_grid_kwh:.2f} kWh",
        f"- PV 利用量: {sim.total_pv_kwh:.2f} kWh",
        f"- PV 自家消費率: {sim.pv_self_consumption_ratio * 100:.1f} %",
        f"- ピーク需要: {sim.peak_demand_kw:.2f} kW",
        f"- CO2 排出: {sim.total_co2_kg:.2f} kg",
        f"- 最低 SOC: {sim.soc_min_kwh:.2f} kWh",
        f"- SOC 違反: {len(sim.soc_violations)} 件",
        "",
        "## infeasible 情報",
        milp.infeasibility_info or "なし",
        "",
        "---",
        "*本レポートは result_exporter.py により自動生成されました。*",
    ]
    with open(run_dir / "experiment_report.md", "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------


def _write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# Excel multi-sheet export (openpyxl)
# ---------------------------------------------------------------------------


def export_excel(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp_result: MILPResult,
    sim_result: SimulationResult,
    run_dir: Path,
    run_label: Optional[str] = None,
) -> Path:
    """
    openpyxl を使い、results.xlsx を run_dir 内に生成する。

    シート構成
    ----------
    Summary          : KPI サマリー（縦持ちキー/値テーブル）
    KPIs             : 10 KPI 一覧（数値）
    VehicleSchedule  : vehicle_schedule.csv と同内容
    ChargingSchedule : charging_schedule.csv と同内容（先頭 10,000 行）
    SitePowerBalance : site_power_balance.csv と同内容

    Returns
    -------
    Path
        生成した xlsx ファイルのパス
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "Excel エクスポートには openpyxl が必要です: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # ---- ヘルパー: シートにテーブルを書く -----------------------------------
    def _write_sheet(
        ws: "openpyxl.worksheet.worksheet.Worksheet",
        headers: List[str],
        rows_data: List[List[Any]],
        freeze: bool = True,
    ) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="1F5C99")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row_vals in enumerate(rows_data, 2):
            for col_idx, val in enumerate(row_vals, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        # 列幅自動調整（最大 40 文字）
        for col_idx, h in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(h)),
                max((len(str(r[col_idx - 1])) for r in rows_data if r), default=0),
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        if freeze:
            ws.freeze_panes = "A2"

    # ---- Sheet 1: Summary --------------------------------------------------
    ws_sum = wb.create_sheet("Summary")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary_rows: List[List[Any]] = [
        ["実行ラベル", run_label or ""],
        ["生成日時", ts],
        ["ステータス", milp_result.status],
        ["目的関数値 [円]", milp_result.objective_value],
        ["MIP ギャップ", milp_result.mip_gap],
        ["計算時間 [秒]", round(milp_result.solve_time_sec, 3)],
        ["BEV 台数", len(ms.K_BEV)],
        ["ICE 台数", len(ms.K_ICE)],
        ["タスク数", len(ms.R)],
        ["充電器数", len(ms.C)],
        ["時間刻み [分]", data.delta_t_min],
        ["時間スロット数", data.num_periods],
        ["PV 有効", data.enable_pv],
        ["V2G 有効", data.enable_v2g],
        ["デマンド料金有効", data.enable_demand_charge],
        ["--- コスト内訳 ---", ""],
        ["電力量料金 [円]", round(sim_result.total_energy_cost, 0)],
        ["デマンド料金 [円]", round(sim_result.total_demand_charge, 0)],
        ["燃料費 [円]", round(sim_result.total_fuel_cost, 0)],
        ["電池劣化費 [円]", round(sim_result.total_degradation_cost, 0)],
        ["運行コスト合計 [円]", round(sim_result.total_operating_cost, 0)],
    ]
    _write_sheet(ws_sum, ["項目", "値"], summary_rows)

    # ---- Sheet 2: KPIs -----------------------------------------------------
    ws_kpi = wb.create_sheet("KPIs")
    charger_utilization = _to_scalar_metric(sim_result.charger_utilization)

    unserved_tasks_value = sim_result.unserved_tasks
    if isinstance(unserved_tasks_value, list):
        unmet_trips = len(unserved_tasks_value)
    else:
        unmet_trips = int(unserved_tasks_value or 0)

    kpi_rows: List[List[Any]] = [
        ["objective_value", milp_result.objective_value],
        ["total_energy_cost", round(sim_result.total_energy_cost, 2)],
        ["total_demand_charge", round(sim_result.total_demand_charge, 2)],
        ["total_fuel_cost", round(sim_result.total_fuel_cost, 2)],
        [
            "vehicle_fixed_cost",
            round(getattr(sim_result, "vehicle_fixed_cost", 0.0), 2),
        ],
        ["unmet_trips", unmet_trips],
        ["soc_min_margin_kwh", round(getattr(sim_result, "soc_min_kwh", 0.0), 3)],
        ["charger_utilization", round(charger_utilization, 4)],
        ["peak_grid_power_kw", round(sim_result.peak_demand_kw, 2)],
        ["solve_time_sec", round(milp_result.solve_time_sec, 3)],
    ]
    _write_sheet(ws_kpi, ["KPI", "値"], kpi_rows)

    # ---- Sheet 3: VehicleSchedule ------------------------------------------
    ws_vs = wb.create_sheet("VehicleSchedule")
    vs_headers = [
        "vehicle_id",
        "vehicle_type",
        "task_id",
        "start_time_idx",
        "end_time_idx",
        "origin",
        "destination",
        "distance_km",
        "energy_kwh_bev",
    ]
    vs_rows: List[List[Any]] = []
    for k in ms.K_ALL:
        assigned = milp_result.assignment.get(k, [])
        if assigned:
            for r_id in assigned:
                task = dp.task_lut.get(r_id)
                vs_rows.append(
                    [
                        k,
                        dp.vehicle_lut[k].vehicle_type,
                        r_id,
                        task.start_time_idx if task else "",
                        task.end_time_idx if task else "",
                        task.origin if task else "",
                        task.destination if task else "",
                        task.distance_km if task else "",
                        task.energy_required_kwh_bev if task else "",
                    ]
                )
        else:
            vs_rows.append(
                [
                    k,
                    dp.vehicle_lut[k].vehicle_type,
                    "(unassigned)",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
    _write_sheet(ws_vs, vs_headers, vs_rows)

    # ---- Sheet 4: ChargingSchedule (先頭 10,000 行に制限) ------------------
    ws_cs = wb.create_sheet("ChargingSchedule")
    cs_headers = [
        "vehicle_id",
        "charger_id",
        "time_idx",
        "z_charge",
        "p_charge_kw",
        "soc_kwh",
    ]
    cs_rows: List[List[Any]] = []
    MAX_CS_ROWS = 10_000
    for k in ms.K_BEV:
        if len(cs_rows) >= MAX_CS_ROWS:
            break
        soc_series = milp_result.soc_series.get(k, [])
        for c in ms.C:
            if len(cs_rows) >= MAX_CS_ROWS:
                break
            pwr_series = milp_result.charge_power_kw.get(k, {}).get(
                c, [0.0] * len(ms.T)
            )
            z_series = milp_result.charge_schedule.get(k, {}).get(c, [0] * len(ms.T))
            for t_idx in ms.T:
                if len(cs_rows) >= MAX_CS_ROWS:
                    break
                soc_val = soc_series[t_idx] if t_idx < len(soc_series) else ""
                cs_rows.append(
                    [
                        k,
                        c,
                        t_idx,
                        z_series[t_idx] if t_idx < len(z_series) else 0,
                        pwr_series[t_idx] if t_idx < len(pwr_series) else 0.0,
                        soc_val,
                    ]
                )
    _write_sheet(ws_cs, cs_headers, cs_rows)

    # ---- Sheet 5: SitePowerBalance -----------------------------------------
    ws_sp = wb.create_sheet("SitePowerBalance")
    sp_headers = [
        "site_id",
        "time_idx",
        "grid_import_kw",
        "pv_used_kw",
        "peak_demand_kw",
    ]
    sp_rows: List[List[Any]] = []
    for site_id in ms.I_CHARGE:
        grid_series = milp_result.grid_import_kw.get(site_id, [0.0] * len(ms.T))
        pv_series = milp_result.pv_used_kw.get(site_id, [0.0] * len(ms.T))
        for t_idx in ms.T:
            sp_rows.append(
                [
                    site_id,
                    t_idx,
                    grid_series[t_idx] if t_idx < len(grid_series) else 0.0,
                    pv_series[t_idx] if t_idx < len(pv_series) else 0.0,
                    milp_result.peak_demand_kw.get(site_id, 0.0),
                ]
            )
    _write_sheet(ws_sp, sp_headers, sp_rows)

    # ---- 保存 --------------------------------------------------------------
    out_path = run_dir / "results.xlsx"
    wb.save(out_path)
    return out_path


def export_excel_bytes(
    data: ProblemData,
    ms: ModelSets,
    dp: DerivedParams,
    milp_result: MILPResult,
    sim_result: SimulationResult,
    run_label: Optional[str] = None,
) -> bytes:
    """
    results.xlsx をファイルに保存せず bytes として返す。
    Streamlit の st.download_button に直接渡せる。

    Parameters
    ----------
    data, ms, dp, milp_result, sim_result : 各パイプライン出力
    run_label : 任意のラベル文字列

    Returns
    -------
    bytes
        xlsx バイナリデータ
    """
    import io
    import tempfile
    from pathlib import Path as _Path

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = _Path(tmp)
        xlsx_path = export_excel(
            data, ms, dp, milp_result, sim_result, tmp_path, run_label
        )
        return xlsx_path.read_bytes()
